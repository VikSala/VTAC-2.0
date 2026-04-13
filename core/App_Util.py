from services.utils import Utils
import xmlrpc
import os
import math
import pandas as pd
import re

#region REGION INFORMACIÓN COMERCIAL
'''
import_comercial_stock()
    ├── abrir_vtac_gui()
    ├── create_excel_base()
    │       ├── parse_madrid_from_dir()
    │       ├── parse_bulgaria_from_dir()
    │       ├── parse_vs_from_dir()
    │       └── odoo_vtac()
    ├── create_vs_vn()
    │       ├── vd_vn_spain()
    │       └── crear_hoja_resultado()
    └── actualizar_stock_comercial()
            ├── update_comercial_stock()
            ├── update_names()
            │       └── update_vz_products()
            └── update_out_of_stock_msg_from_excel()
'''
def update_comercial_stock(excel_path):
    from App_Connection import db, uid, password, models
    import pandas as pd
    from collections import defaultdict

    print("🚀 Update comercial stock (modo PRO)")

    # --- CONFIG ---
    FIELD_MADRID = "x_transit_stock_custom"
    FIELD_BULGARIA = "x_almacen1_custom"
    FIELD_ODOO = "x_almacen_local"

    SHEET_MADRID = "Madrid"
    SHEET_BULGARIA = "Bulgaria"
    SHEET_ODOO = "Odoo"
    SHEET_PROX = "Proximamente"

    # --- HELPERS ---
    def _coerce_sku(x):
        return str(x).strip() if x and str(x).strip() else None

    def _to_float(v):
        try:
            return float(v) if v not in (None, "", "nan") else 0.0
        except:
            return 0.0

    # --- LOAD EXCEL ---
    xls = pd.read_excel(excel_path, sheet_name=[SHEET_MADRID, SHEET_BULGARIA, SHEET_ODOO])

    def norm(df):
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df

    df_mad = norm(xls[SHEET_MADRID])
    df_bul = norm(xls[SHEET_BULGARIA])
    df_odoo = norm(xls[SHEET_ODOO]) if SHEET_ODOO in xls else pd.DataFrame()

    # --- SKUS ---
    all_skus = set(df_mad["SKU"].dropna()) | set(df_bul["SKU"].dropna()) | set(df_odoo.get("SKU", []))
    all_skus = {_coerce_sku(s) for s in all_skus if _coerce_sku(s)}

    # --- INDEX PRODUCTS ---
    CHUNK = 1000
    sku_map = {}

    for i in range(0, len(all_skus), CHUNK):
        batch = list(all_skus)[i:i+CHUNK]

        rows = models.execute_kw(
            db, uid, password, "product.product", "search_read",
            [[("default_code", "in", batch)]],
            {"fields": ["id", "default_code", "product_tmpl_id",
                        FIELD_MADRID, FIELD_BULGARIA, FIELD_ODOO]}
        )

        for r in rows:
            sku = r["default_code"].strip()
            sku_map[sku] = {
                "model": "product.product",
                "id": r["id"],
                "values": {
                    FIELD_MADRID: r.get(FIELD_MADRID),
                    FIELD_BULGARIA: r.get(FIELD_BULGARIA),
                    FIELD_ODOO: r.get(FIELD_ODOO),
                }
            }

    # --- PREPARE CHANGES ---
    grouped = defaultdict(list)
    proximamente = []

    # --- MADRID ---
    for _, row in df_mad.iterrows():
        sku = _coerce_sku(row["SKU"])
        if not sku or sku not in sku_map:
            continue

        new_val = _to_float(row["STOCK"])
        old_val = sku_map[sku]["values"].get(FIELD_MADRID) or 0.0

        if new_val != old_val:
            grouped[("product.product", FIELD_MADRID, new_val)].append(sku_map[sku]["id"])

    # --- BULGARIA ---
    odoo_low = { _coerce_sku(r["SKU"]) for _, r in df_odoo.iterrows() if _to_float(r["STOCK"]) < 1 }

    for _, row in df_bul.iterrows():
        sku = _coerce_sku(row["SKU"])
        if not sku or sku not in sku_map:
            continue

        stock = max(_to_float(row["STOCK"]), 0)
        old_val = sku_map[sku]["values"].get(FIELD_BULGARIA) or 0.0

        if stock != old_val:
            grouped[("product.product", FIELD_BULGARIA, stock)].append(sku_map[sku]["id"])

        # lógica PROX
        undel = row.get("UNDELIVERED ORDER")
        if undel not in (None, "", "nan") and stock <= 0 and sku in odoo_low:
            proximamente.append(sku)

    # --- SIN UPDATE ODOO: x_almacen_local ---

    # --- WRITE BATCH ---
    total_updates = 0

    for (model, field, value), ids in grouped.items():
        models.execute_kw(
            db, uid, password,
            model, "write",
            [ids, {field: value}]
        )
        total_updates += len(ids)

    # --- WRITE PROXIMAMENTE ---
    try:
        df_prox = pd.DataFrame({"SKU": sorted(set(proximamente))})
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_prox.to_excel(writer, sheet_name=SHEET_PROX, index=False)
    except Exception as e:
        print(f"⚠ Error escribiendo Proximamente: {e}")

    print(f"✅ Updates reales: {total_updates}")

    return {"updated": total_updates}

def update_names(excel_path):
    from App_Connection import db, uid, password, models
    import pandas as pd
    import re
    from collections import defaultdict

    print("🚀 update_names PRO")

    SHEET = "Resultado"
    COLS = ("VS", "VN", "VD", "VZ")

    def norm_sku(x):
        if not x:
            return None
        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s if s else None

    # --- 1. LEER EXCEL ---
    df = pd.read_excel(excel_path, sheet_name=SHEET, dtype=str)

    for c in COLS:
        if c not in df.columns:
            df[c] = None

    sku_to_prefix = {}

    for col in COLS:
        for val in df[col]:
            sku = norm_sku(val)
            if not sku:
                continue

            if sku in sku_to_prefix:
                sku_to_prefix[sku] = None  # conflicto
            else:
                sku_to_prefix[sku] = col

    # eliminar conflictos
    sku_to_prefix = {k: v for k, v in sku_to_prefix.items() if v}

    print(f"➡ SKUs válidos: {len(sku_to_prefix)}")

    # --- 2. LEER ODOO MASIVO ---
    skus = list(sku_to_prefix.keys())
    CHUNK = 1000

    product_map = {}

    for i in range(0, len(skus), CHUNK):
        batch = skus[i:i+CHUNK]

        rows = models.execute_kw(
            db, uid, password,
            "product.product", "search_read",
            [[("default_code", "in", batch)]],
            {"fields": ["id", "default_code", "product_tmpl_id", "name"]}
        )

        for r in rows:
            sku = r["default_code"].strip()
            product_map[sku] = {
                "tmpl_id": r["product_tmpl_id"][0],
                "name": r["name"]
            }

    # --- 3. CALCULAR CAMBIOS ---
    grouped = defaultdict(list)
    nuevos_vz = []

    for sku, prefix in sku_to_prefix.items():
        if sku not in product_map:
            continue

        data = product_map[sku]
        old_name = data["name"]

        if "**" in old_name or "VEE25" in old_name:
            continue

        new_name = re.sub(r"^\[[A-Za-z]{2}([^\]]*)\]", rf"[{prefix}\1]", old_name)

        if new_name == old_name:
            continue

        allow = False if prefix == "VZ" else True

        key = (new_name, allow)
        grouped[key].append(data["tmpl_id"])

        if prefix == "VZ":
            nuevos_vz.append(data["tmpl_id"])

    # --- 4. WRITE MASIVO ---
    total = 0

    for (name, allow), ids in grouped.items():
        models.execute_kw(
            db, uid, password,
            "product.template", "write",
            [ids, {
                "name": name,
                "allow_out_of_stock_order": allow
            }]
        )
        total += len(ids)

    print(f"✅ Productos actualizados: {total}")
    print(f"➡ Nuevos VZ: {len(nuevos_vz)}")
    print("50%")

    # --- 5. POST PROCESO ---
    def update_vz_products(vz_tmpl_ids):
        from App_Connection import db, uid, password, models
        from collections import defaultdict

        ctx = {"active_test": False, "lang": "es_ES"}

        DESCATALOGADO_BINARIO = '''
                iVBORw0KGgoAAAANSUhEUgAAAhkAAAIWCAYAAAD6TJ9eAAAACXBIWXMAAA7EAAAOxAGVKw4bAAAK+mlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0IGJlZ2luPSLvu78iIGlkPSJXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQiPz4gPHg6eG1wbWV0YSB4bWxuczp4PSJhZG9iZTpuczptZXRhLyIgeDp4bXB0az0iQWRvYmUgWE1QIENvcmUgOS4xLWMwMDIgNzkuZjM1NGVmYywgMjAyMy8xMS8wOS0xMjo0MDoyNyAgICAgICAgIj4gPHJkZjpSREYgeG1sbnM6cmRmPSJodHRwOi8vd3d3LnczLm9yZy8xOTk5LzAyLzIyLXJkZi1zeW50YXgtbnMjIj4gPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9IiIgeG1sbnM6eG1wPSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvIiB4bWxuczpkYz0iaHR0cDovL3B1cmwub3JnL2RjL2VsZW1lbnRzLzEuMS8iIHhtbG5zOnBob3Rvc2hvcD0iaHR0cDovL25zLmFkb2JlLmNvbS9waG90b3Nob3AvMS4wLyIgeG1sbnM6eG1wTU09Imh0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC9tbS8iIHhtbG5zOnN0RXZ0PSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvc1R5cGUvUmVzb3VyY2VFdmVudCMiIHhtbG5zOnN0UmVmPSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvc1R5cGUvUmVzb3VyY2VSZWYjIiB4bXA6Q3JlYXRvclRvb2w9IkFkb2JlIFBob3Rvc2hvcCAyMi40IChXaW5kb3dzKSIgeG1wOkNyZWF0ZURhdGU9IjIwMjQtMDEtMDRUMTE6NDc6MjArMDE6MDAiIHhtcDpNb2RpZnlEYXRlPSIyMDI0LTAzLTIwVDExOjE2OjI3KzAxOjAwIiB4bXA6TWV0YWRhdGFEYXRlPSIyMDI0LTAzLTIwVDExOjE2OjI3KzAxOjAwIiBkYzpmb3JtYXQ9ImltYWdlL3BuZyIgcGhvdG9zaG9wOkNvbG9yTW9kZT0iMyIgeG1wTU06SW5zdGFuY2VJRD0ieG1wLmlpZDoxODE4ZTgwNy05MmVkLTNiNDUtYjc0OC01ZGY5OTEwN2VjZjEiIHhtcE1NOkRvY3VtZW50SUQ9ImFkb2JlOmRvY2lkOnBob3Rvc2hvcDo5N2ZjN2EyMC1jZDBiLTBkNDAtYTQwZi0yZmVhYjE2NDljNmQiIHhtcE1NOk9yaWdpbmFsRG9jdW1lbnRJRD0ieG1wLmRpZDowMjA3ZGE1ZC1jOTgzLWRhNDQtOTU2My1lY2VlYTFkNzkxYTYiPiA8eG1wTU06SGlzdG9yeT4gPHJkZjpTZXE+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJjcmVhdGVkIiBzdEV2dDppbnN0YW5jZUlEPSJ4bXAuaWlkOjAyMDdkYTVkLWM5ODMtZGE0NC05NTYzLWVjZWVhMWQ3OTFhNiIgc3RFdnQ6d2hlbj0iMjAyNC0wMS0wNFQxMTo0NzoyMCswMTowMCIgc3RFdnQ6c29mdHdhcmVBZ2VudD0iQWRvYmUgUGhvdG9zaG9wIDIyLjQgKFdpbmRvd3MpIi8+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJzYXZlZCIgc3RFdnQ6aW5zdGFuY2VJRD0ieG1wLmlpZDo4N2FmZDNjNC0zMjA0LWVjNDctODQ4NC03YjkxNjA3MjZlZDYiIHN0RXZ0OndoZW49IjIwMjQtMDMtMjBUMTE6MTY6MTMrMDE6MDAiIHN0RXZ0OnNvZnR3YXJlQWdlbnQ9IkFkb2JlIFBob3Rvc2hvcCAyNS41IChXaW5kb3dzKSIgc3RFdnQ6Y2hhbmdlZD0iLyIvPiA8cmRmOmxpIHN0RXZ0OmFjdGlvbj0iY29udmVydGVkIiBzdEV2dDpwYXJhbWV0ZXJzPSJmcm9tIGltYWdlL3BuZyB0byBhcHBsaWNhdGlvbi92bmQuYWRvYmUucGhvdG9zaG9wIi8+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJkZXJpdmVkIiBzdEV2dDpwYXJhbWV0ZXJzPSJjb252ZXJ0ZWQgZnJvbSBpbWFnZS9wbmcgdG8gYXBwbGljYXRpb24vdm5kLmFkb2JlLnBob3Rvc2hvcCIvPiA8cmRmOmxpIHN0RXZ0OmFjdGlvbj0ic2F2ZWQiIHN0RXZ0Omluc3RhbmNlSUQ9InhtcC5paWQ6YzIyZDg0ZmEtNjE4OC0wYzQzLTk5NDgtN2MwMjdkM2VmM2I2IiBzdEV2dDp3aGVuPSIyMDI0LTAzLTIwVDExOjE2OjEzKzAxOjAwIiBzdEV2dDpzb2Z0d2FyZUFnZW50PSJBZG9iZSBQaG90b3Nob3AgMjUuNSAoV2luZG93cykiIHN0RXZ0OmNoYW5nZWQ9Ii8iLz4gPHJkZjpsaSBzdEV2dDphY3Rpb249InNhdmVkIiBzdEV2dDppbnN0YW5jZUlEPSJ4bXAuaWlkOmU4NzczNWRiLWIxNTMtZDE0ZC1iZmY2LWFjM2M2ZDI0MjUyZSIgc3RFdnQ6d2hlbj0iMjAyNC0wMy0yMFQxMToxNjoyNyswMTowMCIgc3RFdnQ6c29mdHdhcmVBZ2VudD0iQWRvYmUgUGhvdG9zaG9wIDI1LjUgKFdpbmRvd3MpIiBzdEV2dDpjaGFuZ2VkPSIvIi8+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJjb252ZXJ0ZWQiIHN0RXZ0OnBhcmFtZXRlcnM9ImZyb20gYXBwbGljYXRpb24vdm5kLmFkb2JlLnBob3Rvc2hvcCB0byBpbWFnZS9wbmciLz4gPHJkZjpsaSBzdEV2dDphY3Rpb249ImRlcml2ZWQiIHN0RXZ0OnBhcmFtZXRlcnM9ImNvbnZlcnRlZCBmcm9tIGFwcGxpY2F0aW9uL3ZuZC5hZG9iZS5waG90b3Nob3AgdG8gaW1hZ2UvcG5nIi8+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJzYXZlZCIgc3RFdnQ6aW5zdGFuY2VJRD0ieG1wLmlpZDoxODE4ZTgwNy05MmVkLTNiNDUtYjc0OC01ZGY5OTEwN2VjZjEiIHN0RXZ0OndoZW49IjIwMjQtMDMtMjBUMTE6MTY6MjcrMDE6MDAiIHN0RXZ0OnNvZnR3YXJlQWdlbnQ9IkFkb2JlIFBob3Rvc2hvcCAyNS41IChXaW5kb3dzKSIgc3RFdnQ6Y2hhbmdlZD0iLyIvPiA8L3JkZjpTZXE+IDwveG1wTU06SGlzdG9yeT4gPHhtcE1NOkRlcml2ZWRGcm9tIHN0UmVmOmluc3RhbmNlSUQ9InhtcC5paWQ6ZTg3NzM1ZGItYjE1My1kMTRkLWJmZjYtYWMzYzZkMjQyNTJlIiBzdFJlZjpkb2N1bWVudElEPSJ4bXAuZGlkOmMyMmQ4NGZhLTYxODgtMGM0My05OTQ4LTdjMDI3ZDNlZjNiNiIgc3RSZWY6b3JpZ2luYWxEb2N1bWVudElEPSJ4bXAuZGlkOjAyMDdkYTVkLWM5ODMtZGE0NC05NTYzLWVjZWVhMWQ3OTFhNiIvPiA8L3JkZjpEZXNjcmlwdGlvbj4gPC9yZGY6UkRGPiA8L3g6eG1wbWV0YT4gPD94cGFja2V0IGVuZD0iciI/PkID1SQAAYaBSURBVHic7P13dCTZfecLfm5EpLdIeI9yXdW+u9i+2SSbVnRqGtFII2kkzTyN9o1mRm9n387uO/t2zjszq6M38ySNDDUaUhQl+qbomqbJZvtme1PVVdVd3hdQ8EB6GxF3/4iMQAJIVAFVqEIBuJ8+2ZXIiLhxIzIj7jd+92eElBKFQqFQKBSK1UZb6w4oFAqFQqHYmCiRoVAoFAqF4oqgRIZCoVAoFIorghIZCoVCoVAorghKZCgUCoVCobgiGGvdgYsxmZvsL9cq8YpZiY9mxm5a6/4oFAqFQnEt0BJOnkuEEucA+lN9b691f5pxTYmMydxk/5npc/cenzr54FRhavuJ9OkH6osCa9oxhUKhUCiuXSoAXZGOtweTA6/2Jrr3DLUOvngtCA+x1nkyJnOT/XvP7f+Nvef3fXasMHEjSlAoFAqFQnG5VKK+yMQdvbu/dmP3rp9s79j20lp0Yk1ERsWsGAeG3/7Uk8ef+Q9KWCgUCoVCcUWpRH2RiQ/ueO9/urXv5u/FQ/GZq7XjqyoyKmbFePXU67/7i2NP/e/5WqH/qu1YoVAoFAoFQOU9Qw/86YPXPfCnV0NsXDWR8dbI27/y8L7vf1GJC4VCoVAo1pzKp274+L+5a8sdXwkYAfNK7eSKi4zJ3GT/w3u+9/d1J041LaJQKBQKxTVCV6Rjz6dveegPr5TPxhUVGa+f3vPZb+z7zldR4kKhUCgUimuVyod3fOA/vvu6d/7pals1rojIqJgV4wdv/vivXhl5/XdRAkOhUCgUimuebcmhJz63+9O/1x5rP7daba66yMiWsqmvvvrNh0+kT79/Ndq7ueMGWsOtpMJJAIZaB1ejWYVCoVAo1i3j2QlKtRKlWpnhzAinZs+QrxUuu92oL3Lu9+/53Q+vVo6NVRUZk7nJ/r/85d++cDnOnduSQ9wzeBeDrf20x9pXrW8KhUKhUGxksqUsZ2fO8dboIV4Zef1ymqr8izt+6xM39d7488vt06qJjMsRGFFfhA/ueC+39t1MPBRflf4oFAqFQrFZqZgVjo0f56eHHmOsMHFJTayG0FgVkXGpAsMVF3dtuYOAoVw3FAqFQqFYbd4aeftSxcZlC43LFhkVs2L8t2e+8MpYYWL3Srb78I4P8O7r3qnEhUKhUCgUV4HXT+/hkYM/XZHvRtQXOfdvH/iD+y/VGfSyRcZfP/u3j6/EybMr0sFv7P4s/am+y9qvQqFQKBSKlZEtZfnumz/kwMTBZW8T9UXO/a8P/rvbLiVDqLbSDRr5xcEn/0NDpdSLcnfvHfzRe/61EhgKhUKhUKwB8VCc37v3t/nUDR9f9jb5WqH/q69+8+FL2d8lWzKOT5y49wsvfelplpkH4z1DD/DQrR+9pH0pFIplUr+evetaSpCSWqUC5Qpkc5jlItVShVI+h6wU0EtljJkMlIrYpoms1rCsCjXbxqpaYNeQ0oaqBQIkNiDqL+m8hI4wDIQGaDqaP4huaBi6DoYPYfgJ+EJUUwnMWBTh9xOIRAiGw2jBECIeQwSDGH4/QtNACABE/V/3b4VCsXq8NfI2X379a8tdvfKpGz7+bx7Ycf+XVrKPSxIZFbNi/Odf/JeTy3X0/Bd3/BY39d644v0oFIrlYZXL1DIZZC5HaXICM5/Hmk5Tzmcxc3nK2SyRfJ5AsYiUEoHEkjYSG01KbCGQlgUIhNTQpQWAjQANNGykLZz1sOt71fBEhic66suE7i3TpECgIXUdpO2ICEBqEoEGaFSjYUqRMIFokkAkihGLoqeS6PEIoWSKQKoN4jGMUEgJDoViFTk3M8wXX/7Kcv00Kv/be//9jpX4Z1ySyPjFwSf/w8+OPf4ny1n3Uzd8nAd23L/ifSgUm5oGS4S0baxKhWqxiCyVkLkCpZkZzLFx9LExKtlZqrNpRL6MzwYhLWxpoQkbUwNpS4SUjhXCNT4sgXBWQ3N3LxxrgibBlhLLbWMBWn1bm/nLNQl6vS1L1GVHw/KFTQk5tz+p6aBrCARSQi0SwEgmCUZaEF2diJ5OQi0ptHgcOxLGFwrhC4dBCGUBUShWwPGJE3zhpeUZKLYlh574w3f/wQeW2/aKRcZkbrL/j5/602MsY5pETZEoFMtDWjbStsC2sbI5SuPj2OeHqZwbpjQxTWF2hlqlQLhsETGrCGFhS4FEQwi7Lhw0JBIpHIsEUiAahnF7GeOtwBnom/ZRNNcnjdssFBEwJ1wWbi8Aza7bP4Tz0pa8HUnnSCTIBuEggLKuMxuN4vcFSCZa0DtSxHv6MAYG0Nvb8CXiCL0uWISmhIdC0YQVTJ2sKKx1xSLj71/66sMHJg5+9mLrbUsO8Yfv/oMVta1QbHikdKYlbBtpWZilEqXJSazTZ8mePYkcHkMbn0SaFTSrhiZMED4sYSCEjbAFCLtu6ahPUQi7PrI7ZgoppCcwPMuF+2Av6wO9aBAGYmlhMa/rDSLBFRauqGi0fEicv933rsiA+SLE/VxIR2DIC4qMC/VLOPtEUD90kGAJC90IINpaob+bYF8//oEtBLp7MMIhNMMATXMEiBIeCgW/OPgkPzv2+EXXi/oi5/4/H/x/bl1OMTVjJR04NzN844GJgw8towP89l2/sZKmFYqNiZRY1SpUK1jZAtXRcWoTY8iZGUoj56idP49VLBGQNiGcqY2KDpoATdMAAwHo0gZbILUGfwghARtkw/u6T4WoCw6bucWa9GQJsmHgx31fFyPLGedFQ1sW860kWoN4sC4wdnui5TLHdyGlI04azClS1EPnahW08+cRI8NUeZ2S0LGDQey+LsJ9/WitrQSSKXx9PfiSLYhgEM3n8/xGFIrNxAdveB9HJ49xIn36guvla4X+V0+9/rvLcQJdkch44eTLf8gypkkeuuGjKj24YtMiq1XsUhkzX2Dm/AiFE4fxHzmGPTyGZtZAWGjoaEJDFxK9PgVQpe6PYNcjRJwJhSUG4vqjv6hHesiGzxcohXlTII2miAXvPYvGEoP+ImtH3ZLhun961hJ3mZhvxWhcNv8DlqdsLkCzKRq3cUtQd0R1dmSUS3DiBNVTxwEwpYElbejqIn/rLlp7txPt68cfjyLCYbSAShio2Dx8bven+eOn/vSi67145pU/WI7IWPZ0SbaUTf3HX/zxeS4iMm7uuIHfu/e3l9WmQrERsMpl7GIRKhUqM1PUjh6jcOQEtdPnKIoaujQJmaYTZSGFN75LAZY255dwpbiAblDUEfX/FQwI1QyEYRDs6cG8bhvRHTvw9XShB4KIUAgjHF7r7ioUV5RfHnuB7x/88cVWW5ZvxrItGS+feu1/YhlWjA/sfO9ym1Qo1i1WLkctk6YyMkZu+BzyyFH0qTHMmQwgQLcwECTQHYuE0Ob8FcCZopCg2xfez2qgBMbFcYSYIGKCholtWVTPnYHh06SffBI9HEZ2dGL29xLdso1ofz9GMoGeTCh/DsWG464td/CLY09dLKw18OrZN3531UTG3vP7LurseXPHDSqbp2LDYmYyVKemKJ05hXnoKPkzJ9CmZvFZFprP74SKCg1bCJA6dj2/hGZLz3KxyBGTy54pUKwSdt2qayHBlnWH2nr0Sz4P+SLayeMUn3uCSiCOf+t1+K7bib59CKOjlWBrmxIcig1BwAjwwR3vvag148DEwYeypWzqQunGlyUyJnOT/WOFiYtm01JWDMVGwq7VHMfMmRnKE5OUjx6ldOIEeiYNQuJDYGs2FhpS2vVoDnc6RGI7AaWeiHBnJj3jRUNop+Iao+E7kYAlJEJadf8WgVkpUDt8AA7tp+rzIbdtoWXHTgJ9AxiJOP6uToxYbK16r1BcNsu1ZhwdP/7+O4Z2f2epFZYlMg6PHf0VLjJV0hXpUFYMxYagPHKe/JmzmMdOUDt8CGtsAqlZ6LaFHx0pdCzNdp54nWxVgHDkhJDO9EhD1qslRcQyIzkUa4/rs+HkMPU5H0obkPhrFbSDhykdOkpegIjH8W3fRmDXjQQH+gkPDqD5/WvXeYXiEggYAe7o3c0zp395wfX2j7716csWGfvOH7joVMn7tr9nOU0pFNcktelZCiePUz12nOrRU9TOniZg20jA1gVSq+d+wMaJ+LAcfYHjzOlMgVj11pzwR6EmQzYo88NjJI4Dr+dok5+htneG6p495FNt6DuvI3zdTiJbtxDo6VHhsYp1w31b77qoyLhYWotliYzlVFq9rnP7cppSKK4ZaufHyB45RP70WYyxSapnTmKUi2hS4BOCmtCcxFbYCLth2sPNR2EDiHoxMtngayGdvBT23NqLQjUb/77QsmbrLne9i7W51HvVZvO/Gz5rlrzMNV7ZUqunZrfRpyawpyfJvvAipaEB7O4uRP8AbbtuILRlaHEjCsU1RHusna5IB2OFiQuud25m+Mb+VN/bzZZdVGScmxm+qC/GtuSQyouhWBdI0yRz+DD5fQcwjpygdO4USBMLDV2ALWysuuOm40Rhz4WcOqYLAIQt5jJbetmshOdwIetCg/rHniOGO3DZy1zWbN3lrnexNpd6r9ps/nf9syVdaBoSlDl/CkwcQaIhsU+fwTx7BvuVV5ns7ME30E/0llsI3noLvkhkqVYVijVlV/vOi4mMwOnpM/ddssgYz07cyEX8Ma5r33GxZhSKNUPaNqWREWqHTzD71gHk8FmYnsGycdJ2axLbS82tN/hZOHPuQL2SqKM7XB+LxgyaNnOJp7xIEq8D7gRK/TPZOBA1vK9vIJqsO+/B+wLrLafNxjoj12ybDX9fsM1mfVnlNj2t4X7fcsH3y/z1FyIb/tVsgU9KrPPDVMaGqb61n9qLWwhdt5PUzl0EhgZU8i/FNcW2tqGLTpnMFNNDSy27qMgo1UqJi63Tk+i62CoKxVXHLBYpvL6H7BuvUpuaQp+awajUnIgPIbD1hiHJrQXSUAPEHV7cBFrIuvZoHF0ahIY7AAGL63bI+YNQ40DVWECssU3Z0Kasp9X0anMstV69nYu16c7uNPZltdvEHYgXCiV3PdGwbGGbYv7+l2xTLK9N6ssupc154sE9V4sWLJd6JVvqJ7RUxPf2W5gHDjHZ3o7WEid4w3VE3/lugu3tl7IDhWJV6Yx3XHSd6eL0kv4SFxUZx6ZOvP9i6yRCF9UhCsVVozYzQ+aF58i9uQ99dBJZynupr616MS2HxpGiLhOEZP5zrvDWbDaoNEaOuOs0PtkvXM/7uGG9hU/Fzdp0V2isEXKhfS+rTa5cm43rLoyu8ZbJ5suuuTabtXVJAmMOx4Ii6r9AgRA29tQ4cnKU3KnjzLx5gNjQAPF77yWy64bL25lCcRm0xy4udk/Nnrl3qWUrql2yFCp0VbHWSMsis2cv+UOH4egR5OQYRrWGjUA2jAiyybulR4zljSRyifeXut6VbnO19n2hNi/096WuezltXsp2Fzv2y6VxiskToUIDC4zhs1TPHmNq/xvkBnfg33k90bvuxt/acoV7pVAsZlty6IJF0/K1wpLmjouKjGK1mLq0bikUVx6rVGL22WepvPgiYiqDLJWQSDRh1gWGw3LLmSsUa41A1iOXAjBbpDazj9qhI+Se+iWhG6+j5cEH8Q0OrHU3FZuIsP/S6/VcVGQsJ3xVobja1LJZpl58nupzT8P4NFgWQmhIdCS2k4Gz7kthaxe3SGgNAqSZmVyhuNo4UzsCS4CoVpBTw5R+eZ7qq69h3Hknsfe8i8iWLWvdTYUCgIpZMQJGwFz4+WVPl2xLDl1uEwrF8pCS3OEj5J97muKRI5j5DIGqiY6GrFstpBsR0uD8MM+xcQGNzpoKxbWK8xPVkDaYpQLm889Rfe0VcoODBO65h8Tdd6MFg2vcS8UmJjCRndzZLIz1skXG5ZhRFIplISXVE6cY/8lPqRw4iM8sITAJ2hKEwK6rBCklUtQ9MGQ9zXf9//N0Rx0BdbP0YudApTsU1zSWhV2wKB0+ROnoEbJPPEnLRz9G7I7dCGNVXO0UilVB/RoV1yZSIk2T0sgwuUd+Qm3PPizbRAMsza67ZDZYKaRACBDSVQ26FyniToXYTawZXmjmBawdCsW1h3Cy0UocwXHuHJm//Uuyg4OEPvJxWm7fjdANhK5SmCvWFiUyFNccdrlMce8B8k88RuHMMUTNBGEg9Hq4nxQIdNAaZ0Xcd/WbqmuiaEBbwjyxUFx4uS0UinWCwEKiYZ06R/YLXyDX3UPovvtIvetdGImEqpeiWDOUyFBcM8hyhcKpU0z/7FHsN/ehCYEmLBBOlgtHYKiJDIViIdKz7UkkGvbYCJXvfYORXz5H+KFPkLr1NoxYtJ5NTqG4eiiRoVhzrFyOytlzlF54hfRzT2AIidR8WAicVJegCaGmNRSKZSCQIDWkHUYbmyb3pa9Q2DZE6/veT3DnDnytrcqyobhqKJGhWDPsSoXymbOkf/RDSvv2ABoIiYkOtls3pP6MpuHVlGjmxKlQKOZw3JSc9Pk6NuLkEdInjmK2dxJ/6BMkb7kZI5lc624qNgFKZCiuOtKyKI2Nkn/ueUpPPIFWLaELDUuAJnXmikmAJzJkQ3pvhUJxcdwU+ZJ6/hgQk9PM/v2Xsa7fQeRDHya4dTtGXFXQVlw5lMhQXFXKJ06S3r+f/LNPY0zPgKZRczw60QAN4ZgshBt8qjG/vJZCoVgR86YXbXy2SentwxTeOkLorruI3XMf4Rt2oqly84orgBIZiqtCbTZN9pXXKfzkJ8jMNH4AJLa0keh1h07pTIlIAdj1ED03WmSteq5QbBycwmw+QKBJQfXl15l86XXi999H6FfeT3jb1rXuomKDoUSG4opiFguU9u8n88LLlPbswxD2XCltRN3PwgJcD4yG+NG62PAWKhSKy2LuMqpfh/WZycwLzzB76gixB99H8tZbCPSpopeK1UGJDMUVI7NvH9MvPId8+VV8NYlPE/PLitf/VUYKheJqMl+xOwZDHTE2RuGb36Ty8ktE3/VO4nfdg5FIrFEfFRsFJTIUq05tJk36qZ+Tffo5rEwODajqjm+FppJ2KxTXFsIV+hpCSKzTJ8mePEXprbdIfvijRHbtWuMOKtYzSmQoVg2rXCb99LOU9uzDOrgXPzo1IeamR6RUxcgUimsV4f7PSdFffWMv48Oj6HffRfs73kF427Y17qBiPaJEhmJVKB45wuxPH6Oy53U0BAIfNhIh6pk6lfVCobi2kXP/WMLxmWJinNJPf8L466+S/NBHSD74HpXIS7EilMhQXBZ2tcrszx8j9/QziIkJfGjYEizd8tJzalJcsEiZQqG49nAfC4K2DefPk/nmNygePkTqIx8ltGVoLbumWEcokaG4ZLIvvUTmyacwjx9Hr5pYuuYEgwgbpO1VOKWx3ojK1qlQXPM0PgvYAEJHVCpUXn6BqROn8d9xN60f/xWnHopCcQGUyFCsGGmaTH77O5ReeB6ZzdRzXGhI28YCBDZCCrDcNFp2PUhVoVCsCxqeBETD35oEa+w8pZ89ysjxo7T9+meJ7Ni+Fj1UrBOUyFCsiNz+/Uz9+EfYR44ipESrx9lbWEgh5t2Q1MyIQrGxkABCotlV5JG3mfqLvyD3nvfS9uEPYUTCa909xTWIEhmKZWEVS4x/52HMF1+GQg4pQK9nDbSQ2G6dhLp3ukKh2LjYbpK82VlKP/kRw/sOkPr8Z4nfqMJdFfNRIkNxUfKHD5H56jcxz54F2wLNCUWVePk4PcdOhUKxeZDYUKvAiSOM/9mfkfnIh+n71Y8hfL617priGkGJDMWSWNksk48/TuHJJzDSaaTQnJwXEi+p1nxtoULbFIrNiNQkRjmP+cgPGTt8iPhnfo3Ijh0glFVzs6NEhmIxUlI+cYLJr3yF6qlzgEQKiRQmFrqT90I6Dp5umOpctVRQ8SMKxSajbtkUdo3qoQOM/uejtH/6M8Q//AGE37/WvVOsIUpkKOZhV6tk33yTzFe/CVPTGEJiC6teDVXOdzXHfe+KC7Hgc4VCsVmo5wrFlqCZVaa/+TDlmRlaf+0TGNGosmpsUpTIUHiUz59n8gffo/jqKwRMia1rjrUCiURDCg0Na05rIEAazHf2VLVJFIrNiqznxBECpG5S/MVPKR88QMunfo3YnbsRhhpyNhtqEl2BtCyyh44y8ld/Re2FFwmYFmAjMGm0VAhpO/VHWBhJ4goMGyUwFAoF1A0XmoYcPsv0X/05kz97DKtQWOtuKa4ySlZucqx8nuwvX2Ly298lUEl7jp3SdbHw3iwUEAsFRmOsiQpj3ci48nKhnFz4jcsmnyk2H1LTkFKS+cbXqZ0+TsfnPo+vo3Otu6W4SiiRsYmpnj3HxCM/pPr8s2i6Rk3X0F2dYMslRohmH9pNPlMWjbVELPieFtaMcf/UJAjpVN2UdbumNlf7e976bpiyxPH1nfcNy7kpd1H/OUgBdj0CSc6tBrIelSQdv2Ehwa5r2XkePrazksTpmwCwHfOr2wfF+kAgMISk+uKLTJw4SfI3fpPw7bepUNdNgBIZm5TC8ROk//ZLWGfPgKZ7hczmIRe+WXhXt5n/TKvu+mvJPDtTXSeIehZWv6ahC4EUAq3+NWlINNsRGZaQWBpgS2wbpCc0nH+FlMi6ijBFXRTYc3vUAKP++xGaRLcdgaEDaAIhnCJ5om7kksIREZoUYAtMYVMVEtu2kbI+t+/WvBHOW1eQ1D9SMnZd4X5bGtXxSc7/1Rdo+eyvkfrgB9ACgTXtmeLKokTGJsMsFsk/9yKTj/8MRs+j6c6NXq8PAHM37uUKBnWrXwsWyjpNCHxCYAgNTdcQmsAUBsICW5pUNahoGhVdx9Z1irpGVgNbM8AIUPHp2LqGphuUdQ0L572maQjNyY8i3X+b9EdIiWZLsCXStpC2xLZNNCkJWBbCstAsG59VQ1QrCGwSEiIVG8O20KSFJiWRmsAydCp+QcA08VsmppSYto1lSyQSGydseuHknWK9YKObVTLf+jbmTIbUe9+Dv793rTuluEIokbGJsPIFzn3nG9R+/gy6JuvmacccrTeMWko2XBu4UwsuWl1EGLqGIQS6Y6oAwLQlVSAb8FPQDYqBALOGH9MfxA4HsKMR9GiUUDRKMBbHCIeIRcL4A0ECgQC6YTjWBk1D1Nts/HfhZ82wbWeeRErpRRm4792/a9Uq5XIZs1qlUiiSKxQoF3KUsllksUQgU8SulECaJMslEpUKwWqNSM0kKiSa5lpBHLdkS0pMy8K06+IDMWfuUFyjCAQSiUnusR9hHdhH8l/+HpFdKiX5RkSJjE1CeXiEmR/8iMpLT6MbmlONXTpmbm+OGzmvKrvi6uP6KiAkuhD4dZ2ArmNoGpaUFC2LvKFT1TSqukEm6KcYDGJFYtQSCXytKcKJJOFEkt5YlEA4jN/vx6iHDtq2jW3bmKaJZVnYtk21VsOuVJz9LxAI8/p2kR9HowBpJlQANE3DHwgQDIUwWlvRdR1N09DqzoGWZVGtVKgWi+SzWSazWYrpNPbsLL5MBpHLEqnWiFRrhGpVQqZJCIO438AnJTXbpoJJ1bKwGv2QV3AciquEBE1AZfgUo1/6Eh2f+zzxO+9Q+TQ2GEpkbAJmX32DzA8foXbiIIF6avB5l7Ga4F4zpOcBiSMqNJ2QYTiOuJZNUcJkOMhsOEIuHMFuSWK0tqKFQvjDYcLJJC3xOKFwGF/dic62bWq1GpZlUalUKJVKdV+Hpb/kC1koVrLO3HE1FyqmaV5wO13X0XUdIxSiLRbDGBxE0zRv21KxSG42zWw2w0ShgF0qUZ2dxZhNkyzkSebzxEoW8YCfIDqmZVOwHdHhOqF6Bo9lH43iSuDk9RMgDLSREab/9gtUTn6E1Cd+FT0YXOvuKVYJJTI2OLMvvsTsV76Bnp7AJ0DWU6O4FmXLiyhQXA0cI4Uzyum6RsjQCWk6GoKygDyCkWCATDxBIRZD6+4m3NFOrLWVzmQSfyCAruueaDBNk1qtRqFQwLKsucgNANxpDjdB0toPqxfrg2tpqVari7bTdR2fz0dbdxedfb3e1I5t25SKRQrZLOOTk5wYH4epKWKzs7TmS8TLZaIBi4StYWJTtEwqNYta3cfjGjgtmxtZjz8qFsn/8LuUSxl6fu3z6LHYWvdMsQookbFBsUtlJh/9KdmfPkogV8LWHLd+W0iEaEin5YYlKl+MK0ajk2ZA0wj7Avh0g4pZY8xvMBOJkAuFqSST+Ls6iXV3k0ql6IvFPOtErVajWq1SKpc934dGvMFbLAxeXZ/fajMxYlmWY80oleZ9rus6hmEQT6Vo7eyEm26iWq1SyGbJzM4yfH6M2sQogXSWWLFEW6lIm6iS1A1qtkmhWqFm1T06lOJYIwQSgRCS2mO/YHJshpbPf47AlqG17pjiMlEiYwNi5nJMf+UbFF54GkNKQHf8LQTIusBwLReqpNmVo37qCaIRMXT8uk7GlBwLR5ltT1FrayM8OEgk1UJ3KESw7j8hpaRWq1GpVCgUCg2pmi/ufLnRaXbstm1TqVSo1P1KhBAYhkEkHieRSiG2b6daqVAul6kWS4ylZzl27hy+8XFaZ2bpnU3TETIwLYt8rYJpOdW+rEWCTXGlkVIgpEZp7x5qE1PE/6ffIXHDjWvdLcVloETGBqM8Nsb4V7+G9doeNGwQGrZuOXkPnDRLiiuEGw2iC42ozyBg6Fi2pGjaHE0kGe/ugu4ekv0DdKVaCIRCaJqGbdtYlkWtVqPcYKlQomL5LDxH7jQSOM6muq4TCAYJRyKkOjuwd+ygWi6TT6fZd+4c5sgwqfPj9M7OEjM0NF3HrlUomSa1emI6Zw+yHhuhuFLIepKX2vlhJv72C+if/02i99231t1SXCJKZGwgyocOM/aNr2EeO4Iu5uJGpJxzspJafQBTd8nLYG5As4WTCVMKScinE9MDmBacDQYY6W6n2tGJv62NZHc3W1Ip9HqUh2malMtlLGsuXftaioqFDpoLp2RWEpGxMMqkWdTJ1cDdV6Pvios7xdLS2UlLZyfmbbdRSGc4fH6E8vQMxuQE3aPn6S+VSehQqlYp2BbYEiE1R2h4ZkCpxPsVQGJjjE0y9ZWvYk7Pkvz4R9e6S4pLQImMDULu5VfIfvUb2NOj6NJ95mrIA40ELC/ls+LSmBu4nEi7sKYTMAxMTWdC8/FWVzu1rVtp2b6VvmQSzTDmh2eWSvMGbOFm1bxCuPtqFA2NYaqNx+WGkrrTDe57N7qjMYdGM1xnVCll0/eWZXmfCSHmTQM1y8nh/r1aNLblOpdWKhXP0hFLtRBLtTh9NU1yszO8duYs+vHjDI5N0lWz0GWNaq1KSUpq9ePQVEjsFaH+i8DOzJL5p4epZkp0/OavrXGvFCtFiYwNwMwLzzP75X9Ay2eQQiIuECuinrguHbcEXBBBMOhDShgXgtG2FkpbttFyww10tbaiNeSkqNVqTR01V7VfDUKiUTy4g6cQgkAg4IkEN2+Gz+fDMAwMw/BCRxu3aWZZudCg37jvxiRcrtBw83K44bWudcF9uULEjTBx3zf242JCZ6W4Ysc0TUzT9I5b0zQSbe0k2zswb7mFmdlZTh45RvT4Udqnp2m3beK6Rq5aw3Snt1atV4pGBBJZKlP+8Y85Xc4x+Hv/HKGpeLj1ghIZ6xhpmkz96FHyP/whopLDVnMgVwS3uH1QCPzBANMIjnR0UtyxncS2raQSCQLBIFZ9kK9WqwssBavrXttoJQBHTPh8Pvx+P4FAgGDQyeLp9/vx+XyecPB6cwHLwdVgYQ6NhQnAGsNYXWtDqVTy/rYsywnXZc66slrH0CiKYC50NtnRQUtHB7W772Q2m2H4zDkCh4/TNzJJSq9Qq5UpemJDXYerjRRQkyXMpx7j/MwI3f/zH6FFo2vdLcUyUCJjnWKXK8w+/H3yj/4UYZvYwvZubbYy364YrWGQcsN5JeBDENJ1hNA4GwwxMTSI75ab6BjagqZpmPXBrnpRi8WlfScLrRO6rhMIBAgEAkQiEUKhkPf3wgH3WnUYXW7/QqHQIiFimibVemryYrFIqVSiUqksshi55+Fyz4G7T7dNTdeJp1pJtXfAHe/g/PAwR/a+Sd+JM3QXK6BZlM0ytm0hpPCKyikuHwEYlk35jb2c/y9/Rtf//d9hJBNr3S3FRVAiYx1iFYrMfP8HzD76A6ewmffsJBCq/vUl45QPdyqHGkIjaBiMBIIcSSapDA7SsXs3Q6kUtbppnbqPwfJaXsZaC3wl/H4/wWAQv9/vWSkikQiBQOCaFxKrxcLjdEVWLBbzzpWb2bRYLHqhrK4Qacwwermiw53OcfcJkOzupnNoiEI2x2v79qOfOkXb2Bg91QLCrFGzwULUnYMvedcK6lWFJYBO7chBJv/iL0n9q98n0NW5xj1TXAglMtYZtXSayS9/ldrzz6IZNrbmFADQbB1hO/OUUk2brBgB2LoA3Yem+5jUDUa3DBK/7z4G+/qQUlIuVyiVyjS41F4WjU/prrNlMBgkFosRjUYJN6QKVyzGFQyuX0kkEvGWOd9XmVKpRD6fJ5/PU6lUPOfTyxVp7nfn+nJoPoOt99+L7z3vIj01xd5XXiZ1+G16ijV8lRo2JnO/G6U2Lhcpbcpv72fsj/+U9j/6nwlvHVrrLimWQImMdUQtk2X0m9+m9tLTWIZAr5e7Fu6/9feghMZy0RBoQsPnA9sXYtQXY/SG7fTefw87W1spFovk83lgztv9Us7sQqdI17kwGAwSiUSIx+NEo1ElKlYJIQShUIhQKEQqlQKgXC6Ty+XI5/MUCgUqlUpTS9SlCA/btimVShSLRfzhMLs+9nGKD76Xt197g+Te/XRn0/jNMrYlsaWGxYVruCgujpACbfgMU3/+F3T+h/+VQF/PWndJ0QQlMtYJ1ekZxr71HcxnHkNoc2m1hJPGE4HmOCjO2RQVS+JUONWFjiYMCoEgR1rjlK+/noHbb+fm1lZyuRyzs7PzBhzZ8P9l7WWBpULXdUKhkDflEYlECIfDG37K41ohGAwSDAZpb29HSkk+n6dYLFIulykUCl4itIURLStBCIFpmqTTaQzDYNeD76Z4524OHXgL+9AheofP01G10C0bE4mlcmxcMraQVHUJY+cY/cu/pv1f/h6RHdtUFddrDCUy1gG12TSn/9t/Qz90gJqhEbRFvRCWk4IXREMhEhtljl0aQ+gYmkATNulAiNNd3UTe9QA7rt8FQDabZXp6+pLm7xf6VAgh8Pl83hRIMpkkHA6v6vEoLg0hBLFYjFhDEa5CoUA6nfbEhys43PUb/11O+5ZlMTs7i2EY7LznHrT77mPk1CneeOkVBo8doa1awY/Asm1qVzjMeWMjsE4dYfS//hc6//W/JXbbTWvdIUUDSmRc41Snppn41j/he3s/tgZ+y3HuFNTT7yLq5ZJdgdFYjksBTj0EnyYJ6Dqm7ScdjHG6K07wnrvYecut2LbNzMyMU5GTledhaBQXbiZJV1gkEgklLNYJkUjE8+vI5/Nks1ny+TylUsnL6+H+NrRl5mlwxUY6nUYIQXtvL/2/+RucOXyYkRdfpH94mJZqhYglqdo1qraTtlyxUjSM2QyZv/sHjH//bwhtGVzrDinqKJFxDVMaHWf8i1/GePM1LAMEhuN/IQUIzSkZLt1S3kpcNENDw2f4MagxFotxest2Erfews7rdqDrOul02nHc07T6zX1lkSBSSnRdJxgMEo1GicfjhMNhgsHglT0wxRUlGo0SredhKBaLFItFT3S4uTpWEibrruP6g3Rv3Ypv507OnzrFmTf30XriFP3pGaJSUjFNarIhwZdUV/VyEFIiz59m4s//gpbf+W3iu29b6y4pUCLjmsXM5Rn/q/+O7+ABqgYIqdf9MJySyNRDV+cEhuv6qZAIDAFRXcfWNEbDYc52dRF/1/2844YbqFQqZLNZb6BY7lPpwiRNPp+PSCRCIpEgHo8rYbFBCYfDhMNh2traKJVKZLNZ0uk0xWLRy1S63Gyk7vJsNutYNvr6GNq5k7HhEd54+hkGTp+lO5MmrJkULQvLVJaN5WILiS0EYuQcs3/zPxB/9G+J3XT9Wndr06NExjVIeWycie/9AO3gm1R1DdARtgBNA2zmFyDRVJ12GmSWhJDPwBCCsVCAk11d+Hfv5rpbb0UIwfj4BFLaK/K5cPMjuCm53UEnFosRj8eV4+Ymwo1YaW9vJ5vNksvlvKRgbqZXNy37hWgUG9lslmgizm2f/yxnDh/h3Guv0TN8jr50Ab8uKZpVTKRyaFwGEkds+GYnmfna19H/xe8Svm77WndrU6NExjVGZXKK0S/+HeabryIMzSl2Jl3fC9kgMNwbjq1uPvUCcEGfTkz3kdZ13ujvI3zXnWzfuYtAIEA2m6VWq61IXDRGGjRaLKLR6Lw03YrNh6ZpJJNJkskklmWRz+fJZDJkMhlK9SJ4bkTRhRDCsVPkczlKpRJ9O7YzdP0uRk6e5LWXXmfr0eNsq+QpUSVjOrcAp7ayohlumLklBNrxI0z/5V8j/+D3idx0wxr3bPOiRMY1hFUuM/mP30LseRV0Uc/kWR8QhXRS+cj6tIibPlBY1DNkrFGvrz7zqmlKm4CmE/eHyNlV9rW1kL35Vrbddx/BUMjLi7BccdGY1VHXdeLxOMlkUjlwKpZE13USiQSJRMLz3ZidnaVYLHoZRy9k3ZCA0DTPAdkwDDoHBujdupXje/Yy9vJLbJkYpdWEcq1GybK8K95uUlFXATYSG4k+OszUl7+C/u//iGBf71p3a1OiRMY1Qm1mltGHv0P1pWewda2eLhzmnlnkEu83J7YEISSJYBAbgzdbWpi9/jp6br+Nvu5u8vn8ikJRG2tURCIRWlpaCIfDXk4LhWI5uFNpLS0tFItFCoWCJzjgwmIDHAuJKzZ8Ph9bbr2F/NYtHHn7bfQ393PjmXO0BA2ylarjH7R8X+VNhXNaBLYUiDMnGf+b/0HHH/w+oYG+te7apkOJjGsAWakw/bVvUnn2cdBAswWiLjKczJ1N7iJik8bVSxDYRP0GcX+UU5qP4zdeR+e9d3NLTw/VatUJR21IHX0h3JLi4EQUpFIpEonEvBTVCsVKcYvWtbS00NLSQjabZWpqikKhADhi40IOx5qmYVkWMzMzhEIhbrjvPjLXX8/+PXuJ73mDW9I5bNtixqrUQ6/dQLOFbqKb1yFcQN0CLLAPHmTiL75Az//67/B1da111zYVSmSsNVIy+08/ovjs09gaaF4ODADJ8i2hG/hxRjjRNFJKDE2QCBhkgH1dHdi7d7Prdsepc3Z21vOhuJjAcMWFYRjedEgymSQUCl2dY1JsGtz8G4lEgkwm4yX8qtVqFxUbQghKpRKlUolwOMz1732QkS1DPPvcc/SfGaW3KqjYJSq2jRQCTcq6rHAeTqQ3seJeDxv4PrEkzsSzdfIQM//4j7T963+NrsrEXzWUyFhLpGTi298m8+gjWFjotqDxdrOCBNar37drCGGBJSDqM4j4dfa0JCjf9y56b7gRfzBIqVRatlOnOy2i6zqtra20t7cTDofx+/1X6WgUmxV3KiWVSlEsFpmenmZ2dpZarYZhGEv+dt3P3RTobT09xD/5SWZGxxh+6hnuOH2GkL/KbLWCxKhXFahbQIVAZQGW2JojNYqvvsyYXaXrj/4f6OqB4qqgRMYaMvrTR0n/5AdolSK61Jw8GKLxqUPhOsWlIgGylsneW26n7Z3309veQbVapVAozKs1sWQ7dXGhaRotLS10dnYSiURUQTLFVcedSolGo7S1tTE2NkYmk8G2bXw+3wXFhltzJRAI0LN1C4W2Vl574w16977BznyFdDFPSVqgi3rZAVG3alg4cSmb994iqTt2v/IqY1/6Or3/5l+qyLyrgBIZa8TMSy9QePif8Jdq2OhoXlo/9aN3sSUEfRo+TbCnqwvzjjvYestt2FKSy+WWNTXiRotomkZHR4cXhqrEhWKt8fl8JJNJIpEIhUKBqakpZmdnsSzrgg6iQgiq1SrVahV/IMC2Bx5gcssWnn3+FW4/fIiUVmWiWkHDUIm8FiCFoKYZmM88yVh3iq7PfHqtu7ThUSJjDci98hrFL3wRvZBBoqMLgXRDURv8tISsT4Q0TKeKhlpo7jI39bC8yHpr3mZ9Vmdem41/N6ABsZCPCd3g3APvoe/uu9EMg0q1uuypEdehM5VK0dHRofJbKK5JXLERjUbp6upibGzMi4y62O+1Wq2iaRqtPT20fuYTvH3kegLPPMHN6VmyZZOqtHFtGcKefy1e7BpmOesud73ltLnSfV9g3Wb3r/ldLlP6/j+RDsZIfvyDS59gxWWjRMZVpnT8BCNf/QcCuTRSFyDtebUJBMxz9vQS78i5ZaJxWcPF07h8qfWuhTZh6aRCwgbDEISEwcFUO+b7P8DWXbuoVGtUKpVlTY3Yto2UklAoRHd3Ny0tLUpcKK55DMMgGo2yZcsWWlpaGBsbo1AoXDT1vW3bVCoVDMOg54YbyHR08OLjj3PL8cNEpU2xKhHCmSppvBabXcO2XOAm2mTdhdfwUuutuM26SLiibeIuE4hSlfTXvopIRUjcfz+KK4MSGVeR6sgo01/6Cv7zw9i687PXkI7nMzSErTrraw1P/oI5tS8F6I1PJe6y+n5E3ZJgCWfQbtoml96mttw2AX2BJUPDmR1216OhHSEkUcPPVMjPm/ffT98dd+IPhigUSyBt5zxoYkk/V3dqJBwO09vbSzweX1aKZ4XiWkLXdVKpFMlkkkwmw9jYGPl8/qK1UUzTxLIsIskk8c99joOHDxN48glunU2Tq1rUsJnvBDp3IbnXsNv8Qguke703XreuWcBbtsT9a9ltNrFwrrRNt19LtWnXl8v6h1o1T/pvv4je1UF0244lz63i0lEi4yphzswy+dVvYh55GxBO9k4c1W03pN2SDWGr7u2gbuyYU+myYZmcv0y67dTfzPubpbdz22QZbdrMf2JYsk0Wt+keqASkEJ71QxOCoN/HoVQ71oc/zI5t26hUKvV6ELa33VK44ajd3d10dHQonwvFusadKnHztkxNTXH+/Hmq1SqGsfRtW0pJrVbD1nUGbrqJXE8Pr/zscW48cYqwlaNUs0EYznVs2/OmKhutqAvvG7Dg+r7Qsib3hWW1Kedf4pfS5lL3xLlG8T6w3bf5HJN//hcE/uN/xNfevtSpVVwiyys/qbgsrEKRqW/9E7WXn8d25UWDCGikMcWWXPD3xZbJJu9X2ubC7a5km86DhiSi69TCfl7feQPx3/s9OoaGKBaLWJY1JzCWwJ0+SaVS7Nq1i97eXiUwFBsKXdfp7Oxk586ddHR0oOu6Vw14Kdx6KsFYjKHf/DxvvvedTAQjRP0BdCHqDzba3CDN0veTlS7jAsuulTYbP5NI5PlzTPzZn2Kl0yhWFyUyrjDSsph+6hnyj//Mce5sUNyb3YivCUHUb3CyLcXpT32WHZ//PELXqVQqF93WTaYVj8fZtWsX27dvV7VFFBuacDjMli1buO6664jH48h63ZKlape4USjFYpGd73435d//V7zZvxVd8+MXGkJKb/phsyOkRvXw25z/0peQ9fICitVBiYwriLRt8q+8Sv7vv4whpaOaxea+qiUCGx2fLokHwrzV0Yf+W/+coZtuolgsYteLPy25vZRePoGBgQF27NihxIViUxEOh9m5cyeDg4P4/f4LCg1wBHkmkyHe2srWf/Uv2XPbzdgBP2Gfjr7pH3UcHGuOgfnyK8z+8BElNFYR5ZNxBcnse4vMX38BW9pY2gKnqU2IBDRhEzV8jIfC7L/lFgbe9wF0v49cLnfhbes3UsMwaGlpoauri2AweHU6rlBcg7S3txOLxZiYmGB2dpZqtbpkaLcQglwuRzAYZNevfZqjWwYJPvtLdkxNIqRNzbI3rdxoFGi2ZTHz7YeR0SitH/ogKlnX5aMsGVeI2uQ4U//4DcrFEhLNiayQ2qa9kEHiF4JowM/5eIixX32I23/9NxCGTqlUvuCWbkhqJBJh27ZtDA0NKYGhUADBYJCBgQG2b99ONBr1LH3N0DSNSqVKLpfj+nc+QPJf/T57h4bAHyRwAWfSzYRAolUrzP79FymfG17r7mwIlMi4ApiFPFMPfwP95BHQLKSwEWhObLbcXKfcfUYIGQb+gJ8D3f0Uf+Ofs233bsbHx6lUKkuG5TVOjbS3t7Njxw7i8fjV67xCsU6IRCJcd911dHd3YxgGlmU1FRtu1YLRsVEC4TCDv/M77L9hF4VAiIjhcwaEZtnxNhFSCOxihZlHfoxZKK51d9Y9Sr6uMtK0OP/d71B78inQtbpjlSMsvHhu5gbfjYwbRhb2G6SDYY7s2MG2hx4iHI0yO5sGli7H7t4g4/E4XV1dJBKJq9VthWJdous6fX19xONxxsfHyWQyXoryeUjQNZ1sNkskEuGm3/wt3n7qafqef4HOQp5yzcKWm9gnQQh0n5/co49Rnpml/4/+EJ+6/1wym+ux+iqQe+11rO9+D0064WFOMhkvA0b9v42PBHShEQ+EmIhEOf7AA9z6W7+FLxBgenp6gXPK3Hu3kJmb82Lr1q1KYCgUKyAej7N9+3YvnLtWqzW1amiaRrFYJJPJcOsHP0D61z7Fwc5uIgGdgG6DEPVMoZsNCdLGF9AQb7zM1I++u9YdWtdsxl/QFaN0+iznv/FNJBqWnMtWt9mcPSXg13XiPj/HWlKMffTj3PbhD5PL5chms06K5CYmWdu2sW2bWCzG0NAQfX19KueFQnEJCCHo7u5my5YtJJNJhBBeLZ+F61mWxfj4OIM33EDs1z/L/t4eNF0nYPiQbGKHUCS2rpF/5lnye/asdXfWLUpkrBLV2TTn/+ZvMc6cxtI2b0llG0FA1/EbBm8MDSJ+/TPseMduJicnKZfLDTUY5isvtwx7T08P27Zto6Wl5ep3XqHYYCSTSc+qYRgGtVqt6XpCCKanp0m2tdH2G7/Ja7vvpRgIETP8cw9LV7fr1wZSYEzNMvs/vkjxxKm17s26RPlkrBKj3/gW9qEjCK2xaE/jZblxzRmNfhUxnx8NjVd37mDrZz9LKBJhZmYGKWXTIk9uGuRQKERfXx9tbW1Xs+sKxYbHnXoMBoOcP3+eXC6HruuLrkchhOenseuTD3EkmeDm516iU+iM1kqAhmDp6JWNio2gMjbK5Fe/Qc//8of4ksm17tK6QlkyVoGpF16k+tQv0OdV5Nl8RA2DnE/n5Tt20//JT+IPBpmamgJo6uDpFnTq6upi+/btSmAoFFeQlpYWtm7dSnd3N+BcfwsRQlAoFKhUKux633t568EHOBlP0BaIIOq5ajZbwUFZv6ebb77G9M9+htxkIutyUSLjMikcP8XMl7+GbdWwhcVGtlgshSYlMcNgRg9w/IMf4oZPfxJfIMDs7OySJaprtRq6rtPf38/AwADRaPQq91qh2HyEQiH6+/sZGhoiEAhgmuaibKFuOvJCocCuB9/N1Cc/zoGWFO3BEEIKbHvxPW6jyw6JQEqN7E9/Sv6Vl9e6O+sKNV1yGdTyeUb/4UtoE8NIbbMEps4/SoFN3B9mNBTm1Lveyc3veTfZbJZqtXrB6ZFoNEpfX5/nlKZQKK4OmqbR3t5OKBTizJkz5HI5DMOYdx26DqH5fJ7+XbsY1nUO/ODH3ATMVPJUbXfyxKnyKOqvjZoGSCBBCrRMhtl//CpGbw+hgaG17ta6YIP+JK4Ok9/5Dva+vU7uC7fe8GZ41UNybWyi/gBvt7Yw9ulPsu3++0in015644VYloVlWbS3t3vOnUpgKBRrQzQaZdu2bfT19QHz02u7WJZFJpOhY9Bx4n5xaJCQL0pA82FLiZASYdfvC+IauDdd0fuehZRgnjvHxDe/jqkqti4LJTIukfSePRSffALD3qyDpKQtGOF8pBU+/WkGbrieWrWKZZpL+l8IIejr62NoaEgVNVMorgGCwSA9PT309/d7OWoWYts2pVKJtt5euj/7GV4fGqA15COsaUjbGUSkgI1/K3Ss1bYQ1F5+nfwbr691h9YFSmRcAqWREaa+9o/IdBprkyWrcZ91Yr4QR1tS5D7zEG19feQLeScOf4HAcKdH/H4/Q0NDXtpjhUJxbaBpGh0dHWzdupVwOLxIaAghkFKSzWYJhMN0fOoTPLN1G5oRIqL5AInQNrzC8JAAZo2pr/0D2aNH1ro71zyba4RcBexqlYmvfBOOnYG6wHArhG6Gl7BtWn1+9g0MYPzOb9M6NEi5XEbaizOZSimxLItEIsHOnTtpa2tb0hFUoVCsHUIIWltb2blzJ8lkcsl8GoVCgWAkwo7f/C1eu3M3+P1EjABOuL5Y8/vTlX65SEBMzTD2hf9OeWz86nxJ6xR1x18hs8+/RHn/XqTYfLn9NSmIBeMcbO8k+cmHCMfjVKvVeRefiyswkskk27ZtIxwOK/8LheIax+/3s3XrVtra2qjVak2v7UqlgmXb7PjYR3jj9lspGz58+mYbSgQWGsbxI6SffRZ7CVGmUCJjRRSOnyT9vR8gihk2SSBJHaeGbCjg543tWzH+2W8QSaUcC8YSAgPwsnf6/f6r3WGFQnGJ+Hw+tmzZwpYtWxBCNE2+VavVCIXD6KkWSrpESJvNdVMUCAQ2GoWHv0f2kJo2WQo1Ob5MrGKJyW/9E5w86Zn8xSa5qDQpiYdjvDrUS8/nPo0/GKRUKjVd103W09fXR0dHh7JeKBTrEF3X6ejoQNd1Tp8+Pa+aazAYxGcYHPj+I+x66TUCtQply6rnypgTJNq8eFY3Km1jIBFIIZACfOUcw3/3N4T+438i0Nq61l275lCWjOUgJbOP/hzz5WeRuonwcsBdcKMNsAw0TRAOBdnb3UHLQw/hCwSWFBi2bXsJtjo7O5XAUCjWMUII2tra2LZtG4FAACkl4XCYUi7Hia//I7c99wzhcp6qZXqVpgE06bzm2DjiwkUg0aSNJiWWkPhOHWfm29/BLpfXumvXHMqSsQzyx0+Qf/ppNCGxcWbjQNalRn0gdUutevVKZMOy+t9Nl9Xfe9fhwmVr0aazngEE/HFe3Xkdg5/9NXyBAMVisek5sm2bUCjkeagrFIqNQUtLC5FIhLHRUY7u2UPgRz/hrswUGVnF9p5THQuGc3sRTsoM934i3OqvgmbVl9c7EkHA1Cn+9PsU776L6B3vWOsuXVMoS8ZFsKs1pn78M8SJ04i6+W9xHIXz6XzFfqHiaAuX2RdYtlSbC+dJl7vdxdp0lmsIoqEQB7vb6fzYRwiGwxQKBZph2zbBYJAtW7YogaFQbEB8mkb20GG6v/0trp/NMFkV2LZYYNF17ivuc4xodpvcgAgJtgYaOjPf/hq1XG6tu3RNoUTGBZCmyfRjT2I9+QxSWEhsJ5GuEHUF78xRIhoGfO+91vB3fWD3CqhpTj6Jxu28gV9b0M7CNmXD3xdqs1FMLNxu6TYlYGiCQCDI6x29BD/9ayTb2shms4umP6R0KjImEgm2b99OJBJZ6lQqFIp1Si2TIf+DH9L+t39HolRkppbHL2voNmi2RJMSTQo0qSFw/BScV/3eJwVIbUNaMaB+eAik0CgfOcrIDx7BbpLUbLOipksuQO74CXIPP4xmFbF0gZDOS3oTjq6AaDADisZldeox5M5A7ky4zNvONSuuqE1XHMhltslF25RIdKGR8Md4s7ODzt/9HRKtrUxPTy8pMJLJJNu3b1f5LxSKDUjp7DlGv/5VrOdeQKs/U+lCc1KJs1zdsDHFRTM0BDM//A6R63fSfueda92dawI1MiyBtCyyL76ENj6KpUts4VgwEBqaBCFsRzR4k4+N/hn23KtxWd3XwbNu1OcvHVEgG7a7UJuNy7TVaxPQhCDuC3I6ksD+0IdoaW+/oMBwS0crgaFQbDyKRw8z/Of/DfHL5xDG3FSIe1vboIaJS0YCwpakSjVyX/0a0rIuus1mQFkyliD92usUHnsUw5irNAj1rG9NC67Kef8sxrUyLMqLeYHtlrHsctsUjl+4hiCo+TkYT1F96CNsu/XmJQWGZVl0dHQwMDDghbUpFIqNgVUokH76STI/+iFieBhT0+oFIDdinMjqYwmBdeo4E7/4OZ0f/uhad2fNUSKjCZXxCaa+9g20bBopdERdsst5A/RyHTsXLl9q2YW2uXJtOv9B3K8zGghTfuhXufneuzl//jzStmmsRWLbNrZt097ezuDgoLJgKBQbDDObZeKrXyX/s0cxbAlizpdinsO7UhsXRLMh+5W/I37TzYT6B9a6O2uKGiUWYNdqzD7+OL7jRxBigz+l1/VDxDDI+cKcuu8edr7jdsbGxpwsf0pgKBSbhsLxY4x/4W8o/eQRNFvWq6oujCBRLBc9V2Lq77+MXa2sdVfWFGXJWED+/DCzjz5CEG1Dzj3OOxQp8BsaWb+fI+96Nzd++ENkMhmvLLuLbdtYlkVXVxf9/f1qikSh2GDMPPsssw9/C06cwCv8uMZ9Wu9YQke+8hqzzz1H6/s/sNbdWTOUyGhAWibjj/4E0jlqnuPFxpqIbBQPhqYRMILsufEm3vGJXyWdTlMul+dZKVyB0dHRQV9fnxIYCsUGQpomEz/4CfmHv43IzWLXrbf2BrvvrQ0WltSY/dFjtLzr3WibtIaTsnm7SEn+pVeQjz6OIQHNashxsfHQEAQNnX2DA/R/9KPk83lKpdIigWHbNp2dnQwMDGAYSpMqFBuFytgEw3/535n5+79H5rJYmsASYG3clBZXH2FjnjrE2A9/sNY9WTPUqFGnMjXJ1Be/hFEFU7PnUuHKjfnkHjT8HOnqo/XXP08kHieXzc4TGG4USXt7O/39/UpgKBQbiOyeN5n5xsNU3noTDcf/QtCQRmdNe7exEFaV7I8fIXbzrcSu37XW3bnqKEsGgJRkn38Zc2IcS6viTJFodYHR/HITcv5ruQhWZ7vl3gQat9Hqsz9hXWciFMD4zKdItLaSz+cXOXKapklra6uyYCgUG4z0E08y9hd/jv3W62hCognHx9vNHezea1wW3uvW6v61Xu6zC/cnpIYxOU36Ww8vvyMbCDV6ALWpKdLf+y4gkNRLFtctGHLBL8ar+bOgjYv9jhdeuCvZ7lL3t3A7icCv6ZR0jfMf+iDXDQ6QTqcXtVer1Whra2NwcBCfz3eRPSkUivVAbTbN9Le/Q+mJXyAKeWyhefcHLydwo7iApveeRq7m/Wu52631fXbxdo6ZqHL4LdKvvkzyrnsu0urGQlkypGTmB9+FyVEva4SQzonRaKZKQV+Bwm3cTmNhCeTlbac36ccFt6F5P3UkhvBz8IH3sP3+d5LNZhdtW6vViMfj9Pf3K4GhUGwQyiMjDP+f/5Xsj36IyOcxJF7+n2a49xBthZaARe3I5vfRZW13Cfu+nPvsSvu5knNkAzKXZ/rRnyKtzVXXZNOLjOzxY0w++jNAB2nXLzwxV19MNJgQ69vI+v8aL1FxkRc4iTklcz9IuYLtYPEP+ULbNPbTXTfsD7D3xp1s/dAHKZZKTvbSBizLIhaLMTQ0RCAQuPCJUygU64LM3r2c/r/+BGvva2jY2Jpzl2hM5tvsHiLq6wix9EPVcu5fXjsr3a7Jvi+2nWjY6Yr3t8J+eusss5+2kMi9+8j8/DE2E5t6usSuVJj8wt8QLJaRQqtXWXV+Eu40SVPTW5MrbjkKeOF2rqltpdutfH82UV+At1NddH3qE1j1sNRGLMsiGAwyNDSkyrUrFBsAu1pl8pvfo/ijf0IvZEDzgRDOw44rNOrrNhUR4sJ5hZfcbpltXHDbJbZbjgW58X657O2uQj81CaJSJf/TR/HfeCPhoaGL9WpDsKlFRv7QYexDbyNEvaLpRkRC0GeQNnR46MMEIxEqlfkZ6GzbRtd1ent7Vbl2hWIDYGZznPnK17B+8Rh6rYhAQ0o5V09xrTu4SZHYVE4dY/KZZxj8nd9Z6+5cFTatyJCWRebb30eTop4+d+MhBRi6TtEIcPwjH2Zox3Yq1er8daRE13X6+/tpbW1do54qFIpVQUryhw4z9d+/iHXkCLZuIbW5yLK5Yo+KtUEAGubeNym/7xzB/v617tAVZ9OKjPRTz2G+uRe5QNMv9FNYj7hZPYUUJIwQL77jNrbdfz/lSmXe8bnve3p6aG9vX5O+KhSK1cGuVpn95fOM/e3fEshkHD8BOf8Ot1KHSMXqI6WNPHSY7KOPEfz9fzHfOWYDsikdP+1KhZnvfBvLqrFRDYc2kDD8HI1FSD3wTmqm6RQ9a0BKSSKRoLOzc206qVAoVgUrm2X0m99g+k/+BH+mUPcps9GkrDuzb8pb/TWMRva1V8m//fZad+SKsyl/eWPf+i7y3PCiHBgbBVtC1AhyPpSg9NlPk+zowDTnh03Ztk0ymWTLli3z6pkoFIp1hJSUz5xh7L/+OcWvPVwPczCZmxOpxzZstEqP6x1hYQ+fYfbRR7ErG7tK66YTGXa1SubpZ9BqFnNxqhsIKQjjQ9cDDH/wXfTu2kW5XJ43TWLbNsFgkC1btqhsngrFekVKMq+/wej/6/9L5aWXQXNqEgkpF2SebBZ4qVhrNFtSffkFcm++vtZduaJsuhFm9sknCUyNYesgcMNUN8rFJxDYJAJ+nt/ST+87dlNp4ofhRpIogaFQrE/MTIbc63uY+esvYBdyoINmayBtpGiSMUc4wmOjPVOtbzS0XJnigQPEbr9zw1Zp3VSjjJnPM/tP30PUXPOUmw5rYyCRtPh1jsR8hD/8AYLhMIVCYW65lAghGBgYIJVKrWFPFQrFpVKZmmLyf3yZ2lO/QGg6op7zQgrby0BpLdAZmgoquSaRQPbJpwnvvofE7tvWujtXhE01XTL2yM8R52ecMC4k9USya9yr1UEiCesaJSPG+HveT9/27ZRKpXnruFVV29ra1qiXCoXiciiePsP5P/tvlJ95ipqmYYNnoZBi8Qvm0l5vUBe0dYxT/dY/NUPu2WewF6QX2ChsGpFRmZik8vgTUCs7FVY3EBLQhSCk67y5cwdb7r2XUqk0L6unmzK8t7d37TqqUCguCWmazDz7HOf/0x9jvvYiQtq4D0jebKhsqIYgQdhzn9XfKq5BLAS5F58jt++Nte7KFWHTTJdkfvEztPOnMd1AcSnYMJedhFQgxMlQmPj73k0gFGRmZsYr325ZFn6/n8HBQXRdX+POKhSKlWDmc0z94PsUv/Z1NNPG0g3seq0M5w42fy5k4V1NLvmH4tpAEJjNk336WaI33oq+wco6bKxH+iWwKxWKew9Ss6uAxVxh4/WPRBI2NAq6zvB999CzdSvZbNYLS3X9MPr7+1XKcIVinVEePsfo//grsl/9GlggtU1xy95kSCSS8mt7qJ4fXevOrDqbwpIx88SzVI4cQwixITJ6NqJroBk2r+3axrYPvI9KpYJlWZ7IqNVq6ypluJTSe9m2jW3b8z5rRAgx76VpGpqmzWU8XZD/w93ebXNh+wtpbLfxX/d1ucfY2I+Ffbjc/a7kPDYea7Njdpev5LgWHtPC76fxWJbqq0vjdpfSnwt910t9xwv3sdLz2YyF+2k8F81Iv/IG01//GhzYh9R1Kkiw5aLHJDfVlhBiLli1/l6xPrAEMDtD/sUXCfT3oW2gKtgbXmRYhQKlXzyKVkp7lQcd1r/YsJEk9CAnYglaf+VXCEciTE1NeTfiWq1GLBajo6NjjXu6PB7Z/xOeOfX8RddLBZLc2nYzAtjdfRu6rqPrOj6fD8MwvFfjTTxdTPN/PP4ny+rH7rZbSQTidEba6U/0YRjGovZ1XZ836F0MKSUHRw8xXZhh78h+TmfOLlrnHe23kfDH2J7aSiKYmHcsPp/PO86LDbTLPY8AW+NDDMb6EcCO1DaS4SS6rs87TvdcXmhQ/JMn/ozxwsS8z+7q2s2Dg+8C8M6f3+/32tU0jT3n3uTrex9e1N7/7bZ/QcwfRdM0DMPA7/d7579RCCxESsnJqdOMpEc4NzvC66N7F63TG+lmZ3I7QSPIDe275h1v43G73/GPDvx02eezGVvjQ3xi+0e9/bjnwf1OFx6PtG1e/P5XGD28h+O3G4y//+5FbW47PEk4U6bn0CTRqRK6EBhCOP9qGroETQg0VzBdcu8VVwUJQmrkH/kBsXffT3Bw61r3aNXY8CIj/dyzlE8ccWLH17+uaEAS0qAmBGM33cLurduYnZ1d5IcxMDCAf4PFX89U0jw98ksAnhr5JbfEbuAdbbcRDAbnvRoHtJVYsPZM7fPet/lS3Nt6J12xTq/dUCg0b9C70NOolJLnT7zE99/+0UX3+8bkm94x9Qa6uDl5I/3x3kXHtJzBdrmczJ7mZPa0t9+A5ue2xM3c2n4TgUBg3mvhfufve/H5LRaLjI2NAeD3+wmHw0QiEUKhEIFA4IICbWpqkoKWR9d1gsEg0WiUcDhMMBjE5/MtWt89z48dfZJCrdCkxTlGCqOMFByz9KNnHueW2I3satlBayRFIBAgGAwSCAS838/lWj+LpSLDw8P4fD6CwSCRSGTeeXDPpZSSZ996kkdOPgE+4ObYkm2e2OXUGjpwdz8tIxkGXzlL+9EZAprmvXya5gkPTVk2rnmkkNiZNNPP/ZLuz/ejNfmdr0c2tMioZbOkn3kGUa4BG0tjCCBg+Njf28PgB95XjyaxEWJ+Pox4PL7WXb3i7M8dZDh7nrvDu4nFYsRiMeLxuHcj9/l8XOo4MVWb4WdjT/KOiZvpinUSj8e9fYTDYQKBwJID/htn9/LDt39CvnrhQa8ZI5UxRsbH6J7o4PbYLbTEksRiMSKRyLzB1n3aXq3U8BW7yiuzb3AgfZC7wrfTHe8iGo16+3UF1sIn8GanN5vPcip9CiEE4XCYlpYWbNtGCOFZCpYawM8ND+OznKf+eDxOa2srUko0TVu07zfO7uWHb/2E/EXExVLsz73N/tzbDOp93Jq4kZZYiydqQqHQopT8KyVfyHNk4gjhcJhYLEZLSwttbW3zzsMbZ/fy/X0/pGivPMX0bG+C2U/dTPzEJFsfOUisVCNiGIR0nZCm4a8LjsYpFcW1iMRGI/vIIyTueyfRbdvWukOrwoYWGcUTx7FPn0aXAnNDBYlLIoaf2WAM+eCDdHR3MT4+7t10q9UqbW1tJJPJte3mCrmcJ8YZkeaFsVfYOjpAW1sbbW1ttLa2Yts2kUgEy7Yu3sgSmJi8XtnH1tE+OlIdXvstLS3AfH8Gl2+9/k+8OnL5IWmjcoKJ9FPcMLmD3kQPqVSKZDJ5QZGzGn5HRVnimcKL7JzaSn+sl5aWFlpaWojH40SjUYLB4EWtKbPZNOWzBYQQxONxenp60DQNn8/nWUaW4uSpk9gFi2AwSFtbG7Zte9MmroWhVC3xzTe+w1sThy77eAHOWMMMT49yw+QOBhP9JJNJEokElerl1ZbIFrK8+eabxONx2tra6OnpQQjhHc8P3v4xr59fPK2z4v1sa2ffv76fzm++TvuZDAmfj7hhEDEMglLi1zR0UFaNaxgJ+GfT5A++RWRoCLEBogE3tsh48UW06Qy21EBc+iBzraFrGgVD4/gtN3Hj3XcxMzPj3egtyyISidDT07Pu0oZf7uCYTRQ4uPcQXTOdFAoFarWa9/Rrict7GrV8NkeLJ5iZmKGnp4dKpeI9lTf6K1TMCl966R84OXv6svY3b9+azYHAEabOTjEw209HRwft7e20tLR4A77P5/P6sJrOzUdCJ5k9NUPvdA+dnZ10dHRgWRZSynn7bWbKyBfyjJ/OoGkaLS0t+Hw+EokE8Xgc0zQv6DQ5PDJMZbZMOBymVqsRjUaJx+PE43Fs22a2kOaLL/8944XJVTtWAEtYHNAPMzE8wbaZLbS3t1OwipfVZrlSZvLseVKpFLZtEw6HSaVS5Io5vnf8EYYLqxdRIAMGY797D+lvvkrHwUla/X5Stk3c5yMMBDQNg40hNDZWvuY5bCEoP/Uk9nveix5bespsvbC+RqEVkD92nMqBN7GrVcQGmdsC56IKaRqno2HaHnw3lm1TM2sIhDdN0tfXRzQaXeuurgp2yeLwV/YzPT1NqVRC0zS2PLiD7lt7CbctDsmdDqcpnMhTrVaRUnqOdiLYfP5/4unznH7tBLlcDsuySPa3sOXd22m/oWvRurJX48SjJyiXy1iW5TkzNvoqfOP1hy8oMLRxSXG2wPgr5ykWi1SrVSzLIr69hUhXlI7butH9zZ9eRjumyR/Nk8vlKJVK3rbuQO0N+Ms8j36/n60fuA5fyEfnbT1L9nmsbZrcoRz5fN4TVzA/WqKZyihXyoyOjqLrOqZp0tLSMu+YL8Tk1BTFyTyRSIRAIEBXVxflcplarUapWuIb+75zQYEhpiWlSec8N54rPWTQvruLcHuExJaWJbcfT85QPlGmP5ulu6OD7bEhALLZLCMjI5zOncW8fvEQd/ixtzn++OF5lopAIEA8HicQCFAul72+PH72SYYrY0v2obbvHMWZHMPffJmSbVOzbWwpST1wHdGhdnru34kebH5vK//GXZz+4rMUTqYpBwLUpMT2+cAwYAMIDQmEdB1TSkzbYqPJDfOtE6QPHSF15zvWfZXsDSsyZl56FevEaTRdw95A3hiGJrA0GLv+JnYPDjI9Pe0JDNM06erq8sz4GwFpS9LpNJlMhmq1it/v59wvTzF7cIrb/9XdiwfkmMa5c+c8gREKhRwfBhFq2n6xVCCTyZDP550pj2HB0R8cRNM0WnfNj8rR/DqTpWmss47ACIVC83wknjnxyyVN9/6MQXbfDBOnx8lms/MsIbquUzydo3y2wPSr47Tf003nHc0H/XRfkfKBY5TLZc8aAHNTNktNQTQ7j4ZhMPLcGQKBABO/PE/fh7bQct3iUGctpDMVnaVyouz5J7j99qwnTfZZrdVIp9Pouo7f76dYLFKpVDBNc1GY6kLyhRzZdAbTNMnlcp7AsG2bnx5+jFPpM023M/Iauf1pJo6NkcvlFp1nqpLpV8eZEYKJlvN03t9LfDC5qJ3xX46QmfFh9da86cdgMIhpms6Uj+HDZLF1rGbWyGazaJrmOXW6U0uuA63f7+fQzJucC043P/ij40z+8HUmj0+QN02qto2b31MXgszzx8i9cJzRb7xE7z+7l9733dS0Gftzd3L2zx6nUiphNVqNDAOhaV6o63pDIInpPs4FfQTNCtEqmBvnFo9EoFkmMz9+jOTNN6KHmt+71gsbUmQUjh6l9uxTSMtEoiGlDevwYlqIBuiaztu9g2x/6CHy+bxnvTBNk3A4TEdHx7LDKtcDEkm1WsU0Te+GnUgkiIfjlEcKRLbMd2z1R/1MTk6i67rnbJhKpSDQ/Ps3TdObVvH5fMRiMZLJJNaoCbsWr1/z1xgfHycYDBKPx0mlUiQSCSzD5rHjTzbdR+C8zvirI0xMTJDL5byByo1UcZ0p3UgV64zFTHGc5L3taL7536Ue1sl05KidPOlZA1xnSHfQbzZ2L3Uek8kk4XAYn8+HfaBCOj9Fcvfi2jaR6+KMHD7jVfFt9KtYalrOti3K5TK6rlOpVLz9u5aQC1GpVimVShiGQbVapVarYVkWJ9On2Tu+v+k2xphg/IVhJicnvfPsDu7ueXYHfHdaqfJqnpnzFZLvaEPz61hVi+OPHiJ/IksymfSO03V2NQyDWCxGQk9QorlIcPPU6LpOJBIhlUrR3t7uTTfZ545zanfz47aePczIt19hulqjYFlY9es7oGkEF0SNCKD6T28wcm6Grs/fu8iqYSTCFD55G+e+/IKXV0PUI00aI07Wy51R4jxk+TQfLw4NULv7HqoHDvC+40exCqV69dmNgS1s9Befp/DWB4nfcce6Hr82pMjIv/461qmTCEQ9iX9Dfn8vuz/MXV4LZ/fkgmXuustdbzltrnTfAp+ukwsF4b3vIZZIOFYMIbBtG13X6ezsJLzBUtKCk2PBHdDa29vp7++nvb0dLe6n1iR7a7FYJJ1OMz09zezsLLlcjkA82LRt9+lf13Xi8Ti9vb10d3cT7Y1TZrFJ3zRNitk809PTzMzMkM1mKRaLPD3xy6btByYMxl4ZZnR0lFwuh23bnkBxBZAbCeP3+z3BWCqVyB7OUbt5cZvhLTHO7z8NZ5jnDOkO+EtZCBaex4GBATo7O0kkEgSDQWzbplAokB7OQ9/iKRt90M/I2yP4/X7PiuOG1V5y+M4KeWHs5aaf+8YE5395jrGxMfL5vHeeE4mE57SaSCSIRqNe+Kxt21QqFfL5PNm9OcztFpOvjmKOVOsRSZJarTZPhEajUWKxGLIAY3KxyDB0wxNtLS0t9Pb20tfXR09PD+2pFsSe/QzfKoHEom3tV09y7tuvMFmpUqhPg/k1jahhkDAMEj4fUcOJGPFpGgKBJSXl/WOUys8T/oMHF7WZuG2QY937YDSPjsQQ4BNgINEMHYSGPm+661q5JwK4U3ECzYawXiVn+HjhhpsJ3XUHuhDoN93MieFRdtQks7USulsVbk5WrfB4VtrPK9OmRMOiysyjPyJy883oweb3r/XAhhMZZj5H+ehhdNvGnpd8y8UtFwSObcArKcTcRSObLGOJ7bQLLGtss3H/jcuW16YmNBA6J4a2c9PddzM7O+uFo9mWRXtHx4atrur6D7gOcz09PXT3d3O+ZQqYX7mwkitj2zalUolcLjfPf6EZAuH5U8TjcTo7O+nv74c+nVEWz/kXpwpUKhUKhQK5XI5CocBsMc3xzKlF6/rKOlOvjDE2NkYmk0FKSTQapa2tja6uLrq6urwooEgk4viO1EVGuVwmm81yauYc6VRuUdvGliAjB0a8p3Q33DIQCDiWu2Wcx+7ubgYGnGicSCSCEIJisUgmk+Fg7TiWb347ia0tHH5mP8Fg0HmaTyS8KJerITFmrFlmq5lFn+sljfEXR7zzDCw6z+3t7d559vv9nsioVqsUCgWy2SzTk9MEYzrRgYhngXGPLxwOE41GPeuGmbdh/K1FffE1TIvE43Ha29vp6+2lKxbD9/wv8c+e562PLE6uxUSW8//wS6YqVfKmiQTCuk6Lz0eb3097wEfS5yNu+AjpAp9w8nxa0qZi2+TPZpl67ijmu65b1HT0PbsY/4fn8WsQ0AQhXSOggU+T6JqBJjTEmtwTm7XZeE8EiY4QkohhMNbSxeFbbyZy3S7KFSfixx+NcPLGXQy9/CpBKbCkU5bK+T1aCOlU2pbCXmLfF+vnXCG6pZetdpuOzJAIsi+9TGVyinB/H+uVDScyMkeOU33rAJqYuxCElMyfNW40QS80GDb+vXDZhba7WJuXsp2zTAKaJsiEo4QfuN9LayyEoGaaRGMxuru7N2zxM9fB0E1mFOgKMdOVpaItFg7FSSdXgmVZ1Gq1eab2pXCzSgaDQcJtEap9NtPhmUXrWVWL4rgTktnY9rnSSNN2K0eKTExMkMlksCyLaDRKR0cH/f39DAwM0NPTQ2trK/F4fF6SqcYn7NaZVp6rvExVr81ru3VnG28/s5dQKEQ0Gp034C81HeE6I7oiobW11RuE4/G4ExZaF2fF6QonmO/3EIgF0dsMpqenmZiYoLW11ev/1UjXP2ku/k4ASgdzTExMkM1msW2baDTqicX+/n56enpoa2sjHo97eVPc6ZJarUapVCKfzzM7O+tZv9ypSFestLa20tLSQjKZJBAIUPSVYXxxX1z/k0AgQDgcJplIEJmdQvvZo6TKZUbft73pMWR/vp+Zao28aWJJSUTXafX56A4G6AkE6Qj4Sfp8RA3DiRBpCFeuSknBtEg9cZwjt/VjxefP4bfv3sKZLz1DUKsR0QziRo2ophPRDAI490ZNrMU98cJtSkAXkkAwyKHBQSKf/Tw3tbTw1ltvefc/y7YJbNnCviNvc39aMluuYAvHEiDsessS5CX3c+2W6UiC1Sqlva8R7uuBed/R+mFDiQxpWWT37ofZ3Iaan9MF6Lrk3NAgu269hUwmgyYEdj08s6WlhdA6dw5aCiPi477/bb4ZOMPSIYXjb5xfUftbP76TrR/f6f1tAtOkm647tX+sqaPcrLX46VpYgolDY6TTaWq1GoFAgFQqRV9fH1u2bGFwcNBz0g2Hw96cPzgiw50yicViXDcyyVvlw/Pa1wMGWqvO9PQ0iUTCGwgTiQTWBUSG60/hCo1kMklrayvJZBK/30+1WqVYLFLzm5wYO7OojXB3lPSb00xPO690Ol1PlHXliw5OWounJ0QFxg+Oks1mqdVqhEIh2travPM8MDBAd3e3d54DgYCXwExKiWVZVOv+H8lkklQqRTabpVAoYNs2gUBg3pRLLBYjEAgQbfKdA2jaXIpyv2Ggj44h39pH1DCIBoLkO5uEJFZqjD9/jKxpYtanSBI+H13BAH3BEL3BAG1+Pwmfj2B9qsR9nLABS0pihk3E1MkeGGX4/vkpqY2gD/+tfaQPnGfa0Gmp+kgaPmKGRVDTuCbzAQtJUAiMUJD9W3fQ8Zv/jGAkQrVaI5VKMT4+4U0VG8Eg49t2Yb6+j6CsUnA2R5PzmkOuwyFBQ1D86aO0fOgj67aeyYYSGYVjx7GefQrsGvY6Ds9aiC5gPBql5Vc/5nnYuxeYmw1RAelDM+RGs4vCB12fi8vBKpqMvjriRa24fhB+v58Ra/G0ij1TI5PJUC6XEUJ4Voze3l76+/vp7e2lvb3d8xForEniDn6uo+IN1i7eOn140T78ySC58QwzMzNMT0+TyWTqg2Nzq40bgdJYRyQYDHpTAYFAANM0PUfUX0w8Q2VBBkoj6KNSqZDL5Uin055PytWwZBTlYnFZm6qSzWbnTW90dHTQ19dHf38/fX19tLe3E4vFPOfaxnwitm1jWZY3JRKPxykWi16Yss/n8/xPGjPI+peI4vHq2dRq1E6fppLL4YuGCegGQU1jpm2xz1T19BQ5y6RSF4dhXafV76PTH6A7GKAzECDl8xHWdSehVsO9TQK2lJ5T6NCZNMP3L+5XsDPB7JvnyNRM0maNvGVSsgyihoEl5TUTzupYLwRRn85IOMbwu97Djg9+AMu2KRaLaJpGW1ub47OUzQLOtFd053W8ODLMu0drFKqVecdiC66JY7sUJBqVwyfJvfE6ifuafLHrgA0jMqRtk3tzH9bZ0wgh12VoVjM0LKqBIKN33cfubVsZHx93kkvVb4Dd3d0E1qnCXU2Kp/Oc/NkRTNP0wkuj0aiXHvpy6rfYJYtzvzhFbjoLQCAQIBKJeNMTeXtxOuvSVIlisYhlWZ4DYkdHB11dXXR0dMybJnGfrBtpTJ/d39EHpxf3K9gSYqrshMS6A36pVMK8SA6KZjSGpLr9ifhCVCrzRUa4JYxlWRSLRXI5J3fG1RAZtmZTlbVFn5fTRYrFIrZtEwqF5p3nzs5O2traSCQSi4Sce8yNkTmuhScajXpRMO40XWOm0YvVjLErFSrTU9jShli0PlHrrJ+OLr5WiyMzlC0bWzrJsmKGTqvPT3vAT6vPN2+apJkY0IRAq0ehdKYXnyOAUHucMdsmb5lkTZO8aVKu596wNI01zyTk+OdjGIIQAc4k20h/+iFuu/suxxpYrXrCUNM0OjudhHuWZVGtVGhrb8d874OMfufrxGwo1U+DFHULhly/QsMWMPnDHxC/9751Oa5tGJFRmpxk5rXX8GnOHKMtbebPfa0/NCHQdI2xUJTB976XdDo97ybphk9uZuyazfSecYZfPEOh4PhLRCIRrz5EKpXyntIvhdyJDOeePkl6Io1pmp753PVHSCQSC31PAajVqp6zqesEmEqlvCgHN8lUM4EBcwOg3++nJd7StG+acMSm60dRKBQolUrY2qVnt3XFht/vpy3cykwlvWC5Ns+XoVQqUalUroLIaN5+zXSiP4QQ874b99pwz7MbttqMhQIrEAh4eTwaM7oup06MXSlTGxslYNkQCILUwHasU/YSp6gKVG0bicQnBBFdJ24YxHSdiK4TXOCDsdSZ1oFoqXlmW4EzrVK2bPKmScGyKFkWFdvGlBLLljT1k79KGAICPgPTH+TN/i3EP/1Jtm8ZYnJy0vseXFxBmUwmGR8fJ5FIOJF1g4McP3CAe/bvp0hlLj9S3bdyPabSEEhsDXKH3yJ/6hSxreuvOuuGERnFU6cwD+zBj8Bet5p1MQF8TN5yOze3tDAzM+NZMUKhEF1di7NSbgbMisnYnhFkzWbitVHPh6ExL0F3dzfd3d20t7c7T7IrDAGr5Mrs/btXqWQds3nj/HxHRwfd3d10dnY6ic+aOACaloVlWZ5QcC0f0Wh0XpGxCw1Y7gBniSV8LDTnya5arVIul71skrbv8vwj3OmmsrW4ZoemCS8CplKpUKlUnBDPK6znDbP5dJf73bjJr9wQU9eCdSEh18jC+jONosn9bDlPkVapiKjWQNcX+AAsPcTZUmI7ror4NI1QXVyEdZ2gpuMTAidg9QL9x3kosSPNxbTzMC+p1aNRKrZNzZaYtsSWErkoxPLqIJHoAiK6n3Q4wuHdd3DdJz+BPxhkdnbW6XuT816r1ejo6PAsT4ZhoBkGvnfcyeT+w8SEJEcFfT0qiwYkjj9JLFcg+6NHiP3R/7LWXVoxG0JkSNtGHj9OtFLD1HUatb7WIGbXFxK/LhgLh+h58D0UCoV6oiXnSFpaWjbFNIlZqPHc//kE09PTVCoV/H6/N1C7T6ju9EggECCZTNLZ2UlfX583Hx+Px5dMxnXskYP0v2cLwcR8x9lALEgtaDJ9atrLHBqJRGhvb6enp4e+vj7PcTOVbmGmMjtve39rwBv83IiOxlLtyxn4wLnB5iqLQ1gB7IpjsXCdF2u1mmPmNy7/1y6EoGSWF3/OXFlyN5GZaZrgu/JXmA+D2oIsm0bM75nQ3cRZ7nl2fxsrMTGvRFA03R6QXgZUucjZMJ6rkI3Nv26DvSlsCZoAo554y0u6pdWTZi2jOwIotjYX01bRsapZEk9cmFJisbRl5EojAb/QSPr9nIm0MPrRX+H29z1IsVhkdna2aVJB11/J7/czNDSEpmkcOXKEarWKYRh0bN3C6R1bufvYIQplC4HuWDEWnD836LQZjf4uLo2BqGuBJjXK+w+s0d4vj/U9n1CnNjlB4YUXkA0/HQHoEjS7/mL9vfxIzt11F6nubs/0LqX0BrvNhGvSdp0u3UHfjazo7u5maGiIbdu2sW3bNrZu3Upvby9tbW1Eo9ElHfUATr10vOnnN3zkZkqlkndTSyaTdHV1ecmVOjo6iMfj9MUWpwAPpuZEi2tubyyktpLy7CenTzf9vDiWB/AcGN3XatwKy7Uyk8WpRZ/XcjVvn437jerNa+Vc6mBdnizOO0dCCMJNUsNHuqOLpjUaX5fTh8vBC86Uc0GKGtA+tdh/J9yd9N5rgI5wIsq8v+feX+w1PpSkGflTE/U+Se8/kE7/5NW9r7nnJqTpBLQAB7oHmfjcp7n1Qx/wUvw3Exjuby0ej7NlyxYv2ieZTHph3/FkEnH3HYz7dSKGH1k//42zbY1jwkK0huUuYsE2a/Gy0bAnxqmcPdek19c2G0JkFA8donbwLaSQaNJGSBtdQqOVWbgydB28pISQMDiTTDH4oV/x6mq4KZ03m7NnozXALZftJpLaunUrO3bs4Prrr+f6669n165d7NixwwsRTSQShEIhjAtEA5x44iilzOLIhdbBdrbcvx3DMLwU5Y3poVtaWohEIvQnexdt64v6ab3eEYILRYA7QC/Xj+GF0y8t+kzWbHJns/MG2LlB+fIH1aWETTW72LoBS+xylQZ397haxWLfFF/UT9uNHYvO8Zzg4or4i1jli5d/n/dN1LvQOrlYZPgTYTrevcvzG7Dr/huy4X4gl3nveGvX4ocPu1Ijs+dMPSmrO/UiPAHk9e8qvYSEuKEjfX7euOcuYn/wL9lxxzuYmJigWq02FRhuIcD29na2bdtGMpkEnHuDW3HaTWI3dPvtnLjhNsK+AIbQkHVnE8093gWminnhrgt+KoK5ccTNubEWY4JAohWLTP7op4vOzbXOuhcZZrHI1CuvIaTmHYyG88MQAmzN8c5d/dvMlUTiNzQm776XSCLhmKOlc7OMxWIbqgDacnCd8WKxGO3t7QwODrJz505uuukmbr31Vm677TZuvvlmbrjhBk9gdHZ2Lsrw2Aw3rfipF5tbM3a998Z54Z6RSIRoNOqFM/r9fm7sur7ptoPv3oZt29RqNSqViucz4SYHW87g9/zxF5tWGy0NF7wEY4uqwa7C4H588kTTzzMn0vOEjRuZ0WyPun9pR0n32Eu10gX70Wj1adWa/+6HljjPpmku+zw39mspUehSS6eZ/Mkjy2hs8UeDB5tXjt32sd3YgCklVc9nwnaqp17A2bOR/e/sIx1b/PCRO3Qeu95/HWdKxlevX6ILcdUGAQloApJ+P/lgiH3vehc3/s5vk2hrZXZ21gvNX0itVkPXdfr6+hgYGFgUKRYOh516Q5ZFpVIhGAoTuutOxgI+oj6trg4aHEcbxgVXYLgCxFumNVg1mPts7cYRibRsss/8ArOwWKhey6x7kVEcPk/phZcRUiDterySdJKy2vVHALlOXgASgV/XOO/XSb3jDorFYr3KpWPF6OjoWJdhTJdDo/Ok69Q5ODjI9u3bue6669ixYwfbtm1jcHCQ3t5eOjo65gkMZxBcIrIAx8nxzDMnKTd5Sk8NtNJ9u5PSt1nEgaZppKIpdiS3Ldo23BZh4L1bvdTV+XyeQqHgVRS92AA4mh7je2//qOmy9PEZr9iZG3rp+iKIyyyQN5oe45lTzy/6vDZVpZwtzfOBcH0f4r7FSaZ8sbnBwB245/3WpaRUW3zO7dpcKfnGc93mayXWZFom3BZl20d3ellS3dBaN/JlOUJjND3Gl1/6R6Zy01Sr1aai0LZtZl55iZH/33+i8kbzGir1o63/X9QtEPXJCSmJTJXoOb84kVekK8m2f/4AVdumaFkU6hEgFcuiattY9TLvS907pnqjPH9n89TTM3tOYUpnMtn1+fDXhUbj9MWVuJ+5WGj4hE7c0DnY3cexX/sMuz71SSrVKjMzzTO5uk7NkUiELVu20NPT0zTfjaZpdHR0eHleisUiXdu2c6qvl4CwPOuDdB1c3f41mIgkEtd85PXfnptaWuuxQUqBLQT+TIbZp5++wG/v2mNdiwxpmpT27iU0O40U7oXsvLyZabkaxuOrg/tzTxh+zt54K8n2du9pVUonxXE8Hr9wIxuQpdJhuw6YfX19dHd309HRQSqVIhaLeUmT5qYRlmrcsQQEg0HG9zbPFjp0/1YvZNNN1OQ+Kbui471b39V024H7ttB2TxfZbJaZmRlmZma8eWd3AGyWBnw0PcbfvPSlpm2a41WmDk14ETWuh71bsOxyEo+Va2W+u/+HTZcVzjhZNYF5zqyBQIBkILFofV/Ehy/q9wSGVY+4abQQHJk8uvj4CjVvkFoYQnpz/IamfevZ3U/7vd1kMhnvPKfTaQqFwrzy8s14e/QQf/Pil3hr4hBfevkrjEydJ5PJeK98Pk+5XGbqZz9j4v/6r1gvv3KRG2ejPV4uWnLbc6ebbrXl/TfT+5v3kbcsMqZJxjTJmSbFutBwolAWM90X40efaG5NKx8ZZfy5I17CrYCmEdZ1QvXEXj4hrngiLgmEdY2oL8ibO3bh/91/zg3vfjfVWs1zaF+Ia5lqa2tj69atTiXlC+CmkndDuiOxKNYNNzGr+fALDQvbcf608fxQRP2rsWkop9awzP32bPcg1tQc7gxsds0i/cSja9mRFbOuo0usbI7qy47D51xWt/U1MdKIBIKaIKNrxB9417wnF3fucbNZMaB5OuxEIkEqlSKVShEOh+cNRiuNEHBzJOQOzGLdaaGH5w/Sib4Wolvj5DI5ZmdnmZ2dJZvNesm0/H4/PYlu7um4g5cnXl/Ufu/dAxQH80wen6JlfNxxRG0o0rUwUdTzJ17iBwd/3LyzFky8fJ5CoYCUEr/fv6hYmV64NJExk5/l6298m1PpxenEKUvGXhuZl5PCzYIZDAZpCS0+15pfp/+BIU486iRJcy0ErnXgjbN7OZU+u2i7Wq7atIy9YRhcn7qOc+WRpvVivPN8eork2JiXH0UI0TTjZ6la4heHn+TZ0y94bUyWpvnavm/xvvZ3E9KcejIBoVF65hn0R3+GnsvUI9guDQ1oO5fjur0jHL19sS/PwPtvJr+9i/TDrzI9USKs6/jqSbhswF+f4nB/2wfe2c+LdzW3YMhKjeHvvUrJspDg5eBw82+E6m1fyYpHNjZJ3U9Fg5fv3k3Pxz5GrB6O70ZfLcQV793d3fT29no1fS6EEILW1lamp6cpFotUKhX6dr+DE6/t5caRYUqWM+XsRRuu19uotODMGYonThDetth6ei2yvkVGrkDx7cPYmsQx/Ml1Y7VohpCSiM/Hy7fextDgANXanCd/R0cH0WhzD/6NzlLpsBufpC+3bTd9tDxbg12Lb7vdd/Vz+ntHmZqaYnJy0kvEFQ6HvVDJdw3dz4n0KSari2tshLui0AVnps9TmC17+U7caZ10NcO57DDPnHmeQm3p2iz5vbNMnBr3QvZcHx23ZHw4HEYvN3/OdgVrY/G4crnMq8Ovo03oPHvuhabbAWT3z5DP570Mpq5VzRU2XaluQmPBRWGvnbf14Iv7mdk76VU7HU+Pc6hylLeyi1OlgzMVZFnWvO/cMAzvu3+w7508fOIHVOzFWdDc83x2epTCbBlLWF5ODzdvxhvn9zJbzvDi8CvN91/L8sj5n3FLdRehvIn56M+JnTtNCgjpBj4psS/xRiOEQAdueew4Y/0Jsm2Lr+noUDvR//BRMkdGscdzaL84hiUlccMgpGkUh5JMbEmx//YeKv6lJcLUI3uYODpGTUp0IYgYBnGfzykZr9dLxl9BS4ZE0OoLMRYNc/rDH2brffciNJ1sNlsPy23ufxEKhejt7aWlpWVFVrlAIEBnZyenTp2iUCjQ1tXFyJ3vIDs1TNiaywK6npFoiDyk97ypRMaVRkrJ7IkTyFLZkafSMWytVzuGBAKawYyhE73/fm8O131a3ayJt64Gjdk1Q9N+tKKPSnj+HSnRlyRxayszp2aYnJz0som6BbOMeonvXx36MN8//mNm7eYFtGSrYJJZHq89B+dwXsuk9FaOs6+eJp93Qlfdku0dHR1eKfNwOIyWXnxjNiI+7v1/v2feZ4c5xeHsKcheeL/VUyWGXz9LueyII9fRzs2qGQ6HiYdjvLPnHh4/+8yi7VNb20htbQPgNKOcro3CEjf82lSFsT0jWJbllUx3c6C4tWgikQgf6/sQPzr7c2pLNCRbBVOkecZ8CYZxXivAP6Zx9M0XsPfsIZlJ0+33oweCzm9FUL/fLMECnwR3/t/9V8OxKtz19b28+M9up9je/OHB2NlNaWc3B5uUb78Ysz/dy+mf7aNoOnlFQppGwjBoNQxaDIOYW83V7e+K97AYL5EZEkNAImDwdmc/8tMfZ9t11zlir1hEE2LR/lz/i1gsxtDQELFYk0Jyy9i/O1XqZqPtv/MORl57iZuHh6kIy3kUdS3El3/Ia4BAVouUDu0HPr3WnVkW69Ynw8oXyD7zNDqWM0IvkRVxvSCRxAyDU1t2kurrm+eL4Zp6FVeORr+P9lLz+d/u3b1ks1nPmjE9Pe0V55JS4vP5aIm38MGuB2mRi30ULof867Oceu446XTay/jq+qZ0d3fPK2WuX6bjZyPmcIUzT5zwytUHg0EvlNcVNm7q7nsG76I71HkZO5OMPHOWUsmJOHGdfd3sna7VKhgMMtDaz3tbHyBor/J1YUL5tRzHH36ewz97lLGRYaZNk7RtU7IsTNeJ8xKbd/Nl+IUgXrK49auvEzo3u3r9Bya/8zLHv/saOdPExqmHkjQM2n0+2v1+WuqWjKCmYdSnYlYHZ/AOaQECWpSXd+7C+L3fpGPHDsrlMpVKpe7EPh/XR6ejo4Nt27ZdksBwcavwAhQKBVLt7eRu3U3O8BPQnGl1ue6iDRuRSGlhHTlGabxJquFrkHUrMsyxCcQvn8fWVi0cf82QCPwILNuidOP1CF33HOM0TVNWjKtAozWjTU/RYi8WCYF4kMStzrzvxMQEU1NTpNNpisWi54QZDofpaO3gwdZ30lu+jAG3jjVTY/RHZzj54pzACIfDtLW1eY6vPT09tLW1eVYVIVbnsi4ey3Hsx4eYnnYiLgKBAC0tLXR2dtLV1eWlbHdDef1+P5/e+RBtvgs76TVDmpKRx84wfmKUWq2GYRhEo1GvxHo0Gp0nMiKRCFvaB3lf7F20lFbHGbo2VuHsw0d4++FfcG7/fmZyeXKWTcWyqVlOjQ9TSi8c9FLRhMDQNIKaRqpsc+OXXib23LHL7n/lzBTH//gRTjy6j5xpYkmnOmuLz0dHwKnm2hEI0OLzEdF1/AuiSy4X24agzyAfCbLvEx+h/3d/j3hrK+VyGdtuXrTSDVsdGBhgy5YthMOLq9SuBCGEZ2F0nbNbbriBkZYkhmFgCbDEOvbJoP59TU6Tee65te7KsliX0yXSNCkceBNpVrE1wRK1k9YNUgp8AZ3jqRhtu3Z65jxX3V/uhadYHq41IxQKsSO4lVerexet0727l0Nv7mNqaoqJiQlvyiAUCnnZSN18HrdXbyF67hgnK2epda3M0lbLVEkfmGJi3xilUsnL+BqNRkmlUvT09DA0NMTAwIBXQ8UN2b1c5+DKRInxV84zfXzSy3jqWjC6u7vnCRvX+dXn82FZFvFwjE8MfpSnTj7LSXuxU2czyhNFzj1xiplz05RKJScktp50raOjY56Vxk3LDhCPx+no6OCO6m0cHjnCsBxFdq5cYJVGCkzvHWf68CjZ4ycwiwV8QhAyDAwEPgSGcF6r4b/gWTM0jYiuk/L52Pr4Mc7uOcvkvUMYd69srr0ynmHimYOM/vwAFcuiVhdBYV0naRh0BAL0BgL0BAK0+nwkDMNz+lwVK4YEIWxSPoOxaJRzn/40W++4k2rd72cp3ErSAwMDpFKpVXNqd618bjhz1+Agh1LdbDk/jY7k0ssHXhtIQCsXqL76InzmM2vdnYuyLkWGVSqTe/wpLCHxCgfJ9atOfViYRpCpO+5ha2srZj2uPxAI0N3dvdbdW1OEEASDQW+OfmGEwOXemHyGb5ETaSgUIhVPcS4zwmhtYt76gXiQtnd0UjxbJJvNetVPXcdCN+W5mxzItm385/yc3XuWCf80RbNEZFscf3SxmT99dJpqrkruTIbc2YwX8un+FlwfDDe1+cDAAL29vbS3t3tWDKdWx8rOQTVfJX3EcVadPTxFYTzvJbLSNM07nq6uLgYGBhgYGKCnp4fW1tb5uUjqUSfRaJQ723fTdqqFt7KHyIo8Ruf8463mKkwfmqQ4nmfm8JS3P3f7zs5Oent76enpmWcxcZ1sASKRyLyw2NC5IGdfPstsMkexVCS2I4G/SWn1Wr7KbP14x14YxrIsylNTlCcmEJUK/noURtLwkfT5Sfj8Xql1n5scbIlzqQtHPLjrugmvFq6vCYEPx1fCNgwn5H6qjPG9/Yx+41XyH7qeStAgecdWAonFDxlTe09TmcmTeesc6T1nsHGqrEqc/QfrPhjtfj+ddYHRGQiQarRirIJgkoBflxgixJ5duwh87MMMDAxQqk8hNt2m/gCVSCQYGBggGAyuetRcS0sL6XSadDrt+A3dfzeT544Rz1QwpVz/D6YIymfHyB05RmznjrXuzgVZlyKjfH6M3MG3MBrn1tapwJBAxDAY84dI3rYbTdeRtRq2bdPa2rqpfDE+duOHeaDvPsbHxzl9+jQHDx7k6NGj3oDtlu52n5yXyuLZjGQ4yX9+///O1NQUw8PDHDt2jLfffht9RpBMJr3Q2EgkQjgcJhqN8uG4U0vh9OnTHD16lFOnTjE9PY2U0utTtVr1zLKu6df1JXBvsm7p9NBIiLHJMaYfmyKTyVAqlZZMyuUKKFf4uFMHHR0d9PT00NPT4/liJBIJgsGgV8682Xk8cuQI586dY3rasRa4U0OGYcwTbW7eD9eiEwwGSSQSXmG4/v7+eYXnGvfrtulaW9ykY2fOnGFkzwiTk5Nks1mq1eqiY3b3F41GaW1tpaenxxNRjVNBjX11s8A29jkQCHD+/HnGp8eZOTZBNpulVCp5ImYhAqiMjiGnpwgBQcNwLAA+H60+H51+Px11P4ZIPbeELgTdR2f4zf/yS2ZrNcYqFU6VShwvFilUKsTrgsSL4KhvIxbsV4h61VXm7mNaPUnW+BNHmapWOf/wa+Qsi7JtY9q2I0aa/VbAS7IVrpeKb633vcPv93wxYobhHcNq3DJDQsPwh3jjzjvY+rnPoPt8lIpLR0e5fe/q6qK3t/eycrpciEAgQDwe93KdDNx2KyeefYK2dBZb0jTfyHrCFgKmp8i8+KISGauNtCwybx3AsGyEAKk5Fgwvr/w6QyDRDYPhm29ksMNJJmNLiWEYtLa2brq8GO6A7EYwtLe3YxgGtm0Ti8VIpVJe/gN3cFsujdkx3aRebhin3++ntbWVeDzuCQ1XyLhJuEzTJBgMUq1WCYVCjpNl/SbZWCvDTYPuCgV38ItEIiSTSSYnJ72kXG5yL1ekNA6YrlXEFRju9IEbTdLS0jJvuqIxR0iz81ir1QgEAhSLRc/fpzGrpvuv299QKEQ8Hqe1tZWOjg66urro7OycF8mycL+u82wikfAGdVdAxGIxZmZmvORWZj3yoTGjazKZpKOjw8uR0NXV5eVCcXOLLNyX24Ybhuz+TiYnJ72cJm7uhFpdwCNB1mpUz5/Hn04TMAwCukbEMEjoPlp8PlI+ZyqjxQgQN3TCuuZlyQRnysMQwpv2SBoGpm1TrouRFp+PqGE4DpZ1y0EjmvOjxK9pXp0Tvf53SNeJ6zozhuEl5SrX0417zqf134peFybBeqKtuGHQUhdJrT4fLX4/CcMguooJuDQBAV1jJhzm7Hs/wLYPfAAbSaFYcH5TTe7Ftm2j6zq9vb10dl6+v9LFiMViBINByuUyLS0t2LffRf70GKFSnoq9vidNhASqFapv7ccr4XuNsi5FRuHYYcckLOb7YzT+sNfD1IkEIprBuCEIvGM3voCfYrHo5U/YTEXQYH6ERzwe9wbGeDyOlJJwOExnZ+c80/lKrBnuQNSYntzv91OpVPD5fKRSKVpbW71oBrcke61W85JQxeNxyuUyhmGQTCYXTFMIbz+NNTdcweAOou3t7czOznrZJBuzUkopvdwQrsCIx+Mkk0nPCTKZTHpiKBAILBrolzqPwWDQs6AszH7pJjJrFBjRaJREIuHl4UilUt4xu4N+43E3WjMikQgwJ7hc64Z73MVikVqt5h2vKw6SySRtbW20t7fT1tbmOX0uPMfuvnw+37z3bg4PN7TXFRn5fH5OaFSrVEbHKB87hl4q4g8FCeoaYU0nahjEDYO4bpAwDC9xVcDQvURYbgimJoQ3tdJiGJQDAfyaRs22MerTFS0+H2HXAsLiBHGu0BB1oeEKBldkpHw+R2RYFkXLomRZVKV0pkbqWTx9Yi6TZ9QwiOu6kw+j3v9oPb+Gv0HsXM7t0QeEDB9H+nrhc5/n5uuv96KPBI55uVFjuD5mkUiEwcFB77dxpXEfJEZGRshms2y5917OvPQSu06WENis1yDWRuzxCUpnThPasmWtu7Ik605k2JUK1v6DaNqFQ9XXA0JC2NA43NVLorPTuxiFEHR0dKxoAN0IuINcKBRynjxsm3A47D15BwIBksnkPF+AlVgyGgutmaaJpmm0trZ60QyRSMSbfnCTbLlOhq5vQkdHB5VKxfvbHfQXip5GodEoGtynbNeXozFNuTuF4A6Y7mDvhnG6Lzec001StdA3pdl5DIVCtLe3e4P7QpHRaMVo3Hfjft19u8KmcdBv3DfgWXJci4o7DeJaFdwpDHcb1+fEFVWJRGLe/ppZrRotGgutP/F4nFQq5Yk4V2CU83kyhw5TPHUGDQMRjWEIZ6ohoOuENUdsuBkx3ekOQ5sbnN1e6PXpjoiu0+r3owlBi2Fg1pNfherWjXjdmqEv8VvV6sci6o6YRt0qEdF1EpZFqi4uSrZNtV44zQJPZLjWlKCmEaqLDfcVqvuHGA1TJJciMCROnZ+gLhH+KPt2bKf1Nz5PqrNzydojMCcwkskkQ0NDy8reuZq01LOLFotF+vr6ODzYR234FFoVrAXjh/uQuh4eUMFJ3ihzBQrnx5TIWE2Khw9jnDmL1J2LZT3rDNd8WtwyRFd94DNN03ua3mw0PoG7g20kEvE81Bvn7VdqyVj4lO36TriDsDvQuamy3SdndwBzn7ZdPwrAEw6un8jCAd/dzk2N3fiknUwmvdwBjT4d7nG767sJqRodU90Bvtkgv9R5DIfDi8TMwm3c/jYKjcZ9u5Ed7jEtJfAahUbjtE08HqdUKi0qENcoqoLBoDcVtVDMLHWsjVM+jfuLxWJeobNKpUJ+bIzJ119Df/oJumwBkSCaNmc98AlBUGh13wYNn+aEmnoiYOFx4jhZgnMv8msalXqNEVG3crgD/cX8IARz/hiuePHbNhFNI24YVFyBUbdiuJVZtfp2rtBwnU4D9b+NBuvL5U2PCAK6j1wowIn77mPnpz6NpmvMzMyiLWGqtyzLm4ZcqrjZlca1Hrq/u/B1Oym+9jrxWhGr4RpYWPZ+XQgNIZDpGfL79tF2/71r3ZslWV8iw7YpPPcUUpholu75Y6xHJBDRdUb8IaK33EogGCSbzaLrOl1dXZvOiuHiCoHGqY3GOhbu02rjU/xyaRx8XatGY5Ezd4DySqZrmjcIutu5VoDGwdFNd91s0G8cBBuFRiQS8USl65ToWhcarQoLX40D7oWOvdl5bNxPo8hYOAXRbN+uqLmQuFi4/4UDfygUwjTNeVVN3fPo9tMVOO4+l7u/pQSdu7/M2wcpffe7pPa+QVL3g895SnHTfLsDtQ+BwVxEiCsuFj7MuMsM52CdQdi2vQgPd+rDqAsGw/3OLnQMDe26/bE0jVA9P4flvsDL1eGKE3dfjf82iqNLv01KAkIjpAmOd7eS+8hD7H7XA860W76EpumLzo4bPRIIBBgaGiKRWN3EdCulcZpu6+53cPoXj5HMn1n8vbp/rKcxxbYoHXwTs1DAuErTUCtlXYkMs1ph/O23CUE9TS/zfiVNfzDXIM6FL4kIjf3bttE1MODdeF2Hus1IowOiO1A0DoiNT6yXEsLqttvojNk4bbBU240DZWP55ZX0p5mPhntsC0ugLzwXC308Luc8NivD3bide7zN9rtSJ+RmAss93qW+14X7XSmNYsMwDHxCo/bGG+T/+gskzp6l5gvOiyyYN7AzNyi7y7yU201OmZvvwifBQOAX2ryHHlcAuE/JYpnzu6LhpSOQQjjlxgWeBUMuWL9xKkcwt89LnVN2z0NIM9A0nUMDg/h++9e5edf1jI+Pe+HNC0+Mez1Fo1H6+vquiarRrmVsZmaG9vZ2yjfdQnVkDL3qhLM6Jd0bjmTNK64uD1fM+obHKR54m/g9d611l5qyrkRG/vgJKmfPEBQalu5cP+tJdM4hCWo+pvUg1ZtuJBqPk81msW2bVCq1aa0YwLzBZbXNq5farjvwrWYfVrPNpfZzpc7j5fTlau6zOj3D5A9+QuU738BXymH7dXxoqxaF5lkI3GNb5WOcJ3au0D4uRMDwkQ/GOHnHrbT9yodo7+lmdHR0yd+uW9TOjUK6VsLvdV2npaWFbDZLJpOh9/77GX/+Fbqr6yMt98WwpmbJ79+vRMZqUHnpFeIVExutfuHhZJtb436tGCkJGT7e7Omh7fpd2LbthRdGo9GrfkNWKDYa5fOjjP/1f8d++lmkVkNqYIu6k+U6eEpdK6SUTmSMP8hINMjpD36QOx76VUqlElNTU0sKY9M08fl8dHd3X5MJBFtaWpicnCSXy9HT18feLQP0zU5iYzsVddfxb0JISenYEexK5f/P3n/HSXLe9534+3mqqnPu6cmbsMAikhBzJkVRFKlkixIpWVa07ucg++5syz7f3evOutPd+Sz/ZAVTkpXvJFGJpCJJMQAgRRJgAEjkXSx2F5t3dmYnT+dQVc/9UWGqe3p2t3t6ZnqA/uDVmJ3pqqeeqnqe7/N5vhE5hBGJB4ZkKLNF6ZvfQLPAFrZPLA6eT4Zjv7WVSenwJIcKBT+kMJ/Pv+LCVkcYYdAof/3rLHzsLxBfewxLCmwE2LarBDjoaZh2CwJbOYXbDKHxwuwMpXe9nQff/W2srq76EVWd8ExfiUSC6elpstnsPvT91tB1nWw2S6VcoV6vE3nTG2icPIVRr2Lb6sAyTxuQykZeuUb1pZdI3H//fndpCw4MyWgsLmKePefvRHxyoQ4WCVXYxKTBhmVjHDnsOHyWShiGQT6ff0WbSkYYYaeY/8gfUf7ox7HXVlC+z4BDLKRbfvPgbUx2GwKlLOK6jS6jPP36NzL5wQ9waHzcL8rXKZeEEH50kJcN1kuMNqzI5/OsrKxSKpWYuu8+FhJxDjVqKHHAFpEO2ELRWligfPLFEcnYCZo3FqHZAPDL9UJH6NEBgAIiUmMulyM+OYVt25imydjY2NBP0hFGGFZYtRoLH/kIG3/wh0jlLojCdL/VAOGkBwZQ9sESGrsKhVI2WSNGKax49i1v59gPfJBILMbamlOCvpv5ttlsIqX009vvt9/P7cDJ1xJnealKLJ3m3Pgkh1bXEMpsq6x70MzVAoFlmdTm5va7K11xYEhG9dxFlLTAdvwxvDFxsIaD4y0OkpXDhzh8+BDVahVd1xkbG0PXD8zrGGGEoUHl1Gmufuwj8Mgj6LYb2g5sSgc3o8RBExa7DBtFCEEqEuFSpsCN7/g27nzn27GVYn19fdvF1jRNotEoU1NTFAqFA7UoJ5NJVlZWnBTnDz6Ade4sonXwiEUQCscvw1pdGkq/jAOjmy9+8xtotuZbVCUHqPMuFIqIlKyEQlj33kMylaJarRKLxfYs1e4II7ycsP7lr7Hwn/4T6pHPI4SGLQWbAYmelPA8+2xGPhkOlIKIkCSiIZ654ziVf/Qj3POed9MyTWq1WtdF17Ztms0m6XSau+66i/Hx8QO3OKfTab9+z+zrXstSMoL2MjBRCxT2/By1uWv73ZUtOBBP16xWaZ55AUsdiO7eFGFdZymfIXv8Dj83RjweH2kxRhihR1z9s49x4xd+Efulc0iloZQE5ZpGgPaMEyP4UIqMrpE0wjz+uteT+Uc/yczdd1Mul2k2m9sSDNu2GR8f59ixYwc2l4+u66RSKZrNJtlCgbUjx9EPgKnndmBfuULj4uX97sYWHIiVrfz8C+jrRZxcdwcXOhJhK1YPTXPP7CyVSgVd10mn0wfCpjnCCMOA5sI813/3d6l/4TFkveL+NUAmvLLMB9mbbxegENhYjIXjLCaSXHz3Oznyre8iHI1SLpf9BHOdsCwLXdc5evQo2Wz2wG+ICoUCKysrDqF63etonDyFMFvuaOlMc3ZwICpVqmfPknvPt+13V9pwIEZL6emnkablKEHt7QTH8AoUb95GpGQjrCPvvItINMra+jrpdHpkKhlhhNtE8ZlnWP61D9N44QWkku3Gj2DKxs50wK9AOJlMnX/bQEgY5MIxXhgbo/4PPsQ9999Ho9mkWq12PV8phWVZRCIRjhw5QiaT2auu7yq8+kelUomp++9lORVnYrlGyx0zmxlxu5esH0Z43bQvXcQqldCSyX3tTxAHgmS0Tj7rlMyTQcHRmUFlmEeDM9N1TWMxmyd35500Gg2UUqTT6aHJjDfCCMOMlS99keXf/i24cBGB7Nh5Cjb9LYZZFuw1nLL0Uc0p+vbNo8cwPvR9HDp+nGq16tcF6gbLsshmsxw6dIhoNLp3Xd4D5HI5NjY2SCWTzN15J+OrK2AGo47UgSEYHqRSNK/OUbk+T+ruEcm4bbQWlzAvXEIJCye9p40TjuYlULHZdPIKxLYOGYQAUxgs5Sa4e3KSSqVCNOaUpD5ozlMjjLCXaK2tce23f4PGI19AlkuAjVAKW0gQ7iKp3EJdwmJTDozmlY0ibhhUojGee8+3c+x97wOp+eaRbvD+PjMzc2DCU3tFLpdjfn4e07KQ9z+A9rWv0cIJBwUnwZXAduvGDO+64kHgdHF1YY7Q3DVSd5/Y7y75GHqSUTn7ElazjKYslAiWXQ2qRId7ACil0DRJS1mY2RRR1waaz+dfdjuEEUYYJFrLy1z9hV+k/uhXEKqFwnaS8XmazLaqVgz9YrCXEEqQ0yOsRpOc/Xvv5573vIdGs0mr2dyWYHj+F0eOHCGXy71sN0BCCD/KJDJeoAKEwMkOq5Q7vjgwSjEnjBX0ZpPK5Uv73Z02DD3JKD93Eq1hYwknFlipYDgaOBqM4TabKCEICY1iNEbsruP+30cOnyOM0B3Ktik+9hWWfvt3sc6fBWyUsp1qpEoGsjQGc2EEfw/+7ZUDjxToQmDoGs/ecQzjQz/A3XfdSaVaRdn2tk/FsiwSiQSHDh0iOUQ2/d1COp2mVCqRLoyzNDHJkYUbNJUNynaDoA/WWFJArNUifOYcrbV1jGxmv7sEHACSsXHxFMJq4rxyT2vh/fQ+wx114pViXsylGb/7hJ+Aa+SLMcIIW2E3Gsz/zSep/sr/jVIKU+hId0+hkHtaifSgQeAQDAyD517/eo782I8RicUpFovbaiW88uwTExPMzs6+YjY+2WyWhYUFQrEY107cjT6/SB17yPXiN4dA0Lpwnub1uaEhGcOdeEIpWFtDSYX01XueDwa0E43hZZqaEJhCUhqfIlMoUK1WiUajo2JoI4zQgebSEgu/8l9Y+6VfoGXrtITmbi6CG4sROiEUCGETkYJaOsOp7/lu7vmn/wyh6WxsbGxLMDyzydTUFIcPH37FEAwAwzBIJpO0Wi1C995DVUjENmakgwJbSLS5RZrr6/vdFR9Drclora4QXl7HND3nG2hPtNNNRTpcUEBIQMMIwV13EgqFsCyLTCYzIhkjjOBC2TbV0y9y7Rd/CXHyKTRNYBNC2pLNjcXwzvP9hiYlIUNwdWKK+od+mFe94fXbFjcDh1zYtk08Hmd2dpZ0Or0Pvd5/ZDIZ1tfXGbvzOPP5FIdWarTsgzvOhFK0Wk3MYnG/u+JjqElG9dJlrFrTtY1t9+KHm2gIwECyGEuRO3EX9XqdUCg0yo0xwgguVKvF/N89wsYv/SLaWhFLGoDtai83HfAkjl/nwd5rDhYCCEuBFDrP3/dqJn7yx5iZmGBlZQUpxLbZO5VSJJNJjh8//oo226ZSKQzDIJROs3jkKOHiCs2G1baaCNzK3wy/X7ESCkkLceP6fnfFx1CTjOYLL2LXq7TvYg6WiJE4RR9X0imOTE1RqVSIx+OjqJIRRgBayyusfeZTLP3GrxO2nKojGqBszXHyFI4Ww6tAMirVvgkBRKWgnExz6b5XceynfgI9FGJ5eRmpSbpp/r3okVwux+zs7IHP3rlTaJpGMpmkVCyijt+BePYZRKePX0cxzuFfgRS1a9ewa1VkdP/Tvw/1CCudO+OXd/cghtz/IgiF44/RMnTqx+8gnkxSvnGDdDqNYRj73b0RRthXVF+6wPyv/gbWVx8lojnaSoFyU+E4Il0oN1eBcFLkBLNYvlKhAF1AVNO4lsux9sEPcc/b3k61WqVUKjnmEd+07KYsc7N3hkIhDh8+TD6f38c7GC5ks1nW1tZI3H03JcBAYLrJuPx0K4ElZ7iHn6M7r12+Tn1pmdjhw/vdoeElGcqyMK9eQSqFjUS4IasHLQubJgXVaITosSOYpolhGCNTyQiveFROvsD8z/1fWJcvYGsWAokQthOeCtjYaABK+iYST4Nx0GTAIKFQRKWGlDoXCuOYP/kT3P+ab2F5eRnbtjvMI86D8swjiUTiFe1/sR0ikQgAibExlqcmOXptzs+E6hEN3+146MeeQlM2jWtz1DaK7L8eY4hJRn1+HlZXcNK7KoS7yxn6dxyAQ4IlG2GD2MQ4zWaTaDQ6MpWM8IqFubHB9c9+jtqf/Bli7hqW7ugvHOHtZO1U0t0t2iPTiAfl/i+lG5QjYV587ZvIf+d7mTp8iMXFRcQ2/hfeYjk2NsbMzMzI2bwLDMMgEonQFILrM7PcMzdPVVltY+8gDUMlFNbGCq2FBXjVA/vdneElGbXzF7FqJgSqFDg604NEM0AqQVkzyGbSmKZJMpl8xdtBR3hlwlxZYf43fpPq33wCIRSWdPTQIhCK7uwaFdLdQdoHxxC+a1AKDGzSRpjVeJoXvu2dPPDBH0ApddPw1FarRSgUYmpqisnJyZdt9s6dQghBKpViaXGRVjaHJnQUTZT7vCSby85BIL2mEIStFly6DLYNXaKL9hJDu9pVLl/GqrfQbJdYKBsbgfDLOA8/HGc1k0YuRyqTpVqtjnYSI7wiUTp1ktWP/BHVL3weDeHMZUC0sYhAmnDvz3bH768w2MompEkyMsyFqSkWvuM7uO9b30W1WnVKlXchDkopTNMkFosxMzMz8r+4DaRSKZaXl9HGC1TREGJTe6S6DMthhVfSS0NRn7+CWSmjJ1P72qehJRnm+UuIZhXpl94N1Fw8ALsaIQRSSFp6FHXHHUQiEarV6sgfY4RXHJY/9bcsfeQjqEuXQCg3p2I7ewhOaQF+ZESgKOZBmPYDg3BjHBJ6BEPXeeKee0j8wPfxwIkTlEqlbQmG539RKBSYnJwcyZvbRDQaJRQKkZiZZSmkk7YkdRTSbh91Q5+ry3VSNYVAXryIub4xIhldYduIuSsIZbpeGJ7bTWedkuGGEJJyOEXo8FFarRaRSGQ06Ud4xcCq1Vj4yEcpf/SP0TdKzt+kH4xKe5KtYd8j7j3yIYO6HuUbb3srhz7wPaTcKIitDp4OTNNECMHExAQzMzMjs2wPkFISj8eJZzNUc3km5kvULfPAWeg3HaQV5tXrNMtVIvvcp6Echc2VFerFNUdNpTqF0MF440IIbAkbUY3MoVkajQaJROIVlbZ3hFcu6vPzzP/Ob1H71ENIZTl+FignWgTYqoA+GPN6t6EQaFKRCYdYiSV46b3v5a7v/E6kprG+vo5SalsHz3A4zMTEBOPj4yM50yOEECSTSSKxGIszk4QWrgLmfnerZ/hht0rQKpWwV1f2u0vDSTJaFy4iSjWUje9849tsDwC8iH/dVlRSMY6MF2iZJolEYr+7NsIIu46NL36Jhd//fVovvIDmqZuVwhROaOomRtqLIBQQFhDXI5y84zjyAx/grhMnME2TRrW6bQSJ6cqWY8eOEYsNQ9DiwUQ8HicciVA9PIv55DedBVsdQDud21/dsjGvXnMdNfZvrg0lyWjMz2PXGrc+cIghUNhS0JyeJBKNYpbLI1PJCC973Pjkpyh++MOwtowQBpa00dxcNwe4JMSuw0aSlJK0Bo/fex+5n/xxxmamKRZLtFrbO3halkUul+PQoUN+vocR+oOmaUSjUYyJAkXbwlDQOEDJH9sh0LCo3ZjHbjSQ+zg2hpNklEqoZsMp73xAtX4SaEbC2EcOo5QiEomMIktGeNmitbzM/G//LsUvPEJovQToSEyUUig3NHWE7pBKkdAVS8k4z3z7dzD9rneSzGRc/4vtzSNCCA4dOkShUBhlEB4AhBBEo1Fi+QJr2QxTyys0bQupaCPIByGcVQBNadBYmmeiViU0IhntsCplNKsJUmArtZnpjwPihCMcp8+W0EkeOoRpmsTj8VGc+ggvSzQuX+bqf/g/aD5zCk1ZKOXt/pxwc6HEyDDSBQqQQpHVQyyk0yz/+I/wqje/hUq1SqVSuan/hWEYHD16lHQ63bXK6gj9IRQKEU+nKObzHF5dR5puHZNAttlgwbRhhQRaAmrFGmazyX6WwBtKkkGxiDRb2ELzX64c5jfaFYpKJEx6fNyPWR9hhJcbVr/2OOu/+MuYl85v1nhwdwLCy2szQleEBIR1jWfvOEr4R36MY3cep1gqY1uOw2EnwfDyXySTyZH/xS5ACEEkEiEUCVOZnESdv+SQikDl34OyDilAUzatSgW7Zd3y+N3EUJKM6tIytpBYwkklLt0yz3SorYYVQoCQgrV8nmPZHKZpjVKJj/CyglWtMvebv0X5b/4GUS0hkWjKCUq1XLIx2l93hwIiQqAiBt9857s49qEfRDcMyuWym4hha8SNl/9iamqK6enpkXlklxAOh4nG4qiJSSefiyZooZwECmqTR9ty+LXqSoAsrqA36vvaj6EjGWatSmVjnZDSIeCJ7idBCc7BIYVQAiE07MIY4XAYy6qNhMIILxu01ta4/Mu/SvNzn0JaTup/f3E8qH5yewChHN+UqJSU4wmu/MD3ce+3fzuNZpNmowFKdYT3OvD8L44ePUqhUBiZXXcRQgji8TiRsTFMGwwhML3FR23moRh6goFTtTg9twCVyr72ZehIRm1pBVWtOU/IFVwKu33aHYAXbElBM5MBwDD0Udz6CAceyrIoP/88i//5V2idftEtWCh9+7QSqi1T55BP0z1BGyEQgoQIcWFynOaP/zB3PfggpVLJzQXUHbZt++nBM5nMiGDsMoQQxGIxjGyGYiRMutpE2SbCrQMuvez3B+A1CBQ0GtTWNvY1IdfQkYzK4hKqXEEq0xFgB1BUCQGmMLBTTjrXUejqCAcddqPB8t9+mpVf/jCiUnQ2AchAHhsHB0D27j0EaEJg6DrPvOZBJn/8xyikUmxsFLdNX+CFp46NjXH48OGRJnSPIIQgkUgQisUoxULkKipQJXiTXIgDQTQcql9ZXCS7j7kyho5kWMvL6CWvZokJaAfgZbZDARYQijvJt0bx6yMcZFjlMgu//l/Z+LOPo6GwNSvo34ktDlDk1x5CARKBISWtaJzn3vkO7v6RH6bZbFKtVreV+ZZlIaVkamqKmZmZkRZ0jxEKhQhHo6yHw34FVk9b5/liHIQlyZmOErG6iN1qIUP7E2MydCQjPH+dcKWCUhKwOwhGNyk2fK/bsAU1CYnpSWzbHmkyRjiYsG2qL77E0u/9LpUvPgJCc8wi7qT0ScXwTcF9hwI0ARFN5+rULOvvfw93vvMdVKtVbNvufo5S2LZNNBplenp6VD11nyCEIJXJMJ9IIqWGbbWQXtXgA5AjIwihBFyfR9VrMCIZDlprq9BodBFcquMntB8U/N4rpNb5fafXtujyXbdjb79NgUADivEI8VwOIcQoCdcIBw9Ksfp3X2Dp5/8z9uoSUihQXiic8H9K5bln32xO2WytP3Q7863bd4OYw9ud12+bHmmQ/hERKYloGi8cO0Lqp3+aBw4f9oubdYMXPZJMJjl8+PBoY7LPSGcyyGwaKRQyEMMqtozrYAxV8PdbjaXOY293zN1Om5vzTSpFdWUVq9nct7yWQ0cyamtrCOHEiW/SCeXagBWoboFxivaKjjZbBVvwpW4Kha3fdTu2tzZNKWmkUuTjTkG0kbPWCAcJrbVV1h/9Kqs//3+hmlWE0pFKtNVDBq8QEwjlGAid3V23OaXYOlduZ751Eo5BzeHtzuu3TS8mRCGBmNSppxI8e9c95H/0h0kXCiwvL28rB7zokUKhwOzs7Mj/YgiQSCYx8zksYaPZ3hv2xk/nuKbL7zcbuyrwt87vbjXmuI02N+ebEgprZQnVbN3GXe8OhopkKNumtbbuhHKJDvbmx493Y3oeRMe/tzunG1Psdmy3827dphCSRipLJBod+WOMcKBQv3aN+d/4NZqf+SxKhLEJI4Xl1FjyDvKd3lQgnE90RJR0zpHOOdXL3OzW5u3O4c42Oxf6ftrc2r6tBCEkOQ0WMjku/b3v4e7vfB+maVIqlbatPWLbNqFQiKmpKSYmJrYcM8L+IBqNwvg4TctC88JYPXuJvxZ1LvCepizgxNE2/m82loJjcKfr1uaxSthYy6uI+v7VAhsekqEUZqmEqFbbaYMf3tUpBGB7gtF5fLfvbnZe/20KQEiNRj6NrusjkjHCgUHlhRe4/gs/j3j2WSQSS7XaxGj7vKTzL45sDf5+y3/f6vdOreUg2rzd727/OBsn90VUC3FmapzShz7I/W97G+VymXq9vi3B8KqnHj58mJQbiTbCcEDXdRL5MVp2YBQqcIzhnmqjU6slAQs/9ER1kgYPwXE9qLVpmzYFaBsrWNXq1pvcIwwPyQDsYgmtWqfVVa14UEwOElsK9PwYUsqRP8YIQw+7Xmftsw+x9nu/hz13BRWI6BpFjNwcCkVKM0BKnnrN/WR+8B9wz/HjrK2t0Wq1utYV8fwyCoUCExMTJBKJve72CLeAEIJELkctniRdKdOwWk40kJIgPI3FpqnM+WnhEwol2TSn7R+EEliWTaNcYr9yTg8PyVCK1voGdq22perdQYJE0ZSC0HhhaEiGl+zHU88Gf3ofD0II/yOlbPvpfb8b/fM+3fq3n317uaO1usrS7/8BG3/yJ+iub4CCNu/5VwrRCC4Xys2+aQf+3Xm0JiCr69SiMU6/+10c/+D3E40nWFlZwbbtrgTDNE00TWN6eprJycnbLm52u3PYmwOd8yP4GeH2EEulWM6kyNRrCLvD/6FtUgSJhACl02Y2CfgS9jbGXP2IEP6VhRBddfrbwyE69fWN2z5j0BgakqGUolVex2pVEdiobQSb796iAq44ytlR7BQC4dQdgS0v1fn+5vDcfGrKQk+ksCyLRqOxZbHcUR87FtpOIeL3xb3eNy8/RbVZY6WywqNXvrZtu3ekjnIkeQgB3D9+H7FQFE3T0HXd/2iatuWa/UIpxYXlS8ytz1FpVHno/Bf2rW9BgtMpwPtFt3fUea1uJG+30G3cVM+cYfF3f4fGFx4BoK4kljuTlCs3+yEYAoEUzlzYTijudB57c9XbL/YufNv7cvYdR2lEDIqZGC/cP7XtsQ889hIAEy8tc/TqGmeSaZa/+3u47zveS8uyWJ+fxzCMrsTXNE3C4TDT09OMj4/ful/us3n0pa+ilOKZ689xaeNK12Nz4QwPjr0KARxOH2I8Ueg6R7x50s/83Yt5sl3fvGsEr9ttI9LrtYN96CRl8Xici+kU5vIKsqG5NXm8AATwVwnPV8P2RqDFphbDI4f27Y+xR8+BEBTOLZJ5aQVNCOcD6EIg3X/fzphX3jHF/UstPjQkA6WwN0rY9RrtYUKbKE8k+JOffvvALhlqmJx44iIEiMTMM3MkFkpInCx93gu91Yvd2rfH+fTTjw+srwA/eNcHOJSa8YVGKBTCMAxfgHiT5tT1F3j2+km+MffUbbd9oXiJC8VLAHxh7lFSepJ7Uye4L383iUiCcDhMOBzues1ehNVaZY0nLj/JZ889suO+fUvhVU7inEDfQqGQL0xvp2+/+uXf5MLapdvuy62QC2f4yft/BCkluq5jGAafv/glvnptsGOhF3QbN7/1d7/KdUrwrjC867t3fI1Esca3/adHEDiCUBcCQwh08OeP9yZ2Mo8fePQcxz/7oj8/jcB1pLdo3GZb8/dPcPG+SU7eROB34uTb72z7ObGa5J6p41y9ehXbtolEIpim6T9rT1NhmiZjY2NMTk6STCZveZ1T11/ga5e+waml07fVr9XGOn8396jzyxyEZYgH0w9wb+4EyWhyyxzpdf7uxTzxZEtw/gKsV9f5uYd/fmDXjmoR3jz5BmecuPd/In8n2VgGKaV/fT0UohWJY1lg4yTRb4+xCpAbu9MR1CEY8/dPcvHeSU7eP33b/Tv5jrucf7hj7PiXzzL5zDWS8yVCUhIKjnt3XgXnVxuUU2CUtbUentBgMTwkA5AbRWStNgCdxO2hGdY3X6iL590Xe+wbl8ifX6Lw3PyWl6q76qttX+wuYXHxBrLkVgqMRonH404xn0gEIQQ3iot8+vTnOLl4e4LpZiiaJR5ffZLHV5/k7vBxXpt7kFgsRjweJxqNEo1GfaFwOzujWrPGQy9+ni9efGygfXtV5F4eyN5LPB4nFosRi8WIRCKEw2FfyN9MJb1YWdpxf4JYbaxz/vJ5YiGnH7FYjGazOdBr9IrguImEw9T++hMs37kOocFFzpdTUa5GJJFSg5CQRKUgKiURKQlLCUqhDUBVv6EUC80WhhREhHONqHsNA27rGqWJBN/8jrs5d7yw4/7cyJW40fga6UqCI40Zf254Y9BbLCcmJjh06NAtHcHn1xf4+LN/xcX1yzvqV8Nu8sTaUzyx9hQnQsd5TfZVvryIxWI9z9+9mCde3zx5tluZTmtWfZOQufj81S8B8Lrsg8wmZzicnqXeaNCUOmbTpGpbGFIglEIKEMLzwcC1LQZ8MISiNBHnm++9dyBj7Pw7T3D+nSfIP3eN4x95nKjUiElJRAoiUraR7M4NsEOFJI3S+o770S+GimQ06zVwzQvdYO8Z/YCLbzjKxTccJfnWFY79xdNk5kv+S424bFJ31VgdnHbXcPnKFUqySDweJ51OMzY2hlIKIQRnl17i95/+41257pnGeS7NXeXV2r3MpKdJp9OkUikSiQSxWMzXHgTNAkFcWL7I//uNP6LcHLzK7vn6aS5cu8SDxv1MpMfJZDKkUimfDHm7tm5CtN6q70qfXrhwmoyeJpFIkMlkqFu1gV+jF3jjJqwUrce/RvSbj9P8n9878OtcGIsSXyoS0yRJTSOraShNQwDhgF15J/N42bZoNOrEpCQuNTK6RtpdjISUbabObrj8hlke+p5X9X397bARL3PSOEthMUNWz5BMOtqDZDLJ0aNHmZmZIRQK+fO1G5688jR/9PRHB963s83zXJ6/yqvkPcykp/05kkgkiEajPhnabv7u1TwZGxsD8DcGHsnYC3OihyfXnuXJtWeZ0As8GL6XFdtkrdWkaVuE0QgBuu1kc91c0NsjSC6/bpbP7cIYW3n1LOv/a56pP3uc/LllUppGQtOIuyQ7xCbJ9t6iwrHktMojcwl2q0WjUgHbRjA8ZdFKh/M896+/namPfJX803OkNI24JkloGhEhCEmJxt6QjHMXzhGthslms0xMTGBZFpqmcXb9JT514XO7eu2GbPIN9SzLF1Y4lJlhbGyMfD5PJpPxtSme5iAoqHZLcAZR0Wp8vfUkxy7Mcig7y9jYGLlczu9bUKsR7NtadX1X+nNm4SzR1TC5XI7JyUnK+f0ttXzuwjnES8vIF06SXS9SuGdyV65z7VAGnr5EWtMZM3SUYSA97Z+mOebGHV5jybKYr9VJapKcrtNSBhLazDPbFQU5985jfPE99+ywB9vDDiluTK6yfmaD9GqSyclJZmZmiMVitFot30ei22L+2Etf5S9OfWLX+taQTb7Jc8xfXuCO1aNd529QqxHEXs0T27bRNA3DMHyT7H45qt4wl3jIXELPKwqNqqM1U5AQEJGbcr9dm21z9h3H+OJ77t61flnpKNd+8m2s/96jjJ1dIqfrZHSNpOZoN8Jd+iWxsTZG5hLMWg2rUnEdv4bPA3r+x97Keunz5M8sktN1mko5hAMISzkQx9Nb4dr8NVi0GR8fxzRNYrEYdlTx2ZXtnSYHjYupq5TPFZnemKZardJqtfyQvGg02qYxODV/etcJhgdbU5xPXqV4ocjsxgzVapVms0k+nyeZTPpCNCi0rm/M70pflsur1M+V/fck46F9nWmXz5+l+siT5E0bpRuIw9lduU4xYrDWaJLXbSwUESGJeapdN8HeTmf2ulIsNxpkdY2WUoSEQ/gTSmHdpNLk5TfM7irB8KELaidMGk8tkcvl2NjYIBKJ+LvxbuNwtwlGEHOZRSqXyxzeOESlUqHRaJDP5/3+dSMaezVPotEoiUTC10LeTOuzVzDflOH0U0nGziyT1nUamiStb8p9lPIX9Muvn91VguEjbFD6qXdQ+pXPUblepK4MTKVQuiNkwq5GL/jk7OIougS70cCqVUGIbSNL9hvlH3kLxf/xY1Qsm7ptYxoGuC92L6KhV9dXqc1VsCyLaDTKxMQElzbm9uDKAeiCq/o85bNOoiHLcuySnd7hG7UN/uyZP9/jvsHC2ArVMxVqtVobAfLIT9AZr9rcHTOGSEmuXr1Kq9UiHA6TOTq2rzNtZW2ZYqmCretEhCAe2Z201cZMloVWC1MpQkKQ1XSqtkVLaVheFMAOF42qsrnhX0OS0S3qtk1rmxBAgPJkgi+/994dXbcXSENi3iO4dPoSlmXRarX8BbNzHJ6aP71nBMPD2nSZyvMvUqlUaDabWJbV1r9OErRX82RsbIxCodDWp2HAxj98M8X/6c/JWzo1XcNUBkrXUUoRcWVJZTLBl967ByTWhYgYNP/RO7j47/+SmmVjhmwnUkvXnYykUjqBC7ha9mp5z/rWiaEiGWa9SlMKDGv4NBkAWjrG+odeT/XPnqDpTgAJSF0PZCbdPVRrVVZXVohEIpRKJdbY4Ia5suvX7YRxKMJLX38J03RqzHje4d5HSslnTj+8K3bcW0FL6VzXbtB4qeETIM9bvDN8b6WyO88ulIqwtLSEYRhkMhmizSTsY72rurLZME3CQlCxbRq52K5cJ5xNsGaa6AJSlkbJsqjZiqatsAfkw2cC65aJIQRpe5NgWEq5IbBb8fw776Qe3ltRp6cNFowlmueb2LbtOzJ2jsNPnvr0nvYLQBiStUyRxrlzPgHyfCC6hZHu1TzZ2NigWq3SaDSGimRo6RirH3wtxT/7Bs2QQ9A9ue8t5Cf3YYyFJtKsfs+rufrJZ/1x7/XHJxhCoJSNtVHa074FMTwko1JBq9XQ7OE0l3hIvO4Ozv3hY05ODAGGEISEJLwHE8K0LGq1GvV6nUajQTm0P6litZDGuixy+fJlDMNo8wwPh8MsV1d44tqT+9I3AONwhIufuejvGIPhe0Gb+HJldVeun55IUy6X/Y9HxvYLFlBXiqZSNJSNyuwO40lPZqjaNlXLpmrb1GybprIxUY46dwDXsBU0bOc+mrbTrqW8gMGtVyhPJnoKUR0koicSXPqrS752IBgCrmkaz8w9x40BR23cLhJ3pLjy9EsopXw/iGB4a9Dxcq/mSaVSoV6v02q1ME1zaEgGQNyV+6Da5L4hBPWZ9L6Nscw77+Wlv/iGGzqO3yddCKSmYQAWAlkZOX5ilmuoukK3dZTYyziS3hDOxGh8yyzzz14jJAQx6didk9uUcN5NNGL9VdY7+9iLfO6XP4WmaXz3//R9HHntsZ7bkFmN+VPzxONxMpkMuVyOdDpNPB7n+Run+upXq97isT/8Iqceeg5d13nPf/d+7npb7zbO2FicF9YXMa4avp3Xi4TxnEABLq13T2w0CGSP5mhVWjSbTUzL4uYxD3sDP9vgZHrXrpG7fxbzzDwtl9A0laKlVFv9073EpVfffn6CIFr1Fo/8+md46atneeB9D/Luf9J7NE44GaE5bnH16tW2kHMvouPs4rm++lZdr/DZX/4U109d48633s17/vn7MPoxgU1pXD17lXA4vGWjEPTN2Kt54m2eWq0WlmX5ps5hwHZyPyIFCzsdY4+9yKu+8zV8ax9jLJKJob3/ARY+d4qIDIR0u0EJOk7iOlXfv9ol+y/5XNj1BqqlUGp3tBiXnrrAh7/vP/FrH/j/c/LhZ2nV+y99KyfSrLRMlk2TVdOkaFnU94F1W7rV8znV9Qqf/aVP+sXbnv74N/q6digdYW1tjeXlZZaXl1lfX/dtvF+58vW+2jzz5Rc49dBzGIZDDp7+4ycwG/1pAUKzERYWFpifn2dxcZG1tbU2G3StWaPS2j12P3bnBKZpOteze39PuwVtLAHh3SslPn5iGktBSzlmjJavbVDYA9Jm9IILJ/qrbHrmyy/w0lfPous6Fx97iYUXr/fVjpELc+PGDebn57lx4warq6uUy2UajQaX1q/21eaFJ17i+qlrhEIhFk/Oc+2Z/nJqRMai/hxZWFhgZWWFUqlEvV73NQn1Vn3P5kmj0aDZbPqajJ1mEwVX7n//L/DrH/olTj383MDlfs22uXT3DsbYYy9iCMGFzz3LwulrfbUTOTzGitliqWWyYppsWCZV26Zp2241FYVq7F+unqEhGVa9jm2azgNR2392AuGqkb7xmw9z+rNP991OpJBiw7JYN03WTZOiZVLbA9ata5qf6CeSi2JrvT+PayevomkaoVDICV1TYRql3ssA6xGdSqVCsVhkY2ODYrFItVqlVC1RafXHmotLRTRN8/MLjI2NUV/pr61QOsza2horKyttJMiz9+5WWJ6HxFiSVsvZodnW/pIMDSdPRVgIIrP5Xb1WYjyFqRw/jLrtOEg3bZuW7TimDWIeBwvAKi8VeRdZ0UgY3Cj0V3zMG4uhUIhoNEqj2F+p7OyRHKurqywtLbG4uMjq6iqlUolarcZSdbmvNpcuLbbNE6vcHxFPTKT8jcLS0hKrq6sUi0VqtZpPMvZyntTrder1uqP9c0nGICClxDAMnvzTxznzhRf6bqeb3C8lQqwUbp3BtRuKS0U0HNNLTEqaq/05Z+aPT7JuWqyaJistkw3TomxZNGwb07bdMgH7ZxsYGnOJajVRdgPHtQt2wy9DA1/Vtf7MJfi+N/bVjg3UbJuybVGybCqWTUP1PyFOPvQsf/dbD/tx9J2puz21ZSgU2oxtz/QnPEvLjvCMRCKk02kmJiagZkOP8ySejdNoNKjValQqFd+eulC60Ve/ADYW1tB1nVgsRiaTYXp6GsPsb4hGszEqlQobGxusr69TKpXanMpWK7sbN57IJTBNk1arxcbjKxxbPEY+7yzwy8vLzM3NcfXqVZaWliiVSvy9X/pQX9e59NQFPvF//nlHmm23xgGO81dISj+WPr6LphKAeD6JqRztRct2TCUtpbAAaxvHzN1CMx7q+9xGpe7Pk1QqhWb1tx/TwwalUon19XVWV1dZW1ujXC6zXOqPYADUSzXfFyqbzRLT+6uvGUvHsAyb9fV11tbWWFtb8zUZXmTWanXv5kmj0fC1GaZpDsz50yNkiUSC8sX+Qzm7yf1Kon+tYKNSR3Oz1qY0Da3Wn5ZFjxhUbZuSZbFhWRQti6pL8ONumO0uGQhur3/7d+l2mGYLS5no2KhdUrBowsnYmdQ0wi/173AVzycdm7OtqNk2dWXT3OFc8Nh2Z7rwYMbKUChEMplkfHycdC5NiXpf1/JMJalUivHx8b4qxSrw1ZzeDqTRaKB2oJWrFWv+M8hms0xNTRGPJTD7sObHs3Hq9TqVSsV3LPNyZ5imyUqPzmyteov1+TUKx25d2AognotjWZavBm61HAESCoVIpVI+2QmHw1R26JTl1QuJCCfVsJOUx3FM04QgJARx6SSvihdSPbXd831nE9g4xMJzNPWjP1yNw16hNNZ/CfW1uVXfaTiZTBIS/S0msUyMer1OuVymVCpRLBad913uX+oH50kmkyEV7u2dtvVvIk51sUqpVPKJeBvJ6JGM73SeBDUZgyIZuq77zypJ/2OiU+43bJvKTseYq2FMaRqhWn/CM5aJ07AdZ+uKbVG2LdfpWvmh3fuZbmRoSEajXsNstdDV9jqMnZacDmoysvrObt3ydmuuatjcyWQQ+LumZDJJPp8nl8uRSqX8tN1eIaF4PO7sXnIJSvSXI8MTnolEgnw+T9gI06B3Fm1Zlr8L8QVDawcVGdkkQN5ziEdjbNC7GjGWifu7o1qt5msxPKeyXklGdb1CZa1828IzPZn1n41nY5ZS+qnOPXNVPp+nVttZHoLguE7rkoymE5eSmJSE3NoG3m4pnE/Si+K/1/vOTGWxFP7caNkK0/3YtrOj2uk87qzN4FXTFt7H/a4V3ZnvSdBcEov2H/arpwyf8FYqTg4XGv1L/c55kkwm+9YQGbEQpcY61WqVSqXizxPPXNFr+OpO5kmnX8agnD+DmgwvdXm/sNSmdq6hFM1Y/9oycDYHYSmd+boDJhCaztC4UaJm2dQsm4Zl07KdsG4bnCJp+4ShIRmO+7lXZKY70dgpGfN8MjyBuxPY4KuFLaWwdqgIDu7gp6enmZmZoVAokE6niUajfnx9OBwmHo9jxi0u1XonGeF4GCmlLzxTqRSGbvRFMmzbxrIsn2yYponciUwIhJzGYjHS6TR6ONIXyQDa1LDeDskTXsvV3oRnZa1Ms3r7Ow0jYhDORvw+eCngE4kEkUiE8fFx6vW6nzX1FC/1ens+NFeLkdQkY7rBlGGQ03WSgdT3IXfcL2bjPZGMfu47Mp3BWiz7gtjscPoc1KZKBH52fgCaOyAZAvx5kkgkyKaztPqc4+FkhFZxcxw2Gg1UawfaWjfnhjdPErE4pZ7e6iZCsZC/uNdqtTZzhW3bLPdIMnYyT3ZLk+FpiT0z7E5gs0k0bFth7oBkeHV2Qq6GPbyDmRHJJjBvlHwNoqc99Byu5T6mhRgakmHVGijTAiV3vtXZBtL9GEIQ3aH+yCv467/IHfYtqMkYGxtjdnaWmZkZZzcfj/ux694ivGatQx/OyLnZvB8D78XuS613gRfPxH0HO49o2LZNy9qZF7N3j6FQiFgshtL7I4NGJIRt2227pKDXeq8ko7xSplnrTZBnZrI05+pbNBnZbNY3UXnRLqdWdkAyEISkYxLJahrjhsGkYZB16xkYbihbSEoWMr3tyMsrZZrV3u47e6RAxRV4LVdtawbmybCGp3fC02Tk83li0Sgb9OeErJTCNK22cWibO5M/3hwOhUKEwmHYAcnwspIGTXvePFnqkWTsZJ541+8MY90pzwg+q1gsxk7csD25bytFS6gdtQXOBsEzaRqi/5pd3lrkaRDNznVpH+0lQ0MybMtyS+ZqsI1eYBDCSQrhF23qF/FsAhWI/beAnWr1vEXfY9uFQoGpqSnGx8dJJpO+X4a3CNtl1RfJKBybGFC0jvDj/YMJfKwd0i3vHj3NjSVFXwkWjIjha1o8ouHtjmzb7tmzv7Rc7LmOQv5YgauXL7YlFzIMg0QiQSqVIhRyiJBt27CDpIpCbKpdE5pGRtPI6zpjuk5C0zACWQCrqd6cBEvLxZ73QGPHJ9l4/DwtpfwkYC0cl+5BhrGqwM/gx0N9h+nTNTeay8kZ0b/mM5wMU10p+4u5aZqIHaRAFQh/nkgp0XbQNwX+JiE4R7x5sljtzXdtp/MkqHEcdISJRzRqfcqoTbmvXI0GWDs0yXnaDOnO0X5JSzgVpeE6WLeUwqR9vskRySAwMJ3Y6N0iGSg1kBTgCvyF2vac2nYA0WEqSKVSZLNZ8vm8Y9IwDP84IQT0qaUzIgav/4dv5sojF3zHzbs3jlEoFNA0jfX1dc6fP8+5c+f86IdKpbJlskspyefzZLNZ33ckHA5jyf52VLApPL2PlHJHlMVbwD2i4QvOcu9Ov8XFjZ6FeSKfbNOkWJaFlNK3D8diMYQQO1YJBzV0ESFISEkq8DHcAk5rd/Vujy4ubvTshp0YSzkqZTd81fuYto0p5WBCWKEtbBVXZii1WYhtR8ItUF9E7LCwWzgRoWwX/cXbsiw3s/FOutc+V3aCoDYyOE8Wir1Hiu1kngQ1GW0mkwFptgfyrHDDpv2xN5Cu7diYEUlGWWdTs24F1qURyfBh43Avi2FOK+6hc/e007EW3MF7ZoxIJOI4ncViPsnwMB66Pceqbrj3PQ+QLmRY/uK8H3nhVSrVdd032ZimSSQSoVqtdiUZnuq/UCiQyWSIxWI0tP1NoR2EUqqNaHj/Xq/3Hsa2NreKofe2a/E85z3nT49oBNNMazv0DfIghaMD9MJYQ27mv4hLMgAaY72nE1+bW+1ZSMTzCSw2nT8bbn0RcwA5Mm4HCohKuSPBuoVWDGAh98aj42swPNksO/vmJcHa2Id54kWrBf1C1H56LXZg0HI/iJ2uesE+2QE/qEH3s1cMD8kQCoTtfNh0AB1mBF/ofrzGw6lZrhT7yxI3/epZ8scKVJ8rsbq6SjQaJZ1OI4QgmUwyOTlJJBJhbGzMcVTrWByEEITDYVKpFPl8nvHxcVKpFGUGU7FxECWeg8LT+yil2GgUe2qnVW+xcPo6E0cnezovnt39HAC9oJztjWS06i3mn7vM+KHeEnjFswnHPgw0g45o9K8Ovn0o4pqkpse4duwEm3l3+segxmJwHNoDevc73Z1LKYlEIl1r+/RKxndtnuySj16/aJf7g8EgVjtPk2cHfJ+8xHWj6BLwn8hwDaftoTp+7vU8EEJwV/543yQDHFtx+G1hVtdKyPUFAL+2wsTEBLlczt/ZdAozz7wTjUZJJpO+aWelvLvJe3pFt4yxDas3k051vYKUkuZGA7Nhot9mtcX0ZKbrDs1zatsNkuGFdMrgx/2uGeltulfXK0igfn0ds95Cv00fh/RkBtvVXHhJufwQVlsNLIQ1GE2i3J9ZPUwxkeT8934v099ynAuLj+/sYgPCoLMXBxGphXiXejPxuJMgb25ujtOnT3PhwgUWFhZYX1+n2ezukG0Yhp/gL5FI+D4omqZRr/WWh2e35ok9ZJoM7+d+awi2IOBm4PfNBmmrEclw4LjACARCbe54pCtRFINdyAemJ1GDGWxBdWUw7LJWq/kl1DvtsPcUTvD5i1/a8S3IrM5qtkxl9RKzG5OkYik/CZF37c7so7DprBqJRJxQukQCo7GzuPFBolOQe8J9vdHbDq2yVvYdbnsRngCTr5rBXGh2DaMdaAGo2xiAxUxvTp+VtbIf9m3Wm7dNMgCm33qC2hMXaNqbCbm8QmmD8IkKQgEhpchKjRfuux++6/3c/frX8fSNZwd6nUFg0ATDg2eCE0L42sVarYamaaRSqZuSDC9/xNjYGKlUyicavWoydmueMBiL4mAx6LVogO052otNurHfEV1DQzKcwmgSYWtINtOgCtt5CZYcXKEVP4HPTqHc4jOAvQPW8sB7H+SB9z7o/16mxdd5mq9ffxrcukyvHXuQdx95h5963DAMcuEsdySPcqF0aUe34aGRMznPNXJLSWataT+XRrA6o0c2OqNAvFLRujY0Q4pYLOZHwHhESUrJQnmxp3bKK2U/1r5ZahBJRW773OREiqWr81vC84IFoAahjodbE+fF8d4yQ5ZXykhAR9BYrxDpoUR8ajJLSW1m/mwq2w+rG+yaoQhLiBlRTj34Gmb/yX/D+Ows5XKZljlc/kEeuRg0yRBu3oxQKEQkEsGyLGZmZtB1nXQ6TaVSwdzmWQSzYY6PjzM2NkYikSAUCnGt1FtRuN2aJ9IYPpahUG2J33bY2EAQNI94Ceq89FOW2D8foKFZEYQmQbh+8src+tzV4CbnoFndwEjLTVAsF7l69aovSOLxOKFQiNeNPzgwkuFhtVCiVH+JI8UZPxrC++llrPQWbE+z4YfTDciR0cO9xl28LfcmJ+x0aYkXX3yRM2fOcPnyZb/uh7VNAbKxsTG/BL2XBMswDNaa6z31oVlr+Du0Vo/F5BJjSebNq354Xqe92VwvsvAHfwDHemq2KzwBs6XIoPv9Rg9CH6BZbbihr9BYrcDR2z83MZ7y4/Ybtk3TdohGa4D+SwqICA3bCHPynW/lgQ9+kHg2y8ZGEaXsgZG3QWK3iIaXZ8f7GIZBNpttKwrYDd68jsfjvgYkk8kQjUZZrvUWV71b80TfyQ5uF+C9usGaSwbZ0maklReBZY8KpIEejSJ0wwm5CYypIDsb5GMa6CPfg/e3ur7K2YWzfnirt3hOxiZ4bfbVPLX23ECv14pYvBS5wvpGkbtbd/p+GV6iLE+b4mk1PAxasAdzg8TjcT+HSKvV8ut+bCdAvSJwk5OT5PN5EokEZh/RL0sXF30NktljxUuvAFSwNoO3Q6ucOce1X/pV1PNPwf/5/p771QvKU71Xily6tIh0kwU1V0s9nRvPJzGhrYaJqZwQO32H88XzvwhJSTGeYP7938lr3vVOjGiUarWGYQyNWHPQcb+DJBhBMumZTAzDQClFJBLxNQLbXTMY6RSPx0kmk0SjUapW7w7cuzVPRJ/F6fYCw+KT0dmPtpQKttp7p8EAhmc2CkBaCOE5xA0Xe91vLK0ssfbSEul0mkKh4Ne7SKVSvGH8tSxUFrneXBj4dZfT62xUnuGu80eZKU9TKBS21FXxiMZu7ByDxCaZTPoEIxKJ+Hbn7XwbvJTQ+XyeyclJstksptZ7fMPa3KovPDXZm8DLTGe7es4vffVx1v/sTwldfgljD4zOpT5Knq/NraLhCIle0xJnp3NOrgxUW4SJqRQ7S1/kSAYNQTmRpPWWNzF+4i4spajX6ySTO2394GD2W4+SedMEl5jnUmservTexusT38Lx6FE/usTbOCzXeqvtA7szT1qtFro9PMsUbEYzDwvBuBUEg9ec9YKheXtGJISmCZSwQQ0vc90vbJSLrJxdIJ/PU6/XMQzDX+gjkQjfeejb+eylR5izBk80WnGT0/pLrF1e42j5CLVajWaz6ZScd6vFek6hg0YwiiUUcpxKo9EohULBr/uxHcnwikglEgmy2SxjY2Ocr13quQ8Lp6+TyWScHWK1N9tmLLM1B8DCZz6D+bVvkC8XCckBVAy7DazMpHs+Z/65yyQ1DV0IrKXewn5jmTi2oq2IYFPZmOy8zg9ANR5DHj9OPJttyz/yitqbDGDhaDYa1KRD1KWUvrPl9eJ8z20Nep54hdqsQTpIvxIh2HGyyJ1geEiGbiCExFagOW41+92loUK9UePatWs0Gg3H6TOXo1qtYlkWuq6TSCR499Q7eOLKNzkrLw78+ioM18eXqV7eWtHUU83qO6xs2w0eyfCIjK7rxGIx3wP9ZqGgnhOal0U1mUxi1y/0dH0vLM+7tuhRi+wVgGq1WtRLJRY+/WmSN26QkRJbD2Ej92SsN3pMf+yFr3oJvuwbvZEMI2IQmcn4RZu8CBPHZNJTU1tgRSKE8nk0Q/fTtfsI/nMfd297gUHc3craCvqK9H0ybNtG0zQq9UpP7QxynnSGsW7ntLpf8NMW7Gsvbh8S0Uf5y8FhaEiGMEKgGQdHB7XHaFkmKysr6LpONpulWq3SbDZ9oeDVwXhV7X7CV0KcVucwE4PdAQhdsH64zNnT59qy8nnCPBKJ7KrnvKfSjcfjbemPb0YyPH8OL/ql2OxtsaysVfw+6LqObPSurZm4e4qN5xdZf/45xptNmuEorVAYExs1sJipm6P38NUKArfODwJ1rXf1+eQDh1ldOBmIMHGzfu5kkgvQQiEnousm79627V0hvUOFAUy1y9eusLB03Y8uabVaSClZKvdW22dQ86R0Zn1LhIlt734Kt5c77FEVVpDhCEIzAA1bdY/sHajD1ABmqNeGYoeC83aupRTVanWL57VSyi/77Kk7TdNEv6JxZu4lqjM7q4raCWFIiodrvPT8S1sKGAkhBpv7IdCuV9clHA77i8vt5BwI1kGRUrJS6y1ZWDD23zAMQhhIS2Brt/++c4fzLD7yPC0pqYfCNJXjEGnZCkvY2AM0MynXQ7ozuqSY7idHhiMgdAHachkaLQjfvkYkf2ycRQUt24swsWnZNq2dOn7e4p1bluUsmvFxWN/ZtV7uWFxZpH6uwtjYGK1Wi2g0SiqVYqW+D/PkyBhrp5bbzCWOPBsykuGn6x5MfEl7Cq0BwZcBgNp5tdidYGhIRigaJqTp1JSORov+i97uD/ZTdeaZBTwth+dprs/pXD57mYXsCrIwuFdtZEIsRRexzlv+9XVdR9d1LGvwqs1ghMlOnUuXei7xXtpSU0Yqid3DtA1FQ45Tm2HQCJgOrD1MlLPRY/XV8nLJrw6pC0FISES9heqBZISiYSxUmybDSci1O3fskY/p6WlmZmaYP9dbPpSDhkFsbIrlIovX5v06RZ5T+XqPYd6DnCedhdJ2Q6a80mDvoFLvTjE0JEOLxjG0EC1l7Wu4zbBCCOE7eXarM9DpeBn0R0hcv86V09doHLHRYoMZbPpUmMuPXPYjP7x+tVqDt/552VA922y/moxio0Sl1ZutuVFptCVAC4fDhO0QZg81WuL5hJ9iu6lsGsp2k1MNtv7BdihPJWn0kH0RoFFtOOGr4JIMgbZexUzHbruNeN6rYaI6apgMOOOnOz40TePw4cNMTU0NtP2XM+rNBqurq4TDYd/Pa6NepN5j6v2BzJNcfJsaJkOmyeDg+GMAaAqa4fC+XX9oSIZIxLBiBpImao+c4YYFJx96lr/7rYf9SeqFkgVTeeu63rW0eifRCIVC/r+DTo+JhQTXL1znRnQZ7fjOB1xqJs3ZjZOErjpZQROJBIlEgqYYrHnmVOMMnzr1cF/n/sxr/0WbCvdGuffS1c1Ko22HFgqFiKgwlR6EZ2Yq21aVtGlvhnNa7L6DYmm89/DVRsVLxOX4ZBhCoK9XMY/cfhuZqSwm3e675+7cErFYjDvuuIOxsd7L2b+SYduWb4b1TLAVu9pzO4OaJx7JaHP+NPfTbXErBp0GfLchhWQ91FsivkFieEhG2IBoCCU2U7YCKF89Ptj8asMGT/PgpfH2QkM9jYWXjCqfz7eVVg/mqfDIRjDlt5caPJFIkEwmSS+kuXbqOtVDTbTUzl5/6FCUxSuLfnKwfD5PIz5YkrGTBfjKlSttRGu+1nt47/rCeps5KBwOExNRVnow9scycaLTbqSFvekAaarNiokDQcej8jLRVnK9l3jfWFhzo0sEmoCQEBirFXopmbXdfQ9Se6OUIplMcvToUeLx3u/zIOPKFy5gvlhndnaW0BvjqPRgNmY11VthNBjcPAlnIls0GdYQajLgYCjcBY7p3IrevgZy0BgekhEKYYXDKCH8AkoOY3RrLIoB+scOKAGax2jFTvmP2EwLnEwmyefzWxJeebkivGRchUKBZDJJOBx2BpEbyhksJw1OAaR4PO7/7u00Fi4ssJxdRz/Sv1YjMZnk8jPnWVlZYXl5mbW1NZr68AiEM2fOtOXJWJW9ObMBrFxZatuhGYZBWPb+zDJHClQXiu6O3jOX2JhiwAWzFCih3DnkfBrR3qf5yqUlhEswdASaBKPWO4Hset+DcJZz/Y4SiQTj4+O7kqNl2BHUWGpSwxwQXTXp3QdiUPMkPZPBnG+2O38OmSZjoFABzcggxEAgRbZAgW2DEGjG/hWuHBqSIaNRqiGDqBCOO5wCgQBlu9qMQWoyhq8tT4uRzWZ9x7VCoUA6nfarIhqG0ZZW/NNLj7A2t97TdQp6nuOZw07ynWXJ8tk15In+BmBqKkO5XGZjY4P19XWKxSIyPdiMiztxbnv66afbMn4u5noPwywtFMlkMpv9UQrD0nueOanJLEU3A6bv+KkUtmWjtEHRZ0ewCF/IODuZpcneE3EVr62QdOvQCDdiRfaYkAu23nfTve+dQCmFlE5uh3Q6PZQ1SvYKXmQZ2uBIRoly7+cMaJ4kJ9IsXrnenpCrNYyOn4NkBcGfg2ivvS0NgYqOzCVo4TCtUIQkYCKwvfgSqUDZODVYX77CJKjJGBsbY3Z2lpmZGfL5PPF43E/3GyyQJld7371ZtpNbIhQKkU6nkSXJ6mIRe7z3ZxvLOEmxKpUKpVKJUqlEtNW7/f+m2MHcO3PmDJmMQ4QASsne1cA//dF/1fb7CmVWeKnndhLjKd8vo+F+moOOLhGO4s8hFw4UvefIAPjnf/Xvtvytt8wJDoL33fTuvY92gpBS+n5Lr2QEs+G2dGgMKOVSVdV7FrUDmydjCa53OH+2WoM1wb7SoAtQ8ZG5BCEEyXQK0wgj6zVHpSocjgG4KXSHyQjW2Zed9c1TM8ZiMb8I2NTUFOPj4ySTSQzDaMuoZxgGso8dXL1RZ3l52VfRJxIJlC1YofddqhEJ+VUTq9UqlUoF3Qwz2FIc/T1Xs9Hi6tWrVCpONEkymaTprcD7gOxs3qnl4TlBKhvTdk0mg0zIpQJVGN3/eg1fHSSC9x2MMNkJdiuF/a6hY8wNSvPimUvC4TCqj8J/26Emeifjg0J2Jr81jHUXItb6xzCtQZvoHFEioNGQSqEley+QOCgMDcmQUqKlUpQ0nayCppfoyn+ng7JdbybQ2mkzm4outeOxF0w45ZlEstks+XyeVCqFYRi+U2fQybNXlCplNs4s+86gqVSKvMz2STIMXyDU63VqtRpx0xwoyej3lTdrLZaXnb13IpGg3Kigif1bbOPZBBb4vgkN2/YjTAaTwMwbg8pNFuTMl2Yy3HP46iDh3bcXvmsqm4jYv5j9/UC9tLloe3NWCDEQmeFpdZpyMIufLe2+fDIGhXh2s4ZJsBrr0GDAct9vZ4cN1UvtUTzKTcYlFAhl00iMNBmOQIxHaRoaQgqErcDm5WwhaUMwIsSLiIhEIkSjUWKx2Daq4d4fTq1R49r5C34a4UgkQi6UQ1c6puhduFiW5RONZrOJMIH9C8kOQFGr1Xy7rhWy9qDW6faIZeJOzohAZdJgGOtuoTy5fzsY8AqluUm43ERcc5ncnlw7auyfHTqIWrHaVqV4kD4knsyQUg5kwbNC+1uMrLNQmm8uGVJeKvpwht4N1Iq1NjOpVytQSRtb2LQi+yeUh0rnaKQztMJhFI6ZRDJkHRw69C5VWq0m165dY2FhgdXVVer1uuPrIfqzb9u2jWVZvmAYloqJXi0FP6V4Zn/5tBExyD1waNN0EAjp3Kkj5M2wdmhvFvTtYEQMsg8cQgGWgKtHjiHe8tZ97dN+Ikg2BtWelNJxkh8AzOj+zl8jYpA5mmvTkDabQ6TJOAAQgX9pCFrYhFOpfevP0KzhQkqiySTKCIPterApAbZCeJ+BhfjsrK1W3WGvHnPcr3jpsVi+53Ms22ZlZYW1tTVKpRL1uqPK7VfweWGznkZDaw3HkGrVm201S4Sx//1Kz+QcTYbtmExalo1pK2x7IDrXzXGt8OdMs8fqq7uB9EyOJoKVwgSxe+7BiO6N2SpqDOY6OzXTVgOE1/sYfZL67dAakInD1vZ/k5CaTLdpMprN4dAWeHI/CFndqRuzg506gFdXHed20fZRCCWwFURGJAMQgmg6jYhEcR6PjRK7VeUAxPFC3+c2O1RkQTXVsEOpdjNCq9Vycw70314wN4e0BvMkdirYm7VmuyYjtv9DPX9sfNMvY4/MJaXs/tliPeSPjVNPJjDyeZrNJqVSaU+uuxNzSbTDWdaW/S++5cVSG+GVUhKSgyEZ3jxpDSiyxDL2n2Tkjox1aDIGRzLUDixoQbkvcLXtlf5JhjfGhJdrKdp/LovStRV/HZI4qcSd2kMSU4sQSg446q8HDI1PhhCCSCZDJRxGYiOE7Tj9CVejMWDIOyd2dL4I/BxE74IagaDTU61W83Nk+Nd2F890uPf8B93gaDEG84wNq/8hFU1FKS23C5RqD2mJgwgWbNI0DRHdf5IRioadTJ/gh3O2dtlcUsrsP8kIxcJIw6DZbFKr1dzkSv0b2W9nrgghiGr9azLCyQhFtX5b9XFuhuq6E93UmYm3n0RVHjrniVKKlhqQJiO0/9ETfkFB990OkmTY8Z3dn+j4GPO9O8x7iCSd8blTt09/jBEgGYDEqTlUjEfQEqPoEgBiiQRmMuZkyVDCSSe0C2NeAeJQ76YGD81qh7mEnamEHnjvgzzw3gf938u0+DpP8/XrT8P19mN/8K4PcCg1g67rxGTvtDwUDW0pLmYJi7LdW+Ew2BzcQWhKI0a0L3LQKdgd4dnfDq20WPRD/HRdh8T+k4zsTK5NkxHUZuwWlsb3T03qITuT89XfrVaLVquF7JNk3M5cef/h93B/4V4aVv+LUzjuzC1/HOr9LeKtequNXHifuNY/+UtNZiie35wnlrAGpsnYb58MgOxMti2MdVAkQymFFev//oJyH9zFvNjfJggCY8zrX76/tPitemuzP0IgvQ8KQwjWkzFio2RcDjRNQ2YyWNhuiFcwpdBgYEylmf7BtxF664m+22hWG87OBJdg7GHGwcXFG8gSTjrxRu/XTU9miI8l2naDq3K9r75U1rbu0qSUpGWKqt375Bs7Os7i0/O+6aVpt9hQ/anWzz9+zhfooUS4r5FeXa/wuz/5a054MU41UsP99+v/2Xu5J7DY3Q5i2UCEiafJgF0LGGymI32Frw78vjNxv4quRzLC7J7QW15d4VrjGrquM2bkWG71nuk1WUj51X5t26YZ6m8Rv/HSgj8vvMgxXddJhhJE6mHqdu/q9mQh1TZ/K6H+F7ogZFiitN5lbXW9wu/+1H9tK0boJUt77Q+9kRPvvren9rZEmDR37vegJwxybxnH3EGB3qDcl7h5SlarxG4UqU70TuaThZTvi2ErhT2V7atfN15aCBALMNjUZOhCYCdSxEaajE2IfJ6WZiCtllMSWsEgVPlHX3sH/+gv/u2O2wGnDLY30LzPXtGMy1euUJJF4vE4oT5teO/4x+/mG7/9VUfopyzmIr1XJwVYm1ttE6Ce1iCrpZm3e2/z7nfcS2lundYVJwnPcr73WiMAGwvrlG44aY4NwyCW72/X2Kq3HJUoYAhBTEqiUhARkkgfbzyWiROdyWK5tTxau2wuqU71p8Xoft+SiBB933ckF/UXDodk7B7mF65j2y2n2KARY5neScYdbzjOtW9eptVsod8RphXpz3Nm7oWrWxZfbwE+HJ/lbOl8z23e/Y57WT57A3PFhLRgPdd7GvBu0JL9LQeetsYLvfcKPMZiMWKR3ueeVygtqM3oF0dfewdH/587+j4/iE65LwApBemLS32RjDveeCdXvvIi5ovzJD70Rpjoz/w998JV30RiuBsCQwg0IQgJUJkckT1ytu6GoSMZqlCgZoRIWC3A3jWfjJ2guLjhaDEC7HGvwrjPXThHtBomm82Sy+VIjMcoy95KM0/dM8Pf+6UP+b9bfdY8WL68tCULqa7rjIWzvNA623N7RsTgLT/5Tv/3Rp97/KvPXfar2oZCISJ9Oj82a02kO2FjUpLWJGlNIyUlyT61V17BMM8no7mL5pLqWH+7l633rZGWkpSm9X3f2dkc5YtFn2TsJi5fu8LC0nWn7k8uBrO9txHLxHn///y9O+pHdb3CqYee85PpeXlvwuEwoVCIuyLH+yIZRsTgW//Fe/3f+52/ndCS/TmjNmtNv+SBV4wwn8+TyWTIpDJ9tZmeydC63giYS/Y/UYYn9yUC6RYPFALGHz3L/JuP99xeLBPnu37uh3bUp+p6hZOfeZqElOjC8cEIS0lICEIIpIBKNoMR2r8CaftvqO5AbGwcM5zC6dr+OyF1w9zzVxA4w17DUSfv1YO8Nn+Nc+fOcenSJebn54lV9s/W5pkkvMXcE56F2BiHotP70qdWvcXTn/omuq77Ze4jfdTu8ODtDhJSktM0ZnWDY0aIfJ+L7djxyTbnT0+bsRuo9GnjhS73bRgcM4z+7/tYoU0FvptYXFn058jSlUViG/uTiOj0F0+11RuKxWLE43GfaNxduItsKLMvfesGPdH/ntNLb55KpRgfH+fYsWPcc889TBb6c7DPHyu0OX8OAzy5LwV+dWIDQfzqOjPPXd2XPp3+4ilnrgJhIYhKSVQIIlISlgINEJkMRniUjMtHYWKcWiSC5lqVUMPVxeunr1G6toLmqqOCqqm9wOr6KnNzc8zPz7O6uoqxrKGpvWf5l566QPlG0VeReplJI5EI4XCY1xVes+d9AvjGX3yd0kLRT8+eTCYJpfqfYMLdHcSkJCM1xnWdWcMgX+9PyxKKh/2EXL4mo+/e3RyVHYSvtt+33Pl9x8KYpolpmk7V0F1EsVzk2rVr/hzRrwukvbdyZGNhna9+5Ms+wfBS+KdSqbZ58u7D79jTft0M2g5IhpSScDjsVz2enp7m2LFjTOUn+2ovFAttZhIeggJp28l9HWeBP/7wCxiNvU0atrGwzld//+8wXO1FTEriUpLwTJtSomyFkcnu+py7GYZrBQcShXGqYadbQmnYynWKcT/7jWc//ZSjSgZCOOwx7P6+F6jWqqysrLC+vk65XMastzhc3YE3Uz99WK/w6O//XVutlUQiQTKZJBaLEY1GOZY/whvG9pZonHzoWZ76qyd8u3A6nSaXy/W9Q6uslX1NVdjd1WfcnX3/ZoM8Jps1TOpuHZNBwYs4sJWi3Gf46nb3ndd1Ujswl9i27UT67DLqzQarq6usr69TqVSwyiaF1cyuX9dDq97i87/xOd9EkkwmyWQy5HI5MpkMyWTS12bcN3kPd6fu3LO+3QxavD8pVlkr+xrNaDRKOp2mUCgwOTlJts8U8l6htL3QfN0OtpX7LtFIXV3nvk89u2f98caYZyKJCUFSSlIuyYgKQQSohkNEspk9mXfbYehIRjKfo5FMuIRi/8Opgjj50LOcf+xFRzUVYI4x90XvBUzL8pNpNZtNbNsmr7LcrQbj3HQrtOotvvR7n6d8o+QXWUun02SzWdLpNIlEgmg0SjQa5Z1H3sY98bv2pF8nH3qWL/72I/5uKpvNUigUGB8fh/DO3o0GhFxVZEJK0lIS7/N9x7MJv5ZHa7fNJemd5chov2+NtDBIif5ERjybQEqJZVm7vmjYtkW1Wm0rrpU10xypz+zqdcGZH4/8+mdYOH2dSCRCKpUil8tRKBQoFAqOH1UiQSQS8Z1Av/vO9zFmDC79+7Vnr/R1nh7vfyEKFniMx+N+gcdkor8kUF6hNK8m0n7iZnI/LAQh1wfi8KPnufPTu080vDE2/9xlQu4mIK1pZDWNnOs/FZOSmKazkUoQcqt47xeGjmQkkklkOoOtLDcN2nAQjZMPPcsXf+NzhF0VctJ1hEtJSULTiOxhGGsQns33jthRXq8/uKumk42FdT75H/+SS09c8CvF5vN5xsfHGR8f9wWot0uLRqN866G3c5dxbNf6BPDVP36UL/3O5/1dYz6fZ2Jigunpaaanp2lq/asxnax5wvdR8Byr0sv9efSnJzNYgOWaShq7SDLK6f59UTrvOyIEKSno0cfYR3oy45tL9rKqprf4RaNRjkUOc6+1e1qD6nqFR379M1x64oK/ox8bG2NycpKpqSkmJibI5XIkk0nC4bDvLB2NRvnA0e9hTO6caGzc2ODhX/o08y/O9Xyu3qcmA/CjzDxfKM/JtZAc66u9/RovnbiV3I+6MiHkEo5jnzvNkT9/Ytf6442xi4+9SNjVXmQ1jTFNo6Dr5DWNtKYR0wRxBHooiQyF91WTMXTRJeFwGDOVQiEReCRj/6JLrp++xrOfforLj71IzB1oWU0j777YvK6TdG1g+wEviiIWi3FMJoivRnl643mKid6Ta90MJx9+li//zhcwDINkMkkqlWpbyCcnJ8nn8ySTSX+XJqUkkUjw5sk3EL0c4XTzHK3I4DwQli4u8ujv/x2LZxZ8k00ul2NqaorDhw9z+PBhMrNZYL6v9oPJd6SXF2UATr65Bw5hnr5OC3y/jEGjcl9/tnDYet86oOmSueMzaO9+L7DQV7vZozlalZab8XPvEIx8SGtpkmsJnm+ephYZnEPh2cde5OH/8mkMw/B38oVCgZmZGQ4dOsTs7KxjPshm/arKmqYhhCAcDpNMJnnvxLfyxJVvct7oTxOxsbDOp/7jX2FZFovnFpi65/Y1N+lj/eVogMB4CeTK8f5t76Cw036NF7h9uR+VkrBw8lFoQhADDn/5PLy0yNUffjP2kf5IVjecfexFHvrFT6ADUVebknP9pSZ1nQm3bykpiSpJVEArYRAJOxqz/cLQkQyAVjKBKSW62nuCce6xF2nWHOHztd94yHH0ATKaRsS1eeU1zXmpHnPcQ5Kha5oTMeE6jnmRHbFYzN8dqZbi3JWXmDPnkXf0P7iq6xUufOMlHvu9L/rkIhqN+hqM6elpjhw5wuHDh5mamiKXyzn5O0IhdF1HSkk8HiebzXJ3/S4il0O8ePUsq6EN9In++3X99DXOPHqac198EcMwfDNNNptlYmKC2dlZjh49ypEjR5BJSZ+ZyWnVGn5+iLC7W9HZedzT5Ilplk5fH4gaUQARt2++EzLQLPSffCd431EpCOs6N+67j9f8o58iOp3kxfX+SMbk3dNsPL860Cqk3SCl1uZcGXS+9P923eDFubMsimXkdP87+JMPP8u5r5xh+dyir8Xz5odHdo8ePcrs7Czj4+NOWG00iq7r/kIcCoWIx+PkcjleXX+A8OUQ5+sXaYzffm6Os4+9yCMf/oy/4Zh/co77v+NB9FskY/OeVaqPPA8ezHrLz4sRlEuapmHtgGTs1XgBOPeVMzRrDQSCr//Bo4hqDQ11S7kfDfhlKJy5mJSSybki6j/+LdfeeJTKa44Qfc3Rvvt28qFnOfeVF1l8/gox16Ez7vpJFVyCMW0YTOg6OV13HD+lI7PqySzpaAwp989oMZQkQ8/naBk6Wt0iSDIScxv8w5/5OGu2zQ3T5EqzyUvNJldbLZYsiw3LGozq2d2t5jXNz3boOcClXbY4oetM6TpjmkZS00jOFfnQz3ycsmVxw7K40mxy0TS5NjvNejTO2traQEKxQqGQH4MedLSMRqPE43Eikc2Q1vCVENceu8ZSc5lW2qbRbDD12tltBc/GwjrXTl4BBKc+8yy11SqappHL5Xzzh2drnZiYYGZmhsOHD3Po0CEmJyfJZDK+APV2M94uzTRNLMtCKcWVK1e49tw1ytkadbuBrdnMvO7wtvc8f3qO1bkVQPD4Hz7mO5nlcjl/8chkMhQKBaanpzl06BCHDh1iPBpF+8qTvP8PPoqpbNYsmzmzxYVmkyutFjdMkw3bpq7srqxBCkHG9cHwnKlC7iKeffIq3//kRynaNgumyaWW0+5cq8WyZVGy7W3zX0TdNiOuI5kAvuVf/BEhIajbNouWxblGgyumyZJpsmFZNLZpy+uj5/AVccdr4UvnOPboeWxg1bK42mr5c8W/720cTqUQZDSNnK4RicUo3n0P93//DxCNx4k0QnxP4r2USiUWFha4fPkyFy9e5Pr16ywvLzuOlpa1RT0rhCAWi5HP5wm1DOSXW0xMTBAOh6lUKly/ft2PmtrpXPHGRjabJZVK+WGjnrYrFoshhEApRfRKmKtfvMpapEhTbyKikvH7t3ekPvXIc4Bg5dIS5x8965s88vm8H82Uy+UYHx9vmx/T09M+CTcMw58fALqu+/4bnonAuKJz+ZnL3IitUKqUGHvN1lDQ+RfnWL22ysnPPENtpUo6nSYSiZBOp8kkMlz4g9PE43E0TaNer1MqlSiVSpTLZer1OpZl+c+qdanBxt8sMjExQTKZpNVqsbS0xMLCAgsLC6yurlKr1brWcdE0jXw+Ty6X8wl/JBJB13WmslP8L2/7t6yvr3PlyhXOnDnD6dOnuXjxIgsLC2xsbHT1udjL8RK8ppSSQiaNqDfQheODEXfna8HVFnhy35tvUilwNZ0GEBMCSwgsTUM9fpH5r5xjwTRpvv9V2GNJtHyCqddu70N36mHHr2P54iLnPvsMuktiMu714q7JJq9pjGsaUy7BKGgaKZeEaEIghETl80Si0RHJ6ETy0CzNkEGkVkMFpL/3L10pQuAnCqraNgLn5e5U9eypwwW0EYyo52UfUJmN6TpZ1wHIwLGza+D3LSGcKIT4zAzJZHJgJCOZTDI+Pk4+n28LiYvFHEc/TdPadkip+RQ3btxgZaXJS8+eolwuU6vV/Cqspmn6KZS9ZxDToqQmkn7+C0+AZjKZNjvz9PS035dEIkE4HPavr5Tyd1Ze216oWzweZ35+nsXFRdbW1jj5tSepVCp+dVhP2NqBhVBKycTEhO9g5hGMdDrt+2F4fUqXq1h//hH0J75OTEhqONnvIoH3aAERy6KJ1lV46u57T7sOVanAzsXzWfD8FZJumw2l0IQgYdtdCa8IENasphF31a1edk3leoqnNY0x9/yoENuSDC3Qx6xrM/aIhoYzZ3Tw7bdt9611v29DCJJSR0snCb/2tZx4y1swQiGfPHjZHYORE/V6HV3XfSLZ7b69BXBsbIx0Ot3mm9BoNLAsCynljueKYRgkEgnGxsb8awWJhpdAzpsjsViM69evs7i4yMrKCqe++BSVSqVtjliW5Y9FbzGamJjw50c0Gt0yP6anp31fDM+U6N2vJ/S9tjwzixcd5PUtei3K9fXr3PjoFVZXVykWi1Sr1bYKyjERJTme8OdDNptlbGyMbDZLPB5H13VqtRrr6+tsbGxQKpWo1WoOmXGfVT6f98+JxWK+XLAsy1/wG41G93mi68Tjcd8/y7tuUBZ4z9l7Pt54SafT25KMvRovwfcghcBaXESsrBJ2F+zt5H5cSj99gafd1IUgDChNQxGQEVKy9NApVi2LVdvmimVRU4p6wPnbBj+C0luH8pq26QvmEoxOP4yCK58yrjxxjpXUNUVrapxoPL4n2qDtMJQkY2xmllokiiyu45Ri3WRh3uIfdXdvBU1DAAkpqQ4ge6JXj0Rz7WxeqFIswCBTLrlJugMt4jJN3AHl9S2rQ80WpI8doxrwdN8JvAntaRO8xd1ztgwWYfIWc29ir6yssLa2RrFY3EI0gouDJ4SDBCMYRTI2NkahUGBsbIx8Pu8LcU8AeAPaK1Dm2QO934PCsFAosLq6ysbGhi9Ag0QjSH48xzKPYHjOp16GwfHxcQqFAtH5OZq/9jtElq4T0iW2UlgKImozuZTCIaU1TdtW+yXBFzI5dyInAjZYwWba7bSUjGsaOpCUktpN0oUbAWKQ99p0hYMmJUmlGNMcB96YEFRv0cew10dX+CTctnT3PYQD93Cr+zYQ6FJQz+VIvPVtHL/nbufduiaHIMnwnv/Y2BjRaNQpJOYuTt3gkd5sNut/QqEQjUbDd85Mp9OUy+UdzRWvrUwmw/j4OGNjY76pz7sHXdf9sRmLxfywy+Ac6SS9luWYL4KF9zwHx2DIdHAselpHT8sYnB8egvMkHncSqHX2bWxsjOXlZdbX132S0Gq1sG3bfx9BU02hUPAXe03TfJJRLBYplUrU63VM02x77ul0mkwm4/jFuflMvOtXKpVtozw014TrPcNCoUAqlSISifgVcT2fmHw+T7VaRdd1MpkMlUpl38eL9w6klAhbUSuW0MIhIkBUOgu3J++7yf1gRW5wFtWIEAhN832aIh45cDXuRU2jbNvUAs7fpks0vLa8gmf+GiQ3o9u8jY8XTZJy/UOcee/M47rQ0PJjRCKRHVUS3imGkmSMT01zMpEiv3iDzugSHSesLiEEtuak7EpJSVXXaQ6oDoRHMtoiClzy4DkBeZnVPHucEMLRsAT6JpDEgFQ+hzExue1k6gVB8uAJH29Ch0KhNpLhZbz0BI8nZDyS4S3ozWazTWvg7ayCSba8ZEJBYRRURXt+GF6xNP9ZBgRo8N9B4VksFv0dVrVa9XdqzWazjfx4ZpKgYPf6lMlkiCuF9Y0nsP7yz4gvlxGajo6JJSCEICFMLE1D4hIBXadh29vGL3kkwtN+eOm1PU2BwNm9JKXE1nV018xQs22aAYGx5R2Kzdh2T6MRkxIdsAJakoTXx5uMa590uyTY62NwbOKqbre7b1sohO1o32qRKIt3HefYm97M2OSkU/AulSKVTvsk1ls0otGoryY3DINGo+HvMLshGOLofXRdxzRNstkslUqFSqXiL+r9Iqgt88Z+JpPx/Za8+RHcYXtjcWNjwye8lUqlLRTWG4texEowk6c3PzKZjD8ePdNBLBbbdn4E+wyO43swSsPbJATJeKlU8hd9j2R49+GFzHqmi6irKm82m5TLZf+evLnlPatYLOZ/DMPANE3Gxsb8+eiZV7Z73p3Xz2azbb4nnlbC04blcjn/Xe/3eIHNIo9Wvc71r32dlBHyk1xFpROyHkMS0drnlucQ7rdDuzZccxf+mG37GoiSZVG2bZ9k1AMVmW3l6O69NcjTpEcCJCMZ+JmUTriqr70UTuVyQ8JGMkU0m/Nl735hKElGKp+jNDGBunAeifLjS5yPIAQgJEI6jC0lFQ1lY6nBBLz6AyRANIJkw3Oyc9KKu4NMbe1bCEW2XiXUaJC5++421X/ffQvs5r2FNqia9IRn8DiPkOTzeV/IeLk2OtXBwbLUniYjuKj7hY8CTl6e+nc7u59HLoKq4aAvRbVabRPo3mIV9OPwzg0K90gk4vdJq1Yo/eEfoT/0CGGh0IREKBBCQ9jKnfwaQtiENEFK2i4p3X7MOGPAGWNtKXuFQFeb4xEhkQIimqB+G+1Kt92QcIqtxYRTZ0ATAhswpJOuOCGkX9vkpn1kkwx7RDiMQENgoxBCIgRd79tR6SoSmk4pmWXuW9/Bu97xdn9h8siERySD7zGdTnPo0CEikcgW7dOtxq43tqSU2Lbtk8pbtXE7COZs8MZZJ8Hw5knnHPEWruAc8Uh4N01Gt/nhzRFPuxj0wbiZsA8e4/UtOE/Gx8fbNgdBTYZ3bDwe96O/vHsWQtBqtajX61s2FUHC5H28d+IlwurchGz3vL156T2DTs2NN48jkQi5XG5oxkvwesUrVzDLZcYNw5mHQhJWAcfvDrnfza/V23yEEG5SOwhLjRiCpJDUhHTJhU3ddgkGClOB7ToFeDVSdIFPNMKurAimDvecvh2NikAop9RXSArWxseIuRuAkU9GByKRCGJqGqREWhbKE+gKLIG/O5NIQtJJy2wricXOq52I4Ec4g0nC5uByf/ohjYFzPfu317ewtGmZJqHlNWYOHRqIyioogLyJ7RUmCwqp4IIe3GV4i4EnPLbzyQgWPfMmuPfxhJEnrG8lPIP99rzOPaIRj8d9YRHslyfUO80lXr+8PoTDYZpnzrH2x39E9NGvEFY6SIXwllAlEEIh1SaBMpQigcRyF+/t3opfcdGdxIYrZLwdDABKEXbbDfXYbpDEeuYXcLV1mobpajBsbk6Egn30/Ig0PALiFEnqdt82EJKSpNK4dPQI/L3v4q2ve50/LlKpVFt5cm/BME2TZDLJ9PQ00WjUf1+d42i7MeCZKryxo5TCsqy2z07mSpAoG4bhj2Hvut54DY7FoGmgc450+gd1tt85PzxiEQxTvd2dpDf3gu17RMMjCR5RCG4MgjVSgoXYvOtbluXnnPDO69xU3Oyd3M57Dc5N7/6DDq6RSMQ3lw7TePGuhVKsnDmDLRRxw3BThgs05VRb3U7ud23P+wiH7OsoQpokoqAlpR+67ldiBn9eeud3yp5Qxycoi9r6pBS6kFTz4xRicX887ReGkmQAhI4fpR4LES6ZCJT/ACWghOMo55lOvBczSKtT0M7maTZEYCHoNsi8Y7y+2cJxrNPn5sglk8gB5Y/vJBLddkmdwiq4y/AmuLeQe5M0SDI6y7cH7djbXbOXfgd3gp4A7OyTJ4CCJCN4rq7rrP/FX1L6yB8Tu3QJIQVoNkLYgAXuUiuxUQIMFBrOBPXGyq3GTOc48MiFPx7FZhXeQbQL7tgBbG8HuIO2cP/W2T8bx49JSHj+LW8j94M/yL3HjlIqldA0jXjAWcx73977mJycpFAoEIvFsG2bcDjsv6dbCfvgGAiOneD5gyLj282RTp+h4ByxLItEIuEvxkGyGyQZnQtgtznS6/zo1vdOQu5tCrbzV/I2Bp0ERynl30PnM+72XKC3d3IrmdSpORq28QKgmk2stQ0nEaRrvpQ2CCn8Elqyx0t1kg0DhSUEMVdzYXkEA4UKaD+DG109uNF1N74am+tR5wiTQlBH0pycJJqI++NgvzCUJMO2bVKHD1EzDCLuIBLKEbaeOsh7uHvKzwKD+WZjra1vQmBUyxj1OuFsdnf7160vgYmvudEEfn0LV+B4wjNIMjqFhvfT+36nfYJ2wRMKhbb06WbCUFWqLP2X36Ty8T/FqJVQmnJVl8I9x1Na+mf7poUd5UTtEGjekxhkuzse190iAAKERQI5TbIRinD6gx/g7g98gFgiSbFY9BfMTpimiRCCmZkZJicn0VzHVO/nQYa3mHv3rpQiEonc1lgMzo1ui+FO++V9PPLQ2a/gsbvZl50iOOeHFbbU0OYWiKiA5HBt4cJ91DuhM578cSJRnGg3T+Np4yxqquN42NzkBjc4TrfUlj4pICR0ykqiTU4QjkT2VYsBQ0oyAPJTUyyFk8A6CLDdJ6mGY87cNhQ2ql6neW2e8NTeFjLrxDAJHQ+du6fbEUKVc+dY+i//hdKnPw+ARKKE5ZBPfyEITqzhuuf9gkJioMhIydz4DNc+9AHu/873YVoWa2tr244P0zSJRqNMTExQKBT2XWjtJjp338OAYevPyxVWvUb9qW84lb93UWQEtQ+dWscdQ0FEKVZSKUKFQlu49H5hKEmGlJJsocBLU7NMLcyD3fLVVTtIILcvEEjsG+uUnjtF8g2v3e/uHHgsfepvWfnVD9N66SyWNBC2QscCJV0C6u3VvcmrGKwh7WBCoQgLRUiEePqNryX+Yz/KPXcco+76wnQjF15IaiaT4ciRI34elhFGeDmifvES9WuXkQd5TyIUIQkbE2PEx8bafGL2C0NJMrwwTev4Eeznn4JGc7OCyQFbL5SQqFaD0gvP7XdXDjwW/uAjrP7qr2GuLmEjUCjH30IJbIGjPlSwVYF4wAbNIBDwH7IFJDRBWMLT73grh376p4mnUpTLZd8M0gnPfj8+Ps7s7CzhAfkTjTDCMEJZNsWnn0OzbVdrflCZhkIJKE+MMZVK+RE9+4mhJBmejTR84jgNQ8doCoS91f50IKBsBC3sC+epz88T2WeTyUFE49oc13/hFyj93RcRxRJKCgQKaTsmEitgJtn8/4EbKQNFwAWYBBZL+QkWf+gfMPuOtxOKRikWi350QSe8/AlHjx4ln8/vawXHEUbYE9gW5S9+BWUb2MJCHFD5IYXAtG1ak5OEY9E2X7r9wlBLj+zRo9TCBqHyppnkoPlkgLMAti5fonL27Ihk9Ijyi2e5/C/+W+xL5xC27UZcCIQ7EILmM88peAQHAojqkmuTs9j/5l9xz333+8mVoLsDr+d/ceTIEVKp1L7vgkYYYS/QWlyk9eRj2PLgmeQ9OCkUBDUlsXM536F+vzcJQ+vBpWka8VSK5WSElubEBmjqYA4AWynEShHzwsX97sqBwsrffZHz//j/h3nhLLYtUUg3X4oT0ixVh1JTHMzxMUjY7gPRbScl8pkHX4382f+ViRN3UyqVts2MqJSi2WySSqW48847SafTI4IxwisGpedPYleKoJocZC2ooRS1aBgtkfDzEu33PB5aTUYoFCKeSlHPZJFXryJVN8/qTbW4COxuFWBLN6ZdbY1ttoXs2kYnBtWmEoCwKX7hEfLf/wH0dHq72x4BUGaLK//3z7P6sT/FqJRBeYWU3XchNmO9hEs0bMmm3w7t33W9BttrPW56ntheBAm2Jzk3PW8A/fSOEwpCQkPqBs996AMc/YEfQAsZ21bQhE3zyOzsLJOTkxhG/6XPRxjhwEEplh95GIlToFCpbTKbBtcDobDdyS7srbv1vtaYQJvS3ioTbt6mo+HVNI1ioUByYsJPGb/fGFpNhq7rJFMpxMwxou4W1RYEIgg2cyrKDg2HErYjgG+p+ej+8v1dcts73GGbyqZx+RLFk6dudvIrHvUbN7j8L/4xxT/8Paha2BhsPlOHaDiajGDaq/Z/C5Sr5eh8F1uPDf7e/bytx3Vr0xkv3Y7d/tp991Ox5TwncY9JXFiUcuNc/Lf/irt/7EdBkzSbrZsSDF3XOXbsGLOzsyOCMcIrDma1SvNrX9nWD6PreuDOw5ttELzjbnuNCbR5c3Rr0yk5L4D1mWlShYKfX2W/MbQkw0uME737HpqhOLYEy+utsEFYzsLvZgPdvBEbie0ec7MBoPw2trxV1R7LDBZih20qAdbcMrXnX7zdR/CKgjJNNp58ihd/6icoPfR5hCnQlIlPKL3nuuXfzu9CWQ75EM678j4ICwieZ7Wd57wn99+q8zxr6/W6tCnd45zcsx1tYnU9b0f9xEL65znnSGESM0JcO3qc6r//dxx51zupVqtuvY2tUstL5hSLxbjjjjvI5/P7rlYdYYT9QHPuOvLG9e39udz1wF9j/LmpumbcbD/x5mvMoNpUgI5CKIE1OUMsmRwakrH/PbgF8vfdw0YkRrLWQCgb6VapU8JRM3kZERTKF6UOS9xeP+0zVuUd3fm9BzfTqJ9Stv82AWhVYGFumzt95cKu1pj/2EdZ+cVfQCuXUW75IalcbZXYNJU46Hy+rsrR9fxsPzK4P2k/z1FTum0rgWhLEupN9c5re+e58BxQ266nAueLbc/ru5++DsOh2JqUGJEYZ77t2xj/0R8lmkxSqVTYDl7GyEKhwOHDh0dJnkZ4RaP23AsoS8fC9KPUukkbXxvobjTF9laQ215jnCbVQNoMCUklFEK/4yiGW9tpRDJuAV3XSU1McG1igtzqOi1PTe4Ka6k2y207j99ZkIQKVjvZis2skNtzxs21RgVe6s7aFLZF48Z1zFIJPZnctq1XElrFItf/4y9T/NPfRwi19b0J19Oiy8Le+e+g/4z7lw4rxDYLvnesf65oJxuBNm55XrAD215v5/20hTP+wwIaqQynf+D7uecHP0jTrba5HWzbRkrJxMQE09PT+x7eNsII+43iFz+PtC1Mzx9CiG6KB3fqeevBzefNba8xIkgcdtKmQtc0FsYLJGZnAPyqtfuNoSYZ4XCYRDJJ6a67EOdeQpnmpg1MCd8zoz2/o8Df+N0St6se7kWNfBMiooVoPPEUG1//Gvn3fkcPbb78oCyL0vMnmf+d36XxmU8i2Sx057w8i96e+ysHCggBhq5x6dhxaj/wAU5867dSc8t/b5e90zOPTE9Pk8vl9rzfI4wwbGitrNB4/MuY2H7l4luvHYNZDwbZpnT1HGtTBaYmJrAsi0gk0kObu4f9pzk3geeEFnvNgzR0DcFmiJ4Hb6PnczvlVbDcu37eNgSwco3qs0/ud0/2FcqyWPmbv+bST/0kzb/9NNJ2C695B/geVbZjq1JDPUz3DDZOGG9MQlyPcPaNbyL5c/8br/7u76ZWq900e6dSinQ6zfHjx0cEY4QRXKx+4e9orRadyCy7e1THQYCGs+41jhwlkckADI0T91BrMjRNwzRNsnfcwXo0ylitjLLZjDJxtRq2CmoyHAjoVoxyX6GEQrWaWGcu0lpZxci/8oR9a2WF9U/9LSv/8T+gVatY0nBflOdlwKZGEgLfvXIh3BLtAkFMapjpKKff/FZmfuqnSGQy3LhxY1unTc88ks/nmZ2dHQob7QgjDAuKD38GYdl+Ij/bFT4HSeIoQJeKWjxO+I7jfmXrEcnoAelCgbN33cXEEyu0sJw/BiP8aCcXwzpCFIAWo/HN59h4/AnGvuv9+92lPUX1xTPM/+qvUv7bT6A5DwOhAtEPXd/bkL7MPYNwzSOCjFRcmSxw4x/+CPd+5/totVpsbGxsSzA8lenU1BSFQmFvuz3CCEOO2tw8tW8+C8rCc8NWgf8fFAggJATzE5Nk7rgD0zRJp9NDs6EYej10LBZD13Xse+4BTds0j6gD0PlOKEAKzJU5yt98AmVZ+92jPcP6449z/d/9O6xPfAKhwBRi079ReKGYI3RCoYgIiEuNc3ffS+lf/0se/PvfS61Wp1QqbWseMU2TeDzOsWPHRgRjhBG6YOnhz0OpeCDNI0FIHJ+M9VSa5NgYpmkSCoX2u1s+hoPq3ATxeJyNjQ1CR2Zp2LYbRrq58w0kfxx6uHEBCDRaT52ifvkq0TuO7m+ndhl2rc7SR/+cld/6VeTVq1hSA+Ek1FIEPaZx3+tBn/KDgmM6SmgaphHmyW99F5M/8qMcOnqEpaUlLNvq6jnuZe/M5/NMT0+PyrOPMMI2aDz6KLJedzIIH2AIAaZS1PM5QpEIzWZzaLQYcEBIBkBsepq6ESbVqPqhRkG+cVDg+DOa2GdfoDV/7WVNMpoLN1j48H9l/Y/+AI0aLak778tWbuRWt7jvA/ZCdwEK0JRDMNayY1z+/g9w4gPfRygSYWlpyamsKLYSjFarha7rHDp0iLGxsVH+ixFG2AZWtYp95jRC2TdNdzDscJJwCaqGjnbnHYRCIVqt1tD4Y8ABsDh4TizZiUmKhw4jNXXTOhDDDwGaRFXXqb54br87s2soPX+KS//6Zyh95LcQqoWFge0ON6k6i5sd3Ek+aHjhqTndYH56loX//r/jwR/9EZCS1dXVbf0vWq0WkUiEo0ePMjExMSIYI4xwE6w/9hjmygq2FNtn+jwgMJBs5MfInDiBaZqEw+GhqFniYehJBjjajHAkQuV1r8HUBLY4uAJUobCExLY01j/7EM2FG/vdpYFj4aMf59q//dc0vvIlbKG5cSMSRXtWzVd6xdQgPLezsLDRtRDPvfNdlP/dv+Hud72TYrFIuVzuah6xbZtWq8XY2BjHjx9nbGxsr7s+wggHDsWP/gV2rRLIE30woQGaEKzNHqJw7BjVatX3YxwWDE9PboJkMsny8jKR176G5sc/hmFb25SdORhQysayFebTX6dx6SKhyYn97tJAoCyLKz//8xT/7KPI9WVAQ/lpYgLhqQd857AbUCiSQqOlhTj5Qx/krh/+YUKxKOvr676fRRBCCEzTRCnF5OQk09PTQ+XsNcIIw4r6xcu0nn0BaZko4ZlLDsR+ews0IagbkvqJO0llMiwtLQ2dH9aBIBmRSAQhBLkjR6lMzjB59TLVg+wkqJxkU7LSoPSFh4k/+CAyGt3vXu0ItfnrzH3416n/8R8gFG7+vM18+yOtxfZQChK6YDWb4+oP/BAnvu97EJrGRnEDgehqImk0GoTDYWZmZkb+FyOM0ANWPvHXmCtzwGZ+jIMKHcFKJkXigfsxTRPDMIYm06eHA0EyNE0jGo2CUsy/7vVMXb2KOuhpp4XAlorVv/gr0h/4QeL33r3fPeoba3/7Ga79yn9Gnb2E8HcEI1ZxKyhAwyYd0jh9/6sJ/ZN/yp3HjmHaNs16AyEknc9RKYVpmqRSKQ4dOkTSrbY4wggj3BqtjQ1KDz0CrRpKSlCaO8UOprwSQrA8NcvM3XdTLpeJxWIjktEvkskk5XIZ47UPYn7ir8EWmxXsDiiElNhLy6x+9WvE7r4LMQTFbHrF0l/+Ncs/938gVhZRbmbKgzxpdwtBIuBlo40IiGo6z735LWT/8T8lNz3dkR68/Rnato1lWYyNjTE7Ozt0wmSEEYYda098A+vqHCiB7Sd0PJjGd11KbGHQuvNuMrk8i4uLFAqFodNqHhiSEYlEsCyL3Im7WJqepnD9OjW7dRDHhg+lBNK2qD/6RZrf+92Exw9O0qT64hJL//U3Wfvzv4T1VZxKop52abSzvhlsbDLSoJhIcuoHf4Aj3/VdGNEolUoFpVRXzYTnlzE7O8vk5ORQOXaNMMJBQfXxJ1BrKyC8St32gZVWOlDKJAg9cD+2bW9q/IcMB0ZShcNhhBDEUiku3nMPM/OL1Gjtd7d2DAXUv/oV6nPXDgzJqF+5wvl//W9QTz4Bpucbc3Cdp/YSQkFK07g+Mcn6P/tnnHjj67Fsm2azue05pmkSjUY5cuQIyWRyKMo3jzDCQUNjeYXa178GqoUl5IH3E9MEFGNh4odmaDQaRCKRoXP6hAO0KoRCISKRCEopKodmsZQ68OFHAEoIZKVG9bHHsFvDT5o2vvRlzv3Ej2E+8RjKbCGVjVReenTJSIvRHUIpdCwyRojzr34dzZ/73zn85jfSbLUwW62uT00pRavV8tODp9PpEcEYYYQ+sf7Q57DPn8UOaArb8/UcHAgEtoDyWJ6x6Wnq9TqRSGQoI8wOjCYDIJ1Os7CwQP7Vr6KoKTRbYB+0lJ9doJRi7WMfI/Vd3038+B373Z2uUJbFjd/5HVZ+7dcwN9YQtsSpNuJl7Xzl1GG5XWyaPRQhIbGMGM+859so/PiPEc9mqdfrqG3Gr/f3qakppqamhlJ4jDDCQYHdbFL69GexN0pY0tW72pvuYwcrwkQgpaAcjVN/3etJplLMz88TCoWG0gn8QJGMVCrlkIyjR1i+5y7ueO4UxQOdFNaBJQTN06eoPPU0sWNHh84BtHF9gav/x/9O7VN/g4WjZjzoz3yvIFAkhGQtX2D+v/lvuPt976XeaNBoNLY9x7OvTk9PMzk5uYe9HWGElydWPvc5Gs89hS0AhE8wBAeNYAAoQspiPZVm/E1voVqtomvaUJpK4ACZS8Bx/gyHwxihMNXXvxVdGIiXhdEEpGZQ/tM/wlxZ2e+u+FCmSfXsOS789H9L+ZN/hYlAqZFJ5HahIYgbEa5P38HCf/ffcvd3vo9KtUqj2Z1gKKWwbZtIJMIdd9wxIhgjjDAAKMti4+FHaK6uoJBI282vK8CSBy92QCiJEmHWp6aYOHqUcrlMJBIZSqdPOGAkQ0pJMpnEtCxSb3o9S/EEIXEQmehWCF2n8uUvUz75/FBUfbNrdRb/5E85/8HvxfzmowihIbzsnW0FzUbohrAQGOEop97zHsK//B+5653vYGNjwwlP7fLslFIopUin05w4cYJMJrP3nR5hhJchyqdPUz/5HFJJd7EQXefgQYEmFK1oCPsd7yAai2FZFtFYbKTJGASEEKTTaZRSZCYnWPiW+0nIENr+r8kDgRaKsPzZz2LV6vvaD3N9g7lf+RUW/uf/BZaLoATCdpwXBQohNtOEH9ypujvQsIkLiZ3I8eLf//vc9TP/isz4OBsbGwBdbaa2bSOEYHJykhMnTgxVcaMRRjjoKH/161gXriCUcDZJQqHEwc1ELIRgPZNk8k1vpFQqIaX0q5UPIw4UyQCIxWKEQiGkrqPe+CZsIdGlgC6lrw8aWkLR/NvP0rhydd+0GZVTp7j0sz/Lxq/9MgKLltCxhIbqsjgKxcg/w4cirGxiKOYOH+LCP/+n3P8v/hk2sLa21pVceOaRcDjMoUOHOHTo0N53e4QRXsZoLS1R/uxnEc2Gsy0SFmAfOBOJB4lAorN81wkKMzNUq1XC4QjJZHK/u7YtDtzKrGka8XgcWylSD9zLQjqMLqWzyw4kbgsufsItLU7gd4/Bits4dhBtis42g99t/oa1vsjqH/4/KGuPozWUYu0zn+PiT/wk9Y9/HGFrCFuhY+KUTm1fJH2C0XFvr0QonOydYV3w0uvegvjZn+V1H/p+yuUylUpl2+qptm2TSqU4fvw44+Pje9/xEUZ4mWP5rz5B9ZtPo5ABQXUwM3wCaMpGGhJ13wNomoZt28Ri0aE1lcABiy7xEIlEWFtbI5JMsXzkTqafeY6GUr6nsMJlTwpsAmmiAiQA3O+8Rd491lsOvL/fqk0VOO5WbdrBNgm0aYPtOSBZGut/9jcU/uW/IjyxN45/VqnEyl/9NTf+0y8h1xaxhV/WDKeIkO3YMIPkiHZfmFcq0ZBAWINGKMqL73k3R//JPyM9Nsbc3JzzfReCYbkEMpfLcfTo0VH2zhFG2AU0llcofe5zGK0GphSudtjf+u1n1/qC404iKMcMUvffQ7VaRUpJIpHY767dFAdOugkhSCaTLCwsEI0naLzhTdhPn0JiOc5z4GS47ijSqjpjoQPfdzvWX1MDfo43Pe422/T70qVNj5yoWpGFP/5jDv/3/xKxywtQ7eJlbvzaf6XyJ38MQrm5LzbzO9BBorbDwZuyO4dUioSuMzc9zY3v/R7u++AH0XSdxcUlNK27krDVahGJRJiYmGB8fHzo6gyMMMLLBUsPPUzjm0+hK4FQtvvXgy2phDS4due93Hv/A2xsbGAYxlD7Y8ABNJcARKNRpziUgNzrX8tCJk1USky5qXVQYvNje6FKgd89zYH3b9v9zvK+844NtBk8rt827W3aRLn5MgVIDdb+8y/RWl/f1edYfOyrXPm3/xPlP/1jLGlh+7ahzY844J7YuwEbMIQgrRmcv+ceyv/mZ3j9T/wELdNkbW29K8HwsndGo1GOHj3K1NTUiGCMMMIuwao3qH7lqxiNGtYQJqjqBxJAF4j3vBvDMDBNE13XhzZ01cOBJBlSStLpNK1Wi/TEBEtv+Baioj1jhucXAXSum8P5wdVwKLCUQBeClUc+P7Bn1gbbZv1vPsXcv/kfsR/7MhamU5jLdmiFch2jFNL9qfb98ezjawnAYY0xJGgaz73pLej/5me49y1vYWlpiUa9jpTCO9KHZVnYtk02m+X48eOj8NQRRthlbHzjG5hf/SoWCnDMk/stS3b0UWAoydLYGNNveiPFYhHNTcA17ObW4e7dNhBCkMlknFSq0SjirW+l/IVHCVsmrg8xynb9FbeuFEOJtkXNViglWfrlDzP2vu/AyGYHdp3m0hKLH/5FNv78k1gbGygJjh1JOU/O74Si3XvklQsvMsQWNimhU0ykuPgjP8ih972PZDbLjRs3nGMCOybviZmmiZSSqakpJiYmMAxjH+5ghBFeOWhtFFn76Mfg+pyjxVAHP2GjUAqhC5bf/nYOT0yyuLiIpmkHYsNyIDUZ4FRljcfjtFotxu69l2v3HCcupT+YPDPGQR1dtrARl89z42MfH1ibzYUFXvoffobV//f3sIqreB4YzjDw3VPZfGjBf79yYQOaUozLEOtjeS7/83/MA//wHxJPpykWi9vWC2i1WoTDYY4ePcrMzMyIYIwwwh5g7cmnqH32sygkttw0ax9kSAG1SJjcu99Nve7kUQqHw6TT6X3u2a1xIDUZ4JhMcrkcV65cIZHJcPU1D3L81HPoSjjsVai2qJLgILvVstltPPZzzu2c2+mLGoQtBBu//duMfed3Ejm8sxwKq5/6W5b/n9/D/trjCDaLbW1e04kkaScYmz28WT9vhYAbac/n3Oq8rWrFzV2LwqVPYtOrpJsTa1BXYytFMHhYoIgLSVQL8dzr34j1fd/NXW90kuA0Go0tBc58rYdtk06nOXToEKlUamu/3QyfXq4M27bb2pJS+h8hxI4LHwWvFfzZeT3vWsFrbtdX73M7EEL47d/ufW3X5+A1O9vt7PugcLNnEOxLt/543/XSrofOcdDt2XT2JdiP3Xwmw4jG0jIrf/InqHLZ32SqYPSqi37lWT+ybBDXM5BcH59l6uhRqtWqHwBxEPy6DizJCD5kTdcJ3XU3a0aYsN2kprYmiQoOspux2psll+p2nqcsuWkIp9g6sLqet+U4AXNXWfiD3+fov//3N7nAzXHtV3+Htd/9MGpxEanc3bQw8XJ2KuFUJOx2fYFEIRAqsBR3uZ/t0PY8b+O87Z5n57NfePMxVg/neeLb7mv7++yL82Sur3L0T59wBCugK4UuBBqgu3/zPr//mz++pQ8PfOEUR//sGyggI0Bg8I1v/zaW/8EbOdc4DWdP+8eGhMFbQq8nKiK+EDcMg6mpKWZnZ7vGrz/20lf5i1Of2PL3//FN/8p5BkJgGAahUAjDMJwxrml9LRIXli9ybe06f/XCJ9v+/qr8fYzHCrxm4tVIKdE0DcMwMAwDXdfRdR0pJZ88+Wm+ePGxnq/r4Y7UUb7vzu9G0zR0Xd9yX90WzxfmT7NSWeXpuee4tHFlS5uvK3wL6VCSO3N3kI6k/f567XvPywsf7ndxVUpxYfkSc+tzXF2b45vzT285ZiY+xd2ZO4noEe4r3OPfZ/ATfH/BvvQzDr5w7kt85uzDXfv7bbPv5LWTD/rv0jvPe5evBJJRffFFzIcfwpI60jadjaa6WW6i3rQct7uO+Md3Oe925Wewn1KTlN/wWiLRKPX1dXRdJztAM/pu4sCSDABd1wmFQrRaLQp33snCvfdyz3MnaZgWUnjqi9vcaXFzorDZUnub/Wa87EzCtV0/lYDSQ5/B+h/+HVqkt3TTVqnM3C/+Assf+ROMatHdsVv4sbMCNtU83fU3Tj+DuoLbw+bz7M1mdavn2cjFeOIn3s75u6e6fn/tnimu3TPFi289wbG/+SaFh18gBESEICwEIQQGDtnYDlUFS5ZJWurMh6Pc+MDf58T3/X3WrAvQUdusqVo8Uz3J8cphwFFhHjlyhGw2ixDCz4nRLV9GJ27cuIFt2+i6TiTiZPHzIql6XSRqzRp//dyneGLuya7fP7/yAqzAY3Nf482Z13MsdcQvsuRdMxQKYdt21/NvF9ValWvXrmEYBpFIhHg8TjweJxqNEg6H/YVXKcVj57/GX3ZZdDvx5NIzAHxh7lFmwpO8KnM/h1IzRCIR/7OTBdbry+fOfp5Kq3LTY+cq88xV5gH49OWHeXXyfu7J3kU+niMcDvtFHUOhkE84btWf7cbB6eUz2xKMu+N3Mm7mmZ+fJxwOE4vF/OcMDvFVSr2siYZVq7H22c9gNJpYUuLk+Gm/X3n7osiFk4q8nzxA268pN5eJznmBfkubasgg963vcoorCvxxcRBwoEmG5/gyNzdHfmqKhbe/C/XMi0QxqWO5jp+erszVmwEIV3A6Xo8IoZDKC9V08tor9ztEYBEOnieUW8tDumfZLqcJnEfAMUTg/1sI5ebRF93PC/TTAsTlKyz81m8x8y//+9t+NsVvPsmN//M/0HjycTS7GUgLbrntS6RSKGEjneDZrdoWBQIrcH/B+7nd52lhbzkv+Dzd7272PN20opUjY3zup7+dYvrW2e3MiMG5H3oL8ycmmf7wwySkIC4VcSRRJGEgvM26X0QRVYIrYznM7/keTrzhDdRbLRpm9+qpG1qJM7WXyJXTHDlyBNu22djY8AV6cDGF7cXLpUuXsG2bUChEKpXCNE2ALTvzW2G9us5vfvX3uFFZuuWxDbvJl1a/yrnll3h17H5SqZT/icfjPknqF+VKmTOLZ4jFYiSTSbLZLGNjYwgh/Pt68srT/PWpT1Fu3nxB74a5xgJzNxaYWhznNclXk01mSCaTxONxYrEYkUjE1wLcjsngyStP89cnP0X5FuRiOzxXOsVzpVMc0WZ5MH0/2WSWRCJBLBbziZXXH3sbU1O3cbBYWeKjJ/+y6/EpEmSWE5xfPE84HPafs1KqTbPycsfa15+g8vGPYQiFhoWSnqxzZI1Um9JlU8azRX61yW5hIbac111+da4xjvx0ikoq1Kb8FFaX8zbbdLJJOzJaAWFpc+Ytb+H4iRNUazUEgkQicWAI44EeeV4o6/z8PEopxl73Gq4dP8rMxbNULRtduQtnB5v1WaK70xbgDhzn9825367XEt5A8M5TgWbajt08ZvM81fa3Tf2A8PvX3q/AsY0m5Y/+Cc0f+iFCkxO3fC4rn/wUyz//HzDPX8ASOqDheyr4bav2a6Jck8hWKD/y5Bb97Po8JcE6LEKJwPu4vefpXe/rP/zW2yIYQZRfc4xz/9sHyP3sn5PVJFlpkxKCZMBPoxMVpfjqzDSRN76ByUSChRs3aJkmZtzc/jr5OmrRRFwRtFotWq2Wv3P0dq6e/XQ7P4YXX3zRqagYjVIoFIBNbV0oFOp6Tjf8xbOfuC2CEcR1ucjq2jr3LBynUCgwNjZGLpej2Wr21E4nipUizzzzDKlUirGxMaanpxFC+Pf1V6c+yTevbzVD9Ip5tcji+he4b+kuZtLT5HI5MhmHcMRiMcLh8E21GrVmjT958mOcXDzdpfXecdm6xrWVee5buosj6UNkMhnS6bSvlQiHwyjVXUvUOQ6adovHqo93PTZkGYTOCU6WTvom5LGxMSzL8rVHzrX62IofILSWV1j6nd9BlEqYeOu/RClvnjtyx5MtaotcZ4t8dn53iEf7eW6bHevBZhvB8x0ZH1yHhO1unNrOa7+u8rTMKFpKYrzv/b7PmRBi6BNwBXGgSQY4u8RUKkWpVGLs0CFefMNrmbx8HsOyaJ/CAeNEpzFNde4wg99vc547QLqfs/X39nO3P6/zOOePAvPCVRb+r/+bw7/2X9gOdrPJ9V/5VYq/+9tQXMeZDJY7aLdbUjdJRre76P7XbfrZ8XvnPW733DcP3v65XP22+7h6ZGxr928D6miB6z/+Nsp/8BgNKbE01ynO7v5M5qJxinccI1qt0rp6Fdu2iUQiWFFrW1uO0AVrkxVWn1ylVqth2zZSSt/PwdtJO9qI7gL/8uXLtFotEokElmW1aRVu12xxav40JxdfuK1jO1GPNjm5/CKHNjaoVquYpkm91V17c9ttNuosXblOLpdz6yzEyOVylKol/vLCJ7lamttR+0FY0ub58BmWryxzeO0Q4+PjFAoFsllHm+BpNTqJRi+an576Iyye115k8doix1ePUSgUyOfzPvkxze5aos5xcC25wDobW46TtsR8vs5Ll67RbDYJh8OMjY0RDofJ5XK+Furl7vSpLIv1hx/C/NIX24iBp0B2Dtpexvcrl7qtB51t3FLO3/Q8RURTnLv/1cy86tXU63WUUkQika7O5MOKA08yDMMgl8uxurqKkU6TecfbWX/sixSuXmfDdW18OcBWLapf+gKVp58m/prXbPm+dvEiF3/25+DvvgBmE0sIx9SiukfYHDRcevBw179bXznD3K99jlXbJvKe+zn0995AfDLTdszcw89S+r0vUnBDnDUhHHPJNp7ZKhZnde0GWrFILpcjnU5Tr9edhf4mztzhQoSl8XUuXryIlJJQKOT7B/x/7b13fCR3ff///Mxs39XuSqtdrbruznfn8527jW2wAYM7NsUGQ+jNlNBMCQRCQvINIfklpBBCQs+XLyFgQokDhG6KaTa2we160an3lbaXmfn8/tiduZW0utNVnaTP8/HYO2l3dvazo5n5vD7vupzYgLGxMSqVCk1NTfh8PrLZLKVSCcMwlr0SfWzkiYbP+yZdTD0wzsTEBKJNp+3KTgKx+auhyUfGMPeWMNtLTkzA9vatPGXTJZTLZaanp+nv72f/zEHK5y4ez+7vP8n+H+6eZ6mwFwFer5disUi5XMY0TX40/DMG80sLDG1ckk/lGH9ghHw+77wvfE4zwWSIxEXt6J7Gf4zRxDTZvVkymQyFQsF5r30M64VGsVLkiw995agCQ0xLCpPVsdTvT/e7iF+SJBAPEtnQvOT7x6MzFA8U6U6nKRQKjhvMMCoNt68/D/StXpZyWJm7iww9PkA6nQZwrCX28V9uDMhqpzw1ReozX0AzC1hCny80VjUCj2Vh3ngzbp8Po1AAIBwOryr31+oZ6VGw/a/5fJ5oVxej3RvoGBxZMwIDAKFRmZpg8iv/MV9kSMncQw8z+EfvxNizD5dlIWpt72vGPKohUKv7kiv5GtSYGJtl+F++z7hpkpaSyg8eY+CHj3PF372cSF+1q+mu/7yfiW/+Fr+oBnz6a66SiBA0LTFxlypl0uk0LpfLmRwrlQrWEubtepq3t3J47z48g555wXder9exZiylF2ZnZ6ufY1lks9kjn7sgrfFo5Mv5Rc+5ijqTvxljeHiYVCpFebDM0CMDXPTay2nqjABw4Ht7mHl8klAo5PRDiEajhMNhJ2iw3jJTZrEbpWJUSKfTaJrmxKH4fD7nONoBkPuNQwyUhxqO3zPnIv3oDBP946TTaUqlquCx3U35/gzFgRzTD44Tv7Kdtss6Gh/LrjzFx/dRLBbniTR7Ve92uxFC8M3HvsXBVH/DfbiyGpnHZpnYN+akLNePhbJk+sFxZoRgonmEtqd1Eu6NLtrP+P3DzM24MTurLjT7eFSWEBn2eRDcHMZsHN+M6DcZeOgwk5OTlMtl/H4/uq7j9/sJhULOeefxeJYdj7IakZZF6sGHKD2xC0vzIjAQQsIyrtWzHaFL0nqQ4JYtVCrVc0XXdWKx2AqP7PhYEyLD6/XS3NzMyMgIsViMfRedT+q3v8Zbtiix/Fz+sx0hJel7v8fUtTfTetN1mNkck/d8lYl/+WcYO4yGhkSbVysCibMSctyAq5BG4zZ9bqYsi1kpKUmJDniA3e/9Etv/9mWM/eBRMj96Aq8QjtAyag8TcC1xz7VMi3K5jGVV/zcMo2rFWMZ55PK6aL26neEfDhMMBp2gv/oAxKX+Crao8Hq98z73ZDM8DAzGx8eZnp6mUCg45YgH7z3Ihju2MPPoJOVDBfx+P5qmYVmWE1dim9xdLheBQIBIJELEiJCj8crfNE1nEg4Gg7S0tBCPx52GcP7mAE+WH2v4Xu+IzviDw0xMTJDJZJxqqfVZL7ZVSAiBedhkJj9O9Ko4mnt+UKwe0JlLZKgcPDgvw6c+iHbX2G4eHGqcfeMaE4z/cojJyUlnLLY4qM/AsccipaT0YJaZkRLRS1vRPDpm2WT//+4ieyBNNBp1xFskEiEajWLIxjaKYrGIK+omcVVjAaVPwuCvBpicnCSXyzn7bWlpobW1ldbWVqLRKMFgcJ64XYsUDg8y/uEPI0QRiY6wNKSzpFp937k+ODwsNZ54+tPo7u7CqLUmsP+uq4k1ITLq292apkn31U9n5Bc/Z/MTOykbq3diXYhEYE1NMfvpzxK+6HyG/+EfyXzh81gaCFwIq86fd4xU0NVGbCTF0IKYDD0apOt9z2XkI99EUk1TbdE0oppG/v1fISIEfl2nXBNdQU0jKAQ+ASEBZmhpE/dCqsWOlrdtpCtK9rw0IwMjhMNhZ2VpBx+erGg4Gp3hjsXBiz5B8zUJ+r/Yj5SSQCBAIpEgFovhfVKjw50ktqHZ8fk2NTXNC1C0gwd9Ph+6rlPwFBkxFosMl14VIm63m+bmZjo7O+nq6qKjo8MRGbv0fTQwguCdcDH2wBCjo6NkMhknDiYcDtPc3ExLS4uT8eLxeBBCYBgGhUKB9O4MlfMX7zOwoYmRx/rh8PwAWju19X92frfhMXSPCUbuH2RsbIxsNuuMJRKJ0NzcTHNzM5FIhFAohNfrdYRZqVQim82S/l0G4xyTyQdHMYbLjiWoUqlQKpWOCEit8XngCrrZ8dKLG7qDRA4mfj7C1NQUhUIBl8tFNBqlvb2d7u5uenp6aG9vJxaL0dTUNM+SsdYw8wUmPvkpRP8BADRZy/ag9t8qvvF7kBSEC88ttyJ0HWrifbVZMWCNiAzAyQ1Pp9O09fYyduHFVHbtQjcl1io+2RaiaZL8I79h3223YQ0OoYmq79FCgqgGKC2VJbKaSe4e5dGrNi96PnHJRp7yly/mt396D14hCAtBq6YRFtUmc4aUlAFTSryaRosQxDU3pWQbUy9/OTQIqFuI7SpoCgWYLM8sa7xtF3Vw6MAehoeHaWpqIhQKOZO2eRpFxqbYBti/+PnY1jgXvfpyHv3CQ/j9fpqbm2lvb3dW2IZhOK4Fv99PS0uLM1EFg0F8Ph9SSrxeL2kyMLr4M9x1bpFwOEw8Hqerq4vu7m6SySS+iI+RofHF7yvqTD0wxtjYGHNzc0gpCYVCtLa2kkwmSSaT81bntqvDHnM6nebQzCCzLZlF+3Zt8DH8+LBjgbBTScfzE0w0iMPQCxrjvxp2xgIsGks8HnfG4vF4HJFRLpfJ5XKk02mmJ6fxNemEeoIUi0V0XXcyXZwYHdk4JXnjLVvwNvkWPS8LFlM/GmVyYpJsNosQgnA4TDKZdARGR0cHra2tThyMHQu0Fknv2kn6C59FCKsWcGbWAtxXpxXDRgBe3cXeSy+lbesWJ57I4/GsijLiC1kzIsO+cQ7WsgFan309g795kL79B6mYJiwZPrXKEEA+h3HoIELXaulR1VQnx+8MrGoZ34DEr/fTc9U5DJy72ISc2NbFM/7ldRz4t+/j2juGX0CTJggIgV7Nr8GSENAEYbeXuW3n0f6G17N5xw4e7v/ykp9pl2r2+Xy0t7eTducXrcKNXIVcKkuka75VxOV1kXh6J+M/HiccDs+brI1KY1/8qaAv1suF8R08Ork4ALT1nARXvv3pjPzwMG63m0Ag4GQ66LrudIv1eDxOrYV6sztUiwBFy9GGIkPXdSfY0953PB53JueDxcMNx1zak2diYoK5uTlM0yQUCpFIJOZNnLFYjHA47LidgHnWg9hMjJ+XfkNZn39sY1tbefKnv3NiFew00kPm4kqiAIWdGSYmJkin01iWRSgUoq2tje7ubrq7u+dN4n6/3xE8tqWiUCiQzWZJpVJMT0+TSqXIZrOOcLKzXYLBIK5i48DV5o2LV6tWxWL0vgFmBqbI5/PO/uzj1NvbS2dnJ/F43DlOtitnLWIWCox+6pNo0sSqb8HlLLBW5/1PAi50Si4d4/m34vJ6KFfKWFISDAZXRRnxhawZkWF3Zp2cnCSVStG7bRu/ufxyOg4ewmVKKrYPYa2giZqoqE/Vqq6Q19C3nMd533iQybfcQKFBrYxQMsqO99/OxBd/Rvmnu5BS4hZVoeFB4tM0rKYo01dfw3mvejnRWIxMZvHK18YWbD6fj5aWFvx+P2kWB1UC7Lr3cS59/VW4vfODUyNdUfLbsowOjjrWjGAwSEmeXFro0RBCcO2Gazg8O8hsZbGVJhgPsfGF51J4OE25XFVMXq/XCfi03QqBQGBeEan6CWspn7CmHSn8ZAuNpqYmJwZhZuD3i8drwsSuMSfY0U6/7OrqYsOGDfT29pJMJmlubnasAPaN1rIsx2XS1NTEluFJnijunrd/3etCi+lMT08TiURobW0lFouxv3Jo8VhKML5zlHQ6TaVSwe/309ra6ozFdkXYY5kfzCsxTZNyuUyhUCAajdLS0kI6nSaXy2FZltPQqqWlhaamJlzG8hvmHf7JAcYeG3biVOrH1tPTQ2dnJ21tbUSjUcdltZazSmZ+9jPK/31vLQ5tKVbnnTCk6+zZfj6dF15UDfqumeJbWlpWeGQnxpoRGYDjN7XL8kYuv5zZH/0v8dEpTt/aUXEmEEDTwSku/MLP+e0bno3ZINtE97lpv+s6cn0J0l/4OV4pCQJNmk4lmST7nFu58sV3YpomxWLRWZ0vRFIteGNPtse6URcm8hz81T62XnveotfiFyYZ7D/AyMiIIzTyocKJHIJlIYSgOdjMc/pu4L/230tZLg6A0D06oauamTk0RzAVdNwIPp9vXkaMHYuxcMLyuBsXBqtvylWfwmqX+h7OLDZ/WDMGc3NzFItFhBDO6ryzs5Pu7m5ndW7HP9RXPrUndnus55nn8kT/7kWf4Yn6yIzPMTMzw/T0NHNzc6T02UXbVaaqWUX17o1EIuG4fLq6uojH406sQ31ast2szDRNAoEAgUCAcDhMPp+nWCw6xbH8fr9TWt2VWd7td3LfOLt+9ISTFWNbZJLJJF1dXXR2dpJMJteNwCiNjzPx4b9FR7L6c0jmI6RACMncpZeS9Hio1IK/7aDr1ciacta5XC5aWlrweDzMzc3RtX0HE+dsBd0ucV2lVqX6qCxnG8WZRQdaHx3kgo9+G9dcY6sCQPD6Cyj8yfOZsSRZqXFg02bmXvs6znv+88jn85iG4Zi5G6EJ4aRbLgdN09h17+PMHJ5a9JqdbTI+Ps7IyAhjY2PkcidWsno52DUSOps7uK39RvxysW/f2XaDm6HoGCMj1UBCe8Vtuz2WW99juUzkF8dAFKYK1b+JaTo+50QiQTKZdIJTw+Hw/FiGmpCxG4H5fD5CoRDdia6Gn+tr9juxG7Ozs6QyKUrWYmtScTZPPp93XEb1Y2lra6O1tZVIJDIvy6W+46otrPx+P+FwmJaWFhKJBB0dHY4QsPcRCATQteWZvuOb2wj0hJyx2da1ZDJJe3s7bW1tTsGxtRzoaXPwwx/G2rOT6hrZLvTHajVcOEjAp1v0t0RoufKKaoxdzWUbj8dXbWzN6hz1UbBXCfl8nkhzFO3ZN5ANRvEKDUStBr2UYB1psbzwYb8urOr/SPVY6YeQR1JUWw5MsPkvv4E+NrvkeeDd1kX+wy/i0e3nULntOTR1djA6OkqhUMBs0FK7HruI0XKx60c8/o3fYZQWlx5v6gjj2RZgdHSUkZERMtn0svd9vNgiw+fzkWxu49rI1fiNpYUGrToDsTEODfUzMjLC9PQ0mUzGCQI9nhodJ0KlUp7ntrEnZzuD41hpmPYE7/F4aA43zhbShIZpmhQKBTKZDJlitvFYjGrarqiJzEgkQiwWcwqy1WcILXXDt9N3bbFR7y6yM1KOVB5dvhB4yiue5ozNttguHNt6EBi5x3dR/s+vYokypqzMu1eLNXCv9moa0zfcSKK317n+PB7PqnWVwBoUGbY5UQhBKpVi0zOeycD2C3BZ4Foi4ri+86eQHOmXY/+jHiv+EBJ0CR4JAQSxqRwb3vtl9F1LV4109yUIvOFqxtMz7N27l6mpKdLpNPl83ily1Zjju0nbk5yZMph4tEFEJBC/oI20mWFkZIS5zOkVGbquO42y2uNJrglcQVNh6Z4vWrPOWHKa/f37GRgYYGxsjJmZGafiqB0QejowTBPTNJ1jGAwG52XjLGfitIWGKRqPUdTil8rlctV1kW3cg8b+nnYxsVAo5IxlYQzG0Vho2WjUjbUqUhrvp5FQDUQCPOVVT3UEbX3tkPqma6t1tbscjFyO4Y/8LeDCkhpgIqREkxwxYpwF96oTeUgJIU0w1NJC9NpnodUCsU3TJBaLraoKnwtZk2ekHV2dy+VoioRxPfc5ZKNhvNJA1vrW2CW2NVl9iPoHYGnVxyrOhFpzCKot2n12qqoQ9H7kXtz3PrTkezzNXvJ9ZYaGqjUYJicnmav15iiVTk0Apm098Pv95B6dw5huEAdR5zZJLyEyHEvaSWJPRIFAgObmZpLJJE8JXUIitXRdEK1JZzw2w/79++nvP2LVqBcaJzu2uH9x1oQn5nUmdtsCU9+qfbkrcyEEmVLjQF6rVM0sswMzy+UyrgbhaK4mD/WVTW1xUN+m/XisBHaVTdu9Y7tWjrWPw7840PD5jVdtpnVHmzP52A+rzjK3VgoPNmLw058n9+N7sfQi1W6lR5oqOvfrVYoOeAVM3fZcEhs3OtY9TdNWtRUD1ljgp00wGCQcDlMsFpmZmeGcq5/K3nM3sfW3U2hSOMFCttio7/4pQAmLsxS7RLqnVg/EsgMAv/YgUwfHSb/mmYjo4syHUHeYieERhoaGnLLQwClbnddbD0KhEN5DOlYU5AKPS1NHmNy2DJlclhCRRfs52bbqS43H/p5CCPaN7mcgPIpsEPPqSfoYHZ/APGA6x2nhBHkyKXRd4U4mC9PznvO1+J2f6z+rfkJe7sR+cLq/4fP5sap7xA7OtCwLn/SSFfMtBsH2EPWdcxc+gDPiihj6xWFCXWHati6uKb715vM4+NU91aJf6TTZbJZ8Pu/EY6zVgM+53z9K+pOfQLOq0XWWENidqVZzT6YqgoCA0XCQpqdchabrGIWCk869VID6amEVa7+l0XWdlpYWvF4v+XyeQChE6eprKfojtf5WwrGvScASRwSGFNVEUBX4eXYiwInNCAlBVNPZpGv0PjGK+Z1+zHTjPKLI9hYnJmJ0dJSZmRn8fn/DbY97TLVJyak94Y3QnGncJTF+QRuGu7G5/lTGQNiCwK6aaRfGOrdrK5tzvWj5xndmz6YAhw8fpr+/n8HBQadmhJ0hcTJj6wgvnjTdIQ+xbdW29vUioH5lvtzP/GX/rxc9JysWmYH0IvHQLBf/fdwhD63bE4vGUV/a/UxYCoQQPPylB6gUF5/LvrCfxNM6SKVSTE5OMjU15dTiOFUWp7MNaVmMfvj/wPQUAg2nxbqou3ev5vu1lPiFYOSqa2ju7nGsGFJKEonEqnaVwBoVGYBTAAkglZpl43XXMdHVg0tUTRerX/2uT5ymb7X6F72aC08kRuGlL+XiG55Hx0QCWVhsofA2+aiEDMbGxpiYmHAm4FM2rroAxGAwyAa9h4jZtGg73esiuqVxaWC7X8qpHI9dsyIUChGLxejs7GRz92a2Vzajlxdf/p6Qh7SeZWhoiKGhIcbHx0mlUuRyuWPEsRybbYmtDZ/vfcYmp19KqVRyurXavVOWM2n+Yv+vGnZSLQzlHAuRnY3idruJicauo74lxmIYxnFP4LZAaiScjoamaZg5gye+/fuGr7duS2AlBePj44yOjjIxMeH8jexzaC0JjdH/+DLGL34FQkOusXu3pNriYLYpiOeZzyLYFHLO+9WctlrPmhUZmqY5JZNzuSwt8VYyt92K6QJdWGvqRF0vSGD/8y/h/j+6BY/UiGlu+reew/S77mbH7bfTHIvRHIoSnl2igVBII52uFqHK5/NO2ehTRb31oKmpiYsCOxr6/pfiVFgLoDq5PXT4Eb788H851hG77kc4HKatrY2NPRvppbPh+02f5UxgY2NjzgRmdyE9UZoDUTZHNy16PtAapOdZG52y3Nlsllwu5zSMO9YxGZ0d4+tP/k/D12b3z2AYhlNjwo75aPMlaNJDDcYSYtNztjqVRDOZDNlslkKhsGxLwejsGJ/79ReYykxTLpcbCqejHUfbKjb64BDDDw823Kb9aV1MzU0xPDw8L9bIPmZrRWTM/fYhUv/0j8iKcSRO0q4cvtq/oqxOwJpL48A1T6frooscIW9ZFolEwqluu5pZ3XaYY9DU1ITP53PK/G577m303/cDun/3+2o4L9Qie1f72br2Gb96Mw8/52Im26rKvvnlVyCH/YRfdAcd7e3k8/lqIarmZorFMhkWZ524fG4qlQrZbJapqSlmZ2fh1Bkz5lkP7GqX2zNbeDS/c1nvL5VKTjOtE+Xg1CG+/eT3ODRbLeEd8zRzYeL8I51kwSkw1mV1cTCzeBITbo25uTmmpqaIxWLMzs469R1ORgQJIXjWxqez75HFgY09T92AruukD6SZmZlhZmbGqaxpVxutj42wGZ0d419//ZmGn2eMl5naNeF0ULXraQSDQQKBABdqO/jF9G8Wva/jkm5kRTK3f84ZS32hK2DJTI4nR3fxld99jWwlx3hmgpdsfyEB3T8voNSuQbLUcbQFka7rzDw0Ttu57biC82/VutdF9Klxxn4zRigUmtfttz5IdTXHZ5TGxxn5yJ9jHj6IhaBR5a3VfOeWQJNwMdbeTuDW2wg0NTE3N+eUs29uXn4Dx7OZNS0y7DS+fD5PuVwmHA5Tuf1FWE/uwZMvUNZW8ym6Pii3Bnnwtc/kwIKeJQ8/azsXGFswcDE4OOjUJnC73RiuJVpop6qr0XQ6zdzcXLWs+CkUGTC/TkIoFOKy0CWMH55krEHX0oWUSiX8fv8JTeLFSpFvPvatRa3Lv3fox4iCoEWPOtYMO60y62pcFCw/k6t2N607ToVC4aTdOZqm0Rnt4MrEZfxmYnFGUOcVPeR7s0zun6J5fHxeMKNdlrt+cv/FgV/zzZ3favxhJkz8ZoRcLofdXMquWWH3kWkPtzNQGGQgv1iQOmPpnyI6NuYE4AkhGlb8LJQL/GD3j/lZ/y+dfUwWpvnio1/m2fFn4Nd8TsyOLXKWCvStt4j5dT/W40W4crHVJbqhhfxIjpGREUc81RcKs/ezWoXG1D3/RfkXv2GtRuJrgAeTqcuewjnbtlEoVCsBm6ZJOBxeE1YMWOMiAyAejzM7O+tU/Ot76lMZ6NnIhj2/R5N6rbmOEhtnK56pHOnmxu6Px1x7CU57CU3Uums2BZgOzzLuWlx5E2CuP4VhGI75+lSlsC6kPrvD6/XyzM5r+Prhb1E5RnH744lBWIjP7WMqN93wte+O/YguI0mP1Vmd4MIB9pUPsjOzt+H2s4dmFh2nejP/yVgyXC4XT+97GgdmDzFZXjzeQDIESTg8PUIuVazGJ5im06httjzHYHqInx7+BbnK0lVfs79LMXFovJqu6nI5Dd/slvH2yv/ZPc/kP/f+FyVrcdqxPZaB6VFyqSKmMJ3jYtfNeHjkd6SKc/xq6IHGx7KS5t6R73JB+Vwi3rDT00RKiWE2DgCuDyQOBoNERBh9xkO2pbho27bL2hn81qFFQqM+/Xc1NtWa/vFPmP7s58DSqBYPX2P3aAkBDcZbmtGeeiUen49cKgVUaz01NzevWnG4kDUvMuzurGNjY5TLZUItLZRvvxXjH55AlGoZJSs9SMVRefpXfsPX33Fjw9dysRK5WAmYPeo+ZvZMUs5URYUdjHc63WT2JOH3++ls7eDC2e08NPf7o76n3qVxItxxwfP4u599rOFrQ64xhhiDCtBYiwBLH6dTcbzqxddz+27mG/u/RcpqHBcjY4JJUvyw8nMYpPpYJoUnMgw82E82W01dDQQCTonv+jbttivm1q4b+Z+B7y0pAmVMMMUsPzV+DUNUH8eBZ0yjf+KQ0wvFsizcbjem0VhkwOKianEzzkFjiIJrvtDQvS7aru1g7HtjBIPBeYXD6nuYrKYiXYX9Bxn76Edh6DCItbkIdEmJF42hK65iy2WXUygUEEJQLpdJJpOnLPPtbGD1nHknQWtrq5NJkMlk2Hj99QxuuwSP5gGO5OKrx1n2QOBF0PzEKH1PnHiQplWxGPrlYaSU6Hq1U6jdcfR0sdBtcmX35SRd8aO+53hTNhfS1pTg1s03ndB74ejH6VT0MKmfOFsiLdyQvJZmeWqj57MPpTj08/3Mzs46zdNisZjT58Nu025PxH6/n55YN8+KXYPPWl6vmmVjQPG3GQZ/2c+hQ4cYGhpiYmKC2dlZ8vk8xlHqotRbM/x+P5FIhB3uxtk5/lgQ11Yvo6OjDA8Pz6vYuuqyTSyL0Y99nNLDv8cSGg0DMVY5EvDpGmPJDgI33kQgFKJYLDric7VX+FzI2vkmR8G+SEulEoZh4AuF8L3hLnLvfheuXAlzDSrl1YlAIJ2/RgDIRFsZfc/dPPWyy/BPPsYuY/9x7VFWLA5/ex8zA9MIIeaVi16qZfmpwnYPaJpGOBw+ptvkVJhHn9J9Kbl8jp8M339c7zvWcTrSb+PkhEZ9PZFELMG1xat5YOQhhn3jJ7xPAHOmwsQvRpjsnyCbzTopgLFYjI6ODrq6uujo6KC1tZWmpqZ5K/1gMMiGeC+uosavph4k5T/5su+VsRJjPxsiM512umjquk5zc7PjfrKOMYHWCw27q2shV+SJ0p5F27Zd1snAyP55bhO7WulqcpuM/b8vk/3GlxFmBVOr9iwSUqypqAy3AAPB8I03cMGVVzA3N4cQgkqlQnt7+2m/L51p1oXIAIhEIkxNTWGaJrlcjo4LL2DfdTew6VvfxFx7YnlVUi2MpqHLCj7hZrRvI+W738bmyy9nZmaGbq2TUqHEYdcwpvvYf7TCeI7DPzzA9OHq372+6VUsFqvmoJ9mfSmEWLbb5FQE6WmaxqXtF2EVTB6YfpiiOHbcyXKOk90F9WQnqkVugHici8sXEBrcx8HSAJXk8V2Mlbkys49PMfHomBOcChAKhWhpaaGjo4O+vj56enqcbqULm4nZBcsSiQSXlS9i9/AehuQosu34Db2F4RzTvxsndWjGSUn2eDyOa6beMiTM5ZVLrz9el4UvYXhwjJS52LLX+tR2xv53ZJ7bxBaH9nc9m90m2d8+wtjffRRZKoOQteJaa0leVG83fgRDW7eQuOnGamyOYTjBybFYbFWIweNhXYmMSCTCzMwMUkoqlQq+59xM7kc/wJPLYq6xk3lVIqpN0NyeAHuvvprWP3wzHV1d1VRTqt1R27Q4csTksDlE1swR2Dy/4JVZMhj7/Qj5sSwzu6ecOAe7I2ZbWxvt7e0kk8lqT4AG8QmuWmS/3b/iaD007EnKnpwXZh3YN/Z6t8lQbrhhtklTU5MzMSxsa76sw1e38j2neSOejIvfjT/GVH4KuuZf6sd7nOzsiqNZMlz6/N4j9W6WeupX5/X9NzyDHgZ+N8CEZ5q8USC4KYwntNiFMbt3mnKmTObwHJmBOWcfUkq8Xq8Tg5FMJunq6qKnp4fOzk7i8bhjxahf3dsBlvVj8Q/6GPjNAKlohnwhT9PmCJ7Q4vLOlWyZ1J7qSTT2yyFM05xXm8QWB7FYjNbWVke0+f1+9ELjycTn8znCws4Gsru6ejweru+8lm8OfJsK82M6Aq1BQpdGme2fZWZmhlgs5sSf2N/5bCXff5jxT36CSv8u3L4g0oIjMfly1dua7WvGg8ASbqYuuogLN2wgm80ihMAwDCKRyKovId6IdSMyANrb28lkMk6EeMd529lzww1sv/dr5Ex9rYnm1YWopnOZwSD7Xvwytr721ZiWRSaTdSZae0UYiURoGy9ijpQZe/gw+XyeYrFIoVBwVgVwxIpgV+FsbW2ls7OTnp4e2tvbaWlp4e7uN1OpVJiamuLgwYM8/vjjGIdLTiE323/v8Xi4KnE51/ZcQz6fZ2hoiJ07d7J3716i0SimaTottxe6FjRNcyaNcDjMrX03MTAwwIEDB5yGZOPj4yQSiWpZ8mh0QUvw5bkoFq56Y7EYW/Pn4JtxM3DfgFO980SOk7361zSNvlgvf3HtB5iammJwcJA9e/awc+dOzKHqcbMtA/XWj3qhUS+8gsGgMw6nRfqwn7HJMaa/P8Xc3ByFQmHJzBtbyNldTu36AolEgo6ODjo6OpxYjEgk4tSRsN8HOFVRm5qanPLj9nMjIyOMT48zs69aXt0+do3ST+v3ZzfMsyutJhIJOjs757lsrko8hcs7L2FmZoaBgQF2797N7t27iUajVCoVpwOs/d3szqtdrk5eWHku/f397N+/n0OHDjE2VrXk+Hw+EomEcz0sLGhmf7+zCatUYuCjH6X0ja/h8dd67dT1lForSKpN0AbOPYeOF75wnhVD13VaW1vxeE5xXNBZwLoSGfYNfHJyEqEJDKNC/M47mbr/PqJTs5SUylgRdCQuDeaa25h921vZcfPNZLO1ctCiGqlhT0CBQMC5yQshSCQS5HI5pqen501IdvEjeyKrn3g6Oztpa2sjEong8XgolUpOoaZYLEaxWCQYDDpBWPXuAnsyts+leDyOy+XCsiyamprmrfrtyQzmZ5tEo1EnNdQwDDRNIxAIUCwW8Xq9zoq30X6ORr3ICIVCtLa2LmpdfqLHye/3O1aJ+sJStqCIxWJUKhUKhYLzHez3NRJKtuiwa0/UT+zBYNC5TmdmZpyuucVicV5/F/s99irfFhitra0kEgknm6S5udnpzNxoLPZ+7ODw+r+V/TednJwklUqRTqedDr4Ly6zbwbJ2QGk4HHbGE4/HicfjtLa20tzcTCgUcv6u9eK5tbXV8c/b56MtNDwej/MdYrEYpVKJcrnsmNrT6TRCCCf+Q0o5r1vr2Rj8KQ2D8a98icLXvoKuu6le7WIN2C7mIwE3UPR4mXv2dWw+5xwmJiYcK0ZLS4vTBmOtsa5EBkBbWxupVArTNKlUKrR0d7P3la+j9Z8/jlkxMYQK0DiTuAGXx8Xw+RdgvOq1bH3qVczOzjZcsdoTaDAYJBgMcs4551AoFJicnHQmz/qI+voS3/U3e/tG39TU5Exw9gSbTCbRdZ1CoYCu646QsKvH2isNu+lYpVIhHA4jpSQQCDSclO3x10/msVjMWQ17vV5HeOi6TiQSIZlMOhUv7YlxOdRXHLV/tz8zGo2e8HGqdzHYKy/bBRSLxcjn83i9XsrlspPn39ra6gilRu4me6z1IqN+rPF4nFQq5YzXLm1ev/qzxUAoFCIcrtahaG5uprm5mWg0SjgcdtwFS1mF6t1M9T/bwtNOf7VFht351BYatuCx3Wt2wa1IJEI0GnXGZBcCszNb7GNpH/t4PE6pVCIUCjmZMbZIqhe5toXNFjm6rhMIBEin09XAdp9v3rljf9+zTWRIw2DmW//DzF9+CFExsNDRsNCkdBpX2gh5pEP2arxD64BH0zlw6WVsecELSKfTTmq42+0mHo+vmeJbC1l3IsNeYUxPT6PrOuVymU233cbh7/+IrU/uZAqJqFPRa81kd9YgwC808PrYdfPNbPjDNxGKRknVCtI41I5/vUWgvb2dRCJBuVwmk8kQjUZpa2tzekwsXKHbJb7D4bDTOM++0dvYK3+A5uZmyuWyY2Go923bq1W7v0AgECCfzzvxANFolFgs5rgXFq7e6zujQrXwjh0DUS6XnZVoNBp1zOq2m2JZh7UuiyAQCMzvDnsSx6leJNQLvnA4jGmajrm3UqngcrkcIRWNRh3B1Yj62JV60WBbEDKZDLlczrFklMtlZ1VufzfbcmAHPC4ss20HW9a7SBodN8DZrt5CEg6HaWlpcYSOLTDsxmk2dmCn7dawszzqS5nXx/hIKedZtyzLwuPxkM/nnZ/D4fC8jBh7fHb9H9tKFY1GnSZ29vG3hU29i+hsIvfYE4x84C+xpvJIoYGo5tto8kg4hhTV35FHvNln17dYBqKail9oiiBeeAeRWAsjIyPouo5lWcTj8TXRCG0p1p3IgKo1w14tW5aF2+sl/9znwK6duEwwBWg1uawaqZ1aJOCS4BGCQmuCg7fcyLa77kICqVRqyRuhaZq4XC4nGNHulGmvNqs9S450zLQnIdsMbt/4/X7/ohs9HOnn4XK5iEajjhvDXlnbq2F7n4CT/mhnNNT74RdaMmzs9/v9fmf1a0/I9oRlT25NTU3z3A3LxZ6wG63KT/Q4LZyg7cnNXsV7vV4nm6LePdNIpCw1XluE1Y+33rVkj9l2l9Rvb1dXtb9D/diXm7VTX6PF3ne9+KqvgFrfmdXGFqB2oGb9WGwrhP096y1CdjyIpmn4/X7HwmT/TRaeB/bxtj/TFnv28bddMLbwskXG2ZRZUpqYYPzjn0AbOUxZd+GyLCTVDqtWvZqoa4RmiVUoMACXtHBrbg6cu42+Sy9lZmbGKZVv32/Opr/NqWZdioxQKEQ8HmdsbMzxf/Zcey2P/exnnP+rXzNr1Slpyeo8s89CJODBwu9ys2/7DuRrXsUFV1/tdLhsNBHYQtDv99PV1UVzc7MTwFa/8q1UKlQqlUU+e3ub+vTBRjd6O6Lf4/HM62JZ3xOlfpW7MCsBjgiIhdvXU29pqPf91/v36xtpHS2z5WjUC4xTcZwa7dueBO3jZ4sW+xjZ32E5hbzqJ/h6oREMBjEMw3nUdzCtH8fCR70wOpFj10j0GIbhHLv6TBQbe/uFhczsfTU6Bvb5ZR8zv9/vnE/1AbkLBQrgiDdbCNeLnnrBs5RQXCkq0zMc+OAHKd77NYQm0KVJffHw+oWd7SaxRF331VWGW5hMJtvx3fVafAE/09PTaJqGYRhOzM1aZl2KDIBkMkkqlXJu7oFQiMBrX0PqySfwzqapWBKEALnWQpDOLPX+4CYh0DQvey6/jNDd72Djju1MT087boKF2C6J5uZmOjs7nSI1CyckO9jOnjTro+jtiaj+/4VBf/UTisfjmTdx2O9bOEHYk099NcX6z2v0WUf7zIWVPu19HG0/yzn2p+o4He072PutH/9yjsOxxmxPkgtLmzf6nIXjPhWT6cLvaI9l4fGrP4aNzgF7X432D0c6utpiZjnnQf13tt93IufiSjDy8U9hfOUrCE1HCglSVksILBEzYlHXNHuV3IydvxMChI+Rm27hiiuewujoaK0fTzUWY620cz8a61ZkeDwe4vE4Q0NDCCEoFkskt25j54texPmf/TxSq57Qyl1yckhAl4KgrpENhzl4y610v+wPiCaqlqSlUupss3symaSjo6PhhVh/kz4Z6m/Ix8PJfO6JfuaJfhac/HFauM/Tsd/6/dv/r7Qp+XSO5WSOY71VbLUw8ZWvkPq3T4CojtuOeXOWcgtEhFzyl7MfgcCDZODcLZzz0pctCmhfyxkl9axdR9AyaG1tdVajpmmAgK5bb2O0awMRTW+QaVLnIFQcE0lVxbYIjZn2bg6//W1c8K53EG6NMTc7i6DxCs8wDDwej1Opca0rfYViPTD+7W8z8v4/Ri+mq9f+Sg/oNKMjwe0i/6LbicZbyWZzCCFq7l8fbW1tKy6gzwRr/xseBY/HQ1tbm+PHLJfLRONx0i+6HUsDtyawpH2IJItFhlSPho9a4RkkTZpg90UXMPfB93H+859PoVAgnU6DqM/hqWIHcwYCATZs2ODUDFAoFKub/CMPM/re96DPTGMKDSmsmvVi7T68Ag5feDGbr7uhls1YnUuklE5Q9Hpg3bpLbBKJBLOzs05baNOy6L7peh5/7Pdc9JP7SGEga5pbyOo/EgmSasmY2hzomP3qfndcLfZsWr+tqD5d/3N1nyxrn0LW3rPEPus/+8zusyoeghogXDx6x+20vexltPX0MDc3R6VSaajebV93Mpl0Kl8qFIrVT37XbobvfhtieATDbt1ed/8U1J5qdD9Zzv0LjmSi1H5utM/67Za7z3k/L3OfEvAISSrejuetb0F3uzGKRYQQTv2TRCJxMod0VbHuRYbL5aKzs5O9e/cCUCwWiTQ3E3nFKxjfu5v4wEFm0NAtDeGcyWatLh1YtTPOWW/Lunzu2knozM1y8c/2iWvvo7rPI78vtc/67JdG+6z/7DO1TylBExBGYzYcZfCu17Plec/F5fU4PWMaCQw7dbOrq8sphqVQKFY/xZER+t/+doxHn6zWwkA6GSP19xCN6v2j/n5Jg+3qf3fuXyy+ZzXaZ/12x7NPcRz7lBIsIXFJjfGXvYwdF5zP9PSMk8kGOHV01gvr2l1i09TU5NRGAMjlcrRv3cLwTTcjdA2XrQJsS1jthwUGilPCcvcpl3i+0fvOxD4tAW4kEU1juG8DI+/7I87/gxcjNcHc3Fx1JbBE/IXP52PTpk20t7crgaFQrBGMTJb9b3sLxoMPIKmWo6+/ycxbgLH0/Wee8DjKdlC3cDvN+2x8L60uQgOayfA5W+m57daqa7i2d8uyCIVCtLW1HeXT1h7r3pIB1XQwu9x4fdOa5M03s/+Bh9j82O+ZxkAIEFggRV0q1eIU16P9vtzXTnSfJ/rZJ7pPq+YicQmJFzdP3HQzLa99NRs7O0lnMlQqlaq4WLATuxNuc3Mz3d3d60rZKxRrHVkpc+Dtb6Hy/R+iaTpyQZ/r03U/W85+TsU+G4kSSwhcQqDLAOU3vwmPz0c+n3eCPXVdp62tbcnqt2sVZcmoEQqFSCQSVCoVAPL5PInubiov/QMy4TBNmgXUqvtJjbUfG708pJT4dYlPd7PzJXfS974/oqWry+n+uFR6qmEYtLe3s2HDBiUwFIq1hJQM/8WHKHz9awihNViGrUWqbd38QrL7jtvpu+pKioWC86qU0ikFv95QIqOGXZPBrpwnpSSXy9F39VPZe8cdVDQNt7BA6lQFxjoWGgJkLTskKATT3RvZ+b730/emNyA1jVwuh2VZS7pHXC4XGzZsoLu7e17/EIVCsfo5fPfbSX3iU4B73dwiTQEBTbD3okuIv+YVlCsVrJq52+5R09nZuS5SVhey/r7xUfD5fHR1dTlBOuVyGd3loueOOzh8yeU0CQ1B4+JR6wUpahX4kDRpOqPbzsP8m79my3NuQSKcPh6NMAyDYDDI5s2bSSQSKv5CoVhjDLzvfaS++B+YhsHRIx3WDhKBhqAiNOSrXk0o2oxRs4jbFWFbWlqcpojrDSUyFmB3cDQMAyEE2WyWlmQbpRtuYi7UjFfT1sml0xgJaELi113sf9az8P3N39DS3UOxVJrXLGree2rxF9FolE2bNjkt1hUKxdph4I/fTepznwLDnJ82t8YRmEQx2Hv9DXRccAHlcvlI7IaU+Hw+Ojo61u09b31FoCwDj8dDe3u7025Z0zSy2SzbnnsbTw4PsfWLX8Jtlqggkdb6uIjq8UhBKRhi35vfRO9zn4vQdUql4pLb2wFPPT09tLW1KeuFQrHWsCwG3vdeUp/9d6R55J4o5cKKyWsPKQQhDfaefzFtf/hmNJfLieuzU1bb2trWddViZcloQCQSoa2tzUlptVsvb3z5K+i//HLcGCDsnoHrBwG4hJ+J9/wR21/2MqQQjlmwEYZh4Ha72bhxIx0dHUpgKBRrDKtU4vAfv5fpT38Kaa59UVGPBPzSJOdrgj98C5F43JkzAKeyZ2tr68oN8ixAiYwGaJpGPB4nHA47bpNcLk9LIg4vfhFTiQ5cUiLXS1RTHaYson/nXob27K2q86N0T21qamLjxo1Eo9EzP1CFQnFaMbNZBt/7Hmb+7VNggZQGUppIaa0LK4YOBHQfh1/0Qjp3bKdSqcxrgGYXelyPwZ71rO9vfxR8Ph/t7e1omlZ1mwiYnp5m+7XXMvCc51PWNTw0jkFYq0igjMWmRx5G/ukHmD5wAJ/PN8/XaGeVtLW1sW3btnUb7KRQrGXKMzPsf+97SP37Z2uNBNaXVRckUSHYf9ElJF/0oqpVt86KYbdIUPc/JTKOil0J1DTNajFx06RQKHD+a1/JyBVPx4N7XR7AWQs6Du5D/Ml7GX3ySQKBAJqmYZomLpeLjo4Oenp61m2gk0KxlimOjnHgPe+k+P/+E6RnHcoLCAqYirbAG19PSzI5L6uu2mXVv+4qey7Fepwjl43b7aa9vR2/349pmk62iT8YIvK2tzLe24d33hU2f1IVgCbXXqq4BmRNQdvhfrx/9icMPPw7vF4vkUiEvr4+Ojo6VnqICoXiNFDsH+DQ+99P6b++WmsGtr4khkTiRyI8fobe+Aa6LrmYQqHguEmkrJY4SCaT695NYqOOwjEIhUJO+pGdbTI9Pc3mSy5h6uUvpeTz4UIgpIaQR/4Hqg1+1uhDk5CzXLQNjhH/4w8w9utfsXXr1nVZ0U6hWA/kHnqYgde9icp/fRUhdZASIa0VvxedqQcSNClpRmf/1c9kwy23YFmWk00CVWt3PB5f98Ge9SiRsQxaWlqIRCKOz00Iwfj4OOc9//kceNZ1ePCiCQu7TFU91lozY9QhgCwQSE3R86E/JfX1e1d6SAqF4jSQefABBt/yNooP/MKp9rve4jAEENbc7OvuwvOKP8AfCjm9SaAqMAKBgLLkLkCJjGVgN7apd5uUy2UQgr533s3hyy/DhUTWxyBIQFRFxlq+FAUmeSzKqTTjf/hWBj/1OYxsdqWHpVAoThHj3/42g+98O6UnH8dY4/ezpZBACI3plgTpu+9mw8UXk8vl5rlJALq6utZ1TYxGqGJcyyQSiZBMJhkYGEBKiaZppNNp4u1Jcu95O3Nv2k1kboqKKRBSIKVVFRgCrDUYl7GQsgSRmSL37rspHz5M3x+/B11FVisUq5rhz32a8Q/8Ge50moquIaSsxmGsE6UhalYbHwKX0Dh8x/N4yk03MTc355Q3gGpNoHg8rtL1G6AsGcdBa2srzc3N1WyTmtCYnp6ma9s2hl/8EnQkbqxaQNT6QgqwhIaQJtl/+igjH/gQpYGhlR6WQqE4ASpzaUY/+RkmPvB+tNwchhBoyGq66jq6uUlktR6G0NlzxeV0Pu+5lEolSqXSPIERCATo7OxUGXUNUCLjOLCLq3i9XizrSLGZ2dlZdrzmley+/cW4hRspLUxtbcdjLIVAYCGY/vfP0P+GN5L99YMrPSSFQnEclA71s/8t72L8fe+DXA5TCCytGuytW7UgyHVESAgObt6M/13vJNHZydzcnCMmLMvC5XLR29uLz+db4ZGenSh3yXESCARIJBIMDw876UqGYWB5vfS85Q8ZONjPxod/y3RN7ldPRYlcJxemAKTQsIDCT77PyMwUyY//C+ErLl/poSkUimNQOHSYwbe8ldJPfgrCqvYXFSBF9R5m38bW+v1M1O7cQXRmIlFKb3kTOy64gImJiXnblctlurq6iEQiKzHMVYGyZJwAbW1txGIxp42vpmnk83kCTU3wtrcw0RIniI4uqSne9WPSkADSQkiJpbnJP/kYA29+EyP/8m+YudxKD0+hUCzB5H/9Fwde/3pm7/8ZlrCqQqL2EJbz47rwllhYeDDRXW4G7ryTrddey8zMjFPRGHA6S6uiW0dHiYwTQNd1uru7CQQCTlSxXairc8d2Rt/5djQp8LE+LsjGSBACS4KxeyfTf/ZB+t//Qcx0eqUHplAoFjD0sY8z/EfvofzgzxFSOi4Rsxa8vr6oxta5Ndh5/bM497WvplAsUi6X56Wrejweuru7lZvkGCiRcYK43W5HwdY3xSkWi2y54Xr2vP41eHWBvo5lRhWBFBqyVCT7f/+dg294E/knnlzpQSkUCsBIpTj4rvcw8ZG/gskpBAKXNBB17t71FoMhALemceCiy0nc/Q48fh+5XM4RGLYFu6Ojg6amppUd7CpAiYyToKWlhba2Nuekg6rCNS1J76teyZM33kIr/urKQAjnsd6QUmJIiVUpk7v3mxx48YuY+NKXV3pYCsW6ZvbBB9n7oheT/tRnEZk5LGkiTcASizpMr3S1zTNW1ROBVwrmYm143/1u4h0dzM7OzisRXqlUiMVitLa2rsv7+fGiRMZJoGkanZ2dhEKhedaMcrmEy+el+a7X079xIy2ARbVYhrDm9zMR8shq4ciJXvfawt+Psu1yt1upfUoBSDAOHmL63e9k4I//FFlXklehUJwZRv/vfzD0mldT/sX9YJWrJcIBITXqY8gWXttrHbcUGL4QU297GxsuupDZ2dl5r1cqFYLBIMlkEpdL5U0sByUyThJd1+nr68Ptds8TGqVSiVhPN8X3v4vhrj5CaNWQ7LorVlD9A2hy/u/O5Gz/XtvGfo265+1ttbr31m93tu2zGrUtKM+lmf74x9n/mtdT3LvvJP4CCoViuRhzcxz6o/cy+b73YBw8DLWkc22BuFh/CDQElt/HwTffxdbn3sbc3JyTQQhVK7XX66Wrq4tQKLTC4109KJFxCvD7/XR0dMwTGdKSFAtFui++mPQH3kc+GMSvW1hCYkkJVtXFYtVcLRLpuF2O+VhqW2uZ2x3Ptqd6n7X4dENqSMsg842vsv/ZNzL+6c8ha71hFArFqWfmhz9k3/XXk/nUv2FlZqvW1do1aiGRFktf22v4AdWJ0Kd5OPT853Hh619HLp/HNE3n2NnbJZNJ1QTyOFEi4xRhd94z6iZKy7IolUp0XnIx/S97KR6p45LVadbuaSKptlWzC+lZdfts9Jqs/b6wFVv97/XbLblPVnCfEgQmSBMsqEwMM/bud9P/nvdh5vMoFIpTiGUx/JnPMvCaV1F6/FFMw8CyAGliOVJjfePVLA5dsJWuu+5aJDCgWtUzHA4Ti8VUHMZxokTGKUIIQU9PDy0tLfOqgdqN1Da+8pX87pWvQdd8eNeZn/NYSCGwKkXSn/oEO59zK+P3fF3FaigUJ4m0LLK//R17nncnU+98B2J6BokGUqdah1EHFgd5rjc8uBm89Eoif/UR/OEmisXivNdN06SpqYnu7m48Hs8KjXL1okTGKUTXdbq6unC5XI7QEEJQKpVwe72c+8Y3su8FtxHSQJfrq0jX0ane6EypYf7yQQZf+zKG/89HMDOqm6tCcSKY+TzjX/wSB17wAvI//A7CMObVIMb5ef3eh6SQ+LCYTCZp+vM/I9reTqFQmGepsCwLj8dDV1cXwWBwBUe7elEi4xTj8/no7u5GCDFPaBQKBTx+Hy0vfSn7LrgAP9a6T21dSLUkuYVuwsTf/Q37nnc76V/9BqtQPOZ7FQoFyEqF4r797Lv7XQy9+Q0wNVYNaBQ6FqIWh2A7Ma1j7W4NI/EIQSYaJ/OOPyTa0U4+n593H7bjMOLxuCobfhIokXGKEULQ2tpKd3f3vMAigHQ6TeeWLXj/z//h4GWX0SSqOZ3Kc1JFIqs9Emorq8KvfsbBG29k4MN/TWlkdIVHp1Cc3VQmJhn7zBfYe9315L/4eSx0DE134r+O3GjWt8AQErxI8k1NTP7ZB9h6y83kstlF9hwpJfF4nPb29hUZ51pBiYzTRCKRoLW1lcqC2IJUKsXG7dvx/NF7mOjbQERKpGVhWuv3ol8aDSpF5v7u79h76/OZ+d/vYhVLKz0oheLsQkoKT+zkwF1vY/zud2CMjyKkhi5NpGVh1R6KqoxwCxO3HmL8rW9n2403ksvnQIh5iz3TNAmHw3R1dc0rxKU4ftTRO00IIZyys8aC1MzJyUk2XnQRU697DamWOAG0deoVPTrV9ZaOIXTKT/6egy+9k8F/+CdyO3chF0R/KxTrkVL/YUY/9x8cuP3FZL7/dUxhYsc4VWvTKDupjUTgEQL8Pp58wW303XYbmUwGy5wvwOrrYaiCWyePOoKnEZ/Px8aNG9m/fz/FYhFd1xFCYJoms7OzXHznnTzq9RP7m78lMj1GUd0PGiCB6o1TK1SY+vMPMf35/0vynW+h9fY78ChTpmIdYmZzzPzq16T+4m8oPPArTM0ATSKxQDZeO8p1LDikEPgw0T1+nnjJS9jx1rdiSkm5WFwU6Olyuejp6VGBnqcIZck4zQQCAXp6ehBCOBYN++fZ2VkuveMFjLzhdZR8Afzqr7EEtcZEVEuTy8FDDL/73fS/9e2k7rtPFfFSrCvyu3Zz+EMfZuCOl5B/8BcYuoGlSbRaPV1V9WI+EvBIQUS42Pesa9n25jeDplEsFRHa/EBPIQSdnZ2q4NYpRFkyzgDhcJjOzk6GhoYwTdOxaJRLJVKpFDvufBG75tL0feGz+DIFSo6ytmpFf+t6CVA1hh4poLO+HC32t9UtyH/rfzj04K+JvOIVxG+4kfAzn7mSQ1MoTiv5XTuZ+cEPSH31vzAffAQXYGhV4S2QaFZNYgglM6Aa4CkBtyYJIHjsaU+n9W1vQ3e7j2SS1B0oy7JIJpNOd23FqUGJjDOApmm0t7djmiYjIyOOYhaaRj6fJxgMcvHb3sIjLsGmf/kUHrNERdYaFmFVW6UDGrJWxKv6WvX60Gufsr4Cu2RNaomxcTIf/SjlL/0H6T98Ky0vfjGBvr6VHp5Cccow83kmv/Mtpv/541Qe+DVSWEjNjbA0NGlhSdDqjNJCSsS6lxkCKQUuAU1CY88zr6Hjzz9EuDV+pOlZ3SEyDIPm5mY6OztXZLRrGSUyzhB2IGi5XGZychKXy4UQAk3TyOVyAFz0xjfx6OgE53/1K1gCDECT1datWv3y5IjCAGEeeXId3Vecm2gtDdgaGWHyzz5A6r77aHvJS4g+51Y8ifiKjlGhOFnmfvAjJr76n2S//jVkIQ+aQEp3dbEhLKc5oWI+EtCFJKrBnvN30PSudxFNJJiZSS3a1s4k6e3tRdf1xTtTnBRKZJxB7Iqg5XKZ2dlZp0StLTQ0TWPT3W9np8fL5i9+CYsyFSEcC0Z1OrXLAIvaRFtvwVhfrhMbiaAiNIQUWPf9hOEHHmDqG/9N863PIf4HL8EVDq/0EBWK42Lse98j8/VvYvzwxxhD/eiaoKJRs2hW7wmKxkhAExCRsO/yy/G9/wPEe3pJpVJILITQnAWZYRj4/X42bNiA1+td0XGvVZTIOMN4vV56e3sxDINcLofb7QaqQiOTyRAOhznn7rezv1Jm+z1fYk6YGLiqgqLWLRFYWF2n9tz6vfFYTtAbkM9T/P53Gf7Zfcx8//u03XUXrTfftJLDUyiWRebxJ5j49y8y+82v4BocpGqz07GkAGkBFlIataVF7YyX63NxsQhRvQe4gCgm+3acT/BDf07H5s1MT09jWbW27bX7pB0f19PTg8/nW9Ghr2WUyFgBAoEAvb29HDx4kFKp5ORia5pGOp0mFArR/fa38mQux7Zv/zc5LMpCQ5PzBUb1P3WDqVJXObX2g1YoUPrW/zDw+4eZeMbTid/xImLX34jmVU2OFGcXsw88yMy93yDzve9hPrYTXVpYTkVgC4kOFgghQdasmcJiPfceWUjVdSQJC0H/th3ID/wJ7eecM19g1G4OlmXhdrvp7u4mGo2u5LDXPEpkrBDhcJienh4OHTqEYRjzhEYul8Pv99P1x+9ld9DP1nvuwagV2VEsH7s8uRgcovAfX2bgvh8zdt4FtL7mVbTf+ZIVHp1CAXOPPsrkxz5B/pc/o3xwP0JaSAGWcIHQEdJAYNUufVlbVNhxA+vXclmPFNVMEg0ICTh06aV4/uzP6d20kVQqhWVZ8wSGlBLTNNmwYQOtra0rOvb1gBIZK0hzczOGYXD48GEsy3LK1wohKBaL+AIBeu6+m50Vk0u/fg+TmBhCR1P3luNCSqp3odFRimOjjDz6O1Jf+xbtr3olzc+5caWHp1iHZHbuZuxznyb7v99CHDiAkDoaOlLYFgqcOCyg5iqBeQsNdR8A7DgVi2YEA1vOxfOhP6djyxZmZ2cxTXNRsS3DMOjq6qKlpWXlBr2OUCJjBRFCEI/HkVIyMDAwT2gAFAoFfD4fHW97C48CW77+TXKyhIlWLRe8jmMwjpeq0KgF0Y5PUvz6Vxm4/2eMb99G+LbnEH/py3DHVTaK4vQy881vMP3f/0vh/vsxRvrBqPY2ktKqBnULiahllGEdSSebb8NcmK6+vtLXbezifLqUhDSNg5c/BdcHPkB8wwZSqVksy1zUVVVKSU9PD8lkUvUkOUMokbHCCCFIJBKUy2VGR0edGhr2a6VSCV9TE13veTd7w2G2ff4zZLEwVCzGcWFbm4+sBC3MiVFykyOUHvgNqU9/hsjLX07y9XfhiisTquIUIiUT3/kuqY/9I/nfPYIxl8Fj1iq9COFIBCmqDj7NOvL7Ejs8A4M++7Gv5KDUGHzaM4j85Z/TFE+QzeaQ0lokMMrlMp2dnSSTSZWqegZRUu4swC5l29bWtqg9PEC5XEbzuOm863U8fucfEBBu3FTrYyipcbIIzHye8p5dTH34r3jy8qdy8AN/Qmbnkys9MMUqx5idZebLX2Hn05/F0CteTuYn92HOzqHJapCyLXwtAZZWC1yU1dIv6rpeGjvWCiHx42LgkqcQ+OAHiLQlyWSytfvnfIFhmqbTtl0JjDOLsmScJWia5vQ4GR8fB5inxMvlMm6Ph953380T7Qk2fPxjuExl0ThZnMWiFFjFEubQIVL/9FHm/u2f8Z2zjegb3kzizjvQI6rWhmJ5zP32EcY//jGyP/khIpWCQgkEVTcItXo3dZetoBZ/IY+cj5a6rJdECoEmNdwunT233kbPe96NNxQknU5Xa/PVWXpsgdHW1kZnZ6dTMkBx5lAi4yxCCOEU65qZmVmkuCuVCm63m3Nf+zp2eXxs/sd/xFMuUFKt4k+aWmIbugRKFrJkUHjkYcpveTOpv/4IoT+4k/hrX4+vtxs0DaH8uYoa0jSxikVmvv9dxj/1ecq//C0U0liiXBUQNVVRf406zf44khmxpHdEMQ+XNPDqbg5cfxNbPvSnWEBxQTdVm0qlQnNzMx0dHUpgrBBKZJxlaJrGhg0b0HWd6enpRRdOuVxGCI1zXnwne02Tnk//O8G5NAXKysZ6GjDNMsbhA5T++m9Jfepz+G++jtCzbyB69dX4OtvRfL7aLKJYT8hyGTOdJv/YE0zd9xPS3/4fzCcehVrgppAaGhpCyFp89pFYoPraFs5PdRYMSxyxbijqkbgA0x9m93Nvpfttb8VCUK6UGgoM0zRpaWmht7fXqa6sOPMokXEWous6GzZswDAMUqkUmqbNCwYtl4q43G4ueuMb2N23iZYP/39Exw9SAKRyn5xiRM28bWGmJsh8+Uvk//PLTERaabrpWcRe/gqCV16JHgyiqbLEaxvLwshkKY6OMPu1e0j9x5cx9u0HqVXLfGt6LUvEQggLXepQq3vhiAxhR3Vq2NEFjVLS50cVKAQSN2D6gxx+17vY8YqXks/nKZVKizS+7SKJxWL09fU5NYgUK4M6+mcxGzZswLIs0uk0UBejIQSVSoV0Os15N1/Pk2YR8Td/TWR8ihISU67PlLbTT/X4WwjE3CTpe75K+p57cO+4gPCdd9LynFvwdLSjB4LoodAKj1VxKrDKZaxMBiubJf3L3zD1la+R+/H30YuZatxErc26JY8UexKy6r60RP11aFfrtd1slmOpsItJ1btPlMA4gkDiRWBGYxx6xUs57+V/UAvwtBoKDMuyCIfDdHZ2KoFxFqD+AmcxLpeLTZs2MTg4yOTk5CKLhmkYzM7OsuXGm+hPtDH7D//MxocfJC+rHVzVjerUI6Rd6tk+ugLz8ceYfexRxv7yz/BcdhmRS64gcu21BC68AHdrHD3ctJJDVhwn0jCoTE6RHxgic/8vyN/3Y8qP/57KyEg1tkJKLK3WA0PYQZzyhMWBHZ8B6pqtp2rNEQTRmeztZfbud7DtxuvJ5fPIBgspW2DYLhIVg3F2oETGWY7L5aKnp4dKpUIqlULX9XkWDcMwqhaNpz2NobY2Dn7wg2z69a/JW5Jy3VSoOH0YAEKgl03kLx8g9csHmP6Xf8a9YSPhW55L6FnPJHjeuXjakriikZUerqIB0jAoT0xQ7O8n89BvmfvO9yj89H4wiuhWdUKTQiKpCX17kpNwJM5CcaqQVCenoJSkujZQ+Ku/ZMdTr2JmZqZaJpz51UKklBiGQUtLi3KRnGWov8QqQNd1Nm7cSH9/f7VdsZTzSpADjI6O0tbbS/mDH2Tnxz5O389+jqeQpYy6/Z05xLzUQ6P/ILOf+BipT3wMbdMmws++ntCVV+DbvBl3czOutjjumCr8tRLISoXSyBjm9BSliUnSjz9O7qf3UfzxzxClPJYTKKHVNSOzryUVkXk6kYBXarh1jf4LdiDe8x56Lr2E6ZlpZK0K6kKBYVswuru7lcA4y1B/jVWCy+Vi48aNDAwMMDo6isvlmlcWVwjB5OQk7Zs2kfynf+CxT/wr53zmM/iLeQqLyhIrTjdVC3pdNccD+5k9cIDpz34Gy+0h0NeH97KLCF/5NALbz8PT04O3qwuhouBPG5XpFKXBAYr9h0g/8ntyv3qQ8pOPY46P4JKWkyUkazUtquZ6OT9w05ne1BV1qrGPrFcIAkKw8/nPo+M976KlLeksrha9pyYwEokE3d3dqtDWWYgSGasITdPo7u7GsiwmJyed5+pfT6VSRCIRLnzrW3jc7aL7c58jNDdHfl4cgeJMYyFqqY0SvZTH2LMTa/duMl+6B5qC+C65hOAlV+DfvoPA1g24urvxdXQg1KrshKnMpigPDVPpP8zs7v2UH3uC3EO/orJnN24LatEQCNHY4aGuljOHnU0TAAyfn8df8Hx63v42mppbmJmZafgey7KwLItYLKYExlmMuoOtMnRdp7e3F13XGR8fxzTNeReXpmnMzc0RDAY5/01v4on2dpo//Vk69u0lh4XRwKohUF7lM4EUVs3qLmr1EGTVt59OU/rpTzF++gtmAFdPJ2L7uQQ2b8O3dQvuWDOelhiezi48nZ24VPXRRVRGxigMD2GODFNKz2EMDpHf+STl3Xswd+3GyBfRsJB2rYW696rzfuWo3ns0NGkRFZBKtHP41a9k2yteju52Mzc317AGht3jqa2tjY6ODiUwzmKUyFiF6LpOT08Puq4zPDzcUGjkcjkMw+DSO+9kT08P/R/9J7ofeZCiaWDahYBkLf1upb7IukU2cOsLTExAYgwcRhs4TPq732cWqtVFE214e/tw9W3Ee85GAn296N2dBDq6cHV04I6tj7bV0jSpjI1TGhmlPNBPeXiQzL4D6AeHyB4+gHG4H7IZNKodd6u2iiNC2mbetOWUllexFmcKe6ljAS4haBJuRjZvZO7tb+PiW26mUCjUyoQvFhiWZSGlpKOjg87OzobbKM4elMhYpdhN1WyhYRjGvIAnTdMol8tMTk6y8dJLGf2rv2D3Jz7Bud+6F2lCzs7xl0eqDqpb7MohFxx9e6UtACwLxkYpjY1SeuA3lJBMC4HV0U5TWwd0dEF3B5GOTvTODrTuHoLxOHpHEr2pCd3nO9Nf56SQloVZKGBNzVCamCI/MgJD/ZRGR5kbGsI9NIoxOkF5uB8xN4vOkeNlCwqJxGSxuFCcHdhnexAIYLHriivwvu+P2H7RhaTTaUqlUsNW7JZV7a7a3t5OR0eHEhirACUyVjH2xabrOoODg4uEhhACKSUzMzO09fUReP8H2NXbQ99nv0A4n2FOHomYrxcaSmycPcgFP2vIasqslLiGR8gNj8IjD6EBWTQIBtESSYLhKGZrC0aTH4/fjz8QwAw1YcVjhJpb8ITDaPE4RKPo4TDuaDNuvw/N70ec6sqlUmIVCpiFIpVshkp6DjOdQZuYwkxNU0pnyE5N4U5No2ez5EpFzHwZ18wc5dk5StMTiKlxNMNAB0rV+prYYbUmVSFh11WgZp070tGnvn6mkh1nB5KQEJhuD4+95CXEXvVK2jdsIJVKYZpmQ4FhmqaT0h+LxZTAWCUokbEGSCQSuFwuBgYGKBaLi4rQCCGYnZ0l2BRi611vYP/GTcT/4Z9JHj7ItF2lcCUGrjguFlqbatUbnH8FEpHLIA5lyKMja46xLIIMVYEihWDO58Pl8yGampChENLnR/cH8Lpd6B4PWjCAGQhQCoUR3mY0jxs94AGp4fX4QK99unRVP1lILMOkXLHALGKVypgVA1cuhSeXxsplsQoFDMOgVCxhFfOIQhF9LoPMZzEKRYRRQSxwV5Rr39jFERFhf2drwXlrywiBnGfNOMKRZxUriwZE0UiHm+h/z7vZevsL0d0uZmZmnFiLhVQqFTweD729vcRisTM/aMUJo0TGGqGlpQWXy8Xhw4fJ5XK4XK55F6sQglwuh8fjYeuNN3KopYXs3/8jGx57jKI0qEi1vlvN2FLjiCXqSGijqOUVVZtuSSgUMAoFSKWw14sGUK7bXzVrU0fKWkqtsKolsV3uWucuiyO3DxNpSTQTpDCd+VzIMtm6s6reYgZQqf8sGp1/i2siWAueWZhQulBCyKP8pjizSKr6NCRcDGw7l9m7XseOW2+lVC6TrQV4LiUwgsEg3d3dRKPRMz1sxUmiRMYaIhwOs2nTJgYGBpidnW0oNCqVCqZp0nPJJaT+/u944mtfZ+unP4PbMslIS1k01hgL1/ILWUpYVntomAgK899sFBpu7zgjnCDKxvtsNAY19a99JOAFfJrO7jvvpPWNd7E5kSBXKFApNe6iCtWu05FIhL6+PgKBwBkds+LUsNjxpVjVBAIBNm7cSGtrK4ZhNCxgY1kWxWKRlvZ2Nr3xDex5/59Q9jXRKtQJoTiCPAUPxfpGIrAQNAG65mb/m95E73vfQ0t7O+VymXKpDKJxgGe5XKa1tZUNGzYogbGKUZaMNYjtu/R4PIyPj88rQ24jpSSfz+N2uznnztsZ3rKR8mc+y5Zf/oqSZVBCCQ6FQnFyaEgiQuPwjgsp/OGb6H3qVaBpZLNZwC6yuiCzyjSdFNWOjg7V6GyVo0TGGsXtdtPV1YXf72dwcHDJiO1KpYKmafRefjm5LZt5/N8+yfYvfgm3ZZJTNUIVCsUJ4gZcus7O5z6f+NvfSjKRoFQqYVUqS77Hvk91dXWRSCQa3rMUqwslMtYwmqbR2tqKEIKBgYElhYZlWeRyObzBENve/W4ObN9B/B8+RsvYMFlVrEuhUCwTe1HiB0qBEHve+Q42vvAONN1FsVhs6L61qVQqeL1e+vr6iEajKkV1jaBExhpHCEEsFiMQCDA4OMjc3NySJXhLpRK6rnPObbcytWM7+/71X+n53vcxjQpllPtEoVAcDRc6Fm4NDj/jWrS7XseW88+nXKlglMtHfadpmkSjUbq7uwkGg2dovIozgRIZ6wAhBIFAgM2bN3PgwAFSqdSSQsM0TbLZLLG+Plr/9m/Zc9nldHzyUzSNj1KU0qlXoFAoFDYCgR+TSsDP3le/mk133YXu8VAoNM5GspFSYpomiUTCaZWgWFsokbGO0DSNTZs2MTw8zOTk5JLuE7umhsvl4vxXvJxdXV1MffpzdD72e7RiHgMlNBQKRdU9oguB5nYzvmkjubvuYssNN1ExDMrLEBiappFMJp3KxYq1hxIZ6wy7XXw4HObQoUNUKpWGRXCEEJimyczMDJuvuRrt6dfw+Bf+Hx2f+3dCYyOUwam6qFAo1hPV8m4uJF40yuEoB1/yQja+8Y10h8Ok0xksy1zy3VJKLMsiEAjQ29tLU1PTGRy74kyjRMY6JRKJsHnzZgYGBshmsw3TXKEqNubm5vB4PFz0utey79ytTP7zJ+h+/DH0UgFDLmztpVAo1jICiQeB2+1i9NxtFN/8JrZffz2lUplUKnXUgE078DMcDtPT06PqX6wDlMhYxwSDQTZv3szw8DBTU1OLWsbbaJqGYRhVq8bTnkZ62zYO/udXaP3GfxPtP4iJxdLrFoVCsVZwAV4hSLcnOHDjzXS8+tVs7O0llZqlUikfNeXUNE3cbjdtbW1OvyXF2kf9ldc5LpeL3t5egsEgw8PDFAoFdF1f8mYxNTVFIBDgsrvfzqFnPJ1D//pv9P3kPnzlMkUkRq0PporaUCjWDhoQACyXzsBVV6H/4Vu47OqryeVyTE5OIoRA17TGpestCyklfr+fjo4O1eBsnaFEhgKA1tZWfD4fg4ODZLPZo1o1ioUChUKBrh3baf3o37L305+h5fvfJ3rgEAHTpIiJhersqlCsBbwCdDSmztnE+DXX0HvX60l0djExMTEveLyRwDBNEyEE0WiUrq4u5R5ZhyiRoXAIhUJs3ryZiYkJxsfHKZfL6Lq+2McqqtaKmZkZvF4vF9z9DoZuvJF9995L39e/RnQmRQ5JBeG0IFcoFKsHu2NqQOhkw00MPOtZtLz2NVx03nlUKhVGR0ePavEEMAwDr9dLMpmktbVVuUfWKeqvrpiHy+Wio6MDv9/P0NAQuVxuyZuJEIJSqYRhGHRtO5e2czax+7xtTH75K2x46AF8UpCtBYYKlANFoTjbkVRdI02AEBqHL76U3ItfxLm33kwgGCKVSlGpVI6abmqaJpZlEQqF6O7uJhKJnKnhK85ClMhQNKS5udlpsDYzM4NhGA1XIkIILMtiZmYGt9vNxbffzsjFF/PYV75M4qc/oW3PAaS0yHNEbCgUirMTPwKvEExu2sj4ZZcTe/Wr2HreeaTTaSYmJtC0Y1svXC4Xra2tJJNJ5R5RKJGhWJpgMEhfXx9NTU2Mjo6Sz+dxuVwNU9SEEBiGweTkJM3JJNF33M3YLc9h17e/S9+X76ElM0cWixLKfaJQnC3YV6ILCAlB0eNl1wteQOjOF7H5nM14fF4mJibqUtwbX7uWZWEYBsFgkM7OTpqbm1VzMwWgRIbiGGiaRjweJxAIMDo6yvT0NMBR/auZTKbabn7HDto2beLABecz/vVv0PvTnxKQJhnAQFk1FIqVRFLtlBoEpK5x+LKnkL/jdnqvu45QJEIun2N2dhZN045a+8Iu6JdIJEgmk6r3iGIeSmQolkUwGHSq801OTpLNZo9q1ahUKkxPT+P1etl20w1MX3gBu+//OcHvfpeeX/wSDYucFLUOr8qyoVCcCTTAFCBlVVx4NZ3Biy5k+vrrSF53Hb19GygWi8zMzFS313SWY71IJpO0tLSo4E7FItQZoVg2diGdcDjM6Ogok5OTwNJWDTswtFKpEG6N0fzCFzJ7zdPZ+7OfEf385+k5dIiStMiBSnlVKE4zEoEJeKUkqEEqEmPPq19J4pZbOK+zEyEEmUzGSTs9GoZhAJBMJkkmk/j9/jPwDRSrESUyFMeN3++np6eHYDDI2NgYxWKxcaorVaEhpSSbzaLrOk2tMSIveiGjF17II9/8Ju3f/h+Sk9PkJRRrK6ZqJoqybigUpwI7s8uDRghB1utl93Ofg/v5z2PLBRegu1yUSmUqlXKDPkbzr0MpJYZh4Pf7ndRU1dhMcTSUyFCcEC6Xi0QiQSQSYXx83CnMczSrhmVZjpslsWkjzW97K9PPfz6//8EP6brnKzRPT5GXFhUpVCaKQnGSSI7EXYQ0yHq97LzjhbhvuYnOLVvwBAKUK2WK+RJIjmm9sFNXu7q6aG1txev1HvM9CoUSGYoTRgiBz+ejq6uLSCTC8PAwmUxmyVgN+z2maZLP56vul00bid/1OmZuvJ7hr95D35e/TIspyEmLIpYSGgrFcVPN6vBiEUKjIjT2Pfs63K99FX3nbUN3ezAMg0KhiJTWMfdmWRamaRIOh+no6CAcDqvMEcWyUSJDcdLouk40GiUQCDA2NsbMzAylUmlJF4pNpVJxVkctfX1E3vkuhp51HYe+/b90/fSnRFLTFKWJIQX15byUK0WhWBqBxC90ym4Pu5/xTKznP4/OKy7H4/dTqVQoFwpON9SjIaXENE28Xi8dHR20tbWpwE7FcaPOGMUpw+Px0NXVRSKRcDq7app2zFWPaZpOr5SeSy9BXHIxU696OSP//d8kv3kvTbOzGLKa9qoyURSKxVTLgAv86BR8fvbeeguhO++ga8sWhK5jGAbFYnFZ4gKqAkNKSVtbG+3t7Xg8HuUaUZwQSmQoTimapuHz+diwYQORSISRkZGjFvGqxy5HLIQg1tdH4l3vYuy22xj7v/9O8r4fEywUwIAySmooFNW4JYGOxCs0ym4vA1c/FevlL2frlVdgmCaVSgVpmsclLizLIhAI0N7eTktLixIXipNCiQzFaUHTNFpbW4lGo0xMTDA1NUWpVDqmVcNeQZXL1Uj3+KZNJP/qI4zu38/wj35M63f/l9bBw5hlA4MjwW0KxXpDB7y6m9l4koGbbsB3/bPovOB8EIJ8oQDy+ByLlmU5Dc1isZjKGlGcEpTIUJxW7IZr8Xic4eFhpqennfbQx1oh2WIDoG3TJnq3b2fuZS9l33f+l/A9X6F1cBCtWMIATBWpoVjjCKohnTqgeTykWmIcfsHzSL7oTrZ1dVEulykWi8e1T1vUa5pGW1sbyWQSj8dzOoavWKcokaE4I7jdbvr6+mhubmZsbIxsNuu4RpZjji2VShSLRTyBAOe/5tWkbrmZQ9/9Lp5vf4fE4QFcmTlMw6jeNM/A91EoTjd2GrdWe6DpWKEg44kEmZtupu22W9m+YSPFUoG5ubnjcmvY7hNN0/D7/XR0dKhuqYrTghIZijNKJBIhEokwOzvL+Pg4mUxm2WLDLlc+MzODJxji/Ne8huwLX8jEzp2kv/c9Wn72E1rHxqFYwkBDYFE5Q99LoThViNo/GuCSOpruohyJMvCUS3E//3kkLryADfE4xWKRublZNCHQxPLaDtqWCyEETU1NtLe3Ew6HT+8XUqxrlMhQrAjRaJRIJML09LRTNfT4xEaZ6elpXC4XvZddhuepT+Xwzp3s+9a3CP3kPloHR/FmM7iwqCAwlH1DsQrQAQ8CJFheD5m2BONXXEXwtts478orcLlcZLNZUqkUUKuou4z92uICIBAIEI/HicfjKqhTcdpRIkOxYgghaG1tJRwOMzU1RTqdJpvNLjtmwy7sNTs7C0Dbxo1seP/7GXvpS9n/v99HPPY74nv20Hx4iGC5QBlBpRa7oW6tirMB+1x0UxUXlhCkk51Mbt5A8cILaH32dVx80UVYlsXc3ByGYSzbxQhHskV0XScYDBIOh1W9C8UZRZ1pihXH4/E4xX6mpqaYmpoin887lo1jZaTYN9x0Ok06nSbQ3Mzlb3gdpmkyvGcP+7/3PYK//CXRg4cIzs6gy2oarHKlKFYSd+2hIyj4fYxs3MDsJZcRuv56tjzlcoLBIIVCgenp6eOKX4Jqpogd0BkIBGhtbSUej6uMEcUZR4kMxVmDruu0tbXR3NzM5OSkUznU7gq5XLGRz+fJ5/Pouk775s30bt/O9MtexsBPfwo//xnNe/cRHh4mVCiClBSRTjqssnAoTgf2uaUDfgChUXa7SLUlmN24mdJVV5J49rO4fOtWTNMknU6TyWQATlhc+Hw+YrEY8Xgct9t9mr6ZQnF0lMhQnHV4PB46Oztpbm4mm80yPT1NNpvFMIxjliqHI2LDsixmZ2cRQhBoauKiF7+Y3C3PYXj3bg7t2onr8cdo+f3jxPbvJ4BJRUKJatt528+tRIfiRLHPoWqcRdUdYgrBbLKNiR0XYlx+KZ4tm+k8fwctrXFKpRKTk5NYVrWfyPFmi9huxkgkQktLC6FQSLVgV6w4Jy0yDqUOn4pxKBSLCAQCBAIBotEos7OzTE9PO9kouq4vq0mTfaPO5XLkcjm8Xi+bLrkY9xVPIZfNMrJrFzsfeADPww/T/thjhKem0GTVsmG7VOwWUkpwKI6FLSxcVEWFS4BLCHI+D/3bLyBz8SX4Lr+M3ksvIRqLYVkWuVyOyalJpCWPy2oBVSFtWdY8cdHc3KwsF4qzhpMWGdlK7lSMQ6FYEo/HQyKRIBqNMjMzw9zcHLlcjnK5jKZpy/Iz2zfucrlMqVQCwOv10nfRRWiXXkJqfJLRR3/Pod/+lsCTTxLrHyA0NkwEsCwoIh3BocSGoh7bFVIVFuARAhPBXDTCXFcX2e3b4eKLSVx6KX09PbhcbvL5HBMTE84+jldcmLVS4R6PB7/fT3NzMy0tLUpcKFaKUiIc39PohWOKjE3RvvsPzPZfd+rHpFAcHx6Ph2QySUtLC/l8nrm5OVKpFMViESHEslwpcERwlEolp9S5vynEtuuuo/z0pzM9OsrsyChjhw9hPfw7ko8+SuzAAQJITCkpAEatK6yK41h/1KeM6oAP8AowNBfZWDPD284jc/lTcG/dQqijg3O6ugiEQhiGQS6XwzCqrf6ON33UzhQxTRO/3080GqW5uRmfz4fX6z1l30+hOBG8Lq/R6PljioyAJzBz6oejUJw4Ho8Hj8dDJBIhFouRSqVIpVLk83mklLhcruNypUgpHXeK2+0m1t5OsqcHecVTSF9/I6mhQXbv2oX2298SeexxEocO4bcsKpqgbMlFmSpKdKwtZN3/QlQDN/2yev4YaMy2tjB4/g6yF11M4KKLifX10RVvxePxYJomxWKRVCrlFME6EXFh1KrZ1lstgsHgss5zheJkyZfzJ/zeUxL4OTgzRHdL16nYlUKxbIQQhEIhgsEgsViMubk5MpkM2Wz2uFwp9r4ADMNwovo1TcPr99G9fTvyvPPIPvs65sbG2b3zSSq7nsA9Pk58115iA4NEhKQiJUWqLekNqmJDCY7VSa3oNgILD+AT4JYaRaGTDUcY23Yu6e5OzL4+mi68gObubrpjMdweD4ZhUiwWyOayjkI5EXFhWy1cLpdTvK6pqYlAIKDEheKMcmC2/6ivh9zBiaVeO6bI6Ip0PvL4xM47j7bNXGGObpTIUKwMQggnSLS1tZViscjMzAypVMqJv1iuK8XeH1RXkIVCgUKhgKZpeAJ+OrZuJrl5E+Ubb6BcKJCZnGTPgQPwxBOEfvsQ0d27CZkmLid4tGrpsOr3f2q/vuIUUJ9NpAEeXOhCw6WZFC3JdG8f0xddTOWiCwls3Uy0o4NYMIjb68XldmNZFuVymXyhcMRigTihP7ZpmpimicfjIR6P09LSgs/nU43LFCvCZGbymNtsaO799VKvHVNktASaDx1rm5G5MXZ0bj/mQBSK043b7cbtdhMIBEgkEqRSKafNfH29jRMRHMVikWKxiK7ruFwuApEIwWgUNm+m8qxryaVmmRgc5PCuXbieeJzIk7toOdxPVIJAYkg5L2PFqsV1KNFx5rFFhUatKJaoZoNoQBFBJtLC5PYtFHdsx3PBBTRt2EhHawu+YBAhNCdltGIYFEslR1jA8cdawBGrhRACv99Pa2srkUgEn8+nCmgpVpTx9JJGCodYILZ/qdeOKTLawoknqZYPWDKyaO/kPm7g2ccciEJxptB1Hb/fj9frJRaLUSwWSafTTuzG8QSK2tTX3yiXy87EYrtlIokEkUQCefHFVIpF0tMzHBoewpiegqFhPE88QWTfXmLDo3gQVbeKNKkAJvbEJ2qiww4pFcy3gyhOhPqOpi7AIzQ0QAqNVCBAbuMmsju2Y245By0WI9DeQWuyDV8wiO52OyXsS6Uy0rKw6kQFnJiwAObFWrS0tBAOh/F6vXi9XtVXRHFWcGCq/5jbtASiS250TJHR3dL15DEHMdtPupAm7Ffd/BRnF5qmOTftpqYm2trayGQyTE1NkclkMAwDTdNOyMddLzosy6JSqYZ/6rqO7vXS0tlJS2dHtYmVZVEpFimm0xzq76d48BDuJ58k9OQTRAYGaDJMQGBJgYlEYmFSje3QqE6SqsXb8qiKCYFuR1aIaulugaCiCeZiMbLnnENx+w7Els0ENm0inEgQCwSqrg+AWkMxKSWlmqVi3mecoACw92nXeolEIsTjcZqampYdsKxQnEl2TzbMTK2n1Bfr/dVSLy4r8HM5aax7x/dzWd8ly9mdQrEiaJqGx+OhpaWFaDRKqVRiZmaG2dlZyuUylUrlhAVHPbZPHY5MRpqmoXu9BONxmtraEFdcgWkYFDMZZiYmGDp4CLFnL8aBfXiHh4lk8njSKbzlMi5TB8yai+XIY/2KjqqbCaoCrP6ho6ELQU6DmWgEQmHSLS1UNm1E23ouvi2bCXd0kIhEcPt8AJi1CptSSsqVyryOpacKW4i63W68Xi/hcNiJtTge951CcSaZzEwylju2u+RoxohliYwt8c0/OpbI+M3hB5XIUKwKbFdJIBDA7/fT3t5OuVxmdnaW2dlZ8vn8svulHAt7srJFh/359v/ecBh/tBn93HPhlpuplMsUCwUqhQK5XJbpwUFyAwNw4ADh/gE8I+M0pWbwGyWkVRUeUtYsKkjM2vxrCoF2iifKM41w/hdO3w8dDYHEQqAJqAhIh8OUEgkqPb3kN27CvaEPb3cXvkQCl9dDhz+Ax+tF13VMy8KSNcuTaSCtUy8obGyLhd2krLm5mWg0is/nO6FsE4XiTPOrgw8ec5vzE+fde7TXlyUyLu6+4D+/u++Hf8FR4jIOzParVFbFqsMWHH6/H5/PRzKZJJ/PO5VFi8WiMwmdKlO2vb96AWLUTTjeQABfMIiWSKBtOqf6uVJSKZXIpTOkpiYZHx8nO3AYbWiEwOgY+vgYnulZfNkC3rKJ17IQlBEYCAlSWs76X0rZ0BJiu2SE87tosNXxYMeU2NEmi4NcbfHgBNiKqqAwRe1nTaOiuSj4fZjRCGYsQaGjg3xXO76eHnzJDvzJturkHQggdL3qbDItp1mYlJKKYVAxDOfrSGH/cBJfrwH1fUc8Ho9jsQiFQkpYKFYVJaPEQ8OPHHO7C9p3fP1ory9LZMSb4oPJYOLJsdzEUU0Vvzz4G17S8sLl7FKhOOuwJ4BgMEgwGKSzs5NsNksmk6FYLJLNZqnUzOknk0nQiPrVdCPrh/1ZgWiEplgL+nnnOVkHlmlSyOcp5fIU5+aYmZigMD1JeWoKdyaPyOcgl8UqFvHkCwTTGbR0Gr2QxV0qYRVLYFl4EHhlVQgICUJWpYi0qhO+BmBZi4SCRFTnak1U3wcITWIClnAhBEhRLc1uAprbg+nzUg6EoClEORIm4w8gAj40fwDL76MSCiKizQRicQLxON7WGP6mEIFgCFctELOa4WFhGAamaVAsl1m2LjoF4mKhWHS73YTDYXw+H01NTTQ1NanMEMWq5cFDDy2nbUhpS9s5PzraBssuxnVxx4Vf/e6+Hx5VZDww/BBP23ilsmYo1gSaphEOhwmHqwHNlUqFdDrN3NycIzgadcw8XatVO22ykfgQuu4IENeWzWi6hiY05331wsUsV6iUyxilIqV8nlKxQKlQhEIRVz6PLBapFItU8nmkZSKKRTTDREgwC/lqM5cjn45EoHk9aG5X1bHh82G5XQivG83rw+Wvigcj4Mfl9+MLBHB7/UT8ftweN7rPi6675gm3qogAyzKdoFrLsihXKpTK5aMcpFN6yI/sdoFLxRaabrebYDBIJBJxMkMUitVOySjxg333HXO78xPn3Rv2h49aFVws1x+ZLqRbPvSDj4xwFJdJ7UN57VWvXNY+FYrVSqVSIZfLMTs7O09wWHUT8EqZx5e6puuDUO14k/pHfQ2Ro439WPu3rQzOA4m0pOO+qDb3spt8WfNEkLT9NXWfsdLHsL4cuKZpuFwuQqEQTU1NRCIR1ZRMsea4f98v+cbObx1rs9LrLnvF83d0bv/e0TZatiUj7A/PXNF52b8/MPzQm4623eMTO3li+ElVnEuxpnG73USjUaLRKJZlOYW68vk8uVyOYrHo1ECwJ6xT7WJZimPt3xZCtkXkdAU+LsXRxue8dIaFxULXB+DUP/H7/U5FWZ/PpwpkKdY0k5nJ5QgMksHEk8cSGHCcvUuetvHKf3lg+KHXcAxrxj2PfoOelm5VN0OxLrCzBwKBAC0tLQAUCgWn6Vo+n6dcLjuiw85cqV8hr2RA4HoLRlyYolof2FtfMdaOzQkGgys1VIXijHPPI0eN43R4au8Vn1zOdst2l9h8/tf/755j9TIB2BTt463POKrRQ6FYF9g9UBYKDsMwqFQqNbfBfPfASguPtcJ8V4yc5zJyuVy43W7nf7/fTzAYxO/343Kdkt6RCsWq4gc7f8x39/3wmNuF3MHBD97w3o1LtXev57hFxmRmsvsj9/39Po5hzQB4Zt81PO/C5xzX/hWK9YBpmtWGWvk8hUKBUqlEqVSiUqnUsiXMeW6Wha4WJUCOsNDVYQsL+xjpuo6u604hLK/Xi8/nIxAI4K3Vz1DHU7HeeWL4ST730BeXs+myYjFsjluux5vig8/su+bvf9p//weOte1P+++nM9KuinQpFAuwff1+v995rlwuUyqVKBaL8362rR71dR+AJUXIWpwwF37vpSw/dhyFLSg8Hs88YaGyPxSKxUxmJpcrMNgU7bt/uQIDTsCSAVAySq4P/+BvD2Yrue7lbP+6y16hAkEVihOgXC47D1tomKZJpVKhUqk4z9uprfVZHUtZP85Euu1yaRQbsVRZb1tALHR1eDweXC6XY7HweDzOY6W/n0JxtjOZmeSf7//kcmpiAJTedc1bL11OTzObExIZAPsnDlz1iV9/5icsw20CcPt5t3HN5qed0GcpFIrF2DEdtqXD/t1+zo73ME3TCTqtt4Y0uvZP96S80AJh/78wrVbXdUc4LBQT9sN+XgkJheLEeGL4Se559BvLFhi3n3fb267Z/LTPHM9nnLDIAPjBzh+/71jlxut5Zt813LT9OrwuZbJUKE4ntriof9hWEFt0WJY1zzpSL0KAeTU/GqV4LmShhWShC6fenWFbI4QQjmiof80WGfbPSkgoFKeWZdbCcNgU7fvRW5/xpuuP93NOSmQA/MvPPvnDYzVPq2dTtI8XX3IH8ab4SX2uQqFQKBSK46NklPjSb+/h8Ymdy35PyB0c/KNr33HRsap7NuKkRcbxxmfY3H7ebTxlw2XKqqFQKBQKxRngof5HuHfnd5brHrEpfeBZ794cb4oPnshnnrTIgGpa6z/f/8lfHq/QCLmDvPjC21VQqEKhUCgUp4nBmSHuffzbHJjtP963Hle6aiNOiciAExcaAMlggqf2XsGFXeerKqEKhUKhUJwkJaPEvvH9fGfX9xnLTZzQLk5WYMApFBlwckLD5vzEeVzQvoPeWLeK21AoFAqFYpmkC2kGZgZ5YnQXDww/dDK7OiUCA06xyICq0Ljnka9//niCQZci5A6yobmXrkgnfrcPv9tPWzhxKoapUCgUCsWqZa4wRyo/C8C+qQNM5qZO1GIxj5A7OPiGK19z8/HUwjgap1xkQDUY9Eu/vedLj0/sfB7LTG9VKBQKhUKxcmyK9v3oxZfc8doTDfJsxGkRGTb37/vlXd/Y+a2Po4SGQqFQKBRnK6Vn9l3z9zdtv+5Dy2l6djycVpEB89wn16DEhkKhUCgUZw0hd3DwVZe99MXnJDb9+nTs/7SLDJuH+h+5896d3/noyQSFKhQKhUKhOCWUbt58/YeeseXqvz/V1ot6zpjIgGqsxoOHHnrND/bd96dKbCgUCoVCccYpPbPvmr+/dss1f38iFTyPlzMqMmxKRsn1+NCTt/94/0/fN5ab2I5yoygUCoVCcboohdzBiWv6nvqJKzdc/pkzIS5sVkRk1DM4M7T9kcFHX/rQ8COvyFZyCZTgUCgUCoXiZCkBXNF52b/vaN9276moeXEirLjIqGcyM9l9eHrwqv1TB689PDvwlJqVA5TwUCgUCoXiaJRC7uDEhubeX29u3fSjvljvr05VrYuT4awSGY0YnBnaDjCentheqBQiKz0ehUKhUCjOFvpivb8CSITje05nAOeJctaLDIVCoVAoFKsTbaUHoFAoFAqFYm2iRIZCoVAoFIrTghIZCoVCoVAoTgtKZCgUCoVCoTgtKJGhUCgUCoXitKBEhkKhUCgUitOCEhkKhUKhUChOC0pkKBQKhUKhOC0okaFQKBQKheK0oESGQqFQKBSK04ISGQqFQqFQKE4LSmQoFAqFQqE4LSiRoVAoFAqF4rSgRIZCoVAoFIrTwv8Pyl6mPdjV+ocAAAAASUVORK5CYII='''

        OUT_OF_STOCK_MSG = (
            '<div>'
            '<p style="margin-bottom: 0px;">ESTE PRODUCTO ESTA DESCATALOGADO. YA NO HAY STOCK DISPONIBLE.</p>'
            '<p style="margin-bottom: 0px;">Comuníquese con nosotros para ofrecerle un producto similar con stock inmediato.</p>'
            '<p style="margin-bottom: 0px;">Contacto:</p>'
            '<p style="margin-bottom: 0px;">·&nbsp;Email: <a href="https://comercial@optimaluz.com">comercial@optimaluz.com</a></p>'
            '<p style="margin-bottom: 0px;">·&nbsp;WhatsApp / Telegram / Móvil: (+34) 610 139 920</p>'
            '<p>·&nbsp;Teléfono: (+34) 966 116 649<br></p>'
            '</div>'
        )
        NOTA_INTERNA = "<p>PRODUCTO ZOMBIE (DESCATALOGADO, SIN STOCK Y SIN POSIBILIDAD DE COMPRAR)</p>"

        if not vz_tmpl_ids:
            return

        print(f"🚀 update_vz_products PRO → {len(vz_tmpl_ids)}")

        # --- 1. FILTRAR YA PROCESADOS (OPCIONAL PRO) ---
        templates = models.execute_kw(
            db, uid, password,
            "product.template", "read",
            [vz_tmpl_ids, ["id", "x_icono8"]],
            {"context": ctx}
        )

        to_process = [t["id"] for t in templates if not t.get("x_icono8")]

        if not to_process:
            print("✔ Nada que actualizar")
            return

        # --- 2. WRITE MASIVO BASE ---
        models.execute_kw(
            db, uid, password,
            "product.template", "write",
            [to_process, {
                "allow_out_of_stock_order": False,
                "show_availability": True,
                "available_threshold": 100000,
                "out_of_stock_message": OUT_OF_STOCK_MSG,
                "description": NOTA_INTERNA,
                "x_icono8": DESCATALOGADO_BINARIO,
            }],
            {"context": ctx}
        )

        # --- 3. BUSCAR TODAS LAS IMÁGENES ---
        img_ids = models.execute_kw(
            db, uid, password,
            "product.image", "search",
            [[["product_tmpl_id", "in", to_process]]],
            {"context": ctx}
        )

        images = models.execute_kw(
            db, uid, password,
            "product.image", "read",
            [img_ids, ["id", "sequence"]],
            {"context": ctx}
        )

        # --- 4. AGRUPAR WRITES DE SEQUENCE ---
        seq_groups = defaultdict(list)

        for img in images:
            new_seq = (img.get("sequence") or 0) + 1
            seq_groups[new_seq].append(img["id"])

        for seq, ids in seq_groups.items():
            models.execute_kw(
                db, uid, password,
                "product.image", "write",
                [ids, {"sequence": seq}],
                {"context": ctx}
            )

        # --- 5. CREAR IMÁGENES EN BATCH ---
        new_images = [{
            "product_tmpl_id": tmpl_id,
            "name": "PRODUCTO DESCATALOGADO",
            "image_1920": DESCATALOGADO_BINARIO,
            "sequence": 0,
        } for tmpl_id in to_process]

        models.execute_kw(
            db, uid, password,
            "product.image", "create",
            [new_images],
            {"context": ctx}
        )

        print(f"✅ VZ procesados: {len(to_process)}")

    update_vz_products(nuevos_vz)
    print("75%")

def update_out_of_stock_msg_from_excel(excel_path):
    from App_Connection import db, uid, password, models
    import pandas as pd
    from collections import defaultdict

    print("🚀 update_out_of_stock_msg PRO")

    ctx = {"active_test": False, "lang": "es_ES"}

    # --- MENSAJES ---
    MSG_MADRID = "<p>Haga ya su pedido y recibirá las nuevas existencias que estan llegando a nuestro almacen en un plazo de 1 a 2 días hábiles.</p>"
    MSG_BULGARIA = "<p>Haga ya su pedido y recibirá las nuevas existencias que llegarán desde nuestro almacén europeo en un plazo de 7 a 9 días hábiles.</p>"
    MSG_PROX = (
        '<div>'
        '<p style="margin-bottom: 0px;">STOCK DISPONIBLE PRÓXIMAMENTE.</p>'
        '<p style="margin-bottom: 0px;">Haga ya su reserva y nos pondremos en contacto con usted para indicarle la fecha exacta de entrada del producto o bien ofrecerle un producto similar en stock inmediato.</p>'
        '<p style="margin-bottom: 0px;">Contacto:</p>'
        '<p style="margin-bottom: 0px;">·&nbsp;Email: <a href="https://comercial@optimaluz.com">comercial@optimaluz.com</a></p>'
        '<p style="margin-bottom: 0px;">·&nbsp;WhatsApp / Telegram / Móvil: (+34) 610 139 920</p>'
        '<p>·&nbsp;Teléfono: (+34) 966 116 649<br></p>'
        '</div>'
    )

    def norm_sku(x):
        if not x:
            return None
        s = str(x).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s

    def to_int(v):
        try:
            return int(float(v))
        except:
            return 0

    # --- LOAD EXCEL ---
    df_mad = pd.read_excel(excel_path, sheet_name="Madrid", dtype=str)
    df_bul = pd.read_excel(excel_path, sheet_name="Bulgaria", dtype=str)
    df_prox = pd.read_excel(excel_path, sheet_name="Proximamente", dtype=str)

    madrid = {norm_sku(r["SKU"]): to_int(r["STOCK"]) for _, r in df_mad.iterrows()}
    bulgaria = {norm_sku(r["SKU"]): to_int(r["STOCK"]) for _, r in df_bul.iterrows()}
    proximamente = {norm_sku(r["SKU"]) for _, r in df_prox.iterrows()}

    all_skus = set(madrid) | set(bulgaria) | proximamente
    all_skus.discard(None)

    # --- READ ODOO MASIVO ---
    skus = list(all_skus)
    CHUNK = 1000

    tmpl_map = {}

    for i in range(0, len(skus), CHUNK):
        batch = skus[i:i+CHUNK]

        rows = models.execute_kw(
            db, uid, password,
            "product.product", "search_read",
            [[("default_code", "in", batch)]],
            {"fields": ["default_code", "product_tmpl_id"]}
        )

        tmpl_ids = [r["product_tmpl_id"][0] for r in rows]

        templates = models.execute_kw(
            db, uid, password,
            "product.template", "read",
            [tmpl_ids, ["id", "out_of_stock_message", "allow_out_of_stock_order"]]
        )

        for r, t in zip(rows, templates):
            tmpl_map[r["default_code"].strip()] = {
                "id": t["id"],
                "msg": t.get("out_of_stock_message"),
                "allow": t.get("allow_out_of_stock_order")
            }

    # --- CALCULAR CAMBIOS ---
    grouped = defaultdict(list)

    for sku in all_skus:
        if sku not in tmpl_map:
            continue

        new_msg = None
        new_allow = None

        if madrid.get(sku, 0) > 0:
            new_msg = MSG_MADRID
            new_allow = True

        elif bulgaria.get(sku, 0) > 0 and sku not in madrid:
            new_msg = MSG_BULGARIA
            new_allow = True

        elif sku in proximamente and sku not in madrid:
            new_msg = MSG_PROX
            new_allow = False

        if not new_msg:
            continue

        current = tmpl_map[sku]

        if current["msg"] == new_msg and current["allow"] == new_allow:
            continue

        key = (new_msg, new_allow)
        grouped[key].append(current["id"])

    # --- WRITE MASIVO ---
    total = 0

    for (msg, allow), ids in grouped.items():
        models.execute_kw(
            db, uid, password,
            "product.template", "write",
            [ids, {
                "out_of_stock_message": msg,
                "allow_out_of_stock_order": allow
            }],
            {"context": ctx}
        )
        total += len(ids)

    print(f"✅ Mensajes actualizados: {total}")

def actualizar_stock_comercial(excel_path):
    import threading, traceback, inspect

    # Define aquí el orden y las funciones que quieres ejecutar
    tasks = [
        ("update_comercial_stock", update_comercial_stock),
        ("update_names", update_names),
        ("update_out_of_stock_msg_from_excel", update_out_of_stock_msg_from_excel),
    ]

    results = {k: None for k, _ in tasks}
    results["errors"] = {}

    def run_target(key, fn):
        try:
            sig = inspect.signature(fn)
            if "excel_path" in sig.parameters:
                results[key] = fn(excel_path)
            else:
                results[key] = fn()
        except TypeError:
            # Por si la firma no declara 'excel_path' pero aceptara *args/**kwargs
            try:
                results[key] = fn()
            except Exception as e2:
                results["errors"][key] = f"{e2}\n{traceback.format_exc()}"
        except Exception as e:
            results["errors"][key] = f"{e}\n{traceback.format_exc()}"

    # Secuencial: un hilo por tarea, join inmediato
    for idx, (key, fn) in enumerate(tasks, 1):
        t = threading.Thread(target=run_target, args=(key, fn), name=f"{idx:02d}_{key}")
        t.start()
        t.join()

    print("100%")
    return results

def change_internal_categories_by_prefix():
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    ctx = {"active_test": False, "lang": "es_ES"}

    PARENT_NAME = "Productos descatalogados"#31
    CHILD_VD_NAME = "Con stock (VD)"
    CHILD_VZ_VPD_NAME = "Sin stock (VZ)"

    def get_category_id(child_name):
        # Buscar padre
        parent_ids = models.execute_kw(
            db, uid, password, "product.category", "search",
            [[["name", "=", PARENT_NAME]]],
            {"context": ctx}
        )
        if not parent_ids:
            print(f"❌ Categoría padre no encontrada: '{PARENT_NAME}'")
            return None
        parent_id = parent_ids[0]

        # Buscar hijo bajo ese padre
        child_ids = models.execute_kw(
            db, uid, password, "product.category", "search",
            [[["name", "=", child_name], ["parent_id", "=", parent_id]]],
            {"context": ctx}
        )
        if not child_ids:
            print(f"❌ Categoría hija no encontrada: '{PARENT_NAME} / {child_name}'")
            return None
        if len(child_ids) > 1:
            print(
                f"⚠ Aviso: múltiples coincidencias para '{PARENT_NAME} / {child_name}'. Usando {child_ids[0]}")
        return child_ids[0]

    cat_vd = get_category_id(CHILD_VD_NAME)
    cat_vz_vpd = get_category_id(CHILD_VZ_VPD_NAME)
    if not cat_vd or not cat_vz_vpd:
        return

    # Buscar plantillas por prefijo en name
    vd_ids = models.execute_kw(
        db, uid, password, "product.template", "search",
        [[["name", "ilike", "[VD%"]]],
        {"context": ctx}
    )
    vz_ids = models.execute_kw(
        db, uid, password, "product.template", "search",
        [[["name", "ilike", "[VZ%"]]],
        {"context": ctx}
    )
    vpd_ids = models.execute_kw(
        db, uid, password, "product.template", "search",
        [[["name", "ilike", "[VPD%"]]],
        {"context": ctx}
    )

    vz_vpd_ids = list({*vz_ids, *vpd_ids})  # deduplicar

    print(f"➡ [VD…] encontrados: {len(vd_ids)}  |  [VZ…] + [VPD…] encontrados: {len(vz_vpd_ids)}")

    updated_vd = 0
    updated_vz_vpd = 0

    # Actualizar VD -> Con stock (VD)
    if vd_ids:
        models.execute_kw(
            db, uid, password, "product.template", "write",
            [vd_ids, {"categ_id": cat_vd}],
            {"context": ctx}
        )
        updated_vd = len(vd_ids)

    # Actualizar VZ + VPD -> Sin stock (VZ & VPD)
    if vz_vpd_ids:
        models.execute_kw(
            db, uid, password, "product.template", "write",
            [vz_vpd_ids, {"categ_id": cat_vz_vpd}],
            {"context": ctx}
        )
        updated_vz_vpd = len(vz_vpd_ids)

    print("----- RESUMEN change_internal_categories_by_prefix -----")
    print(f"Plantillas [VD…] movidas a '{PARENT_NAME} / {CHILD_VD_NAME}': {updated_vd}")
    print(f"Plantillas [VZ…] + [VPD…] movidas a '{PARENT_NAME} / {CHILD_VZ_VPD_NAME}': {updated_vz_vpd}")

def import_comercial_stock():
    import time

    # region Helpers de uso general ===============
    def abrir_vtac_gui(comercial_flag: bool = True, autostart: bool = True):
        import os, sys, subprocess, glob
        base_dir = os.path.expanduser(r"~/Documents/SMI Files")
        # Busca VtacGUI*.exe por si lo renombraste (coge el primero que encuentre)
        matches = sorted(glob.glob(os.path.join(base_dir, "VtacGUI*.exe")))
        if not matches:
            raise FileNotFoundError(f"No se encontró VtacGUI.exe en: {base_dir}")
        exe_path = matches[0]

        env = os.environ.copy()
        env.setdefault("SCRAPY_SETTINGS_MODULE", "Scrap.settings")

        args = [exe_path]
        if autostart: args += ["--autostart"]
        args += ["--comercial", "1" if comercial_flag else "0"]

        # Lanza sin bloquear tu app; cwd asegura rutas relativas correctas
        subprocess.Popen(args, cwd=base_dir, env=env, shell=False)

    def _norm(s: str) -> str:
        """Normaliza nombre de columna: minúsculas y solo alfanumérico."""
        return "".join(ch for ch in str(s).lower() if ch.isalnum())

    def _find_col(df: pd.DataFrame, wanted: str, *alts) -> str | None:
        """
        Devuelve el nombre real de la columna en df cuya versión normalizada
        coincide con 'wanted' o cualquiera de 'alts'.
        """
        wanted_norms = {_norm(wanted), *({_norm(a) for a in alts} if alts else set())}
        for col in df.columns:
            if _norm(col) in wanted_norms:
                return col
        return None

    def _to_numeric(series, errors="coerce", fillna=None):
        s = pd.to_numeric(series, errors=errors)
        if fillna is not None:
            s = s.fillna(fillna)
        return s

    def _list_excel_files(dir_path: str, contains: str):
        contains_low = contains.lower()
        paths = []
        for root, _dirs, files in os.walk(dir_path):
            for f in files:
                low = f.lower()
                if contains_low in low and low.endswith((".xlsx", ".xls", ".xlsm")):
                    paths.append(os.path.join(root, f))
        return paths

    def _concat_clean(dfs, cols_keep=None):
        if not dfs:
            return pd.DataFrame(columns=cols_keep or [])
        df = pd.concat(dfs, ignore_index=True)
        if cols_keep:
            # conserva solo columnas requeridas si existen
            keep = [c for c in cols_keep if c in df.columns]
            df = df[keep]
        return df

    # endregion

    # 1) Madrid: ficheros que contienen "In Stock" ===============
    def parse_madrid_from_dir(excel_dir: str) -> pd.DataFrame:
        """
        Buscar ficheros cuyo nombre contenga 'In Stock'.
        Tomar:
          - 'Item No.' -> SKU
          - 'In Stock' -> STOCK
        """
        files = _list_excel_files(excel_dir, "In Stock")
        out = []
        for p in files:
            try:
                df = pd.read_excel(p, dtype=str)  # primera hoja
                col_sku = _find_col(df, "Item No.", "item no", "itemno", "sku")
                col_stk = _find_col(df, "In Stock", "instock", "stock")
                if not col_sku or not col_stk:
                    print(f"⚠ Saltando (columnas no halladas): {p}")
                    continue

                tmp = pd.DataFrame({
                    "SKU": df[col_sku].astype(str).str.strip().str.removesuffix(".0"),
                    "STOCK": _to_numeric(df[col_stk], errors="coerce", fillna=0).astype(int),
                })
                tmp = tmp[tmp["SKU"].astype(str).str.strip().ne("") & tmp["SKU"].notna()]
                out.append(tmp)
            except Exception as e:
                print(f"❌ Error leyendo Madrid '{p}': {e}")

        res = _concat_clean(out, cols_keep=["SKU", "STOCK"])
        res = res.drop_duplicates(subset=["SKU"]).reset_index(drop=True)
        return res

    # 2) Bulgaria: ficheros que contienen "SPRAVKA" ===============
    def parse_bulgaria_from_dir(excel_dir: str) -> pd.DataFrame:

        files = _list_excel_files(excel_dir, "SPRAVKA")
        out = []
        for p in files:
            try:
                df = pd.read_excel(p, dtype=str, header=1)

                # 🔧 Normaliza cabeceras: quita NBSP y espacios
                df.columns = [str(c).replace("\u00a0", " ").strip() for c in df.columns]

                # Parsear: drop fila
                if 0 in df.index:
                    df = df.drop(index=0)

                # Localizar columnas tras strip
                col_sku = _find_col(df, "SKU ")
                col_av = _find_col(df, "AVAILABLE ", "available")
                col_un = _find_col(df, "UNDELIVERED ORDER ", "undelivered order", "undeliveredorder")

                if not col_sku:
                    print(f"⚠ Saltando Bulgaria (SKU no encontrada): {p}")
                    continue

                tmp = pd.DataFrame({
                    "SKU": df[col_sku].astype(str).str.strip().str.removesuffix(".0")
                })

                if col_av:
                    tmp["STOCK"] = _to_numeric(df[col_av].astype(str).str.strip(), errors="coerce",
                                               fillna=0).astype(int)
                else:
                    tmp["STOCK"] = 0

                if col_un:
                    tmp["UNDELIVERED ORDER"] = df[col_un]

                tmp = tmp[tmp["SKU"].ne("") & tmp["SKU"].notna()]
                out.append(tmp)

            except Exception as e:
                print(f"❌ Error leyendo Bulgaria '{p}': {e}")

        res = _concat_clean(out, cols_keep=["SKU", "STOCK", "UNDELIVERED ORDER"])
        for c in ["STOCK", "UNDELIVERED ORDER"]:
            if c not in res.columns:
                res[c] = 0
        return res.drop_duplicates(subset=["SKU"]).reset_index(drop=True)

    # 3) VS: fichero que contiene "SKU_Page" ===============
    def parse_vs_from_dir(excel_dir: str) -> pd.DataFrame:
        """
        Buscar fichero cuyo nombre contenga 'SKU_Page'.
        Tomar columna 'SKU' en hoja VS.
        """
        files = _list_excel_files(excel_dir, "SKU_Page")
        out = []
        for p in files:
            try:
                df = pd.read_excel(p, dtype=str)
                col_sku = _find_col(df, "SKU")
                if not col_sku:
                    print(f"⚠ Saltando VS (SKU no encontrada): {p}")
                    continue
                tmp = pd.DataFrame({"SKU": df[col_sku].astype(str).str.strip().str.removesuffix(".0")})
                tmp = tmp[tmp["SKU"].astype(str).str.strip().ne("") & tmp["SKU"].notna()]
                out.append(tmp)
            except Exception as e:
                print(f"❌ Error leyendo VS '{p}': {e}")

        res = _concat_clean(out, cols_keep=["SKU"]).drop_duplicates(subset=["SKU"]).reset_index(drop=True)
        return res

    # 4) Odoo: productos V-Tac (SKU y STOCK) ===============
    def odoo_vtac(models, uid, db, password) -> pd.DataFrame:
        """
        Extrae de Odoo todos los product.product cuya marca del template sea 'V-Tac',
        devolviendo default_code -> SKU y qty_available -> STOCK.
        """

        ctx = {"active_test": False, "lang": "es_ES"}
        # Variantes cuya plantilla tenga la marca V-Tac y que tengan código
        v_ids = models.execute_kw(
            db, uid, password, "product.product", "search",
            [[
                ["product_tmpl_id.product_brand_id.name", "=", "V-Tac"],
                ["default_code", "!=", False],
            ]],
            {"context": ctx}
        )
        if not v_ids:
            return pd.DataFrame(columns=["SKU", "STOCK"])

        # leer por lotes para no saturar
        def batched(lst, n=200):
            for i in range(0, len(lst), n):
                yield lst[i:i + n]

        rows = []
        for batch in batched(v_ids, 200):
            recs = models.execute_kw(
                db, uid, password, "product.product", "read",
                [batch, ["default_code", "qty_available"]],
                {"context": ctx}
            )
            for r in recs:
                sku = str(r.get("default_code") or "").strip().removesuffix(".0")
                if not sku:
                    continue
                qty = r.get("qty_available") or 0
                try:
                    qty = int(qty)
                except Exception:
                    qty = float(qty)
                rows.append({"SKU": sku, "STOCK": qty})

        df = pd.DataFrame(rows)
        if df.empty:
            return pd.DataFrame(columns=["SKU", "STOCK"])
        # combinar por SKU sumando stock si hay varias variantes con mismo código (raro, pero seguro)
        df = df.groupby("SKU", as_index=False)["STOCK"].sum()
        return df

    # 5) VD / VN desde hoja "Products" externa ===============
    def vd_vn_spain(products_excel_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Lee hoja 'Products' del Excel pasado y separa:
          - 'Descatalogados' -> hoja VD (columna SKU)
          - 'Nuevos Productos' -> hoja VN (columna SKU)
        """
        df = pd.read_excel(products_excel_path, sheet_name="Products", dtype=str)
        col_cat = _find_col(df, "Categoría", "categoria", "category")
        col_sku = _find_col(df, "SKU", "sku", "default_code")
        if not col_cat or not col_sku:
            raise ValueError("No se encontraron las columnas 'Categoría' y/o 'SKU' en la hoja 'Products'.")

        df["__cat__"] = df[col_cat].astype(str).str.strip()
        df["__sku__"] = df[col_sku].astype(str).str.strip().str.removesuffix(".0")
        df = df[df["__sku__"].ne("") & df["__sku__"].notna()]

        vd = df[df["__cat__"] == "Descatalogados"][["__sku__"]].rename(
            columns={"__sku__": "SKU"}).drop_duplicates()
        vn = df[df["__cat__"] == "Nuevos Productos"][["__sku__"]].rename(
            columns={"__sku__": "SKU"}).drop_duplicates()

        return vd.reset_index(drop=True), vn.reset_index(drop=True)

    # 6) Constructor del Excel Base ===============
    def create_excel_base(excel_dir: str, output_path: str):
        """
        - Lee directorio 'excel_dir' para Madrid, Bulgaria, VS.
        - Lee Odoo para Odoo (marca V-Tac).
        - Lee 'products_excel_path' para VD y VN.
        - Escribe a 'output_path' (xlsx) con hojas:
          Madrid, Bulgaria, VS, Odoo, VD, VN
        """
        print(">> Generando Excel Base...")

        # =========================
        # 🔌 Conexión
        # =========================
        def conectar(config):
            common = xmlrpc.client.ServerProxy(f"{config['url']}/xmlrpc/2/common")
            uid = common.authenticate(config['db'], config['user'], config['password'], {})

            models = xmlrpc.client.ServerProxy(f"{config['url']}/xmlrpc/2/object")

            return models, uid

        odoo16 = {
            'url': "http://37.59.66.189:8069/",
            'db': "Real",
            'user': "jcoronado@optimaluz.com",
            'password': "AlAi4ever",
        }
        models, uid = conectar(odoo16)
        db = odoo16["db"]
        password = odoo16["password"]

        # 1) Fuentes de ficheros
        madrid = parse_madrid_from_dir(excel_dir)
        print(f"  Madrid: {len(madrid)} filas")

        bulgaria = parse_bulgaria_from_dir(excel_dir)
        print(f"  Bulgaria: {len(bulgaria)} filas")

        vs = parse_vs_from_dir(excel_dir)
        print(f"  VS: {len(vs)} filas")

        # 2) Odoo
        odoo_df = odoo_vtac(models, uid, db, password)
        print(f"  Odoo (V-Tac): {len(odoo_df)} filas")

        # 3) Guardar
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            madrid.to_excel(writer, sheet_name="Madrid", index=False)
            bulgaria.to_excel(writer, sheet_name="Bulgaria", index=False)
            vs.to_excel(writer, sheet_name="VS", index=False)
            odoo_df.to_excel(writer, sheet_name="Odoo", index=False)

        print(f"✅ Excel Base creado: {output_path}")

    def create_vs_vn():
        print("Buscando...")
        path = os.path.expanduser(r"~/Documents/SMI Files/Comercial/scraped.xlsx")
        if os.path.exists(path):
            print("Encontrado!!!")
            try:
                vd, vn = vd_vn_spain(path)
            except Exception as e:
                print(f"⚠ No se pudieron generar VD/VN desde '{path}': {e}")
                vd, vn = pd.DataFrame(columns=["SKU"]), pd.DataFrame(columns=["SKU"])

            with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                vd.to_excel(writer, sheet_name="VD", index=False)
                vn.to_excel(writer, sheet_name="VN", index=False)

            def crear_hoja_resultado(excel_path, hoja_resultado="Resultado"):
                """
                Lee un Excel con hojas: Madrid, Bulgaria, Odoo, VS, VD, VN
                y crea la hoja 'Resultado' con las columnas: VS, VN, VD, VZ,
                Novedades VN, Novedades VS, Novedades VD, Novedades VZ.
                """
                from pathlib import Path
                excel_path = Path(excel_path)

                # --- Helpers ---
                def norm_sku_series(s):
                    if s is None:
                        return []
                    s = s.astype(str).str.strip()
                    s = s[s.ne("") & s.ne("nan")]
                    return list(dict.fromkeys(s.tolist()))

                def to_num(series):
                    return pd.to_numeric(series, errors="coerce")

                # --- Leer hojas ---
                xls = pd.ExcelFile(excel_path)
                VS_df = pd.read_excel(xls, "VS", dtype={"SKU": str}, usecols=["SKU"])
                VD_df = pd.read_excel(xls, "VD", dtype={"SKU": str}, usecols=["SKU"])
                VN_df = pd.read_excel(xls, "VN", dtype={"SKU": str}, usecols=["SKU"])
                Madrid_df = pd.read_excel(xls, "Madrid", dtype={"SKU": str}, usecols=["SKU", "STOCK"])
                Odoo_df = pd.read_excel(xls, "Odoo", dtype={"SKU": str}, usecols=["SKU", "STOCK"])
                Bulgaria_df = pd.read_excel(
                    xls, "Bulgaria", dtype={"SKU": str}, usecols=["SKU", "STOCK", "UNDELIVERED ORDER"]
                )

                # --- Normalizar ---
                VS_skus = norm_sku_series(VS_df["SKU"])
                VD_skus = norm_sku_series(VD_df["SKU"])
                VN_skus = norm_sku_series(VN_df["SKU"])

                for df in (Madrid_df, Odoo_df, Bulgaria_df):
                    df["SKU"] = df["SKU"].astype(str).str.strip()

                Madrid_df["STOCK"] = to_num(Madrid_df["STOCK"])
                Odoo_df["STOCK"] = to_num(Odoo_df["STOCK"])
                Bulgaria_df["STOCK"] = to_num(Bulgaria_df["STOCK"])

                bul_undel = Bulgaria_df["UNDELIVERED ORDER"].astype(str).str.strip()
                bul_undel = bul_undel.where(~bul_undel.isin(["", "nan"]), other=pd.NA)

                # --- VS final = VS - VD ---
                VD_set = set(VD_skus)
                VS_final = [sku for sku in VS_skus if sku not in VD_set]

                # --- VN final = VN - VS ---
                VS_set = set(VS_skus)
                VN_final = [sku for sku in VN_skus if sku not in VS_set]

                # --- Novedades ---
                Odoo_skus = set(norm_sku_series(Odoo_df["SKU"]))
                Novedades_VN = [sku for sku in VN_final if sku not in Odoo_skus]
                # el resto se calculan más abajo tras VD/VZ

                # --- VD/VZ ---
                madrid_stock = dict(zip(Madrid_df["SKU"], Madrid_df["STOCK"]))
                odoo_stock = dict(zip(Odoo_df["SKU"], Odoo_df["STOCK"]))
                bul_stock = dict(zip(Bulgaria_df["SKU"], Bulgaria_df["STOCK"]))
                bul_undel_map = dict(zip(Bulgaria_df["SKU"], bul_undel))

                VD_final, VZ_final = [], []

                for sku in VD_skus:
                    enviar_a_vz = True

                    # Bulgaria
                    b_stock = bul_stock.get(sku, pd.NA)
                    b_undel = bul_undel_map.get(sku, pd.NA)
                    stock_cond = (pd.isna(b_stock) or (pd.notna(b_stock) and b_stock < 1))
                    undel_cond = pd.isna(b_undel)
                    if stock_cond and undel_cond:
                        enviar_a_vz = True
                    else:
                        enviar_a_vz = False

                    # Madrid
                    m_stock = madrid_stock.get(sku, pd.NA)
                    if pd.notna(m_stock) and m_stock > 0:
                        enviar_a_vz = False

                    # Odoo
                    o_stock = odoo_stock.get(sku, pd.NA)
                    if not (pd.notna(o_stock) and o_stock < 1):
                        enviar_a_vz = False

                    (VZ_final if enviar_a_vz else VD_final).append(sku)

                # --- Novedades para todas las listas ---
                Novedades_VS = [sku for sku in VS_final if sku not in Odoo_skus]
                Novedades_VD = [sku for sku in VD_final if sku not in Odoo_skus]
                Novedades_VZ = [sku for sku in VZ_final if sku not in Odoo_skus]

                # --- Construcción hoja Resultado ---
                max_len = max(
                    len(VS_final), len(VN_final), len(VD_final), len(VZ_final),
                    len(Novedades_VN), len(Novedades_VS), len(Novedades_VD), len(Novedades_VZ)
                )

                def pad(lst, n):
                    return lst + [None] * (n - len(lst))

                resultado_df = pd.DataFrame({
                    "VS": pad(VS_final, max_len),
                    "VN": pad(VN_final, max_len),
                    "VD": pad(VD_final, max_len),
                    "VZ": pad(VZ_final, max_len),
                    "Novedades VN": pad(Novedades_VN, max_len),
                    "Novedades VS": pad(Novedades_VS, max_len),
                    "Novedades VD": pad(Novedades_VD, max_len),
                    "Novedades VZ": pad(Novedades_VZ, max_len),
                })

                # --- Escribir/actualizar hoja Resultado ---
                with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a",
                                    if_sheet_exists="replace") as writer:
                    resultado_df.to_excel(writer, sheet_name=hoja_resultado, index=False)

                return {
                    "VS": VS_final, "VN": VN_final, "VD": VD_final, "VZ": VZ_final,
                    "Novedades VN": Novedades_VN,
                    "Novedades VS": Novedades_VS,
                    "Novedades VD": Novedades_VD,
                    "Novedades VZ": Novedades_VZ,
                }

            crear_hoja_resultado(os.path.expanduser(r"~/Documents/SMI Files/Comercial/Excel_Base.xlsx"))
        else:
            time.sleep(60)
            create_vs_vn()

    # 7) Ejecucion:
    abrir_vtac_gui()

    # Directorio donde están los excels fuente (Madrid, SPRAVKA, SKU_Page)
    excel_dir = os.path.expanduser(r"~/Documents/SMI Files/Comercial")

    # Salida final
    output_path = os.path.expanduser(r"~/Documents/SMI Files/Comercial/Excel_Base.xlsx")

    create_excel_base(excel_dir, output_path)

    create_vs_vn()

    actualizar_stock_comercial(output_path)
#endregion

# region REGIÓN CONTACTOS

puerto = "http://79.72.61.76:8069/"  # http://79.72.56.200:8069
database = "Test"  # odoo0


# ---------------------------------------------------------------------------
# 1) EXPORTAR CONTACTOS DESDE EL ODOO ANTIGUO
# ---------------------------------------------------------------------------

def export_contacts():
    """
    Devuelve un dict:
        {
          "partners":   [ {...}, ... ],   # compañías y personas
          "categories": [ {"id":..,"name":..}, ... ]
        }
    Listo para serializar (JSON, pickle, etc.).
    """
    try:
        import xmlrpc.client
        from App_Connection import db, uid, password, models

        # Campos del partner que vamos a conservar
        PARTNER_FIELDS = [
            "id", "name", "company_type",  # 'company' | 'person' | 'contact'
            "parent_id",
            "street", "street2", "zip", "city",
            "state_id", "country_id",
            "phone", "mobile", "email",
            "vat",
            "customer_rank", "supplier_rank",
            "category_id",  # etiquetas (many2many)
            # Propiedades contables que apuntan a nuestros mapas
            "property_account_position_id",
            "property_payment_term_id",
            "property_supplier_payment_term_id",
            "active",
            "property_product_pricelist",
            # Asientos contables
            "property_account_receivable_id",
            "property_account_playable_id",
        ]

        partners = models.execute_kw(
            db, uid, password,
            "res.partner", "search_read",
            [[("active", "=", True)]],  # quita el filtro si quieres todos
            {"fields": PARTNER_FIELDS}
        )

        # Convertimos relaciones many2one/many2many a números simples
        for p in partners:
            p["parent_id"] = p["parent_id"][0] if p["parent_id"] else None
            p["state_id"] = p["state_id"][0] if p["state_id"] else None
            p["country_id"] = p["country_id"][0] if p["country_id"] else None

            p["category_id"] = p["category_id"] or []

            p["property_account_position_id"] = (
                p["property_account_position_id"][0] if p["property_account_position_id"] else None
            )
            p["property_payment_term_id"] = (
                p["property_payment_term_id"][0] if p["property_payment_term_id"] else None
            )
            p["property_supplier_payment_term_id"] = (
                p["property_supplier_payment_term_id"][0] if p["property_supplier_payment_term_id"] else None
            )
            p["property_product_pricelist"] = (
                p["property_product_pricelist"][0] if p["property_product_pricelist"] else None
            )

        # Exportamos también las etiquetas
        categories = models.execute_kw(
            db, uid, password,
            "res.partner.category", "search_read",
            [[]], {"fields": ["id", "name"]}
        )

        return {"partners": partners, "categories": categories}

    except Exception as e:
        print(f"❌ Error al exportar contactos: {e}")

# ---------------------------------------------------------------------------
# 2) IMPORTAR CONTACTOS EN EL ODOO NUEVO
# ---------------------------------------------------------------------------

def import_contacts_to_odoo(data,
                            url=puerto,
                            db=database,
                            username="admin",
                            password="admin"):
    """
    Importa compañías y contactos en el Odoo destino.
    · Si el partner ya existe (misma clave name+vat/email) ‒> se ACTUALIZA con write().
    · Si no existe ‒> se crea.

    :param data:       dict devuelto por export_contacts()
    :return:           {old_partner_id: new_partner_id}
    """
    import xmlrpc.client

    # 1. Conexión ----------------------------------------------------------------
    common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
    uid = common.authenticate(db, username, password, {})
    if not uid:
        raise RuntimeError("❌ Autenticación fallida")

    models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")

    partners = data["partners"]
    categories = data["categories"]

    # 2. Etiquetas ----------------------------------------------------------------
    cat_name_to_id = {c["name"]: c["id"] for c in models.execute_kw(
        db, uid, password,
        "res.partner.category", "search_read",
        [[]], {"fields": ["id", "name"]}
    )}
    old_cat_id_to_name = {c["id"]: c["name"] for c in categories}

    for name in old_cat_id_to_name.values():
        if name not in cat_name_to_id:
            new_id = models.execute_kw(
                db, uid, password,
                "res.partner.category", "create",
                [{"name": name}]
            )
            cat_name_to_id[name] = new_id

    # 3. Países y estados ---------------------------------------------------------
    country_map = {c["code"]: c["id"] for c in models.execute_kw(
        db, uid, password,
        "res.country", "search_read",
        [[]], {"fields": ["id", "code"]}
    )}

    state_map = {}
    for st in models.execute_kw(
            db, uid, password,
            "res.country.state", "search_read",
            [[]], {"fields": ["id", "code", "name", "country_id"]}
    ):
        key = (st["code"] or st["name"],
               st["country_id"][0] if isinstance(st["country_id"], (list, tuple)) else st["country_id"])
        state_map[key] = st["id"]



    # 4. Partners existentes (clave = name+vat/email) -----------------------------
    def partner_key(p):
        return (p["name"].strip().lower(),
                (p.get("vat") or p.get("email") or "").strip().lower())

    existing_partners = models.execute_kw(
        db, uid, password,
        "res.partner", "search_read",
        [[]], {"fields": ["id", "name", "vat", "email"]}
    )
    existing_map = {partner_key(p): p["id"] for p in existing_partners}

    old2new = {}

    # ---------------------------------------------------------------------------
    # 5. PRIMERA PASADA · COMPAÑÍAS
    # ---------------------------------------------------------------------------
    for p in partners:
        if p["company_type"] != "company":
            continue

        key = partner_key(p)

        vals = _prepare_partner_vals(
            p, parent_new_id=None,
            old_cat_id_to_name=old_cat_id_to_name,
            cat_name_to_id=cat_name_to_id,
            country_map=country_map, state_map=state_map
        )
        vals = _none_to_false(vals)
        vals = _sanitize_for_xmlrpc(vals)

        if key in existing_map:
            new_id = existing_map[key]
            # actualiza (sobrescribe) la información
            models.execute_kw(
                db, uid, password,
                "res.partner", "write",
                [[new_id], vals]
            )
        else:
            new_id = models.execute_kw(
                db, uid, password,
                "res.partner", "create", [vals]
            )
            existing_map[key] = new_id

        old2new[p["id"]] = new_id

    # ---------------------------------------------------------------------------
    # 6. SEGUNDA PASADA · PERSONAS / CONTACTOS
    # ---------------------------------------------------------------------------
    for p in partners:
        if p["company_type"] == "company":
            continue

        parent_new = old2new.get(p["parent_id"])
        key = partner_key(p)

        vals = _prepare_partner_vals(
            p, parent_new_id=parent_new,
            old_cat_id_to_name=old_cat_id_to_name,
            cat_name_to_id=cat_name_to_id,
            country_map=country_map, state_map=state_map
        )
        vals = _none_to_false(vals)
        vals = _sanitize_for_xmlrpc(vals)

        if key in existing_map:
            new_id = existing_map[key]
            models.execute_kw(
                db, uid, password,
                "res.partner", "write",
                [[new_id], vals]
            )
        else:
            new_id = models.execute_kw(
                db, uid, password,
                "res.partner", "create", [vals]
            )
            existing_map[key] = new_id

        old2new[p["id"]] = new_id

    print(f"✅ Sincronizados {len(old2new)} contactos (compañías + personas).")
    return old2new

# ---------------------------------------------------------------------------
# 3) HELPERS
# ---------------------------------------------------------------------------
def _none_to_false(vals):
    return {k: (False if v is None else v) for k, v in vals.items()}

def _sanitize_for_xmlrpc(obj):
    """
    Devuelve una copia de `obj` (dict, list, tuple o valor simple)
    en la que los None se cambian a False recursivamente.
    """
    if obj is None:
        return False

    if isinstance(obj, dict):
        return {k: _sanitize_for_xmlrpc(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [_sanitize_for_xmlrpc(v) for v in obj]

    if isinstance(obj, tuple):
        return tuple(_sanitize_for_xmlrpc(v) for v in obj)

    return obj  # int, str, bool…

def _none_to_false(vals):
    """Devuelve una copia donde cualquier valor None se convierte en False."""
    return {k: (False if v is None else v) for k, v in vals.items()}

def _prepare_partner_vals(p,
                          parent_new_id,
                          old_cat_id_to_name,
                          cat_name_to_id,
                          country_map,
                          state_map):
    """
    Convierte el dict exportado en un vals dict compatible con res.partner.create()
    """
    # -------- País y estado ---------------------------------------------------
    country_id = None
    if p["country_id"]:
        # Para mayor robustez, exporta el ISO code del país en lugar del ID
        # Aquí asumimos que el id coincide o que no pasa nada si no se asigna
        country_id = p["country_id"] if p["country_id"] in country_map.values() else None

    state_id = None
    if p["state_id"] and country_id:
        # Mapping por code o por name
        for key, _id in state_map.items():
            if _id == p["state_id"]:
                state_id = _id
                break

    # -------- Categorías (etiquetas) -----------------------------------------
    new_cat_ids = []
    for old_cat_id in p["category_id"]:
        name = old_cat_id_to_name.get(old_cat_id)
        if name:
            new_id = cat_name_to_id.get(name)
            if new_id:
                new_cat_ids.append(new_id)
    cat_m2m = [(6, 0, new_cat_ids)] if new_cat_ids else False

    # -------- Propiedades contables ------------------------------------------
    country_id = country_id or False
    state_id = state_id or False

    property_account_receivable_id = Utils.get_by_x_id_interno("account.account", p["property_account_receivable_id"])
    property_account_playable_id = Utils.get_by_x_id_interno("account.account", p["property_account_playable_id"])

    return {
        "name": p["name"],
        "company_type": p["company_type"] or "person",
        "parent_id": parent_new_id,
        "street": p["street"],
        "street2": p["street2"],
        "zip": p["zip"],
        "city": p["city"],
        "state_id": state_id,
        "country_id": country_id,
        "phone": p["phone"],
        "mobile": p["mobile"],
        "email": p["email"],
        "vat": p["vat"],
        "customer_rank": p["customer_rank"],
        "supplier_rank": p["supplier_rank"],
        "category_id": cat_m2m,
        "active": p["active"],
        "lang": "es_ES",
        "property_account_receivable_id": property_account_receivable_id,
        "property_account_playable_id": property_account_playable_id
    }

# --- 0) Importar contactos -------------
#contacts_data = export_contacts()  # ya lo tenías
#import_contacts_to_odoo(contacts_data)

# CUENTAS BANCARIAS
def export_partner_bank_accounts(company_name="ALMAITANA DE LUZ, S.L."):
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db = 'Real'
    username = 'jcoronado@optimaluz.com'
    password = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid = common.authenticate(db, username, password, {})
    if not uid:
        print("❌ Error autenticación")
        return []

    models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    accounts = models.execute_kw(
        db, uid, password,
        'res.partner.bank', 'search_read',
        [[]],
        {
            'fields': [
                'acc_number',
                'partner_id',
                'allow_out_payment'
            ]
        }
    )

    resultado = []
    for acc in accounts:
        if not acc.get('partner_id'):
            continue

        resultado.append({
            'acc_number': acc['acc_number'],
            'partner_name': acc['partner_id'][1],
            'allow_out_payment': acc.get('allow_out_payment', False)
        })

    print(f"📤 Exportadas {len(resultado)} cuentas bancarias de contactos")
    return resultado


def import_partner_bank_accounts(accounts):
    from App_Connection import db, uid, password, models

    creadas = 0
    existentes = 0
    errores = 0

    for acc in accounts:
        acc_number = acc['acc_number']
        partner_name = acc['partner_name']
        allow_out = acc.get('allow_out_payment', False)

        print(f"\n🏦 Cuenta {acc_number} → {partner_name}")

        # 1️⃣ Buscar partner por name
        partner = models.execute_kw(
            db, uid, password,
            'res.partner', 'search_read',
            [[('name', '=', partner_name)]],
            {'fields': ['id'], 'limit': 1}
        )

        if not partner:
            print(f"❌ Partner no encontrado: {partner_name}")
            errores += 1
            continue

        partner_id = partner[0]['id']

        # 2️⃣ Evitar duplicados por número de cuenta
        existing = models.execute_kw(
            db, uid, password,
            'res.partner.bank', 'search',
            [[('acc_number', '=', acc_number)]]
        )

        if existing:
            print("⚠️ Cuenta ya existente, saltando")
            existentes += 1
            continue

        vals = {
            'acc_number': acc_number,
            'partner_id': partner_id,
            'allow_out_payment': allow_out,
        }

        try:
            models.execute_kw(
                db, uid, password,
                'res.partner.bank', 'create',
                [vals]
            )
            print("✅ Cuenta bancaria creada")
            creadas += 1
        except Exception as e:
            print(f"❌ Error creando cuenta {acc_number}: {e}")
            errores += 1

    print("\n📊 RESULTADO CUENTAS BANCARIAS")
    print(f"   Creadas: {creadas}")
    print(f"   Existentes: {existentes}")
    print(f"   Errores: {errores}")


# accounts = export_partner_bank_accounts()
# import_partner_bank_accounts(accounts)

# ACTUALIZAR PAYMENT TERMS
# ---------------------------------------------------------
# 1) Exportar payment terms (Odoo origen)
# ---------------------------------------------------------
def export_payment_terms():
    """
    Devuelve una lista de payment terms desde el Odoo origen.
    """
    from App_Connection import db, uid, password, models
    try:
        FIELDS = ["id", "name", "note", "active"]
        pts = models.execute_kw(
            db, uid, password,
            "account.payment.term", "search_read",
            [[]], {"fields": FIELDS}
        )
        print(f"✅ Exportados {len(pts)} payment terms")
        return pts
    except Exception as e:
        print(f"❌ Error al exportar payment terms: {e}")
        return []

# ---------------------------------------------------------
# 2) Importar payment terms (Odoo destino)
# ---------------------------------------------------------
def import_payment_terms(pts):
    """
    Importa en el Odoo destino los payment terms exportados.
    Devuelve un diccionario {old_id: new_id}.
    """
    url = puerto
    db = database
    username = "admin"
    password = "admin"

    try:
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        uid = common.authenticate(db, username, password, {})
        if not uid:
            print("❌ Error de autenticación en Odoo destino")
            return {}

        models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")

        id_map = {}

        for pt in pts:
            try:
                # Buscar si ya existe por nombre
                existing = models.execute_kw(
                    db, uid, password,
                    "account.payment.term", "search",
                    [[["name", "=", pt["name"]]]]
                )
                if existing:
                    new_id = existing[0]
                    print(f"ℹ️ Ya existe: {pt['name']} (id {new_id})")
                else:
                    vals = {
                        "name": pt["name"],
                        "note": pt.get("note", ""),
                        "active": pt.get("active", True),
                    }
                    new_id = models.execute_kw(
                        db, uid, password,
                        "account.payment.term", "create",
                        [vals]
                    )
                    print(f"✅ Creado: {pt['name']} (id {new_id})")

                id_map[pt["id"]] = new_id

            except Exception as e:
                print(f"❌ Error al importar {pt.get('name')}: {e}")

        return id_map

    except Exception as e:
        print(f"❌ Error general en importación de payment terms: {e}")
        return {}

# ---------------------------------------------------------
# 3) Exportar contactos con payment term (Odoo origen)
# ---------------------------------------------------------
def export_contacts_with_pt():
    """
    Devuelve los contactos que tienen un payment term asignado.
    """
    from App_Connection import models, db, uid, password
    try:
        FIELDS = ["id", "name", "vat", "property_payment_term_id"]
        contacts = models.execute_kw(
            db, uid, password,
            "res.partner", "search_read",
            [[["property_payment_term_id", "!=", False]]],
            {"fields": FIELDS}
        )
        print(f"✅ Exportados {len(contacts)} contactos con payment term")
        return contacts
    except Exception as e:
        print(f"❌ Error al exportar contactos con payment term: {e}")
        return []

# ---------------------------------------------------------
# 4) Importar contactos con payment term en Odoo destino
# ---------------------------------------------------------
def import_contacts_with_pt(contacts, pt_id_map):
    """
    Actualiza contactos en Odoo destino para asignarles el payment term correcto.
    - contacts: lista exportada
    - pt_id_map: dict con {old_pt_id: new_pt_id}
    """
    url = puerto
    db = database
    username = "admin"
    password = "admin"

    try:
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        uid = common.authenticate(db, username, password, {})
        if not uid:
            print("❌ Error de autenticación en Odoo destino")
            return

        models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")

        for contact in contacts:
            try:
                # Buscar contacto por VAT o nombre
                if contact.get("vat"):
                    domain = [["vat", "=", contact["vat"]]]
                else:
                    domain = [["name", "=", contact["name"]]]

                existing = models.execute_kw(
                    db, uid, password,
                    "res.partner", "search",
                    [domain], {"limit": 1}
                )

                if not existing:
                    print(f"⚠️ No encontrado en destino: {contact['name']} (VAT {contact.get('vat')})")
                    continue

                dest_id = existing[0]
                old_pt = contact["property_payment_term_id"][0]
                new_pt = pt_id_map.get(old_pt)

                if not new_pt:
                    print(f"⚠️ No hay mapping para payment term {old_pt} → {contact['name']}")
                    continue

                # Actualizar
                models.execute_kw(
                    db, uid, password,
                    "res.partner", "write",
                    [[dest_id], {"property_payment_term_id": new_pt}]
                )
                print(f"✅ Actualizado {contact['name']} con payment term {new_pt}")

            except Exception as e:
                print(f"❌ Error al actualizar {contact['name']}: {e}")

    except Exception as e:
        print(f"❌ Error general en importación de contactos con payment term: {e}")

# 1. Exportar payment terms
#pts = export_payment_terms()

# 2. Importarlos en destino
#pt_map = import_payment_terms(pts)

# 3. Exportar contactos con payment term en origen
#contacts = export_contacts_with_pt()

# 4. Importarlos en destino (usando el mapping)
#import_contacts_with_pt(contacts, pt_map)#'''
# endregion

# region REGION VENDOR PRICELISTS

def export_vendor_pricelists_by_partner():
    """
    Exporta product.supplierinfo usando:
      • partner_key   → VAT o, si falta, nombre del partner
      • product_code  → default_code o name de la plantilla
      • variant_code  → default_code o name de la variante
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models  # conexión ORIGEN

    SUP_FIELDS = [
        "id", "partner_id",
        "product_tmpl_id", "product_id",
        "currency_id", "company_id",
        "price", "min_qty", "delay",
        "date_start", "date_end",
        "product_uom", "sequence",
        "product_code", "product_name"
    ]
    rows = models.execute_kw(
        db, uid, password,
        "product.supplierinfo", "search_read",
        [[]], {"fields": SUP_FIELDS}
    )

    # ---------- Partners (VAT / name) ---------------------------------------
    p_ids = {r["partner_id"][0] for r in rows}
    p_recs = models.execute_kw(
        db, uid, password,
        "res.partner", "read",
        [list(p_ids)], {"fields": ["id", "vat", "name"]}
    )
    pinfo = {p["id"]: p for p in p_recs}

    # ---------- Product codes ------------------------------------------------
    tmpl_ids = {r["product_tmpl_id"][0] for r in rows if r["product_tmpl_id"]}
    variant_ids = {r["product_id"][0] for r in rows if r["product_id"]}

    tmpl_code = {}
    if tmpl_ids:
        for rec in models.execute_kw(
                db, uid, password,
                "product.template", "read",
                [list(tmpl_ids)], {"fields": ["id", "default_code", "name"]}
        ):
            tmpl_code[rec["id"]] = rec["default_code"] or rec["name"]

    var_code = {}
    if variant_ids:
        for rec in models.execute_kw(
                db, uid, password,
                "product.product", "read",
                [list(variant_ids)], {"fields": ["id", "default_code", "name"]}
        ):
            var_code[rec["id"]] = rec["default_code"] or rec["name"]

    # ---------- Resultado ----------------------------------------------------
    export = []
    for r in rows:
        p = pinfo[r["partner_id"][0]]
        export.append({
            "id": r["id"],
            "partner_key": (p["vat"] or p["name"]).strip(),
            "product_tmpl_id": r["product_tmpl_id"] and tmpl_code.get(r["product_tmpl_id"][0]),
            "variant_code": r["product_id"] and var_code.get(r["product_id"][0]),
            "currency_id": r["currency_id"] and r["currency_id"][0],
            "company_id": r["company_id"] and r["company_id"][0],
            "price": r["price"],
            "min_qty": r["min_qty"],
            "delay": r["delay"],
            "date_start": r["date_start"],
            "date_end": r["date_end"],
            "product_uom": r["product_uom"] and r["product_uom"][0],
            "sequence": r["sequence"],
            "product_code": r["product_code"],
            "product_name": r["product_name"]
        })

    return export


def import_vendor_pricelists_by_partner(
        supplier_infos,
        *,
        url="http://79.72.62.113:8069/",
        db="odoo0",
        username="admin",
        password="admin"):
    """
    Importa product.supplierinfo localizando partner y producto por claves estables.
    """
    import xmlrpc.client

    # ---------- Conexión -----------------------------------------------------
    common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
    uid = common.authenticate(db, username, password, {})
    if not uid:
        raise RuntimeError("❌ Autenticación fallida")
    models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")

    main_company_id = models.execute_kw(
        db, uid, password,
        "res.company", "search",
        [[]], {"limit": 1}
    )[0]

    eur_id_list = models.execute_kw(
        db, uid, password,
        "res.currency", "search",
        [[("name", "=", "EUR")]], {"limit": 1}
    )
    if not eur_id_list:
        raise RuntimeError("❌ No se encontró la moneda 'EUR' en la base destino")
    EUR_ID = eur_id_list[0]

    # ---------- Índice de partners (VAT / name) ------------------------------
    partner_idx = {}
    for p in models.execute_kw(
            db, uid, password,
            "res.partner", "search_read",
            [[]], {"fields": ["id", "vat", "name"]}
    ):
        if p["vat"]:
            partner_idx[p["vat"].strip().upper()] = p["id"]
        partner_idx.setdefault(p["name"].strip(), p["id"])

    # ---------- Índices de productos y variantes -----------------------------
    tmpl_idx, var_idx = {}, {}
    for rec in models.execute_kw(
            db, uid, password,
            "product.template", "search_read",
            [[]], {"fields": ["id", "default_code", "name"]}
    ):
        if rec["default_code"]:
            tmpl_idx[rec["default_code"].strip()] = rec["id"]
        tmpl_idx.setdefault(rec["name"].strip(), rec["id"])

    for rec in models.execute_kw(
            db, uid, password,
            "product.product", "search_read",
            [[]], {"fields": ["id", "default_code", "name"]}
    ):
        if rec["default_code"]:
            var_idx[rec["default_code"].strip()] = rec["id"]
        var_idx.setdefault(rec["name"].strip(), rec["id"])

    # ---------- Deduplicación existente --------------------------------------
    existing = models.execute_kw(
        db, uid, password,
        "product.supplierinfo", "search_read",
        [[]], {"fields": ["id", "partner_id", "product_tmpl_id",
                          "product_id", "min_qty",
                          "date_start", "date_end", "price"]}
    )

    def line_key(l):
        return (l["partner_id"][0],
                l["product_tmpl_id"][0] if l["product_tmpl_id"] else None,
                l["product_id"][0] if l["product_id"] else None,
                l["min_qty"], l["date_start"], l["date_end"], l["price"])

    existing_map = {line_key(l): l["id"] for l in existing}

    # ---------- Crear líneas -------------------------------------------------
    old2new, skipped = {}, 0

    for s in supplier_infos:
        # Partner
        pk = s["partner_key"].strip()
        partner_new = partner_idx.get(pk.upper()) or partner_idx.get(pk)
        if not partner_new:
            skipped += 1
            continue

        # Producto / variante
        tmpl_new = s["product_tmpl_id"] and (
                tmpl_idx.get(s["product_tmpl_id"].strip()) or
                tmpl_idx.get(s["product_tmpl_id"].strip().upper())
        )
        var_new = s["variant_code"] and (
                var_idx.get(s["variant_code"].strip()) or
                var_idx.get(s["variant_code"].strip().upper())
        )
        if not tmpl_new and not var_new:
            skipped += 1
            continue

        key = (partner_new, tmpl_new, var_new,
               s["min_qty"], s["date_start"], s["date_end"], s["price"])

        if key in existing_map:
            new_id = existing_map[key]
        else:
            vals = _none_to_false({
                "partner_id": partner_new,
                "product_tmpl_id": tmpl_new or False,
                "product_id": var_new or False,
                "currency_id": EUR_ID,  # s["currency_id"] or False,
                "company_id": main_company_id if s["company_id"] else False,
                "price": s["price"],
                "min_qty": s["min_qty"],
                "delay": s["delay"],
                "date_start": s["date_start"],
                "date_end": s["date_end"],
                "product_uom": s["product_uom"] or False,
                "sequence": s["sequence"],
                "product_code": s["product_code"],
                "product_name": s["product_name"]
            })
            new_id = models.execute_kw(
                db, uid, password,
                "product.supplierinfo", "create", [vals]
            )
            existing_map[key] = new_id

        old2new[s["id"]] = new_id

    print(f"✅ Importadas {len(old2new)} líneas   ⏭️  Omitidas {skipped}.")
    return old2new


# ---------- Utilidad común --------------------------------------------------
def _none_to_false(vals):
    return {k: (False if v is None else v) for k, v in vals.items()}

# 1) ORIGEN
# vendor_data = export_vendor_pricelists_by_partner()

# 2) DESTINO
# vendor_line_map = import_vendor_pricelists_by_partner(vendor_data)

# endregion

#region REGION VENTAS

def export_sale_orders_by_state(state, company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta los pedidos de venta de Odoo 16 filtrados por estado.
    Optimizado para rendimiento: lee todas las líneas en bloque.
    """
    import xmlrpc.client
    from collections import defaultdict

    url = 'https://optimaluz.soluntec.net'  # 'http://79.72.55.217:8069'
    db_old = 'Real'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})

    if not uid_old:
        print("❌ No se pudo autenticar.")
        return

    print(f'🔌 Conectado como {username} (uid: {uid_old})')
    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    FIELDS = [
        "name",
        "partner_id",
        "date_order",
        "currency_id",
        "state",
        "order_line",
        "amount_untaxed",
        "amount_tax",
        "amount_total",
        "user_id",
        "note",
        "origin",
        "company_id",
        "x_comentarios",
        "client_order_ref"
    ]

    print(f"📤 Exportando pedidos de venta en estado '{state}'...")

    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # 1️⃣ Exportar pedidos
    sale_orders = models_old.execute_kw(
        db_old, uid_old, password_old,
        "sale.order", "search_read",
        [[("state", "=", state), ('company_id', '=', company_id_src)]],
        {"fields": FIELDS}
    )

    print(f"   → {len(sale_orders)} pedidos encontrados.")

    # 2️⃣ Reunir todos los IDs de líneas
    all_line_ids = []
    for so in sale_orders:
        all_line_ids.extend(so.get("order_line", []))

    if not all_line_ids:
        print("⚠️  No se encontraron líneas de pedido.")
        return sale_orders

    print(f"   → {len(all_line_ids)} líneas totales detectadas. Leyendo en bloque...")

    # 3️⃣ Leer todas las líneas en bloque
    lines = models_old.execute_kw(
        db_old, uid_old, password_old,
        "sale.order.line", "read",
        [all_line_ids],
        {"fields": [
            "order_id",
            "name",
            "product_id",
            "product_uom_qty",
            "price_unit",
            "price_subtotal",
            "price_total",
            "tax_id",
            "discount"
        ]}
    )

    # Leer SKUs de productos
    product_ids = {l["product_id"][0] for l in lines if l.get("product_id")}
    products = models_old.execute_kw(
        db_old, uid_old, password_old,
        "product.product", "read",
        [list(product_ids)],
        {"fields": ["id", "default_code"]}
    )

    sku_map = {p["id"]: p["default_code"] for p in products}

    for line in lines:
        pid = line.get("product_id")
        line["default_code"] = sku_map.get(pid[0]) if pid else None

    # 4️⃣ Reunir todos los IDs de impuestos de las líneas
    all_tax_ids = set()
    for line in lines:
        for tax_id in line.get("tax_id", []):
            all_tax_ids.add(tax_id)

    tax_name_map = {}
    if all_tax_ids:
        print(f"   → Leyendo {len(all_tax_ids)} impuestos únicos en bloque...")
        taxes = models_old.execute_kw(
            db_old, uid_old, password_old,
            "account.tax", "read",
            [list(all_tax_ids)],
            {"fields": ["id", "name"]}
        )
        for t in taxes:
            tax_name_map[t["id"]] = t["name"]

    # 5️⃣ Añadir nombres de impuestos a cada línea
    for line in lines:
        line["taxes_names"] = [tax_name_map[t] for t in line.get("tax_id", []) if t in tax_name_map]

    # 6️⃣ Agrupar las líneas por pedido
    grouped_lines = defaultdict(list)
    for line in lines:
        order = line.get("order_id")
        if order:
            grouped_lines[order[0]].append(line)

    # 7️⃣ Asignar líneas a cada pedido
    for so in sale_orders:
        so_id = so["id"]
        so["lineas_detalle"] = grouped_lines.get(so_id, [])

    print("✅ Líneas asignadas correctamente a cada pedido.")
    return sale_orders


def import_sale_orders_with_lines(sale_orders, state):
    """
    Importa pedidos de venta en Odoo 18 junto con sus líneas.
    Incluye los impuestos (buscados por nombre).
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    total_creados = 0
    total_existentes = 0

    for so in sale_orders:
        name = so["name"]
        print(f"\n🧾 Procesando pedido: {name}")

        # Comprobar si ya existe
        existing = models.execute_kw(
            db, uid, password,
            "sale.order", "search",
            [[("name", "=", name)]]
        )
        if existing:
            print(f"⚠️  Pedido ya existente: {name}")
            total_existentes += 1
            continue

        # -------------------------------
        # Buscar cliente
        # -------------------------------
        partner_id = None
        if so["partner_id"]:
            partner_name = so["partner_id"][1]
            partners = models.execute_kw(
                db, uid, password,
                "res.partner", "search_read",
                [[("name", "=", partner_name)]],
                {"fields": ["id"], "limit": 1}
            )
            if partners:
                partner_id = partners[0]["id"]
            elif partner_name.startswith("(") and ")" in partner_name:
                partner_name = partner_name.split(")", 1)[1].strip()
                partners = models.execute_kw(
                    db, uid, password,
                    "res.partner", "search_read",
                    [[("name", "=", partner_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if partners:
                    partner_id = partners[0]["id"]
                else:
                    print(f"⚠️  Partner no encontrado: {partner_name}")

        # -------------------------------
        # Buscar moneda
        # -------------------------------
        currency_id = None
        if so["currency_id"]:
            currency_name = so["currency_id"][1]
            currencies = models.execute_kw(
                db, uid, password,
                "res.currency", "search_read",
                [[("name", "=", currency_name)]],
                {"fields": ["id"], "limit": 1}
            )
            if currencies:
                currency_id = currencies[0]["id"]

        # -------------------------------
        # Crear pedido base
        # -------------------------------
        vals_so = {
            "name": name,
            "partner_id": partner_id,
            "date_order": so.get("date_order"),
            "currency_id": currency_id,
            "origin": so.get("origin"),
            "note": so.get("note"),
            "x_comentarios": so.get("x_comentarios"),
            "client_order_ref": so.get("client_order_ref"),
            "state": state,
        }

        try:
            new_so_id = models.execute_kw(
                db, uid, password,
                "sale.order", "create", [vals_so]
            )
            print(f"✅ Pedido creado: {name} (ID {new_so_id})")
            total_creados += 1
        except Exception as e:
            print(f"❌ Error creando pedido {name}: {e}")
            continue

        # -------------------------------
        # Crear líneas
        # -------------------------------
        for linea in so.get("lineas_detalle", []):
            # -------------------------------
            # Producto (por SKU)
            # -------------------------------
            product_id = None
            sku = linea.get("default_code")

            if sku:
                productos = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[("default_code", "=", sku)]],
                    {"limit": 1}
                )
                product_id = productos[0] if productos else None

            # Buscar impuestos por nombre
            impuestos_ids = []
            for tax_name in linea.get("taxes_names", []):
                try:
                    # Extraer número antes del símbolo %
                    match = re.search(r"(\d+(?:\.\d+)?)\s*%", tax_name)
                    porcentaje = match.group(1) if match else None

                    if porcentaje:
                        print(f"🔍 Buscando impuesto con porcentaje exacto: {porcentaje}%")

                        # Obtener todos los impuestos (solo una vez podrías cachearlo fuera del bucle)
                        all_taxes = models.execute_kw(
                            db, uid, password,
                            "account.tax", "search_read",
                            [[], ["id", "name", "description"]],
                            {"limit": 200}  # puedes quitar el limit si tienes pocos
                        )

                        # Buscar coincidencia exacta del número antes del %
                        pattern = rf"(?<!\d){porcentaje}\s*%(\D|$)"  # evita 10 dentro de 210 o 100

                        coincidencias = [
                            t for t in all_taxes
                            if
                            (re.search(pattern, t["name"] or "") or re.search(pattern, t["description"] or ""))
                        ]

                        if coincidencias:
                            impuestos_ids.append(coincidencias[0]["id"])
                            print(f"✅ Impuesto asignado ({porcentaje}%): {coincidencias[0]['name']}")
                        else:
                            print(f"⚠️ No se encontró impuesto exacto con {porcentaje}%")

                    else:
                        print(f"⚠️ No se detectó porcentaje en '{tax_name}'")

                except Exception as e:
                    print(f"❌ Error procesando '{tax_name}': {e}")

            vals_line = {
                "order_id": new_so_id,
                "name": linea.get("name"),
                "product_id": product_id,
                "product_uom_qty": linea.get("product_uom_qty") or 1.0,
                "price_unit": linea.get("price_unit") or 0.0,
                "discount": linea.get("discount", 0.0),
                "tax_id": [(6, 0, impuestos_ids)],
            }

            try:
                models.execute_kw(
                    db, uid, password,
                    "sale.order.line", "create", [vals_line]
                )
                print(f"   ➕ Línea creada: {linea.get('name')}")
            except Exception as e:
                print(f"   ⚠️ Error creando línea: {e}")

    print("\n📊 MIGRACIÓN COMPLETADA")
    print(f"   Total creados: {total_creados}")
    print(f"   Ya existentes: {total_existentes}")


# ----------------------------------------------------------------------
# Funciones por estado
# ----------------------------------------------------------------------

def migrar_pedidos_venta_draft():
    orders = export_sale_orders_by_state("draft")
    import_sale_orders_with_lines(orders, "draft")


def migrar_pedidos_venta_sale():
    orders = export_sale_orders_by_state("sale")
    import_sale_orders_with_lines(orders, "sale")


def migrar_pedidos_venta_cancel():
    orders = export_sale_orders_by_state("cancel")
    import_sale_orders_with_lines(orders, "cancel")


# ----------------------------------------------------------------------
# Función principal opcional
# ----------------------------------------------------------------------

def migrar_pedidos_venta():
    migrar_pedidos_venta_draft()
    migrar_pedidos_venta_sale()
    migrar_pedidos_venta_cancel()

#endregion

#region REGION COMPRAS
def export_purchase_orders_by_state(state, company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta los pedidos de compra de Odoo 16 filtrados por estado.
    Optimizado para rendimiento: lee las líneas y los impuestos en bloque.
    """
    import xmlrpc.client
    from collections import defaultdict

    url = 'https://optimaluz.soluntec.net'
    db_old = 'Real'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})

    if not uid_old:
        print("❌ No se pudo autenticar.")
        return

    print(f'🔌 Conectado como {username} (uid: {uid_old})')
    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    FIELDS = [
        "name",
        "partner_id",
        "partner_ref",
        "date_order",
        "currency_id",
        "state",
        "order_line",
        "amount_untaxed",
        "amount_tax",
        "amount_total",
        "user_id",
        "notes",
        "origin",
        "company_id",
        "x_comentarios"
    ]

    print(f"📤 Exportando pedidos de compra en estado '{state}'...")

    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # 1️⃣ Exportar pedidos
    purchase_orders = models_old.execute_kw(
        db_old, uid_old, password_old,
        "purchase.order", "search_read",
        [[("state", "=", state), ('company_id', '=', company_id_src)]],
        {"fields": FIELDS}
    )

    print(f"   → {len(purchase_orders)} pedidos encontrados.")

    # 2️⃣ Reunir todos los IDs de líneas
    all_line_ids = []
    for po in purchase_orders:
        all_line_ids.extend(po.get("order_line", []))

    if not all_line_ids:
        print("⚠️  No se encontraron líneas de pedido.")
        return purchase_orders

    print(f"   → {len(all_line_ids)} líneas totales detectadas. Leyendo en bloque...")

    # 3️⃣ Leer todas las líneas en bloque
    lines = models_old.execute_kw(
        db_old, uid_old, password_old,
        "purchase.order.line", "read",
        [all_line_ids],
        {"fields": [
            "order_id",
            "name",
            "product_id",
            "product_qty",
            "price_unit",
            "price_subtotal",
            "price_total",
            "taxes_id",
            "date_planned"
        ]}
    )

    # 4️⃣ Reunir todos los IDs de impuestos únicos
    all_tax_ids = set()
    for line in lines:
        for tax_id in line.get("taxes_id", []):
            all_tax_ids.add(tax_id)

    tax_name_map = {}
    if all_tax_ids:
        print(f"   → Leyendo {len(all_tax_ids)} impuestos únicos en bloque...")
        taxes = models_old.execute_kw(
            db_old, uid_old, password_old,
            "account.tax", "read",
            [list(all_tax_ids)],
            {"fields": ["id", "name"]}
        )
        for t in taxes:
            tax_name_map[t["id"]] = t["name"]

    # 5️⃣ Añadir nombres de impuestos a cada línea
    for line in lines:
        line["taxes_names"] = [tax_name_map[t] for t in line.get("taxes_id", []) if t in tax_name_map]

    # 6️⃣ Agrupar las líneas por pedido
    grouped_lines = defaultdict(list)
    for line in lines:
        order = line.get("order_id")
        if order:
            grouped_lines[order[0]].append(line)

    # 7️⃣ Asignar líneas a cada pedido
    for po in purchase_orders:
        po_id = po["id"]
        po["lineas_detalle"] = grouped_lines.get(po_id, [])

    print("✅ Líneas asignadas correctamente a cada pedido.")
    return purchase_orders

def import_purchase_orders_with_lines(purchase_orders, state):
    """
    Importa pedidos de compra en Odoo 18 junto con sus líneas.
    Incluye los impuestos (buscados por nombre).
    """
    from App_Connection import db, uid, password, models

    total_creados = 0
    total_existentes = 0

    for po in purchase_orders:
        name = po["name"]
        print(f"\n📦 Procesando pedido: {name}")

        # Comprobar si ya existe
        existing = models.execute_kw(
            db, uid, password,
            "purchase.order", "search",
            [[("name", "=", name)]]
        )
        if existing:
            print(f"⚠️  Pedido ya existente: {name}")
            total_existentes += 1
            continue

        # -------------------------------
        # Buscar proveedor
        # -------------------------------
        partner_id = None
        if po["partner_id"]:
            partner_name = po["partner_id"][1]
            partners = models.execute_kw(
                db, uid, password,
                "res.partner", "search_read",
                [[("name", "=", partner_name)]],
                {"fields": ["id"], "limit": 1}
            )
            if partners:
                partner_id = partners[0]["id"]
            elif partner_name.startswith("(") and ")" in partner_name:
                partner_name = partner_name.split(")", 1)[1].strip()
                partners = models.execute_kw(
                    db, uid, password,
                    "res.partner", "search_read",
                    [[("name", "=", partner_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if partners:
                    partner_id = partners[0]["id"]
                else:
                    print(f"⚠️  Partner no encontrado: {partner_name}")

        # -------------------------------
        # Buscar moneda
        # -------------------------------
        currency_id = None
        if po["currency_id"]:
            currency_name = po["currency_id"][1]
            currencies = models.execute_kw(
                db, uid, password,
                "res.currency", "search_read",
                [[("name", "=", currency_name)]],
                {"fields": ["id"], "limit": 1}
            )
            if currencies:
                currency_id = currencies[0]["id"]

        # -------------------------------
        # Crear pedido base
        # -------------------------------
        vals_po = {
            "name": name,
            "partner_id": partner_id,
            "partner_ref": po.get("partner_ref"),
            "date_order": po.get("date_order"),
            "currency_id": currency_id,
            "origin": po.get("origin"),
            "notes": po.get("notes"),
            "state": state,
            "x_comentarios": po.get("x_comentarios"),
        }

        try:
            new_po_id = models.execute_kw(
                db, uid, password,
                "purchase.order", "create", [vals_po]
            )
            print(f"✅ Pedido creado: {name} (ID {new_po_id})")
            total_creados += 1
        except Exception as e:
            print(f"❌ Error creando pedido {name}: {e}")
            continue

        # -------------------------------
        # Crear líneas
        # -------------------------------
        for linea in po.get("lineas_detalle", []):
            product_id = None
            if linea.get("product_id"):
                prod_name = linea["product_id"][1]
                productos = models.execute_kw(
                    db, uid, password,
                    "product.product", "search_read",
                    [[("name", "=", prod_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if productos:
                    product_id = productos[0]["id"]

            # Buscar impuestos por nombre
            impuestos_ids = []
            for tax_name in linea.get("taxes_names", []):
                try:
                    # Extraer número antes del símbolo %
                    match = re.search(r"(\d+(?:\.\d+)?)\s*%", tax_name)
                    porcentaje = match.group(1) if match else None

                    if porcentaje:
                        print(f"🔍 Buscando impuesto con porcentaje exacto: {porcentaje}%")

                        # Obtener todos los impuestos (solo una vez podrías cachearlo fuera del bucle)
                        all_taxes = models.execute_kw(
                            db, uid, password,
                            "account.tax", "search_read",
                            [[], ["id", "name", "description"]],
                            {"limit": 200}  # puedes quitar el limit si tienes pocos
                        )

                        # Buscar coincidencia exacta del número antes del %
                        pattern = rf"(?<!\d){porcentaje}\s*%(\D|$)"  # evita 10 dentro de 210 o 100

                        coincidencias = [
                            t for t in all_taxes
                            if
                            (re.search(pattern, t["name"] or "") or re.search(pattern, t["description"] or ""))
                        ]

                        if coincidencias:
                            impuestos_ids.append(coincidencias[0]["id"])
                            print(f"✅ Impuesto asignado ({porcentaje}%): {coincidencias[0]['name']}")
                        else:
                            print(f"⚠️ No se encontró impuesto exacto con {porcentaje}%")

                    else:
                        print(f"⚠️ No se detectó porcentaje en '{tax_name}'")

                except Exception as e:
                    print(f"❌ Error procesando '{tax_name}': {e}")

            vals_line = {
                "order_id": new_po_id,
                "name": linea.get("name"),
                "product_id": product_id,
                "product_qty": linea.get("product_qty") or 1.0,
                "price_unit": linea.get("price_unit") or 0.0,
                "date_planned": linea.get("date_planned"),
                "taxes_id": [(6, 0, impuestos_ids)],
            }

            try:
                models.execute_kw(
                    db, uid, password,
                    "purchase.order.line", "create", [vals_line]
                )
                print(f"   ➕ Línea creada: {linea.get('name')}")
            except Exception as e:
                print(f"   ⚠️ Error creando línea: {e}")

    print("\n📊 MIGRACIÓN COMPLETADA")
    print(f"   Total creados: {total_creados}")
    print(f"   Ya existentes: {total_existentes}")

# ----------------------------------------------------------------------
# Funciones por estado
# ----------------------------------------------------------------------

def migrar_pedidos_compra_draft():
    orders = export_purchase_orders_by_state("draft")
    import_purchase_orders_with_lines(orders, "draft")


def migrar_pedidos_compra_purchase():
    orders = export_purchase_orders_by_state("purchase")
    import_purchase_orders_with_lines(orders, "purchase")


def migrar_pedidos_compra_cancel():
    orders = export_purchase_orders_by_state("cancel")
    import_purchase_orders_with_lines(orders, "cancel")


def migrar_pedidos_compra():
    migrar_pedidos_compra_draft()
    migrar_pedidos_compra_purchase()
    migrar_pedidos_compra_cancel()

#endregion

#region REGION FACTURACION
# ----------------------------------------------------------------------
# Funciones pre-migracion
# ----------------------------------------------------------------------
def export_plan_contable(company_name="ALMAITANA DE LUZ, S.L."):
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db = 'Real'
    username = 'jcoronado@optimaluz.com'
    password = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid = common.authenticate(db, username, password, {})
    if not uid:
        print("❌ Error autenticación")
        return []

    models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    company_id = models.execute_kw(
        db, uid, password,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )[0]

    accounts = models.execute_kw(
        db, uid, password,
        'account.account', 'search_read',
        [[('company_id', '=', company_id)]],
        {
            'fields': [
                'code',
                'name',
                'account_type',
                'reconcile',
                'deprecated'
            ],
            'order': 'code'
        }
    )

    print(f"📤 Exportadas {len(accounts)} cuentas contables")
    return accounts

def import_plan_contable(accounts):
    from App_Connection import db, uid, password, models

    creadas = 0
    existentes = 0

    # Mapeo de tipos (normalizado)
    TYPE_MAP = {
        'asset_receivable': 'asset_receivable',
        'liability_payable': 'liability_payable',
        'asset_cash': 'asset_cash',
        'income': 'income',
        'expense': 'expense',
        'equity': 'equity',
        'asset_current': 'asset_receivable',
        'liability_current': 'liability_payable',
        'asset_current': 'asset_current',
        'asset_fixed': 'asset_current',
        'liability_current': 'liability_payable',
        'liability_non_current': 'liability_payable',
    }

    for acc in accounts:
        code = acc['code']
        name = acc['name']
        acc_type_src = acc['account_type']
        reconcile = acc['reconcile']
        deprecated = acc['deprecated']

        print(f"\n📘 Cuenta {code} – {name}")

        # ¿Existe ya?
        existing = models.execute_kw(
            db, uid, password,
            'account.account', 'search',
            [[('code', '=', code)]]
        )

        if existing:
            print("⚠️ Ya existe")
            existentes += 1
            continue

        acc_type = TYPE_MAP.get(acc_type_src)
        if not acc_type:
            print(f"❌ Tipo no mapeado: {acc_type_src}")
            continue

        # 🔥 Regla obligatoria en Odoo 18
        if acc_type in ('asset_receivable', 'liability_payable'):
            reconcile = True

        vals = {
            'code': code,
            'name': name,
            'account_type': acc_type,
            'reconcile': reconcile,
            'deprecated': deprecated,
        }

        try:
            models.execute_kw(
                db, uid, password,
                'account.account', 'create',
                [vals]
            )
            print("✅ Cuenta creada")
            creadas += 1
        except Exception as e:
            print(f"❌ Error creando cuenta {code}: {e}")

    print("\n📊 PLAN CONTABLE IMPORTADO")
    print(f"   Creadas: {creadas}")
    print(f"   Existentes: {existentes}")

# accounts = export_plan_contable()
# import_plan_contable(accounts)

def limpiar_asientos_automaticos():
    from App_Connection import db, uid, password, models

    print("🧹 Limpiando asientos contables automáticos...")

    # Buscar asientos que NO sean facturas ni pagos
    move_ids = models.execute_kw(
        db, uid, password,
        "account.move", "search",
        [[
            ("state", "=", "posted"),
            ("move_type", "=", "entry"),
        ]]
    )

    print(f"🔎 Asientos encontrados: {len(move_ids)}")

    if not move_ids:
        print("✔️ No hay asientos a borrar")
        return

    try:
        # 1️⃣ Pasar a borrador
        models.execute_kw(
            db, uid, password,
            "account.move", "button_draft",
            [move_ids]
        )
    except Exception as e:
        print("↩️ Asientos pasados a borrador")

    try:
        # 2️⃣ Borrar
        models.execute_kw(
            db, uid, password,
            "account.move", "unlink",
            [move_ids]
        )
    except Exception as e:
        print("🗑️ Asientos eliminados correctamente")

# limpiar_asientos_automaticos()

def export_account_moves_entries(company_name="ALMAITANA DE LUZ, S.L.", limit=None):
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db = 'Real'
    username = 'jcoronado@optimaluz.com'
    password = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid = common.authenticate(db, username, password, {})
    if not uid:
        print("❌ Error autenticación")
        return []

    models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    # Buscar compañía
    company_id = models.execute_kw(
        db, uid, password,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )[0]

    domain = [
        ('company_id', '=', company_id),
        ('move_type', '=', 'entry'),
        ('state', '=', 'posted'),
    ]

    moves = models.execute_kw(
        db, uid, password,
        'account.move', 'search_read',
        [domain],
        {
            'fields': [
                'id', 'name', 'date', 'ref',
                'journal_id', 'line_ids'
            ],
            'limit': limit
        }
    )

    print(f"📤 Exportados {len(moves)} asientos contables")

    # Leer líneas en bloque
    line_ids = [lid for m in moves for lid in m['line_ids']]
    lines = models.execute_kw(
        db, uid, password,
        'account.move.line', 'read',
        [line_ids],
        {
            'fields': [
                'move_id', 'name',
                'account_id', 'partner_id',
                'debit', 'credit'
            ]
        }
    )

    lines_by_move = {}
    for l in lines:
        lines_by_move.setdefault(l['move_id'][0], []).append(l)

    for m in moves:
        m['lines'] = lines_by_move.get(m['id'], [])

    return moves

def import_account_moves_entries(moves):
    from App_Connection import db, uid, password, models

    creados = 0
    saltados = 0

    for m in moves:
        name = m['name']
        print(f"\n📘 Asiento {name}")

        # Evitar duplicados
        if models.execute_kw(db, uid, password, 'account.move', 'search', [[('name', '=', name)]]):
            print("⚠️ Ya existe, saltando")
            saltados += 1
            continue

        # Buscar diario
        journal_name = m['journal_id'][1]
        journal = models.execute_kw(
            db, uid, password,
            'account.journal', 'search_read',
            [[('name', '=', journal_name)]],
            {'fields': ['id'], 'limit': 1}
        )
        if not journal:
            print(f"❌ Diario no encontrado: {journal_name}")
            continue

        move_vals = {
            'name': name,
            'move_type': 'entry',
            'journal_id': journal[0]['id'],
            'date': m['date'],
            'ref': m['ref'],
            'line_ids': [],
        }

        for l in m['lines']:
            # Buscar cuenta contable
            account_full_name = l['account_id'][1]
            account_code = l['account_id'][1].split(' ')[0]
            account = models.execute_kw(
                db, uid, password,
                'account.account', 'search_read',
                [[('code', '=', account_code)]],
                {'fields': ['id'], 'limit': 1}
            )
            if not account:
                print(f"❌ Cuenta no encontrada: {account_full_name}")
                continue

            partner_id = None
            if l['partner_id']:
                partner = models.execute_kw(
                    db, uid, password,
                    'res.partner', 'search_read',
                    [[('name', '=', l['partner_id'][1])]],
                    {'fields': ['id'], 'limit': 1}
                )
                partner_id = partner[0]['id'] if partner else None

            move_vals['line_ids'].append((0, 0, {
                'name': l['name'],
                'account_id': account[0]['id'],
                'partner_id': partner_id,
                'debit': l['debit'],
                'credit': l['credit'],
            }))

        try:
            new_id = models.execute_kw(db, uid, password, 'account.move', 'create', [move_vals])
            models.execute_kw(db, uid, password, 'account.move', 'action_post', [[new_id]])
            print(f"✅ Asiento creado y publicado ({new_id})")
            creados += 1
        except Exception as e:
            print(f"❌ Error creando asiento {name}: {e}")

    print("\n📊 RESULTADO")
    print(f"   Creados: {creados}")
    print(f"   Saltados: {saltados}")

# moves = export_account_moves_entries()
# import_account_moves_entries(moves)

# ----------------------------------------------------------------------
# Migrar Invoices
# ----------------------------------------------------------------------

def export_invoices_by_state(state, cliente_T_proveedor_F = True, company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta facturas de Odoo 16 filtradas por estado, optimizado para rendimiento.
    Lee las líneas en bloque para evitar miles de llamadas RPC.
    """
    import xmlrpc.client
    from collections import defaultdict

    url = 'https://optimaluz.soluntec.net'  # 'http://79.72.55.217:8069'
    db_old = 'Real'  # 'odoo1'
    username = 'jcoronado@optimaluz.com'  # 'admin'
    password_old = 'AlAi4ever'  # 'admin'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})

    if not uid_old:
        print("❌ No se pudo autenticar.")
        return

    print(f'🔌 Conectado como {username} (uid: {uid_old})')
    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    tipo_factura = "out_invoice" if cliente_T_proveedor_F else "in_invoice"

    FIELDS = [
        "name",
        "partner_id",
        "ref",
        "invoice_date",
        "invoice_date_due",
        "move_type",
        "state",
        "invoice_line_ids",
        "amount_untaxed",
        "amount_tax",
        "amount_total",
        "currency_id",
        "payment_reference",
        "invoice_payment_term_id",
        "invoice_origin",
        "narration",
        "company_id",
    ]

    print(f"📤 Exportando facturas en estado '{state}'...")

    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # 1️⃣ Exportar facturas
    invoices = models_old.execute_kw(
        db_old, uid_old, password_old,
        "account.move", "search_read",
        [[
            ("move_type", "in", [tipo_factura]),
            ("state", "=", state), ('company_id', '=', company_id_src),
        ]],
        {"fields": FIELDS}
    )

    print(f"   → {len(invoices)} facturas encontradas.")

    # 2️⃣ Reunir todos los IDs de líneas
    all_line_ids = []
    for inv in invoices:
        all_line_ids.extend(inv.get("invoice_line_ids", []))

    if not all_line_ids:
        print("⚠️  No se encontraron líneas de factura.")
        return invoices

    print(f"   → {len(all_line_ids)} líneas totales detectadas. Leyendo en bloque...")

    # 3️⃣ Leer todas las líneas en bloque
    lines = models_old.execute_kw(
        db_old, uid_old, password_old,
        "account.move.line", "read",
        [all_line_ids],
        {"fields": [
            "move_id",
            "name",
            "product_id",
            "quantity",
            "price_unit",
            "price_subtotal",
            "tax_ids",
            "account_id",
        ]}
    )

    all_tax_ids = set()
    for line in lines:
        for tax_id in line.get("tax_ids", []):
            all_tax_ids.add(tax_id)

    taxes_data = {}
    if all_tax_ids:
        taxes_read = models_old.execute_kw(
            db_old, uid_old, password_old,
            "account.tax", "read",
            [list(all_tax_ids)],
            {"fields": ["id", "name"]}
        )
        taxes_data = {t["id"]: t["name"] for t in taxes_read}

    for line in lines:
        line["taxes_names"] = [taxes_data.get(tid) for tid in line.get("tax_ids", []) if tid in taxes_data]

    # 4️⃣ Agrupar las líneas por factura
    grouped_lines = defaultdict(list)
    for line in lines:
        move = line.get("move_id")
        if move:
            grouped_lines[move[0]].append(line)

    # 5️⃣ Asignar líneas a cada factura
    for inv in invoices:
        inv_id = inv["id"]
        inv["lineas_detalle"] = grouped_lines.get(inv_id, [])

    print("✅ Líneas asignadas correctamente a cada factura.")
    return invoices


def import_invoices_with_lines(invoices, state):
    """
    Importa facturas en Odoo 18 junto con sus líneas y mapeo de impuestos.
    Si state='posted', primero crea en draft y luego ejecuta action_post().
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    total_creadas = 0
    total_existentes = 0

    def safe_execute_line(model, method, args, retries=3, wait=5, timeout=60):
        """
        Ejecuta un método XML-RPC con reintentos y timeout de seguridad.
        Evita bloqueos permanentes por errores de red o cuelgues del servidor.
        """
        import xmlrpc.client
        import time
        import socket

        for intento in range(1, retries + 1):
            try:
                # Ajustar timeout del socket
                socket.setdefaulttimeout(timeout)
                result = models.execute_kw(db, uid, password, model, method, args)
                socket.setdefaulttimeout(None)
                return result
            except Exception as e:
                print(f"⚠️ Error ({model}.{method}) intento {intento}/{retries}: {e}")
                if intento < retries:
                    print(f"   ↪ Reintentando en {wait} segundos...")
                    time.sleep(wait)
                else:
                    print(f"   ❌ Línea fallida tras {retries} intentos.")
                    return None

    for inv in invoices:
        name = inv.get("name") or inv.get("payment_reference") or "SIN_NOMBRE"
        print(f"\n🧾 Procesando factura: {name}")

        # Comprobar si ya existe
        existing = models.execute_kw(
            db, uid, password,
            "account.move", "search",
            [[("name", "=", name)]]
        )
        if existing:
            print(f"⚠️  Factura ya existente: {name}")
            total_existentes += 1
            continue

        # -------------------------------
        # Buscar cliente/proveedor
        # -------------------------------
        partner_id = None
        if inv["partner_id"]:
            partner_name = inv["partner_id"][1]
            partners = models.execute_kw(
                db, uid, password,
                "res.partner", "search_read",
                [[("name", "=", partner_name)]],
                {"fields": ["id"], "limit": 1}
            )
            if partners:
                partner_id = partners[0]["id"]
            elif partner_name.startswith("(") and ")" in partner_name:
                partner_name = partner_name.split(")", 1)[1].strip()
                partners = models.execute_kw(
                    db, uid, password,
                    "res.partner", "search_read",
                    [[("name", "=", partner_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if partners:
                    partner_id = partners[0]["id"]
                else:
                    print(f"⚠️  Partner no encontrado: {partner_name}")

        # -------------------------------
        # Buscar moneda
        # -------------------------------
        currency_id = None
        if inv["currency_id"]:
            currency_name = inv["currency_id"][1]
            currencies = models.execute_kw(
                db, uid, password,
                "res.currency", "search_read",
                [[("name", "=", currency_name)]],
                {"fields": ["id"], "limit": 1}
            )
            if currencies:
                currency_id = currencies[0]["id"]

        # -------------------------------
        # Crear factura en borrador siempre
        # -------------------------------
        vals_inv = {
            "name": name,
            "move_type": inv.get("move_type") or "out_invoice",
            "partner_id": partner_id,
            "ref": inv.get("ref"),
            "payment_reference": inv.get("payment_reference"),
            "invoice_date": inv.get("invoice_date"),
            "invoice_date_due": inv.get("invoice_date_due"),
            "currency_id": currency_id,
            "invoice_origin": inv.get("invoice_origin"),
            "narration": inv.get("narration"),
            "state": "draft",
        }

        try:
            new_inv_id = models.execute_kw(
                db, uid, password,
                "account.move", "create", [vals_inv]
            )
            print(f"✅ Factura creada: {name} (ID {new_inv_id})")
            total_creadas += 1
        except Exception as e:
            print(f"❌ Error creando factura {name}: {e}")
            continue

        # -------------------------------
        # Crear líneas
        # -------------------------------
        for linea in inv.get("lineas_detalle", []):
            print(f"Linea: {linea}")
            product_id = None
            if linea.get("product_id"):
                prod_name = linea["product_id"][1]
                productos = models.execute_kw(
                    db, uid, password,
                    "product.product", "search_read",
                    [[("name", "=", prod_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if productos:
                    product_id = productos[0]["id"]

            # Buscar impuestos por número de porcentaje (%)
            impuestos_ids = []
            for tax_name in linea.get("taxes_names", []):
                try:
                    # Extraer número antes del símbolo %
                    match = re.search(r"(\d+(?:\.\d+)?)\s*%", tax_name)
                    porcentaje = match.group(1) if match else None

                    if porcentaje:
                        print(f"🔍 Buscando impuesto con porcentaje exacto: {porcentaje}%")

                        # Obtener todos los impuestos (solo una vez podrías cachearlo fuera del bucle)
                        all_taxes = models.execute_kw(
                            db, uid, password,
                            "account.tax", "search_read",
                            [[], ["id", "name", "description"]],
                            {"limit": 200}  # puedes quitar el limit si tienes pocos
                        )

                        # Buscar coincidencia exacta del número antes del %
                        pattern = rf"(?<!\d){porcentaje}\s*%(\D|$)"  # evita 10 dentro de 210 o 100

                        coincidencias = [
                            t for t in all_taxes
                            if
                            (re.search(pattern, t["name"] or "") or re.search(pattern, t["description"] or ""))
                        ]

                        if coincidencias:
                            impuestos_ids.append(coincidencias[0]["id"])
                            print(f"✅ Impuesto asignado ({porcentaje}%): {coincidencias[0]['name']}")
                        else:
                            print(f"⚠️ No se encontró impuesto exacto con {porcentaje}%")

                    else:
                        print(f"⚠️ No se detectó porcentaje en '{tax_name}'")

                except Exception as e:
                    print(f"❌ Error procesando '{tax_name}': {e}")

            vals_line = {
                "move_id": new_inv_id,
                "name": linea.get("name"),
                "product_id": product_id,
                "quantity": linea.get("quantity") or 1.0,
                "price_unit": linea.get("price_unit") or 0.0,
                "tax_ids": [(6, 0, impuestos_ids)],
            }

            result = safe_execute_line(
                "account.move.line", "create", [vals_line],
                retries=3, wait=5, timeout=90
            )

            if result:
                print(f"   ➕ Línea creada: {linea.get('name')}")
            else:
                print(f"   ⚠️ No se pudo crear la línea: {linea.get('name')}")

        # -------------------------------
        # Publicar si corresponde
        # -------------------------------
        if state == "posted":
            try:
                models.execute_kw(
                    db, uid, password,
                    "account.move", "action_post",
                    [[new_inv_id]]
                )
                print(f"   📤 Factura publicada correctamente.")
            except Exception as e:
                print(f"   ⚠️ Error al publicar factura {name}: {e}")

        elif state == "cancel":
            try:
                models.execute_kw(
                    db, uid, password,
                    "account.move", "button_cancel",
                    [[new_inv_id]]
                )
                print(f"   🚫 Factura cancelada correctamente.")
            except Exception as e:
                print(f"   ⚠️ Error cancelando factura {name}: {e}")

    print("\n📊 MIGRACIÓN COMPLETADA")
    print(f"   Total creadas: {total_creadas}")
    print(f"   Ya existentes: {total_existentes}")


# ----------------------------------------------------------------------
# Funciones por estado
# ----------------------------------------------------------------------

def migrar_facturas_draft(tipo):
    invoices = export_invoices_by_state("draft", tipo)
    import_invoices_with_lines(invoices, "draft")


def migrar_facturas_posted(tipo):
    invoices = export_invoices_by_state("posted", tipo)
    import_invoices_with_lines(invoices, "posted")


# ----------------------------------------------------------------------
# Principal
# ----------------------------------------------------------------------

def migrar_facturas(tipo):
    migrar_facturas_draft(tipo)
    migrar_facturas_posted(tipo)

#endregion

#region FACTURACION/PAGOS
# ---------------------------------------------------------------
# 🔹 MIGRACIÓN DE PAGOS (Clientes y Proveedores)
# ---------------------------------------------------------------

def export_payments(tipo, company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta pagos (account.payment) de Odoo 16.
    tipo puede ser: "clientes", "proveedores" o "todos".
    """
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db_old = 'Real'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})
    if not uid_old:
        print("❌ No se pudo autenticar en origen.")
        return []

    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
    print(f'🔌 Conectado a Odoo 16 como {username}')

    # Buscar compañía
    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # Campos a exportar
    FIELDS = [
        "name",
        "partner_id",
        "amount",
        "payment_type",
        "partner_type",
        "journal_id",
        "ref",
        "state",
        "company_id",
        "currency_id",
        "date"
    ]

    # Construir dominio según tipo
    domain = [("company_id", "=", company_id_src)]
    if tipo == "clientes":
        domain += [("partner_type", "=", "customer")]
    elif tipo == "proveedores":
        domain += [("partner_type", "=", "supplier")]

    # Buscar pagos
    payments = models_old.execute_kw(
        db_old, uid_old, password_old,
        "account.payment", "search_read",
        [domain],
        {"fields": FIELDS}
    )

    # 🔹 Sustituir journal_id por journal_name
    for pay in payments:
        journal = pay.get("journal_id")
        if isinstance(journal, list) and len(journal) == 2:
            pay["journal_name"] = journal[1]
        else:
            pay["journal_name"] = None
        del pay["journal_id"]  # eliminar el ID, ya no sirve

    print(f"📤 {len(payments)} pagos exportados ({tipo}).")
    return payments

def import_payments(payments):
    """
    Importa los pagos exportados desde Odoo 16 a Odoo 18.
    Compatible con clientes (inbound/customer) y proveedores (outbound/supplier).
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    total_creados = 0
    total_existentes = 0

    def find_partner_id(partner_name):
        partners = models.execute_kw(
            db, uid, password,
            "res.partner", "search_read",
            [[("name", "=", partner_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return partners[0]["id"] if partners else None

    def find_journal_id(journal_name):
        journals_dic = {
            "Facturas de cliente": "Customer Invoices",
            "Vendor Bills": "Facturas de proveedores",
            "Bank": "Banco",
            "Operaciones varias": "Miscellaneous Operations",
            "Efectivo": "Cash",
            "Impuestos de base de Efectivo": "Cash Basis Taxes",
            "Diferencia de cambio": "Exchange Difference",
            "Valoración de inventario": "Inventory Valuation",
        }

        if journals_dic[journal_name]: journal_name = journals_dic[journal_name]

        journals = models.execute_kw(
            db, uid, password,
            "account.journal", "search_read",
            [[("name", "=", journal_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return journals[0]["id"] if journals else None

    def find_currency_id(currency_name):
        currencies = models.execute_kw(
            db, uid, password,
            "res.currency", "search_read",
            [[("name", "=", currency_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return currencies[0]["id"] if currencies else None

    print("📥 Iniciando importación de pagos...")

    for pay in payments:
        name = pay.get("name") or pay.get("ref") or "SIN_REF"
        print(f"\n💳 Procesando pago: {name}")

        # Evitar duplicados
        existing = models.execute_kw(
            db, uid, password,
            "account.payment", "search",
            [[("name", "=", name)]]
        )
        if existing:
            print(f"⚠️  Pago ya existente: {name}")
            total_existentes += 1
            continue

        partner_id = None
        if pay.get("partner_id"):
            partner_name = pay["partner_id"][1]
            partner_id = find_partner_id(partner_name)

        journal_name = pay.get("journal_name")
        journal_id=0
        if journal_name:
            journal_id = find_journal_id(journal_name)

        currency_id = None
        if pay.get("currency_id"):
            currency_name = pay["currency_id"][1]
            currency_id = find_currency_id(currency_name)

        vals = {
            "name": name,
            "payment_type": pay.get("payment_type"),  # inbound / outbound
            "partner_type": pay.get("partner_type"),  # customer / supplier
            "partner_id": partner_id,
            "journal_id": journal_id,
            "amount": pay.get("amount"),
            "memo": pay.get("ref"),
            "currency_id": currency_id,
            "date": pay.get("date")
        }
        new_id = 0
        try:
            new_id = models.execute_kw(db, uid, password, "account.payment", "create", [vals])
            print(f"✅ Pago creado correctamente (ID {new_id})")
            total_creados += 1

            # Si el pago estaba validado en origen, validar también en destino
            if pay.get("state") == "posted":
                # 1️⃣ Validar el pago
                models.execute_kw(db, uid, password, "account.payment", "action_validate", [[new_id]])
                print("   📤 Pago validado/publicado.")
            elif pay.get("state") == "cancel":
                models.execute_kw(db, uid, password, "account.payment", "action_cancel", [[new_id]])
                print("   📤 Pago cancelado.")

        except Exception as e:
            # 2️⃣ Restaurar nombre original (ya que Odoo reasigna la secuencia)
            try:
                models.execute_kw(db, uid, password,
                                  "account.payment", "write",
                                  [[new_id], {"name": name}])
            except Exception as e:
                print(f"   ⚠️ No se pudo restaurar el nombre del pago: {e}")
            print(f"❌ Error creando pago {name}: {e}")

    print("\n📊 MIGRACIÓN DE PAGOS COMPLETADA")
    print(f"   Total creados: {total_creados}")
    print(f"   Ya existentes: {total_existentes}")

# ---------------------------------------------------------------
# 🔸 Función principal
# ---------------------------------------------------------------

def migrar_pagos(tipo):
    pagos = export_payments(tipo)
    import_payments(pagos)

# Llamada principal
#migrar_pagos()
#endregion

# region REGION PROYECTOS
# 🔹 PRE PROYECTOS: Cuentas analíticas
def export_cuentas_analiticas(company_name="ALMAITANA DE LUZ, S.L."):
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db = 'Real'
    username = 'jcoronado@optimaluz.com'
    password = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid = common.authenticate(db, username, password, {})
    if not uid:
        print("❌ Error de autenticación")
        return []

    models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

    company_id = models.execute_kw(
        db, uid, password,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )[0]

    accounts = models.execute_kw(
        db, uid, password,
        'account.analytic.account', 'search_read',
        [[('company_id', '=', company_id)]],
        {
            'fields': [
                'name',
                'code',
                'active',
                'partner_id',
            ],
            'order': 'name'
        }
    )

    print(f"📤 Exportadas {len(accounts)} cuentas analíticas")
    return accounts

def import_cuentas_analiticas(accounts):
    from App_Connection import db, uid, password, models

    creadas = 0
    actualizadas = 0

    # 🔹 company_id fijo
    company = models.execute_kw(
        db, uid, password,
        'res.company', 'search_read',
        [[('name', '=', "ALMAITANA DE LUZ, S.L.")]],
        {'fields': ['id'], 'limit': 1}
    )
    if not company:
        print("❌ Compañía no encontrada")
        return
    company_id = company[0]['id']

    PLAN_ID = 1  # definido por ti

    for acc in accounts:
        name = acc['name']
        code = acc.get('code')

        print(f"\n📊 Cuenta analítica: {name}")

        # Buscar por name + company
        existing = models.execute_kw(
            db, uid, password,
            'account.analytic.account', 'search',
            [[('name', '=', name), ('company_id', '=', company_id)]],
            {'limit': 1}
        )

        partner_id = None
        if acc.get('partner_id'):
            partner_name = acc['partner_id'][1]
            partner = models.execute_kw(
                db, uid, password,
                'res.partner', 'search_read',
                [[('name', '=', partner_name)]],
                {'fields': ['id'], 'limit': 1}
            )
            if partner:
                partner_id = partner[0]['id']

        vals = {
            'name': name,
            'code': code,
            'active': acc.get('active', True),
            'partner_id': partner_id,
            'plan_id': PLAN_ID,
            'company_id': company_id,
        }

        try:
            if existing:
                models.execute_kw(
                    db, uid, password,
                    'account.analytic.account', 'write',
                    [existing, vals]
                )
                print("🔄 Existía → actualizada")
                actualizadas += 1
            else:
                models.execute_kw(
                    db, uid, password,
                    'account.analytic.account', 'create',
                    [vals]
                )
                print("✅ Creada")
                creadas += 1
        except Exception as e:
            print(f"❌ Error procesando '{name}': {e}")

    print("\n📊 RESULTADO CUENTAS ANALÍTICAS")
    print(f"   Creadas: {creadas}")
    print(f"   Actualizadas: {actualizadas}")

# accounts = export_cuentas_analiticas()
# import_cuentas_analiticas(accounts)

# ---------------------------------------------------------------
# 🔹 MIGRACIÓN DE PROYECTOS (project.project)
# ---------------------------------------------------------------

def export_projects(company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta todos los proyectos activos de Odoo 16.
    Incluye campos clave como nombre, responsable, cliente y fechas.
    """
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db_old = 'Test'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})
    if not uid_old:
        print("❌ No se pudo autenticar en origen.")
        return []

    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
    print(f'🔌 Conectado a Odoo origen como {username}')

    # Buscar compañía origen
    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # Campos de interés
    FIELDS = [
        "name",
        "active",
        "stage_id",
        "user_id",
        "partner_id",
        "company_id",
        "date_start",
        "date",
        "description",
        "analytic_account_id",
        "privacy_visibility",  # público, seguidores, portal
    ]

    # Buscar proyectos de esa compañía
    projects = models_old.execute_kw(
        db_old, uid_old, password_old,
        "project.project", "search_read",
        [[("company_id", "=", company_id_src)]],
        {"fields": FIELDS}
    )

    print(f"📤 {len(projects)} proyectos exportados correctamente.")
    return projects


def import_projects(projects):
    """
    Importa los proyectos exportados desde Odoo 16 a Odoo 18.
    Mantiene nombre, responsable, cliente y fechas.
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    STAGE_ID_MAP = {
        5: 1,
        7: 2,
        8: 5,
        9: 3,
        10: 4,
        11: 6,
    }

    total_creados = 0
    total_existentes = 0

    def find_user_id(user_name):
        users = models.execute_kw(
            db, uid, password,
            "res.users", "search_read",
            [[("name", "=", user_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return users[0]["id"] if users else None

    def find_partner_id(partner_name):
        partners = models.execute_kw(
            db, uid, password,
            "res.partner", "search_read",
            [[("name", "=", partner_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return partners[0]["id"] if partners else None

    print("📥 Iniciando importación de proyectos...")

    for proj in projects:
        name = proj.get("name") or "SIN_NOMBRE"
        print(f"\n📁 Procesando proyecto: {name}")

        # Comprobar duplicados
        existing = models.execute_kw(
            db, uid, password,
            "project.project", "search",
            [[("name", "=", name)]]
        )
        if existing:
            print(f"⚠️ Proyecto ya existente: {name}")
            total_existentes += 1
            continue

        # Buscar usuario responsable
        user_id = None
        if proj.get("user_id"):
            user_name = proj["user_id"][1]
            user_id = find_user_id(user_name)

        # Buscar cliente
        partner_id = None
        if proj.get("partner_id"):
            partner_name = proj["partner_id"][1]
            partner_id = find_partner_id(partner_name)

        stage_id = None
        if proj.get("stage_id"):
            stage_id_src = proj["stage_id"][0]
            stage_id = STAGE_ID_MAP.get(stage_id_src)

        # Crear diccionario de valores
        vals = {
            "name": name,
            "active": proj.get("active", True),
            "user_id": user_id,
            "partner_id": partner_id,
            "date_start": proj.get("date_start"),
            "date": proj.get("date"),
            "description": proj.get("description"),
            "privacy_visibility": proj.get("privacy_visibility", "followers"),
            "allow_billable": True,
            "stage_id": stage_id,
        }

        # Limpiar valores nulos (evita TypeError: cannot marshal None)
        vals = {k: v for k, v in vals.items() if v is not None}

        try:
            new_id = models.execute_kw(db, uid, password, "project.project", "create", [vals])
            print(f"✅ Proyecto creado (ID {new_id})")
            total_creados += 1

        except Exception as e:
            print(f"❌ Error creando proyecto {name}: {e}")

    print("\n📊 MIGRACIÓN DE PROYECTOS COMPLETADA")
    print(f"   Total creados: {total_creados}")
    print(f"   Ya existentes: {total_existentes}")


# ---------------------------------------------------------------
# 🔸 Función principal
# ---------------------------------------------------------------

def migrar_proyectos():
    proyectos = export_projects()
    import_projects(proyectos)


# Llamada directa (opcional)
# migrar_proyectos()

# ---------------------------------------------------------------
# 🔹 MIGRACIÓN DE TAREAS (project.task)
# ---------------------------------------------------------------

def export_tasks(company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta todas las tareas activas de Odoo 16 con sus campos clave.
    """
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db_old = 'Test'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})
    if not uid_old:
        print("❌ No se pudo autenticar en origen.")
        return []

    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
    print(f'🔌 Conectado a Odoo origen como {username}')

    # Buscar compañía
    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # Campos relevantes
    FIELDS = [
        "id",
        "name",
        "active",
        "project_id",
        "partner_id",
        "company_id",
        "date_deadline",
        "date_assign",
        "date_last_stage_update",
        "description",
        "stage_id",
        "priority",
        "remaining_hours",
        "total_hours_spent",
        "tag_ids",
        "user_ids"
    ]

    tasks = models_old.execute_kw(
        db_old, uid_old, password_old,
        "project.task", "search_read",
        [[("company_id", "=", company_id_src)]],
        {"fields": FIELDS}
    )

    print(f"📤 {len(tasks)} tareas exportadas correctamente.")
    return tasks


def import_tasks(tasks, adjuntos=True):
    """
    Requisito: Desisntalar modulo website_sale_wishlist
    Importa las tareas desde Odoo 16 a Odoo 18.
    Mantiene nombre, responsable, proyecto, cliente y horas.
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    total_creadas = 0
    total_existentes = 0

    # -------------------------------
    # Helpers
    # -------------------------------
    def find_user_id(user_name):
        users = models.execute_kw(
            db, uid, password,
            "res.users", "search_read",
            [[("name", "=", user_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return users[0]["id"] if users else None

    def find_partner_id(partner_name):
        partners = models.execute_kw(
            db, uid, password,
            "res.partner", "search_read",
            [[("name", "=", partner_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return partners[0]["id"] if partners else None

    def find_project_id(project_name):
        projects = models.execute_kw(
            db, uid, password,
            "project.project", "search_read",
            [[("name", "=", project_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return projects[0]["id"] if projects else None

    print("📥 Iniciando importación de tareas...")

    # -------------------------------
    # Bucle principal
    # -------------------------------
    for task in tasks:
        old_task_id = task.get("id")
        name = task.get("name") or "SIN_NOMBRE"

        print(f"\n🗂️ Procesando tarea ORIGEN ID {old_task_id}: {name}")

        # -------------------------------
        # Comprobación por ID de origen
        # -------------------------------
        existing = models.execute_kw(
            db, uid, password,
            "project.task", "search",
            [[("x_old_task_id", "=", old_task_id)]],
            {"limit": 1}
        )

        if existing:
            print(f"⚠️  Tarea ya importada (x_old_task_id={old_task_id}) → se omite")
            total_existentes += 1
            continue

        # -------------------------------
        # Proyecto
        # -------------------------------
        project_id = None
        if task.get("project_id"):
            project_name = task["project_id"][1]
            project_id = find_project_id(project_name)

        # -------------------------------
        # Usuario responsable
        # -------------------------------
        user_id = None
        if task.get("user_id"):
            user_name = task["user_id"][1]
            user_id = find_user_id(user_name)

        # -------------------------------
        # Cliente
        # -------------------------------
        partner_id = None
        if task.get("partner_id"):
            partner_name = task["partner_id"][1]
            partner_id = find_partner_id(partner_name)

        # -------------------------------
        # Valores de la tarea
        # -------------------------------
        vals = {
            "name": name,
            "active": task.get("active", True),
            "project_id": project_id,
            "user_id": user_id,
            "partner_id": partner_id,
            "date_deadline": task.get("date_deadline"),
            "priority": task.get("priority"),
            "remaining_hours": task.get("remaining_hours"),
            "description": task.get("description"),
            "x_old_task_id": old_task_id,  # 🔑 CLAVE
        }

        # Limpiar None
        vals = {k: v for k, v in vals.items() if v is not None}

        try:
            new_id = models.execute_kw(db, uid, password, "project.task", "create", [vals])
            print(f"✅ Tarea creada (ID {new_id})")
            total_creadas += 1
        except Exception as e:
            print(f"❌ Error creando tarea {name}: {e}")

        if adjuntos:
            try:
                def migrar_adjuntos_tarea(origen_task_id, destino_task_id, lote=30, delay=0.1):
                    """
                    Migra adjuntos de una tarea desde Odoo ORIGEN → DESTINO.

                    origen_task_id : ID de tarea en origen
                    destino_task_id : ID de tarea en destino
                    lote : cantidad de adjuntos por lote
                    delay : pausa entre lotes
                    """
                    import xmlrpc.client
                    import math
                    import time
                    from App_Connection import models, db, uid, password  # DESTINO

                    print(f"\n=== MIGRANDO ADJUNTOS TAREA {origen_task_id} → {destino_task_id} ===")

                    # --------------------------------------------------
                    # 🔹 CONEXIÓN ORIGEN
                    # --------------------------------------------------
                    url = 'https://optimaluz.soluntec.net'
                    db_old = 'Test'
                    username = 'jcoronado@optimaluz.com'
                    password_old = 'AlAi4ever'

                    common_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
                    uid_old = common_old.authenticate(db_old, username, password_old, {})

                    if not uid_old:
                        print("❌ No se pudo autenticar en ORIGEN.")
                        return

                    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
                    print(f"🔌 Conectado a ORIGEN como {username} (uid {uid_old})")

                    # --------------------------------------------------
                    # 🔹 LEER ADJUNTOS EN ORIGEN
                    # --------------------------------------------------
                    attach_ids = models_old.execute_kw(
                        db_old, uid_old, password_old,
                        'ir.attachment', 'search',
                        [[
                            ('res_model', '=', 'project.task'),
                            ('res_id', '=', origen_task_id)
                        ]]
                    )

                    total_adjuntos = len(attach_ids)
                    print(f"📎 Adjuntos encontrados en origen: {total_adjuntos}")

                    if total_adjuntos == 0:
                        print("No hay adjuntos que migrar.")
                        return

                    num_lotes = math.ceil(total_adjuntos / lote)

                    # --------------------------------------------------
                    # 🔹 MIGRAR POR LOTES
                    # --------------------------------------------------
                    for i in range(num_lotes):
                        inicio = i * lote
                        fin = inicio + lote
                        lote_ids = attach_ids[inicio:fin]

                        print(f"\nProcesando lote {i + 1}/{num_lotes} ({len(lote_ids)} adjuntos)...")

                        adjuntos = models_old.execute_kw(
                            db_old, uid_old, password_old,
                            'ir.attachment', 'read',
                            [lote_ids, ['name', 'datas', 'mimetype']]
                        )

                        for att in adjuntos:
                            try:
                                models.execute_kw(
                                    db, uid, password,
                                    'ir.attachment', 'create',
                                    [{
                                        'name': att['name'],
                                        'datas': att['datas'],
                                        'mimetype': att['mimetype'],
                                        'res_model': 'project.task',
                                        'res_id': destino_task_id,
                                    }]
                                )
                            except Exception as e:
                                print(f"⚠️ Error importando adjunto '{att['name']}': {e}")

                        print(f"✓ Lote {i + 1}/{num_lotes} completado.")
                        time.sleep(delay)

                    print("\n=== MIGRACIÓN DE ADJUNTOS COMPLETADA ===")

                migrar_adjuntos_tarea(
                    origen_task_id=task.get("id"),
                    destino_task_id=new_id,
                    lote=10,
                    delay=0.1
                )

            except Exception as e:
                print(f"⚠️ Error migrando adjuntos de la tarea {name}: {e}")

    print("\n📊 MIGRACIÓN DE TAREAS COMPLETADA")
    print(f"   Total creadas: {total_creadas}")
    print(f"   Ya existentes: {total_existentes}")


# ---------------------------------------------------------------
# 🔸 Función principal
# ---------------------------------------------------------------

def migrar_tareas():
    tareas = export_tasks()
    import_tasks(tareas)


# Ejecución directa opcional
# migrar_tareas()

# Vincular Elemento del pedido de venta
def sync_task_sale_line_id():
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    # -----------------------
    # ORIGEN
    # -----------------------
    url = 'https://optimaluz.soluntec.net'
    db_src = 'Test'  # o 'Real'
    username = 'jcoronado@optimaluz.com'
    password_src = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_src = common.authenticate(db_src, username, password_src, {})
    if not uid_src:
        print("❌ Error autenticación ORIGEN")
        return

    models_src = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
    print(f"🔌 Conectado a ORIGEN (uid={uid_src})")

    # -----------------------
    # 1️⃣ RESOLVER PRODUCTO DESTINO (UNA VEZ)
    # -----------------------
    PRODUCT_NAME = "Instalación eléctrica* "  # ← el name que ya sabes que coincide

    product_dst_ids = models.execute_kw(
        db, uid, password,
        "product.product", "search",
        [[("name", "=", PRODUCT_NAME)]],
        {"limit": 1}
    )

    if not product_dst_ids:
        print(f"❌ Producto '{PRODUCT_NAME}' no encontrado en DESTINO")
        return

    product_dst_id = product_dst_ids[0]
    print(f"📦 Producto DESTINO resuelto: {PRODUCT_NAME} (ID {product_dst_id})")

    # -----------------------
    # 2️⃣ TAREAS ORIGEN con sale_line
    # -----------------------
    task_ids_src = models_src.execute_kw(
        db_src, uid_src, password_src,
        "project.task", "search",
        [[("sale_line_id", "!=", False)]]
    )

    tasks_src = models_src.execute_kw(
        db_src, uid_src, password_src,
        "project.task", "read",
        [task_ids_src],
        {"fields": ["id", "name", "sale_line_id"]}
    )

    print(f"📋 Tareas origen con sale_line_id: {len(tasks_src)}")

    # -----------------------
    # 3️⃣ ITERAR
    # -----------------------
    for t in tasks_src:
        old_task_id = t["id"]
        task_name = t["name"]

        # 🔹 Leer sale.order.line ORIGEN
        sol_src = models_src.execute_kw(
            db_src, uid_src, password_src,
            "sale.order.line", "read",
            [[t["sale_line_id"][0]]],
            {"fields": ["order_id"]}
        )[0]

        order_name = sol_src["order_id"][1]

        # 🔹 Buscar tarea DESTINO
        task_dst_ids = models.execute_kw(
            db, uid, password,
            "project.task", "search",
            [[("x_old_task_id", "=", old_task_id)]],
            {"limit": 1}
        )

        if not task_dst_ids:
            print(f"⚠️ Tarea destino no encontrada | {task_name}")
            continue

        task_dst_id = task_dst_ids[0]

        # 🔹 Buscar sale.order.line DESTINO (CLAVE FINAL)
        sol_dst_ids = models.execute_kw(
            db, uid, password,
            "sale.order.line", "search",
            [[
                ("order_id.name", "=", order_name),
                ("product_id", "=", product_dst_id)
            ]],
            {"limit": 1}
        )

        if not sol_dst_ids:
            print(
                f"❌ Sale line NO encontrada | "
                f"Tarea='{task_name}' | "
                f"Pedido='{order_name}'"
            )
            continue

        # 🔹 WRITE
        try:
            models.execute_kw(
                db, uid, password,
                "project.task", "write",
                [[task_dst_id], {"sale_line_id": sol_dst_ids[0]}]
            )
            print(f"🔗 Vinculada venta → tarea '{task_name}'")
        except Exception as e:
            print(f"❌ Error escribiendo tarea '{task_name}': {e}")

    print("✅ Sincronización sale_line_id finalizada")

# endregion

#region PARTE DE HORAS
# ---------------------------------------------------------------
# 🔹 MIGRACIÓN DE PARTES DE HORAS (account.analytic.line)
# ---------------------------------------------------------------

def export_timesheets(company_name="ALMAITANA DE LUZ, S.L."):
    """
    Exporta partes de horas desde Odoo 16.
    Incluye referencias a empleado, tarea y proyecto.
    """
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db_old = 'Real'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})
    if not uid_old:
        print("❌ No se pudo autenticar en origen.")
        return []

    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
    print(f'🔌 Conectado a Odoo 16 como {username}')

    # Buscar compañía
    company_id_src = models_old.execute_kw(
        db_old, uid_old, password_old,
        'res.company', 'search',
        [[('name', '=', company_name)]],
        {'limit': 1}
    )
    company_id_src = company_id_src[0] if company_id_src else False

    # Campos relevantes de partes de horas
    FIELDS = [
        "name",
        "employee_id",
        "project_id",
        "task_id",
        "unit_amount",  # Horas trabajadas
        "date",
        "amount",
        "company_id",
        "user_id",
        "account_id",
    ]

    domain = [("company_id", "=", company_id_src)]
    timesheets = models_old.execute_kw(
        db_old, uid_old, password_old,
        "account.analytic.line", "search_read",
        [domain],
        {"fields": FIELDS}
    )

    print(f"📤 {len(timesheets)} partes de horas exportados correctamente.")
    return timesheets

def import_timesheets(timesheets):
    """
    Importa partes de horas en Odoo 18.
    Vincula correctamente proyecto, tarea y empleado.
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    total_creados = 0
    total_existentes = 0
    total_sin_relacion = 0

    def find_project_id(project_name):
        projects = models.execute_kw(
            db, uid, password,
            "project.project", "search_read",
            [[("name", "=", project_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return projects[0]["id"] if projects else None

    def find_task_id(task_name, project_id=None):
        domain = [("name", "=", task_name)]
        if project_id:
            domain.append(("project_id", "=", project_id))
        tasks = models.execute_kw(
            db, uid, password,
            "project.task", "search_read",
            [domain],
            {"fields": ["id"], "limit": 1}
        )
        return tasks[0]["id"] if tasks else None

    def find_employee_id(emp_name):
        employees = models.execute_kw(
            db, uid, password,
            "hr.employee", "search_read",
            [[("name", "=", emp_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return employees[0]["id"] if employees else None

    print("📥 Iniciando importación de partes de horas...")

    for line in timesheets:
        name = line.get("name") or "SIN_DESCRIPCIÓN"
        date = line.get("date")
        hours = line.get("unit_amount", 0.0)

        print(f"\n🕒 Procesando parte: {name} ({hours}h en {date})")

        # Buscar proyecto y tarea
        project_id = None
        if line.get("project_id"):
            project_name = line["project_id"][1]
            project_id = find_project_id(project_name)

        task_id = None
        if line.get("task_id"):
            task_name = line["task_id"][1]
            task_id = find_task_id(task_name, project_id)

        # Buscar empleado
        employee_id = None
        if line.get("employee_id"):
            emp_name = line["employee_id"][1]
            employee_id = find_employee_id(emp_name)

        # Saltar si faltan vínculos clave
        if not employee_id or not project_id:
            print(f"⚠️ Parte ignorado (falta empleado o proyecto).")
            total_sin_relacion += 1
            continue

        # Comprobar duplicados por fecha, empleado, horas y proyecto
        domain_check = [
            ("name", "=", name),
            ("date", "=", date),
            ("employee_id", "=", employee_id),
            ("project_id", "=", project_id),
        ]
        existing = models.execute_kw(
            db, uid, password,
            "account.analytic.line", "search",
            [domain_check]
        )
        '''if existing:
            print(f"⚠️ Parte ya existente para '{emp_name}' en '{project_name}' el {date}")
            total_existentes += 1
            continue'''

        vals = {
            "name": name,
            "employee_id": employee_id,
            "project_id": project_id,
            "task_id": task_id,
            "unit_amount": hours,
            "date": date,
        }

        # Eliminar None antes de enviar
        vals = {k: v for k, v in vals.items() if v is not None}

        try:
            new_id = models.execute_kw(db, uid, password, "account.analytic.line", "create", [vals])
            print(f"✅ Parte creado (ID {new_id})")
            total_creados += 1
        except Exception as e:
            print(f"❌ Error creando parte '{name}': {e}")

    print("\n📊 MIGRACIÓN DE PARTES DE HORAS COMPLETADA")
    print(f"   Total creados: {total_creados}")
    print(f"   Ya existentes: {total_existentes}")
    print(f"   Ignorados (sin empleado o proyecto): {total_sin_relacion}")

# ---------------------------------------------------------------
# 🔸 Función principal
# ---------------------------------------------------------------

def migrar_partes_horas():
    partes = export_timesheets()
    import_timesheets(partes)

# Ejecución directa opcional
#migrar_partes_horas()

#endregion

#region ASISTENCIAS
# ---------------------------------------------------------------
# 🔹 MIGRACIÓN DE ASISTENCIAS (hr.attendance)
# ---------------------------------------------------------------

def export_attendances():
    """
    Exporta todas las asistencias (hr.attendance) de Odoo 16 sin filtrar por compañía.
    Incluye empleado, check_in, check_out y horas trabajadas.
    """
    import xmlrpc.client

    url = 'https://optimaluz.soluntec.net'
    db_old = 'Real'
    username = 'jcoronado@optimaluz.com'
    password_old = 'AlAi4ever'

    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
    uid_old = common.authenticate(db_old, username, password_old, {})
    if not uid_old:
        print("❌ No se pudo autenticar en origen.")
        return []

    models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
    print(f'🔌 Conectado a Odoo 16 como {username}')

    FIELDS = [
        "employee_id",
        "check_in",
        "check_out",
        "worked_hours",
    ]

    attendances = models_old.execute_kw(
        db_old, uid_old, password_old,
        "hr.attendance", "search_read",
        [[]], {"fields": FIELDS}
    )

    print(f"📤 {len(attendances)} asistencias exportadas correctamente.")
    return attendances

def import_attendances(attendances):
    """
    Importa asistencias (hr.attendance) en Odoo 18.
    Asocia correctamente al empleado.
    """
    import xmlrpc.client
    from App_Connection import db, uid, password, models

    total_creadas = 0
    total_existentes = 0
    total_sin_empleado = 0

    def find_employee_id(emp_name):
        employees = models.execute_kw(
            db, uid, password,
            "hr.employee", "search_read",
            [[("name", "=", emp_name)]],
            {"fields": ["id"], "limit": 1}
        )
        return employees[0]["id"] if employees else None

    print("📥 Iniciando importación de asistencias...")

    for att in attendances:
        emp_name = att["employee_id"][1] if att.get("employee_id") else None
        check_in = att.get("check_in")
        check_out = att.get("check_out")
        hours = att.get("worked_hours", 0.0)

        if not emp_name:
            print(f"⚠️ Asistencia ignorada (sin empleado).")
            total_sin_empleado += 1
            continue

        employee_id = find_employee_id(emp_name)
        if not employee_id:
            print(f"⚠️ Empleado '{emp_name}' no encontrado en destino.")
            total_sin_empleado += 1
            continue

        # Comprobar duplicado exacto (empleado + check_in + check_out)
        domain_check = [
            ("employee_id", "=", employee_id),
            ("check_in", "=", check_in),
            ("check_out", "=", check_out),
        ]
        existing = models.execute_kw(
            db, uid, password,
            "hr.attendance", "search",
            [domain_check]
        )
        if existing:
            print(f"⚠️ Asistencia ya existente para {emp_name} el {check_in}")
            total_existentes += 1
            continue

        vals = {
            "employee_id": employee_id,
            "check_in": check_in,
            "check_out": check_out,
            "worked_hours": hours,
        }

        # Eliminar None
        vals = {k: v for k, v in vals.items() if v is not None}

        try:
            new_id = models.execute_kw(db, uid, password, "hr.attendance", "create", [vals])
            print(f"✅ Asistencia creada (ID {new_id}) - {emp_name} {check_in}")
            total_creadas += 1
        except Exception as e:
            print(f"❌ Error creando asistencia de {emp_name}: {e}")

    print("\n📊 MIGRACIÓN DE ASISTENCIAS COMPLETADA")
    print(f"   Total creadas: {total_creadas}")
    print(f"   Ya existentes: {total_existentes}")
    print(f"   Ignoradas (sin empleado): {total_sin_empleado}")

# ---------------------------------------------------------------
# 🔸 Función principal
# ---------------------------------------------------------------

def migrar_asistencias():
    asistencias = export_attendances()
    import_attendances(asistencias)

# Ejecución directa opcional
#migrar_asistencias()

#endregion

# region TRANSFERENCIA
def update_public_info(excel_path):
    from App_Connection import models, db, uid, password

    print("🔎 Leyendo Excel...")
    df = pd.read_excel(excel_path)

    for index, row in df.iterrows():
        default_code = str(row["default_code"]).strip()
        name = str(row["name"]).strip() if "name" in row else ""
        excel_categ_id = str(row["categ_id"]).strip() if "categ_id" in row else ""
        excel_categories = str(row["public_categ_ids"]).strip()

        if not default_code:
            continue

        # Buscar producto en Odoo
        product_ids = models.execute_kw(
            db, uid, password,
            'product.template', 'search',
            [[('default_code', '=', default_code)]]
        )

        if not product_ids:
            print(f"⚠️ Producto no encontrado en Odoo: {default_code}")
            continue

        product_id = product_ids[0]

        # ----------------------------
        # 1. Actualizar nombre si procede
        # ----------------------------
        if name:
            models.execute_kw(
                db, uid, password,
                'product.template', 'write',
                [[product_id], {'name': name}]
            )
            # print(f"✏️ Nombre actualizado para {default_code} → {name}")

        # ----------------------------
        # 2. Actualizar categoría interna si procede
        # ----------------------------
        if excel_categ_id:
            try:
                # Buscar por nombre de categoría
                categ_ids = models.execute_kw(
                    db, uid, password,
                    'product.category', 'search',
                    [[('name', '=', excel_categ_id)]],
                    {'limit': 1}
                )
                if categ_ids:
                    categ_id = categ_ids[0]
                else:
                    categ_id = models.execute_kw(
                        db, uid, password,
                        'product.category', 'create',
                        [{'name': excel_categ_id}]
                    )
                    # print(f"🆕 Creada categoría interna '{excel_categ_id}' con ID {categ_id}")

                models.execute_kw(
                    db, uid, password,
                    'product.template', 'write',
                    [[product_id], {'categ_id': categ_id}]
                )
                # print(f"📂 Categoría interna actualizada para {default_code} → {excel_categ_id}")

            except Exception as e:
                print(f"⚠️ Error al asignar categoría interna para {default_code}: {e}")

        elif excel_categories:
            category_paths = [path.strip() for path in excel_categories.split(",")]
            parts = [p.strip() for p in category_paths[0].split("/")]
            padre = parts[0]

            # Buscar por nombre de categoría
            categ_ids = models.execute_kw(
                db, uid, password,
                'product.category', 'search',
                [[('name', '=', padre)]],
                {'limit': 1}
            )
            if categ_ids:
                categ_id = categ_ids[0]
            else:
                print(f"No existe {padre}")

            models.execute_kw(
                db, uid, password,
                'product.template', 'write',
                [[product_id], {'categ_id': categ_id}]
            )

        # ----------------------------
        # 3. Actualizar categorías públicas
        # ----------------------------
        if excel_categories:
            # Obtener categorías actuales del producto
            product_data = models.execute_kw(
                db, uid, password,
                'product.template', 'read',
                [product_id], {'fields': ['public_categ_ids']}
            )
            odoo_categories = product_data[0]['public_categ_ids']

            # Procesar múltiples rutas de categorías (separadas por comas)
            category_paths = [path.strip() for path in excel_categories.split(",")]
            new_categ_ids = []

            for path in category_paths:
                parts = [p.strip() for p in path.split("/")]
                parent_id = False
                last_categ_id = None

                for part in parts:
                    domain = [('name', '=', part)]
                    if parent_id:
                        domain.append(('parent_id', '=', parent_id))
                    else:
                        domain.append(('parent_id', '=', False))

                    categ_ids = models.execute_kw(
                        db, uid, password,
                        'product.public.category', 'search',
                        [domain]
                    )

                    if categ_ids:
                        last_categ_id = categ_ids[0]
                    else:
                        vals = {'name': part}
                        if parent_id:
                            vals['parent_id'] = parent_id
                        last_categ_id = models.execute_kw(
                            db, uid, password,
                            'product.public.category', 'create',
                            [vals]
                        )
                        # print(f"🆕 Creada categoría '{part}' (padre={parent_id}) con ID {last_categ_id}")

                    parent_id = last_categ_id

                if last_categ_id:
                    new_categ_ids.append(last_categ_id)

            if new_categ_ids:
                if set(new_categ_ids) != set(odoo_categories):
                    models.execute_kw(
                        db, uid, password,
                        'product.template', 'write',
                        [[product_id], {'public_categ_ids': [(6, 0, new_categ_ids)]}]
                    )
                    # print(f"🔄 Actualizado {default_code} → {excel_categories}")
                # else: print(f"✅ Categorías ya correctas para {default_code}")

        print(f"Actualizando {index} / 9164")

def export_import_v_tac_products():
    """
    Importa productos V-Tac a Odoo desde un diccionario exportado.
    Convierte imágenes de iconos (x_icono1...x_icono8) a base64 antes de subirlas.
    """

    def export_v_tac_products(download_media=False,
                              islimited=False,
                              limit=1,
                              output_file=None):
        """
        Exporta productos V-Tac de Odoo a Excel y los devuelve como diccionario products.
        """

        try:
            import pandas as pd
            import base64
            import os
            from pathlib import Path
            from App_Connection import models, db, uid, password

            desired_fields = [
                'image_1920', 'name', 'website_description', 'description', 'description_purchase',
                'description_sale', 'list_price', 'standard_price', 'volume', 'weight',
                'barcode', 'default_code', 'product_brand_id', 'out_of_stock_message',
                'x_url', 'x_codigo_de_familia', 'x_almacen1_custom', 'x_almacen2_custom',
                'x_transit_stock_custom', 'x_transit', 'x_icono1', 'x_icono2', 'x_icono3',
                'x_icono4', 'x_icono5', 'x_icono6', 'x_icono7', 'x_icono8'
            ]

            # --- Dominio base ---
            domain = [
                ('product_brand_id.name', '=', 'V-Tac')
            ]

            # --- Buscar productos (con paginación si es necesario) ---
            if islimited:
                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [domain],
                    {'limit': limit}
                )
                products = models.execute_kw(
                    db, uid, password,
                    'product.template', 'read',
                    [product_ids],
                    {'fields': desired_fields}
                )
            else:
                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [domain]
                )
                batch_size = 100
                products = []
                for i in range(0, len(product_ids), batch_size):
                    batch_ids = product_ids[i:i + batch_size]
                    batch = models.execute_kw(
                        db, uid, password,
                        'product.template', 'read',
                        [batch_ids],
                        {'fields': desired_fields}
                    )
                    products.extend(batch)

            # --- Procesar productos ---
            for product in products:
                product_id = product['id']

                # Limpieza y formateo
                for key, value in list(product.items()):
                    if value is False or value is None:
                        product[key] = ""
                    elif isinstance(value, list):
                        if key == 'product_brand_id' and len(value) == 2:
                            product[key] = value[1]

                # --- Atributos ---
                atributos_dic = {}

                # 1. Obtener líneas de atributos del template
                attribute_lines = models.execute_kw(
                    db, uid, password,
                    'product.template.attribute.value', 'search_read',
                    [[('product_tmpl_id', '=', product_id)]],
                    {'fields': ['attribute_id', 'product_attribute_value_id']}
                )

                for line in attribute_lines:
                    attr_name = ""
                    value_name = ""

                    # Nombre del atributo (ej: Color, Potencia...)
                    if isinstance(line.get('attribute_id'), list) and len(line['attribute_id']) == 2:
                        attr_name = line['attribute_id'][1]

                    # Valor del atributo (ej: Rojo, 20W...)
                    if isinstance(line.get('product_attribute_value_id'), list) and len(
                            line['product_attribute_value_id']) == 2:
                        value_name = line['product_attribute_value_id'][1]

                    if attr_name:
                        if attr_name not in atributos_dic:
                            atributos_dic[attr_name] = []
                        if value_name and value_name not in atributos_dic[attr_name]:
                            atributos_dic[attr_name].append(value_name)

                product['Atributos'] = atributos_dic

                # Galería
                galeria = []
                gallery_imgs = models.execute_kw(
                    db, uid, password,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', '=', product_id)]],
                    {'fields': ['image_1920', 'video_url']}
                )

                if download_media:
                    default_code = product.get('default_code', '').strip()
                    if default_code:
                        ruta_base = os.path.normpath("D:/BANCOS/OPTIMA/src")
                        ruta_producto = os.path.join(ruta_base, default_code)
                        os.makedirs(ruta_producto, exist_ok=True)

                        for idx, img in enumerate(gallery_imgs):
                            if img.get('image_1920'):
                                try:
                                    raw = Utils.image_url_to_base64(
                                        img['image_1920'])  # base64.b64decode(img['image_1920'])
                                    file_path = os.path.join(ruta_producto, f"{idx + 1}.jpg")
                                    with open(file_path, "wb") as f:
                                        f.write(raw)
                                except Exception as e:
                                    print(f"❌ Error al guardar media {idx + 1} para {default_code}: {e}")
                            elif img.get('video_url'):
                                galeria.append(img['video_url'])
                    else:
                        print(f"⚠️ Producto sin SKU, se omite galería: {product.get('name')}")

                product['Galeria'] = gallery_imgs

            # --- Exportar a Excel ---
            if output_file:
                df = pd.DataFrame(products)
                df.to_excel(output_file, index=False)

            print(f"✅ Productos V-Tac exportados correctamente a: {output_file} (total: {len(products)})")
            return products

        except Exception as e:
            print(f"❌ Error al exportar productos V-Tac: {e}")

    try:
        import xmlrpc.client

        products = export_v_tac_products()

        # 1. Conexión
        url = 'http://79.72.55.217:8069'
        db = 'odoo1'
        username = 'admin'
        password = 'admin'

        common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
        uid = common.authenticate(db, username, password, {})

        if not uid:
            print("❌ No se pudo autenticar.")
            return

        print(f'🔌 Conectado como {username} (uid: {uid})')
        models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

        # --- Iterar sobre productos ---
        for product in products:
            vals = product.copy()

            # Limpieza de campos no válidos para create/write
            for k in ["Galeria", "Atributos", "id"]:
                vals.pop(k, None)

            # Verificar si ya existe por default_code
            default_code = vals.get("default_code")
            existing_ids = []
            if default_code:
                existing_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [[('default_code', '=', default_code)]],
                    {'limit': 1}
                )

            if existing_ids:
                # --- Actualizar ---
                try:
                    models.execute_kw(
                        db, uid, password,
                        'product.template', 'write',
                        [existing_ids, vals]
                    )
                    print(f"✏️ Actualizado producto {default_code}")
                except Exception as e:
                    print(f"❌ Error al actualizar {default_code}: {e}")
            else:
                # --- Resolver Many2one: product_brand_id ---
                if vals.get("product_brand_id"):
                    brand_name = vals["product_brand_id"]
                    brand_ids = models.execute_kw(
                        db, uid, password,
                        'product.brand', 'search',
                        [[('name', '=', brand_name)]],
                        {'limit': 1}
                    )
                    if brand_ids:
                        vals["product_brand_id"] = brand_ids[0]
                    else:
                        # si no existe, lo creamos
                        new_brand_id = models.execute_kw(
                            db, uid, password,
                            'product.brand', 'create',
                            [{'name': brand_name}]
                        )
                        vals["product_brand_id"] = new_brand_id

                vals["allow_out_of_stock_order"] = True
                vals["is_storable"] = 1
                vals["invoice_policy"] = "delivery"
                vals["show_availability"] = True
                vals["available_threshold"] = 100000
                vals["name"] = str(vals["default_code"])
                # --- Crear ---
                try:
                    new_id = models.execute_kw(
                        db, uid, password,
                        'product.template', 'create',
                        [vals]
                    )

                    # --- Importar Atributos ---
                    atributos = product.get("Atributos", {})
                    for attr_name, values in atributos.items():
                        try:
                            # 1. Buscar o crear el atributo
                            attr_ids = models.execute_kw(
                                db, uid, password,
                                'product.attribute', 'search',
                                [[('name', '=', attr_name)]],
                                {'limit': 1}
                            )
                            if not attr_ids:
                                attr_id = models.execute_kw(
                                    db, uid, password,
                                    'product.attribute', 'create',
                                    [{'name': attr_name}]
                                )
                            else:
                                attr_id = attr_ids[0]

                            # 2. Asegurar que existen los valores
                            value_ids = []
                            for val in values:
                                val_ids = models.execute_kw(
                                    db, uid, password,
                                    'product.attribute.value', 'search',
                                    [[('name', '=', val), ('attribute_id', '=', attr_id)]],
                                    {'limit': 1}
                                )
                                if not val_ids:
                                    val_id = models.execute_kw(
                                        db, uid, password,
                                        'product.attribute.value', 'create',
                                        [{'name': val, 'attribute_id': attr_id}]
                                    )
                                else:
                                    val_id = val_ids[0]
                                value_ids.append(val_id)

                            # 3. Crear (o actualizar) la línea de atributo en el template
                            line_ids = models.execute_kw(
                                db, uid, password,
                                'product.template.attribute.line', 'search',
                                [[('product_tmpl_id', '=', new_id), ('attribute_id', '=', attr_id)]],
                                {'limit': 1}
                            )

                            if not line_ids:
                                models.execute_kw(
                                    db, uid, password,
                                    'product.template.attribute.line', 'create',
                                    [{
                                        'product_tmpl_id': new_id,
                                        'attribute_id': attr_id,
                                        'value_ids': [(6, 0, value_ids)]
                                    }]
                                )
                            else:
                                # Si ya existe, añadimos valores sin machacar
                                models.execute_kw(
                                    db, uid, password,
                                    'product.template.attribute.line', 'write',
                                    [line_ids, {'value_ids': [(4, vid) for vid in value_ids]}]
                                )

                        except Exception as e:
                            print(f"⚠️ Error al importar atributo {attr_name}: {e}")

                    # --- Importar Galería ---
                    galeria = product.get("Galeria", [])
                    for idx, img in enumerate(galeria):
                        try:
                            vals_img = {'product_tmpl_id': new_id, 'sequence': idx, 'name': f"img{idx + 1}"}
                            if img.get("image_1920"):
                                vals_img['image_1920'] = img['image_1920']
                            if img.get("video_url"):
                                vals_img['video_url'] = img['video_url']

                            models.execute_kw(
                                db, uid, password,
                                'product.image', 'create',
                                [vals_img]
                            )
                        except Exception as e:
                            print(f"⚠️ Error al importar galería en {product.get('default_code')}: {e}")

                    print(f"➕ Creado producto {default_code} (ID {new_id})")
                except Exception as e:
                    print(f"❌ Error al crear {default_code}: {e}")

        print("✅ Importación finalizada")

    except Exception as e:
        print(f"❌ Error en importación: {e}")

def export_import_products_by_excel_brands(excel_path, islimited=False, limit=1, output_file=None):
    """
    Exporta e importa productos a Odoo en base a las marcas listadas en un Excel (columna 'Marca').
    Convierte imágenes de iconos (x_icono1...x_icono8) a base64 antes de subirlas.
    """

    def export_products(islimited=False,
                        limit=1,
                        output_file=None):
        """
        Exporta productos de Odoo según las marcas encontradas en el Excel.
        Devuelve una lista de diccionarios 'products'.
        """
        try:
            import pandas as pd
            import os
            from App_Connection import models, db, uid, password

            # --- 1. Leer Excel ---
            df = pd.read_excel(excel_path)
            if "Marca" not in df.columns:
                print("❌ El Excel no tiene columna 'Marca'")
                return []

            marcas = df["Marca"].dropna().unique().tolist()
            print(f"🔍 Se encontraron {len(marcas)} marcas en el Excel: {marcas}")

            desired_fields = [
                'image_1920', 'name', 'website_description', 'description', 'description_purchase',
                'description_sale', 'list_price', 'standard_price', 'volume', 'weight',
                'barcode', 'default_code', 'product_brand_id', 'out_of_stock_message',
                'x_url', 'x_codigo_de_familia', 'x_almacen1_custom', 'x_almacen2_custom',
                'x_transit_stock_custom', 'x_transit', 'x_icono1', 'x_icono2', 'x_icono3',
                'x_icono4', 'x_icono5', 'x_icono6', 'x_icono7', 'x_icono8'
            ]

            # --- 2. Dominio dinámico ---
            domain = [('product_brand_id.name', 'in', marcas)]

            # --- 3. Buscar productos ---
            if islimited:
                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [domain],
                    {'limit': limit}
                )
                products = models.execute_kw(
                    db, uid, password,
                    'product.template', 'read',
                    [product_ids],
                    {'fields': desired_fields}
                )
            else:
                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [domain]
                )
                batch_size = 100
                products = []
                for i in range(0, len(product_ids), batch_size):
                    batch_ids = product_ids[i:i + batch_size]
                    batch = models.execute_kw(
                        db, uid, password,
                        'product.template', 'read',
                        [batch_ids],
                        {'fields': desired_fields}
                    )
                    products.extend(batch)

            # --- 4. Procesar productos ---
            for product in products:
                product_id = product['id']

                # Limpieza de valores nulos/listas
                for key, value in list(product.items()):
                    if value is False or value is None:
                        product[key] = ""
                    elif isinstance(value, list):
                        if key == 'product_brand_id' and len(value) == 2:
                            product[key] = value[1]

                # Atributos
                atributos_dic = {}
                attribute_lines = models.execute_kw(
                    db, uid, password,
                    'product.template.attribute.value', 'search_read',
                    [[('product_tmpl_id', '=', product_id)]],
                    {'fields': ['attribute_id', 'product_attribute_value_id']}
                )
                for line in attribute_lines:
                    attr_name, value_name = "", ""
                    if isinstance(line.get('attribute_id'), list) and len(line['attribute_id']) == 2:
                        attr_name = line['attribute_id'][1]
                    if isinstance(line.get('product_attribute_value_id'), list) and len(
                            line['product_attribute_value_id']) == 2:
                        value_name = line['product_attribute_value_id'][1]
                    if attr_name:
                        if attr_name not in atributos_dic:
                            atributos_dic[attr_name] = []
                        if value_name and value_name not in atributos_dic[attr_name]:
                            atributos_dic[attr_name].append(value_name)
                product['Atributos'] = atributos_dic

                # Galería
                gallery_imgs = models.execute_kw(
                    db, uid, password,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', '=', product_id)]],
                    {'fields': ['image_1920', 'video_url']}
                )
                product['Galeria'] = gallery_imgs

            # --- 5. Exportar a Excel si se pide ---
            if output_file:
                pd.DataFrame(products).to_excel(output_file, index=False)
                print(f"✅ Exportados {len(products)} productos a {output_file}")

            return products

        except Exception as e:
            print(f"❌ Error al exportar productos: {e}")
            return []

    # -------------------------------------------------
    # Importación (idéntica a tu export_import_v_tac_products pero genérica)
    # -------------------------------------------------
    try:
        import xmlrpc.client

        products = export_products(islimited=islimited, limit=limit,
                                   output_file=output_file)

        url = 'http://79.72.61.76:8069/'
        db = 'odoo0'
        username = 'admin'
        password = 'admin'

        common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
        uid = common.authenticate(db, username, password, {})

        if not uid:
            print("❌ No se pudo autenticar en Odoo destino.")
            return

        print(f'🔌 Conectado como {username} (uid: {uid})')
        models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

        # --- Iterar sobre productos (igual que en tu script original) ---
        for product in products:
            vals = product.copy()
            for k in ["Galeria", "Atributos", "id"]:
                vals.pop(k, None)

            default_code = vals.get("default_code")
            existing_ids = []
            if default_code:
                existing_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [[('default_code', '=', default_code)]],
                    {'limit': 1}
                )

            if existing_ids:
                try:
                    models.execute_kw(db, uid, password, 'product.template', 'write', [existing_ids, vals])
                    print(f"✏️ Actualizado producto {default_code}")
                except Exception as e:
                    print(f"❌ Error al actualizar {default_code}: {e}")
            else:
                # Resolver Many2one: product_brand_id
                if vals.get("product_brand_id"):
                    brand_name = vals["product_brand_id"]
                    brand_ids = models.execute_kw(
                        db, uid, password,
                        'product.brand', 'search',
                        [[('name', '=', brand_name)]],
                        {'limit': 1}
                    )
                    if brand_ids:
                        vals["product_brand_id"] = brand_ids[0]
                    else:
                        new_brand_id = models.execute_kw(
                            db, uid, password,
                            'product.brand', 'create',
                            [{'name': brand_name}]
                        )
                        vals["product_brand_id"] = new_brand_id

                vals["is_storable"] = 1
                vals["invoice_policy"] = "delivery"
                vals["name"] = str(vals["default_code"])
                #vals["show_availability"] = True
                #vals["available_threshold"] = 100000
                #vals["allow_out_of_stock_order"] = True

                try:
                    new_id = models.execute_kw(db, uid, password, 'product.template', 'create', [vals])

                    # --- Importar Atributos ---
                    atributos = product.get("Atributos", {})
                    for attr_name, values in atributos.items():
                        try:
                            # 1. Buscar o crear el atributo
                            attr_ids = models.execute_kw(
                                db, uid, password,
                                'product.attribute', 'search',
                                [[('name', '=', attr_name)]],
                                {'limit': 1}
                            )
                            if not attr_ids:
                                attr_id = models.execute_kw(
                                    db, uid, password,
                                    'product.attribute', 'create',
                                    [{'name': attr_name}]
                                )
                            else:
                                attr_id = attr_ids[0]

                            # 2. Asegurar que existen los valores
                            value_ids = []
                            for val in values:
                                val_ids = models.execute_kw(
                                    db, uid, password,
                                    'product.attribute.value', 'search',
                                    [[('name', '=', val), ('attribute_id', '=', attr_id)]],
                                    {'limit': 1}
                                )
                                if not val_ids:
                                    val_id = models.execute_kw(
                                        db, uid, password,
                                        'product.attribute.value', 'create',
                                        [{'name': val, 'attribute_id': attr_id}]
                                    )
                                else:
                                    val_id = val_ids[0]
                                value_ids.append(val_id)

                            # 3. Crear (o actualizar) la línea de atributo en el template
                            line_ids = models.execute_kw(
                                db, uid, password,
                                'product.template.attribute.line', 'search',
                                [[('product_tmpl_id', '=', new_id), ('attribute_id', '=', attr_id)]],
                                {'limit': 1}
                            )

                            if not line_ids:
                                models.execute_kw(
                                    db, uid, password,
                                    'product.template.attribute.line', 'create',
                                    [{
                                        'product_tmpl_id': new_id,
                                        'attribute_id': attr_id,
                                        'value_ids': [(6, 0, value_ids)]
                                    }]
                                )
                            else:
                                # Si ya existe, añadimos valores sin machacar
                                models.execute_kw(
                                    db, uid, password,
                                    'product.template.attribute.line', 'write',
                                    [line_ids, {'value_ids': [(4, vid) for vid in value_ids]}]
                                )

                        except Exception as e:
                            print(f"⚠️ Error al importar atributo {attr_name}: {e}")

                    # --- Importar Galería ---
                    galeria = product.get("Galeria", [])
                    for idx, img in enumerate(galeria):
                        try:
                            vals_img = {'product_tmpl_id': new_id, 'sequence': idx, 'name': f"img{idx + 1}"}
                            if img.get("image_1920"):
                                vals_img['image_1920'] = img['image_1920']
                            if img.get("video_url"):
                                vals_img['video_url'] = img['video_url']

                            models.execute_kw(
                                db, uid, password,
                                'product.image', 'create',
                                [vals_img]
                            )
                        except Exception as e:
                            print(f"⚠️ Error al importar galería en {product.get('default_code')}: {e}")

                    print(f"➕ Creado producto {default_code} (ID {new_id})")
                except Exception as e:
                    print(f"❌ Error al crear {default_code}: {e}")

        print("✅ Importación finalizada")

    except Exception as e:
        print(f"❌ Error en export_import_products_by_excel_brands: {e}")

def actualizar_descripciones_desde_excel(ruta_excel):
    from App_Connection import models, db, uid, password
    import pandas as pd

    print("📌 Leyendo Excel…")
    df = pd.read_excel(ruta_excel, sheet_name="Sheet1")

    if "SKU" not in df.columns:
        raise Exception("❌ La hoja Sheet1 debe contener la columna 'SKU'.")

    # Normalizar SKU
    df["SKU"] = df["SKU"].astype(str).str.strip()

    productos_actualizados = 0
    productos_no_encontrados = 0

    print(f"🔍 Total SKUs: {len(df)}")

    for sku in df["SKU"]:
        if not sku or sku.lower() == "nan":
            continue

        # Buscar producto por default_code
        prod_ids = models.execute_kw(
            db, uid, password,
            "product.template", "search",
            [[["default_code", "=", sku]]]
        )

        if not prod_ids:
            productos_no_encontrados += 1
            print(f"⚠️ Producto NO encontrado: {sku}")
            continue

        prod_id = prod_ids[0]

        # Leer descripción actual
        prod_data = models.execute_kw(
            db, uid, password,
            "product.template", "read",
            [prod_ids, ["website_description"]]
        )[0]

        descripcion_actual = prod_data.get("website_description", "") or ""

        # Generar descripción nueva con tu función
        try:
            nueva_descripcion = Utils.update_product_description(sku, descripcion_actual)
        except Exception as e:
            print(f"❌ Error generando descripción para {sku}: {e}")
            continue

        # Actualizar en Odoo
        try:
            models.execute_kw(
                db, uid, password,
                "product.template", "write",
                [[prod_id], {"website_description": nueva_descripcion}]
            )
            productos_actualizados += 1
            print(f"✔ SKU {sku} → descripción actualizada")

        except Exception as e:
            print(f"❌ Error actualizando descripción de {sku}: {e}")

    print("--------------------------------------------------")
    print(f"🎉 Proceso terminado.")
    print(f"   Productos actualizados: {productos_actualizados}")
    print(f"   No encontrados: {productos_no_encontrados}")

def transferencia_productos_por_marca():
    def migrate_products_by_brands(models_src, db_src, uid_src, pwd_src,
                                   models_dst, db_dst, uid_dst, pwd_dst,
                                   excel_path):
        """
        Migra productos filtrados por marcas obtenidas desde un Excel.
        También migra galería y public_categ_ids.
        """

        import pandas as pd

        # -------------------------------------------------
        # 1️⃣ Leer marcas desde Excel
        # -------------------------------------------------

        df = pd.read_excel(excel_path, sheet_name="Sheet1")

        if "Marca" not in df.columns:
            raise Exception("❌ El Excel no contiene la columna 'Marca'")

        marcas = (
            df["Marca"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not marcas:
            print("ℹ️ No hay marcas en el Excel")
            return

        print(f"🏷️ Marcas detectadas: {marcas}")

        # -------------------------------------------------
        # 2️⃣ Campos a leer
        # -------------------------------------------------

        FIELDS = [
            'image_1920',
            'name',
            'default_code',
            'invoice_policy',
            'standard_price',
            'categ_id',
            'product_brand_id',
            'x_url',
            'description',
            'out_of_stock_message',
            'public_categ_ids'
        ]

        # -------------------------------------------------
        # 3️⃣ Buscar productos archivados por marca
        # -------------------------------------------------

        domain = [
            ('product_brand_id.name', 'in', marcas)
        ]

        product_ids = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'search',
            [domain]
        )

        if not product_ids:
            print("ℹ️ No se encontraron productos para esas marcas")
            return

        products = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'read',
            [product_ids],
            {'fields': FIELDS}
        )

        print(f"📦 Productos a migrar: {len(products)}")

        # -------------------------------------------------
        # 4️⃣ Migración producto a producto
        # -------------------------------------------------

        for product in products:
            try:
                vals = {
                    'image_1920': product.get('image_1920'),
                    'name': product.get('name'),
                    'default_code': product.get('default_code'),
                    'invoice_policy': product.get('invoice_policy'),
                    'standard_price': product.get('standard_price') or 0.0,
                    'x_url': product.get('x_url'),
                    'description': product.get('description'),
                    'out_of_stock_message': product.get('out_of_stock_message'),
                    'list_price': 0.0
                }

                # -------------------------
                # Categoría interna
                # -------------------------
                if product.get('categ_id'):
                    categ_name = product['categ_id'][1]
                    categ_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.category', 'search',
                        [[('name', '=', categ_name)]],
                        {'limit': 1}
                    )
                    if categ_ids:
                        vals['categ_id'] = categ_ids[0]

                # -------------------------
                # Marca
                # -------------------------
                if product.get('product_brand_id'):
                    brand_name = product['product_brand_id'][1]
                    brand_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.brand', 'search',
                        [[('name', '=', brand_name)]],
                        {'limit': 1}
                    )
                    if brand_ids:
                        vals['product_brand_id'] = brand_ids[0]
                    else:
                        vals['product_brand_id'] = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.brand', 'create',
                            [{'name': brand_name}]
                        )

                # -------------------------
                # Categorías web
                # -------------------------
                if product.get('public_categ_ids'):
                    web_categ_ids = []
                    for wc in product['public_categ_ids']:
                        wc_name = models_src.execute_kw(
                            db_src, uid_src, pwd_src,
                            'product.public.category', 'read',
                            [[wc]],
                            {'fields': ['name']}
                        )[0]['name']

                        dst_wc = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.public.category', 'search',
                            [[('name', '=', wc_name)]],
                            {'limit': 1}
                        )

                        if dst_wc:
                            web_categ_ids.append(dst_wc[0])

                    if web_categ_ids:
                        vals['public_categ_ids'] = [(6, 0, web_categ_ids)]

                # -------------------------
                # Crear producto
                # -------------------------
                new_id = models_dst.execute_kw(
                    db_dst, uid_dst, pwd_dst,
                    'product.template', 'create',
                    [vals]
                )

                print(f"➕ Creado {vals.get('default_code')} (ID {new_id})")

                # -------------------------------------------------
                # 5️⃣ Galería
                # -------------------------------------------------

                images = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', '=', product['id'])]],
                    {'fields': ['image_1920', 'video_url', 'sequence']}
                )

                for img in images:
                    models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.image', 'create',
                        [{
                            'product_tmpl_id': new_id,
                            'image_1920': img.get('image_1920'),
                            'video_url': img.get('video_url'),
                            'sequence': img.get('sequence', 0)
                        }]
                    )

            except Exception as e:
                print(f"❌ Error en {product.get('default_code')}: {e}")

        print("✅ Migración por marcas finalizada")

    import xmlrpc.client

    # -----------------------
    # Configuración conexiones
    # -----------------------

    origen = {
        'url': "https://optimaluz.soluntec.net",
        'db': "Test",  # "Real"
        'user': "jcoronado@optimaluz.com",
        'password': "AlAi4ever"
    }

    destino = {
        'url': "http://79.72.61.76:8069",
        'db': "odoo0",
        'user': "admin",
        'password': "admin"
    }

    # -----------------------
    # Conexión ORIGEN
    # -----------------------

    common_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_src = common_src.authenticate(
        origen['db'],
        origen['user'],
        origen['password'],
        {}
    )

    if not uid_src:
        raise Exception("❌ No se pudo autenticar en ORIGEN")

    models_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a ORIGEN (uid={uid_src})")

    # -----------------------
    # Conexión DESTINO
    # -----------------------

    common_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_dst = common_dst.authenticate(
        destino['db'],
        destino['user'],
        destino['password'],
        {}
    )

    if not uid_dst:
        raise Exception("❌ No se pudo autenticar en DESTINO")

    models_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a DESTINO (uid={uid_dst})")

    # -----------------------
    # LLAMADA A LA MIGRACIÓN
    # -----------------------

    migrate_products_by_brands(
        models_src=models_src,
        db_src=origen['db'],
        uid_src=uid_src,
        pwd_src=origen['password'],

        models_dst=models_dst,
        db_dst=destino['db'],
        uid_dst=uid_dst,
        pwd_dst=destino['password'],

        excel_path="ruta"
    )

def transferencia_productos_por_sku():
    def migrate_products_by_sku(models_src, db_src, uid_src, pwd_src,
                                models_dst, db_dst, uid_dst, pwd_dst,
                                excel_path):
        """
        Migra productos filtrados por default_code (SKU) obtenidos desde un Excel.
        También migra galería y public_categ_ids.
        """

        con_atributos = True

        # -------------------------------------------------
        # 1️⃣ Leer SKUs desde Excel
        # -------------------------------------------------

        df = pd.read_excel(excel_path, sheet_name="Sheet1")

        if "SKU" not in df.columns:
            raise Exception("❌ El Excel no contiene la columna 'SKU'")

        skus = (
            df["SKU"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not skus:
            print("ℹ️ No hay SKUs en el Excel")
            return

        print(f"🧾 SKUs detectados: {len(skus)}")

        # -------------------------------------------------
        # 2️⃣ Campos a leer
        # -------------------------------------------------

        FIELDS = [
            'id',
            'image_1920',
            'name',
            'default_code',
            'invoice_policy',
            'standard_price',
            'categ_id',
            'product_brand_id',
            'x_url',
            'description',
            'out_of_stock_message',
            'public_categ_ids'
        ]

        # -------------------------------------------------
        # 3️⃣ Buscar productos por SKU
        # -------------------------------------------------

        domain = [
            ('default_code', 'in', skus)
        ]

        product_ids = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'search',
            [domain]
        )

        if not product_ids:
            print("ℹ️ No se encontraron productos para esos SKUs")
            return

        products = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'read',
            [product_ids],
            {'fields': FIELDS}
        )

        print(f"📦 Productos a migrar: {len(products)}")

        # -------------------------------------------------
        # 4️⃣ Migración producto a producto
        # -------------------------------------------------

        for product in products:
            try:
                vals = {
                    'image_1920': product.get('image_1920'),
                    'name': product.get('name'),
                    'default_code': product.get('default_code'),
                    'invoice_policy': product.get('invoice_policy'),
                    'standard_price': product.get('standard_price') or 0.0,
                    'x_url': product.get('x_url'),
                    'description': product.get('description'),
                    'out_of_stock_message': product.get('out_of_stock_message'),
                    'list_price': 0.0
                }

                # -------------------------------------------------
                # 🔹 Leer líneas de atributos del producto origen
                # -------------------------------------------------

                attribute_lines = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.template.attribute.line', 'search_read',
                    [[('product_tmpl_id', '=', product['id'])]],
                    {'fields': ['attribute_id', 'value_ids']}
                )

                # -------------------------
                # Categoría interna
                # -------------------------
                if product.get('categ_id'):
                    categ_name = product['categ_id'][1]
                    categ_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.category', 'search',
                        [[('name', '=', categ_name)]],
                        {'limit': 1}
                    )
                    if categ_ids:
                        vals['categ_id'] = categ_ids[0]

                # -------------------------
                # Marca
                # -------------------------
                if product.get('product_brand_id'):
                    brand_name = product['product_brand_id'][1]
                    brand_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.brand', 'search',
                        [[('name', '=', brand_name)]],
                        {'limit': 1}
                    )
                    if brand_ids:
                        vals['product_brand_id'] = brand_ids[0]
                    else:
                        vals['product_brand_id'] = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.brand', 'create',
                            [{'name': brand_name}]
                        )

                # -------------------------
                # Categorías web
                # -------------------------
                if product.get('public_categ_ids'):
                    web_categ_ids = []
                    for wc in product['public_categ_ids']:
                        wc_name = models_src.execute_kw(
                            db_src, uid_src, pwd_src,
                            'product.public.category', 'read',
                            [[wc]],
                            {'fields': ['name']}
                        )[0]['name']

                        dst_wc = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.public.category', 'search',
                            [[('name', '=', wc_name)]],
                            {'limit': 1}
                        )

                        if dst_wc:
                            web_categ_ids.append(dst_wc[0])

                    if web_categ_ids:
                        vals['public_categ_ids'] = [(6, 0, web_categ_ids)]

                # -------------------------
                # Crear producto
                # -------------------------
                new_id = models_dst.execute_kw(
                    db_dst, uid_dst, pwd_dst,
                    'product.template', 'create',
                    [vals]
                )

                print(f"➕ Creado {vals.get('default_code')} (ID {new_id})")

                # -------------------------------------------------
                # 5️⃣ Migrar atributos
                # -------------------------------------------------
                if con_atributos:
                    for line in attribute_lines:

                        attribute_name = line['attribute_id'][1]

                        # Buscar atributo en destino
                        attr_ids = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.attribute', 'search',
                            [[('name', '=', attribute_name)]],
                            {'limit': 1}
                        )

                        if attr_ids:
                            attr_id = attr_ids[0]
                        else:
                            # Crear atributo si no existe
                            attr_id = models_dst.execute_kw(
                                db_dst, uid_dst, pwd_dst,
                                'product.attribute', 'create',
                                [{'name': attribute_name}]
                            )

                        # -------------------------------------------------
                        # Migrar valores del atributo
                        # -------------------------------------------------

                        value_ids_dst = []

                        for value_id in line['value_ids']:

                            value_data = models_src.execute_kw(
                                db_src, uid_src, pwd_src,
                                'product.attribute.value', 'read',
                                [[value_id]],
                                {'fields': ['name']}
                            )[0]

                            value_name = value_data['name']

                            # Buscar valor en destino
                            val_ids = models_dst.execute_kw(
                                db_dst, uid_dst, pwd_dst,
                                'product.attribute.value', 'search',
                                [[
                                    ('name', '=', value_name),
                                    ('attribute_id', '=', attr_id)
                                ]],
                                {'limit': 1}
                            )

                            if val_ids:
                                val_id = val_ids[0]
                            else:
                                # Crear valor
                                val_id = models_dst.execute_kw(
                                    db_dst, uid_dst, pwd_dst,
                                    'product.attribute.value', 'create',
                                    [{
                                        'name': value_name,
                                        'attribute_id': attr_id
                                    }]
                                )

                            value_ids_dst.append(val_id)

                        # -------------------------------------------------
                        # Crear línea de atributo en el nuevo template
                        # -------------------------------------------------

                        models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.template.attribute.line', 'create',
                            [{
                                'product_tmpl_id': new_id,
                                'attribute_id': attr_id,
                                'value_ids': [(6, 0, value_ids_dst)]
                            }]
                        )

                # -------------------------------------------------
                # 5️⃣ Galería
                # -------------------------------------------------

                images = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', '=', product['id'])]],
                    {'fields': ['image_1920', 'name', 'video_url', 'sequence']}
                )

                for img in images:
                    models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.image', 'create',
                        [{
                            'product_tmpl_id': new_id,
                            'image_1920': img.get('image_1920'),
                            'name': img.get('name'),
                            'video_url': img.get('video_url'),
                            'sequence': img.get('sequence', 0)
                        }]
                    )

            except Exception as e:
                print(f"❌ Error en {product.get('default_code')}: {e}")

        print("✅ Migración por SKU finalizada")

    import xmlrpc.client
    #region LLAMADA
    # -----------------------
    # Configuración conexiones
    # -----------------------

    origen = {
        'url': "https://optimaluz.soluntec.net",
        'db': "Real",  # "Test"
        'user': "jcoronado@optimaluz.com",
        'password': "AlAi4ever"
    }

    destino = {
        'url': "http://79.72.61.76:8069",
        'db': "odoo0",  # "odoo0",
        'user': "admin",
        'password': "admin"
    }

    # -----------------------
    # Conexión ORIGEN
    # -----------------------

    common_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_src = common_src.authenticate(
        origen['db'],
        origen['user'],
        origen['password'],
        {}
    )

    if not uid_src:
        raise Exception("❌ No se pudo autenticar en ORIGEN")

    models_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a ORIGEN (uid={uid_src})")

    # -----------------------
    # Conexión DESTINO
    # -----------------------

    common_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_dst = common_dst.authenticate(
        destino['db'],
        destino['user'],
        destino['password'],
        {}
    )

    if not uid_dst:
        raise Exception("❌ No se pudo autenticar en DESTINO")

    models_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a DESTINO (uid={uid_dst})")

    # -----------------------
    # LLAMADA A LA MIGRACIÓN
    # -----------------------

    migrate_products_by_sku(
        models_src=models_src,
        db_src=origen['db'],
        uid_src=uid_src,
        pwd_src=origen['password'],

        models_dst=models_dst,
        db_dst=destino['db'],
        uid_dst=uid_dst,
        pwd_dst=destino['password'],

        excel_path='ruta'
    )
    #endregion

def transferencia_productos_archivados():
    def migrate_archived_products(models_src, db_src, uid_src, pwd_src,
                                  models_dst, db_dst, uid_dst, pwd_dst):
        """
        Migra productos archivados (active=False) desde origen a destino,
        manteniéndolos archivados y usando solo los campos solicitados.
        """

        FIELDS = [
            'image_1920',
            'name',
            'default_code',
            'invoice_policy',
            'standard_price',
            'categ_id',
            'product_brand_id',
            'x_url',
            'description',
            'out_of_stock_message'
        ]

        # 🔍 Buscar productos archivados en origen
        product_ids = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'search',
            [[('active', '=', False)]]
        )

        if not product_ids:
            print("ℹ️ No hay productos archivados para migrar")
            return

        products = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'read',
            [product_ids],
            {'fields': FIELDS}
        )

        print(f"📦 Productos archivados encontrados: {len(products)}")

        for product in products:
            try:
                vals = {}

                # --- Campos simples ---
                vals['image_1920'] = product.get('image_1920')
                vals['name'] = product.get('name')
                vals['default_code'] = product.get('default_code')
                vals['invoice_policy'] = product.get('invoice_policy')
                vals['standard_price'] = product.get('standard_price') or 0.0
                vals['x_url'] = product.get('x_url')
                vals['description'] = product.get('description')
                vals['out_of_stock_message'] = product.get('out_of_stock_message')

                # 🔒 Forzar archivado y precio
                vals['active'] = False
                vals['list_price'] = 0.0

                # --- Categoría ---
                if product.get('categ_id'):
                    categ_name = product['categ_id'][1]
                    categ_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.category', 'search',
                        [[('name', '=', categ_name)]],
                        {'limit': 1}
                    )
                    if categ_ids:
                        vals['categ_id'] = categ_ids[0]

                # --- Marca ---
                if product.get('product_brand_id'):
                    brand_name = product['product_brand_id'][1]
                    brand_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.brand', 'search',
                        [[('name', '=', brand_name)]],
                        {'limit': 1}
                    )
                    if brand_ids:
                        vals['product_brand_id'] = brand_ids[0]
                    else:
                        new_brand = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.brand', 'create',
                            [{'name': brand_name}]
                        )
                        vals['product_brand_id'] = new_brand

                # --- Buscar si existe en destino ---
                existing_ids = []
                if vals.get('default_code'):
                    existing_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.template', 'search',
                        [[('default_code', '=', vals['default_code'])]],
                        {'limit': 1}
                    )

                new_id = models_dst.execute_kw(
                    db_dst, uid_dst, pwd_dst,
                    'product.template', 'create',
                    [vals]
                )
                print(f"➕ Creado (archivado): {vals['default_code']} (ID {new_id})")

                print("✅ Migración de productos archivados finalizada")
            except Exception as e:
                print(f"❌ Error en {product.get('default_code')}: {e}")

    import xmlrpc.client

    # -----------------------
    # Configuración conexiones
    # -----------------------

    origen = {
        'url': "https://optimaluz.soluntec.net",
        'db': "Test",  # "Real"
        'user': "jcoronado@optimaluz.com",
        'password': "AlAi4ever"
    }

    destino = {
        'url': "http://79.72.61.76:8069",
        'db': "odoo0",
        'user': "admin",
        'password': "admin"
    }

    # -----------------------
    # Conexión ORIGEN
    # -----------------------

    common_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_src = common_src.authenticate(
        origen['db'],
        origen['user'],
        origen['password'],
        {}
    )

    if not uid_src:
        raise Exception("❌ No se pudo autenticar en ORIGEN")

    models_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a ORIGEN (uid={uid_src})")

    # -----------------------
    # Conexión DESTINO
    # -----------------------

    common_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_dst = common_dst.authenticate(
        destino['db'],
        destino['user'],
        destino['password'],
        {}
    )

    if not uid_dst:
        raise Exception("❌ No se pudo autenticar en DESTINO")

    models_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a DESTINO (uid={uid_dst})")

    # -----------------------
    # LLAMADA A LA MIGRACIÓN
    # -----------------------

    migrate_archived_products(
        models_src=models_src,
        db_src=origen['db'],
        uid_src=uid_src,
        pwd_src=origen['password'],

        models_dst=models_dst,
        db_dst=destino['db'],
        uid_dst=uid_dst,
        pwd_dst=destino['password']
    )

# endregion

#region REGION UTIL QUIMI
def get_precio_compra(excel_path, output_missing="skus_no_encontrados.xlsx"):
    from App_Connection import models, db, uid, password
    import openpyxl

    try:
        # --- Abrir Excel en modo edición ---
        wb = openpyxl.load_workbook(excel_path)
        if "Sheet1" not in wb.sheetnames:
            print("❌ El Excel no tiene hoja 'Sheet1'")
            return
        ws = wb["Sheet1"]

        # --- Localizar cabeceras ---
        headers = {cell.value: idx for idx, cell in
                   enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
        required = ["SKU", "STOCK", "PRECIO DE COMPRA", "VALOR STOCK"]
        for col in required:
            if col not in headers:
                print(f"❌ Falta la columna obligatoria: {col}")
                return

        col_sku = headers["SKU"]
        col_stock = headers["STOCK"]
        col_precio = headers["PRECIO DE COMPRA"]
        col_valor_stock = headers["VALOR STOCK"]

        # --- Lista para SKUs no encontrados ---
        skus_no_encontrados = []

        # --- Iterar filas ---
        for row in ws.iter_rows(min_row=2, values_only=False):
            sku_cell = row[col_sku - 1]
            stock_cell = row[col_stock - 1]
            precio_cell = row[col_precio - 1]
            valor_stock_cell = row[col_valor_stock - 1]

            sku = str(sku_cell.value).strip() if sku_cell.value else ""
            if not sku:
                continue

            try:
                # Buscar producto en Odoo
                prod_ids = models.execute_kw(
                    db, uid, password,
                    "product.template", "search",
                    [[("default_code", "=", sku)]],
                    {"limit": 1}
                )

                if not prod_ids:
                    print(f"⚠️ Producto no encontrado: {sku}")
                    skus_no_encontrados.append(sku)
                    continue

                prod_id = prod_ids[0]

                # Buscar supplierinfo
                supp_ids = models.execute_kw(
                    db, uid, password,
                    "product.supplierinfo", "search_read",
                    [[("product_tmpl_id", "=", prod_id)]],
                    {"fields": ["price"], "limit": 1}
                )

                if supp_ids:
                    precio_compra = supp_ids[0]["price"]
                    precio_cell.value = precio_compra

                    try:
                        stock_val = float(stock_cell.value) if stock_cell.value else 0
                        valor_stock_cell.value = round(precio_compra * stock_val, 2)
                    except Exception:
                        valor_stock_cell.value = ""
                    print(f"✅ {sku} → Precio: {precio_compra}, Valor Stock actualizado")
                else:
                    print(f"⚠️ Sin supplierinfo para: {sku}")
                    skus_no_encontrados.append(sku)

            except Exception as e:
                print(f"❌ Error procesando {sku}: {e}")
                skus_no_encontrados.append(sku)

        # --- Guardar Excel principal actualizado ---
        wb.save(excel_path)
        print(f"✅ Archivo actualizado: {excel_path}")

        # --- Exportar SKUs no encontrados ---
        if skus_no_encontrados:
            df_out = pd.DataFrame({"SKU": skus_no_encontrados})
            df_out.to_excel(output_missing, index=False)
            print(f"⚠️ SKUs no encontrados guardados en: {output_missing}")
        else:
            print("✅ Todos los SKUs fueron encontrados en Odoo")

    except Exception as e:
        print(f"❌ Error general: {e}")

def detectar_cambios_excel(ruta_excel):
    """
    Compara las hojas Odoo16 y Odoo18 y crea la hoja 'Cambios' con:
      - default_code
      - Solo las columnas que difieren (valores tomados de Odoo16)
        Si el valor fue eliminado (vacío en 16 pero no en 18) se marca con '*'.
    Las demás columnas quedan vacías.
    """

    # Leer las hojas
    df16 = pd.read_excel(ruta_excel, sheet_name='ODOO16')
    df18 = pd.read_excel(ruta_excel, sheet_name='ODOO18')

    if 'default_code' not in df16.columns or 'default_code' not in df18.columns:
        raise ValueError("Falta la columna 'default_code' en una o ambas hojas.")

    # Limpiar duplicados y vacíos
    df16 = df16.dropna(subset=['default_code']).drop_duplicates(subset=['default_code'])
    df18 = df18.dropna(subset=['default_code']).drop_duplicates(subset=['default_code'])

    # Indexar por default_code
    df16 = df16.set_index('default_code')
    df18 = df18.set_index('default_code')

    comunes = df16.index.intersection(df18.index)
    columnas_comunes = [col for col in df16.columns if col in df18.columns]

    cambios = []

    for code in comunes:
        fila16 = df16.loc[code]
        fila18 = df18.loc[code]

        dif_cols = [
            col for col in columnas_comunes
            if str(fila16[col]).strip() != str(fila18[col]).strip()
        ]

        if dif_cols:
            fila_resultado = {col: "" for col in columnas_comunes}  # vacío por defecto
            fila_resultado['default_code'] = code

            for col in dif_cols:
                val16 = str(fila16[col]).strip()
                val18 = str(fila18[col]).strip()

                # Si el valor fue eliminado (antes existía y ahora está vacío)
                if val16 in ["", "nan", "None"] and val18 not in ["", "nan", "None"]:
                    fila_resultado[col] = "*"  # marca eliminación
                else:
                    fila_resultado[col] = fila16[col]  # valor normal de Odoo16

            cambios.append(fila_resultado)

    if not cambios:
        print("✅ No se detectaron diferencias entre Odoo16 y Odoo18.")
        return

    # Crear DataFrame con las mismas columnas
    df_cambios = pd.DataFrame(cambios)
    columnas_finales = ['default_code'] + [c for c in columnas_comunes if c != 'default_code']
    df_cambios = df_cambios[columnas_finales]

    # Guardar resultados
    with pd.ExcelWriter(ruta_excel, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_cambios.to_excel(writer, sheet_name='CAMBIOS', index=False)

    print(f"💾 {len(df_cambios)} productos con diferencias guardados en 'CAMBIOS'.")

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    def resaltar_cambios_no_default_code(
            ruta_excel,
            hoja="CAMBIOS",
            columna_default_code="default_code"
    ):
        wb = load_workbook(ruta_excel)
        ws = wb[hoja]

        # Estilo amarillo
        amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # ---------------------------------------
        # 1️⃣ Detectar índice de columna default_code
        # ---------------------------------------
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        if columna_default_code not in headers:
            raise ValueError(f"No existe la columna '{columna_default_code}'")

        col_default = headers[columna_default_code]

        # ---------------------------------------
        # 2️⃣ Obtener todos los valores default_code
        # ---------------------------------------
        default_codes = set()
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_default).value
            if val not in (None, ""):
                default_codes.add(str(val).strip())

        # ---------------------------------------
        # 3️⃣ Recorrer resto de columnas
        # ---------------------------------------
        for col_name, col_idx in headers.items():
            if col_name == columna_default_code:
                continue

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value

                if val in (None, ""):
                    continue

                if str(val).strip() not in default_codes:
                    cell.fill = amarillo

        # ---------------------------------------
        # 4️⃣ Guardar cambios
        # ---------------------------------------
        wb.save(ruta_excel)

        print("✅ Valores resaltados correctamente en amarillo")

    resaltar_cambios_no_default_code(ruta_excel)

def aplicar_cambios_desde_excel(
        ruta_excel,
        hoja_cambios="CAMBIOS",
        columna_ref="default_code"):
    """
    Lee la hoja CAMBIOS, detecta celdas en amarillo (FFFF00),
    y actualiza SOLO esas columnas en Odoo, buscando por default_code.
    """
    from App_Connection import models, db, uid, password
    from openpyxl import load_workbook

    COLOR_AMARILLO = "FFFF00"

    # === 1. Cargar Excel ===
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb[hoja_cambios]

    # Obtener cabeceras
    headers = [c.value for c in ws[1]]
    col_index = {headers[i]: i + 1 for i in range(len(headers))}

    # === 2. Recopilar cambios detectando celdas amarillas ===
    cambios = []

    for row in ws.iter_rows(min_row=2):
        ref_cell = row[col_index[columna_ref] - 1]
        ref_value = str(ref_cell.value).strip() if ref_cell.value else ""

        if not ref_value:
            continue  # ignoramos filas sin referencia

        for col_name, idx in col_index.items():
            cell = row[idx - 1]
            fill = cell.fill

            # Detectar el amarillo
            is_yellow = (
                    fill
                    and fill.fgColor
                    and fill.fgColor.rgb is not None
                    and fill.fgColor.rgb[-6:].upper() == COLOR_AMARILLO
            )

            if is_yellow:
                cambios.append({
                    "default_code": ref_value,
                    "columna": col_name,
                    "valor": cell.value,
                })

    if not cambios:
        print("No se encontraron cambios resaltados en amarillo.")
        return

    # === 3. Procesar cambios en Odoo ===
    for cambio in cambios:
        ref = cambio["default_code"]
        col = cambio["columna"]
        val = cambio["valor"]

        # Buscar el producto por default_code
        product_ids = models.execute_kw(
            db, uid, password,
            "product.template", "search",
            [[["default_code", "=", ref]]]
        )

        if not product_ids:
            print(f"⚠ No encontrado en Odoo → {ref}")
            continue

        # Actualizar SOLO la columna modificada
        try:
            models.execute_kw(
                db, uid, password,
                "product.template", "write",
                [product_ids, {col: val}]
            )
            print(f"✔ Actualizado {ref}: {col} = {val}")

        except Exception as e:
            print(f"❌ Error actualizando {ref} ({col}): {e}")

    print("✔ Todos los cambios han sido aplicados a Odoo correctamente.")

#PREPARAR PRODUCTOS B2C
def borrar_notas_internas(sku_excel=False, excel_path=None):
    """
    Elimina la descripción interna de productos.
    - Si sku_excel=False: actúa sobre TODOS los productos (funcionamiento original)
    - Si sku_excel=True: actúa solo sobre productos cuyo default_code esté en el Excel
    """
    from App_Connection import models, db, uid, password
    import pandas as pd

    print("\n=== INICIANDO ELIMINACIÓN DE DESCRIPCIONES INTERNAS ===")

    # -------------------------------------------------------
    # 1) MODO FLEXIBLE: leer default_code desde Excel
    # -------------------------------------------------------
    product_ids = []

    if sku_excel:
        if not excel_path:
            print("❌ Debes indicar excel_path cuando sku_excel=True")
            return

        try:
            df = pd.read_excel(excel_path)

            if "default_code" not in df.columns:
                print("❌ El Excel no contiene columna 'default_code'")
                return

            default_codes = (
                df["default_code"]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )

            print(f"📄 Se encontraron {len(default_codes)} default_code en el Excel.")

            if not default_codes:
                print("⚠ No hay default_code válidos en el Excel.")
                return

            # Buscar solo productos coincidentes
            product_ids = models.execute_kw(
                db, uid, password,
                "product.product", "search",
                [[("default_code", "in", default_codes)]]
            )

            print(f"🔍 Productos encontrados por default_code: {len(product_ids)}")

        except Exception as e:
            print(f"❌ Error leyendo Excel: {e}")
            return

    else:
        # -------------------------------------------------------
        # 2) MODO ORIGINAL: borrar notas internas de TODOS
        # -------------------------------------------------------
        product_ids = models.execute_kw(
            db, uid, password,
            "product.product", "search",
            [[]]
        )

        print(f"🔍 Total productos encontrados: {len(product_ids)}")

    # -------------------------------------------------------
    # 3) Validación
    # -------------------------------------------------------
    if not product_ids:
        print("⚠ No se encontraron productos para procesar.")
        return

    # -------------------------------------------------------
    # 4) Write masivo
    # -------------------------------------------------------
    try:
        models.execute_kw(
            db, uid, password,
            "product.product", "write",
            [product_ids, {"description": False}]
        )

        print(f"✔ Descripciones internas eliminadas de {len(product_ids)} productos.")

    except Exception as e:
        print(f"❌ Error eliminando descripciones: {e}")

    print("\n✔ Proceso completado.")

def duplicar_precios_productos(sku_excel=False, excel_path=None):
    from App_Connection import models, db, uid, password
    import pandas as pd

    print("\n=== INICIANDO ACTUALIZACIÓN DE PRECIOS ===")

    # ======================================
    # 1) Obtener product_ids
    # ======================================
    if sku_excel:
        if not excel_path:
            print("❌ Debes indicar excel_path cuando sku_excel=True")
            return

        df = pd.read_excel(excel_path)

        if "default_code" not in df.columns:
            print("❌ El Excel no tiene columna default_code")
            return

        default_codes = (
            df["default_code"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        print(f"📄 SKUs encontrados en Excel: {len(default_codes)}")

        product_ids = models.execute_kw(
            db, uid, password,
            "product.product", "search",
            [[("default_code", "in", default_codes)]]
        )
    else:
        product_ids = models.execute_kw(
            db, uid, password,
            "product.product", "search",
            [[]]
        )

    if not product_ids:
        print("⚠ No se encontraron productos para actualizar.")
        return

    # ======================================
    # 2) Leer precios
    # ======================================
    productos = models.execute_kw(
        db, uid, password,
        "product.product", "read",
        [product_ids, ["id", "default_code", "standard_price"]]
    )

    # ======================================
    # 3) Actualizar precios
    # ======================================
    for p in productos:
        prod_id = p["id"]
        sku = p.get("default_code", "")
        sp = p["standard_price"] or 0.0

        nuevo_precio = sp * 2

        try:
            models.execute_kw(
                db, uid, password,
                "product.product", "write",
                [[prod_id], {
                    "standard_price": nuevo_precio,
                    "list_price": nuevo_precio
                }]
            )
            print(f"✔ {sku} | standard_price & list_price → {nuevo_precio}")

        except Exception as e:
            print(f"❌ Error actualizando {sku}: {e}")

    print("\n✔ Proceso completado correctamente.")

def create_pricelist_lines_and_activate_mto(
        company="ALMAITANA DE LUZ, S.L.",
        sku_excel=False,
        excel_path=None):

    from App_Connection import models, db, uid, password
    import pandas as pd

    # ======================================
    # 1. Buscar proveedor
    # ======================================
    partner_ids = models.execute_kw(
        db, uid, password,
        'res.partner', 'search',
        [[['name', '=', company]]]
    )

    if not partner_ids:
        print(f"❌ Proveedor '{company}' no encontrado")
        return

    supplier_id = partner_ids[0]
    print(f"✔️ Proveedor encontrado: {company} (ID {supplier_id})")

    # ID ruta MTO
    MTO_ROUTE_ID = 1

    # ======================================
    # 2. Obtener productos
    # ======================================
    if sku_excel:
        if not excel_path:
            print("❌ Debes indicar excel_path cuando sku_excel=True")
            return

        df = pd.read_excel(excel_path)
        if "default_code" not in df.columns:
            print("❌ El Excel no tiene columna default_code")
            return

        default_codes = (
            df["default_code"].dropna().astype(str).str.strip().unique().tolist()
        )

        product_ids = models.execute_kw(
            db, uid, password,
            'product.template', 'search',
            [[('default_code', 'in', default_codes)]]
        )

        print(f"📄 Productos filtrados por Excel: {len(product_ids)}")

    else:
        product_ids = models.execute_kw(
            db, uid, password,
            'product.template', 'search',
            [[]]
        )
        print(f"🔎 Productos encontrados: {len(product_ids)}")

    # ======================================
    # 3. Procesar productos
    # ======================================
    for pid in product_ids:

        product = models.execute_kw(
            db, uid, password,
            'product.template', 'read',
            [pid],
            {'fields': ['name', 'default_code', 'list_price', 'seller_ids', 'route_ids']}
        )[0]

        name = product.get('name')
        default_code = product.get('default_code') or ""
        seller_ids = product.get('seller_ids', [])
        pvp = product.get('list_price') or 0.0
        route_ids = product.get('route_ids', [])

        print(f"\n🔧 Procesando: {name}")

        # 4A. Crear línea de compra si no existe
        if not seller_ids:
            vals = {
                'partner_id': supplier_id,
                'product_tmpl_id': pid,
                'min_qty': 1,
                'product_name': name,
                'product_code': default_code,
                'price': pvp / 2
            }

            supplierinfo_id = models.execute_kw(
                db, uid, password,
                'product.supplierinfo', 'create',
                [vals]
            )

            print(f"   ✔️ Línea de compra creada → ID {supplierinfo_id}")
        else:
            print("   ⏭️ Ya tenía proveedor")

        # 4B. Activar ruta MTO si no está
        if MTO_ROUTE_ID not in route_ids:
            models.execute_kw(
                db, uid, password,
                'product.template', 'write',
                [[pid], {'route_ids': [(6, 0, route_ids + [MTO_ROUTE_ID])] }]
            )
            print("   ✔️ Ruta MTO activada")
        else:
            print("   ⏭️ Ruta MTO ya estaba activa")

    print("\n🎉 PROCESO COMPLETADO.")

def preparar_productos_b2c_cliente(sku_excel=False, excel_path=None):
    duplicar_precios_productos(sku_excel=sku_excel, excel_path=excel_path)
    create_pricelist_lines_and_activate_mto(sku_excel=sku_excel, excel_path=excel_path)

def publicar_todos_productos(batch_size=200):
    from App_Connection import models, db, uid, password
    import math

    try:
        # Buscar todos los productos
        product_ids = models.execute_kw(
            db, uid, password,
            'product.template', 'search',
            [[]]
        )

        total = len(product_ids)
        if not total:
            print("⚠️ No se encontraron productos en la base de datos.")
            return

        print(f"🔄 Publicando {total} productos en bloques de {batch_size}...")

        # Calcular número de bloques
        num_batches = math.ceil(total / batch_size)
        publicados = 0

        # Procesar por lotes
        for i in range(num_batches):
            start = i * batch_size
            end = min(start + batch_size, total)
            batch = product_ids[start:end]

            # Actualizar lote
            models.execute_kw(
                db, uid, password,
                'product.template', 'write',
                [batch, {'is_published': True}]
            )

            publicados += len(batch)
            print(f"✅ Bloque {i + 1}/{num_batches} → {len(batch)} productos publicados ({publicados}/{total})")

        print(f"\n🎯 Proceso completado: {publicados} productos publicados correctamente.")

    except Exception as e:
        print(f"❌ Error al publicar productos: {e}")

def importar_stock_desde_excel(ruta_excel, ubicacion='Almai/Stock Almacén'):#Almai/Stock Almacén #WH/Stock
    from App_Connection import models, db, uid, password
    import pandas as pd

    try:
        df = pd.read_excel(ruta_excel)

        if not {'SKU', 'STOCK'}.issubset(df.columns):
            print("❌ El Excel debe tener las columnas 'SKU' y 'STOCK'.")
            return

        print(f"📦 Importando stock a la ubicación '{ubicacion}'...")
        total = len(df)
        actualizados = 0

        # Buscar la ubicación por nombre
        location_ids = models.execute_kw(
            db, uid, password,
            'stock.location', 'search',
            [[('complete_name', '=', ubicacion)]]
        )
        if not location_ids:
            print(f"❌ No se encontró la ubicación '{ubicacion}'.")
            return

        location_id = location_ids[0]

        for i, row in df.iterrows():
            sku = str(row['SKU']).strip() if pd.notna(row['SKU']) else None
            qty = float(row['STOCK']) if pd.notna(row['STOCK']) else 0

            if not sku:
                continue

            try:
                # Buscar producto por default_code
                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.product', 'search',
                    [[('default_code', '=', sku)]]
                )
                if not product_ids:
                    print(f"⚠️ [{i + 1}] Producto no encontrado: {sku}")
                    continue

                product_id = product_ids[0]

                # Buscar si ya existe un stock.quant para ese producto y ubicación
                quant_ids = models.execute_kw(
                    db, uid, password,
                    'stock.quant', 'search',
                    [[('product_id', '=', product_id), ('location_id', '=', location_id)]]
                )

                if quant_ids:
                    # Actualizar cantidad existente
                    models.execute_kw(
                        db, uid, password,
                        'stock.quant', 'write',
                        [quant_ids, {'quantity': qty}]
                    )
                else:
                    # Crear nuevo registro de stock.quant
                    models.execute_kw(
                        db, uid, password,
                        'stock.quant', 'create',
                        [{
                            'product_id': product_id,
                            'location_id': location_id,
                            'quantity': qty,
                        }]
                    )

                actualizados += 1
                print(f"✅ [{i + 1}] {sku} → cantidad = {qty}")

            except Exception as e:
                print(f"⚠️ Error procesando {sku}: {e}")

        print(f"\n🎯 Proceso completado: {actualizados}/{total} productos actualizados correctamente.")

    except Exception as e:
        print(f"❌ Error general: {e}")

#STOCK LEDSPRESS: 1. rellenar_stock_odoo16 2. rellenar_stock_ledspress 3. importar_stock_desde_excel 4. create_pricelist_lines_and_activate_mto 5. publicar_todos_productos
def rellenar_stock_odoo16(ruta_excel):
    from App_Connection import models, db, uid, password
    import pandas as pd
    from openpyxl import load_workbook

    try:
        # Leer Excel
        df = pd.read_excel(ruta_excel)

        if 'SKU' not in df.columns:
            print("❌ No existe la columna 'SKU' en el Excel.")
            return

        # Asegurar columna de destino
        if 'ODOO16' not in df.columns:
            df['ODOO16'] = ""

        total = len(df)
        encontrados = 0

        print(f"🔍 Buscando stock de {total} productos en Odoo...")

        # Iterar por los SKUs
        for i, sku in enumerate(df['SKU'], start=1):
            if pd.isna(sku):
                continue

            try:
                # Buscar producto por default_code
                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.product', 'search_read',
                    [[('default_code', '=', str(sku))]],
                    {'fields': ['qty_available']}
                )

                if product_ids:
                    stock = product_ids[0]['qty_available']
                    df.at[i - 1, 'ODOO16'] = stock
                    encontrados += 1
                    print(f"✅ [{i}/{total}] {sku} → Stock: {stock}")
                else:
                    print(f"⚠️ [{i}/{total}] SKU no encontrado: {sku}")

            except Exception as e:
                print(f"⚠️ Error procesando {sku}: {e}")

        # Guardar resultados en el mismo archivo Excel
        with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False)

        print(f"\n📦 Proceso completado. {encontrados}/{total} productos encontrados en Odoo.")
        print(f"💾 Archivo actualizado: {ruta_excel}")

    except Exception as e:
        print(f"❌ Error general: {e}")

def rellenar_stock_ledspress(ruta_excel):
    from App_Connection import models, db, uid, password

    try:
        # Leer Excel principal
        df = pd.read_excel(ruta_excel, sheet_name="Sheet1")

        required_cols = ["SKU", "ODOO16", "M_STOCK", "B_STOCK", "UNDELIVERED ORDER"]
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"❌ Falta la columna obligatoria '{col}' en el Excel.")

        print(f"📦 Procesando {len(df)} productos para actualizar stock y mensajes en Odoo...")

        # Mensajes predefinidos
        DESC_LOCAL = """
        <br>
        <br>
        Almacén local / 1-2 Días entrega"""
        DESC_MUNDIAL = """
        <br>
        <br>
        Almacén mundial / 7-9 Días entrega
        """
        DESC_PROX = "<br><br>Disponible próximamente"
        DESC_NO = "<br><br>No Disponible"

        MSG_LOCAL = "Haga ya su pedido y recibirá las nuevas existencias que estan llegando a nuestro almacén en un plazo de 1 a 2 días hábiles."
        MSG_MUNDIAL = "Haga ya su pedido y recibirá las nuevas existencias que estan llegando a nuestro almacén en un plazo de 7 a 9 días hábiles."
        MSG_PROX = (
            "STOCK EN PROCESO DE PRODUCCIÓN Y DISPONIBLE PRÓXIMAMENTE.<br>"
            "Haga ya su reserva y nos pondremos en contacto con usted para indicarle la fecha exacta de entrada del producto "
            "o bien ofrecerle un producto similar en stock inmediato.<br>"
            "Reserva de productos:<br>"
            "- Email: ventas@ledspress.com<br>"
            "- WhatsApp / Telegram: (+34) 678 863 635<br>"
        )
        MSG_NO = (
            "STOCK NO DISPONIBLE.<br>"
            "Comuníquese con nosotros para ofrecerle un producto similar en stock inmediato.<br>"
            "Contacto:<br>"
            "- Email: ventas@ledspress.com<br>"
            "- WhatsApp / Telegram: (+34) 678 863 635<br>"
        )

        actualizados = 0

        for i, row in df.iterrows():
            sku = str(row["SKU"]).strip() if pd.notna(row["SKU"]) else None
            if not sku:
                continue

            try:
                # Buscar producto en Odoo
                products = models.execute_kw(
                    db, uid, password,
                    "product.template", "search_read",
                    [[("default_code", "=", sku)]],
                    {"fields": ["id", "description", "out_of_stock_message"]}
                )

                if not products:
                    print(f"⚠️ SKU no encontrado: {sku}")
                    continue

                product_id = products[0]["id"]
                description = products[0].get("description") or ""
                out_msg = ""

                # Valores desde el Excel
                qty_odoo16 = row["ODOO16"] if pd.notna(row["ODOO16"]) else 0
                qty_madrid = row["M_STOCK"] if pd.notna(row["M_STOCK"]) else 0
                qty_bulg = row["B_STOCK"] if pd.notna(row["B_STOCK"]) else 0
                undelivered = row["UNDELIVERED ORDER"] if pd.notna(row["UNDELIVERED ORDER"]) else None

                # Campos a actualizar
                vals = {}

                # --- ODOO16 ---
                vals["qty_available"] = qty_odoo16
                if qty_odoo16 > 0:
                    vals["description"] = DESC_LOCAL

                # --- M_STOCK ---
                vals["x_transit_stock_custom"] = qty_madrid
                if qty_madrid > 0:
                    vals["description"] = DESC_LOCAL
                    vals["out_of_stock_message"] = MSG_LOCAL

                # --- B_STOCK ---
                vals["x_almacen1_custom"] = qty_bulg
                if qty_bulg > 0:
                    desc_actual = vals.get("description", description)
                    vals["description"] = desc_actual + DESC_MUNDIAL
                    if qty_madrid <= 0:
                        vals["out_of_stock_message"] = MSG_MUNDIAL

                # --- UNDELIVERED ORDER ---
                if undelivered and qty_madrid <= 0 and qty_bulg <= 0:
                    if qty_odoo16 <= 0: vals["description"] = DESC_PROX
                    vals["out_of_stock_message"] = MSG_PROX
                elif not undelivered and qty_madrid <= 0 and qty_bulg <= 0:
                    if qty_odoo16 <= 0: vals["description"] = DESC_NO
                    vals["out_of_stock_message"] = MSG_NO

                # Actualizar en Odoo
                models.execute_kw(
                    db, uid, password,
                    "product.template", "write",
                    [[product_id], vals]
                )

                actualizados += 1
                print(f"✅ [{i + 1}] {sku} actualizado correctamente.")

            except Exception as e:
                print(f"⚠️ Error procesando {sku}: {e}")

        print(f"\n✅ Proceso completado: {actualizados}/{len(df)} productos actualizados en Odoo.")

    except Exception as e:
        print(f"❌ Error general: {e}")
#END PREPARAR PRODUCTOS B2C

#Comprobar Stocks de una lista de SKUs en hojas: ODOO1, MADRID2, BULGARIA3
#Hojas y columnas:
#ALMACEN: SKU, ODOO, MADRID, BULGARIA
#ODOO1: SKU, STOCK
#MADRID2: SKU, STOCK
#BULGARIA3: SKU, STOCK, UDELIVERED ORDER
def rellenar_stocks(ruta_excel):
    # Leer todas las hojas necesarias
    df_almacen   = pd.read_excel(ruta_excel, sheet_name="ALMACEN")
    df_odoo      = pd.read_excel(ruta_excel, sheet_name="ODOO1")
    df_madrid    = pd.read_excel(ruta_excel, sheet_name="MADRID2")
    df_bulgaria  = pd.read_excel(ruta_excel, sheet_name="BULGARIA3")

    # --- Normalizar SKU a string (evita errores con números) ---
    for df in [df_almacen, df_odoo, df_madrid, df_bulgaria]:
        df["SKU"] = df["SKU"].astype(str).str.strip()

    # Asegurar que las columnas a rellenar existen
    for col in ["ODOO", "MADRID", "BULGARIA"]:
        if col not in df_almacen.columns:
            df_almacen[col] = ""

    # ---------------------------------------------------------
    # RELLENAR COLUMNA ODOO
    # ---------------------------------------------------------
    if "STOCK" not in df_odoo.columns:
        raise ValueError("La hoja ODOO1 debe contener una columna STOCK.")

    df_almacen["ODOO"] = df_almacen["SKU"].map(
        df_odoo.set_index("SKU")["STOCK"]
    )

    # ---------------------------------------------------------
    # RELLENAR COLUMNA MADRID
    # ---------------------------------------------------------
    if "STOCK" not in df_madrid.columns:
        raise ValueError("La hoja MADRID2 debe contener una columna STOCK.")

    df_almacen["MADRID"] = df_almacen["SKU"].map(
        df_madrid.set_index("SKU")["STOCK"]
    )

    # ---------------------------------------------------------
    # RELLENAR COLUMNA BULGARIA (con reglas especiales)
    # ---------------------------------------------------------
    if "STOCK" not in df_bulgaria.columns:
        raise ValueError("La hoja BULGARIA3 debe contener una columna STOCK.")
    if "UNDELIVERED ORDER" not in df_bulgaria.columns:
        raise ValueError("La hoja BULGARIA3 debe contener columna UNDELIVERED ORDER.")

    # Mapeo base (STOCK)
    bulg_stock = df_bulgaria.set_index("SKU")["STOCK"]
    bulg_undelivered = df_bulgaria.set_index("SKU")["UNDELIVERED ORDER"]

    def get_bulgaria_value(sku):
        stk = bulg_stock.get(sku, None)
        und = bulg_undelivered.get(sku, None)

        # Caso especial:
        # Si UNDELIVERED ORDER tiene valor
        # y STOCK está vacío o <= 0 => usar UNDELIVERED ORDER
        if pd.notna(und) and (pd.isna(stk) or stk <= 0):
            return und

        # Caso normal: usar STOCK
        return stk

    df_almacen["BULGARIA"] = df_almacen["SKU"].apply(get_bulgaria_value)

    # ---------------------------------------------------------
    # GUARDAR EN EL MISMO EXCEL SIN BORRAR OTRAS HOJAS
    # ---------------------------------------------------------
    with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_almacen.to_excel(writer, sheet_name="ALMACEN", index=False)

    print("✔ Hoja ALMACEN actualizada correctamente.")

def actualizar_stock_custom(ruta_excel):
    from App_Connection import models, db, uid, password

    # Leer las hojas necesarias
    df_madrid = pd.read_excel(ruta_excel, sheet_name="MADRID2")
    df_bulgaria = pd.read_excel(ruta_excel, sheet_name="BULGARIA3")

    # Normalizar SKU
    df_madrid["SKU"] = df_madrid["SKU"].astype(str).str.strip()
    df_bulgaria["SKU"] = df_bulgaria["SKU"].astype(str).str.strip()

    # ---- MADRID2 → x_transit_stock_custom ------------------------------------
    if "STOCK" not in df_madrid.columns:
        raise ValueError("MADRID2 debe contener columna STOCK")

    print("\n=== Actualizando x_transit_stock_custom desde MADRID2 ===")

    for _, row in df_madrid.iterrows():
        sku = row["SKU"]
        stock = row["STOCK"]

        # Buscar producto por default_code
        product_ids = models.execute_kw(
            db, uid, password,
            "product.product", "search",
            [[["default_code", "=", sku]]]
        )

        if not product_ids:
            print(f"⚠ Producto con SKU {sku} no encontrado en Odoo.")
            continue

        try:
            models.execute_kw(
                db, uid, password,
                "product.product", "write",
                [product_ids, {"x_transit_stock_custom": stock}]
            )
            print(f"✔ SKU {sku} → x_transit_stock_custom = {stock}")
        except Exception as e:
            print(f"❌ Error actualizando SKU {sku}: {e}")

    # ---- BULGARIA3 → x_almacen1_custom --------------------------------------
    if "STOCK" not in df_bulgaria.columns:
        raise ValueError("BULGARIA3 debe contener columna STOCK")

    print("\n=== Actualizando x_almacen1_custom desde BULGARIA3 ===")

    for _, row in df_bulgaria.iterrows():
        sku = row["SKU"]
        stock = row["STOCK"]

        product_ids = models.execute_kw(
            db, uid, password,
            "product.product", "search",
            [[["default_code", "=", sku]]]
        )

        if not product_ids:
            print(f"⚠ Producto con SKU {sku} no encontrado en Odoo.")
            continue

        try:
            models.execute_kw(
                db, uid, password,
                "product.product", "write",
                [product_ids, {"x_almacen1_custom": stock}]
            )
            print(f"✔ SKU {sku} → x_almacen1_custom = {stock}")
        except Exception as e:
            print(f"❌ Error actualizando SKU {sku}: {e}")

    print("\n✔ Proceso completado.")

#endregion