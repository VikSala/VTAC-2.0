import math
from collections import namedtuple

from pywin.Demos.ocx.ocxtest import test1

from services.campos_odoo import ClavesExcel, excel_a_odoo, FISCAL_POSITION_MAP
from services.utils import Utils
from collections import defaultdict
import xmlrpc.client
import pandas as pd
import numpy as np
import ast

from services.utils_download_files import ruta_excel

is_connected = False
atributos_cache = {}
valores_cache = {}

#region Importar Productos
def assign_variant_skus(product_template_id, referencias, attribute_variable, params):
    """
    Asigna los SKUs a las variantes generadas por Odoo, respetando el orden.
    """
    models, db, uid, password = params

    # 1) Leer variantes creadas por Odoo
    variants = models.execute_kw(
        db, uid, password,
        'product.product', 'search_read',
        [[('product_tmpl_id', '=', product_template_id)]],
        {'fields': ['id', 'product_template_attribute_value_ids']}
    )

    # 2) Ordenar las variantes por el valor del atributo variable
    # Esto es CLAVE para que coincidan con referencias[0], referencias[1], etc.
    variants_sorted = sorted(
        variants,
        key=lambda v: v['product_template_attribute_value_ids']
    )

    # 3) Asignar SKU por orden
    for variant, ref in zip(variants_sorted, referencias):
        models.execute_kw(
            db, uid, password,
            'product.product', 'write',
            [[variant['id']], {'default_code': ref}]
        )
        print(f"🔧 Asignado SKU {ref} → variante {variant['id']}")

'''def attributes_to_odoo(product_template_id, specifications, params):
    if isinstance(specifications, str):
        try:
            specifications = ast.literal_eval(specifications)
        except Exception as e:
            print(f"❌ Error convirtiendo especificaciones: {e}")
            specifications = {}

    # Agregar atributos
    try:
        if isinstance(specifications, dict):
            for key, value in specifications.items():
                #print(f"🛠️ Key: {key} | Valor: {value}")
                if value:
                    attr_id = Utils.get_or_create_attribute(key, atributos_cache, params)
                    value_id = Utils.get_or_create_attribute_value(attr_id, value, valores_cache, params)
                    Utils.create_attribute_line(product_template_id, attr_id, value_id, params)
    except Exception as e:
        print(f"❌ Error convirtiendo especificaciones: {e}")'''
def attributes_to_odoo(product_template_id, specifications, params, row=None):
    # Convertir string a dict/list si viene del Excel
    if isinstance(specifications, str):
        try:
            specifications = ast.literal_eval(specifications)
        except Exception:
            specifications = {}

    # ---------------------------
    # CASO 1 → sin variantes
    # ---------------------------
    if isinstance(specifications, dict):

        for key, value in specifications.items():
            if not value:
                continue

            attr_id = Utils.get_or_create_attribute(key, atributos_cache, params)
            value_id = Utils.get_or_create_attribute_value(attr_id, value, valores_cache, params)

            Utils.create_attribute_line(product_template_id, attr_id, value_id, params)

        return  # FIN sin variantes

    # ---------------------------
    # CASO 2 → con variantes
    # ---------------------------
    if isinstance(specifications, list):

        if len(specifications) == 0:
            return

        comunes = specifications[0]

        # 1) Crear atributos comunes
        for key, value in comunes.items():
            attr_id = Utils.get_or_create_attribute(key, atributos_cache, params)
            value_id = Utils.get_or_create_attribute_value(attr_id, value, valores_cache, params)

            Utils.create_attribute_line(product_template_id, attr_id, value_id, params)

        # 2) Detectar atributo variable
        atributo_variable = list(specifications[1].keys())[0]

        attr_var_id = Utils.get_or_create_attribute(atributo_variable, atributos_cache, params)

        # 3) Crear valores del atributo variable
        valores_variantes_ids = []
        for variante_dict in specifications[1:]:
            valor = list(variante_dict.values())[0]
            value_id = Utils.get_or_create_attribute_value(attr_var_id, valor, valores_cache, params)
            valores_variantes_ids.append(value_id)

        # 4) Crear línea que generará las variantes
        params.models.execute_kw(
            params.db, params.uid, params.password,
            'product.template.attribute.line', 'create',
            [{
                'product_tmpl_id': product_template_id,
                'attribute_id': attr_var_id,
                'value_ids': [(6, 0, valores_variantes_ids)]
            }]
        )

        # 5) Asignar SKUs a variantes generadas por Odoo
        referencias = row["SKU"] if "SKU" in row else None

        if referencias:
            try:
                referencias = ast.literal_eval(referencias)
                assign_variant_skus(product_template_id, referencias, atributo_variable, params)
            except Exception as e:
                print(f"⚠️ No se pudieron asignar SKUs de variantes: {e}")


# === IMPORT: Leer Excel y crear Productos en Odoo ===
def import_to_odoo(excel_path, limite = 0, import_all = False, event_manager=None):
    from App_Connection import models, db, uid, password

    NOVEDADES = False
    LEDXPRESS = False
    ADVANCE = False
    ODOO16 = False #True

    ConnectionParams = namedtuple("ConnectionParams", ["models", "db", "uid", "password"])
    params = ConnectionParams(models, db, uid, password)

    def emit_progress(message):
        if event_manager:
            event_manager.emit('progress_update', message)
            print(message)
        else:
            print(message)

    if import_all:
        df = pd.read_excel(excel_path, sheet_name="Products")
    else:
        df = pd.read_excel(excel_path, sheet_name="Import")

    total = len(df)
    emit_progress(f"📦 Encontrados {total} productos para importar")

    # === Iterar filas y crear productos ===
    for index, row in df.iloc[limite:].iterrows():
        try:
            current_product = index + 1
            product_name = row.get(ClavesExcel.NOMBRE.value, f"Producto {current_product}")

            emit_progress(f"📦 Importando producto {current_product}/{total}: {product_name}")

            name = row[ClavesExcel.NOMBRE.value]
            product_data = Utils.build_product_data_from_row(row, Utils, df)

            print("1. Categorías")# Categorías
            #if not ODOO16: Utils.preparar_categorias_para_producto(row, product_data, params)

            print("2. Marca")#Marca
            Utils.preparar_marca_para_producto(row, product_data, params)

            #Iconos Novedades
            if NOVEDADES:
                print("3. Iconos")
                atributos = ast.literal_eval(row["Atributos"])
                product_data['x_icono1'] = Utils.image_url_to_base64("http://79.72.55.217:8069/web/image/1011-1dd7909f/V-TAC.png")
                icons_detected = Utils.ico_match(product_data['name'], False)
                if isinstance(icons_detected, str):
                    icons_detected = icons_detected.split(',')
                    ico_count = 2
                    for icon in icons_detected:
                        if icon != '':
                            product_data['x_icono' + str(ico_count)] = Utils.image_url_to_base64(icon)
                            ico_count += 1
                    for attr_name, val_name in atributos.items():
                        if not attr_name or not val_name:
                            continue

                        ico_str = Utils.ico_match(attr_name, val_name)  # Devuelve ico_url o None
                        if ico_str and ico_count < 9:
                            product_data['x_icono' + str(ico_count)] = Utils.image_url_to_base64(ico_str)
                            ico_count += 1
            elif LEDXPRESS:
                product_data['x_icono1'] = Utils.image_url_to_base64("http://79.72.55.217:8069/web/image/1011-1dd7909f/V-TAC.png")
                led_icons = ast.literal_eval(row["Iconos"])
                ico_count = 2
                for icon in led_icons:
                    product_data['x_icono' + str(ico_count)] = Utils.image_url_to_base64(icon)
                    ico_count += 1
            elif ADVANCE:
                product_data['x_icono1'] = Utils.image_url_to_base64(row["Logo"])

            if LEDXPRESS: product_data['image_1920'] = Utils.image_url_to_base64(row["Imagen principal"])

            print("3,5. Crear producto base...")
            product_template_id = models.execute_kw(
                db, uid, password,
                'product.template', 'create',
                [product_data]
            )

            # 3.6 Crear la regla de precio (pricelist item) con el descuento
            if LEDXPRESS:
                descuento = row["Descuento"]
                if descuento and not (isinstance(descuento, float) and math.isnan(descuento)):
                    try:
                        print("3,6. Crear descuento...")
                        # ID de la tarifa a la que quieres aplicar el descuento
                        # Puedes usar 1 si es la 'Public Pricelist' (tarifa pública)
                        pricelist_id = 1

                        # Leer el valor del descuento desde tu Excel
                        discount_value = float(descuento)

                        # Crear la regla de descuento
                        pricelist_item_id = models.execute_kw(
                            db, uid, password,
                            'product.pricelist.item', 'create',
                            [{
                                'pricelist_id': pricelist_id,
                                'applied_on': '1_product',  # Aplicar a un producto concreto
                                'product_tmpl_id': product_template_id,  # ID del producto recién creado
                                'compute_price': 'percentage',  # Tipo de cálculo: porcentaje
                                'percent_price': discount_value,  # Ej: -20 para un -20%
                            }]
                        )

                        print(f"✅ Pricelist creada (ID: {pricelist_item_id}) con descuento {discount_value}%")

                    except Exception as e:
                        print(f"⚠️ Error creando pricelist para producto {product_template_id}: {e}")

            print("4. Atributos")# Atributos
            specifications = row[ClavesExcel.ATRIBUTOS.value]
            attributes_to_odoo(product_template_id, specifications, params, row)

            print("5. Videos")# Agregar imágenes adicionales si existen
            galeria = ClavesExcel.GALERIA.value
            video = ClavesExcel.VIDEOS.value
            is_video = False
            if video in row and pd.notna(row[video]):
                url_video = ast.literal_eval(row[video])
                if url_video:
                    url_video = ast.literal_eval(row[video])[0]
                    url_img = "https://i.ytimg.com/vi/" + url_video.split("=")[1] + "/sddefault.jpg"
                    is_video = True
                    image_id = models.execute_kw(
                        db, uid, password,
                        'product.image', 'create',
                        [{
                            'name': f"{name}-yt",
                            'product_tmpl_id': product_template_id,
                            'image_1920': Utils.image_url_to_base64(url_img),
                            'video_url': url_video,
                            'sequence': 0
                        }]
                    )

            print("5. Imágenes")
            if galeria in row and pd.notna(row[galeria]):
                try:
                    extra_urls = ast.literal_eval(row[galeria])
                    for i, url in enumerate(extra_urls, start=1):

                        #if "http" not in url: url = "http://143.47.53.74:8070/" + url
                        campo_odoo = 'image_1920' if "youtu" not in url else 'video_url'
                        valor_odoo = Utils.image_url_to_base64(url.strip()) if "youtu" not in url else url
                        ext = ".webp" if "youtu" not in url else ""

                        if valor_odoo:
                            ctx = {
                                "active_test": False,
                                "lang": "es_ES",
                                "interactive": False,
                            }
                            image_id = models.execute_kw(
                                db, uid, password,
                                'product.image', 'create',
                                [{
                                    'name': f"{name}-extra-{i}{ext}",
                                    'product_tmpl_id': product_template_id,
                                    campo_odoo: valor_odoo,
                                    'sequence': i+1 if is_video else i
                                }],
                                {"context": ctx}
                            )
                except Exception as e:
                    msg = f"⚠️ Error procesando imágenes adicionales para {product_name}: {e}"
                    emit_progress(msg)
                    print(msg)

            emit_progress(f"✅ Producto {current_product}/{total} importado: {product_name}")

        except Exception as e:
            msg = f"❌ Error importando producto {current_product}/{total}: {e}"
            emit_progress(msg)
            print(msg)
            continue

    if import_all and not NOVEDADES and not LEDXPRESS and not ADVANCE and not ODOO16:
        '''def extraer_iconos_por_sku(excel_file_path, batch_size=100):
            # 1. Conexión a Odoo
            url_b2b = 'https://optimaluz.soluntec.net'
            db_b2b = 'Real'
            username_b2b = 'Usuario programación'
            password_b2b = 'Programación'

            common_b2b = xmlrpc.client.ServerProxy(f'{url_b2b}/xmlrpc/2/common', allow_none=True)
            uid_b2b = common_b2b.authenticate(db_b2b, username_b2b, password_b2b, {})

            if not uid:
                print("❌ No se pudo autenticar.")
                return

            print(f'🔌 Conectado como {username_b2b} (uid: {uid})')
            models_b2b = xmlrpc.client.ServerProxy(f'{url_b2b}/xmlrpc/2/object', allow_none=True)

            try:
                # Leer Excel y extraer SKUs únicos
                df = pd.read_excel(excel_file_path)
                if 'SKU' not in df.columns:
                    raise ValueError("❌ El Excel debe contener una columna llamada 'SKU'.")

                skus = df['SKU'].dropna().astype(str).unique().tolist()
                resultado = {}

                # Procesar en batches
                for i in range(0, len(skus), batch_size):
                    batch_skus = skus[i:i + batch_size]
                    print(f"🔄 Procesando batch {i + 1}-{i + len(batch_skus)} / {len(skus)}")

                    # Buscar productos en Odoo por default_code
                    domain = [('default_code', 'in', batch_skus)]
                    fields = ['default_code'] + [f'x_icono{i}' for i in range(1, 9)]

                    products = models_b2b.execute_kw(
                        db_b2b, uid_b2b, password_b2b,
                        'product.template', 'search_read',
                        [domain],
                        {'fields': fields}
                    )

                    # Mapear iconos por SKU
                    for product in products:
                        sku = product['default_code']
                        iconos = [product.get(f'x_icono{i}') for i in range(1, 9) if product.get(f'x_icono{i}')]
                        resultado[sku] = iconos

                    # Añadir SKUs no encontrados (vacíos)
                    encontrados = {p['default_code'] for p in products}
                    no_encontrados = set(batch_skus) - encontrados
                    for sku in no_encontrados:
                        resultado[sku] = []
                        print(f"⚠️ SKU no encontrado en Odoo: {sku}")

                print(f"\n✅ Iconos extraídos correctamente para {len(resultado)} productos.")
                return resultado

            except Exception as e:
                print(f"❌ Error extrayendo iconos: {e}")
                return {}

        iconos_por_sku_ = extraer_iconos_por_sku(excel_path)

        def actualizar_iconos_por_sku(excel_file_path, iconos_por_sku, batch_size=100):
            # 2. Leer Excel con columna SKU
            df = pd.read_excel(excel_file_path)
            if 'SKU' not in df.columns:
                print("❌ El Excel debe contener una columna llamada 'SKU'.")
                return

            skus = df['SKU'].dropna().astype(str).unique().tolist()
            total = len(skus)
            actualizados = 0
            no_encontrados = 0

            # 3. Procesar por lotes
            for i in range(0, total, batch_size):
                batch_skus = skus[i:i + batch_size]

                # Buscar productos por default_code en lote
                productos = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search_read',
                    [[('default_code', 'in', batch_skus)]],
                    {'fields': ['id', 'default_code']}
                )

                sku_to_id = {p['default_code']: p['id'] for p in productos}

                for sku in batch_skus:
                    template_id = sku_to_id.get(sku)
                    iconos = iconos_por_sku.get(sku, [])

                    if not template_id:
                        print(f"❌ Producto con SKU '{sku}' no encontrado en Odoo.")
                        no_encontrados += 1
                        continue

                    if not iconos:
                        print(f"⚠️ Sin iconos para SKU: {sku}")
                        continue

                    # Construir campos para write()
                    valores_iconos = {}
                    for j, icono_b64 in enumerate(iconos):
                        if j >= 8:
                            break  # Solo hasta x_icono8
                        campo = f'x_icono{j + 1}'
                        valores_iconos[campo] = icono_b64

                    try:
                        models.execute_kw(
                            db, uid, password,
                            'product.template', 'write',
                            [[template_id], valores_iconos]
                        )
                        actualizados += 1
                        print(f"✅ Iconos actualizados para SKU: {sku}")
                    except Exception as e:
                        print(f"❌ Error al actualizar SKU '{sku}': {e}")

            print(f"\n🔁 Proceso completado: {actualizados} productos actualizados, {no_encontrados} no encontrados.")

        actualizar_iconos_por_sku(excel_path, iconos_por_sku_)'''

        def extraer_media_por_sku(excel_path, batch_size=100):
            import pandas as pd
            import xmlrpc.client

            url = 'https://optimaluz.soluntec.net'
            db_o = 'Real'
            user = 'Usuario programación'
            pwd = 'Programación'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid_o = common.authenticate(db_o, user, pwd, {})
            models_o = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            df = pd.read_excel(excel_path)
            skus = df['SKU'].dropna().astype(str).unique().tolist()

            resultado = {}

            for i in range(0, len(skus), batch_size):
                batch = skus[i:i + batch_size]

                templates = models_o.execute_kw(
                    db_o, uid_o, pwd,
                    'product.template', 'search_read',
                    [[('default_code', 'in', batch)]],
                    {'fields': ['id', 'default_code', 'image_1920'] + [f'x_icono{i}' for i in range(1, 9)]}
                )

                tmpl_by_id = {t['id']: t for t in templates}
                sku_to_tmpl = {t['default_code']: t['id'] for t in templates}

                images = models_o.execute_kw(
                    db_o, uid_o, pwd,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', 'in', list(tmpl_by_id))]],
                    {'fields': ['product_tmpl_id', 'image_1920', 'sequence']}
                )

                images_by_tmpl = {}
                for img in images:
                    images_by_tmpl.setdefault(img['product_tmpl_id'][0], []).append(img)

                for sku, tmpl_id in sku_to_tmpl.items():
                    t = tmpl_by_id[tmpl_id]
                    resultado[sku] = {
                        "iconos": [t.get(f'x_icono{i}') for i in range(1, 9) if t.get(f'x_icono{i}')],
                        "main": t.get('image_1920'),
                        "gallery": images_by_tmpl.get(tmpl_id, [])
                    }

            return resultado

        def aplicar_media_por_sku(excel_path, media_por_sku, batch_size=100):
            import pandas as pd

            df = pd.read_excel(excel_path)
            skus = df['SKU'].dropna().astype(str).unique().tolist()

            for i in range(0, len(skus), batch_size):
                batch = skus[i:i + batch_size]

                templates = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search_read',
                    [[('default_code', 'in', batch)]],
                    {'fields': ['id', 'default_code']}
                )

                sku_to_id = {t['default_code']: t['id'] for t in templates}

                for sku in batch:
                    media = media_por_sku.get(sku)
                    tmpl_id = sku_to_id.get(sku)

                    if not media or not tmpl_id:
                        continue

                    valores = {}

                    # Iconos
                    for i, icono in enumerate(media['iconos'][:8]):
                        valores[f'x_icono{i + 1}'] = icono

                    # Imagen principal
                    if media.get("main"):
                        valores['image_1920'] = media['main']

                    if valores:
                        models.execute_kw(
                            db, uid, password,
                            'product.template', 'write',
                            [[tmpl_id], valores]
                        )

                    # Galería
                    for img in media.get("gallery", []):
                        if not img.get("image_1920"):
                            continue
                        models.execute_kw(
                            db, uid, password,
                            'product.image', 'create',
                            [{
                                'product_tmpl_id': tmpl_id,
                                'image_1920': img['image_1920'],
                                'sequence': img.get('sequence', 10)
                            }]
                        )

        media_por_sku_ = extraer_media_por_sku(excel_path)
        aplicar_media_por_sku(excel_path, media_por_sku_)

    emit_progress("🎉 Importación de productos completada.")

# endregion

# region Actualizar Productos

def update_from_merge_to_odoo(excel_path, event_manager):
    """
    Actualiza todos los productos a partir de los datos en la hoja 'Update' del Excel.
    """
    from App_Connection import models, db, uid, password
    ConnectionParams = namedtuple("ConnectionParams", ["models", "db", "uid", "password"])
    params = ConnectionParams(models, db, uid, password)

    def emit_progress(message):
        if event_manager:
            event_manager.emit('progress_update', message)
            print(message)
        else:
            print(message)

    def update_odoo_product(product_sku, update_values):
        """
        Actualiza un product.template en Odoo por sku.
        """
        # Buscar producto por sku
        product_ids = models.execute_kw(
            db, uid, password,
            'product.template', 'search',
            [[['default_code', '=', product_sku]]]
        )

        if product_ids:
            # Actualizar el primer producto encontrado
            success = models.execute_kw(
                db, uid, password,
                'product.template', 'write',
                [[product_ids[0]], update_values]
            )
            return success
        else:
            print(f"Producto '{product_sku}' no encontrado.")
            return False

    df = pd.read_excel(excel_path, sheet_name="Update")
    if df.empty or df.dropna(how='all').empty:
        emit_progress("ℹ️ No hay productos para actualizar")
        return
    # Limpiar filas vacías o sin datos válidos
    df = df.dropna(subset=["SKU", "Actualizar"])
    total = len(df)
    emit_progress(f"🔄 Encontrados {total} productos para actualizar")

    for index, row in df.iterrows():
        sku = row["SKU"]
        current_product = index + 1
        campos_a_actualizar = row["Actualizar"]
        emit_progress(f"🔄 Actualizando producto {current_product}/{total}: {sku}")

        # Si es string, evalúa el diccionario (en caso de que Excel lo haya guardado como texto)
        if isinstance(campos_a_actualizar, str):
            try:
                campos_a_actualizar = eval(campos_a_actualizar)
            except Exception as e:
                print(f"⚠️ Error al interpretar los campos de '{sku}': {e}")
                continue

        # Si se incluye ATRIBUTOS, tratarlo por separado
        atributos = ClavesExcel.ATRIBUTOS.value
        if atributos in campos_a_actualizar:
            product_id = models.execute_kw(
                db, uid, password,
                'product.template', 'search',
                [[['default_code', '=', sku]]]
            )
            if product_id:
                attributes_to_odoo(product_id[0], campos_a_actualizar[atributos], params)
                campos_a_actualizar.pop(atributos)  # Evitar actualizarlo también por write()

        # Si hay más campos, hacer una única actualización
        campos_odoo = {}

        for clave_excel, valor in campos_a_actualizar.items():
            clave_excel_norm = clave_excel.strip()
            if clave_excel_norm in excel_a_odoo:
                clave_odoo = excel_a_odoo[clave_excel_norm]
                campos_odoo[clave_odoo] = valor
            else:
                print(f"⚠️ Clave no encontrada en mapeo: '{clave_excel_norm}'")

        if campos_odoo:
            actualizado = update_odoo_product(
                product_sku=sku,
                update_values=campos_odoo
            )

            if actualizado:
                print(f"✅ Producto '{sku}' actualizado correctamente.")
            else:
                print(f"❌ Fallo al actualizar '{sku}'.")

    emit_progress("🎉 Actualización de productos completada.")

def delete_from_merge_to_odoo(excel_path, event_manager=None):
    from App_Connection import models, db, uid, password

    def clean_value(val):
        if isinstance(val, (np.generic, pd._libs.tslibs.nattype.NaTType)):
            return val.item()
        elif pd.isna(val):
            return None
        return val

    def emit_progress(message):
        if event_manager:
            event_manager.emit('progress_update', message)
            print(message)
        else:
            print(message)

    try:
        df = pd.read_excel(excel_path, sheet_name='Quantity')
    except Exception as e:
        emit_progress(f"⚠️ No se encontró hoja 'Quantity' o error leyendo: {e}")
        return

    if df.empty or df.dropna(how='all').empty:
        emit_progress("ℹ️ No hay productos para eliminar")
        return

    productos_a_eliminar = df['Producto Eliminado'].dropna().unique()
    total = len(productos_a_eliminar)

    if total == 0:
        emit_progress("ℹ️ No hay productos marcados para eliminación")
        return

    emit_progress(f"🗑️ Encontrados {total} productos para eliminar")

    for index, sku in enumerate(productos_a_eliminar):
        try:
            sku = clean_value(sku)
            current_product = index + 1
            emit_progress(f"🗑️ Eliminando producto {current_product}/{total}: {sku}")

            # Buscar producto por nombre
            product_ids = models.execute_kw(
                db, uid, password,
                'product.template', 'search',
                [[('default_code', '=', sku)]]
            )

            if product_ids:
                # Desactivar en lugar de eliminar (más seguro)
                models.execute_kw(
                    db, uid, password,
                    'product.template', 'write',
                    [[product_ids[0]], {'active': False}]
                )
                emit_progress(f"✅ Producto {current_product}/{total} desactivado: {sku}")
            else:
                emit_progress(f"⚠️ Producto no encontrado {current_product}/{total}: {sku}")

        except Exception as e:
            emit_progress(f"❌ Error eliminando producto {current_product}/{total}: {e}")
            continue

    emit_progress("🎉 Eliminación de productos completada.")

#Publicar por Categoria/SKUs
def main_publish(por_categorias: bool, event_manager=None):
    from App_Connection import models, db, uid, password

    def emit_progress(message):
        if event_manager:
            event_manager.emit('progress_update', message)
            print(message)
        else:
            print(message)

    try:
        emit_progress("📁 Seleccionando archivo Excel...")
        ruta_excel = Utils.seleccionar_excel()

        if not ruta_excel:
            emit_progress("❌ Publicación cancelada - No se seleccionó archivo.")
            if event_manager:
                event_manager.emit('operation_error', "No se seleccionó archivo para publicar")
            return

        emit_progress("📊 Leyendo datos del archivo Excel...")
        df = pd.read_excel(ruta_excel)

        if por_categorias:
            if "Familia" not in df.columns:
                emit_progress("❌ No se encontró la columna 'Familia' en el Excel.")
                if event_manager:
                    event_manager.emit('operation_error', "Columna 'Familia' no encontrada")
                return

            categorias = df["Familia"].dropna().unique().tolist()
            total_categorias = len(categorias)
            emit_progress(f"🏷️ Publicando productos por categoría ({total_categorias} categorías)...")

            productos_publicados = 0
            for i, categoria in enumerate(categorias, 1):
                emit_progress(f"🏷️ Procesando categoría {i}/{total_categorias}: {categoria}")

                # Buscar ID de categoría por nombre
                categoria_ids = models.execute_kw(
                    db, uid, password,
                    'product.public.category', 'search',
                    [[('name', '=', categoria)]]
                )

                if not categoria_ids:
                    emit_progress(f"⚠️ Categoría '{categoria}' no encontrada en Odoo.")
                    continue

                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [[('categ_id', 'in', categoria_ids)]]
                )

                if product_ids:
                    models.execute_kw(
                        db, uid, password,
                        'product.template', 'write',
                        [product_ids, {'is_published': True}]
                    )
                    productos_publicados += len(product_ids)
                    emit_progress(
                        f"✅ Publicados {len(product_ids)} productos de la categoría '{categoria}' ({i}/{total_categorias})")
                else:
                    emit_progress(
                        f"⚠️ No se encontraron productos para la categoría '{categoria}' ({i}/{total_categorias})")

            emit_progress(f"🎉 Publicación por categorías completada: {productos_publicados} productos publicados")

        else:
            if "SKU" not in df.columns:
                emit_progress("❌ No se encontró la columna 'SKU' en el Excel.")
                if event_manager:
                    event_manager.emit('operation_error', "Columna 'SKU' no encontrada")
                return

            skus = df["SKU"].dropna().astype(str).tolist()
            total_skus = len(skus)
            emit_progress(f"🏷️ Publicando productos por SKU ({total_skus} valores)...")

            productos_publicados = 0
            for i, sku in enumerate(skus, 1):
                emit_progress(f"🏷️ Procesando SKU {i}/{total_skus}: {sku}")

                product_ids = models.execute_kw(
                    db, uid, password,
                    'product.template', 'search',
                    [[('default_code', '=', sku)]]
                )

                if product_ids:
                    models.execute_kw(
                        db, uid, password,
                        'product.template', 'write',
                        [product_ids, {'is_published': True}]
                    )
                    productos_publicados += 1
                    emit_progress(f"✅ Producto SKU '{sku}' publicado ({i}/{total_skus})")
                else:
                    emit_progress(f"⚠️ Producto con SKU '{sku}' no encontrado ({i}/{total_skus})")

            emit_progress(f"🎉 Publicación por SKU completada: {productos_publicados} productos publicados")

        # Emitir completion al final
        if event_manager:
            event_manager.emit('operation_completed', "Publicación finalizada correctamente")

    except Exception as e:
        emit_progress(f"❌ Error en publicar_productos: {e}")
        if event_manager:
            event_manager.emit('operation_error', f"Error en publicación: {e}")


# endregion

# region Ventana Conexión
import os
def abrir_ajustes():
    from tkinter import Toplevel, messagebox
    from pathlib import Path
    import tkinter as tk
    import json

    config_path = os.path.expanduser(
        "~/Documents/SMI Files/Data/perfiles.json")  # Path(__file__).parent / "config.json"
    ruta_script = Path(__file__).parent / "App_Connection.py"

    ajustes_win = Toplevel()
    ajustes_win.title("Ajustes")
    ajustes_win.geometry("400x300")
    ajustes_win.resizable(False, False)

    ajustes_win.transient()
    ajustes_win.grab_set()
    ajustes_win.focus_force()

    config_backup_path = os.path.expanduser("~/Documents/SMI Files/Data/perfiles_backup.json")

    def cargar_config():
        ruta = os.path.expanduser("~/Documents/SMI Files/Data")
        os.makedirs(ruta, exist_ok=True)
        if config_path:
            with open(config_path, "r") as f:
                try:
                    data = json.load(f)
                    perfil = data.get("current_profile", "default")
                    perfiles = data.get("profiles", {})
                    if not perfiles:
                        raise ValueError
                    return perfiles.get(perfil, {}), perfiles, perfil
                except:
                    pass

        # Si no hay archivo o está mal formado, crear uno por defecto
        default_config = {
            "current_profile": "default",
            "profiles": {
                "default": {
                    "URL": "",
                    "Database": "",
                    "Usuario": "",
                    "Password": ""
                }
            }
        }
        with open(config_path, "w") as f:
            json.dump(default_config, f, indent=4)

        return default_config["profiles"]["default"], default_config["profiles"], "default"

    # Cargar configuración y perfiles
    valores_actuales, perfiles, perfil_actual = cargar_config()

    perfil_var = tk.StringVar(value=perfil_actual or "default")

    # Selector de perfil arriba a la izquierda
    tk.Label(ajustes_win, text="Perfil:").place(x=10, y=10)
    perfil_dropdown = tk.OptionMenu(ajustes_win, perfil_var, *perfiles.keys())
    perfil_dropdown.place(x=60, y=5)

    # Entry + botón para crear nuevo perfil
    nuevo_perfil_var = tk.StringVar()
    tk.Entry(ajustes_win, textvariable=nuevo_perfil_var, width=12).place(x=240, y=10)
    tk.Button(ajustes_win, text="Nuevo", command=lambda: crear_nuevo_perfil()).place(x=330, y=5)

    # Frame central para los campos
    frame_central = tk.Frame(ajustes_win)
    frame_central.place(relx=0.5, rely=0.52, anchor="center")

    # Campos
    campos = {
        "URL": tk.StringVar(value=valores_actuales.get("URL", "")),
        "Database": tk.StringVar(value=valores_actuales.get("Database", "")),
        "Usuario": tk.StringVar(value=valores_actuales.get("Usuario", "")),
        "Password": tk.StringVar(value=valores_actuales.get("Password", ""))
    }

    def actualizar_guardar_si_cambios(*args):
        colocar_botones()

    for var in campos.values():
        var.trace_add("write", actualizar_guardar_si_cambios)

    for idx, (label_text, var) in enumerate(campos.items()):
        tk.Label(frame_central, text=label_text + ":").grid(row=idx, column=0, padx=10, pady=6, sticky='e')
        entry = tk.Entry(frame_central, textvariable=var, width=30)
        if label_text.lower() == "password":
            entry.config(show="*")
        entry.grid(row=idx, column=1, padx=10, pady=6)

    # Mensaje informativo
    mensaje_label = tk.Label(frame_central, text="", fg="green")
    mensaje_label.grid(row=4, column=0, columnspan=2, pady=5)

    def colocar_botones():
        y = 250
        x0 = 70 if perfil_var.get() != "default" else 100
        if hubo_cambios():
            guardar_btn.place(x=x0, y=y)
        else:
            guardar_btn.place_forget()

        conectar_btn.place(x=x0 + 100, y=y)

        if perfil_var.get() != "default":
            eliminar_btn.place(x=x0 + 200, y=y)
        else:
            eliminar_btn.place_forget()

    def hubo_cambios():
        return any(campos[k].get() != valores_actuales.get(k, "") for k in campos)

    def actualizar_campos_desde_perfil(*args):
        nonlocal valores_actuales
        perfil = perfil_var.get()
        datos = perfiles.get(perfil, {})
        for campo, var in campos.items():
            var.set(datos.get(campo, ""))
        valores_actuales = datos
        colocar_botones()

    perfil_var.trace_add("write", actualizar_campos_desde_perfil)

    def crear_nuevo_perfil():
        nuevo = nuevo_perfil_var.get().strip()
        if not nuevo:
            messagebox.showerror("Error", "Debes ingresar un nombre para el nuevo perfil.")
            return
        if nuevo in perfiles:
            messagebox.showinfo("Perfil existente", "Ese perfil ya existe.")
            return

        perfiles[nuevo] = {"URL": "", "Database": "", "Usuario": "", "Password": ""}
        perfil_var.set(nuevo)
        nuevo_perfil_var.set("")

        guardar_config(perfiles[nuevo], False, nuevo)
        perfil_dropdown["menu"].add_command(label=nuevo, command=tk._setit(perfil_var, nuevo))
        actualizar_campos_desde_perfil()

    def eliminar_perfil():
        perfil = perfil_var.get()
        if perfil == "default":
            messagebox.showwarning("Aviso", "No se puede eliminar el perfil 'default'.")
            return

        confirm = messagebox.askyesno("Eliminar perfil", f"¿Eliminar perfil '{perfil}'?")
        if confirm:
            del perfiles[perfil]
            nuevo = "default" if "default" in perfiles else list(perfiles.keys())[0]
            perfil_var.set(nuevo)
            guardar_config(perfiles[nuevo], False, nuevo, True)
            perfil_dropdown["menu"].delete(0, "end")
            for p in perfiles:
                perfil_dropdown["menu"].add_command(label=p, command=tk._setit(perfil_var, p))
            actualizar_campos_desde_perfil()

    def guardar_config(config, set_connection, perfil="default", was_deleted = False):

        data = {"current_profile": perfil, "profiles": perfiles}
        perfiles[perfil] = config
        with open(config_path, "w") as f:
            json.dump(data, f, indent=4)
        if not was_deleted:
            with open(config_backup_path, "w") as f:
                json.dump(data, f, indent=4)
            #ruta = os.path.expanduser("~/Documents/SMI Files/Data/perfiles_backup.json")
            #shutil.copy(config_backup_path, ruta)


        if set_connection:
            with open(ruta_script, "w", encoding="utf-8") as f:
                f.write(f"""import xmlrpc.client

url = '{config["URL"]}'
db = '{config["Database"]}'
username = '{config["Usuario"]}'
password = '{config["Password"]}'

common = None
uid = None
models = None

def conectar():
    global common, uid, models
    try:
        common = xmlrpc.client.ServerProxy(f'{{url}}/xmlrpc/2/common', allow_none=True)
        uid = common.authenticate(db, username, password, {{}})
        if uid:
            print(f'Conectado como {{username}} (uid: {{uid}})')
            models = xmlrpc.client.ServerProxy(f'{{url}}/xmlrpc/2/object', allow_none=True)
        else:
            print('Error de autenticación')
        return uid
    except Exception as e:
        print(f'❌ Error en la conexión: {{e}}')
        return None
""")

    def guardar_ajustes():
        config = {campo: var.get() for campo, var in campos.items()}
        perfil = perfil_var.get()
        if not all(config.values()):
            mensaje_label.config(text="Todos los campos deben estar completos.", fg="red")
            return
        guardar_config(config, False, perfil)
        mensaje_label.config(text="Ajustes guardados correctamente.", fg="green")
        actualizar_campos_desde_perfil()

    def conectar():
        global is_connected

        config = {campo: var.get() for campo, var in campos.items()}
        perfil = perfil_var.get()
        db = config["Database"]
        cambios = hubo_cambios()
        try:
            guardar_config(config, True, perfil)

            import importlib.util
            import sys
            spec = importlib.util.spec_from_file_location("App_Connection", str(ruta_script))
            App_Connection = importlib.util.module_from_spec(spec)
            sys.modules["App_Connection"] = App_Connection
            spec.loader.exec_module(App_Connection)

            uid = App_Connection.conectar()
            is_connected = uid
            if uid:
                mensaje_label.config(text=f"Conectado a: {db}", fg="green")
                ajustes_win.update()
                if cambios:
                    mensaje_label.config(text="Reiniciando...", fg="blue")
                    ajustes_win.update()
                    import time; time.sleep(1.2)
                    os.execl(sys.executable, sys.executable, *sys.argv)
                else: ajustes_win.destroy()
            else:
                mensaje_label.config(text=f"No se pudo conectar a: {db}", fg="red")

        except Exception as e:
            mensaje_label.config(text=f"Error al conectar: {e}", fg="red")

    # Botones
    guardar_btn = tk.Button(ajustes_win, text="Guardar", command=lambda: guardar_ajustes())
    conectar_btn = tk.Button(ajustes_win, text="Conectar", command=lambda: conectar())
    eliminar_btn = tk.Button(ajustes_win, text="Eliminar", command=lambda: eliminar_perfil())
    colocar_botones()
    ajustes_win.wait_window()

# endregion


def main_import_with_event_manager(event_manager):
    global is_connected, atributos_cache, valores_cache

    def emit_progress(message):
        if event_manager:
            event_manager.emit('progress_update', message)
            print(message)
        else:
            print(message)

    emit_progress("🔧 Iniciando proceso de conexión...")
    abrir_ajustes()

    if is_connected:
        from App_Connection import models, db, uid, password

        ConnectionParams = namedtuple("ConnectionParams", ["models", "db", "uid", "password"])
        params = ConnectionParams(models, db, uid, password)

        emit_progress("✅ Conexión establecida con Odoo")

        ruta = Utils.seleccionar_excel()

        atributos_cache = Utils.cargar_atributos_existentes(params)
        valores_cache = Utils.cargar_valores_atributos_existentes(params)

        url_src = 'http://37.59.66.189:8069/'#
        db_src = 'Real'#Test
        username_src = 'jcoronado@optimaluz.com'
        password_src = 'AlAi4ever' #AlAi4ever@optimaluz.com
        common = xmlrpc.client.ServerProxy(f"{url_src}/xmlrpc/2/common")#""#
        uid_src = common.authenticate(db_src, username_src, password_src, {})#""#
        models_src = xmlrpc.client.ServerProxy(f"{url_src}/xmlrpc/2/object")#""#

        def migrar_adjuntos_modelo(modelo, usar_lotes=True, limitar_fecha = False):
            """
            Migra adjuntos (ir.attachment) desde cualquier modelo.

            modelo: string del modelo (ej: 'project.task')
            usar_lotes: bool → si True migra en lotes de x
            """

            import math

            print(f"\n📎 Iniciando migración de adjuntos para modelo: {modelo}")

            total_migrados = 0
            total_existentes = 0
            total_sin_destino = 0

            domain = [("res_model", "=", modelo)] if not limitar_fecha \
                else [("res_model", "=", modelo), ("create_date", ">=", "2026-01-01"), ("create_date", "<",  "2027-01-01")]

            # ---------------------------------------
            # 🔹 1. Obtener IDs de adjuntos en origen
            # ---------------------------------------
            attach_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "ir.attachment", "search",
                [domain]
            )

            total_adjuntos = len(attach_ids)
            print(f"📤 Adjuntos encontrados en origen: {total_adjuntos}")

            if total_adjuntos == 0:
                return

            # ---------------------------------------
            # 🔹 2. Configuración de lotes
            # ---------------------------------------
            if usar_lotes:
                lote = 30
            else:
                lote = total_adjuntos  # todo en uno

            num_lotes = math.ceil(total_adjuntos / lote)

            # ---------------------------------------
            # 🔹 3. Migración
            # ---------------------------------------
            for i in range(num_lotes):

                inicio = i * lote
                fin = inicio + lote
                lote_ids = attach_ids[inicio:fin]

                #print(f"\n🔹 Procesando lote {i + 1}/{num_lotes} ({len(lote_ids)} adjuntos)")

                # Leer SOLO el lote actual
                adjuntos = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "ir.attachment", "read",
                    [lote_ids, ["name", "datas", "res_id", "mimetype"]]
                )

                for adj in adjuntos:

                    nombre = adj.get("name")
                    datos = adj.get("datas")
                    res_id_origen = adj.get("res_id")
                    mimetype = adj.get("mimetype")

                    if not datos:
                        continue

                    # Buscar registro equivalente en destino
                    registro_destino = models.execute_kw(
                        db, uid, password,
                        modelo, "search",
                        [[("x_id_interno", "=", res_id_origen)]],
                        {"limit": 1}
                    )

                    if not registro_destino:
                        total_sin_destino += 1
                        continue

                    res_id_destino = registro_destino[0]

                    # Verificar duplicado por nombre
                    existente = models.execute_kw(
                        db, uid, password,
                        "ir.attachment", "search",
                        [[
                            ("res_model", "=", modelo),
                            ("res_id", "=", res_id_destino),
                            ("name", "=", nombre)
                        ]],
                        {"limit": 1}
                    )

                    if existente:
                        total_existentes += 1
                        continue

                    # Crear adjunto
                    vals = {
                        "name": nombre,
                        "datas": datos,
                        "res_model": modelo,
                        "res_id": res_id_destino,
                        "mimetype": mimetype,
                        "type": "binary",
                    }

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "ir.attachment", "create",
                            [vals]
                        )
                        total_migrados += 1
                    except Exception as e:
                        print(f"❌ Error creando adjunto {nombre}: {e}")

            # ---------------------------------------
            # 🔹 Resumen final
            # ---------------------------------------
            print("\n📊 MIGRACIÓN DE ADJUNTOS COMPLETADA")
            print(f"   Migrados: {total_migrados}")
            print(f"   Ya existentes: {total_existentes}")
            print(f"   Sin registro destino: {total_sin_destino}")

        def sincronizar_tareas_desde_origen_por_proyecto():
            """
            1️⃣ Exporta las tareas del Odoo 16 (Test)
            2️⃣ Busca las líneas, usuarios y etiquetas equivalentes por nombre en Odoo 18
            3️⃣ Actualiza las tareas destino por coincidencia (nombre + proyecto)
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models
            import time

            # --- Conexión al Odoo origen (16) ---
            url_old = 'https://optimaluz.soluntec.net'
            db_old = 'Test'
            username_old = 'jcoronado@optimaluz.com'
            password_old = 'AlAi4ever'

            common_old = xmlrpc.client.ServerProxy(f'{url_old}/xmlrpc/2/common', allow_none=True)
            uid_old = common_old.authenticate(db_old, username_old, password_old, {})
            if not uid_old:
                print("❌ No se pudo autenticar en Odoo origen.")
                return

            models_old = xmlrpc.client.ServerProxy(f'{url_old}/xmlrpc/2/object', allow_none=True)
            print(f"🔌 Conectado a Odoo 16 ({db_old}) como {username_old}")

            # --- Exportar tareas con su proyecto ---
            FIELDS = ["id", "name", "project_id", "sale_line_id", "user_ids", "tag_ids"]
            tareas = models_old.execute_kw(db_old, uid_old, password_old,
                                           "project.task", "search_read",
                                           [[], FIELDS])
            print(f"📦 {len(tareas)} tareas exportadas desde Odoo 16.")

            total_actualizadas = 0
            total_no_encontradas = 0

            for t in tareas:
                name = t["name"]
                project_name = t["project_id"][1] if t.get("project_id") else None

                # --- Buscar proyecto en destino ---
                project_id = None
                if project_name:
                    projects = models.execute_kw(
                        db, uid, password,
                        "project.project", "search_read",
                        [[("name", "=", project_name)]],
                        {"fields": ["id"], "limit": 1}
                    )
                    if projects:
                        project_id = projects[0]["id"]

                if not project_id:
                    print(f"⚠️ Proyecto '{project_name}' no encontrado, se omite tarea '{name}'.")
                    total_no_encontradas += 1
                    continue

                # --- Buscar tarea destino (por nombre + proyecto) ---
                domain = [("name", "=", name), ("project_id", "=", project_id)]
                task_dest = models.execute_kw(
                    db, uid, password,
                    "project.task", "search",
                    [domain],
                    {"limit": 1}
                )
                if not task_dest:
                    print(f"⚠️ Tarea '{name}' no encontrada en proyecto '{project_name}'.")
                    total_no_encontradas += 1
                    continue

                task_id = task_dest[0]
                vals = {}

                # --- sale_line_id ---
                sale_line_name = t["sale_line_id"][1] if t.get("sale_line_id") else None
                if sale_line_name:
                    sale_line = models.execute_kw(
                        db, uid, password,
                        "sale.order.line", "search",
                        [[("name", "=", sale_line_name)]],
                        {"limit": 1}
                    )
                    if sale_line:
                        vals["sale_line_id"] = sale_line[0]
                    else:
                        print(f"⚠️ Línea de venta '{sale_line_name}' no encontrada para tarea '{name}'.")

                # --- user_ids ---
                user_ids = []
                if t.get("user_ids"):
                    users = models_old.execute_kw(
                        db_old, uid_old, password_old,
                        "res.users", "read",
                        [t["user_ids"]], {"fields": ["name"]}
                    )
                    for u in users:
                        found_user = models.execute_kw(
                            db, uid, password,
                            "res.users", "search",
                            [[("name", "=", u["name"])]],
                            {"limit": 1}
                        )
                        if found_user:
                            user_ids.append(found_user[0])
                if user_ids:
                    vals["user_ids"] = [(6, 0, user_ids)]

                # --- tag_ids ---
                tag_ids = []
                if t.get("tag_ids"):
                    tags = models_old.execute_kw(
                        db_old, uid_old, password_old,
                        "project.tags", "read",
                        [t["tag_ids"]], {"fields": ["name"]}
                    )
                    for tg in tags:
                        found_tag = models.execute_kw(
                            db, uid, password,
                            "project.tags", "search",
                            [[("name", "=", tg["name"])]],
                            {"limit": 1}
                        )
                        if found_tag:
                            tag_ids.append(found_tag[0])
                if tag_ids:
                    vals["tag_ids"] = [(6, 0, tag_ids)]

                # --- Actualizar tarea ---
                if vals:
                    try:
                        models.execute_kw(db, uid, password, "project.task", "write", [[task_id], vals])
                        print(f"✅ Tarea '{name}' ({project_name}) actualizada correctamente.")
                        total_actualizadas += 1
                    except Exception as e:
                        print(f"❌ Error actualizando tarea '{name}' ({project_name}): {e}")
                else:
                    print(f"ℹ️ No hay datos para actualizar en tarea '{name}' ({project_name}).")

                time.sleep(0.1)  # ligera pausa entre tareas

            print(f"\n📊 SINCRONIZACIÓN COMPLETADA:")
            print(f"   ✅ Tareas actualizadas: {total_actualizadas}")
            print(f"   ⚠️ No encontradas o sin proyecto: {total_no_encontradas}")

        #sincronizar_tareas_desde_origen_por_proyecto()

        #Fallo: Crear lineas en PVs en vez de en Proyectos
        #Corregir Servicios

        def sincronizar_tareas_con_creacion():
            """
            Sincroniza tareas desde Odoo 16 a Odoo 18:
              - Busca tarea por nombre + proyecto
              - Busca pedido existente por prefijo de línea (ej: PV-OPT/24/1107)
              - Si el pedido existe → crea la línea dentro
              - Si no existe → omite con aviso
              - Busca usuarios y etiquetas por nombre
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models
            import time

            # --- Odoo origen (16) ---
            url_old = 'https://optimaluz.soluntec.net'
            db_old = 'Test'
            username_old = 'jcoronado@optimaluz.com'
            password_old = 'AlAi4ever'

            common_old = xmlrpc.client.ServerProxy(f'{url_old}/xmlrpc/2/common', allow_none=True)
            uid_old = common_old.authenticate(db_old, username_old, password_old, {})
            if not uid_old:
                print("❌ No se pudo autenticar en Odoo origen.")
                return

            models_old = xmlrpc.client.ServerProxy(f'{url_old}/xmlrpc/2/object', allow_none=True)
            print(f"🔌 Conectado a Odoo 16 ({db_old}) como {username_old}")

            FIELDS = ["id", "name", "project_id", "sale_line_id", "user_ids", "tag_ids"]
            tareas = models_old.execute_kw(db_old, uid_old, password_old,
                                           "project.task", "search_read",
                                           [[], FIELDS])
            print(f"📦 {len(tareas)} tareas exportadas desde Odoo 16.")

            total_actualizadas = 0
            total_creadas = 0

            for t in tareas:
                name = t["name"]
                project_name = t["project_id"][1] if t.get("project_id") else None
                if not project_name:
                    print(f"⚠️ Tarea '{name}' sin proyecto, omitida.")
                    continue

                # Buscar proyecto destino
                project = models.execute_kw(
                    db, uid, password,
                    "project.project", "search_read",
                    [[("name", "=", project_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if not project:
                    print(f"⚠️ Proyecto '{project_name}' no encontrado, omitida '{name}'.")
                    continue
                project_id = project[0]["id"]

                # Buscar tarea destino (nombre + proyecto)
                domain = [("name", "=", name), ("project_id", "=", project_id)]
                task_dest = models.execute_kw(db, uid, password,
                                              "project.task", "search",
                                              [domain], {"limit": 1})

                vals = {}

                # --- sale_line_id ---
                sale_line_name = t["sale_line_id"][1] if t.get("sale_line_id") else None
                if sale_line_name:
                    # Detectar pedido por prefijo (ej: PV-OPT/24/1107)
                    pedido_name = sale_line_name.split(" - ")[0]
                    pedido = models.execute_kw(
                        db, uid, password,
                        "sale.order", "search_read",
                        [[("name", "=", pedido_name)]],
                        {"fields": ["id"], "limit": 1}
                    )
                    if pedido:
                        order_id = pedido[0]["id"]
                        # Buscar si ya existe línea en ese pedido
                        existing_line = models.execute_kw(
                            db, uid, password,
                            "sale.order.line", "search_read",
                            [[("order_id", "=", order_id), ("name", "=", sale_line_name)]],
                            {"fields": ["id"], "limit": 1}
                        )
                        if existing_line:
                            vals["sale_line_id"] = existing_line[0]["id"]
                        else:
                            try:
                                new_line = models.execute_kw(
                                    db, uid, password,
                                    "sale.order.line", "create",
                                    [{
                                        "order_id": order_id,
                                        "name": sale_line_name,
                                        "product_uom_qty": 1,
                                        "price_unit": 0.0
                                    }]
                                )
                                vals["sale_line_id"] = new_line
                                print(f"➕ Creada línea '{sale_line_name}' en pedido '{pedido_name}'")
                            except Exception as e:
                                print(f"⚠️ Error creando línea '{sale_line_name}': {e}")
                    else:
                        print(f"⚠️ Pedido '{pedido_name}' no encontrado, no se crea línea de venta.")

                # --- user_ids ---
                user_ids = []
                if t.get("user_ids"):
                    users_old = models_old.execute_kw(
                        db_old, uid_old, password_old,
                        "res.users", "read", [t["user_ids"]],
                        {"fields": ["name"]}
                    )
                    for u in users_old:
                        user_name = u["name"]
                        found_user = models.execute_kw(
                            db, uid, password,
                            "res.users", "search_read",
                            [[("name", "=", user_name)]],
                            {"fields": ["id"], "limit": 1}
                        )
                        if found_user:
                            user_ids.append(found_user[0]["id"])
                        else:
                            print(f"⚠️ Usuario '{user_name}' no encontrado en destino, omitido.")
                if user_ids:
                    vals["user_ids"] = [(6, 0, user_ids)]

                # --- tag_ids ---
                tag_ids = []
                if t.get("tag_ids"):
                    tags_old = models_old.execute_kw(
                        db_old, uid_old, password_old,
                        "project.tags", "read",
                        [t["tag_ids"]], {"fields": ["name"]}
                    )
                    for tg in tags_old:
                        tag_name = tg["name"]
                        found_tag = models.execute_kw(
                            db, uid, password,
                            "project.tags", "search_read",
                            [[("name", "=", tag_name)]],
                            {"fields": ["id"], "limit": 1}
                        )
                        if found_tag:
                            tag_ids.append(found_tag[0]["id"])
                        else:
                            new_tag = models.execute_kw(
                                db, uid, password,
                                "project.tags", "create",
                                [{"name": tag_name}]
                            )
                            tag_ids.append(new_tag)
                            print(f"➕ Etiqueta '{tag_name}' creada.")
                if tag_ids:
                    vals["tag_ids"] = [(6, 0, tag_ids)]

                # --- Crear o actualizar tarea ---
                if task_dest:
                    try:
                        models.execute_kw(db, uid, password,
                                          "project.task", "write",
                                          [[task_dest[0]], vals])
                        print(f"✅ Tarea '{name}' ({project_name}) actualizada.")
                        total_actualizadas += 1
                    except Exception as e:
                        print(f"❌ Error actualizando '{name}' ({project_name}): {e}")
                else:
                    vals["name"] = name
                    vals["project_id"] = project_id
                    try:
                        models.execute_kw(db, uid, password,
                                          "project.task", "create", [vals])
                        print(f"➕ Tarea '{name}' creada en '{project_name}'.")
                        total_creadas += 1
                    except Exception as e:
                        print(f"❌ Error creando '{name}' ({project_name}): {e}")

                time.sleep(0.1)

            print(f"\n📊 SINCRONIZACIÓN FINALIZADA:")
            print(f"   ✅ Tareas actualizadas: {total_actualizadas}")
            print(f"   ➕ Tareas creadas: {total_creadas}")

        def eliminar_por_sku(excel_ruta):
            from App_Connection import models, db, uid, password
            import pandas as pd

            # -------------------------------------------------
            # LEER EXCEL
            # -------------------------------------------------
            df = pd.read_excel(excel_ruta, sheet_name="Sheet1")

            if "SKU" not in df.columns:
                raise ValueError("La hoja Sheet1 debe contener una columna llamada 'SKU'.")

            # Normalizar SKUs a string
            df["SKU"] = df["SKU"].astype(str).str.strip()

            # -------------------------------------------------
            # RECORRER SKUs
            # -------------------------------------------------
            for sku in df["SKU"].dropna().unique():
                if sku == "" or sku.lower() == "nan":
                    continue

                # Buscar producto por default_code
                product_ids = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[["default_code", "=", sku]]]
                )

                if not product_ids:
                    print(f"⚠️ No encontrado en Odoo: {sku}")
                    continue

                # -------------------------------------------------
                # ELIMINAR PRODUCTO
                # -------------------------------------------------
                try:
                    models.execute_kw(
                        db, uid, password,
                        "product.product", "unlink",
                        [product_ids]
                    )
                    print(f"✔ Producto(s) eliminado(s) correctamente: {sku} (IDs: {product_ids})")

                except Exception as e:
                    print(f"❌ Error eliminando {sku}: {e}")

            print("✔ Proceso completado.")

        def resetear_campos_custom():
            from App_Connection import models, db, uid, password

            # Buscar todos los productos
            product_ids = models.execute_kw(
                db, uid, password,
                "product.product", "search",
                [[]]  # vacío → todos los productos
            )

            if not product_ids:
                print("⚠️ No hay productos en la base de datos.")
                return

            # Campos a actualizar
            valores = {
                "x_almacen1_custom": 0,
                "x_almacen2_custom": 0,
                "x_transit_stock_custom": 0,
                "x_transit": 0,
            }

            # Actualizar en lote
            try:
                models.execute_kw(
                    db, uid, password,
                    "product.product", "write",
                    [product_ids, valores]
                )
                print(f"✔ Campos custom reseteados correctamente en {len(product_ids)} productos.")

            except Exception as e:
                print(f"❌ Error actualizando productos: {e}")

        import re
        import unicodedata

        def reemplazar_nombre_producto():
            from App_Connection import models, db, uid, password

            print("\n=== INICIANDO REEMPLAZO INTELIGENTE DE NOMBRES ===")

            product_ids = models.execute_kw(
                db, uid, password,
                "product.template", "search",
                [[]]
            )

            productos = models.execute_kw(
                db, uid, password,
                "product.template", "read",
                [product_ids, ["name", "default_code"]]
            )

            def normalizar(texto):
                # Normaliza caracteres unicode
                t = unicodedata.normalize("NFKC", texto)

                # Sustituye comillas raras
                t = t.replace("’", "'").replace("`", "'")

                # Sustituye espacios raros (NBSP, EN SPACE, EM SPACE…)
                t = re.sub(r"\s+", " ", t)

                return t.strip()

            # Regex que detecta todas las variantes de "3 EN 1"
            patron = re.compile(r"3\s*[-]?\s*EN\s*[-]?\s*1", re.IGNORECASE)

            for p in productos:
                prod_id = p["id"]
                nombre_original = p["name"] or ""
                sku = p.get("default_code", "")

                # Normalizamos solo para detectar
                nombre_norm = normalizar(nombre_original).upper()

                # Si contiene la estructura "3 EN 1" en cualquier variante
                if patron.search(nombre_norm):
                    # Hacemos el reemplazo sobre el nombre
                    nuevo_nombre = patron.sub("CCT", normalizar(nombre_original))

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "product.template", "write",
                            [[prod_id], {"name": nuevo_nombre}]
                        )
                        print(f"✔ {sku} | '{nombre_original}' → '{nuevo_nombre}'")
                    except Exception as e:
                        print(f"❌ Error actualizando {sku}: {e}")

            print("\n✔ Reemplazo completado correctamente.")

        def actualizar_coste_desde_excel(ruta_excel):
            from App_Connection import models, db, uid, password

            print("\n=== ACTUALIZANDO standard_price DESDE EXCEL ===")

            # Leer hoja Sheet1
            df = pd.read_excel(ruta_excel, sheet_name="Sheet1")

            # Normalización de columnas
            if "SKU" not in df.columns or "COSTE" not in df.columns:
                raise ValueError("La hoja debe contener columnas SKU y COSTE.")

            df["SKU"] = df["SKU"].astype(str).str.strip()

            for _, row in df.iterrows():
                sku = row["SKU"]
                coste = row["COSTE"]

                # Buscar producto en Odoo por default_code
                product_ids = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[["default_code", "=", sku]]]
                )

                if not product_ids:
                    print(f"⚠ SKU {sku} no encontrado en Odoo.")
                    continue

                try:
                    models.execute_kw(
                        db, uid, password,
                        "product.product", "write",
                        [product_ids, {"standard_price": coste}]
                    )
                    print(f"✔ SKU {sku} → standard_price = {coste}")

                except Exception as e:
                    print(f"❌ Error actualizando SKU {sku}: {e}")

            print("\n✔ Proceso completado.")

        def aplicar_cambios_desde_excel(
                ruta_excel = ruta,
                hoja_cambios="CAMBIOS",
                columna_ref="default_code"):
            """
            Lee la hoja CAMBIOS, detecta celdas en amarillo (FFFF00),
            y actualiza SOLO esas columnas en Odoo, buscando por default_code.
            """
            from App_Connection import models, db, uid, password
            from openpyxl import load_workbook

            COLOR_AMARILLO = "FFFF00"

            # === 1. Cargar Excel ===
            wb = load_workbook(ruta_excel, data_only=True)
            ws = wb[hoja_cambios]

            # Obtener cabeceras
            headers = [c.value for c in ws[1]]
            col_index = {headers[i]: i + 1 for i in range(len(headers))}

            # === 2. Recopilar cambios detectando celdas amarillas ===
            cambios = []

            for row in ws.iter_rows(min_row=2):
                ref_cell = row[col_index[columna_ref] - 1]
                ref_value = str(ref_cell.value).strip() if ref_cell.value else ""

                if not ref_value:
                    continue  # ignoramos filas sin referencia

                for col_name, idx in col_index.items():
                    cell = row[idx - 1]
                    fill = cell.fill

                    # Detectar el amarillo
                    is_yellow = (
                            fill
                            and fill.fgColor
                            and fill.fgColor.rgb is not None
                            and fill.fgColor.rgb[-6:].upper() == COLOR_AMARILLO
                    )

                    if is_yellow:
                        cambios.append({
                            "default_code": ref_value,
                            "columna": col_name,
                            "valor": cell.value,
                        })

            if not cambios:
                print("No se encontraron cambios resaltados en amarillo.")
                return

            # === 3. Procesar cambios en Odoo ===
            for cambio in cambios:
                ref = cambio["default_code"]
                col = cambio["columna"]
                val = cambio["valor"]

                # Buscar el producto por default_code
                product_ids = models.execute_kw(
                    db, uid, password,
                    "product.template", "search",
                    [[["default_code", "=", ref]]]
                )

                if not product_ids:
                    print(f"⚠ No encontrado en Odoo → {ref}")
                    continue

                # Actualizar SOLO la columna modificada
                try:
                    models.execute_kw(
                        db, uid, password,
                        "product.template", "write",
                        [product_ids, {col: val}]
                    )
                    print(f"✔ Actualizado {ref}: {col} = {val}")

                except Exception as e:
                    print(f"❌ Error actualizando {ref} ({col}): {e}")

            print("✔ Todos los cambios han sido aplicados a Odoo correctamente.")


        def aplicar_cambios_atributos_old_new(
                ruta_excel,
                hoja="Sheet1"):
            """
            Lee pares Old/New desde el Excel y actualiza atributos en productos Odoo
            usando el método SEGURO (idéntico a cambiar_cri_en_producto):

                1) eliminar PTAV antigua
                2) eliminar PTAL del atributo
                3) crear valor NEW si no existe
                4) crear PTAL nueva con valor NEW

            Esto evita validaciones de variantes y funciona en todos los casos.
            """

            import pandas as pd
            from App_Connection import models, db, uid, password

            # === 1. Leer Excel ===
            df = pd.read_excel(ruta_excel, sheet_name=hoja)
            df.columns = df.columns.str.strip()

            # === 2. Detectar pares Old/New ===
            old_cols = [c for c in df.columns if c.startswith("Old: ")]
            new_cols = [c for c in df.columns if c.startswith("New: ")]

            atributos = []
            for old_col in old_cols:
                atributo = old_col.replace("Old: ", "").strip()
                new_col = f"New: {atributo}"
                if new_col in new_cols:
                    atributos.append((atributo, old_col, new_col))

            if not atributos:
                print("❌ No se encontraron columnas Old/New")
                return

            print(f"✔ Atributos detectados: {[a[0] for a in atributos]}")

            # === 3. Procesar cada atributo ===
            for atributo, old_col, new_col in atributos:

                print("\n" + "=" * 70)
                print(f"🔧 Procesando atributo: {atributo}")
                print("=" * 70)

                # 3.1 Buscar atributo en Odoo
                attr_ids = models.execute_kw(
                    db, uid, password,
                    "product.attribute", "search",
                    [[["name", "=", atributo]]]
                )

                if not attr_ids:
                    print(f"❌ Atributo '{atributo}' no existe en Odoo. Se salta.")
                    continue

                attr_id = attr_ids[0]

                # === 4. Filas del Excel ===
                for _, row in df.iterrows():

                    valor_old = str(row[old_col]).strip() if pd.notna(row[old_col]) else ""
                    valor_new = str(row[new_col]).strip() if pd.notna(row[new_col]) else ""

                    if not valor_old or not valor_new:
                        continue  # ignorar instrucciones incompletas

                    print(f"\n➡ Cambio: '{valor_old}' → '{valor_new}'")

                    # === 4.1 Buscar valor OLD (si no existe, ignorar) ===
                    old_val_ids = models.execute_kw(
                        db, uid, password,
                        "product.attribute.value", "search",
                        [[
                            ["name", "=", valor_old],
                            ["attribute_id", "=", attr_id]
                        ]]
                    )

                    if not old_val_ids:
                        print(f"   ⚠ OLD '{valor_old}' no existe en Odoo. Se continúa.")
                        continue

                    old_val_id = old_val_ids[0]

                    # === 4.2 Buscar productos que usan ese valor OLD ===
                    ptav_ids = models.execute_kw(
                        db, uid, password,
                        "product.template.attribute.value", "search",
                        [[["product_attribute_value_id", "=", old_val_id]]]
                    )

                    if not ptav_ids:
                        print(f"   ⚠ Ningún producto usa OLD '{valor_old}'")
                        continue

                    # Leer productos afectados
                    ptavs = models.execute_kw(
                        db, uid, password,
                        "product.template.attribute.value", "read",
                        [ptav_ids, ["product_tmpl_id"]]
                    )

                    productos = {p["product_tmpl_id"][0] for p in ptavs}
                    print(f"   ✔ Productos afectados: {len(productos)}")

                    # === 4.3 Buscar o crear valor NEW ===
                    new_val_ids = models.execute_kw(
                        db, uid, password,
                        "product.attribute.value", "search",
                        [[
                            ["name", "=", valor_new],
                            ["attribute_id", "=", attr_id]
                        ]]
                    )

                    if new_val_ids:
                        new_val_id = new_val_ids[0]
                        print(f"   ✔ Valor NEW existente")
                    else:
                        new_val_id = models.execute_kw(
                            db, uid, password,
                            "product.attribute.value", "create",
                            [{
                                "name": valor_new,
                                "attribute_id": attr_id
                            }]
                        )
                        print(f"   🆕 Valor NEW creado")

                    # === 4.4 ACTUALIZAR producto por producto ===
                    for prod_id in productos:

                        # ============================================
                        # 🔹 PRIMERA MODIFICACIÓN: PRINT DEL SKU
                        # ============================================
                        try:
                            # Intentar leer el default_code del producto
                            prod_data = models.execute_kw(
                                db, uid, password,
                                "product.template", "read",
                                [[prod_id], ["default_code"]]
                            )[0]
                            sku = prod_data.get("default_code", "SIN SKU")
                        except:
                            sku = "SIN SKU"

                        print(f"   → Procesando producto {prod_id} → SKU: {sku}")

                        # ============================================
                        # 🔹 SEGUNDA MODIFICACIÓN: NO INTERRUMPIR FLUJO
                        # ============================================
                        try:
                            # 1) Buscar PTAV específica del producto
                            ptav_prod_ids = models.execute_kw(
                                db, uid, password,
                                "product.template.attribute.value", "search",
                                [[
                                    ["product_tmpl_id", "=", prod_id],
                                    ["product_attribute_value_id", "=", old_val_id]
                                ]]
                            )

                            # 2) Buscar PTAL del atributo para el producto
                            ptal_prod_ids = models.execute_kw(
                                db, uid, password,
                                "product.template.attribute.line", "search",
                                [[
                                    ["product_tmpl_id", "=", prod_id],
                                    ["attribute_id", "=", attr_id]
                                ]]
                            )

                            # === EL MÉTODO QUE SÍ FUNCIONA ===
                            # 🔥  BORRAR PTAV
                            if ptav_prod_ids:
                                try:
                                    models.execute_kw(
                                        db, uid, password,
                                        "product.template.attribute.value", "unlink",
                                        [ptav_prod_ids]
                                    )
                                except Exception as e:
                                    print(f"      ⚠ Error borrando PTAV en SKU {sku}: {e}")

                            # 🔥  BORRAR PTAL
                            if ptal_prod_ids:
                                try:
                                    models.execute_kw(
                                        db, uid, password,
                                        "product.template.attribute.line", "unlink",
                                        [ptal_prod_ids]
                                    )
                                except Exception as e:
                                    print(f"      ⚠ Error borrando PTAL en SKU {sku}: {e}")

                            # 🔥  CREAR PTAL NUEVA CON NEW
                            try:
                                models.execute_kw(
                                    db, uid, password,
                                    "product.template.attribute.line", "create",
                                    [{
                                        "product_tmpl_id": prod_id,
                                        "attribute_id": attr_id,
                                        "value_ids": [(6, 0, [new_val_id])]
                                    }]
                                )
                                print(f"      ✔ {valor_old} → {valor_new} aplicado en SKU {sku}")
                            except Exception as e:
                                print(f"      ❌ Error creando nueva PTAL en SKU {sku}: {e}")

                        except Exception as e:
                            print(f"   ❌ Error procesando producto {prod_id} (SKU {sku}): {e}")
                            print("   Continuando con el siguiente producto…")
                            continue

                    # === 4.5 eliminar OLD si ya no se usa ===
                    still_used = models.execute_kw(
                        db, uid, password,
                        "product.template.attribute.value", "search",
                        [[["product_attribute_value_id", "=", old_val_id]]]
                    )

                    if not still_used:
                        models.execute_kw(
                            db, uid, password,
                            "product.attribute.value", "unlink",
                            [[old_val_id]]
                        )
                        print(f"   🗑 Valor OLD '{valor_old}' eliminado del sistema")

            print("\n🎉 FINALIZADO: Todos los cambios OLD/NEW aplicados correctamente.")

        def desarchivar_productos(batch_size=1):
            """
            Desarchiva TODOS los productos (product.template) en Odoo,
            procesándolos en lotes para evitar que el servidor se bloquee.
            """

            from App_Connection import models, db, uid, password

            print("🔍 Buscando productos archivados...")

            # === 1. Buscar productos archivados (active=False) ===
            product_ids = models.execute_kw(
                db, uid, password,
                "product.template", "search",
                [[["active", "=", False]]]
            )

            total = len(product_ids)

            if total == 0:
                print("✔ No hay productos archivados.")
                return

            print(f"📦 Productos archivados encontrados: {total}")
            print(f"🚀 Procesando en lotes de {batch_size}...\n")

            # === 2. Procesar en lotes ===
            for i in range(0, total, batch_size):
                batch = product_ids[i: i + batch_size]

                try:
                    models.execute_kw(
                        db, uid, password,
                        "product.template", "write",
                        [batch, {"active": True}]
                    )
                    print(f"   ✔ Lote {i + 1}-{i + len(batch)} desarchivado correctamente.")

                except Exception as e:
                    print(f"   ❌ Error desarchivando lote {i + 1}-{i + len(batch)}: {e}")
                    print("   Continuando con el siguiente lote...")

            print("\n🎉 FINALIZADO: Todos los productos han sido procesados.")

        def eliminar_asteriscos(ruta_excel):
            from App_Connection import models, db, uid, password
            import pandas as pd

            print("📌 Leyendo Excel...")
            df = pd.read_excel(ruta_excel, sheet_name="Sheet1")

            if "default_code" not in df.columns:
                raise Exception("❌ La columna 'SKU' no existe en Sheet1")

            # Normalizar SKU (evitar espacios, NaN, etc.)
            df["default_code"] = df["default_code"].astype(str).str.strip()
            skus = df["default_code"].unique()

            print(f"🔍 Total SKUs encontrados: {len(skus)}")

            productos_modificados = 0

            for sku in skus:
                if not sku or sku.lower() == "nan":
                    continue

                # Buscar producto en Odoo
                prod_ids = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[["default_code", "=", sku]]]
                )

                if not prod_ids:
                    print(f"⚠️ Producto no encontrado en Odoo: {sku}")
                    continue

                prod_id = prod_ids[0]

                # Leer el nombre actual
                prod_data = models.execute_kw(
                    db, uid, password,
                    "product.product", "read",
                    [prod_ids, ["name"]]
                )[0]

                nombre_original = prod_data["name"]
                nombre_nuevo = nombre_original.replace("*", "")

                if nombre_original != nombre_nuevo:
                    # Actualizar en Odoo
                    models.execute_kw(
                        db, uid, password,
                        "product.product", "write",
                        [[prod_id], {"name": nombre_nuevo}]
                    )
                    productos_modificados += 1
                    print(f"✔ Modificado SKU {sku}: '{nombre_original}' → '{nombre_nuevo}'")

            print("--------------------------------------------------")
            print(f"🎉 Proceso completado. Productos modificados: {productos_modificados}")

        def actualizar_stock_custom(ruta_excel, opt_stock = False):
            from App_Connection import models, db, uid, password

            # Leer las hojas necesarias
            df_madrid = pd.read_excel(ruta_excel, sheet_name="MADRID2")
            df_bulgaria = pd.read_excel(ruta_excel, sheet_name="BULGARIA3")

            # Normalizar SKU
            df_madrid["SKU"] = df_madrid["SKU"].astype(str).str.strip()
            df_bulgaria["SKU"] = df_bulgaria["SKU"].astype(str).str.strip()

            # ---- MADRID2 → x_transit_stock_custom ------------------------------------
            if "STOCK" not in df_madrid.columns:
                raise ValueError("MADRID2 debe contener columna STOCK")

            print("\n=== Actualizando x_transit_stock_custom desde MADRID2 ===")

            for _, row in df_madrid.iterrows():
                sku = row["SKU"]
                stock = row["STOCK"]

                # Buscar producto por default_code
                product_ids = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[["default_code", "=", sku]]]
                )

                if not product_ids:
                    #print(f"⚠ Producto con SKU {sku} no encontrado en Odoo.")
                    continue

                try:
                    models.execute_kw(
                        db, uid, password,
                        "product.product", "write",
                        [product_ids, {"x_transit_stock_custom": stock}]
                    )
                    print(f"✔ SKU {sku} → x_transit_stock_custom = {stock}")
                except Exception as e:
                    print(f"❌ Error actualizando SKU {sku}: {e}")

            # ---- BULGARIA3 → x_almacen1_custom --------------------------------------
            if "STOCK" not in df_bulgaria.columns:
                raise ValueError("BULGARIA3 debe contener columna STOCK")

            print("\n=== Actualizando x_almacen1_custom desde BULGARIA3 ===")

            for _, row in df_bulgaria.iterrows():
                sku = row["SKU"]
                stock = row["STOCK"]

                product_ids = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[["default_code", "=", sku]]]
                )

                if not product_ids:
                    #print(f"⚠ Producto con SKU {sku} no encontrado en Odoo.")
                    continue

                try:
                    models.execute_kw(
                        db, uid, password,
                        "product.product", "write",
                        [product_ids, {"x_almacen1_custom": stock}]
                    )
                    print(f"✔ SKU {sku} → x_almacen1_custom = {stock}")
                except Exception as e:
                    print(f"❌ Error actualizando SKU {sku}: {e}")

            if opt_stock:
                df_opt = pd.read_excel(ruta_excel, sheet_name="ODOO1")
                df_opt["SKU"] = df_opt["SKU"].astype(str).str.strip()

                # ---- ODOO1 → x_almacen3_custom --------------------------------------
                if "STOCK" not in df_opt.columns:
                    raise ValueError("ODOO1 debe contener columna STOCK")

                print("\n=== Actualizando x_almacen3_custom desde BULGARIA3 ===")

                for _, row in df_opt.iterrows():
                    sku = row["SKU"]
                    stock = row["STOCK"]

                    product_ids = models.execute_kw(
                        db, uid, password,
                        "product.product", "search",
                        [[["default_code", "=", sku]]]
                    )

                    if not product_ids:
                        #print(f"⚠ Producto con SKU {sku} no encontrado en Odoo.")
                        continue

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "product.product", "write",
                            [product_ids, {"x_almacen3_custom": stock}]
                        )
                        print(f"✔ SKU {sku} → x_almacen3_custom = {stock}")
                    except Exception as e:
                        print(f"❌ Error actualizando SKU {sku}: {e}")

            print("\n✔ Proceso completado.")

        def migrar_tarea(destino, origen_task_id, destino_task_id, lote=30, delay=0.1):
            from App_Connection import models, db, uid, password
            import time
            """
            Migra una tarea de project.task desde Odoo origen → destino.

            origen: dict con url, db, user, password
            destino: dict con url, db, user, password
            origen_task_id : ID de tarea en origen
            destino_task_id : ID de tarea en destino
            lote : cantidad de adjuntos por lote (default 30)
            delay : segundos entre lotes para evitar saturación (default 0.2)
            """

            print(f"\n=== MIGRANDO TAREA {origen_task_id} → {destino_task_id} ===")

            # --- LOGIN DESTINO ---
            dest_common = xmlrpc.client.ServerProxy(f'{destino['url']}/xmlrpc/2/common', allow_none=True)
            dest_uid = dest_common.authenticate(destino['db'], destino['user'], destino['password'], {})

            if not dest_uid:
                print("❌ No se pudo autenticar.")
                return

            print(f'🔌 Conectado como {destino['user']} (uid: {dest_uid})')
            dest_models = xmlrpc.client.ServerProxy(f'{destino['url']}/xmlrpc/2/object', allow_none=True)

            # -----------------------------
            # 3) MIGRAR ADJUNTOS
            # -----------------------------
            print("\nLeyendo adjuntos del origen...")

            attach_ids = models.execute_kw(
                db, uid, password,
                'ir.attachment', 'search',
                [[['res_model', '=', 'project.task'], ['res_id', '=', origen_task_id]]]
            )

            total_adjuntos = len(attach_ids)
            print(f"Total adjuntos encontrados: {total_adjuntos}")

            if total_adjuntos == 0:
                print("No hay adjuntos que migrar.")
                return

            # Calculamos lotes
            num_lotes = math.ceil(total_adjuntos / lote)

            for i in range(num_lotes):
                inicio = i * lote
                fin = inicio + lote
                lote_ids = attach_ids[inicio:fin]

                print(f"\nProcesando lote {i + 1}/{num_lotes} ({len(lote_ids)} adjuntos)...")

                adjuntos = models.execute_kw(
                    db, uid, password,
                    'ir.attachment', 'read',
                    [lote_ids, ['name', 'datas', 'mimetype']]
                )

                # Crear adjuntos en destino
                for att in adjuntos:
                    try:
                        dest_models.execute_kw(
                            destino['db'], dest_uid, destino['password'],
                            'ir.attachment', 'create',
                            [{
                                'name': att['name'],
                                'datas': att['datas'],
                                'mimetype': att['mimetype'],
                                'res_model': 'project.task',
                                'res_id': destino_task_id,
                            }]
                        )
                    except Exception as e:
                        print(f"⚠️ Error importando {att['name']}: {e}")

                print(f"✓ Lote {i + 1}/{num_lotes} completado.")

                time.sleep(delay)

            print("\n=== MIGRACIÓN COMPLETADA CORRECTAMENTE ===")

        destino = {
            'url': "http://79.72.61.76:8070/",
            'db': "odoo1",
            'user': "admin",
            'password': "admin"
        }

        '''migrar_tarea(
            destino,
            origen_task_id=1426,
            destino_task_id=1523,
            lote=1,  # tamaño del lote
            delay=1  # descanso entre lotes
        )'''

        def revertir_facturas_cliente_a_borrador():
            from App_Connection import db, uid, password, models
            f=0
            ctx = {
                "active_test": False,
                "lang": "es_ES",
            }

            # =========================
            # 1️⃣ FACTURAS DE CLIENTE
            # =========================
            invoice_domain = [
                ("move_type", "in", ["out_invoice", "out_refund"]),#["in_invoice", "in_refund"]),#
                ("state", "=", "posted"),
            ]

            invoice_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [invoice_domain],
                {"context": ctx}
            )

            print(f"➡ Facturas cliente encontradas: {len(invoice_ids)}")

            for inv_id in invoice_ids:
                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.move", "button_draft",
                        [[inv_id]],
                        {"context": ctx}
                    )
                except Exception as e:
                    f+=1
                    print(f"Factura {inv_id}: {f}")
            print("✅ Proceso finalizado")

        def revertir_pagos_clientes_a_borrador():
            from App_Connection import db, uid, password, models

            ctx = {
                "active_test": False,
                "lang": "es_ES",
            }
            p=0

            # 1️⃣ Buscar pagos de clientes (inbound) que NO estén en borrador
            payment_ids = models.execute_kw(
                db, uid, password,
                "account.payment", "search",
                [[
                    ("payment_type", "=", "inbound"),#("payment_type", "=", "outbound"),
                    ("state", "!=", "draft")
                ]],
                {"context": ctx}
            )

            print(f"➡ Pagos de clientes a revertir: {len(payment_ids)}")

            if not payment_ids:
                print("ℹ️ No hay pagos que revertir.")
                return

            # 2️⃣ Revertir a borrador
            revertidos = 0
            for payment_id in payment_ids:
                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.payment", "action_draft",
                        [[payment_id]],
                        {"context": ctx}
                    )
                    revertidos += 1
                except Exception as e:
                    p+=1
                    print(f"Pago {payment_id}: {p}")

            print(f"✅ Pagos revertidos correctamente: {revertidos}/{len(payment_ids)}")

        def revertir_todos_los_asientos_a_borrador():
            from App_Connection import db, uid, password, models

            print("🔄 Buscando asientos contables en estado POSTED...")

            # 1. Buscar todos los asientos publicados
            move_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[("state", "=", "posted")]]
            )

            print(f"📊 Asientos encontrados: {len(move_ids)}")

            if not move_ids:
                print("✅ No hay asientos para revertir")
                return

            # 2. Revertir a borrador
            for move_id in move_ids:
                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.move", "button_draft",
                        [[move_id]]
                    )
                except Exception as e:
                    print(f"❌ Error en asiento ID {move_id}: {e}")

            print("✅ Todos los asientos han sido revertidos a BORRADOR")

        def borrar_todos_los_asientos_en_borrador():
            from App_Connection import db, uid, password, models

            print("🗑️ Buscando asientos en estado BORRADOR...")

            move_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[("state", "=", "draft")]]
            )

            print(f"📊 Asientos en borrador encontrados: {len(move_ids)}")

            if not move_ids:
                print("✅ No hay asientos para borrar")
                return

            for move_id in move_ids:
                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.move", "unlink",
                        [[move_id]]
                    )
                except Exception as e:
                    print(f"❌ Error borrando asiento ID {move_id}: {e}")

            print("✅ Todos los asientos en borrador han sido eliminados")



        def export_conciliacion_factura(invoice_name, company_name="ALMAITANA DE LUZ, S.L."):
            import xmlrpc.client

            url = 'https://optimaluz.soluntec.net'
            db_old = 'Test'
            username = 'jcoronado@optimaluz.com'
            password_old = 'AlAi4ever'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid = common.authenticate(db_old, username, password_old, {})

            if not uid:
                print("❌ No se pudo autenticar")
                return []

            models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            # 1️⃣ Buscar compañía e Invoice
            invoice_ids = models.execute_kw(db_old, uid, password_old, 'account.move', 'search', [[
                ('name', '=', invoice_name),
                ('move_type', '=', 'out_invoice'),
                ('company_id.name', '=', company_name)
            ]])

            if not invoice_ids:
                print(f"❌ Factura {invoice_name} no encontrada")
                return []

            # 2️⃣ Buscar la línea 'Receivable' con el campo correcto de Odoo 16
            # En Odoo 16 el account_type para clientes es 'asset_receivable'
            lines = models.execute_kw(db_old, uid, password_old, 'account.move.line', 'search_read', [[
                ('move_id', '=', invoice_ids[0]),
                ('account_id.account_type', '=', 'asset_receivable')
            ]], {'fields': ['id', 'matched_debit_ids', 'matched_credit_ids']})

            if not lines:
                print("❌ No se encontró línea de cuenta a cobrar")
                return []

            line = lines[0]

            # 3️⃣ Obtener los IDs de conciliación parcial
            # Estas tablas vinculan la línea de la factura con la línea del pago
            partial_ids = line['matched_debit_ids'] + line['matched_credit_ids']

            if not partial_ids:
                print("⚠️ Factura no conciliada")
                return []

            # 4️⃣ Buscar las líneas de pago (contrapartida) a través de account.partial.reconcile
            partials = models.execute_kw(db_old, uid, password_old, 'account.partial.reconcile', 'read', [partial_ids],
                                         {'fields': ['debit_move_id', 'credit_move_id', 'amount']})

            resultado = []
            for p in partials:
                # Identificar cuál de las dos líneas NO es la de nuestra factura (esa es el pago)
                pago_line_info = p['credit_move_id'] if p['debit_move_id'][0] == line['id'] else p['debit_move_id']

                resultado.append({
                    "invoice_name": invoice_name,
                    "pago_asiento": pago_line_info[1],  # Nombre del asiento de pago/banco
                    "pago_line_id": pago_line_info[0],
                    "importe_conciliado": p['amount']
                })
                print(f"   ↔ Conciliado con {pago_line_info[1]} | Importe: {p['amount']}")

            return resultado

        #export_conciliacion_factura("FV1OP/25/00857")

        def export_plan_contable():
            import xmlrpc.client

            company_id = 2

            accounts = models_src.execute_kw(
                db_src, uid_src, password_src,
                'account.account', 'search_read',
                [[('company_id', '=', company_id)]],
                {
                    'fields': [
                        'code',
                        'name',
                        'account_type',
                        'reconcile',
                        'deprecated'
                    ],
                    'order': 'code'
                }
            )

            print(f"📤 Exportadas {len(accounts)} cuentas contables")
            return accounts

        def import_plan_contable(accounts):
            creadas = 0
            existentes = 0

            # Mapeo de tipos (normalizado)
            TYPE_MAP = {
                'asset_receivable': 'asset_receivable',
                'liability_payable': 'liability_payable',
                'asset_cash': 'asset_cash',
                'income': 'income',
                'expense': 'expense',
                'equity': 'equity',
                'asset_current': 'asset_receivable',
                'liability_current': 'liability_payable',
                'asset_current': 'asset_current',
                'asset_fixed': 'asset_current',
                'liability_current': 'liability_payable',
                'liability_non_current': 'liability_payable',
            }

            for acc in accounts:
                code = acc['code']
                name = acc['name']
                acc_type_src = acc['account_type']
                reconcile = acc['reconcile']
                deprecated = acc['deprecated']

                print(f"\n📘 Cuenta {code} – {name}")

                # ¿Existe ya?
                existing = models.execute_kw(
                    db, uid, password,
                    'account.account', 'search',
                    [[('code', '=', code)]]
                )

                if existing:
                    print("⚠️ Ya existe")
                    existentes += 1
                    continue

                acc_type = TYPE_MAP.get(acc_type_src)
                if not acc_type:
                    print(f"❌ Tipo no mapeado: {acc_type_src}")
                    continue

                # 🔥 Regla obligatoria en Odoo 18
                if acc_type in ('asset_receivable', 'liability_payable'):
                    reconcile = True

                vals = {
                    'code': code,
                    'name': name,
                    'account_type': acc_type,
                    'reconcile': reconcile,
                    'deprecated': deprecated,
                }

                try:
                    models.execute_kw(
                        db, uid, password,
                        'account.account', 'create',
                        [vals]
                    )
                    print("✅ Cuenta creada")
                    creadas += 1
                except Exception as e:
                    print(f"❌ Error creando cuenta {code}: {e}")

            print("\n📊 PLAN CONTABLE IMPORTADO")
            print(f"   Creadas: {creadas}")
            print(f"   Existentes: {existentes}")

        #accounts = export_plan_contable()
        #import_plan_contable(accounts)

        def limpiar_asientos_automaticos():
            from App_Connection import db, uid, password, models

            print("🧹 Limpiando asientos contables automáticos...")

            # Buscar asientos que NO sean facturas ni pagos
            move_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[
                    ("state", "=", "posted"),
                    ("move_type", "=", "entry"),
                ]]
            )

            print(f"🔎 Asientos encontrados: {len(move_ids)}")

            if not move_ids:
                print("✔️ No hay asientos a borrar")
                return

            try:
                # 1️⃣ Pasar a borrador
                models.execute_kw(
                    db, uid, password,
                    "account.move", "button_draft",
                    [move_ids]
                )
            except Exception as e:
                print("↩️ Asientos pasados a borrador")

            try:
                # 2️⃣ Borrar
                models.execute_kw(
                    db, uid, password,
                    "account.move", "unlink",
                    [move_ids]
                )
            except Exception as e:
                print("🗑️ Asientos eliminados correctamente")

        #limpiar_asientos_automaticos()

        def export_account_moves_entries(company_name="ALMAITANA DE LUZ, S.L.", limit=None):
            import xmlrpc.client

            url = 'https://optimaluz.soluntec.net'
            db = 'Test'
            username = 'jcoronado@optimaluz.com'
            password = 'AlAi4ever'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid = common.authenticate(db, username, password, {})
            if not uid:
                print("❌ Error autenticación")
                return []

            models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            # Buscar compañía
            company_id = models.execute_kw(
                db, uid, password,
                'res.company', 'search',
                [[('name', '=', company_name)]],
                {'limit': 1}
            )[0]

            domain = [
                ('company_id', '=', company_id),
                ('move_type', '=', 'entry'),
                ('state', '=', 'posted'),
            ]

            moves = models.execute_kw(
                db, uid, password,
                'account.move', 'search_read',
                [domain],
                {
                    'fields': [
                        'id', 'name', 'date', 'ref',
                        'journal_id', 'line_ids'
                    ],
                    'limit': limit
                }
            )

            print(f"📤 Exportados {len(moves)} asientos contables")

            # Leer líneas en bloque
            line_ids = [lid for m in moves for lid in m['line_ids']]
            lines = models.execute_kw(
                db, uid, password,
                'account.move.line', 'read',
                [line_ids],
                {
                    'fields': [
                        'move_id', 'name',
                        'account_id', 'partner_id',
                        'debit', 'credit'
                    ]
                }
            )

            lines_by_move = {}
            for l in lines:
                lines_by_move.setdefault(l['move_id'][0], []).append(l)

            for m in moves:
                m['lines'] = lines_by_move.get(m['id'], [])

            return moves

        def import_account_moves_entries(moves):
            from App_Connection import db, uid, password, models

            creados = 0
            saltados = 0

            for m in moves:
                name = m['name']
                print(f"\n📘 Asiento {name}")

                # Evitar duplicados
                if models.execute_kw(db, uid, password, 'account.move', 'search', [[('name', '=', name)]]):
                    print("⚠️ Ya existe, saltando")
                    saltados += 1
                    continue

                # Buscar diario
                journal_name = m['journal_id'][1]
                journal = models.execute_kw(
                    db, uid, password,
                    'account.journal', 'search_read',
                    [[('name', '=', journal_name)]],
                    {'fields': ['id'], 'limit': 1}
                )
                if not journal:
                    print(f"❌ Diario no encontrado: {journal_name}")
                    continue

                move_vals = {
                    'name': name,
                    'move_type': 'entry',
                    'journal_id': journal[0]['id'],
                    'date': m['date'],
                    'ref': m['ref'],
                    'line_ids': [],
                }

                for l in m['lines']:
                    # Buscar cuenta contable
                    account_full_name = l['account_id'][1]
                    account_code = l['account_id'][1].split(' ')[0]
                    account = models.execute_kw(
                        db, uid, password,
                        'account.account', 'search_read',
                        [[('code', '=', account_code)]],
                        {'fields': ['id'], 'limit': 1}
                    )
                    if not account:
                        print(f"❌ Cuenta no encontrada: {account_full_name}")
                        continue

                    partner_id = None
                    if l['partner_id']:
                        partner = models.execute_kw(
                            db, uid, password,
                            'res.partner', 'search_read',
                            [[('name', '=', l['partner_id'][1])]],
                            {'fields': ['id'], 'limit': 1}
                        )
                        partner_id = partner[0]['id'] if partner else None

                    move_vals['line_ids'].append((0, 0, {
                        'name': l['name'],
                        'account_id': account[0]['id'],
                        'partner_id': partner_id,
                        'debit': l['debit'],
                        'credit': l['credit'],
                    }))

                try:
                    new_id = models.execute_kw(db, uid, password, 'account.move', 'create', [move_vals])
                    models.execute_kw(db, uid, password, 'account.move', 'action_post', [[new_id]])
                    print(f"✅ Asiento creado y publicado ({new_id})")
                    creados += 1
                except Exception as e:
                    print(f"❌ Error creando asiento {name}: {e}")

            print("\n📊 RESULTADO")
            print(f"   Creados: {creados}")
            print(f"   Saltados: {saltados}")

        #moves = export_account_moves_entries()
        #import_account_moves_entries(moves)

        def export_partner_bank_accounts(company_name="ALMAITANA DE LUZ, S.L."):
            import xmlrpc.client

            url = 'https://optimaluz.soluntec.net'
            db = 'Test'
            username = 'jcoronado@optimaluz.com'
            password = 'AlAi4ever'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid = common.authenticate(db, username, password, {})
            if not uid:
                print("❌ Error autenticación")
                return []

            models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            accounts = models.execute_kw(
                db, uid, password,
                'res.partner.bank', 'search_read',
                [[]],
                {
                    'fields': [
                        'acc_number',
                        'partner_id',
                        'allow_out_payment'
                    ]
                }
            )

            resultado = []
            for acc in accounts:
                if not acc.get('partner_id'):
                    continue

                resultado.append({
                    'acc_number': acc['acc_number'],
                    'partner_name': acc['partner_id'][1],
                    'allow_out_payment': acc.get('allow_out_payment', False)
                })

            print(f"📤 Exportadas {len(resultado)} cuentas bancarias de contactos")
            return resultado

        def import_partner_bank_accounts(accounts):
            from App_Connection import db, uid, password, models

            creadas = 0
            existentes = 0
            errores = 0

            for acc in accounts:
                acc_number = acc['acc_number']
                partner_name = acc['partner_name']
                allow_out = acc.get('allow_out_payment', False)

                print(f"\n🏦 Cuenta {acc_number} → {partner_name}")

                # 1️⃣ Buscar partner por name
                partner = models.execute_kw(
                    db, uid, password,
                    'res.partner', 'search_read',
                    [[('name', '=', partner_name)]],
                    {'fields': ['id'], 'limit': 1}
                )

                if not partner:
                    print(f"❌ Partner no encontrado: {partner_name}")
                    errores += 1
                    continue

                partner_id = partner[0]['id']

                # 2️⃣ Evitar duplicados por número de cuenta
                existing = models.execute_kw(
                    db, uid, password,
                    'res.partner.bank', 'search',
                    [[('acc_number', '=', acc_number)]]
                )

                if existing:
                    print("⚠️ Cuenta ya existente, saltando")
                    existentes += 1
                    continue

                vals = {
                    'acc_number': acc_number,
                    'partner_id': partner_id,
                    'allow_out_payment': allow_out,
                }

                try:
                    models.execute_kw(
                        db, uid, password,
                        'res.partner.bank', 'create',
                        [vals]
                    )
                    print("✅ Cuenta bancaria creada")
                    creadas += 1
                except Exception as e:
                    print(f"❌ Error creando cuenta {acc_number}: {e}")
                    errores += 1

            print("\n📊 RESULTADO CUENTAS BANCARIAS")
            print(f"   Creadas: {creadas}")
            print(f"   Existentes: {existentes}")
            print(f"   Errores: {errores}")

        #accounts = export_partner_bank_accounts()
        #import_partner_bank_accounts(accounts)


        def export_apuntes_analiticos(analytic_name=None, company_name="ALMAITANA DE LUZ, S.L."):

            company_id = models_src.execute_kw(
                db_src, uid_src, password_src,
                'res.company', 'search',
                [[('name', '=', company_name)]],
                {'limit': 1}
            )[0]

            # ------------------------------
            # Dominio dinámico
            # ------------------------------
            domain = [('company_id', '=', company_id), ("employee_id", "=", False)]

            if analytic_name:
                domain.append(('name', '=', analytic_name))

            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                'account.analytic.line', 'search_read',
                [domain],
                {
                    'fields': [
                        'name',
                        'account_id',
                        'date',
                        'amount',
                        'ref',
                        'partner_id',
                        'unit_amount',
                        'product_id',
                        'move_line_id'
                    ],
                    'order': 'date',
                }
            )

            # --------------------------------------------------
            # 🥈 FASE 2: Enriquecer con datos de account.move.line
            # --------------------------------------------------

            # 1️⃣ Extraer IDs únicos de move_line
            move_line_ids = list({
                l['move_line_id'][0]
                for l in lines
                if l.get('move_line_id')
            })

            if move_line_ids:
                # 2️⃣ Leer move lines reales
                move_lines = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    'account.move.line', 'search_read',
                    [[('id', 'in', move_line_ids)]],
                    {
                        'fields': [
                            'id',
                            'move_name',
                            'name',
                            'account_id',
                            'balance',
                            'debit',
                            'credit',
                        ]
                    }
                )

                # 3️⃣ Indexar por ID para acceso rápido
                move_line_map = {ml['id']: ml for ml in move_lines}

                # 4️⃣ Inyectar datos en cada apunte analítico
                for l in lines:
                    if l.get('move_line_id'):
                        ml_id = l['move_line_id'][0]
                        ml = move_line_map.get(ml_id)
                        if ml:
                            l['move_line_data'] = ml

            print(f"📤 Exportados {len(lines)} apuntes analíticos")
            return lines

        def import_apuntes_analiticos(lines):
            #from App_Connection import db, uid, password, models

            def extraer_codigo_analitico(texto):
                """
                Extrae un código tipo PV-OPT/24/0374 desde un string largo.
                Busca la parte que contiene '/' y corta en el primer espacio.
                """
                if not texto:
                    return None

                for parte in texto.split():
                    if '/' in parte:
                        return parte.strip()

                return None

            creados = 0
            omitidos = 0

            PRODUCT_UOM_ID = 28  # fijo

            # 🔹 Obtener la ÚNICA compañía
            company = models.execute_kw(
                db, uid, password,
                'res.company', 'search_read',
                [[]],
                {'fields': ['id', 'name'], 'limit': 1}
            )
            if not company:
                print("❌ No se encontró ninguna compañía")
                return
            company_id = company[0]['id']

            print(f"🏢 Usando compañía: {company[0]['name']} (ID {company_id})")

            for l in lines:
                name = l.get('name')
                print(f"\n📈 Apunte analítico: {name}")

                # --------------------------------------------------
                # 1️⃣ Cuenta analítica (account_id) — OBLIGATORIA
                # --------------------------------------------------
                if not l.get('account_id'):
                    print("❌ Sin cuenta analítica, se omite")
                    omitidos += 1
                    continue

                acc_name_raw = l['account_id'][1]
                acc_name = extraer_codigo_analitico(acc_name_raw)

                if not acc_name:
                    acc_name = "Interno"#print(f"❌ No se pudo extraer código analítico de: {acc_name_raw}")
                    #omitidos += 1
                    #continue

                acc = models.execute_kw(
                    db, uid, password,
                    'account.analytic.account', 'search_read',
                    [[('name', 'ilike', acc_name)]],
                    {'fields': ['id', 'company_id'], 'limit': 1}
                )
                if not acc:
                    print(f"❌ Cuenta analítica no encontrada: {acc_name}")
                    omitidos += 1
                    continue

                account_id = acc[0]['id']

                # 🔥 Forzar company_id en la cuenta analítica
                models.execute_kw(
                    db, uid, password,
                    'account.analytic.account', 'write',
                    [[account_id], {'company_id': company_id}]
                )

                # --------------------------------------------------
                # 2️⃣ Partner (opcional)
                # --------------------------------------------------
                partner_id = None
                if l.get('partner_id'):
                    partner_id_origen = l['partner_id'][0]
                    partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)

                # --------------------------------------------------
                # 3️⃣ Producto (opcional)
                # --------------------------------------------------
                def extraer_sku(display_name):
                    import re
                    if not display_name:
                        return None
                    m = re.search(r'\[(.*?)\]', display_name)
                    return m.group(1) if m else None

                product_id = None

                if l.get('product_id'):
                    display = l['product_id'][1]
                    sku = extraer_sku(display)

                    if sku:
                        productos = models.execute_kw(
                            db, uid, password,
                            'product.product', 'search_read',
                            [[('default_code', '=', sku)]],
                            {'fields': ['id'], 'limit': 1}
                        )
                        if productos:
                            product_id = productos[0]['id']
                        else:
                            productos = models.execute_kw(
                                db, uid, password,
                                "product.product", "search",
                                [[("default_code", "ilike", sku)]],
                                {"limit": 1, "context": {"active_test": False}}
                            )
                            if productos:
                                product_id = productos[0]
                            else:
                                print(f"⚠️ Producto no encontrado (SKU={sku})")

                # --------------------------------------------------
                # 4️⃣ Apunte contable (move_line_id) — opcional pero crítico
                # --------------------------------------------------
                move_line_id = None
                if l.get('move_line_id'):
                    move_line_name = l['move_line_id'][1]
                    '''
                    ml = models.execute_kw(
                        db, uid, password,
                        'account.move.line', 'search_read',
                        [[('name', '=', move_line_name)]],
                        {'fields': ['id', 'company_id'], 'limit': 1}
                    )'''
                    ml = l.get('move_line_data')
                    if not ml:
                        print("⚠️ Sin datos de move line enriquecidos")
                        continue

                    move_name = ml['move_name']
                    line_name = ml['name']
                    balance = ml['balance']

                    domain = [
                        ('move_name', '=', move_name),
                        ('name', '=', line_name),
                        ('balance', '=', balance),
                        ('company_id', '=', company_id),
                    ]

                    ml = models.execute_kw(
                        db, uid, password,
                        'account.move.line', 'search_read',
                        [domain],
                        {'fields': ['id'], 'limit': 1}
                    )
                    if ml:
                        move_line_id = ml[0]['id']

                        # 🔥 Forzar company_id en el apunte contable
                        models.execute_kw(
                            db, uid, password,
                            'account.move.line', 'write',
                            [[move_line_id], {'company_id': company_id}]
                        )
                    else:
                        print(f"⚠️ Apunte contable no encontrado: {move_name} / {line_name}")

                # --------------------------------------------------
                # 5️⃣ Crear apunte analítico (YA sin conflicto)
                # --------------------------------------------------
                vals = {
                    'name': name,
                    'account_id': account_id,
                    'company_id': company_id,
                    'date': l.get('date'),
                    'amount': l.get('amount'),
                    'ref': l.get('ref'),
                    'partner_id': partner_id,
                    'unit_amount': l.get('unit_amount'),
                    'product_id': product_id,
                    #'product_uom_id': PRODUCT_UOM_ID,
                    'move_line_id': move_line_id,
                }

                try:
                    models.execute_kw(
                        db, uid, password,
                        'account.analytic.line', 'create',
                        [vals]
                    )
                    print("✅ Apunte creado")
                    creados += 1
                except Exception as e:
                    print(f"❌ Error creando apunte '{name}': {e}")
                    omitidos += 1

            print("\n📊 RESULTADO APUNTES ANALÍTICOS")
            print(f"   Creados: {creados}")
            print(f"   Omitidos: {omitidos}")

        #lines = export_apuntes_analiticos()
        #import_apuntes_analiticos(lines)

        def export_sale_orders_by_state(state):
            """
            Exporta los pedidos de venta de Odoo 16 filtrados por estado.
            Optimizado para rendimiento: lee todas las líneas en bloque.
            """
            company_id_src = 1  # OSE: 1 / ALM: 2
            # ------------------------------------------------
            # 1️⃣ Buscar pedidos de venta
            # ------------------------------------------------
            so_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "sale.order", "search",
                [[("id", "=", 76117), ('company_id', '=', company_id_src)]],
                {"order": "date_order asc"}
            )

            if not so_ids:
                print("ℹ️ No se encontraron pedidos de compra")
                return []

            FIELDS = [
                "id",
                "name",
                "partner_id",
                "date_order",
                "currency_id",
                "state",
                "order_line",
                "amount_untaxed",
                "amount_tax",
                "amount_total",
                "user_id",
                "note",
                "origin",
                "company_id",
                "x_comentarios",
                "client_order_ref",
                "fiscal_position_id",
                "user_id",
                "create_date"
            ]

            print(f"📤 Exportando pedidos de venta en estado '{state}'...")

            sale_orders = models_src.execute_kw(
                db_src, uid_src, password_src,
                "sale.order", "read",
                [so_ids],
                {"fields": FIELDS}
            )

            '''# 1️⃣ Exportar pedidos
            sale_orders = models_src.execute_kw(
                db_src, uid_src, password_src,
                "sale.order", "search_read",
                [[("state", "=", state), ('company_id', '=', company_id_src)]],
                {"fields": FIELDS}
            )'''

            print(f"   → {len(sale_orders)} pedidos encontrados.")

            # 2️⃣ Reunir todos los IDs de líneas
            all_line_ids = []
            for so in sale_orders:
                all_line_ids.extend(so.get("order_line", []))

            if not all_line_ids:
                print("⚠️  No se encontraron líneas de pedido.")
                return sale_orders

            print(f"   → {len(all_line_ids)} líneas totales detectadas. Leyendo en bloque...")

            # 3️⃣ Leer todas las líneas en bloque
            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "sale.order.line", "read",
                [all_line_ids],
                {"fields": [
                    "order_id",
                    "name",
                    "product_id",
                    "product_uom_qty",
                    "price_unit",
                    "price_subtotal",
                    "price_total",
                    "tax_id",
                    "discount",
                    "x_nota_interna",
                    "display_type"
                ]}
            )

            # Leer SKUs de productos
            product_ids = {l["product_id"][0] for l in lines if l.get("product_id")}
            products = models_src.execute_kw(
                db_src, uid_src, password_src,
                "product.product", "read",
                [list(product_ids)],
                {"fields": ["id", "default_code"]}
            )

            sku_map = {p["id"]: p["default_code"] for p in products}

            for line in lines:
                pid = line.get("product_id")
                line["default_code"] = sku_map.get(pid[0]) if pid else None

            # 6️⃣ Agrupar las líneas por pedido
            grouped_lines = defaultdict(list)
            for line in lines:
                order = line.get("order_id")
                if order:
                    grouped_lines[order[0]].append(line)

            # 7️⃣ Asignar líneas a cada pedido
            for so in sale_orders:
                so_id = so["id"]
                so["lineas_detalle"] = grouped_lines.get(so_id, [])

            print("✅ Líneas asignadas correctamente a cada pedido.")
            '''# 8️⃣ Filtrar pedidos por name
            names_filtrar = {
                "PV-OPT/23/1950",#Draft
                "PV-OPT/23/2145",
                "PV-OPT/23/2148",
                "PV-OPT/23/2249",
                "PV-OPT/22/2407",#PV
                "OPT/25/2007",#Cancel
            }

            sale_orders_filtrados = [
                so for so in sale_orders
                if so.get("name") in names_filtrar
            ]

            print(f"🎯 Pedidos filtrados por name: {len(sale_orders_filtrados)}")'''

            return sale_orders

        def import_sale_orders_with_lines(sale_orders, state):
            """
            Importa pedidos de venta en Odoo 18 junto con sus líneas.
            Incluye los impuestos (buscados por nombre).
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models

            total_creados = 0
            total_existentes = 0

            is_from_16 = True

            for so in sale_orders:
                name = so["name"]

                # Comprobar si ya existe
                '''existing = models.execute_kw(
                    db, uid, password,
                    "sale.order", "search",
                    [[("x_id_interno", "=", so["id"])]]
                )
                if existing:
                    total_existentes += 1
                    continue'''

                print(f"\n🧾 Procesando pedido: {name}")

                partner_id = None
                currency_id = None
                fiscal_position_id = None
                user_id = None

                # -------------------------------
                # Buscar moneda
                # -------------------------------
                if so["currency_id"]:
                    currency_name = so["currency_id"][1]
                    currencies = models.execute_kw(
                        db, uid, password,
                        "res.currency", "search_read",
                        [[("name", "=", currency_name)]],
                        {"fields": ["id"], "limit": 1}
                    )
                    if currencies:
                        currency_id = currencies[0]["id"]

                if is_from_16:
                    # -------------------------------
                    # Buscar cliente
                    # -------------------------------
                    if so.get('partner_id'):
                        partner_id_origen = so['partner_id'][0]
                        partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)
                        if not partner_id:
                            print("Debug")

                    # -------------------------------
                    # Buscar posición fiscal (con mapa ES → EN)
                    # -------------------------------
                    if so.get("fiscal_position_id"):
                        try:
                            fiscal_position_id_origen = so['fiscal_position_id'][0] if so['fiscal_position_id'][
                                                                                           0] != 27 else 26
                            fiscal_position_id = Utils.get_by_x_id_interno("account.fiscal.position",
                                                                           fiscal_position_id_origen, db, uid, password,
                                                                           models)
                        except Exception as e:
                            fp_es = so["fiscal_position_id"][1]
                            fp_en = FISCAL_POSITION_MAP.get(fp_es)

                            if not fp_en:
                                print(f"⚠️ Posición fiscal sin mapear: '{fp_es}'")
                            else:
                                fps = models.execute_kw(
                                    db, uid, password,
                                    "account.fiscal.position", "search_read",
                                    [[("name", "=", fp_en)]],
                                    {"fields": ["id"], "limit": 1}
                                )
                                if fps:
                                    fiscal_position_id = fps[0]["id"]
                                else:
                                    print(f"⚠️ Posición fiscal destino no encontrada: '{fp_en}'")

                    # -------------------------------
                    # Buscar comercial (user_id)
                    # -------------------------------
                    if so.get("user_id"):
                        user_id_origen = so.get("user_id")[0]
                        #user_id = Utils.get_by_x_id_interno("res.users", user_id_origen, db, uid, password, models)
                        user_name = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "res.users", "read",
                            [[user_id_origen]],
                            {"fields": ["name"]}
                        )[0]["name"]
                        try:
                            user_id = models.execute_kw(
                                db, uid, password,
                                "res.users", "search",
                                [[("name", "=", user_name)]],
                                {"limit": 1}
                            )[0]
                        except:
                            user_id = 43
                else:
                    partner_id = so.get("partner_id")[0]
                    fiscal_position_id = so.get("fiscal_position_id")[0]
                    user_id = so.get("user_id")[0]


                # -------------------------------
                # Crear pedido base
                # -------------------------------
                vals_so = {
                    "x_id_interno": so.get("id"),
                    "name": name,
                    "partner_id": partner_id,
                    "date_order": so.get("date_order"),
                    "currency_id": currency_id,
                    "origin": so.get("origin"),
                    "note": so.get("note"),
                    "x_comentarios": so.get("x_comentarios"),
                    "client_order_ref": so.get("client_order_ref"),
                    "state": state,
                    "fiscal_position_id": fiscal_position_id,
                    "user_id": user_id,
                    "create_date": so.get("create_date"),
                }

                try:
                    new_so_id = models.execute_kw(
                        db, uid, password,
                        "sale.order", "create",
                        [vals_so]
                    )
                    print(f"✅ Pedido creado: {name} (ID {new_so_id})")
                    total_creados += 1
                except Exception as e:
                    print(f"❌ Error creando pedido {name}: {e}")
                    continue

                # -------------------------------
                # Crear líneas
                # -------------------------------
                for linea in so.get("lineas_detalle", []):
                    display_type = linea.get("display_type")

                    # -------------------------------
                    # 📝 NOTA o SECCIÓN
                    # -------------------------------
                    if display_type in ("line_note", "line_section"):
                        vals_line = {
                            "order_id": new_so_id,
                            "name": linea.get("name"),
                            "display_type": display_type,
                            "x_nota_interna": linea.get("x_nota_interna"),
                        }

                        try:
                            models.execute_kw(
                                db, uid, password,
                                "sale.order.line", "create", [vals_line]
                            )
                            print(f"   📝 {display_type} creada: {linea.get('name')}")
                        except Exception as e:
                            print(f"   ⚠️ Error creando nota/sección: {e}")

                        continue
                    # -------------------------------
                    # 📦 Producto (por SKU)
                    # -------------------------------
                    product_id = None
                    sku = linea.get("default_code")

                    if sku:
                        productos = models.execute_kw(
                            db, uid, password,
                            "product.product", "search",
                            [[("default_code", "=", sku)]],
                            {"limit": 1}
                        )
                        product_id = productos[0] if productos else None

                    tax_ids = []
                    if linea.get("tax_id"):
                        tax_origen = linea.get("tax_id")[0]
                        tax_id = Utils.get_by_x_id_interno("account.tax", tax_origen, db, uid, password, models) if is_from_16 else tax_origen
                        tax_ids = [tax_id] if tax_id else []

                    vals_line = {
                        "order_id": new_so_id,
                        "name": linea.get("name"),
                        "product_id": product_id,
                        "product_uom_qty": linea.get("product_uom_qty") or 1.0,
                        "price_unit": linea.get("price_unit") or 0.0,
                        "discount": linea.get("discount", 0.0),
                        "tax_id": [(6, 0, [53])],
                        "x_nota_interna": linea.get("x_nota_interna"),
                    }

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "sale.order.line", "create",
                            [vals_line]
                        )
                        print(f"   ➕ Línea creada: {linea.get('name')}")
                    except Exception as e:
                        print(f"   ⚠️ Error creando línea: {e}")

            #migrar_adjuntos_modelo("sale.order")

            print("\n📊 MIGRACIÓN COMPLETADA")
            print(f"   Total creados: {total_creados}")
            print(f"   Ya existentes: {total_existentes}")

        # Ejecución directa opcional
        # migrar_tareas()

        #Facturas/Pagos clientes/proveedores
        #cuentas analiticas:
        # accounts = export_cuentas_analiticas()
        # import_cuentas_analiticas(accounts)
        # apuntes analiticos:
        # lines = export_apuntes_analiticos()
        # import_apuntes_analiticos(lines)
        #proyectos y tareas
        #comprobar reasignacion: comprobar_reasignaciones_analiticas()
        # Reasignar (si procede)
        # Conciliacion

        def migrar_pedidos_venta_draft():
            orders = export_sale_orders_by_state("draft")
            import_sale_orders_with_lines(orders, "draft")

        def migrar_pedidos_venta_sale():
            orders = export_sale_orders_by_state("sale")
            import_sale_orders_with_lines(orders, "sale")

        def migrar_pedidos_venta_cancel():
            orders = export_sale_orders_by_state("cancel")
            import_sale_orders_with_lines(orders, "cancel")

        # ----------------------------------------------------------------------
        # Función principal opcional
        # ----------------------------------------------------------------------

        def migrar_pedidos_venta():
            migrar_pedidos_venta_sale()
            #migrar_pedidos_venta_draft()
            #migrar_pedidos_venta_cancel()

        #ARCHIVAR: 1, dede, Desconocido, "Invergestión Levante, SL"

        def exportar_tareas_con_adjuntos(company_name="ALMAITANA DE LUZ, S.L."):
            """
            Devuelve tareas de Odoo ORIGEN que tienen al menos un adjunto.
            Se devuelven como dict {task_id: task_data}.
            """
            import xmlrpc.client

            url = 'https://optimaluz.soluntec.net'
            db_old = 'Test'
            username = 'jcoronado@optimaluz.com'
            password_old = 'AlAi4ever'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid_old = common.authenticate(db_old, username, password_old, {})
            if not uid_old:
                print("❌ No se pudo autenticar en ORIGEN")
                return {}

            models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            # Buscar compañía
            company_id = models_old.execute_kw(
                db_old, uid_old, password_old,
                'res.company', 'search',
                [[('name', '=', company_name)]],
                {'limit': 1}
            )
            company_id = company_id[0] if company_id else None

            print("🔍 Buscando tareas con adjuntos en ORIGEN...")

            # 1️⃣ Adjuntos de tareas
            attachments = models_old.execute_kw(
                db_old, uid_old, password_old,
                'ir.attachment', 'search_read',
                [[('res_model', '=', 'project.task')]],
                {'fields': ['res_id']}
            )

            task_ids = sorted({a['res_id'] for a in attachments if a.get('res_id')})

            if not task_ids:
                print("ℹ️ No hay tareas con adjuntos.")
                return {}

            # 2️⃣ Leer datos mínimos de esas tareas
            tasks = models_old.execute_kw(
                db_old, uid_old, password_old,
                'project.task', 'search_read',
                [[('id', 'in', task_ids), ('company_id', '=', company_id)]],
                {'fields': ['id', 'name', 'project_id']}
            )

            print(f"📎 Tareas con adjuntos encontradas: {len(tasks)}")

            return {t['id']: t for t in tasks}

        def migrar_adjuntos_pendientes(task_ids):
            """
            Migra adjuntos solo para tareas que:
            - existen en DESTINO
            - NO tienen adjuntos todavía
            """
            from App_Connection import models, db, uid, password

            print("\n🚀 Iniciando migración inteligente de adjuntos...\n")

            for origen_task_id in task_ids:
                print(f"\n🧩 Procesando tarea ORIGEN ID {origen_task_id}")

                # --------------------------------------------------
                # 1️⃣ Buscar tarea destino por x_origen_id o nombre
                # --------------------------------------------------
                # 🔴 AJUSTA ESTO según cómo mapees tareas
                destino_task = models.execute_kw(
                    db, uid, password,
                    'project.task', 'search_read',
                    [[('x_origen_id', '=', origen_task_id)]],
                    {'fields': ['id', 'name'], 'limit': 1}
                )

                if not destino_task:
                    print("⚠️ Tarea no existe en DESTINO, se omite")
                    continue

                destino_task_id = destino_task[0]['id']
                print(f"   ✔️ Tarea destino encontrada: {destino_task[0]['name']}")

                # --------------------------------------------------
                # 2️⃣ Verificar si ya tiene adjuntos
                # --------------------------------------------------
                adjuntos_destino = models.execute_kw(
                    db, uid, password,
                    'ir.attachment', 'search_count',
                    [[
                        ('res_model', '=', 'project.task'),
                        ('res_id', '=', destino_task_id)
                    ]]
                )

                if adjuntos_destino > 0:
                    print(f"   ⏭️ Ya tiene {adjuntos_destino} adjuntos → se salta")
                    continue

                # --------------------------------------------------
                # 3️⃣ Migrar adjuntos
                # --------------------------------------------------
                print("   🔁 Migrando adjuntos...")
                try:
                    def migrar_adjuntos_tarea(origen_task_id, destino_task_id, lote=30, delay=0.1):
                        """
                        Migra adjuntos de una tarea desde Odoo ORIGEN → DESTINO.

                        origen_task_id : ID de tarea en origen
                        destino_task_id : ID de tarea en destino
                        lote : cantidad de adjuntos por lote
                        delay : pausa entre lotes
                        """
                        import xmlrpc.client
                        import math
                        import time
                        from App_Connection import models, db, uid, password  # DESTINO

                        print(f"\n=== MIGRANDO ADJUNTOS TAREA {origen_task_id} → {destino_task_id} ===")

                        # --------------------------------------------------
                        # 🔹 CONEXIÓN ORIGEN
                        # --------------------------------------------------
                        url = 'https://optimaluz.soluntec.net'
                        db_old = 'Test'
                        username = 'jcoronado@optimaluz.com'
                        password_old = 'AlAi4ever'

                        common_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
                        uid_old = common_old.authenticate(db_old, username, password_old, {})

                        if not uid_old:
                            print("❌ No se pudo autenticar en ORIGEN.")
                            return

                        models_old = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)
                        print(f"🔌 Conectado a ORIGEN como {username} (uid {uid_old})")

                        # --------------------------------------------------
                        # 🔹 LEER ADJUNTOS EN ORIGEN
                        # --------------------------------------------------
                        attach_ids = models_old.execute_kw(
                            db_old, uid_old, password_old,
                            'ir.attachment', 'search',
                            [[
                                ('res_model', '=', 'project.task'),
                                ('res_id', '=', origen_task_id)
                            ]]
                        )

                        total_adjuntos = len(attach_ids)
                        print(f"📎 Adjuntos encontrados en origen: {total_adjuntos}")

                        if total_adjuntos == 0:
                            print("No hay adjuntos que migrar.")
                            return

                        num_lotes = math.ceil(total_adjuntos / lote)

                        # --------------------------------------------------
                        # 🔹 MIGRAR POR LOTES
                        # --------------------------------------------------
                        for i in range(num_lotes):
                            inicio = i * lote
                            fin = inicio + lote
                            lote_ids = attach_ids[inicio:fin]

                            print(f"\nProcesando lote {i + 1}/{num_lotes} ({len(lote_ids)} adjuntos)...")

                            adjuntos = models_old.execute_kw(
                                db_old, uid_old, password_old,
                                'ir.attachment', 'read',
                                [lote_ids, ['name', 'datas', 'mimetype']]
                            )

                            for att in adjuntos:
                                try:
                                    models.execute_kw(
                                        db, uid, password,
                                        'ir.attachment', 'create',
                                        [{
                                            'name': att['name'],
                                            'datas': att['datas'],
                                            'mimetype': att['mimetype'],
                                            'res_model': 'project.task',
                                            'res_id': destino_task_id,
                                        }]
                                    )
                                except Exception as e:
                                    print(f"⚠️ Error importando adjunto '{att['name']}': {e}")

                            print(f"✓ Lote {i + 1}/{num_lotes} completado.")
                            time.sleep(delay)

                        print("\n=== MIGRACIÓN DE ADJUNTOS COMPLETADA ===")

                    migrar_adjuntos_tarea(
                        origen_task_id=origen_task_id,
                        destino_task_id=destino_task_id,
                        lote=10,
                        delay=0.2
                    )
                except Exception as e:
                    print(f"❌ Error migrando adjuntos de tarea {origen_task_id}: {e}")

            print("\n✅ Migración inteligente de adjuntos finalizada")

        # ---------------------------------------------------------------
        # 🔹 MIGRACIÓN DE PROYECTOS (project.project)
        # ---------------------------------------------------------------

        def sincronizar_x_id_proyectos_existentes():
            """
            Busca proyectos ya creados en Odoo destino (por ventas),
            encuentra su equivalente en Odoo origen por nombre,
            y escribe el ID origen en el campo x_id_interno.
            """

            from App_Connection import db, uid, password, models
            import xmlrpc.client
            import re

            def extraer_codigo_opt(texto):
                """
                Extrae secuencia tipo OPT/24/1227 de un string.
                """
                patron = r'OPT/\d+/\d+'
                match = re.search(patron, texto)
                return match.group(0) if match else None

            total_actualizados = 0
            total_no_encontrados = 0

            print("🔎 Sincronizando proyectos existentes...")

            # 1️⃣ Obtener todos los proyectos en destino sin x_id_interno
            proyectos_destino = models.execute_kw(
                db, uid, password,
                "project.project", "search_read",
                [[]],#[[("x_id_interno", "=", 0)]],
                {"fields": ["id", "name"]}
            )

            print(f"📂 {len(proyectos_destino)} proyectos a revisar.")

            for proj_dest in proyectos_destino:
                name = proj_dest["name"]
                codigo = extraer_codigo_opt(name)
                dest_id = proj_dest["id"]

                if name == "Interno": continue

                print(f"\n📁 Procesando: {name}")

                # 2️⃣ Buscar en origen por nombre exacto
                proyectos_origen = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "project.project", "search_read",
                    [[("name", "ilike", codigo)]],
                    {
                        "fields": ["id", "active"],
                        "limit": 1,
                        "context": {"active_test": False}
                    }
                )

                if proyectos_origen:
                    id_origen = proyectos_origen[0]["id"]

                    # 3️⃣ Escribir x_id_interno en destino
                    try:
                        models.execute_kw(
                            db, uid, password,
                            "project.project", "write",
                            [[dest_id], {"x_id_interno": id_origen}]
                        )

                        print(f"✅ Actualizado con ID origen: {id_origen}")
                        total_actualizados += 1

                    except Exception as e:
                        print(f"❌ Error escribiendo proyecto {name}: {e}")

                else:
                    print("⚠️ No encontrado en origen.")
                    total_no_encontrados += 1

            print("\n📊 SINCRONIZACIÓN COMPLETADA")
            print(f"   Actualizados: {total_actualizados}")
            print(f"   No encontrados: {total_no_encontrados}")

        def export_projects():
            """
            Exporta todos los proyectos activos de Odoo 16.
            Incluye campos clave como nombre, responsable, cliente y fechas.
            """
            import xmlrpc.client

            # Buscar compañía origen
            company_id_src = 2

            # Campos de interés
            FIELDS = [
                "name",
                "active",
                "stage_id",
                "user_id",
                "partner_id",
                "company_id",
                "date_start",
                "date",
                "description",
                "privacy_visibility",
                "id"
            ]

            # Buscar proyectos de esa compañía
            projects = models_src.execute_kw(
                db_src, uid_src, password_src,
                "project.project", "search_read",
                [[("company_id", "=", company_id_src), ("active", "in", [True, False])]],
                {"fields": FIELDS}
            )

            print(f"📤 {len(projects)} proyectos exportados correctamente.")
            return projects

        def import_projects(projects):
            """
            Importa los proyectos exportados desde Odoo 16 a Odoo 18.
            Mantiene nombre, responsable, cliente y fechas.
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models

            STAGE_ID_MAP = {
                5: 1,
                7: 2,
                8: 6,
                9: 3,
                10: 4,
                11: 5,
            }

            total_creados = 0
            total_existentes = 0

            def find_user_id(user_name):
                if not user_name: return None
                users = models.execute_kw(
                    db, uid, password,
                    "res.users", "search_read",
                    [[("name", "=", user_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                return users[0]["id"] if users else 43

            def find_partner_id(partner_name, partner_id_src):
                partners = models.execute_kw(
                    db, uid, password,
                    "res.partner", "search_read",
                    [[("name", "=", partner_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                if not partners:
                    partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_src, db, uid, password, models)
                    return partner_id

                return partners[0]["id"] if partners else None

            print("📥 Iniciando importación de proyectos...")

            for proj in projects:
                name = proj.get("name") or "SIN_NOMBRE"
                print(f"\n📁 Procesando proyecto: {name}")

                # Comprobar duplicados
                existing = models.execute_kw(
                    db, uid, password,
                    "project.project", "search",
                    [[("x_id_interno", "=", proj.get("id"))]]
                )
                if existing:
                    total_existentes += 1
                    continue

                # Buscar usuario responsable
                user_id = None
                if proj.get("user_id"):
                    user_name = proj["user_id"][1]
                    user_id = find_user_id(user_name)

                # Buscar cliente
                partner_id = None
                if proj.get("partner_id"):
                    partner_id = proj["partner_id"][0]
                    partner_name = proj["partner_id"][1]
                    partner_id = find_partner_id(partner_name, partner_id)

                stage_id = None
                if proj.get("stage_id"):
                    stage_id_src = proj["stage_id"][0]
                    stage_id = STAGE_ID_MAP.get(stage_id_src)

                # Crear diccionario de valores
                vals = {
                    "name": name,
                    "active": proj.get("active", True),
                    "user_id": user_id,
                    "partner_id": partner_id,
                    "date_start": proj.get("date_start"),
                    "date": proj.get("date"),
                    "description": proj.get("description"),
                    "privacy_visibility": proj.get("privacy_visibility", "followers"),
                    "allow_billable": True,
                    "stage_id": stage_id,
                    "x_id_interno": proj.get("id"),
                }

                # Limpiar valores nulos (evita TypeError: cannot marshal None)
                vals = {k: v for k, v in vals.items() if v is not None}

                try:
                    new_id = models.execute_kw(db, uid, password, "project.project", "create", [vals])
                    print(f"✅ Proyecto creado (ID {new_id})")
                    total_creados += 1

                except Exception as e:
                    print(f"❌ Error creando proyecto {name}: {e}")

            print("\n📊 MIGRACIÓN DE PROYECTOS COMPLETADA")
            print(f"   Total creados: {total_creados}")
            print(f"   Ya existentes: {total_existentes}")

        def actualizar_proyectos_desde_origen():
            """
            Sincroniza proyectos destino con origen:
            - allocated_hours
            - user_id (si está vacío)
            - stage_id según STAGE_ID_MAP
            """

            from App_Connection import db, uid, password, models

            import xmlrpc.client

            STAGE_ID_MAP = {
                5: 1,
                7: 2,
                8: 6,
                9: 3,
                10: 4,
                11: 5,
            }

            def find_user_id(user_name):
                if not user_name:
                    return None

                users = models.execute_kw(
                    db, uid, password,
                    "res.users", "search_read",
                    [[("name", "=", user_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                return users[0]["id"] if users else 43

            print("🔄 Actualizando proyectos desde origen...")

            # 1️⃣ Obtener proyectos destino que tengan x_id_interno
            proyectos_destino = models.execute_kw(
                db, uid, password,
                "project.project", "search_read",
                [[("x_id_interno", "!=", False)]],
                {
                    "fields": ["id", "x_id_interno", "user_id"],
                    "context": {"active_test": False}
                }
            )

            print(f"📂 {len(proyectos_destino)} proyectos encontrados en destino.")

            total_actualizados = 0

            for proj_dest in proyectos_destino:

                dest_id = proj_dest["id"]
                id_origen = proj_dest["x_id_interno"]

                print(f"\n📁 Procesando proyecto destino ID {dest_id} (origen {id_origen})")

                # 2️⃣ Buscar proyecto en origen
                proyectos_origen = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "project.project", "search_read",
                    [[("id", "=", id_origen)]],
                    {
                        "fields": ["allocated_hours", "user_id", "stage_id"],
                        "context": {"active_test": False},
                        "limit": 1
                    }
                )

                if not proyectos_origen:
                    print("⚠️ Proyecto no encontrado en origen")
                    continue

                proj_src = proyectos_origen[0]

                vals_update = {}

                # 3️⃣ allocated_hours
                allocated_hours = proj_src.get("allocated_hours")
                if allocated_hours is not None:
                    vals_update["allocated_hours"] = allocated_hours

                # 4️⃣ user_id (solo si no tiene en destino)
                if not proj_dest.get("user_id") and proj_src.get("user_id"):
                    user_name = proj_src["user_id"][1]
                    user_id_dest = find_user_id(user_name)
                    if user_id_dest:
                        vals_update["user_id"] = user_id_dest

                # 5️⃣ stage_id con mapeo
                if proj_src.get("stage_id"):
                    stage_src_id = proj_src["stage_id"][0]
                    stage_dest_id = STAGE_ID_MAP.get(stage_src_id)

                    if stage_dest_id:
                        vals_update["stage_id"] = stage_dest_id

                # 6️⃣ Escribir si hay algo que actualizar
                if vals_update:
                    try:
                        models.execute_kw(
                            db, uid, password,
                            "project.project", "write",
                            [[dest_id], vals_update]
                        )
                        print(f"✅ Actualizado: {vals_update}")
                        total_actualizados += 1
                    except Exception as e:
                        print(f"❌ Error actualizando proyecto {dest_id}: {e}")
                else:
                    print("ℹ️ Nada que actualizar")

            print("\n📊 ACTUALIZACIÓN COMPLETADA")
            print(f"   Total actualizados: {total_actualizados}")

        # ---------------------------------------------------------------
        # 🔸 Función principal
        # ---------------------------------------------------------------

        def migrar_proyectos():
            proyectos = export_projects()
            import_projects(proyectos)

        # Llamada directa (opcional)
        #migrar_proyectos()

        # ---------------------------------------------------------------
        # 🔹 MIGRACIÓN DE TAREAS (project.task)
        # ---------------------------------------------------------------

        def export_tasks():
            """
            Exporta todas las tareas activas de Odoo 16 con sus campos clave.
            """
            import xmlrpc.client

            # Buscar compañía
            company_id_src = 2

            # Campos relevantes
            FIELDS = [
                "id",
                "name",
                "active",
                "project_id",
                "partner_id",
                "company_id",
                "date_deadline",
                "date_assign",
                "date_last_stage_update",
                "description",
                "stage_id",
                "priority",
                "remaining_hours",
                "total_hours_spent",
                "tag_ids",
                "user_ids"
            ]

            tasks = models_src.execute_kw(
                db_src, uid_src, password_src,
                "project.task", "search_read",
                [[("company_id", "=", company_id_src), ("active", "in", [True, False])]],
                {"fields": FIELDS}
            )

            print(f"📤 {len(tasks)} tareas exportadas correctamente.")
            import_tasks(tasks, adjuntos=True)
            return tasks

        def import_tasks(tasks, adjuntos=True):
            """
            Importa tareas desde Odoo 16 a Odoo 18.
            - Asigna proyecto por x_old_project_id
            - Evita duplicados por x_old_task_id
            - Integra sale_line_id, user_ids y tag_ids en el create
            """
            import time

            def find_user_id(user_name):
                if not user_name: return None
                users = models.execute_kw(
                    db, uid, password,
                    "res.users", "search_read",
                    [[("name", "=", user_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                return users[0]["id"] if users else 43

            def migrar_adjuntos_tarea(origen_task_id, destino_task_id, lote=30, delay=0.1):
                """
                Migra adjuntos de una tarea desde Odoo ORIGEN → DESTINO.

                origen_task_id : ID de tarea en origen
                destino_task_id : ID de tarea en destino
                lote : cantidad de adjuntos por lote
                delay : pausa entre lotes
                """
                import math
                import time

                def safe_value(v):
                    return v[0] if isinstance(v, list) else v

                print(f"\n=== MIGRANDO ADJUNTOS TAREA {origen_task_id} → {destino_task_id} ===")

                # --------------------------------------------------
                # 🔹 LEER ADJUNTOS EN ORIGEN
                # --------------------------------------------------
                attach_ids = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    'ir.attachment', 'search',
                    [[
                        ('res_model', '=', 'project.task'),
                        ('res_id', '=', origen_task_id)
                    ]]
                )

                total_adjuntos = len(attach_ids)
                print(f"📎 Adjuntos encontrados en origen: {total_adjuntos}")

                if total_adjuntos == 0:
                    print("No hay adjuntos que migrar.")
                    return

                num_lotes = math.ceil(total_adjuntos / lote)

                # --------------------------------------------------
                # 🔹 MIGRAR POR LOTES
                # --------------------------------------------------
                for i in range(num_lotes):
                    inicio = i * lote
                    fin = inicio + lote
                    lote_ids = attach_ids[inicio:fin]

                    print(f"\nProcesando lote {i + 1}/{num_lotes} ({len(lote_ids)} adjuntos)...")

                    adjuntos = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        'ir.attachment', 'read',
                        [lote_ids, ['name', 'datas', 'mimetype']]
                    )

                    for att in adjuntos:
                        try:
                            existing = models.execute_kw(
                                db, uid, password,
                                'ir.attachment', 'search',
                                [[
                                    ('res_model', '=', 'account.move'),
                                    ('res_id', '=', destino_task_id),
                                    ('name', '=', att['name']),
                                ]],
                                {'limit': 1}
                            )

                            if existing:
                                continue

                            models.execute_kw(
                                db, uid, password,
                                'ir.attachment', 'create',
                                [{
                                    'name': safe_value(att['name']),
                                    'datas': safe_value(att['datas']),
                                    'mimetype': safe_value(att['mimetype']),
                                    'res_model': 'project.task',
                                    'res_id': destino_task_id,
                                }]
                            )
                        except Exception as e:
                            print(f"⚠️ Error importando adjunto '{att['name']}': {e}")

                    print(f"✓ Lote {i + 1}/{num_lotes} completado.")
                    time.sleep(delay)

                print("\n=== MIGRACIÓN DE ADJUNTOS COMPLETADA ===")

            total_creadas = 0
            total_existentes = 0

            print("📥 Iniciando importación de tareas...")

            # -------------------------------
            # Bucle principal
            # -------------------------------
            for task in tasks:
                old_task_id = task.get("id")
                name = task.get("name") or "SIN_NOMBRE"

                print(f"\n🗂️ Procesando tarea ORIGEN ID {old_task_id}: {name}")

                # -------------------------------
                # Comprobación idempotente
                # -------------------------------
                '''existing = models.execute_kw(
                    db, uid, password,
                    "project.task", "search",
                    [[("x_id_interno", "=", old_task_id)]],
                    {"limit": 1}
                )
                if existing:
                    print(f"⚠️  Tarea ya importada (x_old_task_id={old_task_id}) → se omite")
                    total_existentes += 1
                    continue'''

                # -------------------------------
                # Proyecto (por x_old_project_id)
                # -------------------------------
                project_id = None
                if task.get("project_id"):
                    old_project_id = task["project_id"][0]
                    project_ids = models.execute_kw(
                        db, uid, password,
                        "project.project", "search",
                        [[("x_id_interno", "=", old_project_id)]],
                        {"limit": 1}
                    )
                    if project_ids:
                        project_id = project_ids[0]
                    else:
                        print(f"⚠️ Proyecto destino no encontrado para x_old_project_id={old_project_id}")

                # -------------------------------
                # Usuario responsable (many2one)
                # -------------------------------
                user_id = None
                if task.get("user_id"):
                    user_name_src = task["user_id"][1]
                    find_user_id(user_name_src)
                    #user_id = Utils.get_by_x_id_interno("res.users", user_id_src, db, uid, password, models)

                # -------------------------------
                # Cliente
                # -------------------------------
                partner_id = None
                if task.get("partner_id"):
                    partner_id_origen = task['partner_id'][0]
                    partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)

                # -------------------------------
                # sale_line_id (si existe ya)
                # -------------------------------
                sale_line_id = None
                if task.get("sale_line_id"):
                    sale_line_name = task["sale_line_id"][1]
                    sale_lines = models.execute_kw(
                        db, uid, password,
                        "sale.order.line", "search",
                        [[("name", "=", sale_line_name)]],
                        {"limit": 1}
                    )
                    if sale_lines:
                        sale_line_id = sale_lines[0]

                # -------------------------------
                # user_ids (many2many)
                # -------------------------------
                user_ids = []
                if task.get("user_ids"):
                    users_old = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "res.users", "read",
                        [task["user_ids"]],
                        {"fields": ["name"]}
                    )
                    for u in users_old:
                        uid_found = find_user_id(u["name"])
                        if uid_found:
                            user_ids.append(uid_found)

                # -------------------------------
                # tag_ids (many2many)
                # -------------------------------
                tag_ids = []
                if task.get("tag_ids"):
                    tags_old = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "project.tags", "read",
                        [task["tag_ids"]],
                        {"fields": ["name"]}
                    )
                    for tg in tags_old:
                        found = models.execute_kw(
                            db, uid, password,
                            "project.tags", "search",
                            [[("name", "=", tg["name"])]],
                            {"limit": 1}
                        )
                        if found:
                            tag_ids.append(found[0])

                # -------------------------------
                # Valores de creación
                # -------------------------------
                vals = {
                    "name": name,
                    "active": task.get("active", True),
                    "project_id": project_id,
                    "user_id": user_id,
                    "partner_id": partner_id,
                    "date_deadline": task.get("date_deadline"),
                    "priority": task.get("priority"),
                    "remaining_hours": task.get("remaining_hours"),
                    "description": task.get("description"),
                    "x_id_interno": old_task_id,  # 🔑 clave de idempotencia
                }

                if sale_line_id:
                    vals["sale_line_id"] = sale_line_id
                if user_ids:
                    vals["user_ids"] = [(6, 0, user_ids)]
                if tag_ids:
                    vals["tag_ids"] = [(6, 0, tag_ids)]

                # Limpiar None
                vals = {k: v for k, v in vals.items() if v is not None}

                # -------------------------------
                # Crear tarea
                # -------------------------------
                try:
                    existing = models.execute_kw(
                        db, uid, password,
                        "project.task", "search",
                        [[("x_id_interno", "=", old_task_id)]],
                        {"limit": 1}
                    )
                    if existing:
                        print(f"⚠️  Tarea ya importada (x_old_task_id={old_task_id}) → se omite")
                        total_existentes += 1
                        continue

                    new_id = models.execute_kw(
                        db, uid, password,
                        "project.task", "create",
                        [vals]
                    )
                    print(f"✅ Tarea creada (ID destino {new_id})")
                    total_creadas += 1

                    # -------------------------------
                    # Adjuntos (opcional)
                    # -------------------------------
                    if adjuntos:
                        try:

                            migrar_adjuntos_tarea(
                                origen_task_id=old_task_id,
                                destino_task_id=new_id,
                                lote=10,
                                delay=0.1
                            )

                        except Exception as e:
                            print(f"⚠️ Error migrando adjuntos de la tarea {name}: {e}")

                except Exception as e:
                    print(f"❌ Error creando tarea {name}: {e}")

                time.sleep(0.05)

            # -------------------------------
            # Resumen
            # -------------------------------
            print("\n📊 MIGRACIÓN DE TAREAS COMPLETADA")
            print(f"   Total creadas: {total_creadas}")
            print(f"   Ya existentes: {total_existentes}")

        # ---------------------------------------------------------------
        # 🔸 Función principal
        # ---------------------------------------------------------------

        def migrar_tareas():
            tareas = export_tasks()
            #import_tasks(tareas)

        # Ejecución directa opcional
        #migrar_tareas()

        def export_cuentas_analiticas(company_name="ALMAITANA DE LUZ, S.L."):
            import xmlrpc.client

            url = 'https://optimaluz.soluntec.net'
            db = 'Test'
            username = 'jcoronado@optimaluz.com'
            password = 'AlAi4ever'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid = common.authenticate(db, username, password, {})
            if not uid:
                print("❌ Error de autenticación")
                return []

            models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            company_id = models.execute_kw(
                db, uid, password,
                'res.company', 'search',
                [[('name', '=', company_name)]],
                {'limit': 1}
            )[0]

            accounts = models.execute_kw(
                db, uid, password,
                'account.analytic.account', 'search_read',
                [[('company_id', '=', company_id)]],
                {
                    'fields': [
                        'name',
                        'code',
                        'active',
                        'partner_id',
                    ],
                    'order': 'name'
                }
            )

            print(f"📤 Exportadas {len(accounts)} cuentas analíticas")
            return accounts

        def import_cuentas_analiticas(accounts):
            from App_Connection import db, uid, password, models

            creadas = 0
            actualizadas = 0

            # 🔹 company_id fijo
            company = models.execute_kw(
                db, uid, password,
                'res.company', 'search_read',
                [[('name', '=', "ALMAITANA DE LUZ, S.L.")]],
                {'fields': ['id'], 'limit': 1}
            )
            if not company:
                print("❌ Compañía no encontrada")
                return
            company_id = company[0]['id']

            PLAN_ID = 1  # definido por ti

            for acc in accounts:
                name = acc['name']
                code = acc.get('code')

                print(f"\n📊 Cuenta analítica: {name}")

                domain = [('company_id', '=', company_id)]
                if code:
                    domain.append(('code', '=', code))
                else:
                    domain.append(('name', '=', name))

                # Buscar por name + company
                existing = models.execute_kw(
                    db, uid, password,
                    'account.analytic.account', 'search',
                    [domain],
                    {'limit': 1}
                )

                partner_id = None
                if acc.get('partner_id'):
                    partner_name = acc['partner_id'][1]
                    partner = models.execute_kw(
                        db, uid, password,
                        'res.partner', 'search_read',
                        [[('name', '=', partner_name)]],
                        {'fields': ['id'], 'limit': 1}
                    )
                    if partner:
                        partner_id = partner[0]['id']

                vals = {
                    'name': name,
                    'code': code,
                    'active': acc.get('active', True),
                    'partner_id': partner_id,
                    'plan_id': PLAN_ID,
                    'company_id': company_id,
                }

                try:
                    models.execute_kw(
                        db, uid, password,
                        'account.analytic.account', 'create',
                        [vals]
                    )
                    print("✅ Creada")
                    creadas += 1
                    '''if existing:
                        models.execute_kw(
                            db, uid, password,
                            'account.analytic.account', 'write',
                            [existing, vals]
                        )
                        print("🔄 Existía → actualizada")
                        actualizadas += 1
                    else:
                        models.execute_kw(
                            db, uid, password,
                            'account.analytic.account', 'create',
                            [vals]
                        )
                        print("✅ Creada")
                        creadas += 1'''
                except Exception as e:
                    print(f"❌ Error procesando '{name}': {e}")

            print("\n📊 RESULTADO CUENTAS ANALÍTICAS")
            print(f"   Creadas: {creadas}")
            print(f"   Actualizadas: {actualizadas}")

        def sync_task_sale_line_id():
            import xmlrpc.client
            from App_Connection import db, uid, password, models

            # -----------------------
            # 1️⃣ RESOLVER PRODUCTO DESTINO (UNA VEZ)
            # -----------------------
            '''PRODUCT_NAME = "Instalación eléctrica* "  # ← el name que ya sabes que coincide

            product_dst_ids = models.execute_kw(
                db, uid, password,
                "product.product", "search",
                [[("name", "=", PRODUCT_NAME)]],
                {"limit": 1}
            )

            if not product_dst_ids:
                print(f"❌ Producto '{PRODUCT_NAME}' no encontrado en DESTINO")
                return

            product_dst_id = product_dst_ids[0]
            print(f"📦 Producto DESTINO resuelto: {PRODUCT_NAME} (ID {product_dst_id})")'''
            product_dst_id = 10538 #"Instalación eléctrica* "

            # -----------------------
            # 2️⃣ TAREAS ORIGEN con sale_line
            # -----------------------
            task_ids_src = models_src.execute_kw(
                db_src, uid_src, password_src,
                "project.task", "search",
                [[("sale_line_id", "!=", False), ("company_id", "=", 2)]]
            )

            tasks_src = models_src.execute_kw(
                db_src, uid_src, password_src,
                "project.task", "read",
                [task_ids_src],
                {"fields": ["id", "name", "sale_line_id"]}
            )

            print(f"📋 Tareas origen con sale_line_id: {len(tasks_src)}")

            # -----------------------
            # 3️⃣ ITERAR
            # -----------------------
            for t in tasks_src:
                old_task_id = t["id"]
                task_name = t["name"]

                # 🔹 Leer sale.order.line ORIGEN
                sol_src = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "sale.order.line", "read",
                    [[t["sale_line_id"][0]]],
                    {"fields": ["order_id"]}
                )[0]

                order_name = sol_src["order_id"][1]

                # 🔹 Buscar tarea DESTINO
                task_dst_ids = models.execute_kw(
                    db, uid, password,
                    "project.task", "search",
                    [[("x_id_interno", "=", old_task_id)]],
                    {"limit": 1}
                )

                if not task_dst_ids:
                    print(f"⚠️ Tarea destino no encontrada | {task_name}")
                    continue

                task_dst_id = task_dst_ids[0]

                # 🔹 Buscar sale.order.line DESTINO (CLAVE FINAL)
                sol_dst_ids = models.execute_kw(
                    db, uid, password,
                    "sale.order.line", "search",
                    [[
                        ("order_id.name", "=", order_name),
                        ("product_id", "=", product_dst_id)
                    ]],
                    {"limit": 1}
                )

                if not sol_dst_ids:
                    print(
                        f"❌ Sale line NO encontrada | "
                        f"Tarea='{task_name}' | "
                        f"Pedido='{order_name}'"
                    )
                    continue

                # 🔹 WRITE
                try:
                    models.execute_kw(
                        db, uid, password,
                        "project.task", "write",
                        [[task_dst_id], {"sale_line_id": sol_dst_ids[0]}]
                    )
                    print(f"🔗 Vinculada venta → tarea '{task_name}'")
                except Exception as e:
                    print(f"❌ Error escribiendo tarea '{task_name}': {e}")

            print("✅ Sincronización sale_line_id finalizada")

        def export_purchase_orders_by_state(state):
            """
            Exporta pedidos de compra por estado, incluyendo:
            - fiscal_position_id
            - payment_term_id
            - líneas de producto / nota / sección
            - SKU (default_code) de productos
            """
            company_id_src = 2  # OSE: 1 / ALM: 2

            print(f"📤 Exportando pedidos de compra con estado: {state}")

            # ------------------------------------------------
            # 1️⃣ Buscar pedidos de compra
            # ------------------------------------------------
            po_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "purchase.order", "search",
                [[("state", "=", state), ('company_id', '=', company_id_src),]],
                {"order": "create_date asc"}
            )

            if not po_ids:
                print("ℹ️ No se encontraron pedidos de compra")
                return []

            # ------------------------------------------------
            # 2️⃣ Leer pedidos
            # ------------------------------------------------
            po_fields = [
                "id",
                "name",
                "partner_id",
                "partner_ref",
                "date_order",
                "currency_id",
                "state",
                "order_line",
                "amount_untaxed",
                "amount_tax",
                "amount_total",
                "user_id",
                "notes",
                "origin",
                "company_id",
                "x_comentarios",
                "fiscal_position_id",
                "payment_term_id",
                "priority",
                "create_date",
            ]

            purchase_orders = models_src.execute_kw(
                db_src, uid_src, password_src,
                "purchase.order", "read",
                [po_ids],
                {"fields": po_fields}
            )

            # ------------------------------------------------
            # 3️⃣ Leer todas las líneas de compra
            # ------------------------------------------------
            line_ids = []
            for po in purchase_orders:
                line_ids.extend(po.get("order_line", []))

            line_fields = [
                "order_id",
                "name",
                "display_type",
                "product_id",
                "product_qty",
                "price_unit",
                "price_subtotal",
                "price_total",
                "taxes_id",
                "date_planned",
                "sequence",
                "x_nota_interna",
                "analytic_distribution"
            ]

            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "purchase.order.line", "read",
                [line_ids],
                {"fields": line_fields}
            )

            # ------------------------------------------------
            # 4️⃣ Resolver SKUs de productos
            # ------------------------------------------------
            product_ids = {
                l["product_id"][0]
                for l in lines
                if l.get("product_id")
            }

            sku_map = {}
            if product_ids:
                products = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "product.product", "read",
                    [list(product_ids)],
                    {"fields": ["id", "default_code"]}
                )
                sku_map = {p["id"]: p["default_code"] for p in products}

            # ------------------------------------------------
            # 5️⃣ Agrupar líneas por pedido
            # ------------------------------------------------
            lines_by_order = {}
            for line in lines:
                order_id = line["order_id"][0]
                line["default_code"] = (
                    sku_map.get(line["product_id"][0])
                    if line.get("product_id")
                    else None
                )
                lines_by_order.setdefault(order_id, []).append(line)

            # ------------------------------------------------
            # 6️⃣ Construir estructura final
            # ------------------------------------------------
            result = []

            for po in purchase_orders:
                po_id = po["id"]

                po_data = {
                    "id": po["id"],
                    "name": po["name"],
                    "partner_id": po["partner_id"],
                    "partner_ref": po.get("partner_ref"),
                    "date_order": po.get("date_order"),
                    "currency_id": po.get("currency_id"),
                    "state": po.get("state"),
                    "amount_untaxed": po.get("amount_untaxed"),
                    "amount_tax": po.get("amount_tax"),
                    "amount_total": po.get("amount_total"),
                    "user_id": po.get("user_id"),
                    "notes": po.get("notes"),
                    "origin": po.get("origin"),
                    "company_id": po.get("company_id"),
                    "x_comentarios": po.get("x_comentarios"),
                    "fiscal_position_id": po.get("fiscal_position_id"),
                    "payment_term_id": po.get("payment_term_id"),
                    "lineas_detalle": [],
                }

                for line in sorted(
                        lines_by_order.get(po_id, []),
                        key=lambda l: l.get("sequence", 0)
                ):
                    po_data["lineas_detalle"].append({
                        "name": line.get("name"),
                        "display_type": line.get("display_type"),
                        "default_code": line.get("default_code"),
                        "product_qty": line.get("product_qty"),
                        "price_unit": line.get("price_unit"),
                        "price_subtotal": line.get("price_subtotal"),
                        "price_total": line.get("price_total"),
                        "taxes_id": line.get("taxes_id"),
                        "date_planned": line.get("date_planned"),
                        "sequence": line.get("sequence"),
                        "analytic_distribution": line.get("analytic_distribution"),
                    })

                result.append(po_data)

            print(f"✅ Exportados {len(result)} pedidos de compra")
            return result

        def export_purchase_orders_from_excel(ruta_excel = ruta):
            """
            Exporta pedidos de compra leyendo los IDs desde un Excel (columna 'ID').
            Mantiene exactamente la misma estructura que export_purchase_orders_by_state.
            """

            company_id_src = 2  # OSE: 1 / ALM: 2

            print(f"📤 Leyendo IDs desde Excel: {ruta_excel}")

            # ------------------------------------------------
            # 0️⃣ Leer Excel
            # ------------------------------------------------
            df = pd.read_excel(ruta_excel)

            if "ID" not in df.columns:
                print("❌ El Excel no contiene columna 'ID'")
                return []

            po_ids_excel = df["ID"].dropna().astype(int).tolist()

            if not po_ids_excel:
                print("ℹ️ No hay IDs en el Excel")
                return []

            print(f"🔎 IDs encontrados en Excel: {len(po_ids_excel)}")

            # ------------------------------------------------
            # 1️⃣ Buscar pedidos de compra por ID
            # ------------------------------------------------
            po_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "purchase.order", "search",
                [[("id", "in", po_ids_excel), ('company_id', '=', company_id_src)]],
                {"order": "create_date asc"}
            )

            if not po_ids:
                print("ℹ️ No se encontraron pedidos de compra en origen")
                return []

            # ------------------------------------------------
            # 2️⃣ Leer pedidos
            # ------------------------------------------------
            po_fields = [
                "id",
                "name",
                "partner_id",
                "partner_ref",
                "date_order",
                "currency_id",
                "state",
                "order_line",
                "amount_untaxed",
                "amount_tax",
                "amount_total",
                "user_id",
                "notes",
                "origin",
                "company_id",
                "x_comentarios",
                "fiscal_position_id",
                "payment_term_id",
                "priority",
                "create_date",
            ]

            purchase_orders = models_src.execute_kw(
                db_src, uid_src, password_src,
                "purchase.order", "read",
                [po_ids],
                {"fields": po_fields}
            )

            # ------------------------------------------------
            # 3️⃣ Leer líneas
            # ------------------------------------------------
            line_ids = []
            for po in purchase_orders:
                line_ids.extend(po.get("order_line", []))

            line_fields = [
                "order_id",
                "name",
                "display_type",
                "product_id",
                "product_qty",
                "price_unit",
                "price_subtotal",
                "price_total",
                "taxes_id",
                "date_planned",
                "sequence",
                "x_nota_interna",
                "analytic_distribution"
            ]

            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "purchase.order.line", "read",
                [line_ids],
                {"fields": line_fields}
            )

            # ------------------------------------------------
            # 4️⃣ Resolver SKUs
            # ------------------------------------------------
            product_ids = {
                l["product_id"][0]
                for l in lines
                if l.get("product_id")
            }

            sku_map = {}
            if product_ids:
                products = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "product.product", "read",
                    [list(product_ids)],
                    {"fields": ["id", "default_code"]}
                )
                sku_map = {p["id"]: p["default_code"] for p in products}

            # ------------------------------------------------
            # 5️⃣ Agrupar líneas
            # ------------------------------------------------
            lines_by_order = {}
            for line in lines:
                order_id = line["order_id"][0]
                line["default_code"] = (
                    sku_map.get(line["product_id"][0])
                    if line.get("product_id")
                    else None
                )
                lines_by_order.setdefault(order_id, []).append(line)

            # ------------------------------------------------
            # 6️⃣ Construir resultado
            # ------------------------------------------------
            result = []

            for po in purchase_orders:
                po_id = po["id"]

                po_data = {
                    "id": po["id"],
                    "name": po["name"],
                    "partner_id": po["partner_id"],
                    "partner_ref": po.get("partner_ref"),
                    "date_order": po.get("date_order"),
                    "currency_id": po.get("currency_id"),
                    "state": po.get("state"),
                    "amount_untaxed": po.get("amount_untaxed"),
                    "amount_tax": po.get("amount_tax"),
                    "amount_total": po.get("amount_total"),
                    "user_id": po.get("user_id"),
                    "notes": po.get("notes"),
                    "origin": po.get("origin"),
                    "company_id": po.get("company_id"),
                    "x_comentarios": po.get("x_comentarios"),
                    "fiscal_position_id": po.get("fiscal_position_id"),
                    "payment_term_id": po.get("payment_term_id"),
                    "lineas_detalle": [],
                }

                for line in sorted(
                        lines_by_order.get(po_id, []),
                        key=lambda l: l.get("sequence", 0)
                ):
                    po_data["lineas_detalle"].append({
                        "name": line.get("name"),
                        "display_type": line.get("display_type"),
                        "default_code": line.get("default_code"),
                        "product_qty": line.get("product_qty"),
                        "price_unit": line.get("price_unit"),
                        "price_subtotal": line.get("price_subtotal"),
                        "price_total": line.get("price_total"),
                        "taxes_id": line.get("taxes_id"),
                        "date_planned": line.get("date_planned"),
                        "sequence": line.get("sequence"),
                        "analytic_distribution": line.get("analytic_distribution"),
                    })

                result.append(po_data)

            print(f"✅ Exportados {len(result)} pedidos de compra desde Excel")
            return result

        def import_purchase_orders_with_lines(purchase_orders, state):
            """
            Importa pedidos de compra con:
            - líneas de producto / nota / sección
            - productos por SKU
            - posición fiscal mapeada
            - condiciones de pago
            - restauración correcta del state
            """

            print(f"📥 Importando {len(purchase_orders)} pedidos de compra")

            for po in purchase_orders:
                try:
                    name = po.get("name")
                    #state = po.get("state")
                    # ------------------------------------------------
                    # 1️⃣ PROVEEDOR
                    # ------------------------------------------------
                    partner_id = None
                    if po.get("partner_id"):
                        partner_id_origen = po['partner_id'][0]
                        partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)

                    domain = [
                        ("x_id_interno", "=", po.get("id")),
                    ]

                    if partner_id:
                        domain.append(("partner_id", "=", partner_id))

                    existing_po = models.execute_kw(
                        db, uid, password,
                        "purchase.order", "search",
                        [[domain]],
                        {"limit": 1}
                    )

                    if existing_po:
                        continue

                    # ------------------------------------------------
                    # 2️⃣ POSICIÓN FISCAL
                    # ------------------------------------------------
                    fiscal_position_id = None
                    if po.get("fiscal_position_id"):
                        fiscal_position_id_origen = po['fiscal_position_id'][0] if po['fiscal_position_id'][0] != 27 else 26
                        fiscal_position_id = Utils.get_by_x_id_interno("account.fiscal.position", fiscal_position_id_origen, db, uid, password, models)

                    # ------------------------------------------------
                    # 3️⃣ CONDICIÓN DE PAGO
                    # ------------------------------------------------
                    payment_term_id = None
                    if po.get("payment_term_id"):
                        pt_name = po["payment_term_id"][1]
                        pts = models.execute_kw(
                            db, uid, password,
                            "account.payment.term", "search",
                            [[("name", "=", pt_name)]],
                            {"limit": 1}
                        )
                        if pts:
                            payment_term_id = pts[0]
                        else:
                            print(f"⚠️ Condición de pago no encontrada: {pt_name}")

                    # -------------------------------
                    # Buscar comercial (user_id)
                    # -------------------------------
                    user_id_origen = None
                    if po.get("user_id"):
                        user_id_origen = po.get("user_id")[0]
                    #user_id = Utils.get_by_x_id_interno("res.users", user_id_origen, db, uid, password, models)
                    user_name = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "res.users", "read",
                        [[user_id_origen]],
                        {"fields": ["name"]}
                    )[0]["name"]
                    try:
                        user_id = models.execute_kw(
                            db, uid, password,
                            "res.users", "search",
                            [[("name", "=", user_name)]],
                            {"limit": 1}
                        )[0]
                    except:
                        user_id = 43

                    project_id = None
                    lineas = po.get("lineas_detalle", [])[0].get("analytic_distribution") if po.get("lineas_detalle", []) else False
                    analytic_distribution = lineas
                    if analytic_distribution:
                        analytic_id = int(list(analytic_distribution.keys())[0])

                        project = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "project.project", "search_read",
                            [[("analytic_account_id", "=", analytic_id)]],
                            {"fields": ["id"], "limit": 1, "context": {"active_test": False}}
                        )

                        project_id_src = project[0]["id"] if project else None

                        project_dest = models.execute_kw(
                            db, uid, password,
                            "project.project", "search_read",
                            [[("x_id_interno", "=", project_id_src)]],
                            {"fields": ["id"], "limit": 1}
                        )

                        project_id = project_dest[0]["id"] if project_dest else None

                    # ------------------------------------------------
                    # 4️⃣ CREAR PEDIDO (SIEMPRE EN DRAFT)
                    # ------------------------------------------------
                    vals_po = {
                        "x_id_interno": po.get("id"),
                        "name": name,
                        "partner_id": partner_id,
                        "partner_ref": po.get("partner_ref"),
                        "date_order": po.get("date_order"),
                        "origin": po.get("origin"),
                        "notes": po.get("notes"),
                        "x_comentarios": po.get("x_comentarios"),
                        "fiscal_position_id": fiscal_position_id,
                        "payment_term_id": payment_term_id,
                        "user_id": user_id,
                        "create_date": po.get("create_date"),
                        "priority": po.get("priority"),
                        "project_id": project_id,
                    }

                    try:
                        new_po_id = models.execute_kw(
                            db, uid, password,
                            "purchase.order", "create", [vals_po]
                        )
                        print(f"🧾 Pedido creado (draft): {name}")
                    except Exception as e:
                        print(f"❌ Error creando pedido {name} con id {po.get("id")}: {e}")
                        continue

                    # ------------------------------------------------
                    # 5️⃣ CREAR LÍNEAS
                    # ------------------------------------------------
                    for line in po.get("lineas_detalle", []):
                        display_type = line.get("display_type")

                        # -------------------------------
                        # 📝 NOTA / SECCIÓN
                        # -------------------------------
                        if display_type in ("line_note", "line_section"):
                            vals_line = {
                                "order_id": new_po_id,
                                "name": line.get("name"),
                                "display_type": display_type,
                                "product_qty": 0.0,  # OBLIGATORIO en Compras
                                "sequence": line.get("sequence"),
                            }

                            try:
                                models.execute_kw(
                                    db, uid, password,
                                    "purchase.order.line", "create", [vals_line]
                                )
                            except Exception as e:
                                print(f"⚠️ Error creando nota/sección: {e}")
                            continue

                        # -------------------------------
                        # 📦 PRODUCTO
                        # -------------------------------
                        product_id = None
                        sku = line.get("default_code")

                        if sku:
                            productos = models.execute_kw(
                                db, uid, password,
                                "product.product", "search",
                                [[("default_code", "=", sku)]],
                                {"limit": 1}
                            )
                            if productos:
                                product_id = productos[0]
                            else:
                                productos = models.execute_kw(
                                    db, uid, password,
                                    "product.product", "search",
                                    [[("default_code", "ilike", sku)]],
                                    {"limit": 1, "context": {"active_test": False}}
                                )
                                if productos:
                                    product_id = productos[0]
                                else:
                                    print(f"⚠️ Producto no encontrado (SKU={sku})")

                        # -------------------------------
                        # Buscar impuestos por nombre
                        # -------------------------------

                        tax_ids = []
                        if line.get("taxes_id"):
                            tax_origen = line.get("taxes_id")[0]
                            tax_id = Utils.get_by_x_id_interno("account.tax", tax_origen, db, uid, password, models)
                            tax_ids = [tax_id] if tax_id else []

                        vals_line = {
                            "order_id": new_po_id,
                            "name": line.get("name"),
                            "product_id": product_id,
                            "product_qty": line.get("product_qty") or 1.0,
                            "price_unit": line.get("price_unit") or 0.0,
                            "date_planned": line.get("date_planned"),
                            "sequence": line.get("sequence"),
                            "taxes_id": [(6, 0, tax_ids)],
                        }

                        try:
                            models.execute_kw(
                                db, uid, password,
                                "purchase.order.line", "create", [vals_line]
                            )
                        except Exception as e:
                            print(f"⚠️ Error creando línea de producto: {e}")

                    # ------------------------------------------------
                    # 6️⃣ RESTAURAR STATE ORIGINAL
                    # ------------------------------------------------
                    try:
                        if state == "purchase":
                            models.execute_kw(
                                db, uid, password,
                                "purchase.order", "button_confirm", [[new_po_id]]
                            )
                            print(f"✔ Pedido confirmado: {name}")

                        elif state == "cancel":
                            # confirmar primero (Odoo no cancela desde draft)
                            models.execute_kw(
                                db, uid, password,
                                "purchase.order", "button_confirm", [[new_po_id]]
                            )
                            models.execute_kw(
                                db, uid, password,
                                "purchase.order", "button_cancel", [[new_po_id]]
                            )
                            print(f"⛔ Pedido cancelado: {name}")

                        def sync_picking_estado(pedido_destino_id, pedido_origen_id):
                            """
                            Sincroniza el estado del picking según tu lógica:

                            if estado_origen in ["waiting", "confirmed"]:
                                forzar_estado_correspondiente()
                            elif estado_origen != "assigned":
                                cancelar_y_eliminar_picking()
                            """

                            # -------------------------------------------------
                            # 1️⃣ Obtener picking en origen
                            # -------------------------------------------------
                            picking_origen = models_src.execute_kw(
                                db_src, uid_src, password_src,
                                "stock.picking", "search_read",
                                [[("origin", "!=", False), ("origin", "=", pedido_origen_id)]],
                                {"fields": ["state"], "limit": 1}
                            )

                            if not picking_origen:
                                return

                            estado_origen = picking_origen[0]["state"]

                            # -------------------------------------------------
                            # 2️⃣ Obtener picking destino
                            # -------------------------------------------------
                            picking_destino_ids = models.execute_kw(
                                db, uid, password,
                                "stock.picking", "search",
                                [[("origin", "=", pedido_destino_id)]]
                            )

                            if not picking_destino_ids:
                                return

                            # -------------------------------------------------
                            # 3️⃣ Aplicar TU lógica exacta
                            # -------------------------------------------------
                            for picking_id in picking_destino_ids:

                                if estado_origen in ["waiting", "confirmed"]:

                                    models.execute_kw(
                                        db, uid, password,
                                        "stock.picking", "write",
                                        [[picking_id], {"state": estado_origen}]
                                    )

                                elif estado_origen != "assigned":

                                    # Cancelar
                                    models.execute_kw(
                                        db, uid, password,
                                        "stock.picking", "action_cancel",
                                        [[picking_id]]
                                    )

                                    # Eliminar
                                    models.execute_kw(
                                        db, uid, password,
                                        "stock.picking", "unlink",
                                        [[picking_id]]
                                    )

                        sync_picking_estado(new_po_id, po.get("id"))

                    except Exception as e:
                        print(f"⚠️ Error restaurando estado ({state}) para {name}: {e}")
                except Exception as e:
                    print(f"📥 Importando Fallo: {e} ID: {po.get("id")}")

            migrar_adjuntos_modelo("purchase.order")

            print("✅ Importación de pedidos de compra finalizada")

        def migrar_pedidos_compra_draft():
            orders = export_purchase_orders_by_state("draft")
            import_purchase_orders_with_lines(orders, "draft")

        def migrar_pedidos_compra_purchase():
            orders = export_purchase_orders_by_state("purchase")
            import_purchase_orders_with_lines(orders, "purchase")

        def migrar_pedidos_compra_cancel():
            orders = export_purchase_orders_by_state("cancel")
            import_purchase_orders_with_lines(orders, "cancel")

        def migrar_pedidos_compra():
            migrar_pedidos_compra_purchase()
            migrar_pedidos_compra_draft()
            migrar_pedidos_compra_cancel()

        #migrar_pedidos_compra()

        # ----------------------------------------------------------------------
        # Migrar Invoices
        # ----------------------------------------------------------------------

        def export_invoices_by_state(state, cliente_T_proveedor_F=True, is_rectificativa=False):
            """
            Exporta facturas de Odoo 16 filtradas por estado, optimizado para rendimiento.
            Lee las líneas en bloque para evitar miles de llamadas RPC.
            """
            import xmlrpc.client

            tipo_factura = "out_invoice" if cliente_T_proveedor_F else "in_invoice"
            if is_rectificativa:
                tipo_factura = "out_refund" if cliente_T_proveedor_F else "in_refund"

            FIELDS = [
                "id",
                "name",
                "partner_id",
                "ref",
                "date",
                "invoice_date",
                "invoice_date_due",
                "move_type",
                "state",
                "invoice_line_ids",
                "amount_untaxed",
                "amount_tax",
                "amount_total",
                "currency_id",
                "payment_reference",
                "invoice_payment_term_id",
                "invoice_origin",
                "narration",
                "company_id",
                "invoice_user_id",
                "journal_id",
            ]

            print(f"📤 Exportando facturas en estado '{state}'...")

            # 1️⃣ Exportar facturas
            invoices = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search_read",
                [[
                    ("move_type", "in", [tipo_factura]),
                    ("state", "=", state), ('company_id', '=', 1), # OSE: 1 / ALM: 2
                    ("date", ">=", "2026-01-01"), ("date", "<",  "2027-01-01"),
                ]],
                {"fields": FIELDS}
            )

            print(f"   → {len(invoices)} facturas encontradas.")

            journal_ids = {
                inv["journal_id"][0]
                for inv in invoices
                if inv.get("journal_id")
            }

            journal_code_map = {}

            if journal_ids:
                journals = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.journal", "read",
                    [list(journal_ids)],
                    {"fields": ["id", "code"], "context": {"active_test": False}}
                )

                journal_code_map = {
                    j["id"]: j["code"]
                    for j in journals
                }
            # -------------------------------
            # Añadir journal_code a cada factura
            # -------------------------------
            for inv in invoices:
                if inv.get("journal_id"):
                    inv["journal_code"] = journal_code_map.get(inv["journal_id"][0])
                else:
                    inv["journal_code"] = None

            # 2️⃣ Reunir todos los IDs de líneas
            all_line_ids = []
            for inv in invoices:
                all_line_ids.extend(inv.get("invoice_line_ids", []))

            if not all_line_ids:
                print("⚠️  No se encontraron líneas de factura.")
                return invoices

            print(f"   → {len(all_line_ids)} líneas totales detectadas. Leyendo en bloque...")

            # 3️⃣ Leer todas las líneas en bloque
            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "read",
                [all_line_ids],
                {"fields": [
                    "id",
                    "move_id",
                    "name",
                    "product_id",
                    "quantity",
                    "price_unit",
                    "discount",
                    "price_subtotal",
                    "tax_ids",
                    "account_id",
                    "display_type",
                ]}
            )

            # ------------------------------------------------
            # 4️⃣ Resolver SKUs de productos
            # ------------------------------------------------
            product_ids = {
                l["product_id"][0]
                for l in lines
                if l.get("product_id")
            }

            sku_map = {}
            if product_ids:
                products = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "product.product", "read",
                    [list(product_ids)],
                    {"fields": ["id", "default_code"]}
                )
                sku_map = {p["id"]: p["default_code"] for p in products}

            # 4️⃣ Agrupar las líneas por factura
            grouped_lines = defaultdict(list)
            for line in lines:
                move = line.get("move_id")
                if move:
                    line["default_code"] = (
                        sku_map.get(line["product_id"][0])
                        if line.get("product_id")
                        else None
                    )
                    grouped_lines[move[0]].append(line)

            # 5️⃣ Asignar líneas a cada factura
            for inv in invoices:
                inv_id = inv["id"]
                inv["lineas_detalle"] = grouped_lines.get(inv_id, [])

            print("✅ Líneas asignadas correctamente a cada factura.")
            return invoices

        def import_invoices_with_lines(invoices, state):
            """
            Importa facturas en Odoo 18 junto con sus líneas y mapeo de impuestos.
            Si state='posted', primero crea en draft y luego ejecuta action_post().
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models

            test = False
            total_creadas = 0
            total_existentes = 0

            def safe_execute_line(model, method, args, retries=3, wait=5, timeout=60):
                """
                Ejecuta un método XML-RPC con reintentos y timeout de seguridad.
                Evita bloqueos permanentes por errores de red o cuelgues del servidor.
                """
                import xmlrpc.client
                import time
                import socket

                for intento in range(1, retries + 1):
                    try:
                        # Ajustar timeout del socket
                        socket.setdefaulttimeout(timeout)
                        result = models.execute_kw(db, uid, password, model, method, args)
                        socket.setdefaulttimeout(None)
                        return result
                    except Exception as e:
                        print(f"⚠️ Error ({model}.{method}) intento {intento}/{retries}: {e}")
                        if intento < retries:
                            print(f"   ↪ Reintentando en {wait} segundos...")
                            time.sleep(wait)
                        else:
                            print(f"   ❌ Línea fallida tras {retries} intentos.")
                            return None

            for inv in invoices:
                name = inv.get("name") or inv.get("payment_reference") or "SIN_NOMBRE"
                x_id_interno = inv.get("id")
                print(f"\n🧾 Procesando factura: {name}")

                # Comprobar si ya existe
                existing = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("x_id_interno", "=", x_id_interno)]]
                )
                if existing:
                    # print(f"⚠️  Factura ya existente: {total_existentes}")
                    continue

                # -------------------------------
                # Buscar cliente/proveedor
                # -------------------------------
                partner_id = None
                if inv.get("partner_id"):
                    partner_id_origen = inv.get("partner_id")[0]
                    partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)

                # -------------------------------
                # Buscar moneda
                # -------------------------------
                currency_id = None
                if inv["currency_id"]:
                    currency_name = inv["currency_id"][1]
                    currencies = models.execute_kw(
                        db, uid, password,
                        "res.currency", "search_read",
                        [[("name", "=", currency_name)]],
                        {"fields": ["id"], "limit": 1}
                    )
                    if currencies:
                        currency_id = currencies[0]["id"]

                # -------------------------------
                # Buscar usuario de factura
                # -------------------------------
                user_id_origen = None
                if inv.get("invoice_user_id"):
                    user_id_origen = inv.get("invoice_user_id")[0]
                invoice_user_id = Utils.get_by_x_id_interno("res.users", user_id_origen, db, uid, password, models)

                # -------------------------------
                # Buscar diario contable (por code)
                # -------------------------------
                journal_id = None
                if inv.get("journal_code"):
                    journal_code = inv["journal_code"]

                    journals = models.execute_kw(
                        db, uid, password,
                        "account.journal", "search_read",
                        [[("code", "=", journal_code)]],
                        {
                            "fields": ["id"],
                            "limit": 1,
                            "context": {"active_test": False},
                        }
                    )

                    if journals:
                        journal_id = journals[0]["id"]
                    else:
                        print(f"⚠️ Diario no encontrado: code={journal_code}")

                # -------------------------------
                # Crear factura en borrador siempre
                # -------------------------------
                vals_inv = {
                    "x_id_interno": inv.get("id"),
                    "name": name,
                    "move_type": inv.get("move_type") or "out_invoice",
                    "partner_id": partner_id,
                    "ref": inv.get("ref"),
                    "payment_reference": inv.get("payment_reference"),
                    "date": inv.get("date"),
                    "invoice_date": inv.get("invoice_date"),
                    "invoice_date_due": inv.get("invoice_date_due"),
                    "currency_id": currency_id,
                    "invoice_origin": inv.get("invoice_origin"),
                    "narration": inv.get("narration"),
                    "invoice_user_id": invoice_user_id,
                    "journal_id": journal_id,
                    "state": "draft",
                    "partner_bank_id": "",
                }

                try:
                    new_inv_id = models.execute_kw(
                        db, uid, password,
                        "account.move", "create", [vals_inv]
                    )
                    print(f"✅ Factura creada: {name} (ID {new_inv_id})")
                    total_creadas += 1
                except Exception as e:
                    print(f"❌ Error creando factura {name}: {e}")
                    continue

                # -------------------------------
                # Crear líneas
                # -------------------------------
                for linea in inv.get("lineas_detalle", []):
                    display_type = linea.get("display_type")

                    # -------------------------------
                    # 📝 NOTA / SECCIÓN
                    # -------------------------------
                    if display_type in ("line_note", "line_section"):
                        vals_line = {
                            "x_id_interno": linea.get("id"),
                            "move_id": new_inv_id,
                            "name": linea.get("name"),
                            "display_type": display_type,
                            "quantity": 0.0,
                            "sequence": linea.get("sequence"),
                        }

                        result = safe_execute_line(
                            "account.move.line", "create", [vals_line]
                        )
                        if not result:
                            print(f"⚠️ Error creando nota/sección: {linea.get('name')}")
                        continue

                    # -------------------------------
                    # 📦 LÍNEA NORMAL (PRODUCTO O CONCEPTO)
                    # -------------------------------
                    product_id = None
                    account_id = None

                    sku = linea.get("default_code")

                    # 🔹 Intentar resolver producto por SKU
                    if sku:
                        productos = models.execute_kw(
                            db, uid, password,
                            "product.product", "search",
                            [[("default_code", "=", sku)]],
                            {"limit": 1}
                        )
                        if productos:
                            product_id = productos[0]
                        else:
                            productos = models.execute_kw(
                                db, uid, password,
                                "product.product", "search",
                                [[("default_code", "ilike", sku)]],
                                {"limit": 1, "context": {"active_test": False}}
                            )
                            if productos:
                                product_id = productos[0]
                            else:
                                print(f"⚠️ Producto no encontrado (SKU={sku})")

                    # 🔹 Si NO hay producto → resolver cuenta contable
                    if not product_id:
                        if state in ("out_invoice", "out_refund"):
                            accounts = models.execute_kw(
                                db, uid, password,
                                "account.account", "search",
                                [[
                                    ("account_type", "=", "income"),
                                ]],
                                {"limit": 1}
                            )
                        else:
                            accounts = models.execute_kw(
                                db, uid, password,
                                "account.account", "search",
                                [[
                                    ("account_type", "=", "expense"),
                                ]],
                                {"limit": 1}
                            )

                        if accounts:
                            account_id = accounts[0]
                        else:
                            raise Exception("❌ No se encontró cuenta contable por defecto")
                    cuenta_dest = None
                    try:
                        account_id_origen = linea.get("account_id")
                        if account_id_origen:
                            account_id_origen = account_id_origen[0]

                        # Obtener código cuenta origen
                        acc_origen = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "account.account", "read",
                            [[account_id_origen]],
                            {"fields": ["code"]}
                        )[0]

                        account_code = acc_origen["code"]

                        # Buscar cuenta destino
                        cuenta_dest = models.execute_kw(
                            db, uid, password,
                            "account.account", "search",
                            [[("code", "=", account_code)]],
                            {"limit": 1}
                        )
                    except:
                        pass

                    # -------------------------------
                    # 🔹 Impuestos
                    # -------------------------------
                    tax_ids = []
                    if linea.get("tax_ids"):
                        for tax_origen in linea.get("tax_ids"):
                            tax_id = Utils.get_by_x_id_interno("account.tax", tax_origen, db, uid, password, models)
                            if tax_id:
                                tax_ids.append(tax_id)

                    # -------------------------------
                    # 🔹 Crear línea
                    # -------------------------------
                    vals_line = {
                        "x_id_interno": linea.get("id"),
                        "move_id": new_inv_id,
                        "name": linea.get("name"),
                        "product_id": product_id,
                        "display_type": display_type,
                        "quantity": linea.get("quantity") or 1.0,
                        "price_unit": linea.get("price_unit") or 0.0,
                        "discount": linea.get("discount") or 0.0,
                        "tax_ids": [(6, 0, tax_ids)],
                        "sequence": linea.get("sequence"),
                    }

                    if cuenta_dest: vals_line["account_id"] = cuenta_dest[0]
                    elif account_id: vals_line["account_id"] = account_id

                    result = safe_execute_line(
                        "account.move.line", "create", [vals_line],
                        retries=3, wait=5, timeout=90
                    )

                    if result:
                        print(f"   ➕ Línea creada: {linea.get('name')}")
                    else:
                        print(f"   ⚠️ No se pudo crear la línea: {linea.get('name')}")

                # -------------------------------
                # Publicar si corresponde
                # -------------------------------
                if not test:
                    if state == "posted":
                        try:
                            models.execute_kw(
                                db, uid, password,
                                "account.move", "action_post",
                                [[new_inv_id]]
                            )
                            print(f"   📤 Factura publicada correctamente.")
                        except Exception as e:
                            print(f"   ⚠️ Error al publicar factura {name}: {e}")

                    elif state == "cancel":
                        try:
                            models.execute_kw(
                                db, uid, password,
                                "account.move", "button_cancel",
                                [[new_inv_id]]
                            )
                            print(f"   🚫 Factura cancelada correctamente.")
                        except Exception as e:
                            print(f"   ⚠️ Error cancelando factura {name}: {e}")

            print("\n📊 MIGRACIÓN COMPLETADA")
            print(f"   Total creadas: {total_creadas}")
            print(f"   Ya existentes: {total_existentes}")

        # ----------------------------------------------------------------------
        # Funciones por estado
        # ----------------------------------------------------------------------

        def migrar_facturas_draft(tipo, is_rectificativa):
            invoices = export_invoices_by_state("draft", tipo, is_rectificativa)
            import_invoices_with_lines(invoices, "draft")

        def migrar_facturas_posted(tipo, is_rectificativa):
            invoices = export_invoices_by_state("posted", tipo, is_rectificativa)
            import_invoices_with_lines(invoices, "posted")

        # ----------------------------------------------------------------------
        # Principal
        # ----------------------------------------------------------------------

        def migrar_facturas(tipo, is_rectificativa):
            migrar_facturas_draft(tipo, is_rectificativa)
            migrar_facturas_posted(tipo, is_rectificativa)

        def revertir_todas_las_facturas_a_borrador():
            """
            Revierte a borrador todas las facturas (clientes y proveedores)
            que estén en estado 'posted'
            Compatible Odoo 14+
            """

            ctx = {"active_test": False}

            # Buscar facturas publicadas
            invoice_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[
                    ("move_type", "in", ["out_invoice", "out_refund", "in_invoice", "in_refund"]),
                    ("state", "=", "posted")
                ]],
                {"context": ctx}
            )

            print(f"🔍 Facturas encontradas: {len(invoice_ids)}")

            if not invoice_ids:
                print("ℹ️ No hay facturas que revertir")
                return

            errores = 0

            for inv_id in invoice_ids:
                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.move", "button_draft",
                        [[inv_id]],
                        {"context": ctx}
                    )
                except Exception as e:
                    errores += 1
                    print(f"{errores}: {len(invoice_ids)}")

            print(f"✅ Proceso terminado")
            print(f"📄 Facturas revertidas: {len(invoice_ids) - errores}")
            print(f"⚠️ Errores: {errores}")

        #migrar_facturas(True, False); migrar_facturas(True, True); migrar_facturas(False, False); migrar_facturas(False, True); #migrar_adjuntos_modelo("account.move", True, True)

        #STOCK FLOW
        def existe_picking_destino(name):
            ids = models.execute_kw(
                db, uid, password,
                "stock.picking", "search",
                [[("name", "=", name)]],
                {"limit": 1}
            )
            return ids[0] if ids else None

        def mapear_picking_type(code):
            ids = models.execute_kw(
                db, uid, password,
                "stock.picking.type", "search",
                [[("code", "=", code)]],
                {"limit": 1}
            )
            if not ids:
                raise Exception(f"❌ No se encontró picking_type {code} en destino")
            return ids[0]

        def mapear_location(origen_location_id):
            loc = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.location", "read",
                [[origen_location_id]],
                {"fields": ["complete_name", "usage"]}
            )[0]

            ids = models.execute_kw(
                db, uid, password,
                "stock.location", "search",
                [[("complete_name", "=", loc["complete_name"])]],
                {"limit": 1}
            )

            if not ids:
                raise Exception(f"❌ Ubicación no encontrada: {loc['complete_name']}")

            return ids[0]

        def migrar_picking(origen_id):
            picking = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.picking", "read",
                [[origen_id]],
                {
                    "fields": [
                        "name",
                        "origin",
                        "state",
                        "picking_type_id",
                        "location_id",
                        "location_dest_id",
                        "move_ids_without_package",
                    ]
                }
            )[0]

            # 🔒 Evitar duplicados
            if existe_picking_destino(picking["name"]):
                print(f"⏭️ Picking ya existe: {picking['name']}")
                return

            # 🔎 Tipo
            picking_type = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.picking.type", "read",
                [[picking["picking_type_id"][0]]],
                {"fields": ["code"]}
            )[0]

            picking_type_dest = mapear_picking_type(picking_type["code"])

            # 📍 Ubicaciones
            loc_src = mapear_location(picking["location_id"][0])
            loc_dest = mapear_location(picking["location_dest_id"][0])

            # 🧱 Crear picking
            picking_dest_id = models.execute_kw(
                db, uid, password,
                "stock.picking", "create",
                [{
                    "name": picking["name"],
                    "origin": picking["origin"],
                    "picking_type_id": picking_type_dest,
                    "location_id": loc_src,
                    "location_dest_id": loc_dest,
                    "state": "draft",
                }]
            )

            # 📦 Movimientos
            moves = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.move", "read",
                [picking["move_ids_without_package"]],
                {
                    "fields": [
                        "product_id",
                        "product_uom_qty",
                        "product_uom",
                        "location_id",
                        "location_dest_id",
                    ]
                }
            )

            for m in moves:
                product_dest = Utils.get_by_x_id_interno(
                    "product.product",
                    m["product_id"][0],
                    db, uid, password, models
                )

                models.execute_kw(
                    db, uid, password,
                    "stock.move", "create",
                    [{
                        "name": picking["name"],
                        "product_id": product_dest,
                        "product_uom_qty": m["product_uom_qty"],
                        "product_uom": m["product_uom"][0],
                        "location_id": loc_src,
                        "location_dest_id": loc_dest,
                        "picking_id": picking_dest_id,
                        "state": "confirmed",
                        "procure_method": "make_to_stock",
                    }]
                )

            # 🔐 Forzar estado seguro
            models.execute_kw(
                db, uid, password,
                "stock.picking", "write",
                [[picking_dest_id], {"state": "confirmed"}]
            )

            print(f"✅ Picking migrado: {picking['name']}")

        def migrar_pickings_abiertos():
            print("🚀 Migrando stock.picking abiertos...")

            picking_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.picking", "search",
                [[("state", "not in", ("done", "cancel"))]]
            )

            print(f"📦 Pickings a migrar: {len(picking_ids)}")

            for pid in picking_ids:
                try:
                    migrar_picking(pid)
                except Exception as e:
                    print(f"❌ Error migrando picking {pid}: {e}")

            print("✅ Migración finalizada")

        #migrar_pickings_abiertos()

        # END STOCK FLOW

        # region FACTURACION/PAGOS
        # ---------------------------------------------------------------
        # 🔹 MIGRACIÓN DE PAGOS (Clientes y Proveedores)
        # ---------------------------------------------------------------

        def export_payments(tipo):
            """
            Exporta pagos (account.payment) de Odoo 16.
            tipo puede ser: "clientes", "proveedores" o "todos".
            """
            import xmlrpc.client

            # Buscar compañía
            company_id_src = 2

            # Campos a exportar
            FIELDS = [
                "name",
                "partner_id",
                "amount",
                "payment_type",
                "partner_type",
                "journal_id",
                "ref",
                "state",
                "company_id",
                "currency_id",
                "date"
            ]

            # Construir dominio según tipo
            domain = [("company_id", "=", company_id_src),
                    ("date", ">=", "2026-01-01"),#invoice_date
                    ("date", "<",  "2027-01-01"),]

            if tipo == "clientes":
                domain += [("partner_type", "=", "customer")]
            elif tipo == "proveedores":
                domain += [("partner_type", "=", "supplier")]

            # Buscar pagos
            payments = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.payment", "search_read",
                [domain],
                {"fields": FIELDS}
            )

            journal_ids = {
                pay["journal_id"][0]
                for pay in payments
                if pay.get("journal_id")
            }

            journal_code_map = {}

            if journal_ids:
                journals = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.journal", "read",
                    [list(journal_ids)],
                    {"fields": ["id", "code"], "context": {"active_test": False}}
                )

                journal_code_map = {
                    j["id"]: j["code"]
                    for j in journals
                }
            # -------------------------------
            # Añadir journal_code a cada factura
            # -------------------------------
            for pay in payments:
                if pay.get("journal_id"):
                    pay["journal_code"] = journal_code_map.get(pay["journal_id"][0])

            print(f"📤 {len(payments)} pagos exportados ({tipo}).")
            return payments

        def import_payments(payments):
            """
            Importa los pagos exportados desde Odoo 16 a Odoo 18.
            Compatible con clientes (inbound/customer) y proveedores (outbound/supplier).
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models

            total_creados = 0
            total_existentes = 0
            test = False

            def find_currency_id(currency_name):
                currencies = models.execute_kw(
                    db, uid, password,
                    "res.currency", "search_read",
                    [[("name", "=", currency_name)]],
                    {"fields": ["id"], "limit": 1}
                )
                return currencies[0]["id"] if currencies else None

            print("📥 Iniciando importación de pagos...")

            for pay in payments:
                name = pay.get("name") or pay.get("ref") or "SIN_REF"
                print(f"\n💳 Procesando pago: {name}")

                # Evitar duplicados
                existing = models.execute_kw(
                    db, uid, password,
                    "account.payment", "search",
                    [[("name", "=", name)]]
                )
                if existing:
                    print(f"⚠️  Pago ya existente: {name}")
                    total_existentes += 1
                    continue

                partner_id = None
                if pay.get("partner_id"):
                    partner_id_origen = pay['partner_id'][0]
                    partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)

                # -------------------------------
                # Buscar diario contable (por code)
                # -------------------------------
                journal_id = None
                if pay.get("journal_code"):
                    journal_code = pay["journal_code"]

                    journals = models.execute_kw(
                        db, uid, password,
                        "account.journal", "search_read",
                        [[("code", "=", journal_code)]],
                        {
                            "fields": ["id"],
                            "limit": 1,
                            "context": {"active_test": False},
                        }
                    )

                    if journals:
                        journal_id = journals[0]["id"]
                    else:
                        print(f"⚠️ Diario no encontrado: code={journal_code}")

                currency_id = None
                if pay.get("currency_id"):
                    currency_name = pay["currency_id"][1]
                    currency_id = find_currency_id(currency_name)

                vals = {
                    "name": name,
                    "payment_type": pay.get("payment_type"),  # inbound / outbound
                    "partner_type": pay.get("partner_type"),  # customer / supplier
                    "partner_id": partner_id,
                    "journal_id": journal_id,
                    "amount": pay.get("amount"),
                    "memo": pay.get("ref"),
                    "currency_id": currency_id,
                    "date": pay.get("date")
                }
                new_id = 0
                try:
                    new_id = models.execute_kw(db, uid, password, "account.payment", "create", [vals])
                    print(f"✅ Pago creado correctamente (ID {new_id})")
                    total_creados += 1

                    if not test:
                        # Si el pago estaba validado en origen, validar también en destino
                        if pay.get("state") == "posted":
                            # 1️⃣ Validar el pago
                            models.execute_kw(db, uid, password, "account.payment", "action_validate", [[new_id]])
                            print("   📤 Pago validado/publicado.")
                        elif pay.get("state") == "cancel":
                            models.execute_kw(db, uid, password, "account.payment", "action_cancel", [[new_id]])
                            print("   📤 Pago cancelado.")

                except Exception as e:
                    # 2️⃣ Restaurar nombre original (ya que Odoo reasigna la secuencia)
                    try:
                        models.execute_kw(db, uid, password,
                                          "account.payment", "write",
                                          [[new_id], {"name": name}])
                    except Exception as e:
                        print(f"   ⚠️ No se pudo restaurar el nombre del pago: {e}")
                    print(f"❌ Error creando pago {name}: {e}")

            print("\n📊 MIGRACIÓN DE PAGOS COMPLETADA")
            print(f"   Total creados: {total_creados}")
            print(f"   Ya existentes: {total_existentes}")

        # ---------------------------------------------------------------
        # 🔸 Función principal
        # ---------------------------------------------------------------

        def migrar_pagos(tipo):#"clientes" o "proveedores"
            pagos = export_payments(tipo)
            import_payments(pagos)

        # Llamada principal
        #migrar_pagos("clientes")
        #migrar_pagos("proveedores")
        # endregion

        def importar_stock_desde_excel(ruta_excel, ubicacion='Almai/Stock Almacén'):  # Almai/Stock Almacén #WH/Stock
            from App_Connection import models, db, uid, password
            import pandas as pd

            try:
                df = pd.read_excel(ruta_excel)

                if not {'SKU', 'STOCK'}.issubset(df.columns):
                    print("❌ El Excel debe tener las columnas 'SKU' y 'STOCK'.")
                    return

                print(f"📦 Importando stock a la ubicación '{ubicacion}'...")
                total = len(df)
                actualizados = 0

                # Buscar la ubicación por nombre
                location_ids = models.execute_kw(
                    db, uid, password,
                    'stock.location', 'search',
                    [[('complete_name', '=', ubicacion)]]
                )
                if not location_ids:
                    print(f"❌ No se encontró la ubicación '{ubicacion}'.")
                    return

                location_id = location_ids[0]

                for i, row in df.iterrows():
                    sku = str(row['SKU']).strip() if pd.notna(row['SKU']) else None
                    qty = float(row['STOCK']) if pd.notna(row['STOCK']) else 0

                    if not sku:
                        continue

                    try:
                        # Buscar producto por default_code
                        product_ids = models.execute_kw(
                            db, uid, password,
                            'product.product', 'search',
                            [[('default_code', '=', sku)]]
                        )
                        if not product_ids:
                            print(f"⚠️ [{i + 1}] Producto no encontrado: {sku}")
                            continue

                        product_id = product_ids[0]

                        # Buscar si ya existe un stock.quant para ese producto y ubicación
                        quant_ids = models.execute_kw(
                            db, uid, password,
                            'stock.quant', 'search',
                            [[('product_id', '=', product_id), ('location_id', '=', location_id)]]
                        )

                        if quant_ids:
                            # Actualizar cantidad existente
                            models.execute_kw(
                                db, uid, password,
                                'stock.quant', 'write',
                                [quant_ids, {'quantity': qty}]
                            )
                        else:
                            # Crear nuevo registro de stock.quant
                            models.execute_kw(
                                db, uid, password,
                                'stock.quant', 'create',
                                [{
                                    'product_id': product_id,
                                    'location_id': location_id,
                                    'quantity': qty,
                                }]
                            )

                        actualizados += 1
                        print(f"✅ [{i + 1}] {sku} → cantidad = {qty}")

                    except Exception as e:
                        print(f"⚠️ Error procesando {sku}: {e}")

                print(f"\n🎯 Proceso completado: {actualizados}/{total} productos actualizados correctamente.")

            except Exception as e:
                print(f"❌ Error general: {e}")

        # region PARTE DE HORAS
        # ---------------------------------------------------------------
        # 🔹 MIGRACIÓN DE PARTES DE HORAS (account.analytic.line)
        # ---------------------------------------------------------------

        def export_timesheets():
            """
            Exporta partes de horas desde Odoo 16.
            Incluye referencias a empleado, tarea y proyecto.
            """
            import xmlrpc.client

            # Buscar compañía
            company_id_src = 2

            # Campos relevantes de partes de horas
            FIELDS = [
                "id",
                "date",
                "name",
                "employee_id",
                "project_id",
                "task_id",
                "unit_amount",  # Horas trabajadas
                "time_start",
                "time_stop",
                "company_id",
            ]

            domain = [("company_id", "=", company_id_src), ("employee_id", "!=", False)]
            timesheets = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.analytic.line", "search_read",
                [domain],
                {"fields": FIELDS}
            )

            print(f"📤 {len(timesheets)} partes de horas exportados correctamente.")
            return timesheets

        def import_timesheets(timesheets):
            """
            Importa partes de horas en Odoo 18.
            Vincula correctamente proyecto, tarea y empleado.
            """
            import xmlrpc.client
            from App_Connection import db, uid, password, models

            total_creados = 0
            total_existentes = 0
            total_sin_relacion = 0

            print("📥 Iniciando importación de partes de horas...")

            for line in timesheets:
                name = line.get("name") or "SIN_DESCRIPCIÓN"
                date = line.get("date")
                x_id_interno = line.get("id")
                hours = line.get("unit_amount", 0.0)

                # Comprobar si ya existe
                existing = models.execute_kw(
                    db, uid, password,
                    "account.analytic.line", "search",
                    [[("x_id_interno", "=", x_id_interno)]]
                )
                '''if existing:
                    total_existentes += 1# print(f"⚠️  Factura ya existente: {total_existentes}")
                    continue'''

                print(f"\n🕒 Procesando parte: {name} ({hours}h en {date})")

                # Buscar proyecto y tarea
                project_id = None
                so_line_id = None
                if line.get("project_id"):
                    proyecto = line['project_id'][1]
                    if proyecto == "Interno":
                        project_id = 605
                        #project_id = Utils.get_by_x_id_interno("project.project", project_id, db, uid, password, models)
                    else:
                        project_id_origen = line['project_id'][0]
                        project_id = Utils.get_by_x_id_interno("project.project", project_id_origen, db, uid, password, models)

                    if project_id:
                        project = models.execute_kw(
                            db, uid, password,
                            "project.project", "read",
                            [[project_id], ["sale_line_id"]]
                        )
                        if project and project[0]["sale_line_id"]:
                            so_line_id = project[0]["sale_line_id"][0]

                task_id = None
                if line.get("task_id"):
                    tarea = line['task_id'][1]
                    if tarea == "Ausencias":
                        task_id = 979
                        #task_id = Utils.get_by_x_id_interno("project.task", task_id, db, uid, password, models)
                    else:
                        task_id_origen = line['task_id'][0]
                        task_id = Utils.get_by_x_id_interno("project.task", task_id_origen, db, uid, password, models)

                # Buscar empleado
                employee_id = None
                if line.get("employee_id"):
                    employee_id_origen = line['employee_id'][0]
                    employee_id = Utils.get_by_x_id_interno("hr.employee", employee_id_origen, db, uid, password, models)

                vals = {
                    "name": name,
                    "employee_id": employee_id,
                    "project_id": project_id,
                    "task_id": task_id,
                    #"time_start": start,    #Existe por modulo
                    #"time_stop": stop,     #Existe por modulo
                    "unit_amount": hours,
                    "date": date,
                    "so_line": so_line_id,
                    "x_id_interno": x_id_interno,
                }

                # Eliminar None antes de enviar
                vals = {k: v for k, v in vals.items() if v is not None}

                try:
                    new_id = models.execute_kw(db, uid, password, "account.analytic.line", "create", [vals])
                    print(f"✅ Parte creado (ID {new_id})")
                    total_creados += 1
                except Exception as e:
                    print(f"❌ Error creando parte '{name}': {e}")

            print("\n📊 MIGRACIÓN DE PARTES DE HORAS COMPLETADA")
            print(f"   Total creados: {total_creados}")
            print(f"   Ya existentes: {total_existentes}")
            print(f"   Ignorados (sin empleado o proyecto): {total_sin_relacion}")

        # ---------------------------------------------------------------
        # 🔸 Función principal
        # ---------------------------------------------------------------

        def migrar_partes_horas():
            partes = export_timesheets()
            import_timesheets(partes)

        # Ejecución directa opcional
        #migrar_partes_horas()

        # endregion

        # region ASISTENCIAS
        # ---------------------------------------------------------------
        # 🔹 MIGRACIÓN DE ASISTENCIAS (hr.attendance)
        # ---------------------------------------------------------------

        def export_attendances():
            """
            Exporta todas las asistencias (hr.attendance) de Odoo 16 sin filtrar por compañía.
            Incluye empleado, check_in, check_out y horas trabajadas.
            """
            import xmlrpc.client

            FIELDS = [
                "employee_id",
                "check_in",
                "check_out",
                "worked_hours",
            ]

            domain=[]
            #domain=[("check_in", ">=", "2026-03-03 00:00:00")]
            '''domain = [
                ("check_in", ">=", "2025-09-01 00:00:00"),
                ("check_in", "<=", "2026-04-01 23:59:59"),
                ("employee_id.active", "=", True),
            ]'''

            attendances = models_src.execute_kw(
                db_src, uid_src, password_src,
                "hr.attendance", "search_read",
                [domain], {"fields": FIELDS}
            )

            print(f"📤 {len(attendances)} asistencias exportadas correctamente.")
            return attendances

        def import_attendances(attendances):
            """
            Importa asistencias (hr.attendance) en Odoo 18.
            Asocia correctamente al empleado.
            """

            total_creadas = 0
            total_existentes = 0
            total_sin_empleado = 0

            print("📥 Iniciando importación de asistencias...")

            for att in attendances:
                emp_name = att["employee_id"][1] if att.get("employee_id") else None
                check_in = att.get("check_in")
                check_out = att.get("check_out")
                hours = att.get("worked_hours", 0.0)

                employee_id = None
                if att.get("employee_id"):
                    employee_id_origen = att['employee_id'][0]
                    employee_id = Utils.get_by_x_id_interno("hr.employee", employee_id_origen, db, uid, password,
                                                            models)
                if not employee_id:
                    print(f"⚠️ Empleado '{emp_name}' no encontrado en destino.")
                    total_sin_empleado += 1
                    continue

                vals = {
                    "employee_id": employee_id,
                    "check_in": check_in,
                    "check_out": check_out,
                    "worked_hours": hours,
                }

                # Eliminar None
                vals = {k: v for k, v in vals.items() if v is not None}

                try:
                    new_id = models.execute_kw(db, uid, password, "hr.attendance", "create", [vals])
                    print(f"✅ Asistencia creada (ID {new_id}) - {emp_name} {check_in}")
                    total_creadas += 1
                except Exception as e:
                    print(f"❌ Error creando asistencia de {emp_name}: {e}")

            print("\n📊 MIGRACIÓN DE ASISTENCIAS COMPLETADA")
            print(f"   Total creadas: {total_creadas}")
            print(f"   Ya existentes: {total_existentes}")
            print(f"   Ignoradas (sin empleado): {total_sin_empleado}")

        # ---------------------------------------------------------------
        # 🔸 Función principal
        # ---------------------------------------------------------------

        def migrar_asistencias():
            asistencias = export_attendances()
            import_attendances(asistencias)

        # Ejecución directa opcional
        #migrar_asistencias()

        # endregion

        def recalcular_modelos():
            """
            Fuerza el recálculo de modelos clave tras una migración.
            No modifica datos de negocio, solo recomputa estados y totales.
            Compatible Odoo 14–18.
            """

            ctx = {
                "active_test": False,
                "tracking_disable": True,
                "mail_notrack": True,
            }

            print("🔄 Iniciando recálculo de modelos...")

            # 1️⃣ Proyectos (horas, costes, progreso)
            project_ids = models.execute_kw(
                db, uid, password,
                "project.project", "search",
                [[]],
                {"context": ctx}
            )

            if project_ids:
                print(f"   → Recalculando {len(project_ids)} proyectos...")
                models.execute_kw(
                    db, uid, password,
                    "project.project", "write",
                    [project_ids, {}],
                    {"context": ctx}
                )

            # 2️⃣ Ventas (totales, estados)
            sale_ids = models.execute_kw(
                db, uid, password,
                "sale.order", "search",
                [[]],
                {"context": ctx}
            )

            if sale_ids:
                print(f"   → Recalculando {len(sale_ids)} pedidos de venta...")
                models.execute_kw(
                    db, uid, password,
                    "sale.order", "write",
                    [sale_ids, {}],
                    {"context": ctx}
                )

            # 3️⃣ Compras
            purchase_ids = models.execute_kw(
                db, uid, password,
                "purchase.order", "search",
                [[]],
                {"context": ctx}
            )

            if purchase_ids:
                print(f"   → Recalculando {len(purchase_ids)} pedidos de compra...")
                models.execute_kw(
                    db, uid, password,
                    "purchase.order", "write",
                    [purchase_ids, {}],
                    {"context": ctx}
                )

            # 4️⃣ Stock (solo si existe)
            try:
                picking_ids = models.execute_kw(
                    db, uid, password,
                    "stock.picking", "search",
                    [[]],
                    {"context": ctx}
                )

                if picking_ids:
                    print(f"   → Recalculando {len(picking_ids)} albaranes...")
                    models.execute_kw(
                        db, uid, password,
                        "stock.picking", "write",
                        [picking_ids, {}],
                        {"context": ctx}
                    )
            except Exception:
                print("ℹ️ Módulo stock no presente o no accesible.")

            # 5️⃣ Facturación (facturas y abonos)
            invoice_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[("move_type", "in", ["out_invoice", "out_refund", "in_invoice", "in_refund"])]],
                {"context": ctx}
            )

            if invoice_ids:
                print(f"   → Recalculando {len(invoice_ids)} facturas...")
                models.execute_kw(
                    db, uid, password,
                    "account.move", "write",
                    [invoice_ids, {}],
                    {"context": ctx}
                )

            # 6️⃣ Apuntes analíticos
            aal_ids = models.execute_kw(
                db, uid, password,
                "account.analytic.line", "search",
                [[]],
                {"context": ctx}
            )

            if aal_ids:
                print(f"   → Recalculando {len(aal_ids)} apuntes analíticos...")
                models.execute_kw(
                    db, uid, password,
                    "account.analytic.line", "write",
                    [aal_ids, {}],
                    {"context": ctx}
                )

            print("✅ Recalculo completado correctamente.")

        def conciliar_pagos():
            """
            Conciliación segura de pagos con facturas (clientes y proveedores)
            Compatible Odoo 16–18 (XML-RPC)
            """

            ctx = {
                "active_test": False,
                "tracking_disable": True,
                "mail_notrack": True,
            }

            print("🔗 Iniciando conciliación de pagos...")

            # 1️⃣ Obtener cuentas a cobrar / pagar (Odoo 16+)
            account_ids = models.execute_kw(
                db, uid, password,
                "account.account", "search",
                [[("account_type", "in", ["asset_receivable", "liability_payable"])]],
                {"context": ctx}
            )

            if not account_ids:
                print("⚠️ No se encontraron cuentas receivable/payable.")
                return

            print(f"   → Cuentas receivable/payable: {len(account_ids)}")

            # 2️⃣ Buscar líneas no conciliadas
            aml_ids = models.execute_kw(
                db, uid, password,
                "account.move.line", "search",
                [[
                    ("reconciled", "=", False),
                    ("account_id", "in", account_ids),
                    ("move_id.state", "=", "posted"),
                ]],
                {"context": ctx}
            )

            print(f"   → Líneas pendientes de conciliación: {len(aml_ids)}")

            if not aml_ids:
                print("ℹ️ No hay líneas para conciliar.")
                return

            # 3️⃣ Leer datos clave
            lines = models.execute_kw(
                db, uid, password,
                "account.move.line", "read",
                [aml_ids],
                {"fields": [
                    "id",
                    "partner_id",
                    "account_id",
                    "debit",
                    "credit",
                    "balance",
                    "move_id",
                    "date",
                ]}
            )

            groups = defaultdict(list)

            # 4️⃣ Agrupar por partner + cuenta
            for l in lines:
                if not l.get("partner_id"):
                    continue
                key = (l["partner_id"][0], l["account_id"][0])
                groups[key].append(l)

            conciliadas = 0
            errores = 0

            # 5️⃣ Conciliar por grupo
            for (partner_id, account_id), group_lines in groups.items():
                if len(group_lines) < 2:
                    continue

                line_ids = [l["id"] for l in group_lines]

                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.move.line", "reconcile",
                        [line_ids],
                        {"context": ctx}
                    )
                    conciliadas += 1
                except Exception as e:
                    errores += 1
                    print(
                        f"❌ Error conciliando partner {partner_id} "
                        f"cuenta {account_id}: {e}"
                    )

            print("✅ Conciliación finalizada")
            print(f"   ✔ Grupos conciliados: {conciliadas}")
            print(f"   ⚠️ Grupos con error: {errores}")

        def exportar_ids_contactos_origen():
            """
            Exporta IDs y nombres de contactos desde el Odoo origen
            (incluye archivados)
            """

            url = 'https://optimaluz.soluntec.net'
            db = 'Test'
            username = 'jcoronado@optimaluz.com'
            password = 'AlAi4ever'

            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common', allow_none=True)
            uid = common.authenticate(db, username, password, {})

            if not uid:
                raise Exception("❌ No se pudo autenticar en el Odoo origen")

            models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object', allow_none=True)

            partners = models.execute_kw(
                db, uid, password,
                "res.partner", "search_read",
                [[]],
                {
                    "fields": ["id", "name"],
                    "context": {"active_test": False}
                }
            )

            print(f"📤 Contactos exportados (origen): {len(partners)}")

            # Devuelve un mapa name → id_origen
            return {
                p["name"].strip(): p["id"]
                for p in partners
                if p.get("name")
            }

        def actualizar_id_externo_contactos_por_name(mapa_name_id_externo):
            """
            Actualiza x_id_interno en res.partner
            buscando por name (incluye archivados)
            """

            CTX_ALL = {"active_test": False}

            actualizados = 0
            no_encontrados = 0
            errores = 0

            for name, id_externo in mapa_name_id_externo.items():
                try:
                    partner_ids = models.execute_kw(
                        db, uid, password,
                        "res.partner", "search",
                        [[("name", "=", name)]],
                        {"limit": 1, "context": CTX_ALL}
                    )

                    if not partner_ids:
                        no_encontrados += 1
                        continue

                    partner_id = partner_ids[0]

                    models.execute_kw(
                        db, uid, password,
                        "res.partner", "write",
                        [[partner_id], {"x_id_interno": id_externo}],
                        {"context": CTX_ALL}
                    )

                    actualizados += 1

                except Exception as e:
                    errores += 1
                    print(f"❌ Error contacto '{name}': {e}")

            print("✅ Proceso terminado (contactos)")
            print(f"   ✔ Contactos actualizados: {actualizados}")
            print(f"   ⚠️ Contactos no encontrados: {no_encontrados}")
            print(f"   ❌ Errores: {errores}")

        def migrar_extractos_bancarios():
            print("🔄 Iniciando migración de extractos bancarios...")

            # ---------- OBTENER COMPANY & JOURNAL DESTINO ----------
            company_id = models.execute_kw(
                db, uid, password,
                "res.company", "search",
                [[("name", "=", "ALMAITANA DE LUZ, S.L.")]],
                {"limit": 1}
            )[0]

            journal_id = models.execute_kw(
                db, uid, password,
                "account.journal", "search",
                [[("name", "=", "Caixabank 00221472")]],
                {"limit": 1}
            )[0]

            # ---------- BUSCAR EXTRACTOS ORIGEN ----------
            statement_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.bank.statement", "search",
                [[]],
                #{"limit": 3}
            )

            print(f"📄 Extractos encontrados: {len(statement_ids)}")

            for st in models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.bank.statement", "read",
                    [statement_ids],
                    {"fields": [
                        "name", "date", "balance_start",
                        "balance_end", "balance_end_real",
                        "reference", "first_line_index",
                        "line_ids"
                    ]}
            ):
                vals_statement = {
                    "name": st["name"] or "",
                    "date": st["date"],
                    "balance_start": st["balance_start"],
                    "balance_end": st["balance_end"],
                    "balance_end_real": st["balance_end_real"],
                    "reference": st.get("reference"),
                    "first_line_index": st.get("first_line_index", 1),
                    "company_id": company_id,
                    "journal_id": journal_id,
                }

                new_statement_id = models.execute_kw(
                    db, uid, password,
                    "account.bank.statement", "create",
                    [vals_statement]
                )

                print(f"✅ Extracto creado: {st['name']}")

                # ---------- LÍNEAS DEL EXTRACTO ----------
                lines = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.bank.statement.line", "read",
                    [st["line_ids"]],
                    {"fields": ["date", "payment_ref", "partner_id", "amount"]}
                )

                for line in lines:
                    partner_id = False
                    if line["partner_id"]:
                        partner_name = line["partner_id"][1]
                        res = models.execute_kw(
                            db, uid, password,
                            "res.partner", "search",
                            [[("name", "=", partner_name)]],
                            {"limit": 1}
                        )
                        if res:
                            partner_id = res[0]

                    vals_line = {
                        "statement_id": new_statement_id,
                        "date": line["date"],
                        "payment_ref": line.get("payment_ref"),
                        "partner_id": partner_id,
                        "amount": line["amount"],
                        "company_id": company_id,
                        "journal_id": journal_id,
                    }

                    models.execute_kw(
                        db, uid, password,
                        "account.bank.statement.line", "create",
                        [vals_line]
                    )

                print(f"   ↳ {len(lines)} líneas creadas")

            print("🏁 Migración de extractos bancarios finalizada")

        #Asientos manuales: def migrar_extractos_bancarios, def migrar_nominas, def operation_varias_manuales?, Apertura: APE1/22/001

        # region CONCILIACION
        from datetime import datetime

        LOG_FILE = "errores_conciliacion.txt"

        def log_error(msg):
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"[{timestamp}] {msg}\n")

        def obtener_cuenta_factura_origen(factura_name):
            factura_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [[("name", "=", factura_name)]],
                {"limit": 1}
            )
            if not factura_ids:
                raise Exception(f"Factura origen no encontrada: {factura_name}")

            factura_id = factura_ids[0]

            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[("move_id", "=", factura_id), ("account_id.reconcile", "=", True),]],
                {"fields": ["account_id", "debit", "credit"]}
            )

            for l in lines:
                if not l["account_id"]:
                    continue

                acc = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.account", "read",
                    [[l["account_id"][0]]],
                    {"fields": ["account_type", "code"]}
                )[0]

                if acc["account_type"]:# in ("asset_receivable", "liability_payable"):
                    return {
                        "account_id": l["account_id"][0],
                        "account_code": acc["code"],
                    }

            raise Exception(f"No se encontró cuenta receivable/payable en origen {factura_name}")

        def obtener_linea_factura_destino(factura_dest_id):
            lines = models.execute_kw(
                db, uid, password,
                "account.move.line", "search_read",
                [[("move_id", "=", factura_dest_id),("account_id.reconcile", "=", True),]],
                {"fields": ["id", "account_id", "debit", "credit"]}
            )

            for l in lines:
                if not l["account_id"]:
                    continue

                acc = models.execute_kw(
                    db, uid, password,
                    "account.account", "read",
                    [[l["account_id"][0]]],
                    {"fields": ["account_type", "code"]}
                )[0]

                if acc["account_type"]:# in ("asset_receivable", "liability_payable"):
                    return {
                        "line_id": l["id"],
                        "account_id": l["account_id"][0],
                        "account_code": acc["code"],
                    }

            raise Exception("No se encontró línea principal en factura destino")

        def mapear_cuenta_origen_a_destino_por_code(account_code):
            ids = models.execute_kw(
                db, uid, password,
                "account.account", "search",
                [[("code", "=", account_code)]],
                {"limit": 1}
            )
            if not ids:
                raise Exception(f"Cuenta destino no encontrada para code {account_code}")
            return ids[0]


        '''def obtener_linea_principal_factura_destino(factura_dest_id):
            """
            Devuelve la línea receivable/payable correcta de una factura
            (cliente o proveedor), evitando cuentas alternativas (431, 409, etc.)
            """

            lines = models.execute_kw(
                db, uid, password,
                "account.move.line", "search_read",
                [[("move_id", "=", factura_dest_id)]],
                {"fields": ["id", "account_id", "debit", "credit"]}
            )

            for l in lines:
                if not l["account_id"]:
                    continue

                acc = models.execute_kw(
                    db, uid, password,
                    "account.account", "read",
                    [[l["account_id"][0]]],
                    {"fields": ["account_type"]}
                )[0]

                # 🧾 Factura de CLIENTE → credit en receivable
                if acc["account_type"] == "asset_receivable": #and l["credit"] > 0:
                    return l

                # 🧾 Factura de PROVEEDOR → debit en payable
                if acc["account_type"] == "liability_payable": #and l["debit"] > 0:
                    return l

            raise Exception("❌ No se encontró línea principal de la factura")
'''

        def obtener_linea_principal_factura_destino(factura_dest_id):
            """
            Devuelve la línea conciliable principal de la factura en destino.
            Compatible con 430, 431, 410, 411, etc.
            """

            lines = models.execute_kw(
                db, uid, password,
                "account.move.line", "search_read",
                [[
                    ("move_id", "=", factura_dest_id),
                    ("account_id.reconcile", "=", True)
                ]],
                {
                    "fields": [
                        "id",
                        "account_id",
                        "debit",
                        "credit",
                    ]
                }
            )

            if not lines:
                raise Exception("❌ No se encontró línea conciliable en la factura destino")

            # Si hay varias (raro pero posible), devolver la de mayor importe
            linea_principal = max(
                lines,
                key=lambda x: abs(x["debit"] - x["credit"])
            )

            return linea_principal

        def migrar_adjuntos_factura(origen_invoice_id, destino_invoice_id, lote=5):
            """
            Migra adjuntos PDF de una factura desde Odoo ORIGEN → DESTINO.

            origen_invoice_id  : ID de factura en origen (account.move)
            destino_invoice_id : ID de factura en destino (account.move)
            lote               : cantidad de adjuntos por lote
            """
            import math
            import time

            print(f"\n=== MIGRANDO PDFs FACTURA {origen_invoice_id} → {destino_invoice_id} ===")

             # --------------------------------------------------
            # 🔹 BUSCAR ADJUNTOS PDF EN ORIGEN
            # --------------------------------------------------
            attach_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                'ir.attachment', 'search',
                [[
                    ('res_model', '=', 'account.move'),
                    ('res_id', '=', origen_invoice_id),
                    ('mimetype', '=', 'application/pdf')
                ]]
            )

            total_adjuntos = len(attach_ids)
            print(f"📎 PDFs encontrados en origen: {total_adjuntos}")

            if total_adjuntos == 0:
                print("No hay PDFs que migrar.")
                return

            num_lotes = math.ceil(total_adjuntos / lote)

            # --------------------------------------------------
            # 🔹 MIGRAR POR LOTES
            # --------------------------------------------------
            for i in range(num_lotes):
                inicio = i * lote
                fin = inicio + lote
                lote_ids = attach_ids[inicio:fin]

                print(f"\nProcesando lote {i + 1}/{num_lotes} ({len(lote_ids)} PDFs)...")

                adjuntos = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    'ir.attachment', 'read',
                    [lote_ids, ['name', 'datas', 'mimetype']]
                )

                for att in adjuntos:
                    try:
                        existing = models.execute_kw(
                            db, uid, password,
                            'ir.attachment', 'search',
                            [[
                                ('res_model', '=', 'account.move'),
                                ('res_id', '=', destino_invoice_id),
                                ('name', '=', att['name']),
                            ]],
                            {'limit': 1}
                        )

                        if existing:
                            continue

                        models.execute_kw(
                            db, uid, password,
                            'ir.attachment', 'create',
                            [{
                                'name': att['name'],
                                'datas': att['datas'],
                                'mimetype': att['mimetype'],
                                'res_model': 'account.move',
                                'res_id': destino_invoice_id,
                            }]
                        )
                    except Exception as e:
                        print(f"⚠️ Error importando PDF '{att['name']}': {e}")

                print(f"✓ Lote {i + 1}/{num_lotes} completado.")

            print("\n=== MIGRACIÓN DE PDFs DE FACTURA COMPLETADA ===")

        '''def detectar_pagos_factura_origen(factura_name):
            """
            Devuelve una lista de dicts con las líneas EXACTAS de pago
            conciliadas con la factura en origen, incluyendo el importe.
            """

            resultados = []

            # 1️⃣ Buscar factura
            factura_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [[("name", "=", factura_name)]],
                {"limit": 1}
            )

            if not factura_ids:
                print(f"❌ Factura {factura_name} no encontrada")
                return resultados

            factura_id = factura_ids[0]

            # 2️⃣ Obtener líneas contables de la factura
            factura_lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[("move_id", "=", factura_id)]],
                {
                    "fields": [
                        "id",
                        "account_id",
                        "matched_debit_ids",
                        "matched_credit_ids",
                    ]
                }
            )

            # 3️⃣ Localizar línea receivable/payable de la factura
            linea_factura_id = None

            for l in factura_lines:
                if not l["account_id"]:
                    continue
                acc = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.account", "read",
                    [[l["account_id"][0]]],
                    {"fields": ["account_type"]}
                )[0]

                if acc["account_type"] in (
                        "asset_receivable",
                        "liability_payable",
                ):
                    linea_factura_id = l["id"]
                    matched_ids = set(
                        l["matched_debit_ids"] + l["matched_credit_ids"]
                    )
                    break

            if not linea_factura_id:
                print("⚠️ No se encontró línea receivable/payable en factura")
                return resultados

            if not matched_ids:
                #print("⚠️ La factura no tiene conciliaciones")
                return resultados

            # 4️⃣ Leer partial reconciles
            partials = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.partial.reconcile", "read",
                [list(matched_ids)],
                {
                    "fields": [
                        "debit_move_id",
                        "credit_move_id",
                        "amount",
                    ]
                }
            )

            # 5️⃣ Procesar cada partial reconcile
            for pr in partials:

                if pr["debit_move_id"][0] == linea_factura_id:
                    pago_line_id = pr["credit_move_id"][0]
                elif pr["credit_move_id"][0] == linea_factura_id:
                    pago_line_id = pr["debit_move_id"][0]
                else:
                    continue  # no debería pasar, pero por seguridad

                # 6️⃣ Leer línea de pago
                pago_line = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "read",
                    [[pago_line_id]],
                    {"fields": ["move_id"]}
                )[0]

                pago_move = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move", "read",
                    [[pago_line["move_id"][0]]],
                    {"fields": ["id", "name", "date", "journal_id"]}
                )[0]

                resultados.append({
                    "pago_move_id": pago_move["id"],
                    "pago_move_name": pago_move["name"],
                    "pago_line_id": pago_line_id,
                    "amount": pr["amount"],
                })

            return resultados
'''

        def detectar_pagos_factura_origen(factura_name):
            """
            Devuelve una lista de dicts con las líneas EXACTAS de pago
            conciliadas con la factura en origen, incluyendo el importe.
            """

            resultados = []

            # 1️⃣ Buscar factura
            factura_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [[("name", "=", factura_name)]],
                {"limit": 1}
            )

            if not factura_ids:
                print(f"❌ Factura {factura_name} no encontrada")
                return resultados

            factura_id = factura_ids[0]

            # 2️⃣ Obtener SOLO líneas conciliables de la factura
            factura_lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[
                    ("move_id", "=", factura_id),
                    ("account_id.reconcile", "=", True),
                ]],
                {
                    "fields": [
                        "id",
                        "matched_debit_ids",
                        "matched_credit_ids",
                    ]
                }
            )

            if not factura_lines:
                print("⚠️ No hay líneas conciliables en la factura")
                return resultados

            # Normalmente solo habrá una línea conciliable
            for linea in factura_lines:

                matched_ids = set(
                    linea["matched_debit_ids"] + linea["matched_credit_ids"]
                )

                if not matched_ids:
                    continue

                # 3️⃣ Leer partial reconciles
                partials = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.partial.reconcile", "read",
                    [list(matched_ids)],
                    {
                        "fields": [
                            "debit_move_id",
                            "credit_move_id",
                            "amount",
                        ]
                    }
                )

                for pr in partials:

                    if pr["debit_move_id"][0] == linea["id"]:
                        pago_line_id = pr["credit_move_id"][0]
                    elif pr["credit_move_id"][0] == linea["id"]:
                        pago_line_id = pr["debit_move_id"][0]
                    else:
                        continue

                    pago_line = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.move.line", "read",
                        [[pago_line_id]],
                        {"fields": ["move_id"]}
                    )[0]

                    pago_move = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.move", "read",
                        [[pago_line["move_id"][0]]],
                        {"fields": ["id", "name", "date", "journal_id"]}
                    )[0]

                    resultados.append({
                        "pago_move_id": pago_move["id"],
                        "pago_move_name": pago_move["name"],
                        "pago_line_id": pago_line_id,
                        "amount": pr["amount"],
                    })

            return resultados

        def obtener_lineas_conciliables_destino(move_id):
            """
            Devuelve TODAS las líneas conciliables (account.reconcile=True)
            de un account.move en destino.
            """

            lines = models.execute_kw(
                db, uid, password,
                "account.move.line", "search_read",
                [[
                    ("move_id", "=", move_id),
                    ("account_id.reconcile", "=", True)
                ]],
                {
                    "fields": [
                        "id",
                        "account_id",
                        "debit",
                        "credit",
                        "reconciled",
                    ]
                }
            )

            if not lines:
                raise Exception(
                    f"❌ No se encontraron líneas conciliables en move {move_id}"
                )

            return lines

        def conciliar_factura_con_pagos_destino(factura_dest_id, pagos_dest, factura_name, payment_state):
            try:
                def obtener_linea_destino_por_id_origen(move_id, origen_line_id):
                    """
                    Devuelve la línea receivable/payable en destino
                    que corresponde exactamente a la línea de origen.
                    """

                    def es_linea_de_factura_origen(origen_line_id):
                        line = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "account.move.line", "read",
                            [[origen_line_id]],
                            {"fields": ["move_id"]}
                        )[0]

                        move = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "account.move", "read",
                            [[line["move_id"][0]]],
                            {"fields": ["move_type"]}
                        )[0]

                        return move["move_type"] in (
                            "out_invoice", "out_refund",
                            "in_invoice", "in_refund"
                        )

                    # 🟢 CASO A: línea de asiento contable (pago)
                    try:
                        lines = models.execute_kw(
                            db, uid, password,
                            "account.move.line", "search_read",
                            [[
                                ("move_id", "=", move_id),
                                ("x_id_interno", "=", origen_line_id)
                            ]],
                            {"fields": ["id", "account_id", "balance"]}
                        )
                        if lines:
                            return lines[0]
                    except Exception:
                        pass

                    # 🟡 CASO B: línea de FACTURA (reversión)
                    if es_linea_de_factura_origen(origen_line_id):
                        return obtener_lineas_conciliables_destino(
                            move_id
                        )

                    raise Exception(
                        f"No se pudo resolver línea destino para origen_line_id={origen_line_id}"
                    )

                ctx = {
                    "active_test": False,
                    "tracking_disable": True,
                    "mail_notrack": True,
                }

                line_ids = []

                # 1️⃣ Líneas receivable/payable de la factura
                linea_factura = obtener_linea_principal_factura_destino(factura_dest_id)
                line_ids.append(linea_factura["id"])

                # 2️⃣ Líneas de pago (según estado)
                for pago_move_id, origen_line_ids in pagos_dest.items():
                    for origen_line_id in origen_line_ids:

                        resultado = obtener_linea_destino_por_id_origen(
                            pago_move_id, origen_line_id
                        )

                        # 🔄 Normalizar a lista
                        lineas_dest = resultado if isinstance(resultado, list) else [resultado]

                        for linea_dest in lineas_dest:
                            # 🔑 FILTRO CLAVE
                            linea_estado = models.execute_kw(
                                db, uid, password,
                                "account.move.line", "read",
                                [[linea_dest["id"]]],
                                {"fields": ["reconciled"]}
                            )[0]

                            if linea_estado["reconciled"]:
                                print(
                                    f"⏭️ Línea {linea_dest['id']} "
                                    f"ya conciliada, se omite"
                                )
                                continue

                            line_ids.append(linea_dest["id"])

                # Seguridad: eliminar duplicados
                line_ids = list(set(line_ids))

                if len(line_ids) < 2:
                    raise Exception("Los asientos no son de la misma cuenta")

                models.execute_kw(
                    db, uid, password,
                    "account.move.line", "reconcile",
                    [line_ids],
                    {"context": ctx}
                )

                print("✅ Conciliación exacta (parcial o total) completada")
            except Exception as e:
                msg = str(e)
                if "No se encontró línea destino" in msg:
                    print(f"Revertir manualmente: {factura_name}")
                elif "Los asientos no son de la misma cuenta" in msg:
                    try:
                        print(f"🔧 Corrigiendo cuenta de FACTURA destino: {factura_name}")

                        # 1️⃣ Cuenta correcta desde ORIGEN
                        cuenta_origen = obtener_cuenta_factura_origen(factura_name)
                        cuenta_origen_code = cuenta_origen["account_code"]

                        # 2️⃣ Cuenta actual en DESTINO
                        linea_factura_dest = obtener_linea_factura_destino(factura_dest_id)

                        if linea_factura_dest["account_code"] != cuenta_origen_code:
                            # 3️⃣ Mapear cuenta correcta en DESTINO
                            cuenta_destino_correcta = mapear_cuenta_origen_a_destino_por_code(
                                cuenta_origen_code
                            )

                            # 4️⃣ Reescribir SOLO la línea de la factura
                            models.execute_kw(
                                db, uid, password,
                                "account.move.line", "write",
                                [[linea_factura_dest["line_id"]], {
                                    "account_id": cuenta_destino_correcta
                                }],
                                {"context": {"check_move_validity": False}}
                            )

                            print(
                                f"   ✔ Factura {factura_name}: "
                                f"{linea_factura_dest['account_code']} → {cuenta_origen_code}"
                            )
                        else:
                            log_error(f"Conciliar: {factura_name}")#print("ℹ️ La cuenta de la factura ya coincide con origen")
                            return

                        # 5️⃣ Reintentar conciliación
                        conciliar_factura_con_pagos_destino(
                            factura_dest_id,
                            pagos_dest,
                            factura_name,
                            payment_state
                        )

                    except Exception as e2:
                        print(f"Error corrigiendo factura {factura_name}: {e2}")#log_error(f"Error corrigiendo factura {factura_name}: {e2}")
                elif "allow_none" in msg: print("✅ Conciliación completada")
                else: print(f"Error: {msg}")

        def mapear_cuenta_origen_destino(account_src_id):
            # 1️⃣ Leer cuenta en origen
            acc_src = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.account", "read",
                [[account_src_id]],
                {"fields": ["code", "name"]}
            )[0]

            # 2️⃣ Intentar por code (preferido)
            if acc_src.get("code"):
                domain_code = [("code", "=", acc_src["code"])]
                acc_dest_ids = models.execute_kw(
                    db, uid, password,
                    "account.account", "search",
                    [domain_code],
                    {"limit": 1}
                )
                if acc_dest_ids:
                    return acc_dest_ids[0]

            # 3️⃣ Fallback por name
            domain_name = [("name", "=", acc_src["name"])]
            acc_dest_ids = models.execute_kw(
                db, uid, password,
                "account.account", "search",
                [domain_name],
                {"limit": 1}
            )

            if acc_dest_ids:
                return acc_dest_ids[0]

            # 4️⃣ Error explícito (mejor que fallar silenciosamente)
            raise Exception(
                f"❌ Cuenta destino no encontrada: "
                f"{acc_src['code']} - {acc_src['name']}"
            )

        def migrar_pago_origen_a_destino(move_origen_id):
            """
            Migra un account.move de pago desde origen a destino.
            Devuelve el ID del move creado en destino.
            """

            # 1️⃣ Leer move de origen
            move = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [[move_origen_id]],
                {"fields": ["name", "date", "ref", "journal_id", "line_ids"]}
            )[0]

            move_name = move["name"]

            journal_dest_ids = models.execute_kw(
                db, uid, password,
                "account.journal", "search",
                [[("x_id_interno", "=", move["journal_id"][0])]],
                {"limit": 1}
            )

            # 2️⃣ Crear move en destino (vacío)
            dest_move_id = models.execute_kw(
                db, uid, password,
                "account.move", "create",
                [{
                    "name": move["name"],
                    "date": move["date"],
                    "ref": move["ref"],
                    "journal_id": journal_dest_ids[0],
                    "move_type": "entry",
                }]
            )

            # 3️⃣ Leer líneas del pago en origen
            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "read",
                [move["line_ids"]],
                {
                    "fields": [
                        "id",
                        "name",
                        "account_id",
                        "debit",
                        "credit",
                        "partner_id",
                    ]
                }
            )

            # 4️⃣ Crear líneas en destino
            for l in lines:
                if not l["account_id"]:
                    continue
                account_dest_id = mapear_cuenta_origen_destino(l["account_id"][0])
                vals = {
                    "move_id": dest_move_id,
                    "name": l["name"],
                    "account_id": account_dest_id,
                    "debit": l["debit"],
                    "credit": l["credit"],
                    "x_id_interno": l["id"],
                }

                if l.get("partner_id"):
                    partner_id_origen = l['partner_id'][0]
                    partner_id = Utils.get_by_x_id_interno("res.partner", partner_id_origen, db, uid, password, models)
                    vals["partner_id"] = partner_id

                models.execute_kw(
                    db, uid, password,
                    "account.move.line", "create",
                    [vals],
                    {
                        "context": {
                            "check_move_validity": False,
                            "skip_account_move_synchronization": True,
                        }
                    }
                )

            # 5️⃣ Publicar el pago
            models.execute_kw(
                db, uid, password,
                "account.move", "action_post",
                [[dest_move_id]],
                {"context": {"check_move_validity": False}}
            )

            print(f"✅ Pago migrado: {move_name}")
            return dest_move_id

        def conciliar_factura(id, factura_name, estado):
            try:

                def existe_pago_en_destino(move_name):
                    """
                    Comprueba si un account.move existe en destino por nombre.
                    Devuelve el ID si existe, None si no.
                    """
                    ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", move_name)]],
                        {"limit": 1}
                    )
                    return ids[0] if ids else None


                # 🔎 Obtener factura en destino
                factura_dest_ids = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("x_id_interno", "=", id)]],
                    {"limit": 1}
                )
                if not factura_dest_ids:
                    factura_dest_ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", factura_name)]],
                        {"limit": 1}
                    )

                if not factura_dest_ids:
                    raise Exception(f"❌ Factura {factura_name} no existe en destino")

                factura_dest_id = factura_dest_ids[0]

                pagos_origen = detectar_pagos_factura_origen(factura_name)

                pagos_dest = defaultdict(list)

                for pago in pagos_origen:
                    # 1️⃣ Comprobar si el pago existe en destino
                    pago_dest_id = existe_pago_en_destino(pago["pago_move_name"])

                    # 2️⃣ Si no existe, migrarlo
                    if not pago_dest_id: pago_dest_id = migrar_pago_origen_a_destino(pago["pago_move_id"])

                    # 3️⃣ Asociar la LÍNEA DE ORIGEN al move destino
                    pagos_dest[pago_dest_id].append(pago["pago_line_id"])

                # 🔗 Conciliar factura con pagos en destino
                conciliar_factura_con_pagos_destino(factura_dest_id, pagos_dest, factura_name, estado)

            except Exception as e:
                msg = str(e)
                if "cannot marshal None unless allow_none is enabled" not in msg:
                    print(f"❌ Error al conciliar factura: {factura_name}, Error: {e}")

        def conciliar_todas_las_facturas():

            facturas_a_conciliar_test = [
                "FV1OS/2026/00001",
                "FC1OS/26/0001",
                "FC1OS/26/0003",
                "FC1OS/26/0004",
            ]
            test=False

            print("🚀 Iniciando conciliación masiva de facturas...")

            comp_src = 1 # OSE: 1 / ALM: 2

            # 1️⃣ Buscar facturas cliente y proveedor no pendientes
            factura_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [[
                    ("move_type", "in", [
                        "out_invoice",
                        "in_invoice",
                        "out_refund",
                        "in_refund",
                    ]),
                    ("state", "=", "posted"),
                    ("payment_state", "!=", "not_paid"),
                    ("company_id", "=", comp_src),
                    ("date", ">=", "2026-01-01"), ("date", "<", "2027-01-01"),
                ]]
            )

            if not factura_ids:
                print("ℹ️ No hay facturas para conciliar.")
                return

            # 2️⃣ Leer nombres
            facturas = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [factura_ids],
                {"fields": ["id", "name", "payment_state"]}
            )

            print(f"📄 Facturas a conciliar: {len(facturas)}")

            # 3️⃣ Conciliar una a una
            for f in facturas:
                id = f["id"]
                nombre = f["name"]
                estado = f["payment_state"]

                if test and nombre not in facturas_a_conciliar_test: continue

                print(f"\n🔗 Conciliando factura {nombre} (estado={estado})")

                try:
                    # 🔎 Buscar factura en destino
                    factura_dest_ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("x_id_interno", "=", id)]],
                        {"limit": 1}
                    )

                    if not factura_dest_ids:
                        print(f"⚠️ Factura {nombre} con {id} no existe en destino, se intenta conciliar igualmente")
                        conciliar_factura(id, nombre, estado)
                        continue

                    # 🔗 Solo si NO está conciliada
                    conciliar_factura(id, nombre, estado)

                except Exception as e:
                    print(f"❌ Error conciliando {nombre}: {e}")

            print("✅ Conciliación masiva finalizada")

        #conciliar_todas_las_facturas()#conciliar_factura(39823, "FV1OP/25/00096", "reversed")#

        def limpiar_conciliaciones_clientes_y_proveedores(journal_nominas_id=25, batch_size=500):
            """
            Limpia conciliaciones de Clientes y Proveedores
            EXCLUYENDO nóminas, en batches seguros.
            """

            print("🧹 Limpieza de conciliaciones por batches")
            print(f"📦 Batch size: {batch_size}")

            # 1️⃣ Líneas receivable/payable conciliadas (excluyendo Nóminas)
            line_ids = models.execute_kw(
                db, uid, password,
                "account.move.line", "search",
                [[
                    #("reconciled", "=", True), ("account_id.account_type", "in", ("asset_receivable", "liability_payable")),
                    ("account_id.reconcile", "=", True),
                    ("move_id.journal_id", "!=", journal_nominas_id),
                ]]
            )

            print(f"🔍 Líneas encontradas: {len(line_ids)}")

            if not line_ids:
                print("ℹ️ No hay nada que limpiar")
                return

            # 2️⃣ Buscar partial reconciles asociados
            partial_ids = models.execute_kw(
                db, uid, password,
                "account.partial.reconcile", "search",
                [[
                    "|",
                    ("debit_move_id", "in", line_ids),
                    ("credit_move_id", "in", line_ids),
                ]]
            )

            total = len(partial_ids)
            print(f"🧩 Partial reconciles totales: {total}")

            if not partial_ids:
                print("ℹ️ No hay partial reconciles")
                return

            # 3️⃣ Borrado por batches
            for i in range(0, total, batch_size):
                batch = partial_ids[i:i + batch_size]
                print(
                    f"🧹 Eliminando batch "
                    f"{i + 1}-{min(i + batch_size, total)} / {total}"
                )

                try:
                    models.execute_kw(
                        db, uid, password,
                        "account.partial.reconcile", "unlink",
                        [batch]
                    )
                except Exception as e:
                    print(f"❌ Error en batch {i}-{i + batch_size}: {e}")
                    continue

            print("✅ Limpieza de conciliaciones finalizada")

        #limpiar_conciliaciones_clientes_y_proveedores()

        def eliminar_facturas_destino_not_paid_incorrectas():
            def revertir_y_eliminar_factura_destino(factura_origen_id, nombre):
                """
                NO elimina la factura destino.
                Detecta en ORIGEN los asientos conciliados con la factura
                y elimina esos asientos equivalentes en DESTINO buscándolos por NAME.
                """

                ctx = {"active_test": False}

                print(f"   🔎 Analizando conciliaciones en ORIGEN para {nombre}")

                # ------------------------------------------------------------------
                # 1️⃣ ORIGEN: líneas receivable/payable de la factura
                # ------------------------------------------------------------------
                lineas_origen = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "search",
                    [[
                        ("move_id", "=", factura_origen_id),
                        #("account_id.account_type", "in", ("asset_receivable", "liability_payable")),
                        ("account_id.reconcile", "=", True),
                    ]]
                )

                if not lineas_origen:
                    print("   ℹ️ ORIGEN: la factura no tiene líneas receivable/payable")
                    return []

                # ------------------------------------------------------------------
                # 2️⃣ ORIGEN: conciliaciones parciales
                # ------------------------------------------------------------------
                partials = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.partial.reconcile", "search_read",
                    [[
                        "|",
                        ("debit_move_id", "in", lineas_origen),
                        ("credit_move_id", "in", lineas_origen),
                    ]],
                    {"fields": ["debit_move_id", "credit_move_id"]}
                )

                if not partials:
                    print("   ℹ️ ORIGEN: no hay conciliaciones asociadas")
                    return []

                # ------------------------------------------------------------------
                # 3️⃣ ORIGEN: obtener moves externos + sus names
                # ------------------------------------------------------------------
                moves_origen = {}  # move_id -> name

                for p in partials:
                    for field in ("debit_move_id", "credit_move_id"):
                        line = p.get(field)
                        if not line:
                            continue

                        line_id = line[0]
                        linea = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "account.move.line", "read",
                            [[line_id]],
                            {"fields": ["move_id"]}
                        )[0]

                        move_id = linea["move_id"][0]
                        if move_id != factura_origen_id:
                            move = models_src.execute_kw(
                                db_src, uid_src, password_src,
                                "account.move", "read",
                                [[move_id]],
                                {"fields": ["name"]}
                            )[0]
                            moves_origen[move_id] = move["name"]

                if not moves_origen:
                    print("   ℹ️ ORIGEN: no hay asientos externos")
                    return []

                print(f"   📌 Asientos ORIGEN relacionados: {len(moves_origen)}")

                # ------------------------------------------------------------------
                # 4️⃣ DESTINO: buscar por NAME y eliminar
                # ------------------------------------------------------------------
                eliminados = []

                for move_origen_id, move_name in moves_origen.items():

                    move_dest_ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", move_name)]],
                        {"limit": 1}
                    )

                    if not move_dest_ids:
                        print(f"   ⚠️ Asiento DESTINO no encontrado (name={move_name})")
                        continue

                    move_dest_id = move_dest_ids[0]

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "account.move", "button_draft",
                            [[move_dest_id]],
                            {"context": ctx}
                        )
                    except Exception:
                        pass

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "account.move", "unlink",
                            [[move_dest_id]]
                        )
                        eliminados.append(move_dest_id)
                        print(f"   🗑️ Asiento DESTINO eliminado (name={move_name})")
                    except Exception as e:
                        print(f"   ❌ Error eliminando asiento DESTINO {move_name}: {e}")

                return eliminados

            print("🧹 Eliminando facturas DESTINO not_paid incorrectas")

            origen_ids_eliminados = []

            # 1️⃣ Buscar facturas DESTINO not_paid
            facturas_dest = models.execute_kw(
                db, uid, password,
                "account.move", "search_read",
                [[
                    ("move_type", "in", [
                        "out_invoice", "in_invoice",
                        "out_refund", "in_refund"
                    ]),
                    ("state", "=", "posted"),
                    ("payment_state", "=", "not_paid"),
                ]],
                {"fields": ["id", "name", "x_id_interno"]}
            )

            print(f"📄 Facturas destino not_paid encontradas: {len(facturas_dest)}")

            eliminadas = 0

            for f in facturas_dest:
                factura_dest_id = f["id"]
                nombre = f["name"]
                origen_id = f.get("x_id_interno")

                # 2️⃣ Comprobar estado en ORIGEN
                estado_origen = None
                if origen_id:
                    try:
                        estado_origen = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "account.move", "read",
                            [[origen_id]],
                            {"fields": ["payment_state"]}
                        )[0]["payment_state"]
                    except Exception:
                        estado_origen = None

                # 👉 Si en ORIGEN está not_paid → NO tocar
                if estado_origen == "not_paid":
                    continue
                #else: origen_ids_eliminados.append(origen_id);continue

                print(f"\n🗑️ Eliminando factura {nombre} (origen={estado_origen})")

                if revertir_y_eliminar_factura_destino(origen_id, nombre):
                    eliminadas += 1
                    if origen_id:
                        origen_ids_eliminados.append(origen_id)

            print(f"\n✅ Proceso terminado. Facturas eliminadas: {eliminadas}")
            return origen_ids_eliminados

        def export_invoices_by_ids(origen_ids):
            """
            Exporta facturas de Odoo ORIGEN a partir de una lista de IDs.
            Reutiliza exactamente la misma estructura que export_invoices_by_state.
            """

            if not origen_ids:
                return []

            print(f"📤 Exportando {len(origen_ids)} facturas por ID...")

            FIELDS = [
                "id",
                "name",
                "partner_id",
                "ref",
                "date",
                "invoice_date",
                "invoice_date_due",
                "move_type",
                "state",
                "invoice_line_ids",
                "amount_untaxed",
                "amount_tax",
                "amount_total",
                "currency_id",
                "payment_reference",
                "invoice_payment_term_id",
                "invoice_origin",
                "narration",
                "company_id",
                "invoice_user_id",
                "journal_id",
            ]

            invoices = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [origen_ids],
                {"fields": FIELDS}
            )

            print(f"   → {len(invoices)} facturas leídas.")

            journal_ids = {
                inv["journal_id"][0]
                for inv in invoices
                if inv.get("journal_id")
            }

            journal_code_map = {}

            if journal_ids:
                journals = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.journal", "read",
                    [list(journal_ids)],
                    {"fields": ["id", "code"], "context": {"active_test": False}}
                )

                journal_code_map = {
                    j["id"]: j["code"]
                    for j in journals
                }
            # -------------------------------
            # Añadir journal_code a cada factura
            # -------------------------------
            for inv in invoices:
                if inv.get("journal_id"):
                    inv["journal_code"] = journal_code_map.get(inv["journal_id"][0])
                else:
                    inv["journal_code"] = None

            # 2️⃣ Reunir todos los IDs de líneas
            all_line_ids = []
            for inv in invoices:
                all_line_ids.extend(inv.get("invoice_line_ids", []))

            if not all_line_ids:
                print("⚠️  No se encontraron líneas de factura.")
                return invoices

            print(f"   → {len(all_line_ids)} líneas totales detectadas. Leyendo en bloque...")

            # 3️⃣ Leer todas las líneas en bloque
            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "read",
                [all_line_ids],
                {"fields": [
                    "move_id",
                    "name",
                    "product_id",
                    "quantity",
                    "price_unit",
                    "discount",
                    "price_subtotal",
                    "tax_ids",
                    "account_id",
                    "display_type",
                ]}
            )

            # ------------------------------------------------
            # 4️⃣ Resolver SKUs de productos
            # ------------------------------------------------
            product_ids = {
                l["product_id"][0]
                for l in lines
                if l.get("product_id")
            }

            sku_map = {}
            if product_ids:
                products = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "product.product", "read",
                    [list(product_ids)],
                    {"fields": ["id", "default_code"]}
                )
                sku_map = {p["id"]: p["default_code"] for p in products}

            # 4️⃣ Agrupar las líneas por factura
            grouped_lines = defaultdict(list)
            for line in lines:
                move = line.get("move_id")
                if move:
                    line["default_code"] = (
                        sku_map.get(line["product_id"][0])
                        if line.get("product_id")
                        else None
                    )
                    grouped_lines[move[0]].append(line)

            # 5️⃣ Asignar líneas a cada factura
            for inv in invoices:
                inv_id = inv["id"]
                inv["lineas_detalle"] = grouped_lines.get(inv_id, [])

            print("✅ Líneas asignadas correctamente a cada factura.")
            return invoices

        def conciliar_facturas_eliminadas(origen_ids):

            if not origen_ids:
                print("ℹ️ No hay facturas para reimportar")
                return

            print(f"🔄 Reimportando y conciliando {len(origen_ids)} facturas")

            # 1️⃣ Exportar facturas ORIGEN por ID
            invoices = export_invoices_by_ids(origen_ids)

            if not invoices:
                print("⚠️ No se pudieron exportar las facturas")
                return

            # 3️⃣ Conciliar una a una
            for inv in invoices:
                origen_id = inv.get("id")
                nombre = inv.get("name")
                estado_origen = inv.get("payment_state")

                if estado_origen == "not_paid":
                    continue  # no se concilia

                # 🔎 Buscar factura destino recién creada
                factura_dest_ids = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("x_id_interno", "=", origen_id)]],
                    {"limit": 1}
                )

                if not factura_dest_ids:
                    print(f"⚠️ No se encontró factura destino para {nombre}")
                    continue

                factura_dest_id = factura_dest_ids[0]

                print(f"🔗 Conciliando factura {nombre} (estado={estado_origen})")

                try:
                    conciliar_factura(factura_dest_id, nombre, estado_origen)
                except Exception as e:
                    print(f"❌ Error conciliando {nombre}: {e}")

            print("✅ Reimportación y conciliación finalizadas")

        def corregir_facturas_not_paid_destino():
            """
            Orquesta el flujo completo para corregir facturas DESTINO not_paid incorrectas:
            1️⃣ Elimina facturas destino inconsistentes
            2️⃣ Reimporta desde ORIGEN
            3️⃣ Concilia según estado original
            """

            print("\n🚦 INICIO PROCESO DE CORRECCIÓN DE FACTURAS not_paid EN DESTINO")

            # 1️⃣ Eliminar facturas incorrectas en DESTINO
            print("\n🧹 PASO 1: Eliminando facturas destino incorrectas...")
            origen_ids_eliminados = eliminar_facturas_destino_not_paid_incorrectas()
            origen_ids = origen_ids_eliminados
            if not origen_ids:
                #origen_ids = reconstruir_origen_ids_facturas_eliminadas()
                print("\nℹ️ No se eliminaron facturas. Proceso finalizado.")
                return

            print(f"\n📌 Facturas eliminadas: {len(origen_ids)}")

            # 2️⃣ Reimportar y conciliar
            print("\n🔄 PASO 2 y 3: Obteniendo y conciliando facturas...")
            conciliar_facturas_eliminadas(origen_ids)

            print("\n✅ PROCESO COMPLETO FINALIZADO CON ÉXITO")

        #corregir_facturas_not_paid_destino()

        def detectar_facturas_estado_incoherente_y_exportar_excel():
            """
            Detecta facturas cuyo payment_state (paid / partial)
            no coincide entre ORIGEN y DESTINO y las exporta a Excel.
            """

            ruta_excel = ruta

            print("🔍 Detectando facturas con estado incoherente entre ORIGEN y DESTINO")

            # -------------------------------------------------
            # 1️⃣ Facturas ORIGEN (paid / partial)
            # -------------------------------------------------
            factura_ids_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [[
                    ("move_type", "in", [
                        "out_invoice",
                        "in_invoice",
                        "out_refund",
                        "in_refund",
                    ]),
                    ("state", "=", "posted"),
                    ("payment_state", "in", ["paid", "partial"]),
                ]]
            )

            if not factura_ids_origen:
                print("ℹ️ No se encontraron facturas en ORIGEN")
                return

            facturas_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [factura_ids_origen],
                {"fields": ["id", "name", "payment_state"]}
            )

            print(f"📄 Facturas ORIGEN analizadas: {len(facturas_origen)}")

            resultados = []

            # -------------------------------------------------
            # 2️⃣ Comparar con DESTINO
            # -------------------------------------------------
            for f in facturas_origen:
                origen_id = f["id"]
                name = f["name"]
                estado_origen = f["payment_state"]

                dest_ids = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("name", "=", name)]],
                    {"limit": 1}
                )

                if not dest_ids:
                    estado_destino = "NO_EXISTE"
                else:
                    estado_destino = models.execute_kw(
                        db, uid, password,
                        "account.move", "read",
                        [dest_ids],
                        {"fields": ["payment_state"]}
                    )[0]["payment_state"]

                # -------------------------------------------------
                # 3️⃣ Detectar incoherencia
                # -------------------------------------------------
                if estado_origen != estado_destino and (estado_destino == "partial" or estado_destino == "not_paid" or estado_destino == "paid"):
                    resultados.append({
                        "origen_id": origen_id,
                        "name": name,
                        "payment_state_origen": estado_origen,
                        "payment_state_destino": estado_destino,
                    })

            # -------------------------------------------------
            # 4️⃣ Exportar a Excel
            # -------------------------------------------------
            if not resultados:
                print("✅ No se detectaron incoherencias")
                return

            df = pd.DataFrame(resultados)
            df.to_excel(ruta_excel, index=False)

            print(f"📦 Excel generado: {ruta_excel}")
            print(f"⚠️ Facturas con estado incoherente: {len(df)}")

        def conciliar_facturas_excel_prioritarias():
            """
            Conciliación dirigida por Excel.
            La factura del Excel se concilia PRIMERA,
            el resto de facturas del mismo asiento van después.
            """

            ruta_excel = ruta

            print("🚀 Iniciando conciliación dirigida con prioridad desde Excel")

            # -------------------------------------------------
            # 1️⃣ Leer Excel
            # -------------------------------------------------
            df = pd.read_excel(ruta_excel)

            if df.empty:
                print("ℹ️ El Excel está vacío")
                return

            facturas_excel = df[["origen_id", "name"]].drop_duplicates()
            print(f"📄 Facturas prioritarias: {len(facturas_excel)}")

            bancos_procesados = set()

            # -------------------------------------------------
            # 2️⃣ Procesar cada factura prioritaria
            # -------------------------------------------------
            for _, row in facturas_excel.iterrows():
                factura_origen_id = int(row["origen_id"])
                factura_name = row["name"]

                print(f"\n⭐ Factura PRIORITARIA: {factura_name}")

                # -------------------------------------------------
                # 3️⃣ ORIGEN: detectar asiento de banco
                # -------------------------------------------------
                lineas_factura = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "search",
                    [[
                        ("move_id", "=", factura_origen_id),
                        #("account_id.account_type", "in", ("asset_receivable", "liability_payable")),
                        ("account_id.reconcile", "=", True),
                    ]]
                )

                if not lineas_factura:
                    print("⚠️ Sin líneas conciliables en ORIGEN")
                    continue

                partials = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.partial.reconcile", "search_read",
                    [[
                        "|",
                        ("debit_move_id", "in", lineas_factura),
                        ("credit_move_id", "in", lineas_factura),
                    ]],
                    {"fields": ["debit_move_id", "credit_move_id"]}
                )

                bank_move_id = None

                for p in partials:
                    for field in ("debit_move_id", "credit_move_id"):
                        line = p.get(field)
                        if not line:
                            continue

                        line_id = line[0]
                        linea = models_src.execute_kw(
                            db_src, uid_src, password_src,
                            "account.move.line", "read",
                            [[line_id]],
                            {"fields": ["move_id"]}
                        )[0]

                        if linea["move_id"][0] != factura_origen_id:
                            bank_move_id = linea["move_id"][0]
                            break
                    if bank_move_id:
                        break

                if not bank_move_id:
                    print("⚠️ No se detectó asiento de banco")
                    continue

                if bank_move_id in bancos_procesados:
                    print("⏭️ Asiento banco ya procesado")
                    continue

                bancos_procesados.add(bank_move_id)

                print(f"🏦 Asiento banco detectado (origen_id={bank_move_id})")

                # -------------------------------------------------
                # 4️⃣ ORIGEN: obtener TODAS las facturas implicadas
                # -------------------------------------------------
                facturas_grupo = {}

                bank_lines = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "search",
                    [[("move_id", "=", bank_move_id)]]
                )

                for bl in bank_lines:
                    parts = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.partial.reconcile", "search_read",
                        [[
                            "|",
                            ("debit_move_id", "=", bl),
                            ("credit_move_id", "=", bl),
                        ]],
                        {"fields": ["debit_move_id", "credit_move_id"]}
                    )

                    for p in parts:
                        for field in ("debit_move_id", "credit_move_id"):
                            line = p.get(field)
                            if not line:
                                continue

                            line_id = line[0]
                            linea = models_src.execute_kw(
                                db_src, uid_src, password_src,
                                "account.move.line", "read",
                                [[line_id]],
                                {"fields": ["move_id"]}
                            )[0]

                            move_id = linea["move_id"][0]
                            if move_id != bank_move_id:
                                factura = models_src.execute_kw(
                                    db_src, uid_src, password_src,
                                    "account.move", "read",
                                    [[move_id]],
                                    {"fields": ["id", "name", "payment_state"]}
                                )[0]
                                facturas_grupo[move_id] = factura

                if not facturas_grupo:
                    print("⚠️ No se encontraron facturas asociadas")
                    continue

                # -------------------------------------------------
                # 5️⃣ DESTINO: romper conciliación del grupo
                # -------------------------------------------------
                print("🔨 Rompiendo conciliación del grupo")

                for f in facturas_grupo.values():
                    dest_ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", f["name"])]],
                        {"limit": 1}
                    )

                    if not dest_ids:
                        continue

                    line_ids = models.execute_kw(
                        db, uid, password,
                        "account.move.line", "search",
                        [[
                            ("move_id", "=", dest_ids[0]),
                            #("account_id.account_type", "in", ("asset_receivable", "liability_payable")),
                            ("account_id.reconcile", "=", True),
                        ]]
                    )

                    if not line_ids:
                        continue

                    partial_ids = models.execute_kw(
                        db, uid, password,
                        "account.partial.reconcile", "search",
                        [[
                            "|",
                            ("debit_move_id", "in", line_ids),
                            ("credit_move_id", "in", line_ids),
                        ]]
                    )

                    if partial_ids:
                        models.execute_kw(
                            db, uid, password,
                            "account.partial.reconcile", "unlink",
                            [partial_ids]
                        )

                # -------------------------------------------------
                # 6️⃣ DESTINO: conciliar (PRIORIDAD + RESTO)
                # -------------------------------------------------
                print("🔗 Conciliando grupo")

                # 6.1 PRIORITARIA
                dest_prio = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("name", "=", factura_name)]],
                    {"limit": 1}
                )

                if dest_prio:
                    conciliar_factura(dest_prio[0], factura_name, "paid")

                # 6.2 RESTO
                for f in facturas_grupo.values():
                    if f["name"] == factura_name:
                        continue

                    dest_ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", f["name"])]],
                        {"limit": 1}
                    )

                    if dest_ids:
                        conciliar_factura(dest_ids[0], f["name"], f["payment_state"])

            print("✅ Conciliación dirigida con prioridad finalizada")

        #detectar_facturas_estado_incoherente_y_exportar_excel()
        #conciliar_facturas_excel_prioritarias()
        #conciliar_todas_las_facturas()

        ACCIONES_POR_ESTADO = {
            "paid": 1288,  # Pagado
            "reversed": 1290,  # Revertido
            "partial": 1291,  # Parcial
        }

        def aplicar_acciones_servidor_desde_excel():
            """
            Lee el Excel de facturas incoherentes,
            agrupa por payment_state_origen
            y ejecuta la acción de servidor correspondiente en DESTINO.
            """
            ruta_excel = ruta

            print("🚀 Aplicando acciones de servidor desde Excel")

            df = pd.read_excel(ruta_excel)

            if df.empty:
                print("ℹ️ El Excel está vacío")
                return

            for estado, action_id in ACCIONES_POR_ESTADO.items():

                grupo = df[df["payment_state_origen"] == estado]

                if grupo.empty:
                    continue

                print(f"\n🔧 Estado '{estado}' → acción {action_id}")
                print(f"📄 Facturas a procesar: {len(grupo)}")

                # Buscar facturas DESTINO por name
                dest_ids = []

                for name in grupo["name"].unique():
                    ids = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", name)]],
                        {"limit": 1}
                    )
                    if ids:
                        dest_ids.append(ids[0])

                if not dest_ids:
                    print("⚠️ No se encontraron facturas en DESTINO para este grupo")
                    continue

                print(f"▶️ Ejecutando acción sobre {len(dest_ids)} facturas")

                # Ejecutar acción de servidor
                models.execute_kw(
                    db, uid, password,
                    "ir.actions.server", "run",
                    [[action_id]],
                    {
                        "context": {
                            "active_model": "account.move",
                            "active_ids": dest_ids,
                        }
                    }
                )

            print("\n✅ Acciones de servidor aplicadas correctamente")

        #aplicar_acciones_servidor_desde_excel()
        # endregion

        # region REGIÓN CONTACTOS
        # ---------------------------------------------------------------------------
        # 1) EXPORTAR CONTACTOS DESDE EL ODOO ANTIGUO
        # ---------------------------------------------------------------------------

        def actualizar_costes_desde_excel(ruta_excel=ruta):
            try:
                df = pd.read_excel(ruta_excel)

                if "SKU" not in df.columns or "Coste" not in df.columns:
                    print("❌ El Excel debe contener columnas 'SKU' y 'Coste'")
                    return

                df = df.dropna(subset=["SKU", "Coste"])

                # 1️⃣ Leer TODOS los productos destino una vez
                productos = models.execute_kw(
                    db, uid, password,
                    "product.product", "search_read",
                    [[]],
                    {"fields": ["id", "default_code"]}
                )

                sku_map = {
                    p["default_code"]: p["id"]
                    for p in productos
                    if p["default_code"]
                }

                actualizados = 0
                no_encontrados = 0

                for _, row in df.iterrows():
                    sku = str(row["SKU"]).strip()
                    coste = float(row["Coste"])

                    product_id = sku_map.get(sku)

                    if not product_id:
                        print(f"⚠ SKU no encontrado: {sku}")
                        no_encontrados += 1
                        continue

                    models.execute_kw(
                        db, uid, password,
                        "product.product", "write",
                        [[product_id], {"standard_price": coste}]
                    )

                    actualizados += 1

                print(f"✅ Productos actualizados: {actualizados}")
                print(f"⚠ No encontrados: {no_encontrados}")

            except Exception as e:
                print(f"❌ Error actualizando costes: {e}")

        def update_partner_sales_purchase_fields(partner_name=None):
            """
            Actualiza exclusivamente los campos de la pestaña Venta y Compra
            en Odoo destino, usando x_id_interno como referencia.

            Si partner_name se pasa → solo actualiza ese contacto.
            Si no → actualiza todos.
            """

            try:

                # -------------------------------
                # 1️⃣ EXPORTAR DESDE ORIGEN
                # -------------------------------
                domain = [("active", "=", True)]
                if partner_name:
                    domain.append(("name", "=", partner_name))

                partners_src = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "res.partner", "search_read",
                    [domain],
                    {"fields": [
                        "id",
                        "name",
                        "user_id",
                        "property_account_position_id",
                        "property_payment_term_id",
                        "property_supplier_payment_term_id",
                        "property_product_pricelist",
                        "comment"
                    ]}
                )

                if not partners_src:
                    print("⚠️ No se encontraron partners en origen")
                    return

                # -------------------------------
                # 2️⃣ MAPAS NECESARIOS
                # -------------------------------

                # Usuarios destino por name
                users_dest = models.execute_kw(
                    db, uid, password,
                    "res.users", "search_read",
                    [[]], {"fields": ["id", "name"]}
                )

                user_name_map = {u["name"]: u["id"] for u in users_dest}

                # 🔹 Diccionario manual de listas de precio
                pricelist_id_map = {
                    118: 10,
                    117: 11,
                    116: 12,
                    115: 13,
                    114: 14,
                    113: 15,
                    112: 16,
                    119: 17,
                }

                # 🔹 Payment Terms destino por name
                payment_terms_dest = models.execute_kw(
                    db, uid, password,
                    "account.payment.term", "search_read",
                    [[]], {"fields": ["id", "name"]}
                )
                payment_term_map = {pt["name"]: pt["id"] for pt in payment_terms_dest}

                # -------------------------------
                # 3️⃣ PROCESO DE ACTUALIZACIÓN
                # -------------------------------

                for p in partners_src:

                    # Buscar partner destino por x_id_interno
                    partner_dest = models.execute_kw(
                        db, uid, password,
                        "res.partner", "search",
                        [[("x_id_interno", "=", p["id"])]],
                        {"limit": 1}
                    )

                    if not partner_dest:
                        print(f"⏭ Partner no encontrado en destino: {p['name']}")
                        continue

                    partner_id_dest = partner_dest[0]

                    # -------------------------------
                    # Construir vals
                    # -------------------------------

                    vals = {}

                    # 🔹 user_id por name
                    if p.get("user_id"):

                        user_name = p["user_id"][1]

                        user_id_dest = user_name_map.get(user_name)

                        if not user_id_dest:
                            print(f"⚠️ Usuario destino no encontrado: '{user_name}'")

                        vals["user_id"] = user_id_dest or False

                    # 🔹 property_account_position_id (robusto)
                    if p.get("property_account_position_id"):

                        fiscal_position_id = False

                        try:
                            # ID origen
                            fiscal_position_id_origen = p["property_account_position_id"][0]

                            fiscal_position_id = Utils.get_by_x_id_interno(
                                "account.fiscal.position",
                                fiscal_position_id_origen,
                                db, uid, password, models
                            )

                        except Exception:
                            fiscal_position_id = False

                        # 🔁 Fallback si no se encontró por x_id_interno
                        if not fiscal_position_id:

                            fp_es = p["property_account_position_id"][1]
                            fp_en = FISCAL_POSITION_MAP.get(fp_es)

                            if not fp_en:
                                print(f"⚠️ Posición fiscal sin mapear: '{fp_es}'")
                            else:
                                fps = models.execute_kw(
                                    db, uid, password,
                                    "account.fiscal.position", "search_read",
                                    [[("name", "=", fp_en)]],
                                    {"fields": ["id"], "limit": 1}
                                )

                                if fps:
                                    fiscal_position_id = fps[0]["id"]
                                else:
                                    print(f"⚠️ Posición fiscal destino no encontrada: '{fp_en}'")

                        vals["property_account_position_id"] = fiscal_position_id
                    else:
                        vals["property_account_position_id"] = False

                    # 🔹 property_payment_term_id por name
                    if p.get("property_payment_term_id"):
                        term_name = p["property_payment_term_id"][1]
                        vals["property_payment_term_id"] = payment_term_map.get(term_name, False)

                    # 🔹 property_supplier_payment_term_id por name
                    if p.get("property_supplier_payment_term_id"):
                        term_name = p["property_supplier_payment_term_id"][1]
                        vals["property_supplier_payment_term_id"] = payment_term_map.get(term_name, False)

                    # 🔹 property_product_pricelist por diccionario ID→ID
                    if p.get("property_product_pricelist"):
                        src_pricelist_id = p["property_product_pricelist"][0]
                        vals["property_product_pricelist"] = pricelist_id_map.get(src_pricelist_id, False)

                    # 🔹 notas internas
                    if p.get("comment"):
                        vals["comment"] = p.get("comment")

                    # -------------------------------
                    # WRITE
                    # -------------------------------

                    if vals:
                        models.execute_kw(
                            db, uid, password,
                            "res.partner", "write",
                            [[partner_id_dest], vals]
                        )
                        print(f"✅ Actualizado Venta/Compra: {p['name']}")
                    else:
                        print(f"⏭ Sin cambios: {p['name']}")

                print("🎯 Proceso finalizado correctamente")

            except Exception as e:
                print(f"❌ Error actualizando Venta/Compra: {e}")

        #update_partner_sales_purchase_fields()

        def asignar_padres_categorias_ecommerce(ruta_excel=ruta):

            try:

                df = pd.read_excel(ruta_excel)
                df = df.fillna("")

                # Leer categorías existentes
                categorias_dest = models.execute_kw(
                    db, uid, password,
                    "product.public.category", "search_read",
                    [[]],
                    {"fields": ["id", "name", "parent_id"]}
                )

                def encontrar_categoria(nombre, parent_id):
                    for cat in categorias_dest:
                        if (
                                cat["name"].strip() == nombre and
                                (cat["parent_id"][0] if cat["parent_id"] else False) == parent_id
                        ):
                            return cat["id"]
                    return False

                for _, row in df.iterrows():

                    nombre = row["Nombre"].strip()
                    categoria_padre_path = row["Categoría padre"].strip()

                    if not nombre or not categoria_padre_path:
                        continue

                    # 🔹 Resolver padre por jerarquía
                    partes = [p.strip() for p in categoria_padre_path.split("/")]

                    parent_id = False

                    for parte in partes:
                        parent_id = encontrar_categoria(parte, parent_id)
                        if not parent_id:
                            print(f"⚠️ No se pudo resolver parte '{parte}' en path '{categoria_padre_path}'")
                            break

                    if not parent_id:
                        continue

                    # 🔹 Buscar hijo (SIEMPRE como raíz porque ahora están sin padre)
                    child_id = encontrar_categoria(nombre, False)

                    if not child_id:
                        print(f"⚠️ Categoría hija no encontrada: {nombre}")
                        continue

                    # Evitar autoreferencia
                    if child_id == parent_id:
                        continue

                    models.execute_kw(
                        db, uid, password,
                        "product.public.category", "write",
                        [[child_id], {"parent_id": parent_id}]
                    )

                    print(f"✅ Asignado padre: {categoria_padre_path} → {nombre}")

                print("🎯 Jerarquía asignada correctamente.")

            except Exception as e:
                print(f"❌ Error asignando jerarquía: {e}")

        #asignar_padres_categorias_ecommerce()

        def corregir_cuenta_payable_proveedores(excel_path, cuenta_proveedor_id=810, cuenta_acreedor_id=823, dry_run=True):
            """
            Corrige property_account_payable_id leyendo proveedores desde un Excel.
            Excel:
              - Hoja: Sheet1
              - Columna: Nombre
            """

            resultado = {
                "procesados": 0,
                "actualizados": [],
                "no_encontrados": [],
                "sin_facturas": [],
                "sin_linea_payable": [],
                "ambiguos": [],
                "errores": [],
            }

            # 1️⃣ Leer Excel
            df = pd.read_excel(excel_path, sheet_name="Sheet1")

            if "Nombre" not in df.columns:
                raise ValueError("❌ El Excel no contiene la columna 'Nombre'")

            nombres = (
                df["Nombre"]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )

            print(f"📄 Contactos leídos desde Excel: {len(nombres)}")

            for nombre in nombres:
                try:
                    # 2️⃣ Buscar contacto por nombre exacto
                    partners = models.execute_kw(
                        db, uid, password,
                        "res.partner", "search_read",
                        [[
                            ("name", "=", nombre),
                        ]],
                        {"fields": ["id", "name", "property_account_payable_id"], "limit": 1}
                    )

                    if not partners:
                        resultado["no_encontrados"].append(nombre)
                        continue

                    partner = partners[0]
                    partner_id = partner["id"]

                    resultado["procesados"] += 1

                    # 3️⃣ Buscar UNA factura proveedor posteada
                    facturas = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[
                            ("partner_id", "=", partner_id),
                            ("move_type", "=", "in_invoice"),
                            ("state", "=", "posted")
                        ]],
                        {"limit": 1}
                    )

                    if not facturas:
                        resultado["sin_facturas"].append(nombre)
                        continue

                    factura_id = facturas[0]

                    # 4️⃣ Obtener línea payable
                    linea_payable = obtener_lineas_conciliables_destino(
                        move_id=factura_id
                    )[0]

                    if not linea_payable:
                        resultado["sin_linea_payable"].append(nombre)
                        continue

                    account_id = linea_payable["account_id"][0]

                    # 5️⃣ Clasificación
                    if account_id == cuenta_proveedor_id:
                        cuenta_correcta = cuenta_proveedor_id
                        tipo = "proveedor"
                    elif account_id == cuenta_acreedor_id:
                        cuenta_correcta = cuenta_acreedor_id
                        tipo = "acreedor"
                    else:
                        resultado["ambiguos"].append({
                            "nombre": nombre,
                            "account_id": account_id
                        })
                        continue

                    cuenta_actual = (
                        partner["property_account_payable_id"][0]
                        if partner["property_account_payable_id"]
                        else None
                    )

                    # 6️⃣ Actualizar contacto si procede
                    if cuenta_actual != cuenta_correcta:
                        if not dry_run:
                            models.execute_kw(
                                db, uid, password,
                                "res.partner", "write",
                                [[partner_id], {
                                    "property_account_payable_id": cuenta_correcta
                                }]
                            )

                        resultado["actualizados"].append({
                            "nombre": nombre,
                            "tipo": tipo,
                            "antes": cuenta_actual,
                            "despues": cuenta_correcta
                        })

                except Exception as e:
                    resultado["errores"].append({
                        "nombre": nombre,
                        "error": str(e)
                    })

            # 7️⃣ Resumen
            print("\n📊 RESUMEN CORRECCIÓN PAYABLE (DESDE EXCEL)")
            print(f"✔ Procesados: {resultado['procesados']}")
            print(f"✔ Actualizados: {len(resultado['actualizados'])}")
            print(f"❌ No encontrados: {len(resultado['no_encontrados'])}")
            print(f"⚠ Sin facturas: {len(resultado['sin_facturas'])}")
            print(f"🚫 Sin línea payable: {len(resultado['sin_linea_payable'])}")
            print(f"❓ Ambiguos: {len(resultado['ambiguos'])}")
            print(f"💥 Errores: {len(resultado['errores'])}")

            return resultado

        #migrar_proyectos()
        #migrar_tareas()
        #migrar_pedidos_venta()
        #migrar_pedidos_compra()
        #sync_task_sale_line_id()
        #migrar_asistencias()
        #migrar_partes_horas() #Proyecto: Interno ; Tarea: Ausencias
        #lines = export_apuntes_analiticos()
        #import_apuntes_analiticos(lines)
        #migrar_facturas(True, False); migrar_facturas(True, True); migrar_facturas(False, False); migrar_facturas(False, True)
        #conciliar_todas_las_facturas() #conciliar manualmente: FC1OP/23/1389 y EFEC/2023/12/0035
        #migrar_pagos("clientes")
        #migrar_pagos("proveedores")
        #test_import_comercial_stock()#corregir_cuenta_payable_proveedores(ruta)

        def actualizar_comprador_desde_excel(
                # Corregir Users_ids: Compras, Ventas, Facturacion
                ruta_excel,
                hoja="Sheet1",
                columna_name="name",
                user_id=20
        ):
            # =========================
            # LISTAR USUARIOS EXISTENTES
            # =========================
            usuarios = models.execute_kw(
                db, uid, password,
                "res.users", "search_read",
                [[]],
                {"fields": ["id", "name"], "order": "name"}
            )

            print("📋 Usuarios existentes en Odoo:")
            for u in usuarios:
                print(f" - {u['name']} con id: {u['id']}")

            print("────────────────────────────")

            # =========================
            # LEER EXCEL
            # =========================
            df = pd.read_excel(ruta_excel, sheet_name=hoja)

            nombres_pedidos = (
                df[columna_name]
                .dropna()
                .astype(str)
                .str.strip()
                .unique()
                .tolist()
            )

            if not nombres_pedidos:
                print("⚠️ No se encontraron nombres de pedidos en el Excel")
                return

            actualizados = 0
            no_encontrados = []

            # =========================
            # ACTUALIZAR PEDIDOS
            # =========================
            for nombre in nombres_pedidos:
                po_ids = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("move_type", "=", "out_refund"),("name", "=", nombre)]]
                )

                if not po_ids:
                    no_encontrados.append(nombre)
                    continue

                models.execute_kw(
                    db, uid, password,
                    "account.move", "write",
                    [po_ids, {"invoice_user_id": user_id}]
                )

                actualizados += len(po_ids)

            # =========================
            # RESULTADO FINAL
            # =========================
            print(f"✅ Pedidos actualizados: {actualizados}")

            if no_encontrados:
                print("⚠️ Pedidos no encontrados:")
                for n in no_encontrados:
                    print(f" - {n}")

        # region NOMINAS
        def obtener_nominas_origen(journal_nominas_id=85):
            """
            Devuelve las nóminas (account.move) en ORIGEN
            correspondientes al diario Nóminas.
            """

            nominas = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search_read",
                [[
                    ("journal_id", "=", journal_nominas_id),
                    ("state", "=", "posted"),
                    #("company_id", "=", 2),
                    ("date", ">=", "2026-01-01"),
                    ("date", "<",  "2027-01-01"),
                ]],
                {
                    "fields": [
                        "id",
                        "name",
                        "ref",
                        "date",
                    ]
                }
            )

            return nominas

        def obtener_linea_payable_nomina_origen(nomina_move_id):
            """
            Devuelve la línea PAYABLE principal de una nómina en ORIGEN.
            Es la línea equivalente a la 'línea de factura'.
            """

            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[("move_id", "=", nomina_move_id)]],
                {
                    "fields": [
                        "id",
                        "account_id",
                        "matched_debit_ids",
                        "matched_credit_ids",
                    ]
                }
            )

            for l in lines:
                if not l.get("account_id"):
                    continue

                acc = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.account", "read",
                    [[l["account_id"][0]]],
                    {"fields": ["account_type"]}
                )[0]

                if acc["account_type"] == "liability_payable":
                    return l

            return None

        def detectar_pagos_nomina_origen(nomina_move_id):
            """
            Devuelve las líneas de pago conciliadas con una nómina en ORIGEN.
            Cada elemento identifica exactamente QUÉ línea de pago hay que mapear.
            """

            resultados = []

            # 1️⃣ Línea payable principal de la nómina
            linea_nomina = obtener_linea_payable_nomina_origen(nomina_move_id)
            if not linea_nomina:
                return resultados

            matched_ids = set(
                linea_nomina["matched_debit_ids"] + linea_nomina["matched_credit_ids"]
            )

            if not matched_ids:
                return resultados

            # 2️⃣ Leer partial reconciles
            partials = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.partial.reconcile", "read",
                [list(matched_ids)],
                {
                    "fields": [
                        "debit_move_id",
                        "credit_move_id",
                        "amount",
                    ]
                }
            )

            for pr in partials:
                # 3️⃣ Determinar cuál es la línea de pago
                if pr["debit_move_id"][0] == linea_nomina["id"]:
                    pago_line_id = pr["credit_move_id"][0]
                elif pr["credit_move_id"][0] == linea_nomina["id"]:
                    pago_line_id = pr["debit_move_id"][0]
                else:
                    continue

                # 4️⃣ Obtener el asiento del pago
                pago_line = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "read",
                    [[pago_line_id]],
                    {"fields": ["move_id"]}
                )[0]

                resultados.append({
                    "pago_move_id": pago_line["move_id"][0],
                    "pago_line_id": pago_line_id,
                    "amount": pr["amount"],
                })

            return resultados

        def migrar_nomina_origen_a_destino(nomina_origen_id):
            """
            Migra un asiento de nómina (account.move) desde ORIGEN a DESTINO.
            Usa x_id_interno en account.move.line para trazabilidad perfecta.
            """

            # 1️⃣ Leer nómina ORIGEN
            move = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [[nomina_origen_id]],
                {
                    "fields": [
                        "name",
                        "ref",
                        "date",
                        "journal_id",
                        "line_ids",
                    ]
                }
            )[0]

            # 2️⃣ Mapear diario Nóminas en DESTINO (por x_id_interno)
            journal_dest_ids = models.execute_kw(
                db, uid, password,
                "account.journal", "search",
                [[("x_id_interno", "=", move["journal_id"][0])]],
                {"limit": 1}
            )

            if not journal_dest_ids:
                raise Exception(
                    f"Diario Nóminas no encontrado en destino "
                    f"(x_id_interno={move['journal_id'][0]})"
                )

            journal_dest_id = journal_dest_ids[0]

            # 3️⃣ Crear account.move DESTINO
            nomina_dest_id = models.execute_kw(
                db, uid, password,
                "account.move", "create",
                [{
                    "name": move["name"],
                    "ref": move["ref"],
                    "date": move["date"],
                    "journal_id": journal_dest_id,
                    "move_type": "entry",
                }]
            )

            # 4️⃣ Leer líneas ORIGEN
            lines = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "read",
                [move["line_ids"]],
                {
                    "fields": [
                        "id",
                        "name",
                        "account_id",
                        "debit",
                        "credit",
                        "partner_id",
                    ]
                }
            )

            # 5️⃣ Crear líneas DESTINO (1 a 1)
            for l in lines:
                if not l.get("account_id"):
                    continue

                account_dest_id = mapear_cuenta_origen_destino(
                    l["account_id"][0]
                )

                vals = {
                    "move_id": nomina_dest_id,
                    "name": l["name"],
                    "account_id": account_dest_id,
                    "debit": l["debit"],
                    "credit": l["credit"],
                    "x_id_interno": l["id"],
                }

                if l.get("partner_id"):
                    vals["partner_id"] = Utils.get_by_x_id_interno(
                        "res.partner",
                        l["partner_id"][0],
                        db,
                        uid,
                        password,
                        models,
                    )

                models.execute_kw(
                    db,
                    uid,
                    password,
                    "account.move.line",
                    "create",
                    [vals],
                    {"context": {"check_move_validity": False}},
                )

            # 6️⃣ Publicar nómina en DESTINO
            models.execute_kw(
                db,
                uid,
                password,
                "account.move",
                "action_post",
                [[nomina_dest_id]],
                {"context": {"check_move_validity": False}},
            )

            return nomina_dest_id

        def migrar_y_conciliar_una_nomina(nomina_origen):
            """
            Orquesta la migración y conciliación de UNA nómina.
            Reutiliza el motor de conciliación de facturas.
            """

            nomina_name = nomina_origen["name"]
            nomina_origen_id = nomina_origen["id"]

            print(f"\n🧾 Procesando nómina {nomina_name}")

            # 1️⃣ Buscar nómina en DESTINO (por name)
            nomina_dest_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[("name", "=", nomina_name)]],
                {"limit": 1}
            )

            if nomina_dest_ids:
                nomina_dest_id = nomina_dest_ids[0]
                print(f"ℹ️ Nómina ya existe en destino (ID {nomina_dest_id})")
            else:
                # 2️⃣ Migrar nómina
                nomina_dest_id = migrar_nomina_origen_a_destino(nomina_origen_id)
                print(f"✅ Nómina migrada a destino (ID {nomina_dest_id})")

            # 3️⃣ Detectar pagos conciliados en ORIGEN
            pagos_origen = detectar_pagos_nomina_origen(nomina_origen_id)

            if not pagos_origen:
                print("ℹ️ Nómina sin pagos conciliados en origen")
                return

            # 4️⃣ Preparar pagos DESTINO (mapping exacto por x_id_interno)
            pagos_dest = defaultdict(list)

            for p in pagos_origen:
                pago_origen_move_id = p["pago_move_id"]
                pago_origen_line_id = p["pago_line_id"]

                # 4.1 Buscar pago en DESTINO (por name)
                pago_move = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move", "read",
                    [[pago_origen_move_id]],
                    {"fields": ["name"]}
                )[0]

                pago_name = pago_move["name"]

                pago_dest_ids = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[("name", "=", pago_name)]],
                    {"limit": 1}
                )

                if pago_dest_ids:
                    pago_dest_id = pago_dest_ids[0]
                else: pago_dest_id = migrar_pago_origen_a_destino(pago_origen_move_id)

                # 4.3 Guardar línea ORIGEN exacta (se resolverá por x_id_interno)
                pagos_dest[pago_dest_id].append(pago_origen_line_id)

            # 5️⃣ Conciliar usando el MISMO motor que facturas
            conciliar_factura_con_pagos_destino(
                nomina_dest_id,
                pagos_dest,
                nomina_name,
                payment_state="paid"
            )

            print(f"🔗 Nómina conciliada: {nomina_name}")

        # nomina = { "id": 251556, "name": "NOMIN/2025/12/0012", }; migrar_y_conciliar_una_nomina(nomina)#Un caso
        #nominas = obtener_nominas_origen()#Todos
        #for n in nominas: migrar_y_conciliar_una_nomina(n)

        # endregion

        #region PICKINGS INVENTARIO
        def export_pickings_confirmed_assigned():
            """
            Exporta pickings en estado confirmed y assigned
            junto con sus stock.moves.
            """

            print("📦 Exportando pickings (confirmed / assigned)")

            # -------------------------------------------------
            # 1️⃣ Buscar pickings válidos
            # -------------------------------------------------
            picking_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.picking", "search",
                [[("state", "in", ["confirmed", "assigned", "waiting"]), ('company_id', '=', 2)]]
            )

            if not picking_ids:
                print("ℹ️ No hay pickings para exportar")
                return []

            pickings = models_src.execute_kw(
                db_src, uid_src, password_src,
                "stock.picking", "read",
                [picking_ids],
                {
                    "fields": [
                        "id",
                        "name",
                        "state",
                        "partner_id",
                        "scheduled_date",
                        "origin",
                        "company_id",
                        "picking_type_id",
                        "location_id",
                        "location_dest_id",
                    ]
                }
            )

            print(f"📄 Pickings encontrados: {len(pickings)}")

            resultado = []

            # -------------------------------------------------
            # 2️⃣ Para cada picking, obtener sus moves
            # -------------------------------------------------
            for p in pickings:
                picking_id = p["id"]

                move_ids = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "stock.move", "search",
                    [[("picking_id", "=", picking_id)]]
                )

                moves = []
                if move_ids:
                    moves = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "stock.move", "read",
                        [move_ids],
                        {
                            "fields": [
                                "id",
                                "product_id",
                                "product_uom_qty",
                                "quantity_done",
                                "description_picking",
                                "date",
                                "date_deadline",
                                "location_id",
                                "location_dest_id",
                            ]
                        }
                    )

                resultado.append({
                    "picking": p,
                    "moves": moves,
                })

            print("✅ Exportación finalizada")
            return resultado

        #pickings = export_pickings_confirmed_assigned()

        def import_pickings_confirmed_assigned(data):
            def map_picking_type(picking_type_src):
                # -------------------------------------------------
                # 🧠 Obtener CODE desde ORIGEN
                # -------------------------------------------------
                picking_type_id_src = picking_type_src[0]

                picking_type = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "stock.picking.type", "read",
                    [[picking_type_id_src]],
                    {"fields": ["code"]}
                )[0]

                code = picking_type["code"]

                # -------------------------------------------------
                # 🔗 Mapear por CODE en DESTINO
                # -------------------------------------------------
                ids = models.execute_kw(
                    db, uid, password,
                    "stock.picking.type", "search",
                    [[("code", "=", code)]],
                    {"limit": 1}
                )

                if not ids:
                    raise Exception(f"❌ No se encontró stock.picking.type en destino con code='{code}'")

                return ids[0]

            def map_product(product_src):
                product_id_src = product_src[0]

                prod = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "product.product", "read",
                    [[product_id_src]],
                    {"fields": ["default_code"]}
                )[0]

                default_code = prod["default_code"]

                ids = models.execute_kw(
                    db, uid, password,
                    "product.product", "search",
                    [[("default_code", "=", default_code)]],
                    {"limit": 1}
                )

                if not ids:
                    raise Exception(f"Producto no encontrado en destino: {default_code}")

                return ids[0]

            def map_location(location_src):
                loc_id_src = location_src[0]

                loc = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "stock.location", "read",
                    [[loc_id_src]],
                    {"fields": ["complete_name"]}
                )[0]

                name = loc["complete_name"]

                ids = models.execute_kw(
                    db, uid, password,
                    "stock.location", "search",
                    [[("complete_name", "=", name)]],
                    {"limit": 1}
                )

                if not ids:
                    raise Exception(f"Ubicación no encontrada: {name}")

                return ids[0]

            print("📥 Importando pickings (mapping mínimo)")

            for item in data:
                p = item["picking"]
                moves = item["moves"]

                print(f"\n🚚 {p['name']}")

                existing = models.execute_kw(
                    db, uid, password,
                    "stock.picking", "search",
                    [[("name", "=", p["name"])]],
                    {"limit": 1}
                )

                if existing:
                    print(f"⚠️ Picking ya existe: {p['name']}")
                    continue

                partner_id = Utils.get_by_x_id_interno("res.partner", p['partner_id'][0], db, uid, password, models)
                company_id = 1  # o la que uses en destino
                picking_type_id = map_picking_type(p["picking_type_id"])
                location_id = map_location(p["location_id"])
                location_dest_id = map_location(p["location_dest_id"])

                picking_dest_id = models.execute_kw(
                    db, uid, password,
                    "stock.picking", "create",
                    [{
                        "name": p['name'],
                        "partner_id": partner_id,
                        "scheduled_date": p["scheduled_date"],
                        "origin": p["origin"],
                        "company_id": company_id,
                        "picking_type_id": picking_type_id,
                        "location_id": location_id,
                        "location_dest_id": location_dest_id,
                    }]
                )

                for m in moves:
                    product_id = map_product(m["product_id"])
                    loc_id = map_location(m["location_id"])
                    loc_dest_id = map_location(m["location_dest_id"])
                    move_name = (
                            m.get("description_picking")
                            or f"Movimiento de {m['product_id'][1]}"
                    )

                    move_id = models.execute_kw(
                        db, uid, password,
                        "stock.move", "create",
                        [{
                            "name": move_name,
                            "picking_id": picking_dest_id,
                            "product_id": product_id,
                            "product_uom_qty": m["product_uom_qty"],
                            "description_picking": move_name,
                            "date": m["date"],
                            "date_deadline": m["date_deadline"],
                            "location_id": loc_id,
                            "location_dest_id": loc_dest_id,
                        }]
                    )

                models.execute_kw(
                    db, uid, password,
                    "stock.picking", "action_confirm",
                    [[picking_dest_id]]
                )

                if p["state"] == "assigned":
                    models.execute_kw(
                        db, uid, password,
                        "stock.picking", "action_assign",
                        [[picking_dest_id]]
                    )

                print("✅ Importado")

            print("\n🎉 Importación finalizada")

        #import_pickings_confirmed_assigned(pickings)

        #endregion

        def analizar_asientos_faltantes():
            """
            Analiza los asientos que faltan en DESTINO
            clasificándolos por journal y tipo.
            """

            # Buscar todos en origen
            origen_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [[("state", "=", "posted"), ("company_id", "=", 2),
                    ("date", ">=", "2026-01-01"),
                    ("date", "<",  "2027-01-01"),]]
            )

            origen_moves = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [origen_ids],
                {"fields": ["name", "journal_id", "date", "move_type"]}
            )

            destino_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [[("state", "=", "posted")]]
            )

            destino_moves = models.execute_kw(
                db, uid, password,
                "account.move", "read",
                [destino_ids],
                {"fields": ["name"]}
            )

            destino_names = {m["name"] for m in destino_moves}

            faltantes = [m for m in origen_moves if m["name"] not in destino_names]

            print(f"\n🔴 Total faltantes: {len(faltantes)}")

            # Clasificar por journal
            resumen = {}

            for m in faltantes:
                journal = m["journal_id"][1] if m["journal_id"] else "SIN JOURNAL"

                if journal not in resumen:
                    resumen[journal] = 0

                resumen[journal] += 1

            print("\n📊 Faltantes por journal:")
            for journal, cantidad in sorted(resumen.items(), key=lambda x: x[1], reverse=True):
                print(f"   {journal}: {cantidad}")

        def export_asientos_banco_faltantes():
            """
            Exporta asientos de banco/caja que existen en ORIGEN
            pero no en DESTINO.
            """

            print("🏦 Exportando asientos de banco/caja faltantes")

            # -------------------------------------------------
            # 1️⃣ Domain base
            # -------------------------------------------------
            domain_src = [
                ("state", "=", "posted"),#draft#posted
                ("company_id", "=", 2),
                ("date", ">=", "2026-01-01"),
                ("date", "<",  "2027-01-01"),
                #"|",
                #("journal_id.type", "in", ["bank", "cash"]),
                #("journal_id.code", "=", "VAR"),
            ]

            domain_dest = [
                ("state", "=", "posted"),
                ("date", ">=", "2026-01-01"),
                ("date", "<",  "2027-01-01"),
                #"|",
                #("journal_id.type", "in", ["bank", "cash"]),
                #("journal_id.code", "=", "VAR"),
            ]

            # -------------------------------------------------
            # 2️⃣ Obtener ORIGEN
            # -------------------------------------------------
            origen_ids = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "search",
                [domain_src]
            )

            if not origen_ids:
                print("ℹ️ No hay asientos de banco en ORIGEN")
                return []

            origen_moves = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move", "read",
                [origen_ids],
                {
                    "fields": [
                        "id",
                        "name",
                        "date",
                        "journal_id",
                        "ref",
                        "company_id",
                        "partner_id",
                        "state",
                    ]
                }
            )

            print(f"📄 Asientos banco en ORIGEN: {len(origen_moves)}")

            # -------------------------------------------------
            # 3️⃣ Obtener DESTINO para comparar
            # -------------------------------------------------
            destino_ids = models.execute_kw(
                db, uid, password,
                "account.move", "search",
                [domain_dest]
            )

            destino_moves = models.execute_kw(
                db, uid, password,
                "account.move", "read",
                [destino_ids],
                {"fields": ["name"]}
            )

            destino_names = {m["name"] for m in destino_moves}

            # -------------------------------------------------
            # 4️⃣ Filtrar faltantes
            # -------------------------------------------------
            faltantes = [m for m in origen_moves if m["name"] not in destino_names]

            faltantes_test = [faltantes[0]]

            print(f"❗ Asientos banco faltantes en DESTINO: {len(faltantes)}")

            resultado = []

            # -------------------------------------------------
            # 5️⃣ Exportar cada asiento con sus líneas
            # -------------------------------------------------
            for move in faltantes:#faltantes_test:#
                move_id = move["id"]

                line_ids = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "search",
                    [[("move_id", "=", move_id)]]
                )

                lines = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "read",
                    [line_ids],
                    {
                        "fields": [
                            "id",
                            "account_id",
                            "partner_id",
                            "name",
                            "debit",
                            "credit",
                            "currency_id",
                            "amount_currency",
                        ]
                    }
                )

                resultado.append({
                    "move": move,
                    "lines": lines,
                })

            print("✅ Exportación completada")

            return resultado

        def import_asientos_banco_faltantes(data):
            """
            Importa asientos de banco/caja faltantes en DESTINO.
            No realiza conciliación.
            Crea el asiento con todas las líneas juntas para evitar
            errores de desbalanceo.
            """

            print("📥 Importando asientos banco faltantes")

            for item in data:
                move_src = item["move"]
                lines_src = item["lines"]

                test = False # En la segunda ejecucion: True y comentar la linea de abajo
                if move_src['name'] == "CB472/22/1140" or move_src['name'] == "VAR/25/0120" or move_src['name'] == "VAR/24/0127": continue

                print(f"\n🏦 Importando asiento {move_src['name']}")

                # -------------------------------------------------
                # 1️⃣ Mapear JOURNAL por código
                # -------------------------------------------------
                journal_src_id = move_src["journal_id"][0]

                journal_src = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.journal", "read",
                    [[journal_src_id]],
                    {"fields": ["code"]}
                )[0]

                journal_code = journal_src["code"]

                journal_dest_ids = models.execute_kw(
                    db, uid, password,
                    "account.journal", "search",
                    [[("code", "=", journal_code)]],
                    {"limit": 1}
                )

                if not journal_dest_ids:
                    continue#raise Exception(f"❌ Journal no encontrado en destino: {journal_code}")

                journal_dest_id = journal_dest_ids[0]

                # -------------------------------------------------
                # 2️⃣ Preparar TODAS las líneas antes de crear el move
                # -------------------------------------------------
                line_vals = []
                total_debit = 0.0
                total_credit = 0.0

                for line in lines_src:

                    # Mapear cuenta por código
                    if not line["account_id"]: continue
                    account_src_id = line["account_id"][0]

                    account_src = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.account", "read",
                        [[account_src_id]],
                        {"fields": ["code"]}
                    )[0]

                    account_code = account_src["code"]

                    account_dest_ids = models.execute_kw(
                        db, uid, password,
                        "account.account", "search",
                        [[("code", "=", account_code)]],
                        {"limit": 1}
                    )

                    if not account_dest_ids:
                        raise Exception(f"❌ Cuenta no encontrada: {account_code}")

                    account_dest_id = account_dest_ids[0]

                    # Mapear partner si existe
                    partner_dest_id = False
                    if line["partner_id"]:
                        partner_dest_id = Utils.get_by_x_id_interno("res.partner",line["partner_id"][0],db,uid,password,models)

                    debit = line["debit"] or 0.0
                    credit = line["credit"] or 0.0

                    total_debit += debit
                    total_credit += credit

                    line_vals.append((0, 0, {
                        "x_id_interno": line["id"],
                        "account_id": account_dest_id,
                        "partner_id": partner_dest_id,
                        "name": line["name"] or "/",
                        "debit": debit if not test else 0,
                        "credit": credit if not test else 0,
                        "currency_id": line["currency_id"][0] if line["currency_id"] else False,
                        "amount_currency": line["amount_currency"],
                    }))

                # -------------------------------------------------
                # 4️⃣ Crear MOVE con todas las líneas juntas
                # -------------------------------------------------
                # Mapear partner si existe
                partner_dest_id = False
                if move_src["partner_id"]:
                    partner_dest_id = Utils.get_by_x_id_interno("res.partner", move_src["partner_id"][0], db, uid, password,
                                                                models)

                move_dest_id = models.execute_kw(
                    db, uid, password,
                    "account.move", "create",
                    [{
                        "name": move_src["name"],
                        "partner_id": partner_dest_id,
                        "date": move_src["date"],
                        "journal_id": journal_dest_id,
                        "ref": move_src["ref"],
                        "line_ids": line_vals,
                    }]
                )

                # -------------------------------------------------
                # 5️⃣ Postear asiento
                # -------------------------------------------------
                if not test:
                    models.execute_kw(
                        db, uid, password,
                        "account.move", "action_post",
                        [[move_dest_id]]
                    )

                print("✅ Importado y posteado")

            print("\n🎉 Todos los asientos banco importados")

        #asientos_banco_faltantes = export_asientos_banco_faltantes()
        #import_asientos_banco_faltantes(asientos_banco_faltantes)
        #analizar_asientos_faltantes()

        def actualizar_x_id_interno_lineas_facturas():

            print("🔄 Actualizando x_id_interno por account_code...")

            # -------------------------------------------------
            # 1️⃣ Obtener líneas conciliables de facturas en DESTINO
            # -------------------------------------------------
            lineas_dest = models.execute_kw(
                db, uid, password,
                "account.move.line", "search_read",
                [[
                    ("company_id", "=", 1),
                    ("x_id_interno", "=", 0),
                    ("account_id.reconcile", "=", True),
                    ("move_id.move_type", "in", [
                        "out_invoice",
                        "in_invoice",
                        "out_refund",
                        "in_refund",
                    ]),
                    ("date", ">=", "2026-01-01"),
                    ("date", "<",  "2027-01-01"),
                ]],
                {
                    "fields": [
                        "id",
                        "move_id",
                        "account_id",
                    ],
                    "order": "move_id,id"
                }
            )

            print(f"📊 Líneas conciliables sin x_id_interno: {len(lineas_dest)}")

            # Agrupar destino por move
            grupos_dest = defaultdict(list)

            for linea in lineas_dest:
                if linea.get("move_id"):
                    grupos_dest[linea["move_id"][0]].append(linea)

            actualizados = 0

            # -------------------------------------------------
            # 2️⃣ Procesar factura por factura
            # -------------------------------------------------
            for move_dest_id, lineas_dest_move in grupos_dest.items():

                # Obtener nombre REAL del move destino
                move_dest = models.execute_kw(
                    db, uid, password,
                    "account.move", "read",
                    [[move_dest_id]],
                    {"fields": ["name"]}
                )[0]

                move_name = move_dest["name"]

                if not move_name:
                    continue

                # Buscar move equivalente en ORIGEN
                move_origen = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move", "search",
                    [[
                        ("name", "=", move_name),
                        ("company_id", "=", 1),
                    ]],
                    {"limit": 1}
                )

                if not move_origen:
                    continue

                move_origen_id = move_origen[0]

                # Obtener líneas conciliables en ORIGEN
                lineas_origen = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move.line", "search_read",
                    [[
                        ("move_id", "=", move_origen_id),
                        ("account_id.reconcile", "=", True),
                    ]],
                    {
                        "fields": ["id", "account_id"],
                        "order": "id"
                    }
                )

                if not lineas_origen:
                    continue

                # -------------------------------------------------
                # Agrupar ORIGEN por account_code
                # -------------------------------------------------
                origen_por_code = defaultdict(list)

                for linea in lineas_origen:

                    acc = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.account", "read",
                        [[linea["account_id"][0]]],
                        {"fields": ["code"]}
                    )[0]

                    origen_por_code[acc["code"]].append(linea["id"])

                # -------------------------------------------------
                # Preparar estructuras de asignación
                # -------------------------------------------------

                # Copia completa de todas las líneas origen disponibles
                origen_restantes = []

                for code, ids in origen_por_code.items():
                    origen_restantes.extend(ids)

                asignadas_dest = set()

                # -------------------------------------------------
                # 1️⃣ Primera pasada → asignar por coincidencia account_code
                # -------------------------------------------------
                for linea_dest in lineas_dest_move:

                    account_dest_id = linea_dest["account_id"][0]

                    acc_dest = models.execute_kw(
                        db, uid, password,
                        "account.account", "read",
                        [[account_dest_id]],
                        {"fields": ["code"]}
                    )[0]

                    account_code = acc_dest["code"]

                    if account_code in origen_por_code and origen_por_code[account_code]:

                        origen_line_id = origen_por_code[account_code].pop(0)

                        models.execute_kw(
                            db, uid, password,
                            "account.move.line", "write",
                            [[linea_dest["id"]], {
                                "x_id_interno": origen_line_id
                            }]
                        )

                        if origen_line_id in origen_restantes:
                            origen_restantes.remove(origen_line_id)

                        asignadas_dest.add(linea_dest["id"])
                        actualizados += 1

                # -------------------------------------------------
                # 2️⃣ Segunda pasada → asignar restantes por descarte
                # -------------------------------------------------
                for linea_dest in lineas_dest_move:

                    if linea_dest["id"] in asignadas_dest:
                        continue

                    if not origen_restantes:
                        continue

                    origen_line_id = origen_restantes.pop(0)

                    models.execute_kw(
                        db, uid, password,
                        "account.move.line", "write",
                        [[linea_dest["id"]], {
                            "x_id_interno": origen_line_id
                        }]
                    )

                    actualizados += 1

                if actualizados % 100 == 0:
                    print(f"   🔁 Actualizadas: {actualizados}")

            print(f"✅ Total líneas conciliables actualizadas: {actualizados}")
            print("🏁 Proceso finalizado")

        #actualizar_x_id_interno_lineas_facturas()

        def export_facturas_con_multiples_lineas_conciliables_origen_destino(ruta_excel=ruta):

            print("🔎 Exportando facturas con múltiples líneas conciliables (ORIGEN + DESTINO)...")

            # =================================================
            # 1️⃣ ORIGEN
            # =================================================
            lineas_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[
                    ("company_id", "=", 2),
                    ("account_id.reconcile", "=", True),
                    ("move_id.move_type", "in", [
                        "out_invoice",
                        "in_invoice",
                        "out_refund",
                        "in_refund",
                    ]),
                ]],
                {
                    "fields": [
                        "id",
                        "move_id",
                        "account_id",
                        "debit",
                        "credit",
                    ]
                }
            )

            grupos = defaultdict(list)

            for l in lineas_origen:
                if l.get("move_id"):
                    grupos[l["move_id"][0]].append(l)

            resultados_origen = []
            facturas_afectadas = set()

            for move_id, lineas_move in grupos.items():

                if len(lineas_move) <= 1:
                    continue

                move = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "account.move", "read",
                    [[move_id]],
                    {"fields": ["name"]}
                )[0]

                move_name = move["name"]
                facturas_afectadas.add(move_name)

                for linea in lineas_move:
                    acc = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.account", "read",
                        [[linea["account_id"][0]]],
                        {"fields": ["code"]}
                    )[0]

                    resultados_origen.append({
                        "move_name": move_name,
                        "line_id_origen": linea["id"],
                        "account_code": acc["code"],
                        "debit": linea["debit"],
                        "credit": linea["credit"],
                    })

            print(f"❗ Facturas afectadas detectadas: {len(facturas_afectadas)}")

            # =================================================
            # 2️⃣ DESTINO
            # =================================================
            resultados_destino = []

            for move_name in facturas_afectadas:

                # Buscar move en DESTINO
                move_dest = models.execute_kw(
                    db, uid, password,
                    "account.move", "search",
                    [[
                        ("name", "=", move_name),
                        ("company_id", "=", 1),
                    ]],
                    {"limit": 1}
                )

                if not move_dest:
                    continue

                move_dest_id = move_dest[0]

                # Buscar líneas conciliables en DESTINO
                lineas_dest = models.execute_kw(
                    db, uid, password,
                    "account.move.line", "search_read",
                    [[
                        ("move_id", "=", move_dest_id),
                        ("account_id.reconcile", "=", True),
                    ]],
                    {
                        "fields": [
                            "id",
                            "x_id_interno",
                            "debit",
                            "credit",
                        ]
                    }
                )

                for linea in lineas_dest:
                    resultados_destino.append({
                        "move_name": move_name,
                        "line_id_destino": linea["id"],
                        "x_id_interno": linea.get("x_id_interno"),
                        "debit": linea["debit"],
                        "credit": linea["credit"],
                    })

            # =================================================
            # 3️⃣ Exportar ambas hojas
            # =================================================
            with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:

                df_origen = pd.DataFrame(resultados_origen)
                df_destino = pd.DataFrame(resultados_destino)

                df_origen.sort_values(["move_name", "line_id_origen"], inplace=True)
                df_destino.sort_values(["move_name", "line_id_destino"], inplace=True)

                df_origen.to_excel(writer, sheet_name="ORIGEN", index=False)
                df_destino.to_excel(writer, sheet_name="DESTINO", index=False)

            print(f"📁 Excel generado correctamente en: {ruta_excel}")
            print("🏁 Proceso finalizado")

        #actualizar_x_id_interno_lineas_facturas()
        # export_facturas_con_multiples_lineas_conciliables_origen_destino()

        def debug_replicar_conciliaciones_por_matching(company_id=2, limite_grupos=None):
            from collections import defaultdict

            facturas_a_conciliar = [
                #"FC1OP/22/0786",
                #"FC1OP/22/0797",
                #"FC1OP/22/0798",
                #"FC1OP/22/1512",
                #"FC1OP/22/1519",
                #"FC1OP/23/0273",
                #"FC1OP/23/0274",
                #"FC1OP/23/0644",
                #"FC1OP/23/0645",
                #"FC1OP/23/0960",
                #"FC1OP/23/0961",
                #"FC1OP/23/0962",
                #"FC1OP/23/1397",
                #"FC1OP/23/1455",
                #"FC1OP/24/0143",
                #"FC1OP/24/0770",
                #"FC1OP/25/0472",
            ]

            print("🧪 DEBUG conciliaciones por matching_number")
            print(f"🔢 Límite de grupos a procesar: {limite_grupos}")

            # -------------------------------------------------
            # 1️⃣ Obtener líneas conciliadas conciliables en ORIGEN
            # -------------------------------------------------
            lineas_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[
                    ("reconciled", "=", True),
                    ("account_id.reconcile", "=", True),
                    ("company_id", "=", company_id),
                    ("date", ">=", "2026-01-01"),
                    ("date", "<",  "2027-01-01"),
                ]],
                {
                    "fields": [
                        "id",
                        "move_id",
                        "matching_number",
                        "account_id",
                        "debit",
                        "credit",
                    ]
                }
            )

            print(f"📊 Total líneas conciliadas origen: {len(lineas_origen)}")

            # -------------------------------------------------
            # 2️⃣ Agrupar por matching_number
            # -------------------------------------------------
            grupos = defaultdict(list)

            for l in lineas_origen:
                if l["matching_number"]:
                    grupos[l["matching_number"]].append(l)

            print(f"📦 Total grupos detectados: {len(grupos)}")

            # -------------------------------------------------
            # 3️⃣ Procesar solo N grupos
            # -------------------------------------------------
            grupos_procesados = 0

            for matching, lineas in grupos.items():

                if limite_grupos and grupos_procesados >= limite_grupos:
                    break

                print("\n" + "=" * 60)
                print(f"🔵 Procesando matching_number: {matching}")
                if matching == "P":
                    print("Saltamos Parcial")
                    continue
                print(f"   Líneas en grupo: {len(lineas)}")

                line_ids_destino = []

                for l in lineas:

                    move_name = l["move_id"][1]
                    account_id_origen = l["account_id"][0]
                    move_name = move_name.split("(")[0].strip()

                    if "NOMIN" in move_name: continue

                    print(f"\n   📄 Move ORIGEN: {move_name}")

                    if move_name in facturas_a_conciliar: log_error(f"Mirar: {move_name}")

                    # Obtener código cuenta origen
                    acc_origen = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.account", "read",
                        [[account_id_origen]],
                        {"fields": ["code"]}
                    )[0]

                    account_code = acc_origen["code"]
                    print(f"   💳 Cuenta ORIGEN código: {account_code}")

                    # Buscar move destino
                    move_dest = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", move_name)]],
                        {"limit": 1}
                    )

                    if not move_dest:
                        print("   ❌ Move no encontrado en DESTINO")
                        continue

                    move_dest_id = move_dest[0]
                    print(f"   ✅ Move DESTINO encontrado (id={move_dest_id})")

                    # Buscar cuenta destino
                    cuenta_dest = models.execute_kw(
                        db, uid, password,
                        "account.account", "search",
                        [[("code", "=", account_code)]],
                        {"limit": 1}
                    )

                    if not cuenta_dest:
                        print("   ❌ Cuenta no encontrada en DESTINO")
                        continue

                    account_dest_id = cuenta_dest[0]

                    # Buscar línea destino conciliable
                    linea_dest = models.execute_kw(
                        db, uid, password,
                        "account.move.line", "search",
                        [[
                            ("move_id", "=", move_dest_id),
                            ("account_id", "=", account_dest_id),
                            ("x_id_interno", "=", l["id"]),
                            ("account_id.reconcile", "=", True),
                        ]],
                        {"limit": 1}
                    )

                    if not linea_dest:
                        try:
                            print("   ⚠️ Línea conciliable no encontrada en DESTINO. Corrigiendo cuenta...")

                            # Obtener línea principal de la factura en destino
                            linea_factura_dest = obtener_linea_factura_destino(move_dest_id)

                            # Forzar cambio de cuenta
                            models.execute_kw(
                                db, uid, password,
                                "account.move.line", "write",
                                [[linea_factura_dest["line_id"]], {
                                    "account_id": account_dest_id
                                }],
                                {"context": {"check_move_validity": False}}
                            )

                            # Volver a buscar línea destino conciliable
                            linea_dest = models.execute_kw(
                                db, uid, password,
                                "account.move.line", "search",
                                [[
                                    ("move_id", "=", move_dest_id),
                                    ("account_id", "=", account_dest_id),
                                    ("x_id_interno", "=", l["id"]),
                                    ("account_id.reconcile", "=", True),
                                ]],
                                {"limit": 1}
                            )

                            if not linea_dest:
                                print("   ❌ Línea sigue sin encontrarse tras corrección")
                                continue
                        except:
                            continue

                    print(f"   ✅ Línea DESTINO encontrada (id={linea_dest[0]})")

                    line_ids_destino.append(linea_dest[0])

                print(f"\n   🔎 Total líneas destino encontradas: {len(line_ids_destino)}")

                # -------------------------------------------------
                # Intentar reconciliar si grupo válido
                # -------------------------------------------------
                if len(line_ids_destino) >= 2:
                    print("   🔗 Intentando reconciliar grupo...")

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "account.move.line", "reconcile",
                            [line_ids_destino]
                        )
                        print("   ✅ Grupo conciliado correctamente")
                    except Exception as e:
                        msg = str(e)
                        if "cannot marshal None unless allow_none is enabled" in msg:
                            print("   ✅ Grupo conciliado correctamente")
                        elif "Está tratando de conciliar algunos asientos que ya han sido conciliados" in msg:
                            #log_error(f"Conciliar: {move_name}")
                            print("   ⚠️ Log guardado...")
                        else: print(f"   ❌ Error al reconciliar: {e}")
                else:
                    print("   ⚠️ Grupo incompleto, no se reconcilia")

                grupos_procesados += 1

            print("\n🏁 DEBUG finalizado")

        ''' #PARA EXCEPCIONES
        def debug_replicar_conciliaciones_por_matching(company_id=2, limite_grupos=None):
            from collections import defaultdict

            print("🧪 DEBUG conciliaciones por matching_number")
            print(f"🔢 Límite de grupos a procesar: {limite_grupos}")

            facturas_a_conciliar = [
                "FC1OP/22/0786",
                "FC1OP/22/0797",
                "FC1OP/22/0798",
                "FC1OP/22/1512",
                "FC1OP/22/1519",
                "FC1OP/23/0273",
                "FC1OP/23/0274",
                "FC1OP/23/0644",
                "FC1OP/23/0645",
                "FC1OP/23/0960",
                "FC1OP/23/0961",
                "FC1OP/23/0962",
                "FC1OP/23/1397",
                "FC1OP/23/1455",
                "FC1OP/24/0143",
                "FC1OP/24/0770",
                "FC1OP/25/0472",
            ]

            facturas_excepcion = {
                "FC1OP/25/0376",
                "FC1OP/25/0462",
                "FC1OP/25/0517",
                "FC1OP/25/0557",
                "FC1OP/25/0603",
                "FC1OP/25/0656",
                "FC1OP/25/0715",
                "FC1OP/25/0770",
                "FC1OP/25/0816",
                "FC1OP/25/0843",
                "FC1OP/25/0844",
                "FC1OP/25/0893",
                "FC1OP/25/0939",
                "FC1OP/25/0947",
                "FC1OP/25/0999",
                "FC1OP/25/1001",
                "FC1OP/25/1076",
                "FC1OP/25/1079",
                "FC1OP/25/1130",
                "FC1OP/25/1187",
                "FC1OP/25/1189",
                "FC1OP/25/1259",
                "FC1OP/25/1260",
                "FC1OP/25/1322",
                "FC1OP/26/0036",
            }

            #region Parte INICIAL
            # -------------------------------------------------
            # 1️⃣ Obtener líneas conciliadas conciliables en ORIGEN
            # -------------------------------------------------
            lineas_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[
                    ("reconciled", "=", True),
                    ("account_id.reconcile", "=", True),
                    ("company_id", "=", company_id),
                ]],
                {
                    "fields": [
                        "id",
                        "move_id",
                        "matching_number",
                        "account_id",
                        "debit",
                        "credit",
                    ]
                }
            )

            print(f"📊 Total líneas conciliadas origen: {len(lineas_origen)}")

            # -------------------------------------------------
            # 2️⃣ Agrupar por matching_number
            # -------------------------------------------------
            grupos = defaultdict(list)

            for l in lineas_origen:
                if l["matching_number"]:
                    grupos[l["matching_number"]].append(l)

            print(f"📦 Total grupos detectados: {len(grupos)}")

            # -------------------------------------------------
            # 3️⃣ Procesar solo N grupos
            # -------------------------------------------------
            grupos_procesados = 0
            move_name=""
            #endregion

            for matching, lineas in grupos.items():

                if limite_grupos and grupos_procesados >= limite_grupos:
                    break

                print("\n" + "=" * 60)
                print(f"🔵 Procesando matching_number: {matching}")
                print(f"   Líneas en grupo: {len(lineas)}")

                line_ids_destino = []
                aplicar_modo_excepcion = True #se puede tocar
                modo_excepcion = False #no tocar

                # -------------------------------------------------
                # Detectar si el grupo contiene factura excepción
                # -------------------------------------------------
                for l_check in lineas:
                    move_name_check = l_check["move_id"][1].split("(")[0].strip()
                    if aplicar_modo_excepcion and move_name_check in facturas_excepcion:
                        modo_excepcion = True
                        print("   🔵 Grupo contiene factura EXCEPCIÓN → modo_excepcion=True")
                        break

                # -------------------------------------------------
                # Procesar líneas
                # -------------------------------------------------
                for l in lineas:

                    move_name = l["move_id"][1]
                    account_id_origen = l["account_id"][0]
                    move_name = move_name.split("(")[0].strip()

                    # -------------------------------------------------
                    # FILTRO NORMAL (solo si NO es excepción)
                    # -------------------------------------------------
                    if not modo_excepcion and aplicar_modo_excepcion:
                        if move_name not in facturas_a_conciliar:
                            continue

                    if "NOMIN" in move_name:
                        continue

                    print(f"\n   📄 Move ORIGEN: {move_name}")

                    # Obtener código cuenta origen
                    acc_origen = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.account", "read",
                        [[account_id_origen]],
                        {"fields": ["code"]}
                    )[0]

                    account_code = acc_origen["code"]
                    print(f"   💳 Cuenta ORIGEN código: {account_code}")

                    # Buscar move destino
                    move_dest = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", move_name)]],
                        {"limit": 1}
                    )

                    if not move_dest:
                        print("   ❌ Move no encontrado en DESTINO")
                        continue

                    move_dest_id = move_dest[0]
                    print(f"   ✅ Move DESTINO encontrado (id={move_dest_id})")

                    # -------------------------------------------------
                    # CASO EXCEPCIÓN
                    # -------------------------------------------------
                    if modo_excepcion and move_name in facturas_excepcion:

                        linea_dest = models.execute_kw(
                            db, uid, password,
                            "account.move.line", "search",
                            [[
                                ("move_id", "=", move_dest_id),
                                ("account_id.reconcile", "=", True),
                            ]],
                            {"limit": 1}
                        )

                        if not linea_dest:
                            print("   ❌ No se encontró línea conciliable única en DESTINO")
                            continue

                        print(f"   🔵 Excepción detectada → usando línea única DESTINO (id={linea_dest[0]})")

                        if linea_dest[0] not in line_ids_destino:
                            line_ids_destino.append(linea_dest[0])

                        continue

                    # -------------------------------------------------
                    # FLUJO NORMAL
                    # -------------------------------------------------
                    linea_dest = models.execute_kw(
                        db, uid, password,
                        "account.move.line", "search",
                        [[
                            ("move_id", "=", move_dest_id),
                            ("x_id_interno", "=", l["id"]),
                        ]],
                        {"limit": 1}
                    )

                    if not linea_dest:
                        print(f"   ❌ Línea destino no encontrada para origen_line_id={l['id']}")
                        continue

                    print(f"   ✅ Línea DESTINO encontrada (id={linea_dest[0]})")

                    line_ids_destino.append(linea_dest[0])

                print(f"\n   🔎 Total líneas destino encontradas: {len(line_ids_destino)}")

                # -------------------------------------------------
                # Intentar reconciliar si grupo válido
                # -------------------------------------------------
                if len(line_ids_destino) >= 2:
                    print("   🔗 Intentando reconciliar grupo...")

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "account.move.line", "reconcile",
                            [line_ids_destino]
                        )
                        print("   ✅ Grupo conciliado correctamente")
                    except Exception as e:
                        msg = str(e)
                        if "cannot marshal None unless allow_none is enabled" in msg:
                            print("   ✅ Grupo conciliado correctamente")
                        elif "Está tratando de conciliar algunos asientos que ya han sido conciliados" in msg:
                            #log_error(f"Conciliar: {move_name}")
                            print("   ⚠️ Log guardado...")
                        else: print(f"   ❌ Error al reconciliar: {e}")
                else:
                    print("   ⚠️ Grupo incompleto, no se reconcilia")

                grupos_procesados += 1

            print("\n🏁 DEBUG finalizado")
        '''

        def debug_replicar_conciliaciones_por_matching_v2(company_id=2, limite_grupos=None):
            from collections import defaultdict

            print("🧪 DEBUG conciliaciones por matching_number")
            print(f"🔢 Límite de grupos a procesar: {limite_grupos}")

            # -------------------------------------------------
            # 1️⃣ Obtener líneas conciliadas conciliables en ORIGEN
            # -------------------------------------------------
            lineas_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "account.move.line", "search_read",
                [[
                    ("reconciled", "=", True),
                    ("account_id.reconcile", "=", True),
                    ("company_id", "=", company_id),
                    ("date", ">=", "2026-01-01"), ("date", "<",  "2027-01-01"),
                ]],
                {
                    "fields": [
                        "id",
                        "move_id",
                        "matching_number",
                        "account_id",
                        "debit",
                        "credit",
                    ]
                }
            )

            print(f"📊 Total líneas conciliadas origen: {len(lineas_origen)}")

            # -------------------------------------------------
            # 2️⃣ Agrupar por matching_number
            # -------------------------------------------------
            grupos = defaultdict(list)

            for l in lineas_origen:
                if l["matching_number"]:
                    grupos[l["matching_number"]].append(l)

            print(f"📦 Total grupos detectados: {len(grupos)}")

            # -------------------------------------------------
            # 3️⃣ Procesar solo N grupos
            # -------------------------------------------------
            grupos_procesados = 0

            for matching, lineas in grupos.items():

                if limite_grupos and grupos_procesados >= limite_grupos:
                    break

                print("\n" + "=" * 60)
                print(f"🔵 Procesando matching_number: {matching}")
                if matching == "P":
                    print("Saltamos Parcial")
                    continue
                print(f"   Líneas en grupo: {len(lineas)}")

                line_ids_destino = []

                for l in lineas:

                    move_name = l["move_id"][1]
                    account_id_origen = l["account_id"][0]
                    move_name = move_name.split("(")[0].strip()

                    if "NOMIN" in move_name: continue

                    print(f"\n   📄 Move ORIGEN: {move_name}")

                    # Obtener código cuenta origen
                    acc_origen = models_src.execute_kw(
                        db_src, uid_src, password_src,
                        "account.account", "read",
                        [[account_id_origen]],
                        {"fields": ["code"]}
                    )[0]

                    account_code = acc_origen["code"]
                    print(f"   💳 Cuenta ORIGEN código: {account_code}")

                    # Buscar move destino
                    move_dest = models.execute_kw(
                        db, uid, password,
                        "account.move", "search",
                        [[("name", "=", move_name)]],
                        {"limit": 1}
                    )

                    if not move_dest:
                        print("   ❌ Move no encontrado en DESTINO")
                        continue

                    move_dest_id = move_dest[0]
                    print(f"   ✅ Move DESTINO encontrado (id={move_dest_id})")

                    # Buscar cuenta destino
                    cuenta_dest = models.execute_kw(
                        db, uid, password,
                        "account.account", "search",
                        [[("code", "=", account_code)]],
                        {"limit": 1}
                    )

                    if not cuenta_dest:
                        print("   ❌ Cuenta no encontrada en DESTINO")
                        continue

                    account_dest_id = cuenta_dest[0]

                    # Buscar línea destino conciliable
                    linea_dest = models.execute_kw(
                        db, uid, password,
                        "account.move.line", "search",
                        [[
                            ("move_id", "=", move_dest_id),
                            ("account_id", "=", account_dest_id),
                            ("x_id_interno", "=", l["id"]),
                            ("account_id.reconcile", "=", True),
                        ]],
                        {"limit": 1}
                    )

                    if not linea_dest:
                        try:
                            print("   ⚠️ Línea conciliable no encontrada en DESTINO. Corrigiendo cuenta...")

                            # Obtener línea principal de la factura en destino
                            linea_factura_dest = obtener_linea_factura_destino(move_dest_id)

                            # Forzar cambio de cuenta
                            models.execute_kw(
                                db, uid, password,
                                "account.move.line", "write",
                                [[linea_factura_dest["line_id"]], {
                                    "account_id": account_dest_id
                                }],
                                {"context": {"check_move_validity": False}}
                            )

                            # Volver a buscar línea destino conciliable
                            linea_dest = models.execute_kw(
                                db, uid, password,
                                "account.move.line", "search",
                                [[
                                    ("move_id", "=", move_dest_id),
                                    ("account_id", "=", account_dest_id),
                                    ("x_id_interno", "=", l["id"]),
                                    ("account_id.reconcile", "=", True),
                                ]],
                                {"limit": 1}
                            )

                            if not linea_dest:
                                print("   ❌ Línea sigue sin encontrarse tras corrección")
                                continue
                        except:
                            continue

                    print(f"   ✅ Línea DESTINO encontrada (id={linea_dest[0]})")

                    line_ids_destino.append(linea_dest[0])

                print(f"\n   🔎 Total líneas destino encontradas: {len(line_ids_destino)}")

                # -------------------------------------------------
                # Intentar reconciliar si grupo válido
                # -------------------------------------------------
                if len(line_ids_destino) >= 2:
                    print("   🔗 Intentando reconciliar grupo...")

                    try:
                        models.execute_kw(
                            db, uid, password,
                            "account.move.line", "reconcile",
                            [line_ids_destino]
                        )
                        print("   ✅ Grupo conciliado correctamente")
                    except Exception as e:
                        msg = str(e)
                        if "cannot marshal None unless allow_none is enabled" in msg:
                            print("   ✅ Grupo conciliado correctamente")
                        elif "Está tratando de conciliar algunos asientos que ya han sido conciliados" in msg:
                            #log_error(f"Conciliar: {move_name}")
                            pass#print("   ⚠️ Log guardado...")
                        else:
                            print(f"   ❌ Error al reconciliar: {e}")
                elif len(line_ids_destino) > 2:

                    print("   🔗 Reconciliación secuencial (>2 líneas)...")

                    # Tomamos la primera línea como base
                    linea_base = line_ids_destino[0]

                    for linea in line_ids_destino[1:]:

                        try:
                            models.execute_kw(
                                db, uid, password,
                                "account.move.line", "reconcile",
                                [[linea_base, linea]]
                            )
                            print(f"      ✅ Parcial aplicado entre {linea_base} y {linea}")

                        except Exception as e:
                            msg = str(e)

                            if "ya han sido conciliados" in msg:
                                print("      ⚠️ Ya conciliado")
                            else:
                                print(f"      ❌ Error parcial: {e}")
                else:
                    print("   ⚠️ Grupo incompleto, no se reconcilia")

                grupos_procesados += 1

            print("\n🏁 DEBUG finalizado")

        #debug_replicar_conciliaciones_por_matching_v2()

        def print_lineas_conciliables_facturas():
            print("🔍 Buscando líneas conciliables de facturas...\n")

            # 1️⃣ Buscar facturas (solo customer + vendor)
            moves = models.execute_kw(
                db, uid, password,
                "account.move", "search_read",
                [[
                    ("move_type", "in", ["out_invoice", "in_invoice"]),
                    ("state", "=", "posted"),
                ]],
                {"fields": ["id", "name"]}
            )

            print(f"📄 Facturas encontradas: {len(moves)}\n")

            for move in moves:

                # 2️⃣ Buscar líneas conciliables
                lineas = models.execute_kw(
                    db, uid, password,
                    "account.move.line", "search_read",
                    [[
                        ("move_id", "=", move["id"]),
                        ("account_id.reconcile", "=", True),
                    ]],
                    {
                        "fields": [
                            "id",
                            "name",
                            "account_id",
                            "x_id_interno",
                            "debit",
                            "credit",
                            "balance",
                        ]
                    }
                )

                if lineas:
                    print(f"\n🧾 Factura: {move['name']} (ID {move['id']})")
                    print("   ─────────────────────────────")

                    for l in lineas:
                        print(
                            f"   ➤ Línea ID: {l['id']} | "
                            f"x_id_interno: {l.get('x_id_interno')} | "
                            f"Cuenta: {l['account_id'][1]} | "
                            f"Balance: {l['balance']}"
                        )

        def vincular_tareas_proyectos_desde_excel(ruta_excel=ruta):
            """
            Lee un Excel con columna 'x_id_interno' (ID origen del proyecto),
            busca las tareas asociadas en origen y las vincula en destino
            usando el campo custom x_id_interno tanto en proyecto como en tarea.
            """

            # -----------------------------
            # 1️⃣ Leer Excel
            # -----------------------------
            df = pd.read_excel(ruta_excel)

            if "x_id_interno" not in df.columns:
                print("❌ El Excel no contiene la columna 'x_id_interno'")
                return

            proyectos_origen_ids = df["x_id_interno"].dropna().astype(int).tolist()

            print(f"📌 Proyectos a procesar: {len(proyectos_origen_ids)}")

            # -----------------------------
            # 2️⃣ Buscar proyectos en origen
            # -----------------------------
            proyectos_origen = models_src.execute_kw(
                db_src, uid_src, password_src,
                "project.project", "search_read",
                [[("id", "in", proyectos_origen_ids), ("active", "in", [True, False])]],
                {"fields": ["id", "name"]}
            )

            print(f"📁 Proyectos encontrados en origen: {len(proyectos_origen)}")

            for proyecto in proyectos_origen:

                proyecto_id_origen = proyecto["id"]

                # -----------------------------
                # 3️⃣ Buscar proyecto correspondiente en destino
                # -----------------------------
                proyecto_destino = models.execute_kw(
                    db, uid, password,
                    "project.project", "search_read",
                    [[("x_id_interno", "=", proyecto_id_origen)]],
                    {"fields": ["id", "name"]}
                )

                if not proyecto_destino:
                    print(f"⚠ Proyecto no encontrado en destino para origen ID {proyecto_id_origen}")
                    continue

                proyecto_id_destino = proyecto_destino[0]["id"]

                # -----------------------------
                # 4️⃣ Buscar tareas del proyecto en origen
                # -----------------------------
                tareas_origen = models_src.execute_kw(
                    db_src, uid_src, password_src,
                    "project.task", "search_read",
                    [[("project_id", "=", proyecto_id_origen),
                      ("active", "in", [True, False])]],
                    {"fields": ["id", "name"]}
                )

                print(f"   🔎 Tareas en origen: {len(tareas_origen)}")

                for tarea in tareas_origen:

                    tarea_id_origen = tarea["id"]

                    # -----------------------------
                    # 5️⃣ Buscar tarea equivalente en destino
                    # -----------------------------
                    tarea_destino = models.execute_kw(
                        db, uid, password,
                        "project.task", "search_read",
                        [[("x_id_interno", "=", tarea_id_origen)]],
                        {"fields": ["id", "project_id"]}
                    )

                    if not tarea_destino:
                        print(f"      ⚠ Tarea no encontrada en destino (origen ID {tarea_id_origen})")
                        continue

                    tarea_id_destino = tarea_destino[0]["id"]

                    # -----------------------------
                    # 6️⃣ Vincular tarea al proyecto correcto
                    # -----------------------------
                    models.execute_kw(
                        db, uid, password,
                        "project.task", "write",
                        [[tarea_id_destino], {"project_id": proyecto_id_destino}]
                    )

                    print(f"      ✅ Tarea {tarea_id_destino} vinculada al proyecto {proyecto_id_destino}")

            print("🎯 Proceso finalizado.")

        #vincular_tareas_proyectos_desde_excel()

        def eliminar_pickings_excepto_excel(ruta_excel=ruta):
            """
            Elimina masivamente todos los stock.picking
            excepto los que estén en la columna 'Pickings' del Excel.
            """

            print(f"📖 Leyendo Excel: {ruta_excel}")

            # ------------------------------------------------
            # 1️⃣ Leer Excel
            # ------------------------------------------------
            df = pd.read_excel(ruta_excel)

            if "Pickings" not in df.columns:
                print("❌ El Excel no contiene la columna 'Pickings'")
                return

            pickings_a_conservar = (
                df["Pickings"]
                .dropna()
                .astype(str)
                .str.strip()
                .tolist()
            )

            if not pickings_a_conservar:
                print("⚠️ No hay pickings para conservar. Se eliminarán TODOS.")
            else:
                print(f"🔒 Pickings a conservar: {len(pickings_a_conservar)}")

            # ------------------------------------------------
            # 2️⃣ Buscar todos los pickings
            # ------------------------------------------------
            all_picking_ids = models.execute_kw(
                db, uid, password,
                "stock.picking", "search",
                [[]]
            )

            if not all_picking_ids:
                print("ℹ️ No existen pickings")
                return

            print(f"📦 Total pickings en sistema: {len(all_picking_ids)}")

            # ------------------------------------------------
            # 3️⃣ Leer nombres
            # ------------------------------------------------
            all_pickings = models.execute_kw(
                db, uid, password,
                "stock.picking", "read",
                [all_picking_ids],
                {"fields": ["id", "name", "state"]}
            )

            # ------------------------------------------------
            # 4️⃣ Filtrar los que deben eliminarse
            # ------------------------------------------------
            pickings_eliminar = [
                p for p in all_pickings
                if p["name"] not in pickings_a_conservar
            ]

            if not pickings_eliminar:
                print("✅ No hay pickings para eliminar")
                return

            print(f"🔥 Pickings a eliminar: {len(pickings_eliminar)}")

            ids_eliminar = [p["id"] for p in pickings_eliminar]

            # ------------------------------------------------
            # 5️⃣ Cancelar los que no estén cancelados
            # ------------------------------------------------
            for p in pickings_eliminar:
                if p["state"] != "cancel":
                    try:
                        models.execute_kw(
                            db, uid, password,
                            "stock.picking", "action_cancel",
                            [[p["id"]]]
                        )
                    except Exception as e:
                        print(f"⚠️ Error cancelando {p['name']}: {e}")

            # ------------------------------------------------
            # 6️⃣ Eliminar
            # ------------------------------------------------
            try:
                models.execute_kw(
                    db, uid, password,
                    "stock.picking", "unlink",
                    [ids_eliminar]
                )
                print(f"✅ Eliminados {len(ids_eliminar)} pickings")
            except Exception as e:
                print(f"❌ Error eliminando pickings: {e}")

        def actualizar_website_description_por_sku(excel_path=ruta, campo_odoo="website_description"):
            DESCRIPTION_HTML_1 = """<ul class="nav nav-tabs justify-content-center" role="tablist">
              <li class="nav-item o_not_editable" data-oe-model="ir.ui.view" data-oe-id="3823" data-oe-field="arch" data-oe-xpath="/t[1]/div[1]/ul[1]/li[2]">
                <a class="nav-link active" data-bs-toggle="tab" href="#tp-product-description-tab" role="tab" aria-selected="true">
                  <span class="fa fa-file-text-o me-1"></span> Descripción </a>
              </li>
              <li class="nav-item o_not_editable" data-oe-model="ir.ui.view" data-oe-id="3823" data-oe-field="arch" data-oe-xpath="/t[1]/div[1]/ul[1]/li[3]">
                <a class="nav-link" data-bs-toggle="tab" href="#tp-product-specification-tab" role="tab" aria-selected="false">
                  <span class="fa fa-sliders me-1"></span> Especificaciones </a>
              </li>
            </ul>
            <div class="tab-content">
              <div class="tab-pane fade active show" id="tp-product-description-tab" role="tabpanel">
                <div class="container-fluid">
                  <div class="row m-0 py-2">
                    <div class="col-12">
                      <div itemprop="description" class="oe_structure" id="product_full_description" data-oe-xpath="/t[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]" data-oe-model="product.template" data-oe-id="1715" data-oe-field="website_description" data-oe-type="html" data-oe-expression="product.website_description">
                        <div class="product-description" style="margin-left: 16.6667%;margin-right: 16.6667%;">
                          <div></div> """
            DESCRIPTION_HTML_1_END = """
                        </div>
                      </div>
                    </div>
                  </div>
                </div>
              </div>
              <div class="tab-pane fade" id="tp-product-specification-tab" role="tabpanel"></div>
            </div>"""
            SCRIPT_JS_HTML = """
            <script>
              document.addEventListener("DOMContentLoaded", function() {{
                var el = document.getElementById("product_full_spec") || document.getElementById("product_attributes_simple");
                if (el) {{
                  el.style.marginTop = "40px";
                  el.style.marginLeft = "16.6667%";
                  el.style.borderTop = "none";
                }}
                var elementToMove = document.getElementById("product_attributes_simple");
                var targetContainer = document.getElementById("tp-product-specification-tab");

                if (elementToMove && targetContainer) {{
                    targetContainer.appendChild(elementToMove);
                }}
                document.getElementById("oe_structure_website_sale_product_2").innerHTML = `
                    <hr>
                    <div class="oe_structure oe_empty oe_structure_not_nearest mt16" id="oe_structure_website_sale_product_2" data-oe-model="ir.ui.view" data-oe-id="2195" data-oe-field="arch" data-oe-xpath="/t[1]/t[4]/div[1]/div[3]" data-editor-message="DROP BUILDING BLOCKS HERE TO MAKE THEM AVAILABLE ACROSS ALL PRODUCTS" style="
                        margin-bottom: 16px !important;
                    ">
                      <div id="icons" class="row" style="justify-content: center;">
                      </div>
                    </div>`
                var iconsContainer = document.querySelector("#icons");
                var tempBlock = document.querySelector("#temp_icon_block");
                if (iconsContainer && tempBlock) {
                  iconsContainer.appendChild(tempBlock);
                }
              }});
              document.addEventListener("DOMContentLoaded", function () {{
                function toggleProductFullSpec() {{
                  var descriptionTab = document.querySelector('a[href="#tp-product-description-tab"]');
                  var productFullSpec = document.getElementById("product_full_spec") || document.getElementById("product_attributes_simple");

                  if (descriptionTab && productFullSpec) {{
                    if (descriptionTab.classList.contains("active")) {{
                      productFullSpec.style.display = "none";
                    }} else {{
                      productFullSpec.style.display = "block";
                    }}
                  }}
                }}

                // Ejecutar al cargar
                toggleProductFullSpec();

                // Escuchar cambios de pestaña
                var tabLinks = document.querySelectorAll('[data-bs-toggle="tab"]');
                tabLinks.forEach(function (tab) {{
                  tab.addEventListener("shown.bs.tab", function () {{
                    toggleProductFullSpec();
                  }});
                }});

                function hideIfZero(valueId, boxId) {
                    const el = document.getElementById(valueId);
                    const box = document.getElementById(boxId);
                    if (el && box) {
                      const num = parseInt(el.innerText.trim()) || 0;
                      if (num === 0) {
                        box.style.display = "none";
                      }
                    }
                  }

                  hideIfZero("europeo", "box_europeo");
                  hideIfZero("nacional", "box_nacional");
                  hideIfZero("alicante", "box_alicante");
                  const style = document.createElement("style");
                  style.textContent = "h6.mb-1 span p { display: inline !important; }";
                  document.head.appendChild(style);
              }});
            </script>
            """

            def safe_text_(text):
                return "" if pd.isna(text) else text

            df = pd.read_excel(excel_path)

            for i, fila in df.iterrows():

                sku = fila.get("SKU")
                valor = fila.get("Descripción Web")
                documentos = fila.get("Documentos")

                if not sku:
                    continue

                if not valor:
                    continue

                # ---------------------------------------------------
                # Construcción HTML de PDFs
                # ---------------------------------------------------

                pdf_html = """<div class="row">"""

                if documentos and isinstance(documentos, str):
                    try:
                        pdf_dict = ast.literal_eval(documentos)

                        for i, (pdf_name, pdf_url) in enumerate(pdf_dict.items(), start=1):
                            bloque = f"""
                            <div class="col-md-4 mb-3">
                              <div class="d-flex justify-content-between align-items-center border rounded p-2">
                                <span class="text-truncate me-2" style="max-width: 70%;">{pdf_name}</span>
                                <a target="_blank" class="btn btn-sm btn-outline-secondary" href="{pdf_url}" title="Descargar PDF">
                                  <svg width="20" height="20" viewBox="0 0 20 20" xmlns="http://www.w3.org/2000/svg">
                                    <polyline fill="none" stroke="#000" points="14,10 9.5,14.5 5,10"></polyline>
                                    <rect x="3" y="17" width="13" height="1"></rect>
                                    <line fill="none" stroke="#000" x1="9.5" y1="13.91" x2="9.5" y2="3"></line>
                                  </svg>
                                </a>
                              </div>
                            </div>
                            """
                            pdf_html += bloque

                    except Exception as e:
                        print(f"⚠️ Error procesando documentos PDF en SKU {sku}: {e}")

                pdf_html += "</div>"

                # ---------------------------------------------------
                # Construcción descripción final
                # ---------------------------------------------------

                description = DESCRIPTION_HTML_1 + f"<p>{safe_text_(valor)}</p>{pdf_html}" + DESCRIPTION_HTML_1_END
                website_description = description + SCRIPT_JS_HTML

                # ---------------------------------------------------
                # Buscar producto por SKU
                # ---------------------------------------------------

                product_ids = models.execute_kw(
                    db,
                    uid,
                    password,
                    'product.template',
                    'search',
                    [[('default_code', '=', sku), ("active", "in", [False,True])]],
                    {'limit': 1}
                )

                if not product_ids:
                    print(f"❌ SKU no encontrado: {sku}")
                    continue

                product_id = product_ids[0]

                # ---------------------------------------------------
                # Actualizar producto
                # ---------------------------------------------------

                models.execute_kw(
                    db,
                    uid,
                    password,
                    'product.template',
                    'write',
                    [[product_id], {campo_odoo: website_description}]
                )

                print(f"✅ Actualizado SKU {sku}")

        def eliminar_atributos_de_productos(batch_size=200):

            excel_path=ruta

            try:
                # -------------------------------------------------
                # 🔹 Leer IDs de atributos
                # -------------------------------------------------

                df = pd.read_excel(excel_path)

                if "ID" not in df.columns:
                    raise Exception("❌ El Excel no contiene la columna 'ID'")

                attribute_ids = (
                    df["ID"]
                    .dropna()
                    .astype(int)
                    .unique()
                    .tolist()
                )

                if not attribute_ids:
                    print("ℹ️ No hay IDs de atributos en el Excel")
                    return

                print(f"🔎 Atributos detectados: {len(attribute_ids)}")

                total_eliminadas = 0

                # -------------------------------------------------
                # 🔹 Procesar por lotes
                # -------------------------------------------------

                for i in range(0, len(attribute_ids), batch_size):

                    batch = attribute_ids[i:i + batch_size]

                    # Buscar líneas de atributo en productos
                    line_ids = models.execute_kw(
                        db, uid, password,
                        'product.template.attribute.line',
                        'search',
                        [[('attribute_id', 'in', batch)]]
                    )

                    if not line_ids:
                        continue

                    # Eliminar líneas
                    models.execute_kw(
                        db, uid, password,
                        'product.template.attribute.line',
                        'unlink',
                        [line_ids]
                    )

                    total_eliminadas += len(line_ids)

                    print(f"✅ Líneas eliminadas en lote: {len(line_ids)}")

                print(f"🎯 Total líneas de atributo eliminadas: {total_eliminadas}")

            except Exception as e:
                print(f"❌ Error eliminando atributos de productos: {e}")

        def actualizar_stock_almacen3_por_sku(pwd=password, excel_path=ruta):

            try:
                import pandas as pd

                df = pd.read_excel(excel_path)

                if "SKU" not in df.columns or "STOCK" not in df.columns:
                    print("❌ El Excel debe contener columnas SKU y STOCK")
                    return

                # ----------------------------------------
                # 🔹 limpiar datos
                # ----------------------------------------

                df["SKU"] = df["SKU"].astype(str).str.strip()
                df["STOCK"] = pd.to_numeric(df["STOCK"], errors="coerce").fillna(0)

                sku_stock_map = dict(zip(df["SKU"], df["STOCK"]))

                skus = list(sku_stock_map.keys())

                print(f"🧾 SKUs detectados: {len(skus)}")

                if not skus:
                    print("ℹ️ No hay SKUs")
                    return

                # ----------------------------------------
                # 🔹 leer productos destino
                # ----------------------------------------

                products = models.execute_kw(
                    db, uid, pwd,
                    'product.template', 'search_read',
                    [[('default_code', 'in', skus)]],
                    {'fields': ['id', 'default_code']}
                )

                print(f"📦 Productos encontrados en destino: {len(products)}")

                # ----------------------------------------
                # 🔹 actualizar stock
                # ----------------------------------------

                for p in products:
                    sku = p["default_code"]
                    stock = sku_stock_map.get(sku, 0)

                    models.execute_kw(
                        db, uid, pwd,
                        'product.template', 'write',
                        [[p["id"]], {
                            'x_almacen_local': stock
                        }]
                    )

                    print(f"✅ {sku} → {stock}")

                print("🎯 Actualización de stock finalizada")

            except Exception as e:
                print(f"❌ Error: {e}")

        actualizar_stock_almacen3_por_sku()

        ruta = Utils.seleccionar_excel()

        if not ruta:
            msg = "❌ Import cancelado - No se seleccionó archivo."
            emit_progress(msg)
            print(msg)
            if event_manager:
                event_manager.emit('operation_error', "No se seleccionó archivo para importar")
            return

        def run_import():
            try:
                emit_progress("📋 Preparando archivo Excel para importación...")
                Utils.preparar_excel_import(ruta)
                df_import = pd.read_excel(ruta, sheet_name='Import'); df_update = pd.read_excel(ruta, sheet_name='Update'); df_quantity = pd.read_excel(ruta, sheet_name='Quantity')
                import_all = True if df_import.empty and df_update.empty and df_quantity.empty else False

                emit_progress("🚀 Iniciando importación de productos...")
                import_to_odoo(ruta, import_all=import_all, event_manager=event_manager)

            except Exception as e:
                msg = f"❌ Error en import_to_odoo: {e}"
                emit_progress(msg)
                print(msg)
                if event_manager:
                    event_manager.emit('operation_error', f"Error en importación: {e}")
                return

            # Solo continuar si no hubo errores
            run_update()

        def run_update():
            try:
                emit_progress("🔄 Iniciando actualización de productos...")
                update_from_merge_to_odoo(ruta, event_manager)
            except Exception as e:
                msg = f"❌ Error en update_from_merge_to_odoo: {e}"
                emit_progress(msg)
                print(msg)
                if event_manager:
                    event_manager.emit('operation_error', f"Error en actualización: {e}")
                return

            # Solo continuar si no hubo errores
            run_delete()

        def run_delete():
            try:
                emit_progress("🗑️ Iniciando eliminación de productos obsoletos...")
                delete_from_merge_to_odoo(ruta, event_manager)
                emit_progress("✅ Proceso de importación completado exitosamente")
                if event_manager:
                    event_manager.emit('operation_completed', "Importación finalizada correctamente")
            except Exception as e:
                msg = f"❌ Error en delete_from_merge_to_odoo: {e}"
                emit_progress(msg)
                print(msg)
                if event_manager:
                    event_manager.emit('operation_error', f"Error en eliminación: {e}")

        # Ejecutar en el hilo principal para que sea síncrono
        run_import()
    else:
        msg = "❌ Error en la conexión, no es posible hacer el Import."
        emit_progress(msg)
        print(msg)
        if event_manager:
            event_manager.emit('operation_error', "Error en la conexión con Odoo")
