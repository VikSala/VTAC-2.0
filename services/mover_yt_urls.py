import pandas as pd
import ast
import json

def mover_youtube_a_video_urls(ruta_excel: str):
    print(f"📂 Procesando Excel: {ruta_excel}")
    xls = pd.ExcelFile(ruta_excel)
    if 'OPT' not in xls.sheet_names:
        raise ValueError("La hoja 'OPT' no existe en el archivo.")

    df = pd.read_excel(xls, 'OPT').copy()

    if 'Image_Urls' not in df.columns or 'Video_Urls' not in df.columns:
        raise ValueError("Faltan columnas 'Image_Urls' o 'Video_Urls' en la hoja OPT.")

    def limpiar_urls(row):
        imagenes = []
        videos = []

        try:
            raw_imgs = ast.literal_eval(row['Image_Urls']) if isinstance(row['Image_Urls'], str) else []
        except Exception:
            raw_imgs = []

        try:
            raw_videos = ast.literal_eval(row['Video_Urls']) if isinstance(row['Video_Urls'], str) else []
        except Exception:
            raw_videos = []

        for url in raw_imgs:
            if isinstance(url, str) and ('youtu.be' in url or 'youtube' in url):
                videos.append(url)
            else:
                imagenes.append(url)

        return pd.Series({
            'Image_Urls': json.dumps(imagenes, ensure_ascii=False),
            'Video_Urls': json.dumps(list(set(raw_videos + videos)), ensure_ascii=False)
        })

    df_actualizado = df.apply(limpiar_urls, axis=1)
    df['Image_Urls'] = df_actualizado['Image_Urls']
    df['Video_Urls'] = df_actualizado['Video_Urls']

    # Sobrescribir Excel con la hoja OPT modificada
    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='OPT', index=False)

    print("✅ YouTube links movidos de Image_Urls a Video_Urls en la hoja OPT.")

def rellenar_skus_desde_specifications_y_guardar(excel_path: str, hoja: str = 'ES', col_sku: str = 'SKU', col_specs: str = 'Specifications'):
    """
    Lee un Excel, rellena los SKUs vacíos desde Specifications y guarda el archivo sobrescribiéndolo.

    Args:
        excel_path (str): Ruta al archivo Excel.
        hoja (str): Nombre de la hoja a procesar. Por defecto: 'ES'.
        col_sku (str): Nombre de la columna SKU. Por defecto: 'SKU'.
        col_specs (str): Nombre de la columna que contiene el diccionario. Por defecto: 'Specifications'.
    """
    print(f"📥 Cargando hoja '{hoja}' desde: {excel_path}")
    xls = pd.ExcelFile(excel_path)
    hojas = {name: pd.read_excel(xls, name, dtype=str).fillna('') for name in xls.sheet_names}

    if hoja not in hojas:
        print(f"⚠ La hoja '{hoja}' no existe en el Excel.")
        return

    df = hojas[hoja]
    num_actualizados = 0

    for i, row in df.iterrows():
        sku_val = row.get(col_sku, '').strip()
        if sku_val:
            continue

        specs_raw = row.get(col_specs, '').strip()
        if not specs_raw:
            continue

        try:
            specs_dict = ast.literal_eval(specs_raw)
            if isinstance(specs_dict, dict):
                codigo_orden = specs_dict.get('Código de orden', '').strip()
                if codigo_orden:
                    df.at[i, col_sku] = codigo_orden
                    num_actualizados += 1
        except Exception as e:
            print(f"  ✘ Error en fila {i}: {e}")

    hojas[hoja] = df  # Actualizar hoja modificada

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
        for nombre, df_hoja in hojas.items():
            df_hoja.to_excel(writer, sheet_name=nombre, index=False)

    print(f"✔ Guardado completo. {num_actualizados} SKUs actualizados en la hoja '{hoja}'.")

import openpyxl

def ordenar_hojas_por_sku(ruta_excel, nombre_columna_sku="SKU", hojas_objetivo=None):
    if hojas_objetivo is None:
        hojas_objetivo = ["ES", "UK", "ITA", "OPT"]

    wb = openpyxl.load_workbook(ruta_excel)

    for hoja_nombre in hojas_objetivo:
        if hoja_nombre not in wb.sheetnames:
            print(f"[⚠️] Hoja '{hoja_nombre}' no encontrada, se omite.")
            continue

        ws = wb[hoja_nombre]

        # Obtener encabezados
        encabezado = [cell.value for cell in ws[1]]

        # Encontrar la columna del SKU
        try:
            idx_sku = encabezado.index(nombre_columna_sku)
        except ValueError:
            print(f"[⚠️] Columna '{nombre_columna_sku}' no encontrada en hoja '{hoja_nombre}'.")
            continue

        # Leer todas las filas de datos (excepto encabezado)
        filas_datos = list(ws.iter_rows(min_row=2, values_only=True))

        # Ordenar por SKU (usando el índice de la columna SKU)
        filas_ordenadas = sorted(filas_datos, key=lambda row: str(row[idx_sku]))

        # Borrar todas las filas de datos anteriores
        ws.delete_rows(2, ws.max_row)

        # Reescribir filas ordenadas
        for fila in filas_ordenadas:
            ws.append(fila)

        print(f"[✔️] Hoja '{hoja_nombre}' ordenada por SKU.")

    # Guardar
    ruta_salida = ruta_excel.replace(".xlsx", "_ordenado.xlsx")
    wb.save(ruta_salida)
    print(f"[✅] Archivo guardado como: {ruta_salida}")

def eliminar_imagenes_icono_uk(ruta_excel, columna_urls='Image_Urls', hoja='UK'):
    identificador_icono = '47f05ef62aca7a95e81642a03ad1c70c'

    df = pd.read_excel(ruta_excel, sheet_name=hoja, dtype=str)
    if columna_urls not in df.columns:
        print(f"❌ La columna '{columna_urls}' no se encuentra en la hoja '{hoja}'.")
        return

    def limpiar_array(urls):
        if not isinstance(urls, str) or not urls.strip():
            return ''
        try:
            url_list = ast.literal_eval(urls)  # ← aquí el cambio
            if not isinstance(url_list, list):
                return urls
            url_filtradas = [u for u in url_list if identificador_icono not in u]
            return str(url_filtradas)  # lo devolvemos como string (igual que antes)
        except Exception:
            return urls

    df[columna_urls] = df[columna_urls].apply(limpiar_array)

    ruta_salida = ruta_excel
    with pd.ExcelWriter(ruta_salida, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=hoja, index=False)

    print(f"✅ Imágenes tipo icono eliminadas y archivo guardado como: {ruta_salida}")

import re
def procesar_excel(ruta_excel):
    def extraer_precio(texto):
        if not isinstance(texto, str):
            return None

        promo_match = re.search(
            r'precio\s+(?:promocional|promoción|promo)[^\d€]*€?\s*([\d\.]+,\d{2})',
            texto,
            re.IGNORECASE
        )
        if promo_match:
            return promo_match.group(1)

        normal_match = re.search(
            r'precio[^\d€]*€?\s*([\d\.]+,\d{2})',
            texto,
            re.IGNORECASE
        )
        if normal_match:
            return normal_match.group(1)

        return None
    # Leer todas las hojas
    all_sheets = pd.read_excel(ruta_excel, sheet_name=None, engine='openpyxl')

    # Modificar la hoja 'ITA' si existe
    if 'ITA' in all_sheets:
        df = all_sheets['ITA']

        df['standard_price'] = df['standard_price'].apply(extraer_precio)

        # Reemplazar en el diccionario de hojas
        all_sheets['ITA'] = df
    else:
        print("❌ La hoja 'ITA' no existe en el archivo.")
        return

    # Guardar todas las hojas de nuevo
    output_path = ruta_excel

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, sheet_df in all_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✅ Archivo guardado en: {output_path}")

def detectar_skus_unicos(archivo_excel, nombre_salida='nuevos.xlsx'):
    # Cargar todas las hojas
    hojas = pd.read_excel(archivo_excel, sheet_name=['ES', 'UK', 'ITA', 'OPT'])
    df_es = hojas['ES']
    df_uk = hojas['UK']
    df_ita = hojas['ITA']
    df_opt = hojas['OPT']

    # Extraer SKUs como sets
    skus_es = set(df_es['SKU'].astype(str))
    skus_uk = set(df_uk['SKU'].astype(str))
    skus_ita = set(df_ita['SKU'].astype(str))
    skus_opt = set(df_opt['SKU'].astype(str))

    # Todos los SKUs de comparación
    todos_comparar = {
        'ES': (df_es, skus_es - skus_uk - skus_ita - skus_opt),
        'UK': (df_uk, skus_uk - skus_es - skus_ita - skus_opt),
        'ITA': (df_ita, skus_ita - skus_es - skus_uk - skus_opt)
    }

    # Filtrar filas únicas por hoja
    df_nuevos = pd.DataFrame()
    for nombre_hoja, (df, skus_unicos) in todos_comparar.items():
        filas = df[df['SKU'].astype(str).isin(skus_unicos)]
        df_nuevos = pd.concat([df_nuevos, filas], ignore_index=True)
    return df_nuevos
    # Guardar en nueva hoja "NUEVOS"
    with pd.ExcelWriter(nombre_salida, engine='openpyxl', mode='w') as writer:
        df_nuevos.to_excel(writer, sheet_name='NUEVOS', index=False)

    print(f"Guardado {len(df_nuevos)} SKU(s) únicos en la hoja 'NUEVOS' de '{nombre_salida}'.")



detectar_skus_unicos("GigaExcelToMerge - copia.xlsx")