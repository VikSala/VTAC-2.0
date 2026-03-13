import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog

excel_dinamico = ""


def detectar_archivos_excel(directorio):
    global excel_dinamico
    archivos = [f for f in os.listdir(directorio) if f.lower().endswith(('.xlsx', '.xls'))]

    base_file = None
    secundarios = []

    for f in archivos:
        ruta = os.path.join(directorio, f)
        nombre_lower = f.lower()
        if any(x in nombre_lower for x in ['nave', 'madrid', 'bulgaria']):
            secundarios.append((nombre_lower, ruta))
        elif "final" in nombre_lower:
            excel_dinamico = ruta
        elif "final" not in nombre_lower:
            base_file = ruta

    return base_file, secundarios

def crear_excel_con_hojas():
    # Pedir directorio al usuario
    tk.Tk().withdraw()
    directorio = filedialog.askdirectory(title="Selecciona el directorio con los Excel")

    if not directorio:
        print("❌ No se seleccionó ningún directorio.")
        return

    base_file, secundarios = detectar_archivos_excel(directorio)

    if not base_file or len(secundarios) < 3:
        print("❌ Faltan archivos necesarios.")
        return

    # Leer archivo base
    df_base = pd.read_excel(base_file)
    df_base.columns = [col.strip().upper() for col in df_base.columns]

    # Extraer SKU y limpiar PRODUCTO
    df_base[['SKU', 'PRODUCTO']] = df_base['PRODUCTO'].str.extract(r'\[(.*?)\]\s*\[.*?\]\s*(.*)')
    df_base['SKU'] = df_base['SKU'].astype(str).str.strip()
    df_base['PRODUCTO'] = df_base['PRODUCTO'].str.strip()

    # Crear hoja 1
    df_hoja1 = df_base[['SKU', 'PRODUCTO', 'VENTAS']].copy()
    for col in ['STOCK NAVE', 'STOCK MADRID', 'STOCK BULGARIA', 'PROXIMAMENTE', 'FECHA ENTRADA']:
        df_hoja1[col] = None

    # Inicializar hoja 2 y 3
    df_hoja2 = pd.DataFrame(columns=['SKU', 'STOCK'])
    df_hoja3 = pd.DataFrame(columns=['SKU', 'STOCK'])

    # Buscar secundarios y asignar a hoja 2 y 3
    df_hoja4 = None
    for nombre, ruta in secundarios:
        df_sec = pd.read_excel(ruta)
        df_sec.columns = [col.strip().upper() for col in df_sec.columns]
        df_sec['SKU'] = df_sec['SKU'].astype(str).str.strip()

        if 'nave' in nombre:
            df_hoja3 = df_sec[['SKU', 'STOCK']].copy()
        elif 'madrid' in nombre:
            df_hoja2 = df_sec[['SKU', 'STOCK']].copy()
        elif 'bulgaria' in nombre:
            df_hoja4 = df_sec.copy()

    #UNDELIVERED ORDER, next delivery = PROXIMAMENTE, FECHA ENTRADA
    df_hoja4.columns = [col.strip().upper() for col in df_hoja4.columns]
    df_hoja4.rename(columns={
        "UNDELIVERED ORDER": "PROXIMAMENTE",
        "NEXT DELIVERY": "FECHA ENTRADA"
    }, inplace=True)

    # Crear archivo final con las hojas
    output_path = os.path.join(directorio, "Base_Combinada.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_hoja1.to_excel(writer, sheet_name="1", index=False)
        df_hoja2.to_excel(writer, sheet_name="2", index=False)
        df_hoja3.to_excel(writer, sheet_name="3", index=False)
        df_hoja4.to_excel(writer, sheet_name="4", index=False)

    return output_path

def rellenar_stock_por_sku(ruta_excel):
    # Leer las hojas
    df_1 = pd.read_excel(ruta_excel, sheet_name="1")
    df_2 = pd.read_excel(ruta_excel, sheet_name="2")
    df_3 = pd.read_excel(ruta_excel, sheet_name="3")
    df_4 = pd.read_excel(ruta_excel, sheet_name="4")

    # Normalizar columnas
    df_1.columns = [col.strip().upper() for col in df_1.columns]
    df_2.columns = [col.strip().upper() for col in df_2.columns]
    df_3.columns = [col.strip().upper() for col in df_3.columns]
    df_4.columns = [col.strip().upper() for col in df_4.columns]

    # Normalizar SKUs
    df_1['SKU'] = df_1['SKU'].astype(str).str.strip()
    df_2['SKU'] = df_2['SKU'].astype(str).str.strip()
    df_3['SKU'] = df_3['SKU'].astype(str).str.strip()
    df_4['SKU'] = df_4['SKU'].astype(str).str.strip()

    # Crear diccionarios SKU
    stock_dict_df2 = df_2.set_index("SKU")["STOCK"].to_dict()
    stock_dict_df3 = df_3.set_index("SKU")["STOCK"].to_dict()
    stock_dict_df4_STOCK = df_4.set_index("SKU")["STOCK"].to_dict()
    stock_dict_df4_PROXI = df_4.set_index("SKU")["PROXIMAMENTE"].to_dict()
    stock_dict_df4_FECHA = df_4.set_index("SKU")["FECHA ENTRADA"].to_dict()

    # Rellenar campos
    df_1["STOCK MADRID"] = df_1["SKU"].map(stock_dict_df2)
    df_1["STOCK NAVE"] = df_1["SKU"].map(stock_dict_df3)
    df_1["STOCK BULGARIA"] = df_1["SKU"].map(stock_dict_df4_STOCK)
    df_1["PROXIMAMENTE"] = df_1["SKU"].map(stock_dict_df4_PROXI)
    df_1["FECHA ENTRADA"] = df_1["SKU"].map(stock_dict_df4_FECHA)

    # Guardar resultado
    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        df_1.to_excel(writer, sheet_name="1", index=False)

def actualizar_valores_en_hoja(main_excel, ruta_dinamico):
    from openpyxl import load_workbook

    try:
        # Leer datos desde main_excel (hoja "1")
        df_origen = pd.read_excel(main_excel, sheet_name="1")

        # Cargar el workbook dinámico
        wb = load_workbook(ruta_dinamico)
        ws = wb.worksheets[0]  # primera hoja

        # Leer cabecera de la hoja dinámica
        columnas_existentes = [cell.value for cell in ws[1]]

        # Filtrar columnas comunes
        columnas_comunes = [
            col for col in df_origen.columns
            if col in columnas_existentes
        ]

        if not columnas_comunes:
            print("⚠️ No hay columnas coincidentes.")
            return

        # Borrar filas desde la segunda (sin tocar cabecera)
        ws.delete_rows(2, ws.max_row)

        # Escribir valores en las columnas válidas
        for i, row in df_origen.iterrows():
            for j, col in enumerate(columnas_existentes):
                if col in columnas_comunes:
                    valor = row[col]
                    ws.cell(row=i+2, column=j+1, value=valor)  # +2 para mantener cabecera

        wb.save(ruta_dinamico)
        os.remove(main_excel)

        def calcular_y_ordenar_diferencia(ruta_excel):
            from datetime import datetime
            try:
                # Cargar libro y hoja (primera hoja)
                wb = load_workbook(ruta_excel)
                ws = wb.worksheets[0]

                # Leer cabecera
                columnas = [cell.value for cell in ws[1]]

                # Leer datos a partir de fila 2
                data = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columnas), values_only=True):
                    data.append(row)

                df = pd.DataFrame(data, columns=columnas)

                # Validar columnas necesarias
                if "STOCK NAVE" not in df.columns or "VENTAS" not in df.columns:
                    print("❌ Faltan columnas necesarias: 'STOCK NAVE' y/o 'VENTAS'")
                    return

                # Rellenar vacíos si hace falta
                df["STOCK NAVE"] = pd.to_numeric(df["STOCK NAVE"], errors="coerce").fillna(0)
                df["VENTAS"] = pd.to_numeric(df["VENTAS"], errors="coerce").fillna(0)

                # Calcular y ordenar
                df["DIFERENCIA"] = df["STOCK NAVE"] - df["VENTAS"]
                df.sort_values(by="DIFERENCIA", inplace=True)

                # Escribir resultados de vuelta en la hoja (manteniendo formato)
                for i, fila in enumerate(df.itertuples(index=False), start=2):  # desde fila 2
                    for j, valor in enumerate(fila, start=1):
                        #ws.cell(row=i, column=j, value=valor)
                        cell = ws.cell(row=i, column=j, value=valor)
                        if isinstance(valor, datetime):
                            cell.number_format = "DD/MM/YYYY"

                # Borrar filas sobrantes si hay menos datos que antes
                if len(df) < ws.max_row - 1:
                    ws.delete_rows(len(df) + 2, ws.max_row - len(df) - 1)

                wb.save(ruta_excel)

            except Exception as e:
                print(f"❌ Error durante el cálculo de DIFERENCIA: {e}")

        calcular_y_ordenar_diferencia(ruta_dinamico)

    except Exception as e:
        print(f"❌ Error al actualizar hoja dinámica: {e}")


def main():
    global excel_dinamico
    try:
        main_excel = crear_excel_con_hojas()
        if main_excel:
            rellenar_stock_por_sku(main_excel)
            actualizar_valores_en_hoja(main_excel, excel_dinamico)
            os.startfile(excel_dinamico)
        else:
            print("❌ No se generó el Excel base. Terminando.")
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")

if __name__ == "__main__":
    main()
