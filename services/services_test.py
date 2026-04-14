import pandas as pd
import xmlrpc.client
from utils import Utils

ruta = Utils.seleccionar_excel()

origen = {
    'url': "https://optimaluz.soluntec.net",
    'db': "Test",#"Real",
    'user': "jcoronado@optimaluz.com",
    'password': "AlAi4ever"
}

destino = {
    'url': "http://79.72.61.76:8070/",
    'db': "odoo1",
    'user': "admin",
    'password': "admin"
}

def transferencia_productos_por_sku():
    def migrate_products_by_sku(models_src, db_src, uid_src, pwd_src,
                                models_dst, db_dst, uid_dst, pwd_dst,
                                excel_path):
        """
        Migra productos filtrados por SKU (default_code).
        SOLUCIÓN 1:
        - NO se leen imágenes en el read masivo
        - image_1920 se lee producto a producto
        """

        con_atributos = False

        # -------------------------------------------------
        # 1️⃣ Leer SKUs desde Excel
        # -------------------------------------------------
        df = pd.read_excel(excel_path, sheet_name="Sheet1")

        if "SKU" not in df.columns:
            raise Exception("❌ El Excel no contiene la columna 'SKU'")

        skus = (
            df["SKU"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not skus:
            print("ℹ️ No hay SKUs en el Excel")
            return

        print(f"🧾 SKUs detectados: {len(skus)}")

        # -------------------------------------------------
        # 2️⃣ Campos (SIN image_1920)
        # -------------------------------------------------
        FIELDS = [
            'id',
            'active',
            'name',
            'default_code',
            'invoice_policy',
            'standard_price',
            'categ_id',
            'product_brand_id',
            'x_url',
            'description',
            'out_of_stock_message',
            'public_categ_ids',
            'allow_out_of_stock_order',
        ]

        # -------------------------------------------------
        # 3️⃣ Buscar productos por SKU
        # -------------------------------------------------
        domain = [('default_code', 'in', skus)]

        product_ids = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'search',
            [domain],
            {'context': {'active_test': False}}
        )

        if not product_ids:
            print("ℹ️ No se encontraron productos para esos SKUs")
            return

        products = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'read',
            [product_ids],
            {'fields': FIELDS}
        )

        print(f"📦 Productos a migrar: {len(products)}")

        # -------------------------------------------------
        # 4️⃣ Migración producto a producto
        # -------------------------------------------------
        for product in products:
            try:
                # -------------------------
                # 🔎 Comprobar si ya existe en destino
                # -------------------------
                existing_ids = models_dst.execute_kw(
                    db_dst, uid_dst, pwd_dst,
                    'product.template', 'search',
                    [[('default_code', '=', product.get('default_code'))]],
                    {'limit': 1, 'context': {'active_test': False}}
                )

                if existing_ids:
                    print(f"⏭ SKU {product.get('default_code')} ya existe en destino (ID {existing_ids[0]}) — se omite")
                    continue

                # ---------------------------------------------
                # 🔹 Leer imagen SOLO de este producto
                # ---------------------------------------------
                image_data = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.template', 'read',
                    [[product['id']]],
                    {'fields': ['image_1920']}
                )[0]['image_1920']

                vals = {
                    'active': product.get('active', True),
                    'image_1920': image_data,
                    'name': product.get('name'),
                    'default_code': product.get('default_code'),
                    'invoice_policy': product.get('invoice_policy'),
                    'standard_price': product.get('standard_price') or 0.0,
                    'x_url': product.get('x_url'),
                    'description': product.get('description'),
                    'out_of_stock_message': product.get('out_of_stock_message'),
                    'list_price': 0.0,
                    'is_storable': True,
                    'show_availability': True,
                    'available_threshold': 100.000,
                    'allow_out_of_stock_order': product.get('allow_out_of_stock_order'),
                }

                # -------------------------
                # Categoría interna
                # -------------------------
                if product.get('categ_id'):
                    categ_name = product['categ_id'][1]
                    categ_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.category', 'search',
                        [[('name', '=', categ_name)]],
                        {'limit': 1}
                    )
                    if categ_ids:
                        vals['categ_id'] = categ_ids[0]

                # -------------------------
                # Marca
                # -------------------------
                if product.get('product_brand_id'):
                    brand_name = product['product_brand_id'][1]
                    brand_ids = models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.brand', 'search',
                        [[('name', '=', brand_name)]],
                        {'limit': 1}
                    )
                    if brand_ids:
                        vals['product_brand_id'] = brand_ids[0]
                    else:
                        vals['product_brand_id'] = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.brand', 'create',
                            [{'name': brand_name}]
                        )

                # -------------------------
                # Categorías web
                # -------------------------
                if product.get('public_categ_ids'):
                    web_categ_ids = []
                    for wc in product['public_categ_ids']:
                        wc_name = models_src.execute_kw(
                            db_src, uid_src, pwd_src,
                            'product.public.category', 'read',
                            [[wc]],
                            {'fields': ['name']}
                        )[0]['name']

                        dst_wc = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.public.category', 'search',
                            [[('name', '=', wc_name)]],
                            {'limit': 1}
                        )

                        if dst_wc:
                            web_categ_ids.append(dst_wc[0])

                    if web_categ_ids:
                        vals['public_categ_ids'] = [(6, 0, web_categ_ids)]

                # -------------------------
                # Crear producto
                # -------------------------
                new_id = models_dst.execute_kw(
                    db_dst, uid_dst, pwd_dst,
                    'product.template', 'create',
                    [vals]
                )

                print(f"➕ Creado {vals.get('default_code')} (ID {new_id})")

                # -------------------------------------------------
                # 5️⃣ Migrar atributos
                # -------------------------------------------------
                if con_atributos:
                    attribute_lines = models_src.execute_kw(
                        db_src, uid_src, pwd_src,
                        'product.template.attribute.line', 'search_read',
                        [[('product_tmpl_id', '=', product['id'])]],
                        {'fields': ['attribute_id', 'value_ids']}
                    )

                    for line in attribute_lines:
                        attribute_name = line['attribute_id'][1]

                        attr_ids = models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.attribute', 'search',
                            [[('name', '=', attribute_name)]],
                            {'limit': 1}
                        )

                        attr_id = attr_ids[0] if attr_ids else models_dst.execute_kw(
                            db_dst, uid_dst, pwd_dst,
                            'product.attribute', 'create',
                            [{'name': attribute_name}]
                        )

                        value_ids_dst = []

                        for value_id in line['value_ids']:
                            value_name = models_src.execute_kw(
                                db_src, uid_src, pwd_src,
                                'product.attribute.value', 'read',
                                [[value_id]],
                                {'fields': ['name']}
                            )[0]['name']

                            val_ids = models_dst.execute_kw(
                                db_dst, uid_dst, pwd_dst,
                                'product.attribute.value', 'search',
                                [[
                                    ('name', '=', value_name),
                                    ('attribute_id', '=', attr_id)
                                ]],
                                {'limit': 1}
                            )
                            val_id = val_ids[0] if val_ids else models_dst.execute_kw(
                                db_dst, uid_dst, pwd_dst,
                                'product.attribute.value', 'create',
                                [{
                                    'name': value_name,
                                    'attribute_id': attr_id
                                }]
                            )

                            value_ids_dst.append(val_id)

                            models_dst.execute_kw(
                                db_dst, uid_dst, pwd_dst,
                                'product.template.attribute.line', 'create',
                                [{
                                    'product_tmpl_id': new_id,
                                    'attribute_id': attr_id,
                                    'value_ids': [(6, 0, value_ids_dst)]
                                }]
                            )

                # -------------------------------------------------
                # 6️⃣ Galería (ojo: sigue siendo pesada)
                # -------------------------------------------------
                images = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', '=', product['id'])]],
                    {'fields': ['image_1920', 'name', 'video_url', 'sequence']}
                )

                for img in images:
                    models_dst.execute_kw(
                        db_dst, uid_dst, pwd_dst,
                        'product.image', 'create',
                        [{
                            'product_tmpl_id': new_id,
                            'image_1920': img.get('image_1920'),
                            'name': img.get('name'),
                            'video_url': img.get('video_url'),
                            'sequence': img.get('sequence', 0)
                        }]
                    )

            except Exception as e:
                print(f"❌ Error en {product.get('default_code')}: {e}")

            print("✅ Migración por SKU finalizada")

    import xmlrpc.client
    # region LLAMADA
    # -----------------------
    # Configuración conexiones
    # -----------------------

    origen = {
        'url': "https://optimaluz.com/",
        'db': "odoo1",
        'user': "admin",
        'password': "1324",
    }

    destino = {
        'url': "https://b2b.optimaluz.com/", #"http://82.70.85.127:8069/",#
        'db': "odoo0",
        'user': "admin",
        'password': "1324"
    }

    # -----------------------
    # Conexión ORIGEN
    # -----------------------

    common_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_src = common_src.authenticate(
        origen['db'],
        origen['user'],
        origen['password'],
        {}
    )

    if not uid_src:
        raise Exception("❌ No se pudo autenticar en ORIGEN")

    models_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a ORIGEN (uid={uid_src})")

    # -----------------------
    # Conexión DESTINO
    # -----------------------

    common_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_dst = common_dst.authenticate(
        destino['db'],
        destino['user'],
        destino['password'],
        {}
    )

    if not uid_dst:
        raise Exception("❌ No se pudo autenticar en DESTINO")

    models_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a DESTINO (uid={uid_dst})")

    # -----------------------
    # LLAMADA A LA MIGRACIÓN
    # -----------------------

    migrate_products_by_sku(
        models_src=models_src,
        db_src=origen['db'],
        uid_src=uid_src,
        pwd_src=origen['password'],

        models_dst=models_dst,
        db_dst=destino['db'],
        uid_dst=uid_dst,
        pwd_dst=destino['password'],

        excel_path=ruta
    )
    # endregion

#transferencia_productos_por_sku()

def actualizar_categorias_por_sku(models_src, db_src, uid_src, pwd_src,
                                  models, db, uid, pwd,
                                  excel_path):

    try:
        df = pd.read_excel(excel_path)
        only_cats=False
        skus = (
            df["SKU"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not skus:
            print("ℹ️ No hay SKUs")
            return

        print(f"🧾 SKUs detectados: {len(skus)}")

        # -------------------------------------------------
        # 🔹 Productos ORIGEN
        # -------------------------------------------------

        src_products = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'search_read',
            [[('default_code', 'in', skus)]],
            {'fields': ['default_code', 'name', 'public_categ_ids', 'is_favorite']}
        )

        src_map = {p['default_code']: p for p in src_products}

        # -------------------------------------------------
        # 🔹 Productos DESTINO
        # -------------------------------------------------

        dst_products = models.execute_kw(
            db, uid, pwd,
            'product.template', 'search_read',
            [[('default_code', 'in', skus)]],
            {'fields': ['id', 'default_code', 'public_categ_ids']}
        )

        dst_map = {p['default_code']: p for p in dst_products}

        # -------------------------------------------------
        # 🔹 Categorías públicas DESTINO
        # -------------------------------------------------

        public_dst = models.execute_kw(
            db, uid, pwd,
            'product.public.category', 'search_read',
            [[]],
            {'fields': ['id', 'parent_id', 'x_id_interno']}
        )

        public_map_by_xid = {
            c['x_id_interno']: c
            for c in public_dst
            if c.get('x_id_interno')
        }

        public_map_by_id = {c['id']: c for c in public_dst}

        # -------------------------------------------------
        # 🔹 Categorías internas DESTINO
        # -------------------------------------------------

        internal_cats = models.execute_kw(
            db, uid, pwd,
            'product.category', 'search_read',
            [[]],
            {'fields': ['id', 'name']}
        )

        internal_map = {c['name']: c['id'] for c in internal_cats}

        # -------------------------------------------------
        # 🔹 Función subir a raíz pública
        # -------------------------------------------------

        def obtener_raiz(cat_id):
            while True:
                cat = public_map_by_id.get(cat_id)
                if not cat:
                    return cat_id
                parent = cat['parent_id'][0] if cat['parent_id'] else False
                if not parent:
                    return cat_id
                cat_id = parent

        # -------------------------------------------------
        # 🔹 LOOP PRINCIPAL
        # -------------------------------------------------

        for sku in skus:

            if sku not in src_map:
                print(f"⚠ No existe en origen: {sku}")
                continue

            if sku not in dst_map:
                print(f"⚠ No existe en destino: {sku}")
                continue

            product_src = src_map[sku]
            product_dst = dst_map[sku]

            # 🚀 Si ya tiene categorías públicas → saltar
            if product_dst.get('public_categ_ids') and only_cats:
                print(f"⏭ Ya tiene categorías asignadas: {sku}")
                continue

            new_public_ids = []
            raiz_public_id = False

            for pub_id in product_src.get('public_categ_ids', []):

                if pub_id not in public_map_by_xid:
                    print(f"⚠ Categoría pública no encontrada por x_id_interno: {pub_id}")
                    continue

                cat_dst = public_map_by_xid[pub_id]
                new_public_ids.append(cat_dst['id'])

                if not raiz_public_id:
                    raiz_public_id = obtener_raiz(cat_dst['id'])

            if not new_public_ids and only_cats:
                print(f"⚠ No se pudieron resolver categorías públicas para {sku}")
                continue

            vals = {}

            if new_public_ids:
                vals = {'public_categ_ids': [(6, 0, new_public_ids)]}

                # 🔹 Resolver categ_id (modelo interno) usando nombre de la raíz pública
                if raiz_public_id:

                    raiz_public = public_map_by_id.get(raiz_public_id)

                    if raiz_public:
                        # necesitamos el name real → lo leemos una sola vez
                        raiz_data = models.execute_kw(
                            db, uid, pwd,
                            'product.public.category', 'read',
                            [[raiz_public_id]],
                            {'fields': ['name']}
                        )[0]

                        raiz_name = raiz_data['name']

                        if raiz_name in internal_map:
                            vals['categ_id'] = internal_map[raiz_name]
                        else:
                            print(f"⚠ No existe categoría interna para raíz: {raiz_name}")

            vals['name'] = product_src.get('name')
            #vals['is_favorite'] = product_src.get('is_favorite')

            models.execute_kw(
                db, uid, pwd,
                'product.template', 'write',
                [[product_dst['id']], vals]
            )

            print(f"✅ Actualizado: {sku}")

        print("🎯 Proceso finalizado correctamente")

    except Exception as e:
        print(f"❌ Error: {e}")

def migrar_supplierinfo_por_sku(models_src, db_src, uid_src, pwd_src,
                                models, db, uid, pwd,
                                excel_path):

    try:
        # -------------------------------------------------
        # 1️⃣ Leer SKUs
        # -------------------------------------------------

        df = pd.read_excel(excel_path)

        skus = (
            df["SKU"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not skus:
            print("ℹ️ No hay SKUs")
            return

        print(f"🧾 SKUs detectados: {len(skus)}")

        # -------------------------------------------------
        # 2️⃣ Leer productos ORIGEN
        # -------------------------------------------------

        src_products = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.template', 'search_read',
            [[('default_code', 'in', skus), ('active', 'in', [True, False])]],
            {'fields': ['id', 'default_code']}
        )

        src_map = {p['id']: p['default_code'] for p in src_products}

        if not src_map:
            print("ℹ️ No se encontraron productos en origen")
            return

        # -------------------------------------------------
        # 3️⃣ Leer productos DESTINO
        # -------------------------------------------------

        dst_products = models.execute_kw(
            db, uid, pwd,
            'product.template', 'search_read',
            [[('default_code', 'in', list(src_map.values())), ('active', 'in', [True, False])]],
            {'fields': ['id', 'default_code']}
        )

        dst_map = {p['default_code']: p['id'] for p in dst_products}

        # -------------------------------------------------
        # 4️⃣ Leer supplierinfo ORIGEN
        # -------------------------------------------------

        supplierinfos = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.supplierinfo', 'search_read',
            [[('product_tmpl_id', 'in', list(src_map.keys()))]],
            {'fields': ['product_tmpl_id',
                        'partner_id',
                        'product_name',
                        'product_code',
                        'price', 'min_qty']}
        )

        if not supplierinfos:
            print("ℹ️ No hay supplierinfo en origen")
            return

        print(f"📦 Supplierinfos encontrados: {len(supplierinfos)}")

        # -------------------------------------------------
        # 5️⃣ Procesar uno a uno
        # -------------------------------------------------

        for sup in supplierinfos:

            src_product_id = sup['product_tmpl_id'][0]
            sku = src_map.get(src_product_id)

            if not sku or sku not in dst_map:
                print(f"⚠ Producto destino no encontrado para SKU {sku}")
                continue

            dst_product_id = dst_map[sku]

            # 🔹 Resolver partner por x_id_interno
            if not sup.get('partner_id'):
                print(f"⚠ Supplier sin partner en SKU {sku}")
                continue

            partner_src_id = sup['partner_id'][0]

            partner_dst = models.execute_kw(
                db, uid, pwd,
                'res.partner', 'search',
                [[('x_id_interno', '=', partner_src_id)]],
                {'limit': 1}
            )

            if not partner_dst:
                print(f"⚠ Partner no encontrado en destino (x_id_interno={partner_src_id})")
                continue

            partner_dst_id = partner_dst[0]

            '''# 🔹 Evitar duplicado
            existing = models.execute_kw(
                db, uid, pwd,
                'product.supplierinfo', 'search',
                [[
                    ('product_tmpl_id', '=', dst_product_id),
                    ('partner_id', '=', partner_dst_id),
                    ('product_code', '=', sup.get('product_code'))
                ]],
                {'limit': 1}
            )

            if existing:
                print(f"⏭ Ya existe supplierinfo para SKU {sku}")
                continue'''

            # 🔹 Crear supplierinfo
            vals = {
                'product_tmpl_id': dst_product_id,
                'partner_id': partner_dst_id,
                'product_name': sup.get('product_name'),
                'product_code': sup.get('product_code'),
                'price': sup.get('price') or 0.0,
                'min_qty': sup.get('min_qty') or 1.0,
            }

            models.execute_kw(
                db, uid, pwd,
                'product.supplierinfo', 'create',
                [vals]
            )

            print(f"✅ Supplierinfo migrado para {sku}")

        print("🎯 Migración supplierinfo finalizada")

    except Exception as e:
        print(f"❌ Error migrando supplierinfo: {e}")

def migrar_media_y_web_por_sku(models_src, db_src, uid_src, pwd_src,
                               models, db, uid, pwd,
                               excel_path,
                               batch_size=20):  # 🔥 batch pequeño

    try:
        df = pd.read_excel(excel_path)

        skus = (
            df["SKU"]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        if not skus:
            print("ℹ️ No hay SKUs")
            return

        print(f"🧾 SKUs detectados: {len(skus)}")

        for i in range(0, len(skus), batch_size):

            batch = skus[i:i + batch_size]

            # -------------------------------------------------
            # 🔹 Leer solo IDs ORIGEN (sin binarios)
            # -------------------------------------------------

            templates_src = models_src.execute_kw(
                db_src, uid_src, pwd_src,
                'product.template', 'search_read',
                [[('default_code', 'in', batch)]],
                {'fields': ['id', 'default_code']}
            )

            src_map = {t['default_code']: t['id'] for t in templates_src}

            # -------------------------------------------------
            # 🔹 Leer solo IDs DESTINO
            # -------------------------------------------------

            templates_dst = models.execute_kw(
                db, uid, pwd,
                'product.template', 'search_read',
                [[('default_code', 'in', batch)]],
                {'fields': ['id', 'default_code', 'website_description']}
            )

            dst_map = {t['default_code']: t for t in templates_dst}

            for sku in batch:

                if sku not in src_map or sku not in dst_map:
                    continue

                tmpl_id_src = src_map[sku]
                product_dst = dst_map[sku]

                if product_dst.get('website_description'):
                    print(f"⏭ Ya tiene descripción web: {sku}")
                    continue

                tmpl_id_dst = product_dst['id']

                # -------------------------------------------------
                # 🔹 Ahora sí leer BINARIOS SOLO de este producto
                # -------------------------------------------------

                product_src = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.template', 'read',
                    [[tmpl_id_src]],
                    {'fields': [
                        'image_1920',
                        'website_description'
                    ] + [f'x_icono{i}' for i in range(1, 9)]}
                )[0]

                vals = {}

                if product_src.get('image_1920'):
                    vals['image_1920'] = product_src['image_1920']

                if product_src.get('website_description'):
                    vals['website_description'] = product_src['website_description']

                for idx in range(1, 9):
                    campo = f'x_icono{idx}'
                    if product_src.get(campo):
                        vals[campo] = product_src[campo]

                if vals:
                    models.execute_kw(
                        db, uid, pwd,
                        'product.template', 'write',
                        [[tmpl_id_dst], vals]
                    )

                # -------------------------------------------------
                # 🔹 GALERÍA (leer solo del producto actual)
                # -------------------------------------------------

                images_src = models_src.execute_kw(
                    db_src, uid_src, pwd_src,
                    'product.image', 'search_read',
                    [[('product_tmpl_id', '=', tmpl_id_src)]],
                    {'fields': ['image_1920', 'name', 'video_url', 'sequence']}
                )

                # Borrar galería destino
                existing_imgs = models.execute_kw(
                    db, uid, pwd,
                    'product.image', 'search',
                    [[('product_tmpl_id', '=', tmpl_id_dst)]]
                )

                if existing_imgs:
                    models.execute_kw(
                        db, uid, pwd,
                        'product.image', 'unlink',
                        [existing_imgs]
                    )

                # Insertar nuevas
                for img in images_src:

                    if not img.get('image_1920') and not img.get('video_url'):
                        continue

                    models.execute_kw(
                        db, uid, pwd,
                        'product.image', 'create',
                        [{
                            'product_tmpl_id': tmpl_id_dst,
                            'image_1920': img.get('image_1920'),
                            'video_url': img.get('video_url'),
                            'sequence': img.get('sequence', 10),
                            'name': img.get('name'),
                        }]
                    )

                print(f"✅ Media migrada: {sku}")

        print("🎯 Migración multimedia finalizada sin problemas de memoria")

    except Exception as e:
        print(f"❌ Error migrando multimedia: {e}")

def migrar_ausencias(models_src, db_src, uid_src, pwd_src,
                     models, db, uid, pwd):

    try:

        # -------------------------------------------------
        # 🔹 1️⃣ Leer ausencias ORIGEN
        # -------------------------------------------------

        leaves_src = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'hr.leave', 'search_read',
            [[("employee_id.active", "=", True)]],
            {'fields': [
                'employee_id',
                'department_id',
                'request_date_from',
                'request_date_to',
                'date_from',
                'date_to',
                'name',
                'holiday_status_id',
                'state'
            ]}
        )

        if not leaves_src:
            print("ℹ️ No hay ausencias en origen")
            return

        print(f"🗓 Ausencias detectadas: {len(leaves_src)}")

        # -------------------------------------------------
        # 🔹 2️⃣ Precargar empleados DESTINO
        # -------------------------------------------------

        employees_dst = models.execute_kw(
            db, uid, pwd,
            'hr.employee', 'search_read',
            [[]],
            {'fields': ['id', 'x_id_interno']}
        )

        employee_map = {
            e['x_id_interno']: e['id']
            for e in employees_dst
            if e.get('x_id_interno')
        }

        # -------------------------------------------------
        # 🔹 3️⃣ Precargar departamentos DESTINO
        # -------------------------------------------------

        departments_dst = models.execute_kw(
            db, uid, pwd,
            'hr.department', 'search_read',
            [[]],
            {'fields': ['id', 'name']}
        )

        department_map = {d['name']: d['id'] for d in departments_dst}

        # -------------------------------------------------
        # 🔹 4️⃣ Precargar tipos de ausencia DESTINO
        # -------------------------------------------------

        types_dst = models.execute_kw(
            db, uid, pwd,
            'hr.leave.type', 'search_read',
            [[]],
            {'fields': ['id', 'x_id_interno']}
        )

        type_map = {
            e['x_id_interno']: e['id']
            for e in types_dst
            if e.get('x_id_interno')
        }

        # -------------------------------------------------
        # 🔹 5️⃣ Migración
        # -------------------------------------------------

        for leave in leaves_src:

            # 🔹 Resolver empleado
            if not leave.get('employee_id'):
                print("⚠ Ausencia sin empleado")
                continue

            employee_src_id = leave['employee_id'][0]

            if employee_src_id not in employee_map:
                print(f"⚠ Empleado no encontrado (x_id_interno={employee_src_id})")
                continue

            employee_dst_id = employee_map[employee_src_id]

            # 🔹 Resolver departamento
            department_dst_id = False
            if leave.get('department_id'):
                dept_name = leave['department_id'][1]
                department_dst_id = department_map.get(dept_name)

                if not department_dst_id:
                    print(f"⚠ Departamento no encontrado: {dept_name}")

            # 🔹 Resolver tipo ausencia
            type_src_id = leave['holiday_status_id'][0]

            if type_src_id not in type_map:
                print(f"⚠ Tipo de ausencia no encontrado: (x_id_interno={type_src_id})")
                continue

            type_dst_id = type_map[type_src_id]



            # 🔹 Preparar valores
            vals = {
                'employee_id': employee_dst_id,
                'holiday_status_id': type_dst_id,
                'request_date_from': leave.get('request_date_from'),
                'request_date_to': leave.get('request_date_to'),
                'date_from': leave.get('date_from'),
                'date_to': leave.get('date_to'),
                'name': leave.get('name'),
            }

            if department_dst_id:
                vals['department_id'] = department_dst_id

            # -------------------------------------------------
            # 🔹 Crear en estado borrador
            # -------------------------------------------------

            leave_id = models.execute_kw(
                db, uid, pwd,
                'hr.leave', 'create',
                [vals]
            )

            # -------------------------------------------------
            # 🔹 Ajustar estado
            # -------------------------------------------------

            state = leave.get('state')

            if state == 'confirm':
                models.execute_kw(db, uid, pwd,
                                  'hr.leave', 'action_confirm',
                                  [[leave_id]])

            elif state == 'validate':
                models.execute_kw(db, uid, pwd,
                                  'hr.leave', 'action_validate',
                                  [[leave_id]])

            elif state == 'refuse':
                models.execute_kw(db, uid, pwd,
                                  'hr.leave', 'action_refuse',
                                  [[leave_id]])

            print(f"✅ Ausencia migrada: {leave.get('name')}")

        print("🎯 Migración de ausencias finalizada")

    except Exception as e:
        print(f"❌ Error migrando ausencias: {e}")

def corregir_lineas_qty(models_src, db_src, uid_src, pwd_src,
                        models, db, uid, pwd):

    try:

        print("🔎 Buscando líneas origen con qty=0")

        # -------------------------------------------------
        # 1️⃣ Leer líneas origen qty=0
        # -------------------------------------------------

        lines_src = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'sale.order.line', 'search_read',
            [[('product_uom_qty', '=', 0)]],
            {'fields': ['order_id', 'product_id']}
        )

        if not lines_src:
            print("ℹ️ No hay líneas origen con qty=0")
            return

        # productos origen
        product_ids_src = list({l['product_id'][0] for l in lines_src if l['product_id']})

        products_src = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'product.product', 'read',
            [product_ids_src],
            {'fields': ['default_code']}
        )

        sku_map_src = {p['id']: p['default_code'] for p in products_src}

        # pedidos origen
        order_ids_src = list({l['order_id'][0] for l in lines_src if l['order_id']})

        orders_src = models_src.execute_kw(
            db_src, uid_src, pwd_src,
            'sale.order', 'read',
            [order_ids_src],
            {'fields': ['name']}
        )

        order_map_src = {o['id']: o['name'] for o in orders_src}

        src_keys = set()

        for l in lines_src:

            if not l['product_id'] or not l['order_id']:
                continue

            sku = sku_map_src.get(l['product_id'][0])
            order_name = order_map_src.get(l['order_id'][0])

            if sku and order_name:
                src_keys.add((order_name, sku))

        print(f"📦 Líneas origen relevantes: {len(src_keys)}")

        # -------------------------------------------------
        # 2️⃣ Buscar pedidos destino
        # -------------------------------------------------

        order_names = list({k[0] for k in src_keys})

        orders_dst = models.execute_kw(
            db, uid, pwd,
            'sale.order', 'search_read',
            [[('name', 'in', order_names)]],
            {'fields': ['id', 'name']}
        )

        order_map_dst = {o['id']: o['name'] for o in orders_dst}

        order_ids_dst = list(order_map_dst.keys())

        # -------------------------------------------------
        # 3️⃣ Leer líneas destino
        # -------------------------------------------------

        lines_dst = models.execute_kw(
            db, uid, pwd,
            'sale.order.line', 'search_read',
            [[('order_id', 'in', order_ids_dst)]],
            {'fields': ['id', 'order_id', 'product_id', 'product_uom_qty', 'price_unit']}
        )

        # obtener SKUs destino
        product_ids_dst = list({l['product_id'][0] for l in lines_dst if l['product_id']})

        products_dst = models.execute_kw(
            db, uid, pwd,
            'product.product', 'read',
            [product_ids_dst],
            {'fields': ['default_code']}
        )

        sku_map_dst = {p['id']: p['default_code'] for p in products_dst}

        # -------------------------------------------------
        # 4️⃣ Corregir líneas
        # -------------------------------------------------

        corregidas = 0

        for l in lines_dst:

            if not l['product_id'] or not l['order_id']:
                continue

            order_name = order_map_dst.get(l['order_id'][0])
            sku = sku_map_dst.get(l['product_id'][0])

            if not order_name or not sku:
                continue

            key = (order_name, sku)

            if key in src_keys and l['product_uom_qty'] == 1:

                price = l['price_unit']

                models.execute_kw(
                    db, uid, pwd,
                    'sale.order.line', 'write',
                    [[l['id']], {
                        'product_uom_qty': 0,
                        'price_unit': price
                    }]
                )

                corregidas += 1

                print(
                    f"✔ Corregido pedido {order_name} | "
                    f"SKU {sku} | qty 1→0 | price_unit mantenido {price}"
                )

        print(f"\n🎯 Líneas corregidas: {corregidas}")

    except Exception as e:
        print(f"❌ Error corrigiendo líneas: {e}")

def ejecutar_funciones_transferencia():
    # region CONEXION
    # -----------------------
    # Configuración conexiones
    # -----------------------

    origen = {
        'url': "http://37.59.66.189:8069/",#"http://158.179.220.107:8069/",#
        'db': "Real",#"odoo0",#
        'user': "jcoronado@optimaluz.com",#"admin",#
        'password': "AlAi4ever"#"admin",#
    }

    destino = {
        'url': "https://optimaluz.com/",
        'db': "odoo1",  # "odoo0",
        'user': "admin",
        'password': "1324"
    }

    # -----------------------
    # Conexión ORIGEN
    # -----------------------

    common_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_src = common_src.authenticate(
        origen['db'],
        origen['user'],
        origen['password'],
        {}
    )

    if not uid_src:
        raise Exception("❌ No se pudo autenticar en ORIGEN")

    models_src = xmlrpc.client.ServerProxy(
        f"{origen['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a ORIGEN (uid={uid_src})")

    # -----------------------
    # Conexión DESTINO
    # -----------------------

    common_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/common",
        allow_none=True
    )

    uid_dst = common_dst.authenticate(
        destino['db'],
        destino['user'],
        destino['password'],
        {}
    )

    if not uid_dst:
        raise Exception("❌ No se pudo autenticar en DESTINO")

    models_dst = xmlrpc.client.ServerProxy(
        f"{destino['url']}/xmlrpc/2/object",
        allow_none=True
    )

    print(f"🔌 Conectado a DESTINO (uid={uid_dst})")
    # endregion

    # -----------------------
    # LLAMADA A LA MIGRACIÓN
    # -----------------------

    corregir_lineas_qty(
        models_src=models_src,
        db_src=origen['db'],
        uid_src=uid_src,
        pwd_src=origen['password'],

        models=models_dst,
        db=destino['db'],
        uid=uid_dst,
        pwd=destino['password'],

        #excel_path=ruta
    )


#ejecutar_funciones_transferencia()

def rellenar_nuevos(path_excel=ruta, hoja=0):
    '''
    Detecta “nuevos” comparando las columnas SKU_2 contra SKU_1.
    Concretamente, considera nuevo los SKUs que aparecen en SKU_2 pero no aparece en SKU_1
    '''
    # Leer Excel
    df = pd.read_excel(path_excel, sheet_name=hoja)

    # Normalizar a string y eliminar NaN
    sku1 = df["SKU_1"].dropna().astype(str).str.strip()
    sku2 = df["SKU_2"].dropna().astype(str).str.strip()

    # Convertir SKU_1 a set para comparación rápida
    sku1_set = set(sku1)

    # Detectar nuevos
    nuevos = [sku for sku in sku2 if sku not in sku1_set]

    # Vaciar columna NUEVOS
    df["NUEVOS"] = None

    # Escribir resultados
    df.loc[:len(nuevos)-1, "NUEVOS"] = nuevos

    # Guardar Excel
    df.to_excel(path_excel, index=False)

    print(f"SKUs nuevos encontrados: {len(nuevos)}")

def extraer_skus_pdf_catalogo_a_excel(pdf_path, excel_salida="skus_extraidos.xlsx"):
    import pdfplumber
    import re

    skus = {}

    # patrón SKU válido
    patron_sku = re.compile(r'^\d{4,}(?:-\d+)?$')

    with pdfplumber.open(pdf_path) as pdf:
        for num_pagina, page in enumerate(pdf.pages, start=1):

            text = page.extract_text()

            if not text:
                continue

            for line in text.split("\n"):
                line = line.strip()

                if "|" in line:
                    posible_sku = line.split("|")[0].strip()
                else:
                    match = re.match(r'^\S+', line)
                    if not match:
                        continue
                    posible_sku = match.group()

                # validar SKU
                if patron_sku.match(posible_sku):

                    # guardar solo la primera vez que aparece
                    if posible_sku not in skus:
                        skus[posible_sku] = num_pagina

    # crear dataframe
    df = pd.DataFrame(
        [(sku, pagina) for sku, pagina in skus.items()],
        columns=["SKU", "Pagina"]
    )

    df.sort_values("SKU", inplace=True)

    df.to_excel(excel_salida, index=False)

    print(f"✔ {len(skus)} SKUs extraídos")
    print(f"📁 Excel generado: {excel_salida}")


def actualizar_odoos():
    def actualizar_prods_diariamente(origen, destino):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from datetime import datetime, timedelta
        import re
        import os
        import pandas as pd
        import xmlrpc.client
        import ast

        def conectar(config):
            common = xmlrpc.client.ServerProxy(f"{config['url']}/xmlrpc/2/common")
            uid = common.authenticate(config['db'], config['user'], config['password'], {})

            models = xmlrpc.client.ServerProxy(f"{config['url']}/xmlrpc/2/object")

            return models, uid

        # 🔌 Conexiones
        models_o, uid_o = conectar(origen)
        models_d, uid_d = conectar(destino)

        def get_products_dataframe(models, db, uid, password, domain, fields):
            products = models.execute_kw(
                db, uid, password,
                'product.template', 'search_read',
                [domain],
                {'fields': fields, 'context': {'lang': 'es_ES'}}
            )

            if not products:
                return pd.DataFrame()

            df = pd.DataFrame(products)

            df['categ_id'] = df['categ_id'].apply(
                lambda x: x[1] if isinstance(x, (list, tuple)) else x
            )

            columns_order = [
                'default_code', 'name', 'standard_price',
                'list_price', 'categ_id', 'description', 'public_categ_ids', 'x_almacen_local'
            ]

            return df[columns_order]

        def export_full_product_data():
            # 📅 Rango fechas (solo origen)
            today = datetime.today()
            start_date = today.replace(day=1, hour=0, minute=0, second=0).strftime('%Y-%m-%d %H:%M:%S')
            next_month = (today.replace(day=28) + timedelta(days=5)).replace(day=1, hour=0, minute=0, second=0)
            end_date = next_month.strftime('%Y-%m-%d %H:%M:%S')

            domain = [
                ('write_uid', '=', 33),
                ('write_date', '>=', start_date),
                ('write_date', '<', end_date)
            ]

            fields = [
                'default_code',
                'name',
                'standard_price',
                'list_price',
                'description',
                'categ_id',
                'public_categ_ids',
                'is_published',
                'x_almacen_local',
            ]

            # 📥 NEW → origen (filtrado)
            df_new = get_products_dataframe(
                models_o, origen['db'], uid_o, origen['password'],
                domain, fields
            )

            if df_new.empty:
                print("No hay registros en origen.")
                return False

            # 📥 OLD → destino (🔥 TODOS)
            df_old = get_products_dataframe(
                models_d, destino['db'], uid_d, destino['password'],
                [], fields
            )

            # 📁 Archivo
            filename = os.path.expanduser("~/Documents/SMI Files/cambios_ose.xlsx")

            # 💾 Guardado
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_old.to_excel(writer, sheet_name='OLD', index=False)
                df_new.to_excel(writer, sheet_name='NEW', index=False)

            return filename

        def detectar_cambios_excel(ruta_excel, debug=False):

            def normalize(val):
                # 🔹 Vacíos unificados
                if pd.isna(val) or val in ["", "nan", "None"]:
                    return None

                # 🔹 Strings
                if isinstance(val, str):
                    val = val.strip()

                    # número con coma → float
                    if re.match(r"^\d+,\d+$", val):
                        return round(float(val.replace(",", ".")), 2)

                    # lista en string
                    if val.startswith("[") and val.endswith("]"):
                        try:
                            parsed = ast.literal_eval(val)
                            if isinstance(parsed, list):
                                return sorted(parsed)
                        except:
                            pass

                    return val

                # 🔹 float
                if isinstance(val, float):
                    return round(val, 2)

                # 🔹 lista real
                if isinstance(val, list):
                    return sorted(val)

                return val

            # -----------------------------
            # 📥 Leer Excel
            # -----------------------------
            df_old = pd.read_excel(ruta_excel, sheet_name='OLD')
            df_new = pd.read_excel(ruta_excel, sheet_name='NEW')

            # 🔧 Normalizar default_code (CLAVE)
            df_old['default_code'] = df_old['default_code'].astype(str).str.strip()
            df_new['default_code'] = df_new['default_code'].astype(str).str.strip()

            if 'default_code' not in df_old.columns or 'default_code' not in df_new.columns:
                raise ValueError("Falta la columna 'default_code'.")

            # -----------------------------
            # 🔧 Normalización previa
            # -----------------------------
            for col in ['list_price', 'standard_price']:
                if col in df_old.columns:
                    df_old[col] = df_old[col].fillna(0).astype(float).round(2)
                if col in df_new.columns:
                    df_new[col] = df_new[col].fillna(0).astype(float).round(2)

            # limpiar
            df_old = df_old.dropna(subset=['default_code']).drop_duplicates(subset=['default_code'])
            df_new = df_new.dropna(subset=['default_code']).drop_duplicates(subset=['default_code'])

            df_old = df_old.set_index('default_code')
            df_new = df_new.set_index('default_code')

            comunes = df_old.index.intersection(df_new.index)
            columnas = [c for c in df_old.columns if c in df_new.columns]

            cambios = []

            # -----------------------------
            # 🔍 Comparación
            # -----------------------------
            for code in comunes:
                fila_old = df_old.loc[code]
                fila_new = df_new.loc[code]

                fila_resultado = {"default_code": code}
                hay_cambios = False

                if debug:
                    print("\n" + "=" * 60)
                    print(f"🔎 SKU: {code}")
                    print("-" * 60)

                for col in columnas:
                    val_old_raw = fila_old[col]
                    val_new_raw = fila_new[col]

                    val_old = normalize(val_old_raw)
                    val_new = normalize(val_new_raw)

                    coincide = val_old == val_new

                    if not coincide:
                        hay_cambios = True

                        # lógica de cambios
                        if val_new is None and val_old is not None:
                            fila_resultado[col] = "*"
                        else:
                            fila_resultado[col] = val_new_raw
                    else:
                        fila_resultado[col] = ""

                    # 🔥 DEBUG DETALLADO
                    if debug:
                        estado = "OK" if coincide else "DIFF"
                        print(f"{col}:")
                        print(f"   OLD → {val_old_raw}")
                        print(f"   NEW → {val_new_raw}")
                        print(f"   NORMALIZED → {val_old} | {val_new}")
                        print(f"   RESULT → {estado}")
                        print("-" * 40)

                if debug:
                    if hay_cambios:
                        print(f"❌ RESULTADO FINAL SKU {code}: DIFERENCIAS")
                    else:
                        print(f"✅ RESULTADO FINAL SKU {code}: COINCIDE TOTAL")

                if hay_cambios:
                    cambios.append(fila_resultado)

            # -----------------------------
            # ❌ Sin cambios
            # -----------------------------
            if not cambios:
                print("✅ No hay diferencias.")
                return False

            # -----------------------------
            # 📊 Crear DF resultado
            # -----------------------------
            df_cambios = pd.DataFrame(cambios)

            columnas_finales = ['default_code'] + [c for c in columnas if c != 'default_code']
            df_cambios = df_cambios[columnas_finales]

            # -----------------------------
            # 💾 Guardar Excel
            # -----------------------------
            with pd.ExcelWriter(ruta_excel, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df_cambios.to_excel(writer, sheet_name='CAMBIOS', index=False)

            print(f"💾 {len(df_cambios)} productos con cambios.")

            # -----------------------------
            # 🎨 Resaltado visual
            # -----------------------------
            wb = load_workbook(ruta_excel)
            ws = wb["CAMBIOS"]

            amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            col_default = headers["default_code"]

            default_codes = set(
                str(ws.cell(row=r, column=col_default).value).strip()
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=col_default).value
            )

            for col_name, col_idx in headers.items():
                if col_name == "default_code":
                    continue

                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    val = cell.value

                    if val not in (None, "") and str(val).strip() not in default_codes:
                        cell.fill = amarillo

            wb.save(ruta_excel)

            print("🎨 Cambios resaltados.")

            return ruta_excel

        def aplicar_cambios_desde_excel(
                models, db, uid, password,
                ruta_excel,
                hoja_cambios="CAMBIOS",
                columna_ref="default_code"):

            COLOR_AMARILLO = "FFFF00"

            # =========================
            # 📄 1. Cargar Excel
            # =========================

            wb = load_workbook(ruta_excel, data_only=True)
            ws = wb[hoja_cambios]

            headers = [c.value for c in ws[1]]
            col_index = {headers[i]: i + 1 for i in range(len(headers))}

            # =========================
            # 🟡 2. Detectar cambios
            # =========================

            cambios = []

            for row in ws.iter_rows(min_row=2):
                ref_cell = row[col_index[columna_ref] - 1]
                ref_value = str(ref_cell.value).strip() if ref_cell.value else ""

                if not ref_value:
                    continue

                for col_name, idx in col_index.items():
                    cell = row[idx - 1]
                    fill = cell.fill

                    is_yellow = (
                            fill
                            and fill.fgColor
                            and fill.fgColor.rgb is not None
                            and fill.fgColor.rgb[-6:].upper() == COLOR_AMARILLO
                    )

                    if is_yellow:
                        cambios.append({
                            "default_code": ref_value,
                            "columna": col_name,
                            "valor": cell.value,
                        })

            if not cambios:
                print("No se encontraron cambios resaltados en amarillo.")
                return

            # =========================
            # 🚀 3. Aplicar cambios
            # =========================

            products = models.execute_kw(
                db, uid, password,
                "product.template", "search_read",
                [[]],
                {"fields": ["id", "default_code"]}
            )

            mapa = {p["default_code"]: p["id"] for p in products}

            for cambio in cambios:
                ref = cambio["default_code"]
                col = cambio["columna"]
                val = cambio["valor"]

                product_id = mapa.get(ref)

                if not product_id:
                    print(f"⚠ No encontrado en Odoo → {ref}")
                    continue

                vals = {}

                try:
                    # =========================
                    # 🧠 CAMPOS ESPECIALES
                    # =========================

                    if col == "categ_id":
                        try:
                            if val in [None, "", "nan", "*"]:
                                vals["categ_id"] = False
                            else:
                                if isinstance(val, (list, tuple)):
                                    vals["categ_id"] = val[0] if val else False
                                else:
                                    vals["categ_id"] = int(val)
                        except Exception:
                            print(f"⚠ Error parseando categ_id: {val}")
                            continue

                    elif col == "public_categ_ids":
                        try:
                            if val in [None, "", "nan", "*"]:
                                vals["public_categ_ids"] = [(5, 0, 0)]
                            else:
                                if isinstance(val, str):
                                    ids = ast.literal_eval(val)
                                elif isinstance(val, list):
                                    ids = val
                                else:
                                    ids = []

                                vals["public_categ_ids"] = [(6, 0, ids)] if ids else [(5, 0, 0)]

                        except Exception:
                            print(f"⚠ Error parseando public_categ_ids: {val}")
                            continue

                    # =========================
                    # 🧨 GESTIÓN DE ELIMINACIONES
                    # =========================

                    if val == "*" or val in [None, "", "nan"]:
                        if col == "categ_id":
                            vals["categ_id"] = False

                        elif col == "public_categ_ids":
                            vals["public_categ_ids"] = [(5, 0, 0)]  # eliminar todos

                        else:
                            vals[col] = False  # limpia el campo

                    # =========================
                    # 🟢 VALORES NORMALES
                    # =========================
                    else:
                        if col in ["standard_price", "list_price"]:
                            vals[col] = float(val)
                        else:
                            vals[col] = val

                    # =========================
                    # ✍️ WRITE
                    # =========================

                    models.execute_kw(
                        db, uid, password,
                        "product.template", "write",
                        [[product_id], vals]
                    )

                    print(f"✔ Actualizado {ref}: {col} = {val}")

                except Exception as e:
                    print(f"❌ Error actualizando {ref} ({col}): {e}")

            print("✔ Todos los cambios han sido aplicados a Odoo correctamente.")

        excel_file = export_full_product_data()

        if excel_file:
            excel_file_cambios = detectar_cambios_excel(excel_file)

            if excel_file_cambios:
                aplicar_cambios_desde_excel(models_d, destino['db'], uid_d, destino['password'],
                                            ruta_excel=excel_file_cambios)

    origen = {
        'url': "https://b2b.optimaluz.com/",
        'db': "odoo0",
        'user': "admin",
        'password': "1324",
    }

    destinos = [
        {
            'url': "https://optimaluz.com/",
            'db': "odoo1",
            'user': "admin",
            'password': "1324"
        },
        {
            'url': "http://82.70.85.127:8069/",
            'db': "odoo0",
            'user': "admin",
            'password': "1324"
        }
    ]

    for destino in destinos:
        print(f"\n🚀 Actualizando destino: {destino['url']}")
        actualizar_prods_diariamente(origen, destino)


#actualizar_odoos()


